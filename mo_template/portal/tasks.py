#!/usr/bin/env python3
import requests
import importlib
import dateparser
import os
os.environ[ 'DJANGO_SETTINGS_MODULE' ] = "mo_template.settings"
from celery import Celery,shared_task
from celery.schedules import crontab
from datetime import datetime,timedelta
import sys
sys.path.append(os.getcwd())
import logging
logger = logging.getLogger(__name__)
#for Memurai
#celery_app = Celery('tasks', broker='redis://localhost:6373/0', backend="rpc://")
celery_app = Celery('tasks', broker='redis://localhost:6379/0', backend='redis://localhost:6379/0')
celery_app.conf.update(broker_transport_options={"max_retries": 0, "interval_start": 0, "interval_step": 0.2, "interval_max": 0.5}, accept_content = ['json','pickle'],accept_results = ['json','pickle'],broker_heartbeat = 15,acks_late=False)
from django.conf import settings
FILE_PATH = settings.MEDIA_URL


#*****************************************************************************#
@shared_task
def check_stock(quapi_id,ctrl_id,ctrl_number):
    error,msg = '','' 
    from polls.models import QueryApi,Warehouse
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not cr:
        return 'Quantum connection failed.',msg
        
    query = """SELECT STM_AUTO_KEY FROM STOCK 
        WHERE ROWNUM < 3 AND
        CTRL_ID = '%s' AND CTRL_NUMBER = '%s'
        AND QTY_OH > 0 AND HISTORICAL_FLAG = 'F'
        """%(ctrl_id,ctrl_number)
        
    stm = selection_dir(query,cr)
    #check if there are more than 1 stm_auto_key's for that combo
    if len(stm) > 1:
        error = 'Multiple stock lines found.'   
        
    return error,msg
	
@shared_task
def get_warehouses(quapi_id,session_id):
    error,msg = '','' 
    from polls.models import QueryApi,Warehouse
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not cr:
        return 'Quantum connection failed.',msg
        
    query = """SELECT WAREHOUSE_CODE, WHS_AUTO_KEY, DESCRIPTION
        FROM WAREHOUSE WHERE HISTORICAL = 'F'
        """
    whs = selection_dir(query,cr)
    
    whs_data = []
    for wh in whs:
        whs_data.append(Warehouse( 
            session_id = session_id,
            warehouse_code = wh[0],
            whs_auto_key = wh[1],
            name = wh[2],
            )
        )
        
    try:
        #Warehouse.objects.all().delete()
        new_statii = whs_data and Warehouse.objects.bulk_create(whs_data) or None
    except Exception as exc:
        error = "Error with creating the warehouses: %s"%exc 
        
    return error,msg
    
    
@shared_task
def send_event_notes(sel_groups):
    error,msg = '',''
    from polls.models import MailMail as mail,MailGroup as mg
    from django.conf import settings
    from django.core.mail import EmailMessage
    import ast
    email = None
    mail_groups = mg.objects.filter(id__in=sel_groups)

    for grp in mail_groups:
        
        emails = mail.objects.filter(
            status__in=['draft','failed'],
            mail_group_id=grp
        )
        body = ''
        count = 1
        
        for email in emails:
            body += '\r\n New %s Notification: %s.............................'%(count,email.subject)
            body += '................................................................\r\n'
            body += email.body
            count += 1

        if email:
            recipient_list = email.to_emails.split(",")
            bcc_list = email.cc_field.split(",")
            reply_to_list = email.from_email.split(",") 
            
            try:

                """mail = EmailMessage(
                    subject = email.subject,
                    body = body,
                    from_email = email.from_email,
                    to = recipient_list,
                    bcc = bcc_list,
                    reply_to = reply_to_list,
                    settings.EMAIL_HOST_USER,
                )"""
                
                mail = EmailMessage(
                    email.subject,
                    body, 
                    "admin@mrolive.com", 
                    recipient_list,
                    bcc_list,
                    reply_to=reply_to_list,
                )
                mail.send() 
                msg = 'Message sent!'
                
            except Exception as exc:
                error = "Error with sending email: %s"%exc
            if not error:
                for email in emails:
                    email.status='success'
                    email.save()
            
    return error,msg                                    

@shared_task
def get_uda_status(quapi_id,session_id,app='ro-management'):
    error,msg = '','' 
    from polls.models import QueryApi,StatusSelection as statsel  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not cr:
        return 'Quantum connection failed.',app
        
    
    query="""select distinct udc.attribute_value from uda_checked udc 
    inner join ro_header roh on roh.roh_auto_key = udc.auto_key 
    where udc.uda_auto_key = 11"""
    query="SELECT ATTRIBUTE_VALUE FROM UDA_VALUE_LIST WHERE UDA_AUTO_KEY = '64'"
    recs = selection_dir(query,cr)
    stat_data = []
    
    for status in recs:

        stat_data.append(statsel( 
            session_id = session_id,
            is_dashboard = True,
            name = status[0],
            )
        )
        
    try:
        del_statii = statsel.objects.filter(session_id=session_id,is_dashboard=True).delete()
        new_statii = stat_data and statsel.objects.bulk_create(stat_data) or None
    except Exception as exc:
        error = "Error with creating the statuses: %s"%exc 
    
    return error,app
       
@shared_task
def search_val_stock(quapi_id,user_id,sysur_auto_key,\
    session_id,parameters):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    
    if not cr:
        return 'Quantum connection failed.',''
        
    right_now = datetime.now()
    right_now = right_now.strftime('%m/%d/%Y %H:%M:%S')   
    where_clause = 'AND '
    loc_auto_key = parameters[0]
    rack_auto_key = parameters[1]
    
    if loc_auto_key:
  
        where_clause += """S.LOC_AUTO_KEY = %s
        """%loc_auto_key

    if rack_auto_key:

        if loc_auto_key:
            where_clause += 'AND '  
        where_clause += """S.IC_UDL_005 = %s"""%rack_auto_key
   
    query ="""SELECT DISTINCT S.STM_AUTO_KEY,S.STOCK_LINE,P.PN,P.DESCRIPTION, 
          S.SERIAL_NUMBER,L.LOCATION_CODE,S.REMARKS,CNC.CONSIGNMENT_CODE,
          PCC.CONDITION_CODE,UC.UOM_CODE,UDL.UDL_CODE,S.QTY_OH,
          CASE WHEN W.SI_NUMBER IS NOT NULL THEN W.SI_NUMBER ELSE WB.SI_NUMBER END
          FROM STOCK S
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
            LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
            LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
            LEFT JOIN WO_OPERATION WB ON WB.WOO_AUTO_KEY = WOB.WOO_AUTO_KEY
            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
            LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
            LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_005
            LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
            LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
            LEFT JOIN UOM_CODES UC ON UC.UOM_AUTO_KEY = P.UOM_AUTO_KEY
          WHERE
            S.HISTORICAL_FLAG = 'F'
            AND S.QTY_OH > 0 %s ORDER BY UDL.UDL_CODE,L.LOCATION_CODE"""%where_clause   
    
    recs = selection_dir(query,cr)
    from polls.models import WOStatus as wos
    stock_data = []
    
    for rec in recs:
      
        stock_data.append(wos(
            stm_auto_key = rec[0],
            stock_line = rec[1],
            part_number = rec[2],
            description = rec[3],
            serial_number = rec[4] or 'N/A',
            location_code = rec[5],
            remarks = rec[6],
            consignment_code = rec[7],
            condition_code = rec[8],
            uom_code = rec[9],
            cart = rec[10],
            quantity = rec[11],
            wo_number = rec[12],
            session_id = session_id,
            ))
        
    if recs and stock_data:
        try:
            delete = wos.objects.filter(session_id=session_id).delete()
            wos.objects.bulk_create(stock_data) or []   
        except Exception as exc:
            error = "Error with creating stock locally: %s"%exc
    else:
        error = 'No records found.'
        
    return error,msg

@shared_task       
def transfer_val_stock(quapi_id,user_id,sysur_auto_key,\
    session_id,parameters,stm_keys):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not cr:
        return 'Quantum connection failed.',''
    right_now = datetime.now()
    right_now = right_now.strftime('%m/%d/%Y %H:%M:%S')   
    where_clause = ''
    
    loc_auto_key = parameters[0]
    iq_enable = parameters[1]
    rack_auto_key = parameters[2]
                             

    import ast
    stm_keys = ast.literal_eval(stm_keys) 
    #call update stock rack beta
    stock_recs = get_wos_from_rack_beta(stm_keys=stm_keys,cr=cr)
                                    
                                                 
                                              
    
    msg,error,valid_whs_key,valid_wh = update_stock_rack_beta(\
   
                      
                              
                                          
        sysur_auto_key,\
        user_id,stm_keys,rack_auto_key=rack_auto_key,\
        loc_auto_key=loc_auto_key,iq_enable=iq_enable,\
        stock_recs=stock_recs,cr=cr,con=con)
                      
                                
                                          
                             
                                        
    
                         
                                   
              
                                    
                                 

                                      
                                    
    if not error:
            
        
                
                                                                                              
                                                                                                                                                                            
                                     
                                 
                         
            
                               
                  
                            
        msg = 'Successful Update.'
        
    return error,msg
    
@shared_task
def get_stock_lookup(quapi_id,session_id,parameters):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not cr:
        return 'Quantum connection failed.',''
    right_now = datetime.now()
    right_now = right_now.strftime('%m/%d/%Y %H:%M:%S')   
    where_clause = ''  
                     
    stm_auto_key = parameters[0]
    if stm_auto_key:
        where_clause += """ S.STM_AUTO_KEY = %s"""%stm_auto_key
        
    wo_number = parameters[1]
    if wo_number:
        if where_clause:
            where_clause += ' AND'
        where_clause += """ (UPPER(W.SI_NUMBER) = UPPER('%s')
            OR UPPER(WO.SI_NUMBER) = UPPER('%s')
            OR UPPER(BWO.SI_NUMBER) = UPPER('%s'))"""%(wo_number,wo_number,wo_number)
       
    part_number = parameters[2]
    if part_number:
        if where_clause:
            where_clause += ' AND'
        where_clause += " UPPER(P.PN) = UPPER('%s')"%part_number
        
    user_name = parameters[3]
    if user_name:
        if where_clause:
            where_clause += ' AND'
        sys_clause = """(SELECT SYSUR_AUTO_KEY FROM SYS_USERS
        WHERE UPPER(USER_NAME) = UPPER('%s')
        OR UPPER(USER_ID) = UPPER('%s')
        OR UPPER(EMPLOYEE_CODE) = UPPER('%s'))
        """%(user_name,user_name,user_name)
        
        where_clause += """ S.SYSUR_AUTO_KEY = %s"""%sys_clause
        #--OR W.SYSUR_AUTO_KEY = %s
        #--OR SR.SYSUR_AUTO_KEY = %s)
        #%(sys_clause,sys_clause,sys_clause)
        
    query = """SELECT P.PN,P.DESCRIPTION,S.SERIAL_NUMBER,
        L.LOCATION_CODE,S.STOCK_LINE,PCC.CONDITION_CODE,
        P.UOM_AUTO_KEY,CNC.CONSIGNMENT_CODE,S.QTY_OH,
        SR.STM_AUTO_KEY,S.STM_AUTO_KEY,TO_CHAR(S.NOTES),
        S.CTRL_NUMBER,S.CTRL_ID,
        CASE WHEN W.SI_NUMBER IS NOT NULL THEN W.SI_NUMBER ELSE 
        (CASE WHEN WO.SI_NUMBER IS NOT NULL THEN WO.SI_NUMBER
        ELSE BWO.SI_NUMBER END) END,
        S.STM_ORIGINAL
        FROM STOCK S
        JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
        LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
        LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
        LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
        LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
        LEFT JOIN STOCK_TI STI ON STI.STM_AUTO_KEY = S.STM_AUTO_KEY
        LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = STI.WOB_AUTO_KEY
        LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOB.WOO_AUTO_KEY
        LEFT JOIN WO_BOM BWOB ON BWOB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
        LEFT JOIN WO_OPERATION BWO ON BWO.WOO_AUTO_KEY = BWOB.WOO_AUTO_KEY
        WHERE%s
        """%where_clause
  
    recs = selection_dir(query,cr)    
    
    from polls.models import WOStatus as wos
    stock_data = []
    stms = []
    
    for rec in recs:
    
        if rec[10] not in stms:
                                
            stms.append(rec[10])
            esn = rec[2]
            if rec[15]:
                query = """SELECT SERIAL_NUMBER FROM 
                STOCK WHERE STM_AUTO_KEY = %s"""%rec[15]
                esn = selection_dir(query,cr)
                esn = esn and esn[0] and esn[0][0] or rec[2]
                
            stock_data.append(wos(
                part_number = rec[0],
                description = rec[1],
                serial_number = rec[2] or 'N/A',
                location_code = rec[3],
                stock_line = rec[4],
                condition_code = rec[5],
                slug = esn, #esn
                consignment_code = rec[7],
                qty_oh = rec[8],
                cart = rec[9] and 'green-check.png'\
                    or 'blank.png', #reserved
                stm_auto_key = rec[10],
                notes = len(rec[11])>75 and rec[11][:75] + '...' or rec[11],
                ctrl_number = rec[12],
                ctrl_id = rec[13],
                wo_number = rec[14],
                quantity = rec[8],
                session_id = session_id,
            ))
            
        else:
            continue
        
    if recs and stock_data:
        try:
            delete = wos.objects.filter(session_id=session_id).delete()
            wos.objects.bulk_create(stock_data) or []   
        except Exception as exc:
            error = "Error with creating stock locally: %s"%exc
    else:
        error = 'No records found.'
        
    return error,msg
   
@shared_task
def update_task_mgmt(quapi_id,session_id,sysur_auto_key,parameters,plain_wot_list):
    error,msg = '','' 
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not cr:
        return 'Quantum connection failed.',''
    right_now = datetime.now()
    right_now = right_now.strftime('%m/%d/%Y %H:%M:%S')        
    status = parameters[0]
    sequence = parameters[1]
    description = parameters[2]
    work_required = parameters[3]
    action_taken = parameters[4]
    status_type = parameters[5]    
    wot_list = construct_akl(plain_wot_list)
    
    upd_clause=''
    
    if status:
        upd_clause += "WOT.WOS_AUTO_KEY = %s"%status    
        
        for wot_auto_key in plain_wot_list:
            query="SELECT WOS_AUTO_KEY FROM WO_TASK WHERE WOT_AUTO_KEY = %s"%wot_auto_key
            wos_p = selection_dir(query,cr)
            wos_previous = wos_p and wos_p[0] and wos_p[0][0]
            query = """INSERT INTO WO_TASK_STATUS (SYSUR_AUTO_KEY,
                WOT_AUTO_KEY,WOS_AUTO_KEY,ENTRY_DATE,SYSTEM_DATE,WOS_PREVIOUS)                               
                VALUES(%s,%s,%s,TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),TO_DATE('%s','mm/dd/yyyy'),'%s')
                """%(sysur_auto_key,wot_auto_key,status,right_now,right_now[:10],wos_previous)
            error += insertion_dir(query,cr)        
        
    if sequence:
        if status:
            upd_clause += ', '
        upd_clause += 'WOT.SEQUENCE = %s'%sequence
        
    if action_taken:
        if status or sequence:
            upd_clause += ', '
        upd_clause += "WOT.CORRECTIVE_ACTION = '%s'"%action_taken
        
    if work_required:
        if status or sequence or action_taken:
            upd_clause += ', '
            
        import re    
        for char in work_required.split("\n"):
            work_required = re.sub(r"[^a-zA-Z0-9 .]+", ' ', char)
            
        upd_clause += """WOT.SQUAWK_DESC = '%s', 
            WOT.DESCREPANCY_TEXT = '%s'"""%(work_required,work_required)
    
    for wots in wot_list:    
        query="""UPDATE WO_TASK WOT SET %s 
            WHERE WOT.WOT_AUTO_KEY IN %s"""%(upd_clause,wots)
        error = updation_dir(query,cr)
    
    if error in ['{"recs": ""}','']:
        error = ''
        msg = 'Successful update.'
        orcl_commit(con=con)
    
    return error,msg
      
@shared_task
def get_task_mgmt(quapi_id,session_id,parameters,wot_list = []):
    error,msg = '','' 
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not cr:
        return 'Quantum connection failed.',''
    status = parameters[0]
    sequence = parameters[1]
    description = parameters[2]
    work_required = parameters[3]
    action_taken = parameters[4]
    status_type = parameters[5] 
    wo_number = parameters[6]    
    wh_clause=''
    
    if wot_list:
        wot_lists = construct_akl(wot_list)
        for wots in wot_lists:
            wh_clause += " AND WOT.WOT_AUTO_KEY IN %s"%wots
    
    if status:
        wh_clause += " AND WOT.WOS_AUTO_KEY = %s"%status
        
    if sequence:
        wh_clause += " AND WOT.SEQUENCE = '%s'"%sequence
        
    if description:
        wh_clause += """ AND (UPPER(WOT.SQUAWK_DESC) LIKE UPPER('%s') 
            OR UPPER(WTM.DESCRIPTION) = UPPER('%s'))"""%(description,description)
      
    if action_taken:
        wh_clause += " AND TO_CHAR(WOT.CORRECTIVE_ACTION) = '%s'"%action_taken
        
    if work_required:
        wh_clause += " AND WOT.SQUAWK_DESC = '%s'"%work_required
        
    if status_type:
        wh_clause += " AND WOS.STATUS_TYPE = '%s'"%status_type
        
    if wo_number:
        wh_clause += " AND WOO.SI_NUMBER = '%s'"%wo_number

    task_query = """SELECT WOT.SEQUENCE, WTM.DESCRIPTION,
        WOT.WOT_AUTO_KEY, WOO.SI_NUMBER, WOS.DESCRIPTION,
        WOS.STATUS_TYPE,
        WOT.SQUAWK_DESC, WOT.CORRECTIVE_ACTION,
        CMP.COMPANY_NAME, WOO.WO_UDF_001, WOO.WO_UDF_002, WOO.WO_UDF_003,
        PNM.PN, P.PN, S.SERIAL_NUMBER     
        FROM WO_TASK WOT
        JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WOT.WTM_AUTO_KEY
        JOIN WO_OPERATION WOO ON WOT.WOO_AUTO_KEY = WOO.WOO_AUTO_KEY
        LEFT JOIN WO_STATUS WOS ON WOS.WOS_AUTO_KEY = WOT.WOS_AUTO_KEY
        LEFT JOIN COMPANIES CMP ON CMP.CMP_AUTO_KEY = WOO.CMP_AUTO_KEY
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY = WOO.WOO_AUTO_KEY
        LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
        LEFT JOIN PARTS_MASTER PNM ON PNM.PNM_AUTO_KEY = S.PNM_AUTO_KEY
        LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = WOO.PNM_AUTO_KEY
        WHERE WOT.WOT_AUTO_KEY IS NOT NULL %s
        ORDER BY WOT.SEQUENCE
    """%(wh_clause)  
    
    recs = selection_dir(task_query,cr)
    if recs:                                       
        res = create_mgmt_tasks(session_id,recs)
    else:
        error = 'No tasks found.'
    return error,msg
   
@shared_task
def get_task_statuses(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not cr:
        return 'Quantum connection failed.',''
        
    query = """SELECT DESCRIPTION,WOS_AUTO_KEY,WOS_AUTO_KEY  
        FROM WO_STATUS ORDER BY DESCRIPTION"""
    statuses = selection_dir(query,cr)
    status_data = []
    from polls.models import StatusSelection as stat
    for status in statuses:
        status_data.append(stat(
            name = status[0],
            wos_auto_key = status[1],
            text_wos = str(status[2]),
            session_id = session_id,
            is_dashboard = False,
        ))
        
    try:
        delete = stat.objects.filter(session_id=session_id).delete()
        stat.objects.bulk_create(status_data) or []   
    except Exception as exc:
        error = "Error with creating status types locally: %s"%exc    
        
    query = """SELECT DISTINCT STATUS_TYPE 
        FROM WO_STATUS ORDER BY STATUS_TYPE"""
    status_types = selection_dir(query,cr)
    status_data = []
    for status_type in status_types:
        status_data.append(stat(
            name = status_type[0],
            session_id = session_id,
            is_dashboard = True,
        ))  
     
    try:
        stat.objects.bulk_create(status_data) or []   
    except Exception as exc:
        error = "Error with creating statuses locally: %s"%exc    
    return error

@shared_task
def create_open_labor(session_id,timestamp,recs):
    error = ''
    wtl_data = []
    from polls.models import TaskLabor

    for rec in recs:
        hours = rec[6] 
        if rec[1]:
            start_time = datetime.strptime(rec[1],'%Y-%m-%d %H:%M:%S')
            start_time = start_time + timedelta(hours=1)        
            if not hours:
                diff = timestamp - start_time
                days, seconds = diff.days, diff.seconds
                hours = round(days * 24 + seconds / 3600,2)            
            
        wtl_data.append(TaskLabor(
            session_id = session_id,
            wo_number = rec[2],
            sequence = rec[3],
            task_desc = rec[4], 
            hours = hours,
            user_name = rec[0],
            entry_date = rec[7] and rec[7][:10] or None,
            start_time = start_time,
            ))
        
    try:
        delete = TaskLabor.objects.filter(session_id=session_id).delete()
        TaskLabor.objects.bulk_create(wtl_data) or []   
    except Exception as exc:
        error = "Error with creating open labor locally: %s"%exc 
        
    return error
    
@shared_task
def open_labor(quapi_id,session_id):
    error,msg = '',''
    
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not cr:
        return 'Quantum connection failed.',''

    date_format = '%Y-%m-%d %H:%M:%S'
    today = datetime.now()
    timestamp = today and today.strftime(date_format)    
    #MTU: makes it CST    
    timestamp = today - timedelta(hours=1)        
    """time_query = "SELECT SYSTIMESTAMP FROM DUAL"
    today = selection_dir(time_query,cr)
    today = today and today[0] and today[0][0] and today[0][0][:18] or None
    timestamp = today and datetime.strptime(today,'%Y-%m-%d %H:%M:%S')"""

    query = """SELECT SYS.USER_NAME,WTL.START_TIME,WO.SI_NUMBER,
        WOT.SEQUENCE,WTM.DESCRIPTION,WTL.STOP_TIME,WTL.HOURS,WTL.ENTRY_DATE FROM WO_TASK_LABOR WTL
        LEFT JOIN WO_TASK WOT ON WOT.WOT_AUTO_KEY = WTL.WOT_AUTO_KEY
        LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
        LEFT JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WOT.WTM_AUTO_KEY
        LEFT JOIN SYS_USERS SYS ON SYS.SYSUR_AUTO_KEY = WTL.SYSUR_AUTO_KEY
        WHERE WTL.STOP_TIME IS NULL
        ORDER BY SYS.USER_NAME ASC
        """
    recs = selection_dir(query,cr)
    
    if recs:
        error = create_open_labor(session_id,timestamp,recs)
    else:
        error = 'No open labor found.'

    return error,msg
    
@shared_task
def get_tag_types(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not cr:
        return 'Quantum connection failed.',''
        
    query = """SELECT TAG_TYPE_CODE,TTP_AUTO_KEY FROM TAG_TYPE"""
    tag_types = selection_dir(query,cr)
    tag_data = []
    from polls.models import StockCart as stk_cart
    for tag_type in tag_types:
        tag_data.append(stk_cart(
            name = tag_type[0],
            udl_auto_key = tag_type[1],
            session_id = session_id,
        ))                   
    try:
        delete = stk_cart.objects.filter(session_id=session_id).delete()
        stk_cart.objects.bulk_create(tag_data) or []   
    except Exception as exc:
        error = "Error with creating tag types locally: %s"%exc    
    return error
        
@shared_task
def get_certs(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not cr:
        return 'Quantum connection failed.',''
        
    query = """SELECT CERT_SOURCE FROM CERT_SOURCE"""
    cmps = selection_dir(query,cr)
    cmp_data = []
    from polls.models import Departments as dept_obj
    for cmp in cmps:
        cmp_data.append(dept_obj(
            name = cmp[0],
            session_id = session_id,
        ))                   
    try:
        delete = dept_obj.objects.filter(session_id=session_id).delete()
        dept_obj.objects.bulk_create(cmp_data) or []   
    except Exception as exc:
        error = "Error with creating cert codes locally: %s"%exc    
    return error  
    
@shared_task
def get_companies(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not cr:
        return 'Quantum connection failed.',''
        
    query = """SELECT CMP_AUTO_KEY,COMPANY_NAME FROM COMPANIES"""
    cmps = selection_dir(query,cr)
    cmp_data = []
    from polls.models import Companies as cmp_obj
    for cmp in cmps:
        cmp_data.append(cmp_obj(
            cmp_auto_key = cmp[0],
            name = cmp[1],
            is_customer = True,
            session_id = session_id,
        ))                   
    try:
        delete = cmp_obj.objects.filter(session_id=session_id).delete()
        cmp_obj.objects.bulk_create(cmp_data) or []   
    except Exception as exc:
        error = "Error with creating companies locally: %s"%exc    
    return error        
    
@shared_task
def update_accept_lots(quapi_id,user_id,session_id,\
    sysur_auto_key,stm_keys,filters,parameters,\
    accept_only=False):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not cr:
        return 'Quantum connection failed.',''
        
    stm_lists = construct_akl(stm_keys)
    lot_number = filters[0]
    qty_need_sub = """(SELECT WB.QTY_NEEDED FROM WO_BOM WB,STOCK S,STOCK_TI STI WHERE WB.WOO_AUTO_KEY IN (SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER = '%s') AND S.STM_AUTO_KEY = STI.STM_AUTO_KEY AND STI.WOB_AUTO_KEY = WB.WOB_AUTO_KEY AND ROWNUM <= 1)"""%lot_number         

    hold_line_false = ""
    if accept_only:
        hold_line_false = ", HOLD_LINE = 'F'"
     
    for stms in stm_lists:    
        query = """UPDATE STOCK SET QTY_REC_FROM_LOT = %s%s WHERE STM_AUTO_KEY IN %s"""%(qty_need_sub,hold_line_false,stms) 
        error = updation_dir(query,cr)
        if error != '{"recs": ""}': 
            return error,msg
        
    for stm_auto_key in stm_keys:
        squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
        (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],stm_auto_key)
        error = updation_dir(squery,cr) 
        if error != '{"recs": ""}': 
            return error,msg
      
    if accept_only:
        aud_status = 'failure'
        if not error or error == '{"recs": ""}':
            error = ''
            orcl_commit(con=con)     
            aud_status = 'success'
            msg = 'Successful update.'
            error,filter_msg = get_lots(quapi_id,session_id,filters) 
            right_now = datetime.now()
            right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')                   
            from polls.models import MLApps as maps,QuantumUser as qu
            app_id = maps.objects.filter(code='lot-inpsection')   
            user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
            user_rec = user_rec and user_rec[0] or None
            if user_rec:
                new_val = 'Update to stock with an acceptance - updated qty_rec_from_lot on stock.'
                field_changed = new_val
                field_changed += error
                error += register_audit_trail(user_rec,field_changed,new_val,right_now,app_id,quapi,status=aud_status)  
            else:
                error = 'Incorrect Quantum User ID.'            
    return error,msg
    
@shared_task
def get_locations(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not cr:
        return 'Quantum connection failed.'
    loc_query = """SELECT LOCATION_CODE,
        DESCRIPTION,LOC_AUTO_KEY 
        FROM LOCATION 
        ORDER BY LOCATION_CODE ASC"""
    locations = selection_dir(loc_query,cr)
    loc_data = []
    from polls.models import Location as loca
    for loc in locations:
        loc_data.append(loca(
            location_code = loc[0],
            name = loc[1],
            loc_auto_key = loc[2],
            session_id = session_id,
        ))                   
    try:
        delete = loca.objects.filter(session_id=session_id).delete()
        loca.objects.bulk_create(loc_data) or []   
    except Exception as exc:
        error = "Error with creating locations locally: %s"%exc    
    return error
    
@shared_task
def get_categories(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not cr:
        return 'Quantum connection failed.'
    cat_query = """SELECT RO_CATEGORY_CODE,
        RCT_AUTO_KEY
        FROM RO_CATEGORY_CODES
        ORDER BY SEQUENCE"""
    cats = selection_dir(cat_query,cr)
    cat_data = []
    from polls.models import Categories
    for cat in cats:
        cat_data.append(Categories(
            categ_code = cat[0],
            rct_auto_key = cat[1],
            session_id = session_id,
        ))                   
    try:
        delete = Categories.objects.filter(session_id=session_id).delete()
        Categories.objects.bulk_create(cat_data) or []   
    except Exception as exc:
        error = "Error with creating categories: %s"%exc    
    return error
    
@shared_task
def get_conditions(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not cr:
        return 'Quantum connection failed.'
    cond_query = """SELECT DISTINCT CONDITION_CODE
        FROM PART_CONDITION_CODES WHERE HISTORICAL = 'F'"""
    conditions = selection_dir(cond_query,cr)
    cond_data = []
    from polls.models import PartConditions as part_cond
    for cond in conditions:
        cond_data.append(part_cond(
            condition_code = cond[0],
            session_id = session_id,
        ))                   
    try:
        delete = part_cond.objects.filter(session_id=session_id).delete()
        part_cond.objects.bulk_create(cond_data) or []   
    except Exception as exc:
        error = "Error with creating conditions locally: %s"%exc    
    return error 

@shared_task
def get_stock_status(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not cr:
        return 'Quantum connection failed.'
    query = """select udl_auto_key,udl_code,
        udl_description from user_defined_lookups 
        where UPPER(udl_column_name) = UPPER('ic_udl_004')
        """
    statuses = selection_dir(query,cr)
    stat_data = []
    from polls.models import StatusSelection as stat_sel
    for stat in statuses:
        stat_data.append(stat_sel(
            wos_auto_key = stat[0],
            name = stat[1],
            severity = stat[2],
            session_id = session_id,
        ))                   
    try:
        delete = stat_sel.objects.filter(session_id=session_id).delete()
        stat_sel.objects.bulk_create(stat_data) or []   
    except Exception as exc:
        error = "Error with creating stock statuses locally: %s"%exc    
    return error    
    
def create_bulk_lots(session_id,recs,is_dashboard=False):
    from polls.models import WOStatus as wos
    lot_data = []
    error = ''
    for rec in recs:
        lot_data.append(wos(
            wo_number = rec[0],
            part_number = rec[1],
            description = rec[2],
            serial_number = rec[3],
            status = rec[4],#stock status (IC_UDL_004) 
            qty_needed = rec[5] or 0,#bom qty needed
            qty_oh = rec[6] or 0,#qty_rec_from_lot
            condition_code = rec[7],#bom cond code
            cond_level_zero = rec[8],#part cond code
            notes = rec[9],
            stock_line = rec[10],
            stm_auto_key = rec[11],
            rack = rec[12],
            account_company = rec[13]=='F' and "green-check.png" or 'blank.png',
            next_num  = rec[14],
            consignment_code = rec[15],
            loc_validated_date = rec[16] or None,#S.EXP_DATE
            supdate_msg = rec[17],#S.REMARKS
            session_id = session_id,
            is_dashboard = is_dashboard,
        ))                   
    try:
        delete_lots = wos.objects.filter(session_id=session_id,is_dashboard=is_dashboard).delete()
        lots = lot_data and wos.objects.bulk_create(lot_data) or [] 
    except Exception as exc:
        error = "Error with creating lots locally: %s"%exc 
    return error
    
@shared_task
def get_lots(quapi_id,session_id,filters): 
    error,msg = '','' 
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)    
    lot_number = filters[0]
    part_number = filters[1]
    description = filters[2]
    serial_no = filters[3]
    stock_status = filters[4]
    consignment = filters[5]
    where_clause = " WHERE S.HISTORICAL_FLAG <> 'T' AND S.QTY_OH > 0"   

    if lot_number:
        where_clause += """ AND S.STM_LOT IN (SELECT STM_AUTO_KEY FROM STOCK_RESERVATIONS 
        WHERE WOO_AUTO_KEY IN (SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER = '%s'))
        AND (WB.WOO_AUTO_KEY IN 
        (SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER = '%s'))"""%(lot_number,lot_number)
        
    if part_number:       
        where_clause += " AND UPPER(P.PN) LIKE UPPER('%s%s')"%(part_number,'%')
    if description: 
        where_clause += " AND UPPER(P.DESCRIPTION) LIKE UPPER('%s%s')"%(description,'%')
    if serial_no:       
        where_clause += " AND WB.EXPECTED_SN LIKE UPPER('%s%s')"%(serial_no,'%s')
    if stock_status:       
        where_clause += " AND S.IC_UDL_004='%s'"%stock_status
    if consignment:
        where_clause += """ AND S.CNC_AUTO_KEY=(SELECT CNC_AUTO_KEY FROM
        CONSIGNMENT_CODES WHERE UPPER(CONSIGNMENT_CODE) = UPPER('%s'))"""%consignment 
        
    query = """SELECT      
        W.SI_NUMBER,
        P.PN,
        P.DESCRIPTION,
        WB.EXPECTED_SN,        
        UDL.UDL_CODE,        
        WB.QTY_NEEDED, 
        S.QTY_OH,
        PCC.CONDITION_CODE,
        SPC.CONDITION_CODE,
        TO_CHAR(S.NOTES),
        S.STOCK_LINE,
        S.STM_AUTO_KEY,
        S.SERIAL_NUMBER,
        S.HOLD_LINE,
        S.SERIES_ID,
        CC.CONSIGNMENT_CODE,
        S.EXP_DATE,
        TO_CHAR(S.REMARKS)
        FROM WO_BOM WB
        JOIN WO_OPERATION W ON WB.WOO_AUTO_KEY = W.WOO_AUTO_KEY
        JOIN STOCK_TI STI ON STI.WOB_AUTO_KEY = WB.WOB_AUTO_KEY
        JOIN STOCK S ON S.STM_AUTO_KEY = STI.STM_AUTO_KEY
        LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = WB.PCC_AUTO_KEY
        LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_004
        LEFT JOIN PART_CONDITION_CODES SPC ON SPC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
        LEFT JOIN CONSIGNMENT_CODES CC ON CC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
        JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY%s
        """%(where_clause)
    recs = selection_dir(query,cr)
    #if the stock is not reserved to the woo, then we need to look for the rod
    """
    need to make the same adjustment for the import. 
    if the main component isn't reserved to the WOO it will be on an ROD. 
    if we don't have the woo when we import we don't assign the stm_lot 
    which is what links all stock to the lot they are part of"""
    
    if not recs and lot_number:       
        where_clause = """ WHERE S.STM_LOT IN (SELECT STM_AUTO_KEY FROM STOCK_RESERVATIONS 
        WHERE ROD_AUTO_KEY IN (SELECT ROD_AUTO_KEY FROM RO_DETAIL WHERE WOO_AUTO_KEY IN 
        (SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER = '%s')))"""%lot_number
        
        if part_number:       
            where_clause += " AND UPPER(P.PN) LIKE UPPER('%s%s')"%(part_number,'%')
        if description: 
            where_clause += " AND UPPER(P.DESCRIPTION) LIKE UPPER('%s%s')"%(description,'%')
        if serial_no:       
            where_clause += " AND WB.EXPECTED_SN LIKE UPPER('%s%s')"%(serial_no,'%s')
        if stock_status:       
            where_clause += " AND S.IC_UDL_004='%s'"%stock_status
        
        query = """SELECT      
            W.SI_NUMBER,
            P.PN,
            P.DESCRIPTION,
            WB.EXPECTED_SN,        
            UDL.UDL_CODE,        
            WB.QTY_NEEDED, 
            S.QTY_OH,
            PCC.CONDITION_CODE,
            SPC.CONDITION_CODE,
            TO_CHAR(S.NOTES),
            S.STOCK_LINE,
            S.STM_AUTO_KEY,
            S.SERIAL_NUMBER,
            S.HOLD_LINE,
            S.SERIES_ID,
            CC.CONSIGNMENT_CODE,
            S.EXP_DATE,
            TO_CHAR(S.REMARKS)
            FROM WO_BOM WB
                LEFT JOIN STOCK_TI STI ON STI.WOB_AUTO_KEY = WB.WOB_AUTO_KEY
                LEFT JOIN STOCK S ON S.STM_AUTO_KEY = STI.STM_AUTO_KEY
                LEFT JOIN WO_OPERATION W ON WB.WOO_AUTO_KEY = W.WOO_AUTO_KEY
                LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = WB.PCC_AUTO_KEY
                LEFT JOIN PART_CONDITION_CODES SPC ON SPC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
                LEFT JOIN CONSIGNMENT_CODES CC ON CC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
                LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_004
                JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY%s 
            """%(where_clause) 
        recs = selection_dir(query,cr)
        
    where_clause = ''    
    if lot_number:
        where_clause = """ WHERE S.STM_LOT IN (SELECT STM_AUTO_KEY FROM STOCK_RESERVATIONS 
        WHERE WOO_AUTO_KEY IN (SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER = '%s'))"""%(lot_number)
        
        query = """SELECT      
            '',
            P.PN,
            P.DESCRIPTION,
            '',        
            UDL.UDL_CODE,        
            '', 
            S.QTY_OH,
            '',
            SPC.CONDITION_CODE,
            TO_CHAR(S.NOTES),
            S.STOCK_LINE,
            S.STM_AUTO_KEY,
            S.SERIAL_NUMBER,
            S.HOLD_LINE,
            S.SERIES_ID,
            CC.CONSIGNMENT_CODE,
            S.EXP_DATE,
            TO_CHAR(S.REMARKS)
            FROM STOCK S
            LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_004
            LEFT JOIN PART_CONDITION_CODES SPC ON SPC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
            LEFT JOIN CONSIGNMENT_CODES CC ON CC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
            JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY%s
            """%(where_clause)
        recs += selection_dir(query,cr) 
        
    add_recs = []
    stm_keys = []
    final_recs = []
    series_id = ''
    for srec in recs:
        #loop over all records and get all stock
        #records that have the same series_number 
        #but different series_ids
        if srec[11] in stm_keys:
            #we found a duplicate
            continue
        else:
            stm_keys.append(srec[11])
            final_recs.append(srec)      
            series_id = srec[14]
            
    error = create_bulk_lots(session_id,final_recs,is_dashboard=True)
    return error,msg
    
def get_stock_import(cr,stm_auto_key):
    query="""SELECT MFG_AUTO_KEY,
                              RECEIVER_NUMBER,
                              PCC_AUTO_KEY,
                              LOC_AUTO_KEY,
                              WHS_AUTO_KEY,
                              CNC_AUTO_KEY,
                              IFC_AUTO_KEY,
                              STC_AUTO_KEY,
                              REMARKS,
                              IC_UDF_005,
                              IC_UDF_006,
                              IC_UDF_007,
                              IC_UDF_008,
                              IC_UDF_009,
                              IC_UDF_010,
                              UNIT_COST,
                              UNIT_PRICE,
                              OVHL_COST,
                              TAGGED_BY,
                              IQ_ENABLE,
                              SHELF_LIFE,
                              TAG_DATE,
                              EXP_DATE,
                              TO_CHAR(NOTES),
                              OWNER,
                              PART_CERT_NUMBER,
                              ORIGINAL_PO_NUMBER,
                              SERIAL_NUMBER,
                              CTS_AUTO_KEY,
                              CMP_AUTO_KEY,
                              REC_DATE,
                              VISIBLE_MKT,
                              ALT_ID,
                              MFG_LOT_NUM,
                              AIRWAY_BILL,
                              ORDER_REC_DATE,
                              INSPECT_DUE_DATE,
                              CALIB_REMARKS,
                              CALIB_REF_MASTER,
                              CALIB_REF_INSTR,
                              CMP_CALIB_BY,
                              TAG_NUMBER,
                              HOLD_LINE,
                              REASON_FOR_HOLD,
                              QTY_OH
                              FROM STOCK WHERE STM_AUTO_KEY = %s"""%stm_auto_key
    rec = selection_dir(query,cr)
    return rec
    
def qry_pn_transfer(cr,sysur_auto_key,user_id,stm_auto_key,pnm_auto_key,quantity,factor=1):
    error = ''
    stock_rec = get_stock_import(cr,stm_auto_key)
    stock_rec = stock_rec and stock_rec[0]    
    qty_oh = stock_rec[44] and float(stock_rec[44]) or 0
    qty_update = quantity or qty_oh or 0
    if stock_rec:
        query = """DECLARE CT qc_utl_pkg.cursor_type; 
                    BEGIN CT := QC_IC_PKG2.spi_stock_pn_transfer(
                      %s,       --p_stm NUMBER,
                      %s,       --p_new_pnm   NUMBER,
                      %s,       --p_new_qty   NUMBER,
                      %s       --p_factor    NUMBER,
                    );
                    close CT; END;
                    """%(stm_auto_key,pnm_auto_key,qty_update,factor)
                   
        error = updation_dir(query,cr)
        #update the user on the SA_LOG 
        squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
            (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],stm_auto_key)
        error = updation_dir(squery,cr) 
        
        squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
            (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],stm_auto_key)
        error = updation_dir(squery,cr)
        
    else:
        error = 'No stock record found.' 
                
    return error
    
def qry_stock_import(cr,sysur_auto_key,user_id,stm_auto_key,pnm_auto_key,qty_update):

    error = ''
    stock_rec = get_stock_import(cr,stm_auto_key)
    stock_rec = stock_rec and stock_rec[0]
    
    if stock_rec:
        mfg_auto_key = stock_rec[0]
        receiver_number = stock_rec[1]
        pcc_auto_key = stock_rec[2]
        loc_auto_key = stock_rec[3]
        whs_auto_key = stock_rec[4]
        cnc_auto_key = stock_rec[5]
        ifc_auto_key = stock_rec[6]
        stc_auto_key = stock_rec[7]
        remarks = stock_rec[8]
        ic_udf_005 = stock_rec[9]
        ic_udf_006 = stock_rec[10]
        ic_udf_007 = stock_rec[11]
        ic_udf_008 = stock_rec[12]
        ic_udf_009 = stock_rec[13]
        ic_udf_010 = stock_rec[14]
        unit_cost = stock_rec[15] and float(stock_rec[15]) or 0
        unit_price = stock_rec[16] and float(stock_rec[16]) or 0
        ovhl_cost = stock_rec[17] and float(stock_rec[17]) or 0
        tagged_by = stock_rec[18]
        iq_enable = stock_rec[19]
        shelf_life = stock_rec[20]
        tag_date = stock_rec[21]
        exp_date = stock_rec[22]
        notes = stock_rec[23]
        owner = stock_rec[24]
        part_cert_number = stock_rec[25]
        original_po_number = stock_rec[26]
        serial_number = stock_rec[27]
        cts_auto_key = stock_rec[28]
        cmp_auto_key = stock_rec[29]
        rec_date = stock_rec[30]
        visible_mkt = stock_rec[31]
        alt_id = stock_rec[32]
        mfg_lot_num = stock_rec[33]
        airway_bill = stock_rec[34]
        order_rec_date = stock_rec[35]
        inspect_due_date = stock_rec[36]
        calib_remarks = stock_rec[37]
        calib_ref_master = stock_rec[38]
        calib_ref_instr = stock_rec[39]
        cmp_calib_by = stock_rec[40]
        tag_number = stock_rec[41]
        hold_line = stock_rec[42]
        reason_for_hold = stock_rec[43]
        qty_oh = stock_rec[44] and float(stock_rec[44]) or 0
        
        query = """DECLARE CT qc_utl_pkg.cursor_type; 
                    BEGIN CT := 
                    QC_STOCK_PKG.SPI_IMPORT_STM(%s, --pnm_auto_key
                                  'F', --P_CREATE_PN           VARCHAR2,
                                  'F', --P_CREATE_CC           VARCHAR2,
                                  'F', --P_CREATE_LOC          VARCHAR2,
                                  'F', --P_CREATE_WHS          VARCHAR2,
                                  'F', --P_CREATE_CNC          VARCHAR2,
                                  'F', --P_CREATE_ILS          VARCHAR2,
                                  'F', --P_CREATE_CERT         VARCHAR2,
                                  '%s', --P_MFG_AUTO_KEY        NUMBER, 0
                                  '', --P_PN                  VARCHAR2,
                                  '', --P_DESCRIPTION         VARCHAR2,
                                  %s, --P_QTY_OH              NUMBER, qty_update or 44
                                  '%s', --P_RECEIVER_NUMBER     VARCHAR2, 1
                                  '%s', --P_PCC_AUTO_KEY        NUMBER, 2
                                  '%s', --P_LOC_AUTO_KEY        NUMBER, 3
                                  '%s', --P_WHS_AUTO_KEY        NUMBER, 4
                                  '%s', --P_CNC_AUTO_KEY        NUMBER, 5
                                  '%s', --P_IFC_AUTO_KEY        NUMBER, 6
                                  '%s', --P_STC_AUTO_KEY        NUMBER, 7
                                  '%s', --P_REMARKS             VARCHAR2, 8
                                  '%s', --P_IC_UDF_005          VARCHAR2, 9
                                  '%s', --P_IC_UDF_006          VARCHAR2, 10
                                  '%s', --P_IC_UDF_007          VARCHAR2, 11
                                  '%s', --P_IC_UDF_008          VARCHAR2, 12
                                  '%s', --P_IC_UDF_009          VARCHAR2, 13
                                  '%s', --P_IC_UDF_010          VARCHAR2, 14
                                  %s, --P_UNIT_COST           NUMBER, 15
                                  %s, --P_UNIT_PRICE          NUMBER, 16
                                  %s, --P_OVHL_COST           NUMBER, 17
                                  '%s', --P_TAGGED_BY           VARCHAR2, 18
                                  '%s', --P_IQ_ENABLE           VARCHAR2, 19
                                  '%s', --P_SHELF_LIFE          VARCHAR2, 20
                                  TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss'), --P_TAG_DATE            DATE, 21
                                  TO_DATE('%s','yyyy-mm-dd'), --P_EXP_DATE            DATE, 22
                                  '%s', --P_NOTES               CLOB, 23
                                  '%s', --P_OWNER               VARCHAR2, 24
                                  '%s', --P_PART_CERT_NUMBER    VARCHAR2, 25
                                  '%s', --P_ORIGINAL_PO_NUMBER  VARCHAR2, 26
                                  '%s', --P_SERIAL_NUMBER       VARCHAR2, 27
                                  '%s', --P_CTS_AUTO_KEY        NUMBER, 28
                                  '%s', --P_CMP_AUTO_KEY        NUMBER, 29
                                  TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss'), --P_REC_DATE            DATE, 30
                                  '%s', --P_VISIBLE_MKT         VARCHAR2, 31
                                  '', --P_CONDITION_CODE      VARCHAR2, 
                                  '', --P_LOC_CODE            VARCHAR2,
                                  '', --P_WHS_CODE            VARCHAR2,
                                  '', --P_CONSIGNMENT_CODE    VARCHAR2,
                                  '', --P_ILS_FLAG_CODE       VARCHAR2,
                                  '', --P_CERT_SOURCE         VARCHAR2,
                                  '', --P_SERIALIZED          VARCHAR2,
                                  '', --P_FIXED_ASSET         VARCHAR2,
                                  '%s', --P_ALT_ID              varchar2, 32
                                  '%s', --P_MFG_LOT_NUM         VARCHAR2, 33
                                  '%s', --P_AIRWAY_BILL         VARCHAR2, 34
                                  '%s', --P_ORDER_REC_DATE      DATE, 35
                                  '%s', --P_INSPECT_DUE_DATE    DATE, 36
                                  '%s', --P_CALIB_REMARKS       VARCHAR2, 37
                                  '%s', --P_CALIB_REF_MASTER    VARCHAR2, 38
                                  '%s', --P_CALIB_REF_INSTR     VARCHAR2, 39
                                  '%s', --P_CMP_CALIB_BY        NUMBER, 40
                                  '%s', --P_TAG_NUMBER          VARCHAR2, 41
                                  '%s', --P_HOLD_LINE           VARCHAR2, 42
                                  '%s', --P_REASON_FOR_HOLD     VARCHAR2 43
                                  '',
                                  '',
                                  '',
                                  '',
                                  '',
                                  '',
                                  '',
                                  '',
                                  '',
                                  '',
                                  '',
                                  '',
                                  'F');
            close CT; 
        END;"""%(pnm_auto_key,stock_rec[0],qty_update or qty_oh,stock_rec[1],stock_rec[2],\
            stock_rec[3],stock_rec[4],stock_rec[5],stock_rec[6],stock_rec[7],stock_rec[8],\
            stock_rec[9],stock_rec[10],stock_rec[11],stock_rec[12],stock_rec[13],\
            stock_rec[14],stock_rec[15],stock_rec[16],stock_rec[17],stock_rec[18],\
            stock_rec[19],stock_rec[20],stock_rec[21],stock_rec[22],stock_rec[23],\
            stock_rec[24],stock_rec[25],stock_rec[26],stock_rec[27],stock_rec[28],\
            stock_rec[29],stock_rec[30],stock_rec[31],stock_rec[32],stock_rec[33],\
            stock_rec[34],stock_rec[35],stock_rec[36],stock_rec[37],stock_rec[38],\
            stock_rec[39],stock_rec[40],stock_rec[41],stock_rec[42],stock_rec[43])  
        error = updation_dir(query,cr)

        squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
        (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9])
        error = updation_dir(squery,cr)
        
    else:
        error = 'No stock record found.' 
    return error        

@shared_task
def update_lots(quapi_id,\
    user_id,session_id,sysur_auto_key,\
    stm_keys,parameters,filters,stm_lists=[],\
    is_mgmt=False): 
    error,msg = '','' 
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not cr:
        return 'Quantum connection failed.',''
    stock_status = parameters[0]
    notes = parameters[1]
    quantity = parameters[2]
    
    if quantity:
        try:
            quantity = quantity and float(quantity)
        except Exception as exc:
            return 'Quantity must be a number.',msg
        
    condition = parameters[3]
    serial_number = parameters[4]
    traceable,tag_date,mfg_date,tagged,obtained,hold_flag = '','','','','',False
    manufacturer,ctry_origin,tag_type,tsn_csn,tso_cso,exp_date = '','','','','',''
    insp_date,pn,desc,cons,loc,should_be,loc_auto_key = '','','','','','',''
    remarks,cure_date,trac,alt_pn,pnm_auto_key,cnc_auto_key = '','','','','',''
    traceable_to = ''
    #update the stock now
    set_clause = ''

    if not is_mgmt:
        lot_number = filters[0]
        
        if lot_number:
            qty_need_sub = """(SELECT WB.QTY_NEEDED FROM WO_BOM WB,STOCK S,STOCK_TI STI WHERE WB.WOO_AUTO_KEY IN (SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER = '%s') AND S.STM_AUTO_KEY = STI.STM_AUTO_KEY AND STI.WOB_AUTO_KEY = WB.WOB_AUTO_KEY AND ROWNUM <= 1)"""%lot_number         
            set_clause = ' QTY_REC_FROM_LOT = %s'%qty_need_sub
            
        if len(parameters) > 5:
            """         
                parameters += [upd_traceable,upd_tag_date,upd_mfg_date,upd_tagged]
                parameters += [upd_obtained,upd_hold,upd_mfctr,upd_ctry_origin,upd_tag_type]
                parameters += [upd_tsn_csn,upd_tso_cso,upd_insp_date]
                parameters += [upd_pn,upd_desc,upd_cons,upd_loc,upd_should_be]
                parameters += [upd_trac,upd_alt_pn,upd_remarks,upd_cure_date]
                parameters += [upd_exp_date,]
            """
            traceable_to = parameters[5]
            tag_date = parameters[6]
            mfg_date = parameters[7]
            tagged = parameters[8]
            obtained = parameters[9]
            hold_flag = parameters[10]
            manufacturer = parameters[11]
            ctry_origin = parameters[12]
            tag_type = parameters[13]
            tsn_csn = parameters[14]
            tso_cso = parameters[15]
            insp_date = parameters[16]
            pn = parameters[17]
            desc = parameters[18]
            cons = parameters[19]
            loc = parameters[20]
            should_be = parameters[21]
            trac = parameters[22]
            alt_pn = parameters[23]
            remarks = parameters[24]
            cure_date = parameters[25]
            exp_date = parameters[26]
        
    if not stm_lists:
        stm_lists = construct_akl(stm_keys)
         
    if manufacturer:
        if set_clause:
            set_clause += ','
        set_clause += " IC_UDF_006 = '%s'"%manufacturer

    if ctry_origin:
        if set_clause:
            set_clause += ','
        set_clause += " IC_UDF_020 = '%s'"%ctry_origin

    if tag_type:
        if set_clause:
            set_clause += ','
        tag_type_sub = """(SELECT TTP_AUTO_KEY FROM TAG_TYPE
        WHERE UPPER(TAG_TYPE_CODE) = UPPER('%s') AND ROWNUM <= 1)"""%tag_type
        set_clause += " TTP_AUTO_KEY = %s"%tag_type_sub

    if tsn_csn:
        if set_clause:
            set_clause += ','
        set_clause += " IC_UDF_008 = '%s'"%tsn_csn

    if tso_cso:
        if set_clause:
            set_clause += ','
        set_clause += " IC_UDF_009 = '%s'"%tso_cso

    if exp_date:
        if set_clause:
            set_clause += ','
        set_clause += " EXP_DATE = TO_DATE('%s','MM/DD/YYYY')"%exp_date        
        
    if notes:
        if set_clause:
            set_clause += ','
        set_clause += " NOTES = '%s'"%notes
       
    if stock_status:
        if set_clause:
            set_clause += ','
        set_clause += " IC_UDL_004 = %s"%int(stock_status)
         
    if serial_number:
        if set_clause:
            set_clause += ','
        set_clause += " SERIAL_NUMBER = '%s'"%serial_number
        
    if obtained:
        if set_clause:
            set_clause += ','
        cmp_sub = """(SELECT CMP_AUTO_KEY FROM COMPANIES
        WHERE UPPER(COMPANY_NAME) = UPPER('%s') AND ROWNUM <= 1)"""%obtained
        #set_clause += " CMP_AUTO_KEY = %s"%cmp_sub
        #Unical only:
        set_clause += " IC_UDF_005 = %s"%cmp_sub
        
    if traceable_to:
        if set_clause:
            set_clause += ','
        trace_sub = """(SELECT CTS_AUTO_KEY FROM CERT_SOURCE
        WHERE UPPER(CERT_SOURCE) = UPPER('%s') AND ROWNUM <= 1)"""%traceable_to
        #set_clause += " CTS_AUTO_KEY = %s"%trace_sub
        #Unical only:
        set_clause += " IC_UDF_007 = %s"%trace_sub
        
    if tagged:
        if set_clause:
            set_clause += ','
        tag_sub = """(SELECT CMP_AUTO_KEY FROM COMPANIES
        WHERE UPPER(COMPANY_NAME) = UPPER('%s') AND ROWNUM <= 1)"""%tagged
        set_clause += " CMP_TAGGED_BY = %s, TAGGED_BY = '%s'"%(tag_sub,tagged)
        
    if mfg_date:
        if set_clause:
            set_clause += ','
        set_clause += " MFG_DATE = TO_DATE('%s','MM/DD/YYYY')"%mfg_date 

    if tag_date:
        if set_clause:
            set_clause += ','
        set_clause += " TAG_DATE = TO_DATE('%s','MM/DD/YYYY')"%tag_date

    if hold_flag:
        if set_clause:
            set_clause += ','
        set_clause += " HOLD_LINE = 'T'"
        
    if insp_date:
        if set_clause:
            set_clause += ','
        set_clause += " INSPECT_DUE_DATE = TO_DATE('%s','MM/DD/YYYY')"%insp_date
            
    if should_be:
        if set_clause:
            set_clause += ','
        set_clause += " HOLD_LINE = 'T'"
        
    if trac:
        if set_clause:
            set_clause += ','
        set_clause += " IC_UDF_007 = '%s'"%trac
        
    if remarks:
        if set_clause:
            set_clause += ','
        set_clause += " remarks = '%s'"%remarks
        
    if cure_date:
        if set_clause:
            set_clause += ','
        set_clause += " CURE_DATE = TO_DATE('%s','')"%cure_date
            
    if exp_date:
        if set_clause:
            set_clause += ','
        set_clause += " EXP_DATE = TO_DATE('%s','')"%exp_date
        
    if set_clause:
  
        for stm in stm_keys:
            query = """UPDATE STOCK SET%s WHERE STM_AUTO_KEY = %s"""%(set_clause,stm) 
            error = updation_dir(query,cr)
            if error != '{"recs": ""}':
                break
       
        for stm_auto_key in stm_keys:
            squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
            (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],stm_auto_key)
            error = updation_dir(squery,cr) 
            if error != '{"recs": ""}': 
                break
  
    if alt_pn:
        query="""SELECT PNM_AUTO_KEY FROM PARTS_MASTER WHERE UPPER(PN)=UPPER('%s')"""%alt_pn
        pnm = selection_dir(query,cr)           
        pnm_auto_key = pnm and pnm[0] and pnm[0][0]  
        
        if not pnm:
            error = 'PN not found.'
                    
    if quantity or condition or cons or loc or pnm_auto_key:
        pcc_auto_key = ''
        
        if condition:
            #use the stock transfer procedure to update the pcc_auto_key
            cond_sub = """SELECT PCC_AUTO_KEY FROM 
                PART_CONDITION_CODES 
                WHERE UPPER(CONDITION_CODE) = UPPER('%s')"""%condition
            pcc = selection_dir(cond_sub,cr)
            pcc_auto_key = pcc and pcc[0] and pcc[0][0]

            if not pcc_auto_key:
                error = 'Part condition not found.'
      
        if cons:
            query = """SELECT CNC_AUTO_KEY FROM CONSIGNMENT_CODES
                WHERE UPPER(CONDITION_CODE) = UPPER('%s')
                """%cons
               
            cnc = selection_dir(query,cr)
            cnc_auto_key = cnc and cnc[0] and cnc[0][0]
            
            if not cnc_auto_key:
                error = 'Consignment not found.'
            
        if loc:
            query = """SELECT LOC_AUTO_KEY FROM LOCATION
                WHERE UPPER(LOCATION_CODE) = UPPER('%s')
                """%loc
               
            loc = selection_dir(query,cr)
            loc_auto_key = loc and loc[0] and loc[0][0]
            
            if not loc_auto_key:
                error = 'Location not found.'


        for stm_auto_key in stm_keys:
            
            if quantity or condition or cons or loc:
                srecs = get_wos_from_rack_beta(quapi=quapi,stm_auto_key=stm_auto_key)
                rec = srecs and srecs[0] or []
                qty_update = rec and rec[29] or 1

                if not rec:
                    return 'Stock not found.',msg

                else:
                    """
                    call the stock transfer procedure with user-entered qty and see
                    if new stock record gets created.
                    """
                    if quantity != '':
                    
                        if quantity > rec[29]:
                            error = 'Qty more than expected.'
                            return error,msg
                            
                        else:
                            qty_update = quantity 
                    
                    params=[]
                    params.append(stm_auto_key)#stm_auto_key
                    params.append(qty_update)#qty_oh
                    params.append(rec[30] or 1)#syscm_auto_key
                    params.append(pcc_auto_key or rec[31])#pcc_auto_key                       
                    params.append(cnc_auto_key or rec[32])#cnc_auto_key
                    params.append(loc_auto_key or rec[21] or '')#loc_auto_key 
                    params.append(rec[22] or '')#whs_auto_key
                    params.append(rec[33] or '')#stc_auto_key
                    params.append(rec[34] or '')#dpt_auto_key
                    params.append(rec[20] or '')#str_auto_key  
                    params.append(rec[35])#qty_reserved
                    params.append(rec[36] or '')#sod_auto_key
                    params.append(rec[37] or '')#rod_auto_key
                    params.append(rec[38] or '')#wob_auto_key
                    params.append(rec[39] or '')#pod_auto_key
                    params.append(rec[11] or '')#woo_auto_key                     
                    error = qry_stock_transfer(sysur_auto_key,user_id,params,\
                        quapi,recs=srecs,cr=cr,con=con)
                    
                    if error != '':
                        break 
                        
            if pnm_auto_key:
                error = qry_pn_transfer(cr,sysur_auto_key,user_id,stm_auto_key,pnm_auto_key,quantity)
                if error != '{"recs": ""}':
                    break
        if pnm_auto_key:
            error = synch_new_stms(cr,session_id,len(stm_keys))
                        
    aud_status = 'failure'
    if not error or error == '{"recs": ""}':
        error = ''
        orcl_commit(con=con)     
        aud_status = 'success'
        msg = 'Successful update.'
        if not is_mgmt:
            error,filter_msg = get_lots(quapi_id,session_id,filters)
        
        right_now = datetime.now()
        right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')                   
        from polls.models import MLApps as maps,QuantumUser as qu
        app_id = maps.objects.filter(code='lot-management')   
        user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
        user_rec = user_rec and user_rec[0] or None
        if user_rec:
            new_val = 'Update to stock with a change to notes and/or stock status.'
            field_changed = new_val
            field_changed += error
            error += register_audit_trail(user_rec,field_changed,new_val,right_now,app_id,quapi,status=aud_status)  
        else:
            error = 'Incorrect Quantum User ID.' 
    return error,msg
    
def synch_new_stms(cr,session_id,len_stm_keys):
    error = ''
    query = """SELECT
        W.SI_NUMBER,
        P.PN,
        P.DESCRIPTION,
        WB.EXPECTED_SN,        
        '',        
        WB.QTY_NEEDED, 
        S.QTY_OH,
        PCC.CONDITION_CODE,
        SPC.CONDITION_CODE,
        TO_CHAR(S.NOTES),
        S.STOCK_LINE,
        S.STM_AUTO_KEY,
        S.SERIAL_NUMBER,
        S.HOLD_LINE,
        S.SERIES_ID,
        CC.CONSIGNMENT_CODE,
        S.EXP_DATE,
        TO_CHAR(S.REMARKS)
        FROM STOCK S
        LEFT JOIN STOCK_TI STI ON S.STM_AUTO_KEY = STI.STM_AUTO_KEY
        LEFT JOIN WO_BOM WB ON STI.WOB_AUTO_KEY = WB.WOB_AUTO_KEY
        LEFT JOIN WO_OPERATION W ON WB.WOO_AUTO_KEY = W.WOO_AUTO_KEY
        LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = WB.PCC_AUTO_KEY
        LEFT JOIN PART_CONDITION_CODES SPC ON SPC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
        LEFT JOIN CONSIGNMENT_CODES CC ON CC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
        LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
        WHERE ROWNUM<=%s
        ORDER BY S.STM_AUTO_KEY DESC
    """%len_stm_keys
        
    recs = selection_dir(query,cr)
    if not recs:
        error = 'No new stock records found.'
    else:
        error = create_bulk_lots(session_id,recs,is_dashboard=True)
    return error       

#*****************************************************************************#
    
@shared_task
def user_create(quapi_id,user_id,sysur_auto_key,session_id): 
    error,msg,fail_msg,password  = '','','','' 
    good_rows = {}
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    from polls.models import Document,MLApps as maps,QuantumUser as qu
    import_file = Document.objects.filter(session_id=session_id)
    import_file = import_file and import_file[0] or None
    file_path = import_file and os.path.join(import_file.docfile.path) or ''
    from openpyxl import load_workbook
    wb = load_workbook(filename = file_path)
    sheet = wb.active
    sheet_rows = sheet.iter_rows()
    row_list = []
    row_vals = [[v.value for v in row] for row in sheet_rows]
    col_headings = row_vals[0] 
    count = 0
    for row in row_vals:
        #if it is the first row, then it is the headings
        #assign the headings as keys for each value in row
        if count == 0 or not row[0]:
            count += 1
            continue
        
        dict_row = {}
        col_count = 0
        for col in col_headings:
            dict_row[col] = row[col_count]
            col_count += 1       
        row_list.append(dict_row)
        count+=1 
    if row_list: 
        line_count = 0
        bad_rows = []
        from django.contrib.auth.models import User,Group
        from polls.models import UserProfile as upro
        #loop through each row and process
        position = 'Initial'           
        for row in row_list:                                   
            bad_seq = False
            line_count += 1
            sysur_auto_key = row.get('SYSUR_AUTO_KEY','') 
            user_name = row.get('User Name','') 
            first_name = row.get('First Name','')
            last_name = row.get('Last Name','')            
            import_row = [sysur_auto_key,user_name,first_name,last_name]
            #prepare values for insert
            if not sysur_auto_key:           
                error = 'Line %s has no sysur_auto_key.'%line_count
            if not user_name:           
                error = 'Line %s has no user name.'%line_count
            if error:
                bad_rows.append(user_name) 
            else:            
                #user = User.objects.create_user(user_name, '', '%MROLIVE%')
                #check that there is not already a user with the same username
                ex_user = User.objects.filter(username=user_name)
                if ex_user:
                    continue
                    #error = 'User, %s, already exists.'%user_name
                else:
                    user = User.objects.create(
                        username=user_name,
                        password='%MROLIVE%',
                        first_name=first_name,
                        last_name=last_name,
                        is_superuser = True,
                        )
                      
                    user.set_password('%MROLIVE%')
                    user.save() 
                    user_group = Group.objects.get(name='American Parts Corporation') 
                    user.groups.add(user_group)
                    #create user profile to store the sysur_auto_key, etc.
                    user_profile = upro.objects.create(
                        sysur_auto_key = sysur_auto_key,
                        user = user,
                        )
                    user_profile.save()                    
        if len(bad_rows) < len(row_list):                    
            msg = 'Successfully imported ' + str(len(row_list) - len(bad_rows)) + ' users.  ' + str(len(bad_rows)) + ' users rejected.'
        else:
            msg = 'Could not import any rows. Please see rejection reasons in grid.'   
    #register audit trail record            
    aud_status = 'success'
    app_id = maps.objects.filter(code='user-import')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        field_changed = msg
        new_val = msg
        if error:             
            aud_status = 'failure'
            new_val = error
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,msg,fail_msg

def create_stock_carts(recs,session_id):
    rec_data,error = [],''
    from polls.models import StockCart
    for cart in recs:    
        rec_data.append(StockCart(
        udl_auto_key =  cart[0], 
        name = cart[1], 
        udl_code = cart[2],  
        session_id = session_id,      
        ))
    if rec_data:
        try:
            delete = StockCart.objects.filter(session_id=session_id).delete()
            StockCart.objects.bulk_create(rec_data) or []    
        except Exception as exc:
            error += "Error, %s, creating cart."%(exc)
    return error
    
def create_stock_locations(recs,session_id):
    rec_data,error = [],''
    from polls.models import Location
    for loc in recs:    
        rec_data.append(Location(
        loc_auto_key = loc[0], 
        name = loc[1], 
        location_code = loc[2], 
        iq_enable = (loc[3] == 'T' and True) or False,       
        session_id = session_id,      
        ))
    if rec_data:
        try:
            delete = Location.objects.filter(session_id=session_id).delete()
            Location.objects.bulk_create(rec_data) or []    
        except Exception as exc:
            error += "Error, %s, creating location."%(exc)
    return error
    
def create_whs(recs,session_id):
    rec_data,error = [],''
    from polls.models import Warehouse
    for whs in recs:    
        rec_data.append(Warehouse(
        whs_auto_key = whs[0],
        name = whs[1],
        warehouse_code = whs[2],
        session_id = session_id,      
        ))
    if rec_data:
        try:
            delete = Warehouse.objects.filter(session_id=session_id).delete()
            Warehouse.objects.bulk_create(rec_data) or []    
        except Exception as exc:
            error += "Error, %s, creating whs."%(exc)
    return error
    
@shared_task
def lookup_stock_cart(quapi_id,session_id,cart):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    query = """SELECT UDL_AUTO_KEY,UDL_DESCRIPTION,
    UDL_CODE FROM USER_DEFINED_LOOKUPS WHERE
    UPPER(UDL_CODE) = UPPER('%s')"""%cart
    recs = selection_dir(query,cr)
    if recs:
        error = create_stock_carts(recs,session_id)
    else:
        error = 'Cart not found.'
    return error,msg

@shared_task
def lookup_stock_cart(quapi_id,session_id,cart):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    query = """SELECT UDL_AUTO_KEY,UDL_DESCRIPTION,
    UDL_CODE FROM USER_DEFINED_LOOKUPS WHERE
    UPPER(UDL_CODE) = UPPER('%s')"""%cart
    recs = selection_dir(query,cr)
    if recs:
        error = create_stock_carts(recs,session_id)
    else:
        error = 'Cart not found.'
    return error,msg

@shared_task    
def lookup_location(quapi_id,session_id,location):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    query = """SELECT LOC_AUTO_KEY,DESCRIPTION,
    LOCATION_CODE,IQ_ENABLE FROM LOCATION WHERE
    UPPER(LOCATION_CODE) = UPPER('%s') AND
    HISTORICAL='F'"""%location
    recs = selection_dir(query,cr)
    if recs:
        error = create_stock_locations(recs,session_id)
    else:
        error = 'Location not found.'
    return error,msg

@shared_task
def lookup_warehouse(quapi_id,session_id,warehouse):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    
    query = """SELECT WHS_AUTO_KEY,DESCRIPTION,
        WAREHOUSE_CODE FROM WAREHOUSE WHERE
        UPPER(WAREHOUSE_CODE) = UPPER('%s') AND
        HISTORICAL='F'"""%warehouse
    recs = selection_dir(query,cr)
    whs_auto_key = recs and recs[0] and recs[0][0] or ''
    
    if recs:
        error = create_whs(recs,session_id)
    else:
        error = 'Warehouse not found.'
        
    return error,msg,whs_auto_key
    
def create_skills_bulk(recs,session_id,tlabor_obj):
    skills_data,error = [],''
    for skill in recs:    
        skills_data.append(tlabor_obj(
        wok_auto_key = skill[0],
        description = skill[1],    
        session_id = session_id,      
        ))
    if skills_data:
        try:
            delete = tlabor_obj.objects.all().delete()
            tlabor_obj.objects.bulk_create(skills_data) or []    
        except Exception as exc:
            error += "Error, %s, creating bulk skills."%(exc)
    return error

@shared_task
def get_wo_skills(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi) 
    query = """SELECT WOK_AUTO_KEY,DESCRIPTION FROM WO_SKILLS ORDER BY DESCRIPTION"""
    recs = selection_dir(query,cr)
    from polls.models import TaskSkills as ts
    error = create_skills_bulk(recs,session_id,ts)    
    return error,msg

@shared_task
def labor_modify(quapi_id,session_id,\
    sysur_auto_key,wtl_list,mod_type,\
    user_change='',user_name='',wot_auto_key=0,\
    date_start=None,date_stop=None,wo_skill=0):
    error,msg,set_clause,date_clause = '','','',''
    from polls.models   import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi) 
    """query = "SELECT SYSTIMESTAMP FROM DUAL"
    today = selection_dir(query,cr)
    today = today and today[0] and today[0][0] or None
    today = today and datetime.strptime(today,'%Y-%m-%d %H:%M:%S.%f')
    today = today and today.strftime(date_format)"""
    date_format = "%m/%d/%Y %H:%M:%S"
    date_start = date_start and datetime.strptime(date_start,'%m/%d/%Y %I:%M %p')
    date_stop = date_stop and datetime.strptime(date_stop,'%m/%d/%Y %I:%M %p')
    if date_start and date_stop:
        if date_start > date_stop:
            error = 'Start time must be before stop time.' 
            return 'Start time must be before stop time.' ,msg            
    date_start = date_start and date_start.strftime('%m/%d/%Y %H:%M:%S')
    date_stop = date_stop and date_stop.strftime('%m/%d/%Y %H:%M:%S')
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    
    if mod_type == 'ADD' and wot_auto_key:
        session_id = 'ay8nNoi80920KHOI:jgals82'
        if date_start:
            error,msg = add_wo_labor(quapi_id,\
            session_id,\
            sysur_auto_key,\
            wot_auto_key=wot_auto_key,\
            today=date_start,user_name=user_name)
        if date_stop and not error:
            error,msg = add_wo_labor(quapi_id,\
            session_id,\
            sysur_auto_key,\
            wot_auto_key=wot_auto_key[:-1]+'c',\
            today=date_stop,user_name=user_name)
    else:  
        for wtl_key in wtl_list:
            set_clause,date_clause='',''
            if user_change and user_change != user_name:
                set_clause = """ SYSUR_AUTO_KEY = 
                    (SELECT SYSUR_AUTO_KEY FROM 
                    SYS_USERS WHERE 
                    UPPER(USER_ID)=UPPER('%s'))"""%user_change
            if wo_skill:
                if set_clause:
                    set_clause += ', '
                set_clause +=" WOK_AUTO_KEY = %s"%wo_skill           
            if date_start:
                fdate = datetime.strptime(date_start,date_format)
                if date_stop:
                    tdate = datetime.strptime(date_stop,date_format)
                    if fdate > tdate:
                        error = 'Start time must be before stop time.'
                if set_clause:
                    date_clause += ', '
                date_clause += """START_TIME=
                    TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss')"""%(date_start) 
                    
            if error in ['{"recs": ""}','']:
                #hours calculation dependent on user having set stop/start times 
                diff_query = ''                
                if date_start and date_stop:
                    diff_query = """SELECT 
                    24*(TO_DATE('%s', 'MM/DD/YYYY hh24:mi:ss') - TO_DATE('%s', 'MM/DD/YYYY hh24:mi:ss'))
                    FROM WO_TASK_LABOR WHERE WTL_AUTO_KEY=%s"""%(date_stop,date_start,wtl_key)
                elif date_stop:                    
                    diff_query = """SELECT 
                    24*(TO_DATE('%s', 'MM/DD/YYYY hh24:mi:ss') - START_TIME)
                    FROM WO_TASK_LABOR WHERE WTL_AUTO_KEY=%s"""%(date_stop,wtl_key)
                elif date_start:                    
                    diff_query = """SELECT 
                    24*(STOP_TIME - TO_DATE('%s', 'MM/DD/YYYY hh24:mi:ss'))
                    FROM WO_TASK_LABOR WHERE WTL_AUTO_KEY=%s"""%(date_start,wtl_key)                    
          
                if diff_query:                    
                    diff = selection_dir(diff_query,cr)
                    diff_hours = diff and diff[0] and diff[0][0] and round(diff[0][0],2)or 0
                    if date_stop:
                        if date_clause or set_clause:
                            date_clause += ', '
                        date_clause += """STOP_TIME=TO_TIMESTAMP('%s','MM/DD/YYYY hh24:mi:ss')"""%(date_stop)
                        if diff_hours:                       
                            date_clause += """, HOURS=%s, HOURS_BILLABLE=%s
                            """%(diff_hours,diff_hours)         
                    set_clause += date_clause 
                if set_clause and wtl_key:                    
                    query = """UPDATE WO_TASK_LABOR SET
                        %s                                  
                        WHERE WTL_AUTO_KEY = %s"""%(set_clause,wtl_key)                        
                    error = updation_dir(query,cr)
    aud_status = 'failure'
    if error in ['{"recs": ""}','']:
        error = ''            
        msg += 'Successfully %sed task labor entry(ies).'%mod_type.lower() 
        aud_status = 'success'
        orcl_commit(con=con)
    right_now = datetime.now()
    right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')        
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='labor-mgmt')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:       
        field_changed = '%sed task labor entry. '%mod_type
        new_val = field_changed + error
        field_changed += error
        error += register_audit_trail(user_rec,\
            field_changed,\
            new_val,right_now,\
            app_id,quapi,status=aud_status)  
    else:
        error = 'Incorrect Quantum User ID.'          
    return error,msg
    
@shared_task
def get_wola_recs(quapi_id,user_id,date_from,date_to):
    query = """SELECT WO.SI_NUMBER,WOT.SEQUENCE,WOK.DESCRIPTION,
       WTM.DESCRIPTION,UDL.UDL_CODE,WOT.WOT_AUTO_KEY,WOT.WOS_AUTO_KEY                                    
       FROM WO_TASK WOT
       left join wo_operation wo on wo.woo_auto_key = WOT.woo_auto_key
       LEFT JOIN stock_reservations SR ON SR.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
       left join stock s on s.stm_auto_key = sr.stm_auto_key
       left join wo_task_master wtm on wtm.wtm_auto_key = wot.wtm_auto_key
       left join user_defined_lookups udl on udl.udl_auto_key = s.ic_udl_005 
       left join sys_users sys on sys.sysur_auto_key = wot.sysur_auto_key
       left join wo_skills wok on wok.wok_auto_key = sys.wok_auto_key
       WHERE WOT.WOT_AUTO_KEY = %s"""%wot_auto_key 
    stock_recs = selection_dir(query,cr)
    stock_rec = stock_recs and stock_recs[0] or None
    from polls.models import WOStatus as wos
    error = wo_labor_create(stock_recs,session_id,wos)    
    return error

@shared_task
def get_wo_parts(quapi_id,session_id,wo_number):
    error,msg,serial_no,part_number,description,customer = '','','','','',''   
    from polls.models   import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)             
    if not (cr and con):
        return 'Cannot connect to Oracle',msg,serial_no,part_number,description,customer
    query="""SELECT S.SERIAL_NUMBER, P.PN, P.DESCRIPTION, C.COMPANY_NAME 
        FROM PARTS_MASTER P
        JOIN STOCK S ON S.PNM_AUTO_KEY = P.PNM_AUTO_KEY
        JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
        JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
        JOIN COMPANIES C ON C.CMP_AUTO_KEY = WO.CMP_AUTO_KEY
        ORDER BY P.PN
    """
    recs = selection_dir(query,cr)

    if recs:
        serial_no = recs[0] and recs[0][0] or ''
        part_number = recs[0] and recs[0][1] or ''
        description = recs[0] and recs[0][2] or ''
        customer = recs[0] and recs[0][3] or ''
    else:
        error = 'No parts found.'    
    return error,msg,serial_no,part_number,description,customer

def labor_synch(session_id,recs,is_mgmt=False,systime=None):
    total_hours = 0
    from polls.models import TaskLabor             
    wtl_data = []
    error,start_time = '',None            
    for rec in recs:
        hours = rec[6]
        #MTU Custom -
        stop_time = rec[3] and datetime.strptime(rec[3],'%Y-%m-%d %H:%M:%S') or None
        #stop_time = stop_time and (stop_time + timedelta(hours=1)) or None                      
        if rec[2]:
            
            start_time = datetime.strptime(rec[2],'%Y-%m-%d %H:%M:%S')
            #start_time = start_time + timedelta(hours=1)
            
            if systime and not rec[3]:
                #MTU - CST - CST
                diff = systime - start_time    
                days, seconds = diff.days, diff.seconds
                hours = days * 24 + seconds // 3600 + 1
                total_hours += hours
            #makes it EST
                            
        wtl_data.append(TaskLabor(
            session_id = session_id,
            user_name = rec[0],
            full_name = rec[14],
            entry_date = rec[1] and rec[1][:10] or None,
            start_time = start_time or None,
            stop_time = stop_time or None,
            wo_number = rec[4],
            task_desc = rec[5], 
            hours = hours,
            dept_name = rec[7],
            skill_desc = rec[8],
            pn = rec[9],
            batch_id = rec[10], 
            wtl_auto_key = rec[11], 
            part_desc = rec[12] or '',            
        ))                   
    try:
        #if not is_mgmt:
        #    TaskLabor.objects.all().delete()
        #else:
        TaskLabor.objects.all().delete() 
        error = wtl_data and TaskLabor.objects.bulk_create(wtl_data) or []    
    except Exception as exc:
        error = "Error with creating task labor entries locally: %s"%exc
    return total_hours

@shared_task
def labor_dashboard(quapi_id,session_id,\
    user_auto_key=0,user_id='',date_from='',\
    date_to='',is_mgmt=False,is_detail=False,\
    wo_number=''):
    error,msg,date_clause,user_clause,wo_clause = '','','','',''
    recs,mgmt_recs = [],[]
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)             
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    total_hours = 0
    """time_query = "SELECT SYSTIMESTAMP FROM DUAL"
    today = selection_dir(time_query,cr)
    today = today and today[0] and today[0][0] and today[0][0][:18] or None
    timestamp = today and datetime.strptime(today,'%Y-%m-%d %H:%M:%S')
    timestamp = timestamp and timestamp.strftime('%Y-%m-%d %H:%M:%S')
    ts_today = today and datetime.strptime(today,'%Y-%m-%d %H:%M:%S')
    today = ts_today and ts_today.strftime("%m/%d/%Y")"""
    #m_format = '%m/%d/%Y %H:%M:%S'
    date_format = '%Y-%m-%d %H:%M:%S'
    today = datetime.now()
    timestamp = today and today.strftime(date_format)    
    #MTU: makes it CST    
    ts_today = today - timedelta(hours=1)
    #ts_today = today and datetime.strptime(today,'%Y-%m-%d %H:%M:%S')
    
    if user_id:
        user_clause = """ WTL.SYSUR_AUTO_KEY=
            (SELECT SYSUR_AUTO_KEY FROM 
            SYS_USERS WHERE UPPER(USER_NAME)=UPPER('%s'))"""%(user_id)
    if user_auto_key:
        user_clause = """ WTL.SYSUR_AUTO_KEY=%s"""%(user_auto_key)
    if date_from:
        if date_to:
            fdate = datetime.strptime(date_from,'%m/%d/%Y')
            tdate = datetime.strptime(date_to,'%m/%d/%Y')
            if fdate > tdate:
                error = 'Date from must precede date to.'
        if user_clause:
            user_clause += ' AND'
             
                                   
        date_clause += " TO_DATE('%s','mm/dd/yyyy') <= WTL.ENTRY_DATE"%(date_from[:10])
    if not error:

        if wo_number:
            if date_clause or user_clause:
                wo_clause += ' AND'
            wo_clause += " UPPER(WOO.SI_NUMBER) = UPPER('%s')"%wo_number    
        if date_to:
            if date_clause or user_clause:
                date_clause += ' AND'
                                       
            date_clause += " TO_DATE('%s','mm/dd/yyyy') >= WTL.ENTRY_DATE"%(date_to[:10])
              
        if is_detail:       
            query = """SELECT SUR.USER_NAME, WTL.ENTRY_DATE, 
                    WTL.START_TIME, WTL.STOP_TIME, 
                    WOO.SI_NUMBER, WTM.DESCRIPTION, 
                    WTL.HOURS,DPT.DEPT_NAME, 
                    WOK.DESCRIPTION, PNM.PN, 
                    LBH.BATCH_ID,WTL.WTL_AUTO_KEY,0,
                    WTL.SYSUR_AUTO_KEY,
                    SUR.FIRST_NAME || ' ' || SUR.LAST_NAME                    
                FROM WO_TASK_LABOR WTL
                    JOIN SYS_USERS SUR ON WTL.SYSUR_AUTO_KEY = SUR.SYSUR_AUTO_KEY
                    JOIN WO_TASK WOT ON WTL.WOT_AUTO_KEY = WOT.WOT_AUTO_KEY
                    JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WOT.WTM_AUTO_KEY
                    JOIN WO_OPERATION WOO ON WOT.WOO_AUTO_KEY = WOO.WOO_AUTO_KEY 
                    LEFT JOIN PARTS_MASTER PNM ON WOO.PNM_AUTO_KEY = PNM.PNM_AUTO_KEY
                    LEFT JOIN WO_SKILLS WOK ON WTL.WOK_AUTO_KEY = WOK.WOK_AUTO_KEY
                    LEFT JOIN DEPARTMENT DPT ON WTL.DPT_AUTO_KEY = DPT.DPT_AUTO_KEY
                    LEFT JOIN LABOR_BATCH_DETAIL LBD ON WTL.LBD_AUTO_KEY = LBD.LBD_AUTO_KEY
                    LEFT JOIN LABOR_BATCH_HEADER LBH ON LBD.LBD_AUTO_KEY = LBH.LBH_AUTO_KEY
                WHERE %s
          
                ORDER BY WTL.WTL_AUTO_KEY DESC
                """%(user_clause + date_clause + wo_clause)
            recs = selection_dir(query,cr)
            
        elif not is_mgmt: 
            where_clause = user_clause + date_clause + wo_clause
            where_clause = where_clause and where_clause + ' AND' or ''
            query = """SELECT SUR.USER_NAME, WTL.ENTRY_DATE, 
                    WTL.START_TIME, WTL.STOP_TIME, 
                    WOO.SI_NUMBER, WTM.DESCRIPTION, 
                    WTL.HOURS, DPT.DEPT_NAME, 
                    WOK.DESCRIPTION, PNM.PN, 
                    LBH.BATCH_ID,WTL.WTL_AUTO_KEY,
                    CASE WHEN WTL.STOP_TIME IS NULL THEN 
                    TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss') - WTL.START_TIME
                    ELSE TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss') - WTL.STOP_TIME 
                    END,
                    WTL.SYSUR_AUTO_KEY,
                    SUR.FIRST_NAME || ' ' || SUR.LAST_NAME
                    FROM WO_TASK_LABOR WTL, 
                    SYS_USERS SUR,
                    WO_TASK WOT,
                    WO_TASK_MASTER WTM,
                    WO_OPERATION WOO,  
                    PARTS_MASTER PNM,
                    WO_SKILLS WOK,
                    DEPARTMENT DPT,
                    LABOR_BATCH_DETAIL LBD,
                    LABOR_BATCH_HEADER LBH,
            (SELECT SYSUR_AUTO_KEY,MAX(WTL_AUTO_KEY) WLK       
                FROM WO_TASK_LABOR 
                GROUP BY SYSUR_AUTO_KEY) WTL2
                     WHERE %s
                     WTL.SYSUR_AUTO_KEY = WTL2.SYSUR_AUTO_KEY
                     AND WTL.WTL_AUTO_KEY = WTL2.WLK
                     AND WTL.SYSUR_AUTO_KEY = SUR.SYSUR_AUTO_KEY
                     AND WTL.WOT_AUTO_KEY = WOT.WOT_AUTO_KEY
                     AND WTM.WTM_AUTO_KEY = WOT.WTM_AUTO_KEY
                     AND WOT.WOO_AUTO_KEY = WOO.WOO_AUTO_KEY
                     AND WOO.PNM_AUTO_KEY = PNM.PNM_AUTO_KEY  (+)
                     AND WTL.WOK_AUTO_KEY = WOK.WOK_AUTO_KEY  (+)
                     AND WTL.DPT_AUTO_KEY = DPT.DPT_AUTO_KEY  (+)
                     AND WTL.LBD_AUTO_KEY = LBD.LBD_AUTO_KEY  (+)
                     AND LBD.LBD_AUTO_KEY = LBH.LBH_AUTO_KEY  (+)
                     AND SUR.WO_FLAG = 'T'
                     AND WTL.START_TIME IS NOT NULL
                    ORDER BY SUR.USER_NAME DESC,WTL.WTL_AUTO_KEY DESC
                """%(timestamp,timestamp,where_clause)
                  
            recs = selection_dir(query,cr)
        else:
            query = """SELECT DISTINCT SUR.USER_NAME, '', 
                    '', '', 
                    '', '1', 
                    SUM(WTL.HOURS), '1', 
                    '1', '1', 
                    '1',1,'',
                    '',''                   
                FROM WO_TASK_LABOR WTL
                JOIN SYS_USERS SUR ON WTL.SYSUR_AUTO_KEY = SUR.SYSUR_AUTO_KEY
                JOIN WO_TASK WOT ON WTL.WOT_AUTO_KEY = WOT.WOT_AUTO_KEY
                JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WOT.WTM_AUTO_KEY
                JOIN WO_OPERATION WOO ON WOT.WOO_AUTO_KEY = WOO.WOO_AUTO_KEY 
                LEFT JOIN PARTS_MASTER PNM ON WOO.PNM_AUTO_KEY = PNM.PNM_AUTO_KEY
                LEFT JOIN WO_SKILLS WOK ON WTL.WOK_AUTO_KEY = WOK.WOK_AUTO_KEY
                LEFT JOIN DEPARTMENT DPT ON WTL.DPT_AUTO_KEY = DPT.DPT_AUTO_KEY
                LEFT JOIN LABOR_BATCH_DETAIL LBD ON WTL.LBD_AUTO_KEY = LBD.LBD_AUTO_KEY
                LEFT JOIN LABOR_BATCH_HEADER LBH ON LBD.LBD_AUTO_KEY = LBH.LBH_AUTO_KEY
                WHERE %s
                GROUP BY SUR.USER_NAME
                ORDER BY SUR.USER_NAME
                """%(user_clause + date_clause)
           
            mgmt_recs = selection_dir(query,cr)
        if recs:         
            total_hours = labor_synch(session_id,recs,is_mgmt=False,systime=ts_today)
        if mgmt_recs:
            total_hours = labor_synch(session_id,mgmt_recs,is_mgmt=True)
        if not (recs or mgmt_recs):
            error='No labor entries found.'
            
    return error,msg,total_hours        

@shared_task
def reserve_pnm(quapi_id,session_id,user_id,sysur_auto_key,pnm_list,record):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)             
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    from polls.models import WOStatus
    pnm_set = pnm_list and WOStatus.objects.filter(id__in = pnm_list)
    right_now = datetime.now()
    timestamp = right_now.strftime('%Y-%m-%d %H:%M:%S')
    query = "SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER = '%s'"%record
    woo = selection_dir(query,cr)
    woo_auto_key = woo and woo[0] and woo[0][0] or ''
        
    for pnm in pnm_set:
        woo_auto_key = woo_auto_key or pnm.woo_auto_key or ''
        wob_auto_key = pnm.wob_auto_key or ''
        squery = """INSERT INTO STOCK_RESERVATIONS 
            (STR_AUTO_KEY,STM_AUTO_KEY,WOB_AUTO_KEY,WOO_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY) 
            VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s','%s',%s,%s)"""%(pnm.stm_auto_key,wob_auto_key,woo_auto_key,pnm.quantity,sysur_auto_key)
        error = insertion_dir(squery,cr)
        if not error:
            squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
                (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG
                WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,len(user_id)>9 and user_id[:9] or 'DBA',pnm.stm_auto_key)
            error = updation_dir(squery,cr)
        if error == '{"recs": ""}' or error == '':
            error=''
            msg = 'Successful reservation.'
        if error == '':
            error=''
            aud_status = 'success'
        else: 
            aud_status = 'failure'       
        from polls.models import MLApps as maps,QuantumUser as qu
        app_id = maps.objects.filter(code='parts-request')   
        user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
        user_rec = user_rec and user_rec[0] or None
        if user_rec:
            rec_input = (pnm.ctrl_id and 'Reservation made for Ctrl#: '+ str(pnm.ctrl_number) + ' and Ctrl ID#: ' + str(pnm.ctrl_id)) or ''
            new_val = rec_input
            mode = 'Reserved stock for '
            #add parts_master.pn, wo_task.sequence, wo_operation.si_number
            field_changed = mode + pnm.ctrl_number + pnm.ctrl_id + ', with quantity: ' + str(pnm.quantity)
            field_changed += ', PN: ' + str(pnm.part_number)
            field_changed += error
            error += register_audit_trail(user_rec,field_changed,new_val,timestamp,app_id,quapi,status=aud_status) 
        else:
            error = 'Incorrect Quantum User ID.'
    orcl_commit(con=con) 
    return error,msg        

def create_bom_statuses(session_id,bom_statuses):
              
    from polls.models import StatusSelection as stat_sel
    bstat_data,error = [],''
    for bstat in bom_statuses:    
        bstat_data.append(stat_sel(
        wos_auto_key = bstat[0],
        severity = bstat[2],
        name = bstat[1],
        is_dashboard = bstat[3] and 1 or 0,
        session_id = session_id,      
        ))
    if bstat_data:
        try:
            delete = stat_sel.objects.filter(session_id=session_id).delete()
            stat_sel.objects.bulk_create(bstat_data) or []    
        except Exception as exc:
            error += "Error, %s, creating bom statuses."%(exc)
    return error
 
@shared_task
def get_bom_statuses(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)        
    where_clause,join_wo = '',''      
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
        
    query = """SELECT WBS.BOS_AUTO_KEY,
        WBS.BOM_STATUS_CODE,WBS.DESCRIPTION,WC.BOS_INITIAL
        FROM WO_BOM_STATUS WBS
        LEFT JOIN WO_CONTROL WC ON WC.BOS_INITIAL = WBS.BOS_AUTO_KEY
        """ 
    bom_statuses = selection_dir(query,cr)        
    if not bom_statuses:
        error = 'No statuses found.'
    else:
        error = create_bom_statuses(session_id,bom_statuses)  
    return error,msg
    
@shared_task
def get_part_numbers(quapi_id,session_id,part_char=''):
    error,msg = '',''
    from polls.models import QueryApi,PartNumbers    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)        
    where_clause,join_wo = '',''    
    delete = PartNumbers.objects.filter(session_id=session_id).delete()    
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    if part_char:
        parts_query = """SELECT PN,DESCRIPTION,PNM_AUTO_KEY 
        FROM PARTS_MASTER 
        WHERE PN LIKE '%s%s' ORDER BY PN"""%(part_char,'%')
    else:
        parts_query = """SELECT PN,DESCRIPTION,PNM_AUTO_KEY FROM PARTS_MASTER 
        WHERE ACTIVE_PART = 'T'                                                                                                                     
        AND PNM_AUTO_KEY IN (SELECT PNM_AUTO_KEY FROM WO_BOM WHERE ENTRY_DATE >= TO_DATE('01/01/2015', 'MM/DD/YYYY'))                                                                                                             
        """
    recs = selection_dir(parts_query,cr)
    if recs:
        error = create_parts(session_id,recs)
    return error,msg
    
def create_parts(session_id,parts):
    error = ''
    from polls.models import PartNumbers
    part_data,error = [],''
    for part in parts:    
        part_data.append(PartNumbers(
        part_number = part[0],
        description = part[1],
        pnm_auto_key = part[2],
        session_id = session_id,      
        ))
    if part_data:
        try:
            delete = PartNumbers.objects.filter(session_id=session_id).delete()
            PartNumbers.objects.bulk_create(part_data) or []    
        except Exception as exc:
            error += "Error, %s, creating parts."%(exc)
    return error

@shared_task
def get_priorities(quapi_id,session_id,app=''):
    error = ''
    from polls.models import QueryApi,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id\
        and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg

    query = """
        SELECT PRIORITY_CRITICAL FROM QUANTUM
        UNION
        SELECT PRIORITY_URGENT FROM QUANTUM
        UNION
        SELECT PRIORITY_ROUTINE FROM QUANTUM

    """
    
    recs = selection_dir(query,cr)
    from polls.models import Priority
    req_data,error = [],''
    for row in recs:       
        req_data.append(Priority(
            code = row[0],
            description = row[0],
            session_id = session_id,            
        ))
    if req_data:     
        try:
            delete = Priority.objects.filter(session_id=session_id).delete()
            rec = Priority.objects.bulk_create(req_data) or []    
        except Exception as exc:
            error += "Error, %s, with synching priorities. %s"%(exc,row[0])
    return error,app

@shared_task
def get_ship_vias(quapi_id,session_id,app=''):
    error = ''
    from polls.models import QueryApi,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id\
        and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg

    query = """SELECT SVC_AUTO_KEY,SHIP_VIA_CODE,DESCRIPTION FROM 
        SHIP_VIA_CODES
    """
    recs = selection_dir(query,cr)
    from polls.models import ShipVia
    req_data,error = [],''
    for row in recs:       
        req_data.append(ShipVia(
            svc_auto_key = row[0],
            ship_via_code = row[1],
            description = row[2],
            session_id = session_id,            
        ))
    if req_data:     
        try:
            delete = ShipVia.objects.filter(session_id=session_id).delete()
            rec = ShipVia.objects.bulk_create(req_data) or []    
        except Exception as exc:
            error += "Error, %s, with synching ship via codes. %s"%(exc,row[0])
    return error,app
    
@shared_task
def get_shipping_status(quapi_id,session_id,app=''):
    error=''
    from polls.models import QueryApi,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg

    query = """SELECT SMS_AUTO_KEY,STATUS_CODE FROM 
        SM_STATUS
    """
    recs = selection_dir(query,cr)
    from polls.models import StatusSelection
    req_data,error = [],''
    for row in recs:       
        req_data.append(StatusSelection(
            wos_auto_key = row[0],
            text_wos = str(row[0]),
            name = row[1],
            session_id = session_id,            
        ))
    if req_data:     
        try:
            delete = StatusSelection.objects.filter(session_id=session_id).delete()
            rec = StatusSelection.objects.bulk_create(req_data) or []    
        except Exception as exc:
            error += "Error, %s, with synching shipping statuses. %s"%(exc,row[0])
    
    return error,app
    
@shared_task
def update_shipping(quapi_id,session_id,\
    sysur_auto_key,user_update,smds_list,tote,notes):
    error,msg = '',''
    from polls.models import QueryApi,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id 

    if orcl_conn:
        orcl_conn = oc.objects.filter(id=quapi.orcl_conn_id)
        orcl_conn = orcl_conn and orcl_conn[0]
        cr,con = orcl_conn and orcl_connect(orcl_conn)
        
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
        
    from polls.models import WOStatus
    smds = WOStatus.objects.filter(id__in=smds_list)
    smds_list = smds.values_list('str_auto_key',flat=True)
    upd_smds_list = construct_akl(smds_list)    
    smhs_list = smds.values_list('si_number',flat=True)
    upd_smhs_list = construct_text(smhs_list) 
    upd_clause = ''
    
    if tote:
    
        if upd_clause:
            upd_clause += ', '
        upd_clause += """IC_UDL_005 = (SELECT UDL_AUTO_KEY 
            FROM USER_DEFINED_LOOKUPS 
            WHERE UPPER(UDL_CODE) = UPPER('%s'))"""%tote
            
    if notes:
    
        for upd_smds in upd_smds_list:
            upd_query = """UPDATE SM_DETAIL SET NOTES = NOTES || '; ' || '%s'
            WHERE SMD_AUTO_KEY IN (SELECT SMD_AUTO_KEY FROM STOCK_RESERVATIONS
            WHERE STR_AUTO_KEY IN %s)"""%(notes,upd_smds)
            error = updation_dir(upd_query,cr)
            if error != '{"recs": ""}':
                break
        
    if upd_clause:
    
        for upd_smds in upd_smds_list:
            query = """UPDATE STOCK SET %s
            WHERE STM_AUTO_KEY IN (SELECT STM_AUTO_KEY FROM STOCK_RESERVATIONS
            WHERE STR_AUTO_KEY IN %s)"""%(upd_clause,upd_smds)
            error = updation_dir(query,cr)
            if error != '{"recs": ""}':
                break
                
    if user_update:    
 
        for upd_smds in upd_smds_list:
            query = """UPDATE STOCK_RESERVATIONS 
                SET SYSUR_AUTO_KEY_SCAN = 
                (SELECT SYSUR_AUTO_KEY FROM 
                SYS_USERS WHERE USER_NAME = '%s')
                WHERE STR_AUTO_KEY IN %s
                """%(user_update,upd_smds) 
            error = updation_dir(query,cr)
            if error != '{"recs": ""}':
                break
    
    aud_status = 'failure'  
    
    if error == '{"recs": ""}':     
        orcl_commit(con=con)
        error = ''
        msg = 'Succcessful Update.'
        aud_status = 'success'
        
    #register audit trail record            
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='smd-management')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    
    if user_rec:
        field_changed = 'User changed: ' + user_update 
        field_changed += ', notes added: ' + notes 
        field_changed += ', tote update: ' + tote 
        new_val = 'Update to stock record and/or reservation'
        if error:             
            aud_status = 'failure'
            new_val = error
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'
        
    return error,msg
     
def add_smd_stock(wos_obj,session_id,stock_recs):
    #Cust. Order	Due Date	Status	Part Number	Description	Qty	Seial	Loc
    """SOH.SO_NUMBER, SOD.DUE_DATE, SMS.STATUS_CODE, 
        SMH.SM_NUMBER, 
        STR.QTY_RESERVED, STR.DATE_SCAN, PNM.PN, PNM.DESCRIPTION, 
        CMP.COMPANY_NAME, SYSUR.USER_NAME, 
        STM.STOCK_LINE, STM.STM_AUTO_KEY, STM.SERIAL_NUMBER, 
        LOC.LOCATION_CODE, WHS.WAREHOUSE_CODE, STR.STR_AUTO_KEY
    """
    smd_data,error,msg = [],'',''

    for rec in stock_recs:
        smd_data.append(wos_obj(
            session_id=session_id,
            wo_number = rec[0],#so_number
            due_date = rec[1] and rec[1][:10] or None,
            status = rec[2],#SM Status
            si_number = rec[3],#SM_NUMBER
            qty_reserved = rec[4] or 0,
            start_date = rec[5] and rec[5][:10] or None,#date_scan
            part_number = rec[6],
            description = rec[7],
            customer = rec[8],
            user_id = rec[9],
            stock_line = rec[10],
            stm_auto_key = rec[11] or 0,
            serial_number = rec[12],
            location_code = rec[13],
            wh_code = rec[14], 
            str_auto_key = rec[15],
            priority = rec[16] != '0' and rec[16] or '',
            cart = rec[17],
            notes = rec[18],       
            )
        )
    try:
        
        delete = wos_obj.objects.filter(session_id=session_id).delete()
        wos_obj.objects.bulk_create(smd_data) or None
    except Exception as err:
        logger.error("Error with creation of records for the grid. Message: '%s'",err.args)           
    return error,msg
    

@shared_task
def search_shipping(quapi_id,session_id,sysur_auto_key,filter_list=[],\
    smds_list=[],is_dashboard=False):
    error,msg = '',''
    from polls.models import WOStatus as wos_obj
    stock_recs = get_smd_stock(quapi_id,session_id,sysur_auto_key,\
        filter_list=filter_list,smds_list=smds_list,is_dashboard=is_dashboard)
    if stock_recs:             
        error,msg = add_smd_stock(wos_obj,session_id,stock_recs)        
    elif not error:
        error = 'No records found.'   
    return error,msg

def get_smd_stock(quapi_id,session_id,sysur_auto_key,\
    filter_list = [],smds_list = [],is_dashboard=False):
    error,msg,where_clause = '','',''
    from polls.models import QueryApi,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg

    where_clause = " STM.QTY_OH > 0"

    if smds_list:
        from polls.models import WOStatus
        smds = WOStatus.objects.filter(id__in=smds_list)
        strs_list = smds.values_list('str_auto_key',flat=True)
        upd_strs_list = construct_akl(strs_list)
        
        for strs_in in upd_strs_list:
            where_clause += " AND STR.STR_AUTO_KEY IN %s"%strs_in
    
    if not is_dashboard and filter_list:

        if filter_list[0]:#customer
            cmp = filter_list[0] + '%'
            where_clause += " AND UPPER(CMP.COMPANY_NAME) LIKE UPPER('%s')"%cmp
            
        if filter_list[1]:#SM Order#
            order = filter_list[1] + '%'     
            where_clause += " AND UPPER(SMH.SM_NUMBER) LIKE UPPER('%s')"%order
            
        if filter_list[2]:#entry date
            entry_date = filter_list[2]
            where_clause += " AND (STR.DATE_SCAN <= TO_DATE('%s','MM/DD/YYYY') OR STR.DATE_SCAN IS NULL)"%entry_date
            
        if filter_list[3]:#status
            if filter_list[3] == '0':
                where_clause += " AND SMS.SMS_AUTO_KEY IS NULL"
            else:
                where_clause += " AND SMS.SMS_AUTO_KEY='%s'"%filter_list[3]
                
        if filter_list[4]:#pn
            pn = filter_list[4] + '%'
            #Part Number same filter as RO Edit, but PNM for the STM        
            where_clause += """ AND UPPER(PNM.PN) LIKE UPPER('%s')"""%pn
            
        if filter_list[5]:#description
            desc = filter_list[5] + '%'
            #Part description        
            where_clause += """ AND UPPER(PNM.DESCRIPTION) LIKE UPPER('%s')"""%desc
            
        if filter_list[6]:#ship_via
            ship_via = filter_list[6]
            where_clause += """ AND SVC.SHIP_VIA_CODE = '%s'"""%ship_via
            
        if filter_list[8]:#location
            location = filter_list[8] + '%'
            where_clause += """ AND UPPER(LOC.LOCATION_CODE) LIKE UPPER('%s')"""%location
            
        if filter_list[9]:#whs
            whs = filter_list[9] + '%'     
            where_clause += """ AND UPPER(WHS.WAREHOUSE_CODE) LIKE UPPER('%s')"""%whs
            
        if filter_list[10]:#user
            if filter_list[10] == 'unassigned':
                where_clause += " AND SYSUR.USER_NAME IS NULL"
            else:                
                user = filter_list[10] + '%'
                where_clause += " AND UPPER(SYSUR.USER_NAME) LIKE UPPER('%s')"%user
                
        if filter_list[7]:#priority
            priority = filter_list[7] + '%'
            where_clause += " AND UPPER(SYSUR.USER_NAME) LIKE UPPER('%s')"%priority

    query = """SELECT DISTINCT SOH.SO_NUMBER AS ORDER_NUM, SOD.DUE_DATE AS DUE_DATE, SMS.STATUS_CODE, 
        SMH.SM_NUMBER, STR.QTY_RESERVED, STR.DATE_SCAN, PNM.PN, 
        PNM.DESCRIPTION, CMP.COMPANY_NAME, SYSUR.USER_NAME AS USER_NAME, 
        STM.STOCK_LINE, STM.STM_AUTO_KEY, STM.SERIAL_NUMBER, 
        LOC.LOCATION_CODE AS LOC, WHS.WAREHOUSE_CODE AS WHS, STR.STR_AUTO_KEY, SOH.PRIORITY,
        UDL.UDL_CODE, TO_CHAR(SMD.NOTES)        

        FROM SO_DETAIL SOD

        JOIN SO_HEADER SOH ON SOH.SOH_AUTO_KEY = SOD.SOH_AUTO_KEY
        JOIN STOCK_RESERVATIONS STR ON SOD.SOD_AUTO_KEY = STR.SOD_AUTO_KEY
        JOIN SM_DETAIL SMD ON STR.SMD_AUTO_KEY = SMD.SMD_AUTO_KEY
        JOIN SM_HEADER SMH ON SMD.SMH_AUTO_KEY = SMH.SMH_AUTO_KEY
        JOIN SM_STATUS SMS ON SMH.SMS_AUTO_KEY = SMS.SMS_AUTO_KEY
        JOIN STOCK STM ON STR.STM_AUTO_KEY = STM.STM_AUTO_KEY
        JOIN LOCATION LOC ON STM.LOC_AUTO_KEY = LOC.LOC_AUTO_KEY
        JOIN WAREHOUSE WHS ON STM.WHS_AUTO_KEY = WHS.WHS_AUTO_KEY
        JOIN COMPANIES CMP ON SMH.CMP_AUTO_KEY = CMP.CMP_AUTO_KEY
        JOIN PARTS_MASTER PNM ON STM.PNM_AUTO_KEY = PNM.PNM_AUTO_KEY
        LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = STM.IC_UDL_005
        LEFT JOIN SYS_USERS SYSUR ON STR.SYSUR_AUTO_KEY_SCAN = SYSUR.SYSUR_AUTO_KEY
        WHERE 
        %s
        
        UNION

        SELECT DISTINCT ROH.RO_NUMBER AS ORDER_NUM, SMH.SHIP_DATE AS DUE_DATE, SMS.STATUS_CODE, 
        SMH.SM_NUMBER, STR.QTY_RESERVED, STR.DATE_SCAN, PNM.PN, 
        PNM.DESCRIPTION, CMP.COMPANY_NAME, SYSUR.USER_NAME AS USER_NAME, 
        STM.STOCK_LINE, STM.STM_AUTO_KEY, STM.SERIAL_NUMBER, 
        LOC.LOCATION_CODE AS LOC, WHS.WAREHOUSE_CODE AS WHS, STR.STR_AUTO_KEY, 0,
        UDL.UDL_CODE, TO_CHAR(SMD.NOTES)

        FROM RO_DETAIL ROD
        
        JOIN RO_HEADER ROH ON ROH.ROH_AUTO_KEY = ROD.ROH_AUTO_KEY
        JOIN STOCK_RESERVATIONS STR ON ROD.ROD_AUTO_KEY = STR.ROD_AUTO_KEY
        JOIN SM_DETAIL SMD ON STR.SMD_AUTO_KEY = SMD.SMD_AUTO_KEY
        JOIN SM_HEADER SMH ON SMD.SMH_AUTO_KEY = SMH.SMH_AUTO_KEY
        JOIN SM_STATUS SMS ON SMH.SMS_AUTO_KEY = SMS.SMS_AUTO_KEY
        JOIN STOCK STM ON STR.STM_AUTO_KEY = STM.STM_AUTO_KEY
        JOIN LOCATION LOC ON STM.LOC_AUTO_KEY = LOC.LOC_AUTO_KEY
        JOIN WAREHOUSE WHS ON STM.WHS_AUTO_KEY = WHS.WHS_AUTO_KEY
        JOIN COMPANIES CMP ON SMH.CMP_AUTO_KEY = CMP.CMP_AUTO_KEY
        JOIN PARTS_MASTER PNM ON STM.PNM_AUTO_KEY = PNM.PNM_AUTO_KEY
        LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = STM.IC_UDL_005
        LEFT JOIN SYS_USERS SYSUR ON STR.SYSUR_AUTO_KEY_SCAN = SYSUR.SYSUR_AUTO_KEY
        WHERE
        %s
        
        UNION

        SELECT DISTINCT POH.PO_NUMBER AS ORDER_NUM, SMH.SHIP_DATE AS DUE_DATE, SMS.STATUS_CODE, 
        SMH.SM_NUMBER, STR.QTY_RESERVED, STR.DATE_SCAN, PNM.PN, 
        PNM.DESCRIPTION, CMP.COMPANY_NAME, SYSUR.USER_NAME AS USER_NAME, 
        STM.STOCK_LINE, STM.STM_AUTO_KEY, STM.SERIAL_NUMBER, 
        LOC.LOCATION_CODE AS LOC, WHS.WAREHOUSE_CODE AS WHS, STR.STR_AUTO_KEY, 0,
        UDL.UDL_CODE, TO_CHAR(SMD.NOTES)

        FROM PO_DETAIL POD

        LEFT JOIN PO_HEADER POH ON POH.POH_AUTO_KEY = POD.POH_AUTO_KEY
        JOIN STOCK_RESERVATIONS STR ON POD.POD_AUTO_KEY = STR.POD_AUTO_KEY
        JOIN SM_DETAIL SMD ON STR.SMD_AUTO_KEY = SMD.SMD_AUTO_KEY
        JOIN SM_HEADER SMH ON SMD.SMH_AUTO_KEY = SMH.SMH_AUTO_KEY
        JOIN SM_STATUS SMS ON SMH.SMS_AUTO_KEY = SMS.SMS_AUTO_KEY
        JOIN STOCK STM ON STR.STM_AUTO_KEY = STM.STM_AUTO_KEY
        JOIN LOCATION LOC ON STM.LOC_AUTO_KEY = LOC.LOC_AUTO_KEY
        JOIN WAREHOUSE WHS ON STM.WHS_AUTO_KEY = WHS.WHS_AUTO_KEY
        JOIN COMPANIES CMP ON SMH.CMP_AUTO_KEY = CMP.CMP_AUTO_KEY
        JOIN PARTS_MASTER PNM ON STM.PNM_AUTO_KEY = PNM.PNM_AUTO_KEY
        LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = STM.IC_UDL_005
        LEFT JOIN SYS_USERS SYSUR ON STR.SYSUR_AUTO_KEY_SCAN = SYSUR.SYSUR_AUTO_KEY
        WHERE
        %s
        ORDER BY USER_NAME, WHS, LOC, DUE_DATE ASC, ORDER_NUM
    """%(where_clause,where_clause,where_clause)
    recs = selection_dir(query,cr)   
    return recs
             
@shared_task
def create_shipping(quapi_id,session_id,sysur_auto_key,filter_list):
    error,msg = '',''
    from polls.models import QueryApi,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    query = """SELECT 
    
    FROM 
    """    
    return error,msg            
def lot_import_grid_rows(grid_rows,session_id):
    from polls.models import WOStatus
    req_data,error = [],''
    for row in grid_rows:       
        req_data.append(WOStatus(
            wo_number = row[0],
            part_number = row[2],
            serial_number = row[3],
            description = row[6],              
            stock_line = row[-2],
            slug = row[-1],
            session_id = session_id,            
        ))
    if req_data:     
        try:
            delete = WOStatus.objects.filter(session_id=session_id).delete()
            rec = WOStatus.objects.bulk_create(req_data) or []    
        except Exception as exc:
            error += "Error, %s, with creating rows for grid. %s"%(exc,row[0])
    return error

@shared_task
def lot_create(quapi_id,user_id,sysur_auto_key,session_id): 
    error,msg,fail_msg,cr,got_there = '','','',None,False  
    good_rows = {} 
    from polls.models import QueryApi,Document  
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None                                                                                                      
    import_file = Document.objects.filter(session_id=session_id)
    import_file = import_file and import_file[0] or None
    file_path = import_file and os.path.join(import_file.docfile.path) or ''
    from openpyxl import load_workbook
    wb = load_workbook(filename = file_path)
    sheet = wb.active
    sheet_rows = sheet.iter_rows()
    col_count = 0
    col_headings = []
    line_count = 0
    grid_rows = []        
    #loop through each row and process
    position = 'Initial'
    for row in sheet_rows:
        if not row[0].value:
            break     
        
        if line_count == 0:
            col_headings = row
            line_count += 1
            continue
        
        dict_row = {}
        col_count = 0
        for col in col_headings:
            dict_row[col.value] = row[col_count].value
            col_count += 1                                 
        bad_seq = False
        line_count += 1
        location = dict_row.get('Location','')
        consignment = dict_row.get('Consignment Code','') 
        lot_no = dict_row.get('Lot #','')
        unit_cost = dict_row.get('Unit Cost',0)                                  
        if not unit_cost:
            unit_cost = 0
        if isinstance(unit_cost,str):
            if not unit_cost.isnumeric():
                unit_cost = 0             
        pn = dict_row.get('Part Number','')
        pn = pn.replace("'","''")                                                         
        desc = dict_row.get('Description','')
                         
        if isinstance(desc,str):
            if len(desc) > 49:
                desc = desc[:49]
        serial_no = dict_row.get('Serial Number','')                                                                                 
        if serial_no:
            if isinstance(serial_no,str) and len(serial_no) > 39:
                serial_no = serial_no[:39]
                
        quantity = dict_row.get('Qty','') 
        condition = dict_row.get('Condition','')#condition_code 
        part_cert_num = dict_row.get('Part Cert #','')
        if isinstance(part_cert_num,str):
            if len(part_cert_num) > 39:
                part_cert_num = part_cert_num[:39]  
        part_cert_by = dict_row.get('Certified by (MISCELLANEOUS)','')
        if isinstance(part_cert_by,str):
            if len(part_cert_by) > 39:
                part_cert_by = part_cert_by[:39]       
        notes = dict_row.get('Notes','') 
        from datetime import date                       
        tag_date = dict_row.get('Tag Date','')#tag date
        if not isinstance(tag_date,date):
            tag_date = ''
        remarks = dict_row.get('REMARKS','')
        if isinstance(remarks,str):
            if len(remarks) > 49:
                remarks = remarks[:49] 
        packaging = dict_row.get('Packaging (UDF)','') 
        if isinstance(packaging,str):
            if len(packaging) > 39:
                packaging = packaging[:39] 
        weight = dict_row.get('Weight (UDF)','')
        length = dict_row.get('LENGTH','')
        if isinstance(length,str):
            if not length.isnumeric():
                length = ''
        width = dict_row.get('WIDTH','')
        if isinstance(width,str):
            if not width.isnumeric():
                width = ''
        height = dict_row.get('HEIGHT','')  
        if isinstance(height,str):
            if not height.isnumeric():
                height = ''            
        rec_date = dict_row.get('OnDock (REC DATE)','')
        if not isinstance(rec_date,date):
            rec_date = '' 

        hold_line = dict_row.get('Hold Line','F')  
        lot_alw_precost = dict_row.get('PRECOSTED','F')
        lot_apl_ro_cost = dict_row.get('APPLY RO COST','F')
        traceable_to = dict_row.get('Traceable To','')
        obtained_from = dict_row.get('Obtained From','')
        item_num = dict_row.get('ITEM','')

        import_row = [lot_no,quantity,pn,serial_no,session_id]
        import_row += [notes,desc,location,'',consignment,lot_no]
        import_row += [unit_cost,condition,part_cert_num,notes]
        import_row += [tag_date,remarks,packaging,weight,'']
        import_row += [part_cert_by,rec_date,lot_alw_precost]
        import_row += [lot_apl_ro_cost,length,width,height,desc,hold_line] 
        import_row += [traceable_to,obtained_from,item_num]        
        import_row += [line_count-1]            

        if not lot_no:           
            error += ' Line %s has no lot.'%(line_count-1)
        if not pn:           
            error += ' Line %s has no pn.'%(line_count-1)
        if not quantity:           
            error += ' Line %s has no qty.'%(line_count-1)
        if not location:           
            error += ' Line %s has no location.'%(line_count-1)

        lot_error,msg,show_msg = lot_teardown(quapi_id,\
            sysur_auto_key,user_id,import_row,lot_import=True,\
            line_count=line_count-1)                                        
        error += lot_error
        import_row.append(lot_error)
        
        if lot_error:
            grid_rows.append(import_row)
            
    error = lot_import_grid_rows(grid_rows,session_id) 
    
    if len(grid_rows) < line_count-1:                                          
        msg = 'Successfully imported ' 
        msg += str(line_count -1 -len(grid_rows)) 
        msg += ' rows. ' + str(len(grid_rows))
        msg += ' rows rejected.'
        
    else:
        msg = 'Could not import any rows. Please see rejection reasons.'                                                           
    return error,msg,fail_msg

@shared_task
def update_expiry(quapi_id,conn_string,cmp_auto_key,expiry_date):
    error,msg = '',''
    cr,con = orcl_phone_home(conn_string)
    if not (cr and con):
        return 'Cannot connect to Oracle'
    query = """UPDATE COMPANIES SET 
    APPROVAL_EXPIRE=TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss')
    WHERE CMP_AUTO_KEY = %s   
    """%(expiry_date,cmp_auto_key)
    error = updation_dir(query,cr) 
    if error == '{"recs": ""}':
        error = ''
        msg = 'Expiration date updated.'        
    orcl_commit(con)
    return error,msg

@shared_task
def exchange_stock(quapi_id,session_id,sysur_auto_key,keepers,shippers,wo_number):
    error,msg = '',''
    from polls.models import QueryApi,WOStatus    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    #************************************************************************
    """ 3. Update the Record(s) to Ship
            a. Reduce the Unit Cost on the Record(s) to Ship by the opposite of its Original PO Cost. 
                i. This should call the Cost Adjustment procedure
        4. Then update the Record(s) to Keep 
            a. Direct Update to STOCK[‘CUSTOMER_OWNED_FLAG’] =’F’, 
            b. Cost adjustment procedure to Unit Cost = opposite adjustment of the Record to Ship, 
            c. Direct Update to Original PO Cost = Original PO Cost of Record to Ship
    """
    #************************************************************************
    count = 0
    for keep_stm in keepers:
        oc_keeper = """SELECT REC_COST FROM STOCK 
            WHERE STM_AUTO_KEY = %s"""%keep_stm
        orig_kc = selection_dir(oc_keeper,cr)
        orig_kc = orig_kc and orig_kc[0] and orig_kc[0][0] or 0.00
        orig_kc = -orig_kc
        oc_shipper = """SELECT REC_COST,UNIT_COST FROM STOCK 
            WHERE STM_AUTO_KEY = %s"""%shippers[count]
        shipper_costs = selection_dir(oc_shipper,cr)
        
        orig_sc = shipper_costs and shipper_costs[0] and shipper_costs[0][0] or 0.00
        unit_sc = shipper_costs and shipper_costs[0] and shipper_costs[0][1] or 0.00
        #print('unit: ' + str(unit_sc) + ' orig: ' + str(orig_sc))
        if unit_sc < orig_sc:
            orig_sc = unit_sc             
        neg_orig_sc = orig_sc and -orig_sc or 0.00
        """Call Cost Adjustment Procedure - adjust up the unit_cost of the keeper"""
        squery = """
            DECLARE CT NUMBER;
            BEGIN CT := QC_STOCK_PKG.SPI_STOCK_ADJUST_COST_ORIG(%s, %s, 2, 3);
            END;"""%(shippers[count], neg_orig_sc)
        error = updation_dir(squery,cr)
        #print('cost adj error (if any): ' + error)        
        query = """UPDATE STOCK SET REC_COST = REC_COST + %s,
            CUSTOMER_OWNED ='F'
            WHERE STM_AUTO_KEY = '%s'"""%(orig_sc,keep_stm)
        error = updation_dir(query,cr)
        ###print('upd rec_cost error (if any): ' + error) 
        """
        1. So we insert a new SOH if there isn't already one with SOH['COMPANY_REF_NUMBER'] = SI_NUMBER from the keeper.
            UPDATE WO_OPERATION SET WO_TYPE = 'Internal' WHERE WOO = WOO OF "SHIPPER" 
            UPDATE WO_OPERATION SET CMP_AUTO_KEY = (SELECT CMP_AUTO_KEY FROM COMPANIES WHERE COMPANY_NAME = (SELECT COMPANY_NAME FROM SYS_COMPANIES 
            WHERE SYCM_AUTO_KEY = SYSCM_AUTO_KEY 0F WOO OF "SHIPPER")
        2. Insert new SOD into the SOH where SOH['COMPANY_REF_NUMBER'] = SI_NUMBER of the keeper.  
            The data for the new SOD comes from the shipper rather than the keeper.
        3.   Reserve the shipper STM to the SOD just created in step 2.
        4. For each subsequent exchange with the same SI_NUMBER entered (or left) in the label field, repeat steps 1-3.
        """
        query = """SELECT CMP_AUTO_KEY,CUR_AUTO_KEY,SVC_AUTO_KEY,ATTENTION,
        COMPANY_REF_NUMBER,DUE_DATE,FAX_NUMBER,PHONE_NUMBER,CUR_AUTO_KEY,
        PRIORITY,DPT_AUTO_KEY,SYSCM_AUTO_KEY,EMAIL_ADDRESS,'TAX_AMOUNT',
        'TXT_AUTO_KEY','TAX_METHOD',BILL_ADDRESS1,BILL_ADDRESS2,BILL_ADDRESS3,
        BILL_ADDRESS4,BILL_ADDRESS5,SHIP_ADDRESS1,SHIP_ADDRESS2,SHIP_ADDRESS3,
        SHIP_ADDRESS4,SHIP_ADDRESS5,'TMC_AUTO_KEY','SPN_AUTO_KEY',CST_AUTO_KEY,
        BILL_NAME,SHIP_NAME,GEO_AUTO_KEY,ACT_AUTO_KEY,SVA_AUTO_KEY,'CTH_AUTO_KEY',
        'FOB',SCC_AUTO_KEY,COC_AUTO_KEY,WOO_AUTO_KEY FROM WO_OPERATION WHERE
        UPPER(SI_NUMBER) = UPPER('%s')"""%wo_number
        woo = selection_dir(query,cr)
        #print('woo select: ' + str(woo)) 
        woo = woo and woo[0] or None
        if not woo:
            error = 'No WO found.' 
        else:
            cmp_auto_key = woo[0] or 1
            cur_auto_key = woo[1]
            svc_auto_key = woo[2]
            attention = woo[3]
            company_ref_number = woo[4]
            if not company_ref_number:
                error = 'No customer ref#'
            else:
                query = """SELECT SOH_AUTO_KEY FROM SO_HEADER WHERE UPPER(COMPANY_REF_NUMBER) = UPPER('%s')
                    AND SOS_AUTO_KEY NOT IN (SELECT SOS_CLOSED FROM SO_CONTROL)
                """%company_ref_number
                soh = selection_dir(query,cr)
                soh_auto_key = soh and soh[0] and soh[0][0] or None                
                fax_number = woo[6]
                phone_number = woo[7]
                query = """SELECT EXCHANGE_RATE 
                    FROM CURRENCY WHERE CUR_AUTO_KEY = %s
                    """%woo[8]
                rate = selection_dir(query,cr)
                exchange_rate = rate and rate[0] and rate[0][0] or 1.0
                priority = woo[9]
                dpt_auto_key = woo[10]
                syscm_auto_key = woo[11]
                email_address = woo[12]
                tax_amount = 0.00
                txt_auto_key = ''
                tax_method = ''
                bill_address1 = woo[16]
                bill_address2 = woo[17]
                bill_address3 = woo[18]
                bill_address4 = woo[19]
                bill_address5 = woo[20]
                ship_address1 = woo[21]
                ship_address2 = woo[22]
                ship_address3 = woo[23]
                ship_address4 = woo[24]
                ship_address5 = woo[25]
                tmc_auto_key = ''
                spn_auto_key = ''
                cst_auto_key = woo[28]
                bill_name = woo[29]
                ship_name = woo[30]
                geo_auto_key = woo[31]
                act_auto_key = woo[32]
                sva_auto_key = woo[33]
                cth_auto_key = ''
                fob = ''
                scc_auto_key = woo[36]
                coc_auto_key = woo[37]
                woo_auto_key = woo[38]
                dtc_auto_key = ''
                open_flag = 'T'
                track_changes = 'T'
                date_format = '%Y-%m-%d %H:%M:%S'  
                query = "SELECT SYSTIMESTAMP FROM DUAL"
                today = selection_dir(query,cr)
                today = today and today[0] and today[0][0] and today[0][0][:18]
                today = datetime.strptime(today,date_format)
                right_now = today.strftime("%m/%d/%Y %H:%M:%S")
                if woo[5]:
                    due_date = datetime.strptime(woo[5][:10],'%Y-%m-%d')
                    due_date = due_date.strftime('%m/%d/%Y') 
                else:
                    due_date = right_now[:10]
                entry_date = right_now
                rec_date = right_now
                sysur_modified = sysur_auto_key
                date_created = right_now
                notes ='N/A'
                historical_flag = 'F'
                closed_update = 'F'
                converted = 'F'
                #if no soh found, insert the SOH
                if not soh_auto_key:
                    query = """SELECT NUMBER_PREFIX || TO_CHAR(LAST_NUMBER+1)  FROM SYS_NUMBER_LOG
                        WHERE SYSNLC_AUTO_KEY 
                        = (SELECT SYSNLC_AUTO_KEY FROM SYS_NUMBER_LOG_CODES WHERE LOG_TYPE_CODE = 'SO')
                        """  
                    so_number = selection_dir(query,cr)
                    so_number = so_number and so_number[0] and so_number[0][0]
                    query = """INSERT INTO SO_HEADER (SOH_AUTO_KEY,CMP_AUTO_KEY,SYSUR_AUTO_KEY,CUR_AUTO_KEY,
                           SVC_AUTO_KEY,SO_NUMBER,ATTENTION,COMPANY_REF_NUMBER,DUE_DATE,
                           ENTRY_DATE,FAX_NUMBER,HISTORICAL_FLAG,REC_DATE,
                           NOTES,OPEN_FLAG,PHONE_NUMBER,EXCHANGE_RATE,
                           PRIORITY,CLOSED_UPDATE,DPT_AUTO_KEY,SYSCM_AUTO_KEY,EMAIL_ADDRESS,TAX_AMOUNT,
                           TXT_AUTO_KEY,TAX_METHOD,BILL_ADDRESS1,BILL_ADDRESS2,BILL_ADDRESS3,BILL_ADDRESS4,
                           BILL_ADDRESS5,SHIP_ADDRESS1,SHIP_ADDRESS2,SHIP_ADDRESS3,SHIP_ADDRESS4,SHIP_ADDRESS5,
                           TMC_AUTO_KEY,SPN_AUTO_KEY,CST_AUTO_KEY,BILL_NAME,SHIP_NAME,CONVERTED,GEO_AUTO_KEY,
                           ACT_AUTO_KEY,SVA_AUTO_KEY,CTH_AUTO_KEY,FOB,USR_AUTO_KEY,TRACK_CHANGES,SCC_AUTO_KEY,
                           COC_AUTO_KEY,DTC_AUTO_KEY,SYSUR_MODIFIED,DATE_CREATED)
                    VALUES(G_SOH_AUTO_KEY.NEXTVAL,'%s','%s','%s','%s','%s','%s','%s',TO_DATE('%s','mm/dd/yyyy'),
                    TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),
                    '%s','%s',TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),
                    '%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s',
                    '%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s',
                    '%s','%s','%s','%s','%s','%s',TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'))
                    """%(cmp_auto_key,sysur_auto_key,cur_auto_key,svc_auto_key,so_number,attention,
                    company_ref_number,due_date,entry_date,fax_number,historical_flag,rec_date,notes,
                    open_flag,phone_number,exchange_rate,priority,closed_update,dpt_auto_key,syscm_auto_key,
                    email_address,tax_amount,txt_auto_key,tax_method,bill_address1,bill_address2,bill_address3,
                    bill_address4,bill_address5,ship_address1,ship_address2,ship_address3,ship_address4,
                    ship_address5,tmc_auto_key,spn_auto_key,cst_auto_key,bill_name,ship_name,converted,geo_auto_key,
                    act_auto_key,sva_auto_key,cth_auto_key,fob,sysur_auto_key,track_changes,scc_auto_key,coc_auto_key,
                    dtc_auto_key,sysur_modified,date_created)            
                    error = insertion_dir(query,cr)
                    #print('error (if any) - insert into so header: ' + error)
                    if not error:
                        query = """SELECT SOH_AUTO_KEY FROM SO_HEADER WHERE ROWNUM <= 1 ORDER BY SOH_AUTO_KEY DESC
                        """
                        soh = selection_dir(query,cr)
                        soh_auto_key = soh and soh[0] and soh[0][0] or None
                        query = """UPDATE SYS_NUMBER_LOG SET LAST_NUMBER = LAST_NUMBER + 1 WHERE SYSNLC_AUTO_KEY 
                            = (SELECT SYSNLC_AUTO_KEY FROM SYS_NUMBER_LOG_CODES WHERE LOG_TYPE_CODE = 'SO')
                        """
                        error = updation_dir(query,cr)
                if soh_auto_key:
                    query = """SELECT S.PNM_AUTO_KEY,S.PCC_AUTO_KEY,
                        S.CNC_AUTO_KEY,S.SERIAL_NUMBER,S.QTY_OH,
                        S.STC_AUTO_KEY,P.UOM_AUTO_KEY,P.LIST_PRICE FROM STOCK S
                        LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
                        WHERE S.STM_AUTO_KEY = %s"""%shippers[count]
                        
                    sp_stock = selection_dir(query,cr)
                    #print('SELECT SP STOCK'  + str(sp_stock))
                    sp_stock = sp_stock and sp_stock[0] or None
                    pnm_auto_key = sp_stock[0]
                    pcc_auto_key = sp_stock[1]
                    cnc_auto_key = sp_stock[2]
                    serial_number = sp_stock[3]
                    qty_ordered = sp_stock[4]
                    stc_auto_key = sp_stock[5]
                    uom_auto_key = sp_stock[6]
                    list_price = sp_stock[7]
                    alt_pnm_auto_key = sp_stock[0]
                    entry_date = right_now[:10]
                    ship_date = due_date
                    buy_flag = 'F'
                    route_code = 'S'
                    route_desc = 'Part Sale'            
                    sell_as_type = 'M'
                                
                    #insert the detail line
                    query = """INSERT INTO SO_DETAIL (SOD_AUTO_KEY,PNM_AUTO_KEY,PCC_AUTO_KEY,
                       UOM_AUTO_KEY,CNC_AUTO_KEY,
                       SOH_AUTO_KEY,SCC_AUTO_KEY,ALT_PNM_AUTO_KEY,SYSUR_AUTO_KEY,BUY_FLAG,ENTRY_DATE,
                       EXCHANGE_RATE,LIST_PRICE,SELL_AS_TYPE,SERIAL_NUMBER,SHIP_DATE,QTY_ORDERED,CLOSED_UPDATE,
                       WOO_AUTO_KEY,CONVERTED,ROUTE_CODE,ROUTE_DESC,STC_AUTO_KEY,SYSUR_MODIFIED,DATE_CREATED) 
                    VALUES(G_SOD_AUTO_KEY.NEXTVAL,'%s','%s','%s','%s','%s','%s','%s','%s','%s',
                    TO_DATE('%s','mm/dd/yyyy'),
                    '%s','%s','%s','%s',TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),'%s','%s','%s','%s',
                    '%s','%s','%s','%s',TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'))          
                    """%(pnm_auto_key,pcc_auto_key,uom_auto_key,cnc_auto_key,soh_auto_key,scc_auto_key,
                        alt_pnm_auto_key,sysur_auto_key,buy_flag,entry_date,exchange_rate,list_price,
                        sell_as_type,serial_number,ship_date,qty_ordered,closed_update,woo_auto_key,
                        converted,route_code,route_desc,stc_auto_key,sysur_auto_key,right_now)
                    error = insertion_dir(query,cr)
                    #print('INSERT INTO SO DETAIL: '  + error)
                    sod_sub = "SELECT SOD_AUTO_KEY FROM SO_DETAIL WHERE ROWNUM<=1 ORDER BY SOD_AUTO_KEY DESC"
                    sod = selection_dir(sod_sub,cr)
                    sod_auto_key = sod and sod[0] and sod[0][0] or None
                    
                    if sod_auto_key:
                        squery = """INSERT INTO STOCK_RESERVATIONS 
                            (STR_AUTO_KEY,STM_AUTO_KEY,SOD_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY,QTY_SHIP) 
                            VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s,%s)
                            """%(shippers[count],sod_auto_key,qty_ordered,sysur_auto_key,qty_ordered)
                        error = insertion_dir(squery,cr)
                        ##print('INSERT INTO STOCK_RESERVATIONS: '  + error)                        
    aud_status = 'failure'
    if not error or error == '{"recs": ""}':
        #update all keeper stm CUSTOMER_OWNED = 'F'
        keeper_stock = WOStatus.objects.filter(session_id=session_id,wo_type = 'KEEPER')
        stm_ids = keeper_stock.values_list('stm_auto_key',flat=True)
        stm_ids = construct_akl(stm_ids)
        for stm_list in stm_ids:
            query = """UPDATE STOCK SET CUSTOMER_OWNED='F'
                WHERE STM_AUTO_KEY IN %s AND CUSTOMER_OWNED <> 'F'
                """%stm_list
            error = updation_dir(query,cr)
        keeper_rec = keeper_stock and keeper_stock[0] or None
        woo_key = keeper_rec and keeper_rec.woo_auto_key
        sycm_key = keeper_rec and keeper_rec.syscm_auto_key
        if sycm_key and woo_key:     
            query = """UPDATE WO_OPERATION SET WO_TYPE = 'Internal',CMP_AUTO_KEY = 
                (SELECT CMP_AUTO_KEY FROM COMPANIES 
                WHERE COMPANY_NAME=(SELECT COMPANY_NAME FROM SYS_COMPANIES
                WHERE SYSCM_AUTO_KEY=%s)) WHERE WOO_AUTO_KEY = %s AND 
                (WO_TYPE <> 'Internal' OR CMP_AUTO_KEY <> 
                (SELECT CMP_AUTO_KEY FROM COMPANIES 
                WHERE COMPANY_NAME=(SELECT COMPANY_NAME FROM SYS_COMPANIES
                WHERE SYSCM_AUTO_KEY=%s)))"""%(sycm_key,woo_key,sycm_key)
            error = updation_dir(query,cr)  
        if error == '{"recs": ""}':            
            error = ''
            aud_status = 'success'
            error,msg = get_exchange_stock(quapi_id,session_id,wo_number,cr=cr)
            msg = 'Successful exchange.' 
            orcl_commit(con=con)
    right_now = datetime.now()
    right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')        
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='exchange-portal')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:       
        field_changed = 'Exchanged stock via Exchange Portal app. '
        new_val = field_changed + error
        field_changed += error
        error += register_audit_trail(user_rec,field_changed,new_val,right_now,app_id,quapi,status=aud_status)  
    else:
        error = 'Incorrect Quantum User ID.'          
    return error,msg

def create_exch_stock(session_id,recs,delete_recs=False):
    from polls.models import WOStatus
    stock_data,error = [],''
    for row in recs:     
        stock_data.append(WOStatus(
            session_id = session_id,
            stock_line = row[0],
            part_number = row[1],
            description = row[2],
            serial_number = row[3],
            consignment_code = row[4],
            qty_oh = row[5],
            parts_cost = row[6],
            total_cost = row[7],
            pnm_auto_key = row[8], 
            woo_auto_key = row[9] or 0,
            wo_type = row[10],
            stm_auto_key = row[11],
            syscm_auto_key = row[12] or 1,
            account_company = row[13],            
        ))
    if stock_data:     
        try:
            if delete_recs:
                delete = WOStatus.objects.filter(session_id=session_id).delete()
            WOStatus.objects.bulk_create(stock_data) or []    
        except Exception as exc:
            error += "Error, %s, with saving stock records locally. %s"%(exc,row[0])
    return error

@shared_task
def get_exchange_stock(quapi_id,session_id,wo_number,cr=None):
    error,msg = '',''
    if not cr:
        from polls.models import QueryApi,WOStatus    
        quapi = QueryApi.objects.filter(id=quapi_id)
        quapi = quapi and quapi[0] or None
        cr,con = get_cursor_con(quapi)
    """
    1. Record(s) to Keep
        a. User will enter a WO # to locate STM(s). STM(s) will be reserved to WOO(s).
        b. Display list of STMs reserved to the WO entered
    2. Record(s) to Ship
        a. Display list of STMs where STOCK[‘QTY_RESERVED’] = 0 for the PN found from the User’s WO entry
    """
    query = """SELECT S.STOCK_LINE, P.PN, P.DESCRIPTION,
        S.SERIAL_NUMBER,CC.CONSIGNMENT_CODE,S.QTY_OH,
        S.UNIT_COST,S.REC_COST,P.PNM_AUTO_KEY,
        WO.WOO_AUTO_KEY,'KEEPER',S.STM_AUTO_KEY,WO.SYSCM_AUTO_KEY,
        'comp'
        FROM WO_OPERATION WO
        LEFT JOIN STOCK_RESERVATIONS SR ON WO.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
        LEFT JOIN STOCK S ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
        LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
        LEFT JOIN CONSIGNMENT_CODES CC ON CC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
        WHERE UPPER(WO.SI_NUMBER) = UPPER('%s')
    """%wo_number
    keeper_stock = selection_dir(query,cr)
    if keeper_stock:
        error = create_exch_stock(session_id,keeper_stock,delete_recs=True)
        pnm_auto_key = keeper_stock[0][8]
        woo_auto_key = keeper_stock[0][9]
        if pnm_auto_key:
            query = """SELECT S.STOCK_LINE, P.PN, P.DESCRIPTION,
                S.SERIAL_NUMBER,CC.CONSIGNMENT_CODE,S.QTY_OH,
                S.UNIT_COST,S.REC_COST,P.PNM_AUTO_KEY,
                WO.WOO_AUTO_KEY,'SHIPPER',S.STM_AUTO_KEY,WO.SYSCM_AUTO_KEY,
                C.COMPANY_ABBREV
                FROM STOCK S
                LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
                LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
                JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
                LEFT JOIN CONSIGNMENT_CODES CC ON CC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
                LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = S.CMP_AUTO_KEY
                WHERE P.PNM_AUTO_KEY = %s AND S.QTY_RESERVED = 0
                AND S.QTY_OH > 0
                AND (WO.WOO_AUTO_KEY <> %s OR WO.WOO_AUTO_KEY IS NULL)
            """%(pnm_auto_key,woo_auto_key)    
            shipper_stock = selection_dir(query,cr)
            error = create_exch_stock(session_id,shipper_stock)
        else:
            error = 'Reservation not found.'
    else:
        error = 'Stock Reserved to WO not found.'
    return error,msg
    
@shared_task
def create_new_wo(quapi_id,session_id,user_id,sysur_auto_key,stm_auto_key,si_number,ctrl_number,ctrl_id,is_default_repair):
    error,msg,new_woo_key = '','',''
    from polls.models import QueryApi,WOTask    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    today = datetime.now()
    timestamp = today.strftime('%m/%d/%Y hh:mm:ss')
    datestamp = today.strftime('%m/%d/%Y') 
    if ctrl_number and ctrl_id:    
        woo_sub = """SELECT WO.WOO_AUTO_KEY,WT.WOT_AUTO_KEY,WO.SI_NUMBER,S.STM_AUTO_KEY,
                       S.GLA_AUTO_KEY,WO.GLA_AUTO_KEY,S.PNM_AUTO_KEY,WO.DPT_AUTO_KEY,
                       WO.CMP_AUTO_KEY,WO.SYSCM_AUTO_KEY,WO.KIT_QTY,WO.OPM_AUTO_KEY,WOB.NEED_DATE,
                       WO.ATTENTION,WO.ECD_METHOD,WO.WWT_AUTO_KEY,WO.SVC_AUTO_KEY,
                       WO.NEW_WIP_ACCT,WO.BGS_DEFAULT,WO.CUR_AUTO_KEY,WO.GLA_LABOR,WO.GLA_MISC,
                       WO.LOT_CORE_SETTINGS,S.RECEIVER_NUMBER,WO.PCC_AUTO_KEY,WO.CNC_AUTO_KEY,
                       S.LOC_AUTO_KEY,S.STC_AUTO_KEY,S.CTS_AUTO_KEY,S.WHS_AUTO_KEY,S.STM_ORIGINAL,
                       WO.LOT_APL_RO_COST,WO.LOT_ALW_PRECOST,WO.LOT_REQ_INSPECTION,WO.LOT_COST_DELAYED,
                       S.ORIGINAL_PO_NUMBER,S.SERIAL_NUMBER,WOB.WOB_AUTO_KEY,S.QTY_OH,PM.DESCRIPTION,PM.PN 
                       FROM WO_TASK WT
                       LEFT JOIN WO_BOM WOB ON WT.WOT_AUTO_KEY = WOB.WOT_AUTO_KEY
                       LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WT.WOO_AUTO_KEY
                       LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
                       LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
                       LEFT JOIN PARTS_MASTER PM ON PM.PNM_AUTO_KEY = S.PNM_AUTO_KEY
                       WHERE S.CTRL_NUMBER = %s AND S.CTRL_ID = %s"""%(ctrl_number,ctrl_id)
                       
    elif si_number:
        woo_sub = """SELECT WO.WOO_AUTO_KEY,'',WO.SI_NUMBER,S.STM_AUTO_KEY,
                       S.GLA_AUTO_KEY,WO.GLA_AUTO_KEY,S.PNM_AUTO_KEY,WO.DPT_AUTO_KEY,
                       WO.CMP_AUTO_KEY,WO.SYSCM_AUTO_KEY,WO.KIT_QTY,WO.OPM_AUTO_KEY,WOB.NEED_DATE,
                       WO.ATTENTION,WO.ECD_METHOD,WO.WWT_AUTO_KEY,WO.SVC_AUTO_KEY,
                       WO.NEW_WIP_ACCT,WO.BGS_DEFAULT,WO.CUR_AUTO_KEY,WO.GLA_LABOR,WO.GLA_MISC,
                       WO.LOT_CORE_SETTINGS,S.RECEIVER_NUMBER,WO.PCC_AUTO_KEY,WO.CNC_AUTO_KEY,
                       S.LOC_AUTO_KEY,S.STC_AUTO_KEY,S.CTS_AUTO_KEY,S.WHS_AUTO_KEY,S.STM_ORIGINAL,
                       WO.LOT_APL_RO_COST,WO.LOT_ALW_PRECOST,WO.LOT_REQ_INSPECTION,WO.LOT_COST_DELAYED,
                       S.ORIGINAL_PO_NUMBER,S.SERIAL_NUMBER,WOB.WOB_AUTO_KEY,S.QTY_OH,PM.DESCRIPTION,PM.PN  
                       FROM WO_OPERATION WO
                       LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
                       LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
                       LEFT JOIN PARTS_MASTER PM ON PM.PNM_AUTO_KEY = S.PNM_AUTO_KEY
                       WHERE UPPER(WO.SI_NUMBER) = UPPER('%s')"""%(si_number)
                       # Lot WO's only - AND WO.WO_TYPE = 'Lot'
    wot_data = selection_dir(woo_sub,cr)

    if not wot_data:
        error = 'Cannot create a new WO.'
        return error,msg,new_woo_key
    else:
        woo_auto_key = wot_data[0][0] or ''
        wot_auto_key = wot_data[0][1] or ''
        si_number = wot_data[0][2] or '' 
        stm_auto_key = wot_data[0][3] or ''
        stm_gla_key = wot_data[0][4] or ''
        woo_gla_key = wot_data[0][5] or ''
        pnm_auto_key = wot_data[0][6] or ''
        dpt_auto_key = wot_data[0][7] or ''
        cmp_auto_key = wot_data[0][8] or ''
        syscm_auto_key = wot_data[0][9] or ''
        kit_qty = wot_data[0][10] or ''
        opm_auto_key = wot_data[0][11] or ''
        due_date = wot_data[0][12] or ''
        attention = wot_data[0][13] or ''
        ecd_method = wot_data[0][14] or ''
        wwt_auto_key = wot_data[0][15] or ''
        svc_auto_key = wot_data[0][16] or ''
        new_wip_acct = wot_data[0][17] or ''
        bgs_default = wot_data[0][18] or ''
        cur_auto_key = wot_data[0][19] or ''
        gla_labor = wot_data[0][20] or ''
        gla_misc = wot_data[0][21] or ''
        lot_core_settings = wot_data[0][22] or ''
        receiver_number = wot_data[0][23] or ''
        pcc_auto_key = wot_data[0][24] or ''
        cnc_auto_key = wot_data[0][25] or ''    
        loc_auto_key = wot_data[0][26] or ''
        stc_auto_key = wot_data[0][27] or ''
        cts_auto_key = wot_data[0][28] or ''
        whs_auto_key = wot_data[0][29] or ''
        stm_parent = wot_data[0][30] or ''
        lot_apl_ro_cost = wot_data[0][31] or ''
        lot_alw_precost = wot_data[0][32] or ''
        lot_req_inspection = wot_data[0][33] or ''
        lot_cost_delayed = wot_data[0][34] or ''
        original_po_number = wot_data[0][35] or ''
        serial_number = wot_data[0][36] or ''
        wob_auto_key = wot_data[0][37] or ''
        quantity = wot_data[0][38] or ''
        description = wot_data[0][39] or ''
        pn = wot_data[0][40] or ''
        if not (stm_gla_key and woo_gla_key):        
            query = "SELECT GLA_AUTO_KEY FROM GL_DISTRIBUTION WHERE GLD_AUTO_KEY = (SELECT GLD_INV FROM SYS_COMPANIES WHERE SYSCM_AUTO_KEY = 1)"
            gla = selection_dir(query,cr)
            gla_auto_key = gla and gla[0] and gla[0][0] or None               
            if gla_auto_key:
                if stm_auto_key:
                    upd_qry = "UPDATE STOCK SET GLA_AUTO_KEY = '%s' WHERE STM_AUTO_KEY = '%s'"%(gla_auto_key,stm_auto_key)
                    err_stock = updation_dir(upd_qry,cr)
                    error = update_stock_audit_log(cr,sysur_auto_key,stm_auto_key,user_id,gla_auto_key)
                if woo_auto_key:
                    upd_qry = "UPDATE WO_OPERATION SET GLA_AUTO_KEY = '%s' WHERE WOO_AUTO_KEY = '%s'"%(gla_auto_key,woo_auto_key)
                    err_woo = updation_dir(upd_qry,cr)
                if stm_auto_key and err_stock != '{"recs": ""}' or woo_auto_key and err_woo != '{"recs": ""}':
                    #raise exception because the part is serialized.
                    error = 'Unable to assign GL Account.'
                    return error,msg,new_woo_key
                else:
                    error = ''
        if not (wot_auto_key or woo_auto_key):
            #raise exception because the part is serialized.
            error = 'Task or lot not found.'
            return error,msg,new_woo_key       
        #cond_sub = "SELECT COND_LEVEL,PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE CONDITION_CODE = 'SV'"
        cond_sub = "SELECT COND_LEVEL,PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE PCC_AUTO_KEY = (SELECT PCC_MAIN_OUT FROM WO_CONTROL)"
        pcc_data = selection_dir(cond_sub,cr)
        pcc_auto_key = pcc_data and pcc_data[0] and pcc_data[0][1] or None
        cond_level = pcc_data and pcc_data[0] and pcc_data[0][0] or None
        ins_qty = 0  
        
    # check if there is a reservation to a wob:
    query = """SELECT WOB_AUTO_KEY FROM STOCK_RESERVATIONS
    WHERE STM_AUTO_KEY = %s
    """%stm_auto_key
    res = selection_dir(query,cr)

    if res:
        last_qry ="""SELECT SI_NUMBER FROM WO_OPERATION WHERE SI_NUMBER 
            LIKE '%s%s' AND SI_NUMBER <> '%s' 
            ORDER BY SI_NUMBER DESC,ENTRY_DATE DESC FETCH NEXT 10 ROWS ONLY"""%(si_number,'%',si_number)
        #last_qry = """SELECT VW.WOO_AUTO_KEY,W.SI_NUMBER FROM VIEW_SPS_WO_OPERATION VW 
        #    LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = VW.WOO_AUTO_KEY
        #    WHERE VW.PARENT_WO = '%s'"""%si_number         
        children = selection_dir(last_qry,cr)
        #now we must sort the children by the after-dash part of their SI_NUMBERs
        #ch_sorted = sorted(children,key=itemgetter(1))
        child_list = []
        child_dict = {}
        for child in children:
            si_num_split = '-' in child[0] and child[0].split('-') or []               
            dash_part = si_num_split and si_num_split[-1] or None               
            if dash_part.isnumeric() and child[0] not in child_dict:
                child_dict[child[0]] = int(dash_part)               
        #max_sub = child_dict and max(child_list) or None  
        max_sub = child_dict and max(child_dict, key=lambda key: child_dict[key]) or None
        if max_sub:
            max_dash = child_dict[max_sub]
            incrementer = str(int(max_dash) + 1)                    
            #grab last part of the si_number and remove it
            #add '-n' where n is an integer that is one more than the dash_part value
            num_char = len(str(max_dash))
            si_num = max_sub[:-num_char] + incrementer
        else:
            incrementer = '-001'
            si_num = si_number + incrementer
              
    else:
        # if there is no reservation:
        """
        2. If that STM has no STR, then we get the next SI# in the sequence.
        """
        query = """SELECT NUMBER_PREFIX || TO_CHAR(LAST_NUMBER+1),DPT_AUTO_KEY  FROM SYS_NUMBER_LOG
            WHERE SYSNLC_AUTO_KEY 
            = (SELECT SYSNLC_AUTO_KEY FROM SYS_NUMBER_LOG_CODES WHERE LOG_TYPE_CODE = 'WO')
        """  
        si_num = selection_dir(query,cr)
        si_num = si_num and si_num[0]
        if not si_num:
            return 'No last number found from number log.',msg,new_woo_key
        dpt_auto_key = si_num and si_num[1]        
        query = """SELECT CUR_AUTO_KEY,GLA_LABOR,GLA_MISC FROM SYS_COMPANIES 
            WHERE SYSCM_AUTO_KEY = %s
        """%syscm_auto_key        
        sys = selection_dir(query,cr)
        cur_auto_key = sys and sys[0]
        gla_labor = sys and sys[1]
        gla_misc = sys and sys[2]   
    #split it from the last dash and then take the integer part of that
    if si_num:
        """STM_AUTO_KEY	WOB_AUTO_KEY	WOO_AUTO_KEY	OP_TYPE	
        WO_TYPE	DOC_NUMBER	ADMIN_TYPE	SI_NUMBER	CMP_AUTO_KEY	
        WOO_PARENT	PNM_AUTO_KEY	PNM_MODIFY	SYSUR_AUTO_KEY	
        SYSCM_AUTO_KEY	svc_auto_key	new_wip_acct	bgs_default	
        cur_auto_key	gla_labor	gla_misc	lot_core_settings	
        ENTRY_DATE	dpt_auto_key	kit_qty	opm_auto_key	due_date	
        attention	ecd_method	wwt_auto_key
        """
        query ="""INSERT INTO WO_OPERATION (STM_AUTO_KEY,WOB_AUTO_KEY,WOO_AUTO_KEY,OP_TYPE,WO_TYPE,DOC_NUMBER,ADMIN_TYPE,
            SI_NUMBER,CMP_AUTO_KEY,WOO_PARENT,PNM_AUTO_KEY,PNM_MODIFY,SYSUR_AUTO_KEY,SYSCM_AUTO_KEY,
            ENTRY_DATE,dpt_auto_key,kit_qty,opm_auto_key,due_date,attention,ecd_method,wwt_auto_key,
            svc_auto_key,new_wip_acct,bgs_default,cur_auto_key,gla_labor,gla_misc,lot_core_settings) 
            VALUES('%s','%s',G_WOO_AUTO_KEY.NEXTVAL,'W','Internal','%s','Standard WO','%s','%s','%s','%s','%s','%s','%s',TO_DATE('%s', 'mm/dd/yyyy'),'%s','%s','%s',TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss'),'%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"""%\
            (stm_auto_key,wob_auto_key,si_num,si_num,cmp_auto_key,'',pnm_auto_key,pnm_auto_key,sysur_auto_key,\
            syscm_auto_key,datestamp,dpt_auto_key,kit_qty,opm_auto_key,due_date,attention,ecd_method,\
            wwt_auto_key,svc_auto_key,new_wip_acct,bgs_default,cur_auto_key,gla_labor,gla_misc,lot_core_settings)
        error = insertion_dir(query,cr)               
        sel_qry = """SELECT WO.WOO_AUTO_KEY                     
            FROM WO_OPERATION WO
            WHERE WO.SI_NUMBER = '%s' 
            AND WO.SYSUR_AUTO_KEY = %s"""%(si_num,sysur_auto_key)
        new_woo_key = selection_dir(sel_qry,cr)
        new_woo_key = new_woo_key and new_woo_key[0] and new_woo_key[0][0] or None
        if new_woo_key and stm_auto_key:
            """
               OPERATION_MASTER - WO that is set up ahead of time (template)
               OPERATION_TASKS - almost exact same tasks as WO_TASKS
               OPM has OPTs - OPTs have everything we need to create a new WO_TASK
                  For each new tasks, we insert the sub-woo and pnm from the sub-woo.
               1. pnm_auto_key from WO
               2. default_repair = 'T'
            """
            #1. lookup op master with pnm_auto_key and default_repair = 'T'
            #query = """SELECT OPM_AUTO_KEY FROM OPERATION_MASTER WHERE 
            #    PNM_AUTO_KEY = %s AND DEFAULT_REPAIR = 'T'                            
            #"""%pnm_auto_key
            #opm = selection_dir(query,cr)
            #opm_auto_key = opm and opm[0] and opm[0][0] or None
            #1. lookup op master with pnm_auto_key and default_repair = 'T'
            #automatically apply the default_repair
            where_clause = is_default_repair and "AND DEFAULT_REPAIR = 'T'" or ""
            query = """SELECT OPM.OPM_AUTO_KEY,OPM.OEM_REPAIR,OPM.OPERATION_ID,
                OPM.EXPIRATION_DATE,WRV.VERSION_ID,OPM.DESCRIPTION,OPM.DEFAULT_REPAIR
                FROM OPERATION_MASTER OPM 
                JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = OPM.PNM_AUTO_KEY
                LEFT JOIN WO_REPAIR_VERSION WRV ON WRV.WRV_AUTO_KEY = OPM.WRV_AUTO_KEY                             
                WHERE P.PNM_AUTO_KEY = %s %s                           
            """%(pnm_auto_key,where_clause)
            operations = selection_dir(query,cr)
            opm = operations and operations[0]
            opm_auto_key = opm and opm[0] or ''
            oem_repair = opm and opm[1] or ''
            operation_id = opm and opm[2] or ''
            exp_date = opm and opm[3] or ''
            version_id = opm and opm[4] or ''
            op_desc = opm and opm[5] or ''
            default_repair = opm and opm[6] or ''
            
            if not is_default_repair:
                #msg += 'Select task operation.'
                from polls.models import Operation
                for oper in operations:
                    opm_auto_key = oper[0] or ''
                    oem_repair = oper[1] or ''
                    operation_id = oper[2] or ''
                    exp_date = oper[3] or ''
                    version_id = oper[4] or ''
                    op_desc = oper[5] or ''
                    default_repair = oper[6]=='F' and False or oper[6]=='T' and True
                    from polls.models import Operation
                    operation = Operation.objects.create(
                        session_id=session_id,
                        opm_auto_key=opm_auto_key,
                        operation_id=operation_id,
                        exp_date=exp_date or None,
                        version=version_id,
                        part_number=pn,
                        part_desc=description,
                        op_desc=op_desc,
                        default_repair=default_repair,                                    
                        )
                    operation.save()
            if is_default_repair and opm_auto_key:
                error,msg = create_opm_tasks(session_id,sysur_auto_key,quapi_id,opm_auto_key,new_woo_key,cursor=cr,datestamp=datestamp)
    if not error or error == '{"recs": ""}':
        error = ''
        orcl_commit(con=con)
        msg = 'Successful update.' 
    return error,msg,new_woo_key

@shared_task
def e_signoff(quapi_id,session_id,sysur_auto_key,wot_auto_key):
    error,msg,set_vals = '','',''
    from polls.models import QueryApi,WOTask    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)

    right_now = datetime.now()
    today = right_now.strftime('%m/%d/%Y')
    today_time = right_now.strftime('%m/%d/%Y %H:%M:%S')
    right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')   
    query = """SELECT WT.SYSUR_SIGN_OFF,WT.SYSUR_SIGN_OFF2,
        WOS.STATUS_TYPE,WT.WOS_AUTO_KEY,WT.WOT_AUTO_KEY
        FROM WO_TASK WT
        LEFT JOIN WO_STATUS WOS ON WOS.WOS_AUTO_KEY = WT.WOS_AUTO_KEY
        WHERE WT.WOT_AUTO_KEY = %s
    """%(wot_auto_key)
    wot = selection_dir(query,cr)
    wot = wot and wot[0]
    wot_auto = wot and wot[4]
    if not wot_auto:
        return 'No task found for: %s.'%wot_auto_key,''
        
    signoff1 = wot and wot[0]
    signoff2 = wot and wot[1]
    status_type = wot and wot[2]
    wos_auto_key = wot and wot[3]  

    if signoff1 and signoff2:
        return 'Task already signed off: %s.'%wot_auto_key,''
    
    if status_type != 'Closed':
        """
        a.	Need to check for some things against a WOT 
            before allowing it to be closed when the 
            user is prompted “Task Completed”
            i.	Are there any WWT entries where qty_check_out <> qty_check_in
                1.	If so display “Tools currently checked out to task”
            ii.	Are there any WTL entries from a different sysur_auto_key with no stop_time
            iii.	Are there any wob with qty_needed <> qty_issued

        """
        
        query = """SELECT WTT.WTT_AUTO_KEY
            FROM WO_TASK WT
            LEFT JOIN WO_TASK_TOOLS WTT ON WTT.WOT_AUTO_KEY = WT.WOT_AUTO_KEY
            WHERE (WTT.QTY_CHECKED_OUT <> WTT.QTY_CHECKED_IN OR WTT.WOT_AUTO_KEY IS NULL)
            AND WT.WOT_AUTO_KEY = %s
        """%(wot_auto_key)
        wtt = selection_dir(query,cr)
        wtt = wtt and wtt[0]
        wtt_auto_key = wtt and wtt[0] or ''
        if wtt_auto_key:
            return 'Tools currently checked out to task: %s.'%wot_auto_key,msg 
            
        query = """SELECT WTL.WTL_AUTO_KEY,WT.WOS_AUTO_KEY
            FROM WO_TASK WT
            LEFT JOIN WO_TASK_LABOR WTL ON WTL.WOT_AUTO_KEY = WT.WOT_AUTO_KEY 
            WHERE WTL.STOP_TIME IS NULL 
            AND WT.WOT_AUTO_KEY = %s
        """%(wot_auto_key)
        wot = selection_dir(query,cr)
        wot = wot and wot[0]
        wtl_auto_key = wot and wot[0] or ''
        if wtl_auto_key:
            return 'There is another open labor for this task: %s.'%wot_auto_key,msg

        query = """SELECT WOB.WOB_AUTO_KEY,WOB.QTY_RESERVED
            FROM WO_TASK WT
            LEFT JOIN WO_BOM WOB ON WOB.WOT_AUTO_KEY = WT.WOT_AUTO_KEY 
            WHERE (WOB.QTY_NEEDED <> WOB.QTY_ISSUED OR WOB.WOT_AUTO_KEY IS NULL)
            AND WT.WOT_AUTO_KEY = %s
        """%(wot_auto_key)
        wot = selection_dir(query,cr)
        wot = wot and wot[0]
        wob_auto_key = wot and wot[0] or ''                   
        if wob_auto_key:
            return 'There is an open BOM for this task: %s.'%wot_auto_key,msg
            
        #update the task status to that of 'closed.' 
        sub_woc = "(SELECT WOS_COMPLETE FROM WO_CONTROL FETCH NEXT 1 ROWS ONLY)"       
        query = """INSERT INTO WO_TASK_STATUS (SYSUR_AUTO_KEY,
            WOT_AUTO_KEY,WOS_AUTO_KEY,ENTRY_DATE,SYSTEM_DATE,WOS_PREVIOUS)                               
            VALUES(%s,%s,%s,TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),TO_DATE('%s','mm/dd/yyyy'),'%s')
            """%(sysur_auto_key,wot_auto_key,sub_woc,today_time,today,wos_auto_key)
        error = insertion_dir(query,cr)
        query = """UPDATE WO_TASK SET WOS_AUTO_KEY = %s WHERE WOT_AUTO_KEY = %s"""%(sub_woc,wot_auto_key)
        error = updation_dir(query,cr)
        if error and error != '{"recs": ""}': 
            return error,msg
            
    query = """SELECT CAN_SIGN_OFF1,
        CAN_SIGN_OFF2
        FROM SYS_USERS
        WHERE SYSUR_AUTO_KEY = %s
    """%sysur_auto_key
    wot = selection_dir(query,cr)
    wot = wot and wot[0]
    can_signoff1 = wot and wot[0]
    can_signoff2 = wot and wot[1]

    if can_signoff1 == 'F' and can_signoff2 == 'F':
        return 'User not certified for task sign off.',''
   
    if can_signoff1 == 'T':
        if not signoff1:
            set_qry = """UPDATE WO_TASK SET SYSUR_SIGN_OFF = %s,
            SIGN_OFF_DATE = TO_DATE('%s','mm/dd/yyyy')            
            WHERE WOT_AUTO_KEY = %s"""%(sysur_auto_key,today,wot_auto_key) 
            error = updation_dir(set_qry,cr)

    if can_signoff2 == 'T':
        if not signoff2:
            set_qry = """UPDATE WO_TASK SET SYSUR_SIGN_OFF2 = %s, 
            SIGN_OFF_DATE2 = TO_DATE('%s','mm/dd/yyyy')             
            WHERE WOT_AUTO_KEY = %s"""%(sysur_auto_key,today,wot_auto_key) 
            error = updation_dir(set_qry,cr)     
       
    if error and error != '{"recs": ""}':
        aud_status = 'failure'
        
    elif not error or error == '{"recs": ""}':
        error = ''
        orcl_commit(con=con)     
        aud_status = 'success'
        msg = 'Successful signoff on task: %s. '%wot_auto_key 
                
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='e-signoff')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        new_val = 'User sign off on task: %s. '%wot_auto_key + error
        field_changed = 'Update to signoff user on task: %s'%wot_auto_key
        field_changed += error
        error += register_audit_trail(user_rec,field_changed,new_val,right_now,app_id,quapi,status=aud_status)  
    else:
        error = 'Incorrect Quantum User ID.'    
    return error,msg

@shared_task
def get_signoff_tasks(quapi_id,session_id,wot_auto_key='',wo_number='',woo_auto_key=''):
    error,msg = '',''
    from polls.models import QueryApi,WOTask    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    query="""SELECT DISTINCT WT.WOT_AUTO_KEY, WT.SEQUENCE, WT.SQUAWK_DESC,
        WOS.SEVERITY || ' - ' || CASE WHEN WT.WOS_AUTO_KEY IS NOT NULL 
        THEN WOS.DESCRIPTION
        ELSE 'PENDING' END, 
        WOS.STATUS_TYPE, 
        '', 
        'dummy',WO.WOO_AUTO_KEY,
        WT.WOT_AUTO_KEY,
        WTM.DESCRIPTION,
        WTM.AUTO_SEQ_NUM,
        WT.SKILLS_EST_HOURS           
        FROM WO_TASK WT
        LEFT JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WT.WTM_AUTO_KEY
        LEFT JOIN WO_OPERATION WO ON WT.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
        LEFT JOIN VIEW_WO_STATUS_TASK VTS ON WT.WOT_AUTO_KEY = VTS.WOT_AUTO_KEY
        LEFT JOIN WO_STATUS WOS ON WOS.WOS_AUTO_KEY = WT.WOS_AUTO_KEY
        """    
    if wot_auto_key:
        query+="""
        WHERE WT.WOT_AUTO_KEY = %s
        """%wot_auto_key
    elif wo_number:        
        query+="""
        WHERE UPPER(WO.SI_NUMBER) = UPPER('%s')
        """%wo_number
    elif woo_auto_key:
        query+="""
        WHERE WO.WOO_AUTO_KEY = %s
        """%woo_auto_key           
    #query += " AND (WOS.STATUS_TYPE IS NULL OR WOS.STATUS_TYPE NOT IN ('Closed', 'Cancel'))"
    #query += " ORDER BY WT.SEQUENCE ASC"    
    recs = selection_dir(query,cr)
    
    if recs:
        existing_tasks = WOTask.objects.filter(session_id=session_id)
        del_existings = existing_tasks.delete()
        for rec in recs:
            rec += [1,session_id]
        res = create_tasks_bulk(recs)
        if not res:
            error = 'There was a problem with creating the tasks locally. Try again please.'
    else:
        error = 'No tasks found.'    
    return error,msg
    
@shared_task
def get_part_numbers(quapi_id,session_id,part_char=''):
    error,msg = '',''
    from polls.models import QueryApi,PartNumbers    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)        
    where_clause,join_wo = '',''    
    delete = PartNumbers.objects.filter(session_id=session_id).delete()    
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    if part_char:
        parts_query = """SELECT PN,DESCRIPTION,PNM_AUTO_KEY 
        FROM PARTS_MASTER 
        WHERE ACTIVE_PART = 'T' AND PN LIKE '%s%s' ORDER BY PN"""%(part_char,'%')
    else:
        parts_query = """SELECT PN,DESCRIPTION,PNM_AUTO_KEY 
        FROM PARTS_MASTER WHERE ACTIVE_PART = 'T' ORDER BY PN"""
    recs = selection_dir(parts_query,cr)
    if recs:
        error = create_parts(session_id,recs)
    return error
    
def create_parts(session_id,parts):
    error = ''
    from polls.models import PartNumbers
    part_data,error = [],''
    for part in parts:    
        part_data.append(PartNumbers(
        part_number = part[0],
        description = part[1],
        pnm_auto_key = part[2],
        session_id = session_id,      
        ))
    if part_data:
        try:
            delete = PartNumbers.objects.filter(session_id=session_id).delete()
            PartNumbers.objects.bulk_create(part_data) or []    
        except Exception as exc:
            error += "Error, %s, creating parts."%(exc)
    return error
    
@shared_task
def get_consignment_codes(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi,WOTask    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
        
    where_clause,join_wo = '',''    
    delete = WOTask.objects.filter(session_id=session_id).delete()    
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    cons_query = """SELECT CONSIGNMENT_CODE FROM CONSIGNMENT_CODES ORDER BY CONSIGNMENT_CODE"""
    recs = selection_dir(cons_query,cr)
    if recs:
        error = create_consignments(session_id,recs)
    return error

@shared_task
def get_activities_conditions_tasks(quapi_id,session_id,si_number=''):
    error,msg = '',''
    from polls.models import QueryApi,WOTask    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
        
    where_clause,join_wo = '',''    
    delete = WOTask.objects.filter(session_id=session_id).delete()    
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    if si_number:
        where_clause += " AND UPPER(WOO.SI_NUMBER) = UPPER('%s')"%si_number
        join_wo = " LEFT JOIN WO_OPERATION WOO ON WOT.WOO_AUTO_KEY = WOO.WOO_AUTO_KEY"
        
    act_query = """SELECT DISTINCT ACTIVITY
        FROM WO_BOM"""
        
    cond_query = """SELECT DISTINCT CONDITION_CODE
        FROM PART_CONDITION_CODES WHERE HISTORICAL = 'F'"""
        
    task_query = """SELECT WOT.SEQUENCE, WTM.DESCRIPTION,
        WOT.WOT_AUTO_KEY, WOO.SI_NUMBER, WOS.DESCRIPTION,
        WOS.STATUS_TYPE,
        WOT.SQUAWK_DESC, WOT.CORRECTIVE_ACTION        
        FROM WO_TASK_MASTER WTM
        LEFT JOIN WO_TASK WOT ON WOT.WTM_AUTO_KEY = WTM.WTM_AUTO_KEY
        LEFT JOIN WO_OPERATION WOO ON WOO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
        LEFT JOIN WO_STATUS WOS ON WOS.WOS_AUTO_KEY = WOT.WOS_AUTO_KEY 
        WHERE (WOT.WOS_AUTO_KEY NOT IN 
          (SELECT WOS_AUTO_KEY FROM WO_STATUS 
          WHERE STATUS_TYPE IN ('Closed', 'Cancel')) 
          OR WOS.STATUS_TYPE IS NULL)
        %s ORDER BY WOT.SEQUENCE
    """%(where_clause)

    #activities = selection_dir(act_query,cr)
    activities = [['CONSUM'],['REPLACE'],['REPAIR']]
    activities += [['WO'],['INSPECT'],['ISSUE'],['LOAN']]
    activities += [['RETURN'],['TURNIN'],['EXCHANGE']]
    conditions = selection_dir(cond_query,cr)
    tasks = selection_dir(task_query,cr)
    if activities:
        error = create_activities(session_id,activities)
    if conditions:
        error = create_conditions(session_id,conditions)
    if tasks:
        error = create_tasks(session_id,tasks)
    return error,msg
    
def create_consignments(session_id,consignments):
    error = ''
    from polls.models import Consignments
    cons_data,error = [],''
    for cons in consignments:    
        cons_data.append(Consignments(
        code = cons[0],
        session_id = session_id,      
        ))
    if cons_data:
        try:
            delete = Consignments.objects.filter(session_id=session_id).delete()
            Consignments.objects.bulk_create(cons_data) or []    
        except Exception as exc:
            error += "Error, %s, creating consignments."%(exc)
    return error   
   
def create_mgmt_tasks(session_id,wo_tasks):
    error = ''
    from polls.models import WOTask
    task_data,error = [],''
    
    for task in wo_tasks:    
        task_data.append(WOTask(
        wot_sequence = task[0],
        task_master_desc = task[1],
        wot_auto_key = task[2],
        si_number = task[3],
        wot_status = task[4],
        status_type = task[5],
        wot_description = task[6],
        wot_technician = task[7],
        customer = task[8],#CMP.COMPANY_NAME
        ac_model = task[9],#WOO.WO_UDF_001
        ac_reg = task[10],#WOO.WO_UDF_002
        ac_sn = task[11],#WOO.WO_UDF_003
        part_description = task[12],#PN from S.PNM_AUTO_KEY
        eng_model = task[13],#PN from WOO.PNM_AUTO_KEY
        esn = task[14],#S.SERIAL_NUMBER
        session_id = session_id,      
        ))
        
    if task_data:     
        try:
            delete = WOTask.objects.filter(session_id=session_id).delete()
            WOTask.objects.bulk_create(task_data) or []    
        except Exception as exc:
            error += "Error, %s, creating tasks."%(exc)
    return error
      
def create_tasks(session_id,wo_tasks):
    error = ''
    from polls.models import WOTask
    task_data,error = [],''
    for task in wo_tasks:    
        task_data.append(WOTask(
        wot_sequence = task[0],
        task_master_desc = task[1],
        wot_auto_key = task[2],
        si_number = task[3],
        wot_status = task[4],
        status_type = task[5],
        wot_description = task[6],
        wot_technician = task[7],
        session_id = session_id,      
        ))
    if task_data:     
        try:
            #delete = WOTask.objets.filter(session_id=session_id).delete()
            WOTask.objects.bulk_create(task_data) or []    
        except Exception as exc:
            error += "Error, %s, creating tasks."%(exc)
    return error
    
def create_activities(session_id,activities):
    error = ''
    from polls.models import Activities
    act_data,error = [],''
    for act in activities:    
        act_data.append(Activities(
        activity = act[0],
        session_id = session_id,      
        ))
    if act_data:
        try:
            delete = Activities.objects.filter(session_id=session_id).delete()
            Activities.objects.bulk_create(act_data) or []    
        except Exception as exc:
            error += "Error, %s, creating activities."%(exc)
    return error
    
def create_conditions(session_id,conditions):
    error = ''
    from polls.models import PartConditions
    cond_data,error = [],''
    for cond in conditions:    
        cond_data.append(PartConditions(
        condition_code = cond[0],
        session_id = session_id,      
        ))
    if cond_data:     
        try:
            delete = PartConditions.objects.filter(session_id=session_id).delete()
            PartConditions.objects.bulk_create(cond_data) or []    
        except Exception as exc:
            error += "Error, %s, creating conditions."%(exc)
    return error
    
def create_wo_boms(session_id,wo_number,boms):
    error = ''   
    from polls.models import WOStatus
    bom_data,error = [],''
    for bom in boms:    
        bom_data.append(WOStatus(
        wo_number = wo_number,
        condition_code = bom[0],
        activity = bom[1],
        part_number = bom[2],
        description = bom[3],
        task_master_desc = bom[4],
        qty_needed = bom[5] or 0,
        wob_auto_key = bom[6] or 0,
        notes = bom[7],
        wos_auto_key = bom[8] or 0,
        status = bom[9],
        wot_sequence = bom[10],
        session_id = session_id,      
        ))
    if bom_data:
        try:
            delete = WOStatus.objects.filter(session_id=session_id).delete()
            WOStatus.objects.bulk_create(bom_data) or []    
        except Exception as exc:
            error += "Error, %s, with WO BOM  rows for grid for WO#, %s"%(exc,wo_number)
    return error

@shared_task
def get_wo_bom(quapi_id,session_id,wo_number,wot_auto_key,part_number,description,cond_sel,act_sel,bom_status):
    error,msg = '',''
    from polls.models import QueryApi,WOStatus    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
        
    """
    restriction 1 - pn has to exist
    restriction 2 - qty_need cannot be > wob.qty issued +wob.qty reserved
    restriction 3 - condition should be a dropdown with list of condition codes from part condition code table
    restriction 4 - activity should be a dropdown with a list of wob.activity
    """
    where_clause = ""
    if wo_number:
        where_clause = " UPPER(WO.SI_NUMBER) = UPPER('%s')"%(wo_number) 
    if bom_status:
        bstat_sub = """(SELECT BOS_AUTO_KEY 
            FROM WO_BOM_STATUS 
            WHERE BOM_STATUS_CODE = '%s')"""%bom_status
        where_clause = where_clause and where_clause + " AND" or ""            
        where_clause += " WOB.BOS_AUTO_KEY = %s"%bstat_sub      
    if wot_auto_key:   
        where_clause = where_clause and where_clause + " AND" or ""            
        where_clause += " WOB.WOT_AUTO_KEY = %s"%wot_auto_key
        
    if part_number or description:
        query = """SELECT PNM_AUTO_KEY FROM PARTS_MASTER
            WHERE UPPER(PN) = UPPER('%s') OR UPPER(DESCRIPTION) = UPPER('%s')
        """%(part_number,description)
        pnm = selection_dir(query,cr)
        pnm_auto_key = pnm and pnm[0] and pnm[0][0] or None
        if not pnm:
            return 'Part does not exist.',''
        where_clause = where_clause and where_clause + " AND" or ""
        where_clause += " WOB.PNM_AUTO_KEY = %s"%pnm_auto_key
        
    if cond_sel:
        query = """SELECT PCC_AUTO_KEY FROM PART_CONDITION_CODES
            WHERE UPPER(CONDITION_CODE) = UPPER('%s')"""%cond_sel
        cond = selection_dir(query,cr)
        pcc_auto_key = cond and cond[0] and cond[0][0] or None
        if not cond:
            return 'Condition does not exist.',''
        where_clause = where_clause and where_clause + " AND" or ""            
        where_clause += " WOB.PCC_AUTO_KEY = %s"%pcc_auto_key
        
    if act_sel:
        where_clause = where_clause and where_clause + " AND" or "" 
        where_clause += " WOB.ACTIVITY = '%s'"%act_sel  
    
    query = """SELECT 
        PCC.CONDITION_CODE,
        WOB.ACTIVITY,
        P.PN,
        P.DESCRIPTION,
        WTM.DESCRIPTION,
        WOB.QTY_NEEDED,
        WOB.WOB_AUTO_KEY,
        WOB.NOTES,
        WOB.BOS_AUTO_KEY,
        WBS.BOM_STATUS_CODE,
        WOT.SEQUENCE
        FROM WO_BOM WOB
        LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOB.WOO_AUTO_KEY
        JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = WOB.PCC_AUTO_KEY
        JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = WOB.PNM_AUTO_KEY
        LEFT JOIN WO_TASK WOT ON WOT.WOT_AUTO_KEY = WOB.WOT_AUTO_KEY
        LEFT JOIN WO_TASK_MASTER WTM ON WOT.WTM_AUTO_KEY = WTM.WTM_AUTO_KEY
        LEFT JOIN WO_BOM_STATUS WBS ON WBS.BOS_AUTO_KEY = WOB.BOS_AUTO_KEY
        WHERE%s
    """%(where_clause)
    boms = selection_dir(query,cr)
    if boms:
        error = create_wo_boms(session_id,wo_number,boms) 
    else:
                                                                        
        error = "No BoMs found."    
    return error,msg

@shared_task
def update_wo_bom(quapi_id,session_id,wo_number,wob_id_list,\
    pn,description,condition,activity,qty_need,wo_task,notes,\
    bom_status,sysur_auto_key,task_desc):
    
    error,msg,boms,field_changed,wot_sequence,unver_part = '','','',[],'',''
    part_number,list_price = '',0
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    set_clause,aud_status = '',''    
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
        
    wob_ids = construct_akl(wob_id_list)
    if qty_need:
        error,quantity = qty_to_float(qty_need)
        if error:
            return error,msg

    if wo_task and wo_task.isnumeric():
        query = """SELECT WOT.WOT_AUTO_KEY,WOT.SEQUENCE FROM WO_TASK WOT
            LEFT JOIN WO_TASK_MASTER WTM 
            ON WOT.WTM_AUTO_KEY = WTM.WTM_AUTO_KEY
            WHERE WOT.WOT_AUTO_KEY = %s"""%wo_task
        wot = selection_dir(query,cr)
        wot_auto_key = wot and wot[0] and wot[0][0] or None
        wot_sequence = wot and wot[0] and wot[0][1] or ''
        if not wot:
            return 'Task does not exist.',''        
        set_clause += " WOT_AUTO_KEY = %s,"%wot_auto_key
        
    elif wo_task and not wo_task.isnumeric():
        query = """SELECT WOT.WOT_AUTO_KEY FROM WO_TASK WOT
            LEFT JOIN WO_TASK_MASTER WTM 
            ON WOT.WTM_AUTO_KEY = WTM.WTM_AUTO_KEY
            WHERE WOT.WOT_AUTO_KEY = %s"""%wo_task
        wot = selection_dir(query,cr)
        wot_auto_key = wot and wot[0] and wot[0][0] or None
        if not wot:
            return 'Task does not exist.',''        
        set_clause += " WOT_AUTO_KEY = %s,"%wot_auto_key   

    if pn or description:
        query = """SELECT PNM_AUTO_KEY,ACTIVE_PART,LIST_PRICE FROM PARTS_MASTER
            WHERE UPPER(PN) = UPPER('%s') OR UPPER(DESCRIPTION) = UPPER('%s')
        """%(pn,description)
        pnm = selection_dir(query,cr)
        pnm = pnm and pnm[0]
        pnm_auto_key = pnm and pnm[0] or None
        unver_part = pnm and pnm[1] == 'T' or 'Unverified PN'
        list_price = pnm and pnm[2] or 0
        
        if not pnm:
            return 'Part does not exist.',''
        set_clause += " PNM_AUTO_KEY = %s,"%pnm_auto_key
        
    if condition:
        query = """SELECT PCC_AUTO_KEY FROM PART_CONDITION_CODES
            WHERE UPPER(CONDITION_CODE) = UPPER('%s')"""%condition
        cond = selection_dir(query,cr)
        pcc_auto_key = cond and cond[0] and cond[0][0] or None
        if not cond:
            return 'Condition does not exist.',''        
        set_clause += " PCC_AUTO_KEY = %s,"%pcc_auto_key
        
    if notes:
        set_clause += " NOTES = '%s',"%notes
        
    if bom_status:
        bstat_sub = """(SELECT BOS_AUTO_KEY 
            FROM WO_BOM_STATUS 
            WHERE BOM_STATUS_CODE = '%s')"""%bom_status           
        set_clause += " BOS_AUTO_KEY = %s,"%bstat_sub 
        
    if activity:
        set_clause += " ACTIVITY = '%s',"%activity

    if qty_need: 
        set_clause += " QTY_NEEDED = %s,"%qty_need
    
    set_clause = set_clause and set_clause[:-1] or ''
    for wobs in wob_ids:    
        upd_query = """UPDATE WO_BOM SET %s
            WHERE WOB_AUTO_KEY in %s"""%(set_clause,wobs)
        error = updation_dir(upd_query,cr)
        sel_qry = """SELECT QTY_NEEDED FROM WO_BOM
            WHERE WOB_AUTO_KEY IN %s
        """%(wobs) 
        orig_qtys = selection_dir(sel_qry,cr)
        orig_qtys = orig_qtys and orig_qtys[0] or ''
               
        query = """SELECT 
            PCC.CONDITION_CODE,
            WOB.ACTIVITY,
            P.PN,
            P.DESCRIPTION,
            WTM.DESCRIPTION,
            WOB.QTY_NEEDED,
            WOB.WOB_AUTO_KEY,
            WOB.NOTES,
            WOB.BOS_AUTO_KEY,
            WBS.BOM_STATUS_CODE,
            WOT.SEQUENCE
            FROM WO_BOM WOB
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOB.WOO_AUTO_KEY
            JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = WOB.PCC_AUTO_KEY
            JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = WOB.PNM_AUTO_KEY
            LEFT JOIN WO_TASK WOT ON WOT.WOT_AUTO_KEY = WOB.WOT_AUTO_KEY
            LEFT JOIN WO_TASK_MASTER WTM ON WOT.WTM_AUTO_KEY = WTM.WTM_AUTO_KEY
            LEFT JOIN WO_BOM_STATUS WBS ON WBS.BOS_AUTO_KEY = WOB.BOS_AUTO_KEY
            WHERE WOB.WOB_AUTO_KEY IN %s
        """%wobs
        
        boms = selection_dir(query,cr)
        bom = boms and boms[0] or []

        if bom:
            part_number = bom[2]
            description = bom[3]
            create_wo_boms(session_id,wo_number,boms)
        
    #check to see if error with update   
    if error and error != '{"recs": ""}':
        aud_status = 'failure'
        msg = 'Failed Update.'
        
    elif error == '{"recs": ""}':
        error = ''
        orcl_commit(con=con)     
        aud_status = 'success'
        msg = 'Successful update.'         
       
    right_now = datetime.now()
    right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')        
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='bom-management')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    part_repair,csm_status,eng_status,inv_status = '','','',''
    reg_events=[]
    if activity == 'REPAIR':
        part_repair = 'Activity to Part Repair'
        reg_events += ['Activity to Repair']
    if bom_status == 'CSM':
        csm_status = 'Changed BOM status to CSM'
        reg_events += ['BOM Status to CSM']
    if bom_status == 'ENG':
        eng_status = 'Changed BOM status to ENG'
        reg_events += ['BOM Status to ENG']
    if bom_status == 'INV':
        inv_status = 'Changed BOM status to INV'
        reg_events += ['BOM Status to INV']
        
    cond = 'Condition: ' + condition or ''
    act = 'Activity: ' + activity or ''
    qty_needed = 'Qty Needed: ' + qty_need or ''
    wo_task = 'Task: ' + wo_task or ''
    notes_str = 'Notes: ' + notes or ''
    bom_status = 'BOM Status: ' + bom_status or ''
    field_changed = 'Updates to PN/DESC: %s / %s'%(part_number,description)
    field_changed += ' %s %s %s %s %s %s '%(cond,act,qty_needed,task_desc,notes_str,bom_status)     
    parameters = [unver_part,str(wo_number),str(right_now)]
    parameters += [str(wot_sequence),str(task_desc)]
    parameters += [str(part_number),str(description)]
    parameters += [str(condition),str(orig_qtys),str(qty_need),]
    parameters += [notes,str(list_price),'','',part_repair,csm_status,eng_status]
    
    if user_rec:
        new_val = set_clause + '. ' + error
        field_changed += error
        error += register_audit_trail(
        user_rec,
        field_changed,
        new_val,
        right_now,
        app_id,
        quapi,
        status=aud_status,
        reg_events=reg_events,
        parameters=parameters,
        sysur_auto_key=sysur_auto_key,
        )  
    else:
        error = 'Incorrect Quantum User ID.'    
    return error,msg

@shared_task
def get_att_keys(quapi_id,row):
    error,fail_msg = '',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    key_row = ['PN','DESCRIPTION']
    for att_name in row[2:]:
        if att_name != 'None':
            uda_query = """SELECT UDA_AUTO_KEY FROM 
                USER_DEFINED_ATTRIBUTES WHERE 
                UPPER(UDA_CODE) = UPPER('%s') AND AUTO_KEY_PREFIX='PNM'"""%att_name.strip()          
            uda = selection_dir(uda_query,cr)
            uda = uda and uda[0] and uda[0][0] or ''
            if not uda:
                fail_msg += 'Attribute not found: %s. '%att_name
            key_row.append(uda)
    return key_row,fail_msg          

@shared_task
def get_part_attributes(quapi_id,sysur_auto_key,session_id,pnm_auto_key,stype='one',group='',create_anew=True):
    error,msg,query,group_where = '','','',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)  
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    if group:
        group_where = " AND UDG.GROUP_NAME = '%s'"%group
    if pnm_auto_key and stype == 'one':
        query = """SELECT UDA.AUTO_KEY_PREFIX,UDC.ATTRIBUTE_VALUE,
            UDA.UDA_CODE,UDA.SEQUENCE,UDG.GROUP_NAME FROM UDA_CHECKED UDC
            JOIN USER_DEFINED_ATTRIBUTES UDA ON UDA.UDA_AUTO_KEY = UDC.UDA_AUTO_KEY
            JOIN UDA_GROUPS UDG ON UDG.UDG_AUTO_KEY = UDA.UDG_AUTO_KEY
            WHERE UDA.AUTO_KEY_PREFIX='PNM' AND UDC.ATTRIBUTE_VALUE IS NOT NULL
            AND UDC.AUTO_KEY = %s%s"""%(pnm_auto_key,group_where)
    else:
        query = """SELECT DISTINCT UDA.AUTO_KEY_PREFIX,UDA.UDA_CODE,
            UDA.UDA_CODE,UDA.SEQUENCE,UDG.GROUP_NAME 
            FROM USER_DEFINED_ATTRIBUTES UDA JOIN UDA_GROUPS UDG 
            ON UDG.UDG_AUTO_KEY = UDA.UDG_AUTO_KEY 
            WHERE UDA.AUTO_KEY_PREFIX='PNM'%s"""%group_where
    
    pnm_atts = selection_dir(query,cr)         
    if not pnm_atts and create_anew:
       #create the 'IN HOUSE' attributes 
        udas = ['cleaning','ndt','visual','dimensional']
        error = create_udas(cr,con,group=group,udas=udas)
        query = """SELECT DISTINCT UDA.AUTO_KEY_PREFIX,UDA.UDA_CODE,
            UDA.UDA_CODE,UDA.SEQUENCE,UDG.GROUP_NAME 
            FROM USER_DEFINED_ATTRIBUTES UDA JOIN UDA_GROUPS UDG 
            ON UDG.UDG_AUTO_KEY = UDA.UDG_AUTO_KEY 
            WHERE UDA.AUTO_KEY_PREFIX='PNM'%s"""%group_where
        pnm_atts = selection_dir(query,cr)  
        orcl_commit(con=con)        
    if pnm_atts:
        error = create_attributes(session_id,pnm_auto_key,pnm_atts)            
    return error,msg
    
def create_udas(cr,con,group='',udas=[]):
    udg_auto_key,error = '',''
    if group:
        query = """SELECT UDG_AUTO_KEY FROM UDA_GROUPS
        WHERE UPPER(GROUP_NAME) = UPPER('%s')
        """%group
        groups = selection_dir(query,cr)
        udg_auto_key = groups and groups[0] and groups[0][0] or ''
        if not udg_auto_key:
            #INSERT THE NEW GROUP
            query = """INSERT INTO UDA_GROUPS(UDG_AUTO_KEY,GROUP_NAME,AUTO_KEY_PREFIX)
            VALUES(G_UDG_AUTO_KEY.NEXTVAL,UPPER('%s'),'PNM')
            """%group
            error = insertion_dir(query,cr)
            if not error:
                query = """SELECT UDG_AUTO_KEY FROM UDA_GROUPS
                WHERE UPPER(GROUP_NAME) = UPPER('%s')"""%group
                gps = selection_dir(query,cr)
                udg_auto_key = gps and gps[0] and gps[0][0]

    if udg_auto_key:
        sequence = 1    
        for uda in udas:
            query = """INSERT INTO 
            USER_DEFINED_ATTRIBUTES(AUTO_KEY_PREFIX,UDA_CODE,UDA_DESCRIPTION,SEQUENCE,UDG_AUTO_KEY,HISTORICAL_FLAG,UDA_TYPE)
            VALUES('PNM',UPPER('%s'),UPPER('%s'),%s,%s,'F','Edit')
            """%(uda,uda,sequence,udg_auto_key)
            error = insertion_dir(query,cr)
            sequence += 1
                  
    return error
    
    
def create_attributes(session_id,pnm_auto_key,pnm_atts):
    error = ''
    from polls.models import UserDefAtts as uda
    att_data,error = [],''  
    for att in pnm_atts:
        att_data.append(uda(
        auto_key = pnm_auto_key,
        att_type = att[4],
        att_value = att[1],
        att_name = att[2],
        att_seq = att[3],
        session_id = session_id,      
        ))   
    if att_data:
        try:
            delete = uda.objects.filter(session_id=session_id).delete()
            uda.objects.bulk_create(att_data) or []    
        except Exception as exc:
            error += "Error, %s, creating attributes locally"%exc
    return error
    
@shared_task
def synch_attributes(quapi_id,sysur_auto_key,session_id,row,att_keys):
    error,msg,fail_msg,rec_input = '','','',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    pn = row[0]
    desc = row[1]
    serialized = row[9]
    query="select pnm_auto_key from parts_master where UPPER(pn)=UPPER('%s')"%row[0]
    pnm = selection_dir(query,cr)
    pnm_auto_key = pnm and pnm[0] and pnm[0][0] or None
    if not pnm_auto_key:      
        from portal.tasks import create_pn          
        error,msg = create_pn(quapi_id,session_id,pn,desc,serialized,sysur_auto_key,send_notice=False)
        query = "SELECT PNM_AUTO_KEY FROM PARTS_MASTER WHERE UPPER(PN)=UPPER('%s') AND DESCRIPTION='%s'"%(pn,desc)
        pnm = selection_dir(query,cr)
        pnm_auto_key = pnm and pnm[0] and pnm[0][0] or None
    if pnm_auto_key:
        if row[2] != '' and row[2] != None and att_keys[2]:
            #Cleaning Attribute                                                                          
            #uda_query = """SELECT UDA_AUTO_KEY FROM 
            #    USER_DEFINED_ATTRIBUTES WHERE 
            #    UDA_CODE = 'CLEANING' AND AUTO_KEY_PREFIX='PNM'"""            
            #uda = selection_dir(uda_query,cr)
            #uda_auto_key = uda and uda[0] and uda[0][0] or None
            uda_auto_key = att_keys[2]
            udc_query = """SELECT UDC_AUTO_KEY FROM 
                UDA_CHECKED WHERE AUTO_KEY = %s
                AND UDA_AUTO_KEY = %s"""%(pnm_auto_key,uda_auto_key)
            udc = selection_dir(udc_query,cr)
            udc_auto_key = udc and udc[0] and udc[0][0] or None
            if uda_auto_key and not udc_auto_key:  
                query="""INSERT INTO UDA_CHECKED (UDA_AUTO_KEY,AUTO_KEY,ATTRIBUTE_VALUE) 
                  VALUES(%s,%s,'%s')"""%(uda_auto_key,pnm_auto_key,row[2])
                error = insertion_dir(query,cr)
            elif uda_auto_key and udc_auto_key:
                query="""UPDATE UDA_CHECKED SET ATTRIBUTE_VALUE = '%s'
                    WHERE UDC_AUTO_KEY = %s"""%(row[2],udc_auto_key)
                error = updation_dir(query,cr)
            else:
                error = "Cleaning attribute not defined for PNs."
            if error and error != '{"recs": ""}':               
                fail_msg += '\nFailed insert for CLEANING, pn = ' + str(row[0]) + ' - ' + error
            else:
                rec_input = '\nSynched CLEANING attribute for part, %s'%(str(row[0]) + '(' + str(row[1]) + ')')
        
        if row[3] != '' and row[3] != None and att_keys[3]:    
            #Visual Attribute    
            #uda_query = """SELECT UDA_AUTO_KEY FROM 
            #    USER_DEFINED_ATTRIBUTES WHERE 
            #    UDA_CODE = 'VISUAL' AND AUTO_KEY_PREFIX='PNM'"""            
            #uda = selection_dir(uda_query,cr)
            #uda_auto_key = uda and uda[0] and uda[0][0] or None
            uda_auto_key = att_keys[3]
            udc_query = """SELECT UDC_AUTO_KEY FROM 
                UDA_CHECKED WHERE AUTO_KEY = %s
                AND UDA_AUTO_KEY = %s"""%(pnm_auto_key,uda_auto_key)
            udc = selection_dir(udc_query,cr)
            udc_auto_key = udc and udc[0] and udc[0][0] or None
            if uda_auto_key and not udc_auto_key:  
                query="""INSERT INTO UDA_CHECKED (UDA_AUTO_KEY,AUTO_KEY,ATTRIBUTE_VALUE) 
                  VALUES(%s,%s,'%s')"""%(uda_auto_key,pnm_auto_key,row[3])            
                error = insertion_dir(query,cr)
            elif uda_auto_key and udc_auto_key:
                query="""UPDATE UDA_CHECKED SET ATTRIBUTE_VALUE = '%s'
                    WHERE UDC_AUTO_KEY = %s"""%(row[3],udc_auto_key)
                error = updation_dir(query,cr)
            else:
                error = "Visual attribute not defined for PNs."
            if error and error != '{"recs": ""}':            
                fail_msg += '\nFailed insert for VISUAL, pn = ' + str(row[0]) + ' - ' + error
            else:
                rec_input += '\nSynched VISUAL attribute for pn, %s'%(str(row[0]) + '(' + str(row[1]) + ')') 
                
        if row[4] != '' and row[4] != None and att_keys[4]:
            #Dimensional Attribute    
            #uda_query = """SELECT UDA_AUTO_KEY FROM 
                #USER_DEFINED_ATTRIBUTES WHERE 
                #UDA_CODE = 'DIMENSIONAL' AND AUTO_KEY_PREFIX='PNM'"""            
            #uda = selection_dir(uda_query,cr)
            #uda_auto_key = uda and uda[0] and uda[0][0] or None
            uda_auto_key = att_keys[4]
            udc_query = """SELECT UDC_AUTO_KEY FROM 
                UDA_CHECKED WHERE AUTO_KEY = %s
                AND UDA_AUTO_KEY = %s"""%(pnm_auto_key,uda_auto_key)
            udc = selection_dir(udc_query,cr)
            udc_auto_key = udc and udc[0] and udc[0][0] or None
            if uda_auto_key and not udc_auto_key:  
                query="""INSERT INTO UDA_CHECKED (UDA_AUTO_KEY,AUTO_KEY,ATTRIBUTE_VALUE) 
                  VALUES(%s,%s,'%s')"""%(uda_auto_key,pnm_auto_key,row[4])            
                error = insertion_dir(query,cr)
            elif uda_auto_key and udc_auto_key:
                query="""UPDATE UDA_CHECKED SET ATTRIBUTE_VALUE = '%s'
                    WHERE UDC_AUTO_KEY = %s"""%(row[4],udc_auto_key)
                error = updation_dir(query,cr)
            else:
                error = "Dimensional attribute not defined for PNs."
            if error and error != '{"recs": ""}':              
                fail_msg += '\nFailed insert for DIMENSIONAL, pn = ' + str(row[0]) + ' - ' + error
            else:
                rec_input += '\nSynched DIMENSIONAL attribute for pn, %s'%(str(row[0]) + '(' + str(row[1]) + ')')  
                
        if row[5] != '' and row[5] != None and att_keys[5]:   
            #NDT Attribute    
            #uda_query = """SELECT UDA_AUTO_KEY FROM 
            #    USER_DEFINED_ATTRIBUTES WHERE 
            #    UDA_CODE = 'NDT' AND AUTO_KEY_PREFIX='PNM'"""            
            #uda = selection_dir(uda_query,cr)
            #uda_auto_key = uda and uda[0] and uda[0][0] or None
            uda_auto_key = att_keys[5]
            udc_query = """SELECT UDC_AUTO_KEY FROM 
                UDA_CHECKED WHERE AUTO_KEY = %s
                AND UDA_AUTO_KEY = %s"""%(pnm_auto_key,uda_auto_key)
            udc = selection_dir(udc_query,cr)
            udc_auto_key = udc and udc[0] and udc[0][0] or None
            if uda_auto_key and not udc_auto_key:  
                query="""INSERT INTO UDA_CHECKED (UDA_AUTO_KEY,AUTO_KEY,ATTRIBUTE_VALUE) 
                  VALUES(%s,%s,'%s')"""%(uda_auto_key,pnm_auto_key,row[5])            
                error = insertion_dir(query,cr)
            elif uda_auto_key and udc_auto_key:
                query="""UPDATE UDA_CHECKED SET ATTRIBUTE_VALUE = '%s'
                    WHERE UDC_AUTO_KEY = %s"""%(row[5],udc_auto_key)
                error = updation_dir(query,cr)
            else:
                error = "NDT attribute not defined for PNs."
            if error and error != '{"recs": ""}':             
                fail_msg += '\nFailed insert for NDT, pn = ' + str(row[0]) + ' - ' + error
            else:
                rec_input += '\nSynched NDT attribute for pn, %s'%(str(row[0]) + '(' + str(row[1]) + ')')  
        
        if row[6] != '' and row[6] != None and att_keys[6]:
            #BoM Offset Attribute    
            #uda_query = """SELECT UDA_AUTO_KEY FROM 
            #    USER_DEFINED_ATTRIBUTES WHERE 
            #    UDA_CODE = 'DEFAULTBOMOFFSET' AND AUTO_KEY_PREFIX='PNM'"""            
            #uda = selection_dir(uda_query,cr)
            #uda_auto_key = uda and uda[0] and uda[0][0] or None
            uda_auto_key = att_keys[6]
            udc_query = """SELECT UDC_AUTO_KEY FROM 
                UDA_CHECKED WHERE AUTO_KEY = %s
                AND UDA_AUTO_KEY = %s"""%(pnm_auto_key,uda_auto_key)
            udc = selection_dir(udc_query,cr)
            udc_auto_key = udc and udc[0] and udc[0][0] or None
            if uda_auto_key and not udc_auto_key:  
                query="""INSERT INTO UDA_CHECKED (UDA_AUTO_KEY,AUTO_KEY,ATTRIBUTE_VALUE) 
                  VALUES(%s,%s,'%s')"""%(uda_auto_key,pnm_auto_key,row[6])            
                error = insertion_dir(query,cr)
            elif uda_auto_key and udc_auto_key:
                query="""UPDATE UDA_CHECKED SET ATTRIBUTE_VALUE = '%s'
                    WHERE UDC_AUTO_KEY = %s"""%(row[6],udc_auto_key)
                error = updation_dir(query,cr)
            else:
                error = "Default BoM Offset attribute not defined for PNs."
            if error and error != '{"recs": ""}':            
                fail_msg += '\nFailed insert for DEFAULTBOMOFFSET, pn = ' + str(row[0]) + ' - ' + error
            else:
                rec_input += '\nSynched DEFAULTBOMOFFSET attribute for pn, %s'%(str(row[0]) + '(' + str(row[1]) + ')') 
                
        if row[7] != '' and row[7] != None and att_keys[7]:
            #Offset Attribute    
            #uda_query = """SELECT UDA_AUTO_KEY FROM 
            #    USER_DEFINED_ATTRIBUTES WHERE 
            #    UDA_CODE = 'OFFSET' AND AUTO_KEY_PREFIX='PNM'"""            
            #uda = selection_dir(uda_query,cr)
            #uda_auto_key = uda and uda[0] and uda[0][0] or None
            uda_auto_key = att_keys[7]
            udc_query = """SELECT UDC_AUTO_KEY FROM 
                UDA_CHECKED WHERE AUTO_KEY = %s
                AND UDA_AUTO_KEY = %s"""%(pnm_auto_key,uda_auto_key)
            udc = selection_dir(udc_query,cr)
            udc_auto_key = udc and udc[0] and udc[0][0] or None
            if uda_auto_key and not udc_auto_key:  
                query="""INSERT INTO UDA_CHECKED (UDA_AUTO_KEY,AUTO_KEY,ATTRIBUTE_VALUE) 
                  VALUES(%s,%s,'%s')"""%(uda_auto_key,pnm_auto_key,row[7])            
                error = insertion_dir(query,cr)
            elif uda_auto_key and udc_auto_key:
                query="""UPDATE UDA_CHECKED SET ATTRIBUTE_VALUE = '%s'
                    WHERE UDC_AUTO_KEY = %s"""%(row[7],udc_auto_key)
                error = updation_dir(query,cr)
            else:
                error = "Offset attribute not defined for PNs."
            if error and error != '{"recs": ""}':            
                fail_msg += '\nFailed insert for OFFSET, pn = ' + str(row[0]) + ' - ' + error
            else:
                rec_input += '\nSynched OFFSET attribute for pn, %s'%(str(row[0]) + '(' + str(row[1]) + ')') 
                
        if row[8] != '' and row[8] != None and att_keys[8]:
            #Rank Attribute    
            #uda_query = """SELECT UDA_AUTO_KEY FROM 
            #    USER_DEFINED_ATTRIBUTES WHERE 
            #    UDA_CODE = 'RANK' AND AUTO_KEY_PREFIX='PNM'"""            
            #uda = selection_dir(uda_query,cr)
            #uda_auto_key = uda and uda[0] and uda[0][0] or None
            uda_auto_key = att_keys[8]
            udc_query = """SELECT UDC_AUTO_KEY FROM
                UDA_CHECKED WHERE AUTO_KEY = %s
                AND UDA_AUTO_KEY = %s"""%(pnm_auto_key,uda_auto_key)
            udc = selection_dir(udc_query,cr)
            udc_auto_key = udc and udc[0] and udc[0][0] or None
            if uda_auto_key and not udc_auto_key:  
                query="""INSERT INTO UDA_CHECKED (UDA_AUTO_KEY,AUTO_KEY,ATTRIBUTE_VALUE) 
                  VALUES(%s,%s,'%s')"""%(uda_auto_key,pnm_auto_key,row[8])            
                error = insertion_dir(query,cr)
            elif uda_auto_key and udc_auto_key:
                query="""UPDATE UDA_CHECKED SET ATTRIBUTE_VALUE = '%s'
                    WHERE UDC_AUTO_KEY = %s"""%(row[8],udc_auto_key)
                error = updation_dir(query,cr)
            else:
                error = "Rank attribute not defined for PNs."
            if error and error != '{"recs": ""}':             
                fail_msg += '\nFailed insert for RANK, pn =  ' + str(row[0]) + ' - ' + error
            else:
                rec_input += '\nSynched RANK attribute for pn, %s'%(str(row[0]) + '(' + str(row[1]) + ')') 
                
    orcl_commit(con=con) 
    if error == '{"recs": ""}':
        error = ''    
    if rec_input:    
        aud_status = 'success'
    else:
        aud_status = 'failure'
    msg = 'Successful update.'        
    right_now = datetime.now()
    right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')        
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='part-attributes')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        new_val = rec_input + '. ' + fail_msg
        field_changed = 'Part Attributes App Synch.'
        field_changed += error
        error += register_audit_trail(user_rec,field_changed,new_val,right_now,app_id,quapi,status=aud_status) 
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,rec_input,fail_msg
        
@shared_task
def get_loc_stocks(quapi_id,session_id,loc_text):
    error,msg='',''
    from polls.models import QueryApi,WOStatus    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not (cr and con):
        return 'Cannot connect to Oracle',msg 
    from polls.models import QueryApi      
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None  
    
    query = """SELECT DISTINCT VW.PARENT_WO,WO.SI_NUMBER,P.PN,
        P.DESCRIPTION,WS.SEVERITY ||'-'|| WS.DESCRIPTION,WO.DUE_DATE,WO.RANK,
        WO.SYSUR_AUTO_KEY,UDL.UDL_CODE,L.LOCATION_CODE,
        S.LOC_VALIDATED,S.SERIAL_NUMBER,C.COMPANY_ABBREV,
        CASE WHEN VTL.WOT_AUTO_KEY IS NOT NULL AND VTL.START_TIME IS NOT NULL
        AND VTL.STOP_TIME IS NULL THEN 'green-check.png' ELSE 'blank.png' END,
        WH.WAREHOUSE_CODE,S.STM_AUTO_KEY,S.STOCK_LINE,S.LOC_AUTO_KEY FROM STOCK S         
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
        LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
        LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_005
        LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
        LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY = S.WHS_AUTO_KEY
        LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = WO.WOS_AUTO_KEY
        LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = WO.CMP_AUTO_KEY
        LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
        LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
        LEFT JOIN WO_TASK WT ON WT.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
        LEFT JOIN VIEW_WO_TASK_LABOR VTL ON WT.WOT_AUTO_KEY = VTL.WOT_AUTO_KEY
        WHERE
        S.LOC_AUTO_KEY IN
        (SELECT LOC_AUTO_KEY FROM LOCATION WHERE UPPER(LOCATION_CODE) LIKE UPPER('%s%s'))
        AND S.QTY_RESERVED > 0 AND S.QTY_OH > 0
        ORDER BY DUE_DATE ASC NULLS LAST
    """%(loc_text,'%')
    recs = selection_dir(query,cr)
    stock_recs = []
    poss_dupes = {}
    count = 0
   
    for rec in recs:
    
        if rec[17] and not rec[10]:
        
            query="""SELECT TIME_STAMP FROM SA_LOG WHERE 
                STM_AUTO_KEY = %s
                NEW_LOC_AUTO_KEY=%s
                AND NEW_LOC_AUTO_KEY <> OLD_LOC_AUTO_KEY 
                AND ROWNUM <= 1
                ORDER BY STA_AUTO_KEY DESC
            """%(rec[15],rec[17])
            
            timestamp = selection_dir(query,cr)
            timestamp = timestamp and timestamp[0] and timestamp[0][0]
            rec[10] = timestamp
            
        dupe_pos = poss_dupes and rec[15] in poss_dupes
        dupe_pos = dupe_pos and poss_dupes[rec[15]] or None
        
        if dupe_pos: 
            if rec[13] == 'green-check.png':
                stock_recs.append(rec)
                #if count != dupe_pos:
                stock_recs.remove(recs[dupe_pos])
        else:
            stock_recs.append(rec)
            
        poss_dupes[rec[15]] = count
        count += 1
        
    del_stock = WOStatus.objects.filter(session_id=session_id).delete()      
    if recs:            
        res = create_wo_mgmt_bulk(session_id,stock_recs)      
        if not res:
            error = 'There was a problem creating stock lines'   
    return error,msg

@shared_task
def get_loc_counts(quapi_id,loc_list):
    error,msg='',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not (cr and con):
        return 'Cannot connect to Oracle',msg 
    from polls.models import WOStatus as wos_obj,QueryApi      
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None      
    counts = []
    error = ''   
    for loc in loc_list:       
        query = """SELECT STM_AUTO_KEY FROM STOCK WHERE
        LOC_AUTO_KEY IN 
        (SELECT LOC_AUTO_KEY FROM LOCATION WHERE UPPER(LOCATION_CODE) LIKE UPPER('%s%s%s'))
        AND QTY_RESERVED > 0 AND QTY_OH > 0
        """%('%',loc,'%')
        recs = selection_dir(query,cr)
        loc_count = len(recs)            
        counts.append(loc_count)                
    return error,counts
    
@shared_task
def update_attributes(quapi_id,session_id,sysur_auto_key,att_recs):
    error,msg = '',''
    aud_status = 'failure'
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)   
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    #update the att_recs passed in
    #val,pnm,'pnm','att_code'
    for rec in att_recs:
        uda_sub = """(SELECT UDA_AUTO_KEY FROM USER_DEFINED_ATTRIBUTES WHERE 
                      UPPER(AUTO_KEY_PREFIX) = UPPER('%s') AND UDA_CODE = '%s')"""%(rec[2],rec[3])
        query = """SELECT ATTRIBUTE_VALUE FROM UDA_CHECKED
            WHERE AUTO_KEY = %s AND UDA_AUTO_KEY = %s"""%(rec[1],uda_sub)
        recs = selection_dir(query,cr)
        if recs:
            query = """UPDATE UDA_CHECKED SET ATTRIBUTE_VALUE = '%s'
                WHERE AUTO_KEY = %s AND UDA_AUTO_KEY = %s"""%(rec[0],rec[1],uda_sub)
            error = updation_dir(query,cr)
        else:
            query = """INSERT INTO UDA_CHECKED (ATTRIBUTE_VALUE,AUTO_KEY,UDA_AUTO_KEY) 
                VALUES('%s','%s',%s)"""%(rec[0],rec[1],uda_sub)
            error = insertion_dir(query,cr)
    #if error == '{"recs": ""}' or error == '':
        #error,msg = get_part_attributes(quapi_id,sysur_auto_key,session_id,rec[1],stype='one')
    if not error or error == '{"recs": ""}':
        #if there were no errors, we mark this as a successful update. 
        orcl_commit(con=con)       
        aud_status = 'success'
        msg = 'Successful update.'
        error = ''
    else:
        aud_status = 'failure'           
    right_now = datetime.now()
    right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')       
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='inspection')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        rec_input = 'Update attribute for part: %s - '%(rec[4])
        new_val = rec_input
        field_changed = 'Inspection App Stock Update - '
        field_changed += error
        error += register_audit_trail(user_rec,field_changed,new_val,right_now,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,msg

@shared_task
def update_inspection(user_id,quapi_id,session_id,sysur_auto_key,stm_auto_key,\
        wo_number,ctrl_number,ctrl_id,\
        bag_no='',part_number='',condition='',serial_number='',\
        consignment='',qty_oh=0,notes='',del_previous=True,quantity=0,\
        wot_seq='',wot_desc=''):    
    error,msg,set_clause,pcc_auto_key,cnc_auto_key = '','','',None,None
    description,list_price = '',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)    
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
        
    if bag_no:
        if set_clause:
            set_clause += ','
        set_clause += """IC_UDF_020 = '%s'"""%(bag_no)
        
    if serial_number:
        if set_clause:
            set_clause += ','
        set_clause += """SERIAL_NUMBER = '%s'"""%(serial_number) 
        #if part_number or serial_number:                     
        
    if condition:
        if set_clause:
            set_clause += ','
        set_clause += """ PCC_AUTO_KEY = (SELECT PCC_AUTO_KEY FROM 
        PART_CONDITION_CODES 
        WHERE UPPER(CONDITION_CODE) = UPPER('%s'))"""%(condition)
        
        query = """SELECT PCC_AUTO_KEY,CONDITION_CODE 
            FROM PART_CONDITION_CODES 
            WHERE UPPER(CONDITION_CODE) = UPPER('%s')"""%condition
        pcc = selection_dir(query,cr)
        pcc_auto_key = pcc and pcc[0] and pcc[0][0] or None
        cond_code =  pcc and pcc[0] and pcc[0][1] or ''
        
        if cond_code in ['rej','REJ','Rej','ReJ','reJ','rEJ','rEj','REj']:
        
            query = """SELECT WOB_AUTO_KEY,STR_AUTO_KEY
                FROM STOCK_RESERVATIONS
                WHERE STM_AUTO_KEY = %s"""%stm_auto_key
            wob = selection_dir(query,cr)
            wob_auto_key = wob and wob[0] and wob[0][0]
            str_auto_key = wob and wob[0] and wob[0][1]
            
            if wob_auto_key:
                #update wo_bom activity
                query = """UPDATE WO_BOM SET ACTIVITY = '%s' 
                WHERE WOB_AUTO_KEY = %s"""%('Replace',wob_auto_key)
                error = updation_dir(query,cr)
                
            if str_auto_key:
                query = """DELETE FROM STOCK RESERVATIONS
                WHERE STR_AUTO_KEY = %s
                """%str_auto_key
                error = updation_dir(query,cr) 
                
    if set_clause:            
        query = """UPDATE STOCK SET %s WHERE STM_AUTO_KEY = %s"""%(set_clause,stm_auto_key)
        error = updation_dir(query,cr)
        
    if consignment:
        #if set_clause:
        #    set_clause += ','
        #set_clause += """ CNC_AUTO_KEY = (SELECT CNC_AUTO_KEY FROM CONSIGNMENT_CODES 
        #WHERE UPPER(CONSIGNMENT_CODE) = UPPER('%s'))"""%(consignment)
        query = """SELECT CNC_AUTO_KEY FROM CONSIGNMENT_CODES WHERE UPPER(CONSIGNMENT_CODE) = UPPER('%s')"""%consignment
        cnc = selection_dir(query,cr)
        cnc_auto_key = cnc and cnc[0] and cnc[0][0] or None
                
    pnm_auto_key = None
    if part_number:
    
        query = """SELECT PNM_AUTO_KEY,DESCRIPTION,LIST_PRICE FROM PARTS_MASTER 
        WHERE UPPER(PN) = UPPER('%s')"""%part_number
        
        pnm = selection_dir(query,cr)
        pnm_auto_key = pnm and pnm[0] and pnm[0][0] or None
        description = pnm and pnm[0] and pnm[0][1] or ''
        list_price = pnm and pnm[0] and pnm[0][2] or 0
        
        if not pnm:
            error = 'No PN found.'
                          
    if notes:
        query = """UPDATE STOCK SET NOTES = '%s'
            WHERE STM_AUTO_KEY = %s"""%(notes,stm_auto_key)
        error = updation_dir(query,cr)
        
        if error != '{"recs": ""}':
            return error,msg
           
    if consignment or quantity or pnm_auto_key:
       
        qty_update = qty_oh
        
        if quantity:
            qty_update = quantity
        
        reservations = []        
        if (quantity and qty_oh != quantity) or pnm_auto_key:
            query = """SELECT WOO_AUTO_KEY,WOB_AUTO_KEY,ROD_AUTO_KEY,
            SOD_AUTO_KEY,SMD_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY,
            STR_AUTO_KEY
            FROM STOCK_RESERVATIONS
            WHERE STM_AUTO_KEY = %s
            """%stm_auto_key
            
            reservations = selection_dir(query,cr)
            
            for res in reservations:
                if res[0]:
                    res_type = 'WOO_AUTO_KEY'
                elif res[1]:
                    res_type = 'WOB_AUTO_KEY'
                elif res[2]:
                    res_type = 'ROD_AUTO_KEY'
                elif res[3]:
                    res_type = 'SOD_AUTO_KEY'
                elif res[4]:
                    res_type = 'SMD_AUTO_KEY'
                else:
                    continue
                str_auto_key = res[7]
                
                
                query = """DELETE FROM STOCK_RESERVATIONS
                    WHERE STR_AUTO_KEY = %s"""%str_auto_key   
                error = updation_dir(query,cr)
                
                if error != '{"recs": ""}':
                    return error,msg
                    
                squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                    EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY =
                    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG 
                    WHERE STM_AUTO_KEY = %s 
                    AND EMPLOYEE_CODE = 'DBA')
                    """%(sysur_auto_key,user_id[:9],stm_auto_key)
                error = updation_dir(squery,cr)
                
                if error != '{"recs": ""}':
                    return error,msg
                    
                squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                    EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY =
                    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG 
                    WHERE STM_AUTO_KEY = %s 
                    AND EMPLOYEE_CODE = 'DBA')
                    """%(sysur_auto_key,user_id[:9],stm_auto_key)
                error = updation_dir(squery,cr)
                
                if error != '{"recs": ""}':
                    return error,msg
        
        if consignment or condition or quantity:                               
            
            srecs = get_wos_from_rack_beta(quapi=quapi,stm_auto_key=stm_auto_key)
            rec = srecs and srecs[0] or []
            qty_update = rec and rec[29] or 1

            if not rec:
                return 'Stock not found.',msg

            else:
                """
                call the stock transfer procedure with user-entered qty and see
                if new stock record gets created.
                """
                if quantity != '':
                
                    if float(quantity) > rec[29]:
                        error = 'Qty more than expected.'
                        return error,msg
                        
                    else:
                        qty_update = quantity 
                
                params=[]
                params.append(stm_auto_key)#stm_auto_key
                params.append(qty_update)#qty_oh
                params.append(rec[30] or 1)#syscm_auto_key
                params.append(pcc_auto_key or rec[31])#pcc_auto_key                       
                params.append(cnc_auto_key or rec[32])#cnc_auto_key
                params.append(rec[21] or '')#loc_auto_key 
                params.append(rec[22] or '')#whs_auto_key
                params.append(rec[33] or '')#stc_auto_key
                params.append(rec[34] or '')#dpt_auto_key
                params.append(rec[20] or '')#str_auto_key  
                params.append(rec[35])#qty_reserved
                params.append(rec[36] or '')#sod_auto_key
                params.append(rec[37] or '')#rod_auto_key
                params.append(rec[38] or '')#wob_auto_key
                params.append(rec[39] or '')#pod_auto_key
                params.append(rec[11] or '')#woo_auto_key                     
                error = qry_stock_transfer(sysur_auto_key,user_id,params,\
                    quapi,recs=srecs,cr=cr,con=con)                
                
                squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                    EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY =
                    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG 
                    WHERE STM_AUTO_KEY = %s 
                    AND EMPLOYEE_CODE = 'DBA')
                    """%(sysur_auto_key,user_id[:9],stm_auto_key)

                error = updation_dir(squery,cr)
   
                squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                    EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY =
                    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG 
                    WHERE STM_AUTO_KEY = %s 
                    AND EMPLOYEE_CODE = 'DBA')
                    """%(sysur_auto_key,user_id[:9],stm_auto_key)

                error = updation_dir(squery,cr)
                    
        if pnm_auto_key:
            error = qry_pn_transfer(cr,sysur_auto_key,user_id,stm_auto_key,pnm_auto_key,quantity) 
        
        if (qty_oh != float(quantity) or pnm_auto_key) and reservations:
        
            for res in reservations:
            
                if res[0]:
                    res_type = 'WOO_AUTO_KEY'
                    reserved_to = res[0]
                elif res[1]:
                    res_type = 'WOB_AUTO_KEY'
                    reserved_to = res[1]
                elif res[2]:
                    res_type = 'ROD_AUTO_KEY'
                    reserved_to = res[2]
                elif res[3]:
                    res_type = 'SOD_AUTO_KEY'
                    reserved_to = res[3]
                elif res[4]:
                    res_type = 'SMD_AUTO_KEY'
                    reserved_to = res[4]
                else:
                    continue
                
                qty_reserved = res[5]
                if qty_oh and quantity:                
                    qty_reserved = qty_oh - float(quantity)
                    
                sysur_res = res[6]
                              
                squery = """INSERT INTO STOCK_RESERVATIONS 
                            (STR_AUTO_KEY,STM_AUTO_KEY,%s,QTY_RESERVED,SYSUR_AUTO_KEY) 
                            VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s)
                            """%(res_type,stm_auto_key,reserved_to,qty_reserved,sysur_res)
                error = insertion_dir(squery,cr)
                
                squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                    EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY =
                    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG 
                    WHERE STM_AUTO_KEY = %s 
                    AND EMPLOYEE_CODE = 'DBA')
                    """%(sysur_auto_key,user_id[:9],stm_auto_key)

                error = updation_dir(squery,cr)
   
                squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                    EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY =
                    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG 
                    WHERE STM_AUTO_KEY = %s 
                    AND EMPLOYEE_CODE = 'DBA')
                    """%(sysur_auto_key,user_id[:9],stm_auto_key)

                error = updation_dir(squery,cr)
                
            query = """SELECT STM_AUTO_KEY,CTRL_NUMBER,CTRL_ID FROM STOCK
                WHERE ROWNUM <= 1                
                ORDER BY STM_AUTO_KEY DESC
                """
   
            new_stm = selection_dir(query,cr)
            new_stm = new_stm and new_stm[0]
            new_stm_auto = new_stm and new_stm[0]
            ctrl_number = new_stm and new_stm[1]
            ctrl_id = new_stm and new_stm[2]
            
            if new_stm_auto and quantity and qty_update:
                squery = """INSERT INTO STOCK_RESERVATIONS 
                            (STR_AUTO_KEY,STM_AUTO_KEY,%s,QTY_RESERVED,SYSUR_AUTO_KEY) 
                            VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s)
                            """%(res_type,new_stm_auto,reserved_to,qty_update,sysur_auto_key)
                error = insertion_dir(squery,cr)
                
                if error:
                    return error,msg
                    
                else:
                    squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                        EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY =
                        (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG 
                        WHERE STM_AUTO_KEY = %s 
                        AND EMPLOYEE_CODE = 'DBA')
                        """%(sysur_auto_key,user_id[:9],stm_auto_key)

                    error = updation_dir(squery,cr)
                    
                    if error != '{"recs": ""}':
                        return error,msg 
                        
                    squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                        EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY =
                        (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG 
                        WHERE STM_AUTO_KEY = %s 
                        AND EMPLOYEE_CODE = 'DBA')
                        """%(sysur_auto_key,user_id[:9],stm_auto_key)

                    error = updation_dir(squery,cr)
                    
                    if error != '{"recs": ""}':
                        return error,msg 

    hold_line = False            
    if error == '{"recs": ""}' or error == '':
        #get updated stock line and synch it
        #locally for lookup and display 
        error,msg,hold_line = get_inspection(quapi_id,session_id,\
            wo_number,ctrl_number,ctrl_id,del_previous=True,cr=cr)        
        if not error:
            orcl_commit(con=con)        
            aud_status = 'success'
            msg = 'Successful update.'
        else:
            aud_status = 'failure' 
    else: 
        aud_status = 'failure'
        return error,msg
    
    from polls.models import WOStatus
    stock_rec = WOStatus.objects.filter(session_id=session_id)
    stock_rec = stock_rec and stock_rec[0] or []
   
    if stock_rec:
        notes = stock_rec.notes
        part_number = stock_rec.part_number
        description = stock_rec.description
        wo_number = stock_rec.wo_number
        list_price = stock_rec.misc_cost    
                
    right_now = datetime.now()
    right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')        
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='inspection')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        rec_input = 'Update stock record, %s:'%(stm_auto_key)
        if part_number:
            rec_input += ' with part number = %s'%part_number
        if serial_number:
            rec_input += ' with serial number = %s'%serial_number
        if condition:
            rec_input += 'and condition = %s'%condition
        if consignment:
            rec_input += 'and consignment = %s'%consignment
        new_val = rec_input
        field_changed = 'Inspection App Stock Update'
        field_changed += error
        parameters = ['',str(wo_number),str(right_now)]
        parameters += [str(wot_seq),str(wot_desc)]
        parameters += [str(part_number),description,'','']
        
        reject_part = ''
        if condition in ['Rej','REJ','rej','scrap','SCRAP','Scrap']:
            reject_part = 'Part Rejected'
           
        #parameters += [quantity or qty_oh,notes,str(list_price)]
        parameters += ['',notes,str(list_price)]
        parameters += [hold_line,reject_part,'','','']
        
        if user_rec:
            new_val = set_clause + '. ' + error
            field_changed += error
            error += register_audit_trail(
            user_rec,
            field_changed,
            new_val,
            right_now,
            app_id,
            quapi,
            status=aud_status,
            reg_events= reject_part and ['Condition Code to REJ'] or [],
            parameters=parameters,
            sysur_auto_key=sysur_auto_key,
            ) 
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,msg 
    
@shared_task
def get_inspection(quapi_id,session_id,wo_number,ctrl_number,ctrl_id,del_previous=True,cr=False):
    error,msg = '',''
    recs = []
    if not cr:
        from polls.models import QueryApi    
        quapi = QueryApi.objects.filter(id=quapi_id)
        quapi = quapi and quapi[0] or None
        cr,con = get_cursor_con(quapi)  
        if not (cr and con):
            return 'Cannot connect to Oracle',msg
            
    if ctrl_number and ctrl_id:
    
        where_clause = " WHERE S.CTRL_NUMBER = %s AND S.CTRL_ID = %s"%(ctrl_number,ctrl_id)
        where_clause += """ AND (SR.STR_AUTO_KEY IN (SELECT STR.STR_AUTO_KEY FROM STOCK_RESERVATIONS STR 
              WHERE STR.STM_AUTO_KEY=SR.STM_AUTO_KEY) 
              OR SR.STR_AUTO_KEY IS NULL)
              AND S.HISTORICAL_FLAG = 'F'
              AND S.QTY_OH > 0"""
              
        query = """SELECT S.STM_AUTO_KEY,P.PN,P.DESCRIPTION,S.QTY_OH,PCC.CONDITION_CODE,                                                                        
            WO.SI_NUMBER,CC.CONSIGNMENT_CODE,L.LOCATION_CODE,S.SERIAL_NUMBER,
            S.CTRL_NUMBER,S.CTRL_ID,
            S.STOCK_LINE,TO_CHAR(S.NOTES),S.IC_UDF_008,PCC.DESCRIPTION,P.PNM_AUTO_KEY,
            S.REC_DATE,'','',SR.STR_AUTO_KEY,WB.WOB_AUTO_KEY,PCC.COND_LEVEL,
            S.IC_UDF_020,                                                          
            C.COMPANY_NAME,                                                                                           
            WO.WO_UDF_002,S.HOLD_LINE,WO.SI_NUMBER,P.LIST_PRICE
            FROM STOCK S        
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
            LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
            LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = WO.CMP_AUTO_KEY             
            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
            LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
            LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
            LEFT JOIN CONSIGNMENT_CODES CC ON CC.CNC_AUTO_KEY = S.CNC_AUTO_KEY        
            %s"""%where_clause
            
        recs = selection_dir(query,cr)       
        
        if not recs:
        
            if len(wo_number) > 12:
                ctrl_number = wo_number[:7]              
                ctrl_id = wo_number[8:]
                
            elif len(wo_number) == 12:
                ctrl_number = wo_number[:6]
                ctrl_id = wo_number[7:] 

            where_clause = " WHERE S.CTRL_NUMBER = %s AND S.CTRL_ID = %s"%(ctrl_number,ctrl_id)
            where_clause += """ AND (SR.STR_AUTO_KEY IN (SELECT STR.STR_AUTO_KEY FROM STOCK_RESERVATIONS STR 
                  WHERE STR.STM_AUTO_KEY=SR.STM_AUTO_KEY) 
                  OR SR.STR_AUTO_KEY IS NULL)
                  AND S.HISTORICAL_FLAG = 'F'
                  AND S.QTY_OH > 0"""                

            query = """SELECT S.STM_AUTO_KEY,P.PN,P.DESCRIPTION,S.QTY_OH,PCC.CONDITION_CODE,                                                                        
                WO.SI_NUMBER,CC.CONSIGNMENT_CODE,L.LOCATION_CODE,S.SERIAL_NUMBER,
                S.CTRL_NUMBER,S.CTRL_ID,
                S.STOCK_LINE,TO_CHAR(S.NOTES),S.IC_UDF_008,PCC.DESCRIPTION,P.PNM_AUTO_KEY,
                S.REC_DATE,'','',SR.STR_AUTO_KEY,WB.WOB_AUTO_KEY,PCC.COND_LEVEL,
                S.IC_UDF_020,                                                          
                C.COMPANY_NAME,                                                                                           
                WO.WO_UDF_002,S.HOLD_LINE,WO.SI_NUMBER,P.LIST_PRICE
                FROM STOCK S        
                LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
                LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
                LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
                LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = WO.CMP_AUTO_KEY             
                LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
                LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
                LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
                LEFT JOIN CONSIGNMENT_CODES CC ON CC.CNC_AUTO_KEY = S.CNC_AUTO_KEY        
                %s"""%where_clause
            recs = selection_dir(query,cr)
                 
    elif wo_number or (ctrl_number and not recs):
        if ctrl_number:
            wo_number = ctrl_number + ctrl_id
    
        where_clause = """ WHERE UPPER(W.SI_NUMBER) = UPPER('%s')
            OR UPPER(WO.SI_NUMBER) = UPPER('%s')"""%(wo_number,wo_number)
            
        where_clause += """ AND (SR.STR_AUTO_KEY IN (SELECT STR.STR_AUTO_KEY FROM STOCK_RESERVATIONS STR 
              WHERE STR.STM_AUTO_KEY=SR.STM_AUTO_KEY) 
              OR SR.STR_AUTO_KEY IS NULL)
              AND S.HISTORICAL_FLAG = 'F'
              AND S.QTY_OH > 0"""            
              
        query = """SELECT S.STM_AUTO_KEY,P.PN,P.DESCRIPTION,S.QTY_OH,PCC.CONDITION_CODE,
                '%s',
                CC.CONSIGNMENT_CODE,L.LOCATION_CODE,S.SERIAL_NUMBER,S.CTRL_NUMBER,S.CTRL_ID,
                S.STOCK_LINE,TO_CHAR(S.NOTES),S.IC_UDF_008,PCC.DESCRIPTION,P.PNM_AUTO_KEY,
                S.REC_DATE,'','',SR.STR_AUTO_KEY,WB.WOB_AUTO_KEY,PCC.COND_LEVEL,
                S.IC_UDF_020,
                CASE WHEN C.COMPANY_NAME IS NOT NULL THEN C.COMPANY_NAME ELSE
                CO.COMPANY_NAME END,
                CASE WHEN W.WO_UDF_002 IS NOT NULL THEN W.WO_UDF_002 ELSE
                WO.WO_UDF_002 END,S.HOLD_LINE,'%s',P.LIST_PRICE
                FROM STOCK S                                                                         
                LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
                LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
                LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = W.CMP_AUTO_KEY             
                LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
                LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
                LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
                LEFT JOIN CONSIGNMENT_CODES CC ON CC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
                LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
                LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
                LEFT JOIN COMPANIES CO ON CO.CMP_AUTO_KEY = WO.CMP_AUTO_KEY            
                %s"""%(wo_number,wo_number,where_clause)
                
        recs = selection_dir(query,cr)
        
    stms = []   
    for rec in recs:
    
        if rec[0] not in stms:
            stms.append(rec[0])
            if ctrl_number and ctrl_id and not rec[5]:
                #lookup by sti
                query="""SELECT W.SI_NUMBER FROM WO_OPERATION W
                WHERE W.WOO_AUTO_KEY = (SELECT WOO_AUTO_KEY FROM WO_BOM
                WHERE WOB_AUTO_KEY = (SELECT WOB_AUTO_KEY FROM STOCK_TI
                WHERE STM_AUTO_KEY=%s AND TI_TYPE='T'))
                """%rec[0]
                si=selection_dir(query,cr)
                si_number = si and si[0] and si[0][0] or ''
                rec[5] = si_number
        
    hold_line = ''
    if recs:
        hold_line = recs[0][-1] == 'T' and 'Hold Line' or ''
        error = create_inspections(session_id,recs,wo_number,ctrl_number,ctrl_id,del_previous) 
    else:
        error = 'Record not found.'           
    return error,msg,hold_line
    
def create_inspections(session_id,recs,wo_number,ctrl_number,ctrl_id,del_previous):
    error = ''
    from polls.models import WOStatus

    for stock_rec in recs: 
        ocl_data = []
        ocl_data.append(WOStatus(
            stm_auto_key = stock_rec[0] or 0,
            part_number = stock_rec[1],
            description = stock_rec[2],
            qty_oh = stock_rec[3],
            condition_code = stock_rec[4],
            si_number = stock_rec[5] or stock_rec[26],
            wo_number = stock_rec[5] or stock_rec[26],
            ctrl_number = stock_rec[9],
            ctrl_id = stock_rec[10],
            consignment_code = stock_rec[6],
            location_code = stock_rec[7],
            serial_number = stock_rec[8] or 'N/A',
            stock_line = stock_rec[11],
            notes = stock_rec[12],
            slug = stock_rec[13],
            supdate_msg = stock_rec[14],
            pnm_auto_key = stock_rec[15] or 0,
            arrival_date = stock_rec[16][:10],
            po_number = stock_rec[17],
            ro_number = stock_rec[18],
            cond_level = stock_rec[19] or 0,
            cond_level_gsix = stock_rec[21] and int(stock_rec[21]) >= 6 and 'T' or 'F',
            cond_level_zero = stock_rec[21] and int(stock_rec[21]) == 0 and 'T' or 'F',
            spn_code = stock_rec[22],
            customer = stock_rec[23],
            time_loc = stock_rec[8],
            misc_cost = stock_rec[27],
            session_id = session_id,      
        ))
    if ocl_data:
        try:
            if del_previous:
                delete = WOStatus.objects.filter(session_id=session_id).delete()
            WOStatus.objects.bulk_create(ocl_data) or []    
        except Exception as exc:
            error += "Error, %s, creating record for WO#, %s"%(exc,stock_rec[5])
    return error
    
@shared_task
def create_locations(quapi_id,sysur_auto_key,label,warehouse,whs_auto_key,loc_list):
    error,msg = '',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    locations = []
    if not (cr and con):
        return 'Cannot connect to Oracle',msg         
    count = 0

    if warehouse and not whs_auto_key:
        query = """SELECT WHS_AUTO_KEY FROM 
        WAREHOUSE WHERE WAREHOUSE_CODE = '%s'
        """%warehouse
        
        whs = selection_dir(query,cr)
        whs_auto_key = whs and whs[0] and whs[0][0] or None
        
        if not whs:
            error = 'Warehouse not found.'
            return error,msg,locs
    
    for loc_name in loc_list:     
        count += 1 
        query = """SELECT LOC_AUTO_KEY FROM LOCATION 
            WHERE UPPER(LOCATION_CODE) = UPPER('%s')"""%loc_name 
        ex_loc = selection_dir(query,cr)
        ex_loc = ex_loc and ex_loc[0] and ex_loc[0][0]
        locs = selection_dir(query,cr) 
        
        if ex_loc:    
            loc_name += '_' + str(count)
  
        query = """INSERT INTO LOCATION 
            (LOC_AUTO_KEY,LOCATION_CODE,DESCRIPTION,HISTORICAL) 
            VALUES(G_LOC_AUTO_KEY.NEXTVAL,'%s','%s','F')
            """%(loc_name,loc_name) 
            
        error = insertion_dir(query,cr)

                               
        query = """SELECT LOCATION_CODE,LOC_AUTO_KEY FROM LOCATION 
            WHERE ROWNUM <= 1 ORDER BY LOC_AUTO_KEY DESC"""
        locs = selection_dir(query,cr)
        loc_auto_key = locs and locs[0] and locs[0][1]
        locations.append(locs)
            
                            
                                                                
                                                     
                                               
                                               
                                                    

        if loc_auto_key and whs_auto_key and not error:
            insert_lwh = """INSERT INTO WAREHOUSE_LOCATIONS 
                     (LOC_AUTO_KEY,WHS_AUTO_KEY) 
                     VALUES (%s,%s)        
            """%(loc_auto_key,whs_auto_key)
            error = insertion_dir(insert_lwh,cr)
                
    if not error and count:
        msg = "Successfully added %s new locations"%count
        
    orcl_commit(con=con)        
        
    return error,msg,locations
  
@shared_task
def create_carts(quapi_id,sysur_auto_key,label,cart_list):
    error,msg = '',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    udls = []
    if not (cr and con):
        return 'Cannot connect to Oracle',msg         
    count = 0
    for cart_name in cart_list:     
        count += 1       
        query = """INSERT INTO USER_DEFINED_LOOKUPS 
        (UDL_AUTO_KEY,UDL_COLUMN_NAME,UDL_CODE,UDL_DESCRIPTION,SEQUENCE,HISTORICAL_FLAG) 
        VALUES(G_UDL_AUTO_KEY.NEXTVAL,'%s','%s','%s','%s','F')
        """%('IC_UDL_005',cart_name,cart_name,count)   
        error = insertion_dir(query,cr)  
           
    if not error and count:
        query = "SELECT UDL_AUTO_KEY FROM USER_DEFINED_LOOKUPS WHERE ROWNUM <= %s"%count 
        udls = selection_dir(query,cr)
        msg = "Successfully added %s new carts"%count
        orcl_commit(con=con)        
    return error,msg,udls

@shared_task
def set_overrides(quapi_id,session_id,woo_auto_key,override,sequence,sysur_auto_key):
    error,msg = '',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg       
    query = """UPDATE ORDER_CLAUSES SET     
        OVERRIDE = '%s' WHERE WOO_AUTO_KEY = %s
        AND SEQUENCE = '%s'
    """%(override,woo_auto_key,sequence)
    error = updation_dir(query,cr)
    if error == '{"recs": ""}':
        error = ''
        error,msg = get_overrides(quapi_id,session_id,'',woo_auto_key=woo_auto_key,cr=cr)
        orcl_commit(con=con)        
        aud_status = 'success'
        msg = 'Successful update.'
    else: 
        aud_status = 'failure'  
    right_now = datetime.now()
    right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')        
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='wo-order-clause')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        rec_input = 'Update override to %s for woo_auto_key = %s and sequence = %s'%(override,woo_auto_key,sequence)
        new_val = rec_input
        field_changed = 'Order Clause Override'
        field_changed += error
        error += register_audit_trail(user_rec,field_changed,new_val,right_now,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,msg
    
@shared_task    
def get_overrides(quapi_id,session_id,si_number,woo_auto_key=0,cr=None):
    error,msg,woo_clause = '','',''
    if not cr and si_number:
        from polls.models import QueryApi    
        quapi = QueryApi.objects.filter(id=quapi_id)
        quapi = quapi and quapi[0] or None
        cr,con = get_cursor_con(quapi)
        if not (cr and con):
            return 'Cannot connect to Oracle',msg
        woo_clause = "AND WOO.SI_NUMBER = '%s'"%si_number
    elif cr and woo_auto_key:
        woo_clause = "AND WOO.WOO_AUTO_KEY = '%s'"%woo_auto_key
    if woo_clause:        
        query = """SELECT WOO.SI_NUMBER, CLA.DESCRIPTION, OCL.OVERRIDE,
            WOO.WOO_AUTO_KEY, OCL.SEQUENCE
            FROM ORDER_CLAUSES OCL, CLAUSES CLA, WO_OPERATION WOO
            WHERE WOO.WOO_AUTO_KEY = OCL.WOO_AUTO_KEY
            AND OCL.CLA_AUTO_KEY = CLA.CLA_AUTO_KEY
            %s
            ORDER BY OCL.SEQUENCE DESC"""%woo_clause
        recs = selection_dir(query,cr)
        if recs:
            error = create_woo_ocls(session_id,recs) 
        else:
            error = 'No records found.'
    return error,msg        
    
def create_woo_ocls(session_id,recs):
    error = ''
    from polls.models import WOStatus
    ocl_data,error = [],''
    #WOO.SI_NUMBER, CLA.DESCRIPTION, OCL.OVERRIDE
    for row in recs:    
        ocl_data.append(WOStatus(
        wo_number = row[0],
        description = row[1],
        priority = row[2],
        woo_auto_key = row[3],
        wot_sequence = row[4],
        session_id = session_id,      
        ))
    if ocl_data:
        try:
            delete = WOStatus.objects.filter(session_id=session_id).delete()
            WOStatus.objects.bulk_create(ocl_data) or []    
        except Exception as exc:
            error += "Error, %s, with order clause rows for grid for WO#, %s"%(exc,row[0])
    return error

@shared_task
def correct_counts(quapi_id,session_id):
    error,msg = '',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg 
    query="""SELECT DISTINCT STM.STM_AUTO_KEY,PIH.ENTRY_DATE,PIH.PI_NUMBER,PID.QTY, 
    PID.QTY_FOUND,STM.QTY_OH,STM.CTRL_NUMBER,STM.CTRL_ID,STM.PNM_AUTO_KEY,
    STM.WHS_AUTO_KEY,STM.LOC_AUTO_KEY
    FROM PI_HEADER PIH, PI_DETAIL PID, STOCK STM, LOCATION LOC, STOCK_AUDIT STA
    WHERE PIH.PIH_AUTO_KEY = PID.PIH_AUTO_KEY
    AND PID.STM_AUTO_KEY = STM.STM_AUTO_KEY
    AND STM.LOC_AUTO_KEY = LOC.LOC_AUTO_KEY
    AND STM.STM_AUTO_KEY = STA.NEW_STM_AUTO_KEY
    AND STA.TIME_STAMP <= PIH.CLOSE_DATE
    AND PIH.ENTRY_DATE >= to_Date('07/30/2021','mm/dd/yyyy')
    AND PID.QTY = '1'
    AND PID.QTY_FOUND = '1'
    AND STM.QTY_OH = '0' 
    --AND STA.OLD_QTY_OH = '1'
    --AND STA.NEW_QTY_OH = '0'
    --AND STA.TRAN_TYPE = 'WO Issue'
    ORDER BY PIH.ENTRY_DATE DESC
    """
    sysur_auto_key = 900
    recs = selection_dir(query,cr)
    stms = []
    query ="""INSERT INTO PI_HEADER (PIH_AUTO_KEY,GEO_AUTO_KEY,PI_NUMBER,SYSUR_AUTO_KEY,ENTRY_DATE,OPEN_FLAG,WHS_AUTO_KEY) 
               VALUES(G_PIH_AUTO_KEY.NEXTVAL,1,'11111111',%s,TO_DATE('2022-01-25','YYYY-MM-DD'),'T',1)
        """%sysur_auto_key
    error = insertion_dir(query,cr)
    query = """SELECT PIH_AUTO_KEY FROM PI_HEADER ORDER BY PIH_AUTO_KEY DESC"""
    pih = selection_dir(query,cr)
    pih_auto_key = pih and pih[0] and pih[0][0] or None
    if pih_auto_key:      
        for rec in recs:
            stm_auto_key = rec[0]
            qty_oh = 0
            qty_found = rec[4]         
            ctrl_id = rec[6]
            ctrl_number = rec[7]
            pnm_auto_key = rec[8]
            whs_auto_key = rec[9]
            loc_auto_key = rec[10]
            query = """INSERT INTO PI_DETAIL (STM_AUTO_KEY,PID_AUTO_KEY,SYSUR_AUTO_KEY,QTY,QTY_FOUND,
                CTRL_ID,CTRL_NUMBER,PIH_AUTO_KEY,PNM_AUTO_KEY,WHS_AUTO_KEY,LOC_AUTO_KEY) 
                VALUES ('%s',G_PID_AUTO_KEY.NEXTVAL,%s,'%s','%s','%s','%s',%s,'%s','%s','%s')
                """%(stm_auto_key,sysur_auto_key,qty_oh,qty_found,ctrl_id,ctrl_number,pih_auto_key,pnm_auto_key,whs_auto_key,loc_auto_key)                
            error = insertion_dir(query,cr)
            
    orcl_commit(con=con)
    
@shared_task
def stock_picking(quapi_id,sysur_auto_key,session_id,wo_number,exact_match=True,wob_id_list=[]):
    error,msg,stock_msg = '','',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg      
    where_wo = exact_match and "W.SI_NUMBER = '%s'"%wo_number or ''
    where_wo = not exact_match and "W.SI_NUMBER LIKE '%s%s%s'"%('%',wo_number,'%') or where_wo
    where_clause = "AND %s"%where_wo 
    """
    next retrieve all WOB  and STM 
    where wob.woo_auto_key = (si_number input by user) and  wob.qty_needed <> wob.qty_issued + wob.qty_reserved and 
    wob.activity not in ( 'Turn-In', 'Work Order', 'Repair')
    group by wob.pn, stm.loc
    order by wob.pn, stm.loc
    GROUP BY P.PN,L.LOCATION_CODE
    WOBS: [3352, 4478]"""
    if wob_id_list:
        from polls.models import WOStatus as wos
        wostatus_list = wos.objects.filter(id__in = wob_id_list)

        for wob in wostatus_list:
            query = """SELECT WOB.QTY_NEEDED,S.STM_AUTO_KEY,
                S.CTRL_NUMBER, S.CTRL_ID FROM WO_BOM WOB, 
                PARTS_MASTER P, STOCK S
                WHERE WOB.WOB_AUTO_KEY = %s
                AND S.PNM_AUTO_KEY = P.PNM_AUTO_KEY
                AND P.PNM_AUTO_KEY = WOB.PNM_AUTO_KEY      
            """%(wob.wob_auto_key)
            stm_list = selection_dir(query,cr)
            for stm in stm_list:
                squery = """INSERT INTO STOCK_RESERVATIONS 
                    (STR_AUTO_KEY,STM_AUTO_KEY,WOB_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY) 
                    VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s)"""%(stm[1],wob.wob_auto_key,stm[0],sysur_auto_key)
                error = insertion_dir(squery,cr)
                if not error:
                    stock_msg += str(stm[2]) + '000000' + str(stm[3]) + '| '
            if stock_msg:
                msg += '\r\nBOM: ' + str(wob.wob_auto_key) + ' - reserved stock lines: ' + stock_msg
    else:    
        query = """SELECT W.SI_NUMBER, P.PN, P.DESCRIPTION, WB.QTY_NEEDED, WB.WOB_AUTO_KEY, 
            S.SERIAL_NUMBER, S.STOCK_LINE, S.CTRL_NUMBER, S.CTRL_ID, S.QTY_AVAILABLE 
            FROM WO_OPERATION W, PARTS_MASTER P, WO_BOM WB, STOCK S, LOCATION L
            WHERE S.PNM_AUTO_KEY = P.PNM_AUTO_KEY
            AND P.PNM_AUTO_KEY = WB.PNM_AUTO_KEY
            AND WB.WOO_AUTO_KEY = W.WOO_AUTO_KEY
            AND L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
            AND S.QTY_AVAILABLE > 0
            AND WB.QTY_NEEDED <> WB.QTY_ISSUED + WB.QTY_RESERVED
            AND WB.ACTIVITY NOT IN ( 'Turn-In', 'Work Order', 'Repair')
            %s 
            ORDER BY P.PN,L.LOCATION_CODE"""%where_clause
        recs = selection_dir(query,cr)
        wo_parts = {}

        for row in recs:   
            wo_part = row[0] + '_' + row[1] + '_' + row[2] + '_' + str(row[3]) + '_' + str(row[4])
            if wo_part not in wo_parts:
                wo_parts[wo_part] = [[row[5],row[6],row[7],row[8],row[9]]]
            else:
                wo_parts[wo_part].append([row[5],row[6],row[7],row[8],row[9]])
        if wo_parts:    
            error = synch_stock_picking(session_id,wo_parts)
        else:
            error = 'No records found.'    
    return error,msg
    
def synch_stock_picking(session_id,wo_parts):
    from polls.models import WOStatus
    error = ''
    picking_data = []
    avail_data = []
    for key,value in wo_parts.items():
        parts = key.split('_')            
        picking_data.append(WOStatus(
            is_detail = False,
            session_id = session_id,
            si_number = key,
            wo_number = parts[0],
            part_number = parts[1],
            description = parts[2],
            qty_needed = parts[3],
            wob_auto_key = parts[4],
        ))
        for avail in value:
            avail_data.append(WOStatus(
                is_detail = True,
                session_id = session_id,
                si_number = key,           
                serial_number = avail[0],
                stock_line = avail[1],
                ctrl_number = avail[2],
                ctrl_id = avail[3],
                qty_available = avail[4],
            ))               
    if picking_data:     
        try:
            delete = WOStatus.objects.filter(session_id=session_id).delete()
            WOStatus.objects.bulk_create(picking_data) or []   
            WOStatus.objects.bulk_create(avail_data) or []             
        except Exception as exc:
            error += "Error, %s, with bulk creating stock picking rows."%(exc)              
    return error 

@shared_task
def so_dashboard(quapi_id,session_id,so_number,salesperson,due_date,location,exact_match=False,customer='',status=''):
    error,msg = '',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg        
    where_clause = ''   
    if customer:
        where_cust = exact_match and " AND COMPANY_NAME = '%s'"%customer or ''
        where_cust = not exact_match and " REGEXP_LIKE (COMPANY_NAME, '%s', 'i')"%customer or where_cust
        where_clause += " AND SOH.CMP_AUTO_KEY IN (SELECT CMP_AUTO_KEY FROM COMPANIES WHERE%s)"%where_cust 
    if location:
        where_loc = exact_match and " AND LOCATION_CODE = '%s'"%location or ''
        where_loc = not exact_match and " AND REGEXP_LIKE (LOC.LOCATION_CODE, '%s', 'i')"%location or where_loc
        where_clause += where_loc         
    if status and status != '0':
        where_clause += " AND SOH.SOS_AUTO_KEY = %s"%int(status)
    elif status and status == '0':
        pending = True    
    if salesperson:
        where_spn = exact_match and " AND SALESPERSON_CODE = '%s'"%salesperson or ''
        where_spn = not exact_match and "REGEXP_LIKE (SALESPERSON_CODE, '%s', 'i')"%salesperson or where_spn
        query = "SELECT SPN_AUTO_KEY FROM SALESPERSON WHERE %s"%where_spn
        spn = selection_dir(query,cr) 
        if not spn:
            msg+="\r\nYou have entered a salesperson that doesn't exist: '%s'"%salesperson
            return None
        spn_auto_key = spn and spn[0] and spn[0][0] or None 
        where_clause += " AND SOH.SPN_AUTO_KEY = %s"%spn_auto_key
    if due_date:
        where_clause += " AND SOH.DUE_DATE <= TO_DATE('%s', 'mm-dd-yyyy')"%due_date           
    if so_number:  
        exact_match = not (salesperson or due_date or status or location or customer)        
        where_so = exact_match and " AND SOH.SO_NUMBER = '%s'"%so_number or ''
        where_so = not exact_match and " AND REGEXP_LIKE (SOH.SO_NUMBER, '%s', 'i')"%so_number or where_so
        where_clause += where_so 
    query="""SELECT DISTINCT SOH.SOH_AUTO_KEY,SOH.SO_NUMBER,PNM.PN,PNM.DESCRIPTION, 
                                                                       
             SOH.DUE_DATE,SPN.SPN_AUTO_KEY,SPN.SALESPERSON_CODE,
             CMP.COMPANY_NAME,SOS.DESCRIPTION,SOS.SOS_AUTO_KEY,
             UDL.UDL_DESCRIPTION,LOC.LOCATION_CODE,
             WHS.WAREHOUSE_CODE,SMH.SM_NUMBER,ROH.RO_NUMBER
            FROM SO_STATUS SOS, SO_HEADER SOH, SM_HEADER SMH, 
                                                                    
            PARTS_MASTER PNM, STOCK_RESERVATIONS STR, SM_DETAIL SMD, 
            STOCK STM, SO_DETAIL SOD, USER_DEFINED_LOOKUPS UDL, 
            LOCATION LOC, WAREHOUSE WHS, RO_DETAIL ROD, RO_HEADER ROH,
            COMPANIES CMP, SALESPERSON SPN
            WHERE STM.PNM_AUTO_KEY = PNM.PNM_AUTO_KEY
            AND STM.LOC_AUTO_KEY = LOC.LOC_AUTO_KEY
            AND STM.WHS_AUTO_KEY = WHS.WHS_AUTO_KEY
            AND STM.IC_UDL_005 = UDL.UDL_AUTO_KEY (+)
            AND STM.STM_AUTO_KEY = STR.STM_AUTO_KEY
            AND STR.SMD_AUTO_KEY = SMD.SMD_AUTO_KEY
            AND SMD.SMH_AUTO_KEY = SMH.SMH_AUTO_KEY 
            AND SMD.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY (+)
            AND SOD.SOH_AUTO_KEY = SOH.SOH_AUTO_KEY (+)
            AND SOS.SOS_AUTO_KEY = SOH.SOS_AUTO_KEY
            AND SMD.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY (+)
            AND ROD.ROH_AUTO_KEY = ROH.ROH_AUTO_KEY (+)
            AND SOH.CMP_AUTO_KEY = CMP.CMP_AUTO_KEY (+)
            AND SOH.SPN_AUTO_KEY = SPN.SPN_AUTO_KEY (+)
                                                   
                                                       
                                                       
                                                       
                                                       
                                                      

            AND PNM.MASTER_FLAG = 'T'
            AND SOS.STATUS_TYPE = 'Open' 
            AND SOD.QTY_ORDERED <> SOD.QTY_INVOICED
            %s
            AND ROWNUM <= 5000
            ORDER BY SOH.DUE_DATE NULLS LAST, SMH.SM_NUMBER NULLS LAST         
            """%where_clause   
    """
        b.	Only show stock line that where smd_auto_key and sod_auto_key are not null reserved [checking]
        c.	If user enters a SO# then filter for only that SO. [checking]
        d.	If filtered by location, show all STM that have an STR for an SOD [checking]
    """   
    recs = selection_dir(query,cr) 
    if recs:    
        error = synch_new_sales(session_id,recs)
    else:
        error = 'No records found.'    
    return error,msg
    
def synch_new_sales(session_id,recs):
    from polls.models import Sale
    error = ''
    sale_data = []
    for row in recs:  
        priority = 0   
        due_date = row[4] or None         
        sale_data.append(Sale(
            session_id = session_id,
            soh_auto_key = row[0] or 0,#SOH.SOH_AUTO_KEY
            so_number = row[1],#SOH.SO_NUMBER
            part_number = row[2],#PNM.PN            
            description = row[3],#PNM.DESCRIPTION
            due_date = due_date or None,#SOH.DUE_DATE [4]
            spn_auto_key = row[5] or 0,#SPN.SPN_AUTO_KEY
            spn_code = row[6],#SPN.SALESPERSON_CODE
            customer = row[7],#CMP.COMPANY_NAME [7]
            priority = str(priority),#NEEDS TO BE CALCULATED
            so_status = row[8],#SOS.DESCRIPTION[8]
            sos_auto_key = row[9] or 0,#SOS.SOS_AUTO_KEY
            cart_code = row[10] or '',#UDL.UDL_DESCRIPTION
            loc_code = row[11] or '',#LOC.LOCATION_CODE
            whs_code = row[12] or '',#WHS.WAREHOUSE_CODE
            smd_number = row[13] or '',#,SMH.SM_NUMBER
            ro_number = row[14] or '',#ROH.RO_NUMBER
        ))
    if sale_data:     
        try:
            delete = Sale.objects.filter(session_id=session_id).delete()
            Sale.objects.bulk_create(sale_data) or []    
        except Exception as exc:
            error += "Error, %s, with bulk creating sales rows."%(exc)              
    return error 

#method to look up and make sure the licence is not expired yet.
@shared_task 
def check_exp_date(conn_string,quantum_cmp_key):
    error = ''
    cr,con = orcl_phone_home(conn_string)
    if not (cr and con):
        return 'Cannot connect to Oracle'
    query = "SELECT APPROVAL_EXPIRE FROM COMPANIES WHERE CMP_AUTO_KEY = %s"%quantum_cmp_key
    res = selection_dir(query,cr)
    exp_date = res and res[0] and res[0][0] or None
    right_now = datetime.now()
    #now = right_now.strftime('%Y-%m-%d %H:%M:%S')
    today = right_now.strftime('%Y-%m-%d %H:%M:%S')
    if exp_date and exp_date < today:
        error = 'License expired. Contact MRO Live account manager.'
    return error
	
def orcl_phone_home(conn_string):
	cr,con=None,None
	try:
		con = cx_Oracle.connect(conn_string)
	except Exception as exc:
		error, = exc.args
	cr = con and con.cursor() or None
	return cr,con

def get_grid_recs(cr,pnm_auto_key,session_id):
    error = ''
    query = """
        SELECT S.STOCK_LINE, S.PN, PNM.DESCRIPTION, 
        S.QTY_AVAILABLE, S.QTY_OH, L.LOCATION_CODE, 
        W.WAREHOUSE_CODE, S.SERIAL_NUMBER,
        S.CTRL_NUMBER, S.CTRL_ID, S.STM_AUTO_KEY,
        SR.WOB_AUTO_KEY,SR.WOO_AUTO_KEY
        FROM STOCK S 
        JOIN LOCATION L ON S.LOC_AUTO_KEY = L.LOC_AUTO_KEY 
        JOIN WAREHOUSE W ON S.WHS_AUTO_KEY = W.WHS_AUTO_KEY 
        JOIN PARTS_MASTER PNM ON S.PNM_AUTO_KEY = PNM.PNM_AUTO_KEY 
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
        WHERE 
        S.QTY_OH > 0
        AND (S.PNM_AUTO_KEY IN 
        (SELECT APM.ALT_PNM_AUTO_KEY FROM ALTERNATES_PARTS_MASTER APM WHERE APM.PNM_AUTO_KEY = '%s') 
          OR PNM.PNM_AUTO_KEY = '%s')         
        """%(pnm_auto_key,pnm_auto_key) 
    #wob.qty_needed cannot be > wob.qty_reserved + wob.qty_issued + qty user selects to reserve.
    recs = selection_dir(query,cr)
    from polls.models import WOStatus
    stock_data = []
    for row in recs:    
        stock_data.append(WOStatus(
        session_id = session_id,
        stock_line = row[0], 
        part_number = row[1],
        description = row[2],
        qty_available = row[3],
        quantity = row[4],
        location_code = row[5],
        wh_code = row[6],
        serial_number = row[7],
        ctrl_number = row[8] or '', 
        ctrl_id = row[9] or '', 
        stm_auto_key = row[10] or None,
        wob_auto_key = row[11] or None,
        woo_auto_key = row[12] or None,        
        ))
    WOStatus.objects.filter(session_id=session_id).delete()     
    if stock_data:           
        try:    
            WOStatus.objects.bulk_create(stock_data) or []    
        except Exception as exc:
            error += "Error with creating parts requests for display in grid: %s"%(exc)             
    return error
    
@shared_task
def parts_request(quapi_id,user_name,session_id,sysur_auto_key,\
    part_number,wo_task,quantity,notes,bom_status,\
    wo_task_upd='',yes_request = 'T',pnm_auto_key=None):
    error,msg,show_conf,field_changed,pn_info = '','','F','',[]
    today = datetime.now()
    date_now = today.strftime('%Y-%m-%d %H:%M:%S')  
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not (cr and con):
        return 'Cannot connect to Oracle','',msg,pn_info
    #parts-request - if we don't find a bom with the pnm/wot combo, then we create just like stock reserve.    
    #if we find the bom for the pnm/wot, then does qty_needed = qty_reserved + qty_issued, 
    #then insert a new BoM, if qty_needed > qty_reserved + qty_issued, 
    #then warn user with pop-up with pn that user entered (Is an Open Request. Add (y/n)).
    #if they hit yes, then create the new bom and if no, do nothing.
    #upon submit, clear all fields except Task.
    query = """SELECT BOS_AUTO_KEY FROM WO_BOM_STATUS 
        WHERE UPPER(BOM_STATUS_CODE) = UPPER('%s')"""%bom_status
    bos = selection_dir(query,cr)
    bos_auto_key = bos and bos[0] and bos[0][0] or ''
    
    part_number = part_number and part_number.strip() or ''
    where_clause = "UPPER(P.PN) = UPPER('%s')"%part_number
    if pnm_auto_key:
        where_clause = "P.PNM_AUTO_KEY = %s"%pnm_auto_key
        
    query = """SELECT P.PNM_AUTO_KEY,P.ACTIVE_PART,P.DESCRIPTION, 
        P.PN,M.MFG_CODE
        FROM PARTS_MASTER P
        LEFT JOIN MANUFACTURER M ON M.MFG_AUTO_KEY = P.MFG_AUTO_KEY
        WHERE %s AND ACTIVE_PART <> 'F'"""%where_clause
    pnms = selection_dir(query,cr)

    if not pnms:
        return '','','create-pn',pn_info
    
    if len(pnms) > 1 and yes_request != 'T':
        for pnm in pnms:
            
            pnm_auto_key = pnm[0]
            part_desc = pnm[2]
            part_num = pnm[3]
            mfg_code = pnm[4]
            pn_info.append([pnm_auto_key,part_num,part_desc,mfg_code])
            
        return '','','show_pns',pn_info
    
    else:    
        pnm = pnms and pnms[0]
        pnm_auto_key = pnm and pnm[0] or None
        unver_part = pnm and pnm[1] == 'T' or 'Unverified PN'
        description = pnm and pnm[2] or ''
        part_number = pnm and pnm[3] or ''
        
    error = get_grid_recs(cr,pnm_auto_key,session_id)
    
    query = """SELECT WOB.WOB_AUTO_KEY,WOB.QTY_NEEDED,
        WOB.QTY_RESERVED,WOB.QTY_ISSUED,PCC.CONDITION_CODE,
        WOB.ENTRY_DATE
        FROM WO_BOM WOB
        --JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = WOB.PNM_AUTO_KEY
        LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = WOB.PCC_AUTO_KEY
        WHERE
        WOB.PNM_AUTO_KEY = %s AND WOB.WOT_AUTO_KEY = %s 
        ORDER BY WOB.WOB_AUTO_KEY ASC"""%(pnm_auto_key,wo_task[:-1])
        
    wob = selection_dir(query,cr)
    wob_auto_key = wob and wob[0] and wob[0][0] or ''
    qty_needed = wob and wob[0] and wob[0][1] or 0
    qty_reserved = wob and wob[0] and wob[0][2] or 0
    qty_issued = wob and wob[0] and wob[0][3] or 0
    cond_code = wob and wob[0] and wob[0][4] or ''
    entry_date = wob and wob[0] and wob[0][5] or ''
    
    if qty_needed > qty_reserved + qty_issued and yes_request != 'T':
        return '','Confirm open request.','T',pn_info
        
    wot = None
    if wo_task[-1] in ['s','S','c','C'] and wo_task[:-1].isnumeric():
        query = """SELECT WOT.WOT_AUTO_KEY,WO.WOO_AUTO_KEY,WOT.SEQUENCE,
            WO.SI_NUMBER,WTM.DESCRIPTION FROM WO_TASK WOT
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
            LEFT JOIN WO_STATUS WTS ON WTS.WOS_AUTO_KEY = WOT.WOS_AUTO_KEY
            LEFT JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WOT.WTM_AUTO_KEY
            WHERE WOT.WOT_AUTO_KEY = %s
            AND (WTS.STATUS_TYPE NOT IN ('Closed','Cancel') OR WTS.STATUS_TYPE IS NULL)
            ORDER BY WOT.SEQUENCE ASC"""%wo_task[:-1]
        wot = selection_dir(query,cr)
        
    else:
        query = """SELECT WOT.WOT_AUTO_KEY,WO.WOO_AUTO_KEY,WOT.SEQUENCE,
            WO.SI_NUMBER,WTM.DESCRIPTION FROM WO_TASK WOT
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
            LEFT JOIN WO_STATUS WTS ON WTS.WOS_AUTO_KEY = WOT.WOS_AUTO_KEY
            LEFT JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WOT.WTM_AUTO_KEY
            WHERE UPPER(WO.SI_NUMBER) = UPPER('%s')
            AND (WTS.STATUS_TYPE NOT IN ('Closed','Cancel') OR WTS.STATUS_TYPE IS NULL)
            ORDER BY WOT.SEQUENCE ASC"""%wo_task
        wot = selection_dir(query,cr)
        
    wot_auto_key = wo_task_upd 
    if not wot_auto_key:    
        wot_auto_key = wot and wot[0] and wot[0][0] or None
        wot_sequence = wot and wot[0] and wot[0][2] or ''
        wtm_desc = wot and wot[0] and wot[0][4] or ''
        if not wot_auto_key:
            return 'No task found.','','F',pn_info
            
    else:
        query = """SELECT WOT.SEQUENCE,WTM.DESCRIPTION 
            FROM WO_TASK WOT
            LEFT JOIN WO_TASK_MASTER WTM 
            ON WTM.WTM_AUTO_KEY=WOT.WTM_AUTO_KEY
            WHERE WOT.WOT_AUTO_KEY = %s"""%wot_auto_key
        wtm = selection_dir(query,cr)
        wtm = wtm and wtm[0] or None
        wot_sequence = wtm and wtm[0] or ''
        wtm_desc = wtm and wtm[1] or ''
        
    woo_auto_key = wot and wot[0] and wot[0][1] or None
    if not woo_auto_key:
        return 'No workorder found.','','F',pn_info
        
    si_number = wot and wot[0] and wot[0][3] or ''    
    qty_summed = qty_needed == qty_reserved + qty_issued or False
    
    query = """SELECT WOB.WOB_AUTO_KEY
        FROM WO_BOM WOB
        WHERE 
        WOB.PNM_AUTO_KEY = %s AND WOB.WOO_AUTO_KEY = %s 
        ORDER BY WOB.WOB_AUTO_KEY ASC
        """%(pnm_auto_key,woo_auto_key) 
        
    rec = selection_dir(query,cr)
    wob = rec and rec[0] and rec[0][0] or None
    add_req = 'Add request?  PN %s already found for WO %s '%(part_number,si_number)
    if wob and yes_request != 'T':
        return '',add_req,'pn_found',pn_info    

    
    if not wob_auto_key or (wob_auto_key and (qty_summed or yes_request)):
        today = datetime.now()
        timestamp = today.strftime('%Y-%m-%d %H:%M:%S')
        #need to join on part_condition_codes to get the cond_level
        cond_sub = """SELECT COND_LEVEL,PCC_AUTO_KEY,
            CONDITION_CODE FROM PART_CONDITION_CODES
            WHERE PCC_AUTO_KEY = (SELECT PCC_MAIN_OUT FROM WO_CONTROL)"""
        pcc_data = selection_dir(cond_sub,cr)
        pcc_auto_key = pcc_data and pcc_data[0] and pcc_data[0][1] or ''
        cond_level = pcc_data and pcc_data[0] and pcc_data[0][0] or 0
        cond_code = pcc_data and pcc_data[0] and pcc_data[0][2] or ''
        notes = notes.replace("'",'')
        notes = notes.replace('"','')
        q_wob = """INSERT INTO WO_BOM (BOS_AUTO_KEY,REQUISITION,WOO_AUTO_KEY,SYSUR_AUTO_KEY,WOT_AUTO_KEY,PNM_AUTO_KEY,
            QTY_NEEDED,ACTIVITY,PCC_AUTO_KEY,COND_LEVEL,ENTRY_DATE,NOTES) 
            VALUES('%s','T','%s','%s','%s','%s','%s','%s','%s','%s',
            TO_TIMESTAMP('%s', 'yyyy-mm-dd hh24:mi:ss'),TO_CLOB('%s'))"""%(bos_auto_key,woo_auto_key,sysur_auto_key,wot_auto_key,pnm_auto_key,quantity,'Consumable',pcc_auto_key,cond_level,timestamp,notes)
        error = insertion_dir(q_wob,cr)
        query = "SELECT WOB_AUTO_KEY FROM WO_BOM WHERE PNM_AUTO_KEY = %s ORDER BY WOB_AUTO_KEY DESC"%pnm_auto_key
        wob = selection_dir(query,cr)  
        wob_auto_key = wob and wob[0] and wob[0][0] or None
        #when the user selects the line in the grid, run the reservation of the BoM, etc.        
        if not error:
            orcl_commit(con=con)
            aud_status = 'success'
            msg = 'Successful request.'
        else: 
            aud_status = 'failure' 
            msg = 'Failed request.'
        from polls.models import MLApps as maps,QuantumUser as qu
        app_id = maps.objects.filter(code='parts-request')
        user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
        user_rec = user_rec and user_rec[0] or None

        if user_rec:
            new_val = 'wob_auto_key: ' + str(wob_auto_key)
            field_changed = 'PN: ' + str(part_number)
            field_changed += ',quantity: ' + str(quantity)
            field_changed += ',task: ' + str(wtm_desc)
            field_changed += ',sequence: ' + str(wot_sequence)
            field_changed += ',WO#: ' + str(si_number)
            field_changed += error
            parameters = [unver_part,str(si_number),]
            parameters += [str(entry_date or timestamp)]
            parameters += [str(wot_sequence),str(wtm_desc)]
            parameters += [str(part_number),str(description),str(cond_code),'']
            parameters += [str(quantity),]          
            parameters += ['','','','','','','','']
            right_now = datetime.now()
            now = right_now.strftime('%Y-%m-%d %H:%M:%S')
            error += register_audit_trail(
                user_rec,
                field_changed,
                new_val,
                right_now,
                app_id,
                quapi,
                status=aud_status,
                reg_events=['Parts Request'],
                parameters=parameters,
                sysur_auto_key=sysur_auto_key,
            )            
            return error,msg,'F',pn_info
            
        else:
            error = 'Incorrect Quantum User ID.'
            show_conf = 'F'
            
    return error,field_changed,show_conf,pn_info
    
@shared_task
def tools_checkin(quapi_id,session_id,sysur_auto_key,ctrl_number,ctrl_id,kiosk_auto_key):
    error,msg = '',''
    today = datetime.now()
    date_now = today.strftime('%Y-%m-%d %H:%M:%S')  
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    """TOOL CHECK IN

        1. UPDATE WO_TASK_TOOLS

        WWT_AUTO_KEY = LOOK-UP WWT_BASED ON STM AND STOCK_RESERVATIONS,
        DATE_CHECK_IN = '(SYSTIME/DATE)',
        SYSUR_IN = '(SYSUR_AUTO_KEY FOR USER LOGGED IN)'
    """
    wtt_sub = """SELECT WTT_AUTO_KEY,STR_AUTO_KEY FROM STOCK_RESERVATIONS 
        WHERE STM_AUTO_KEY IN (SELECT STM_AUTO_KEY FROM STOCK WHERE CTRL_NUMBER = %s AND CTRL_ID = %s)
        AND WTT_AUTO_KEY IS NOT NULL
	    """%(ctrl_number,ctrl_id)
    wtt = selection_dir(wtt_sub,cr)
    wtt_auto_key = wtt and wtt[0] and wtt[0][0] or None
    str_auto_key = wtt and wtt[0] and wtt[0][1] or None
    if str_auto_key:
        qury = """DELETE FROM STOCK_RESERVATIONS WHERE STR_AUTO_KEY=%s"""%str_auto_key
        error = updation_dir(qury,cr)
    #else:
        #return 'Not reserved.',''       
    if wtt_auto_key:
        query = """UPDATE WO_TASK_TOOLS SET DATE_CHECK_IN = TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss'),
           SYSUR_IN = %s, QTY_RESERVED = 0, QTY_CHECKED_OUT=1, QTY_CHECKED_IN=1 WHERE WTT_AUTO_KEY = %s"""%(date_now,kiosk_auto_key,wtt_auto_key)
        error = updation_dir(query,cr)
    else:
        return 'Invalid tool.',''
    if not error or error == '{"recs": ""}':
        orcl_commit(con=con)
        aud_status = 'success'
        error = ''
        msg = 'Successful check in.'
    else: 
        aud_status = 'failure'       
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='tools')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        rec_input = (ctrl_id and 'Record with Ctrl#: '+ str(ctrl_number) + ' and Ctrl ID#: ' + str(ctrl_id)) or ''
        new_val = rec_input
        field_changed = "Tool checked in for: " + ctrl_number + ctrl_id
        field_changed += error
        error += register_audit_trail(user_rec,field_changed,new_val,date_now,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,msg
    
@shared_task
def tools_checkout(quapi_id,session_id,sysur_auto_key,ctrl_number,ctrl_id,wo_task,kiosk_auto_key):
    error,msg = '',''
    today = datetime.now()
    date_now = today.strftime('%Y-%m-%d %H:%M:%S')
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    """1. INSERT INTO STOCK_RESERVATIONS
    STM_AUTO_KEY, STR_AUTO_KEY ='(NEXTGEN VAL)',
    QTY_RESERVED = '1', 
    WWT_AUTO_KEY  '(GENERATED DURING STEP 1)', 
    SYSUR_AUTO_KEY = '(USER LOGGED INTO APP)', 
    ENTRY_DATE = '(SYSTIME/DATE)',
    SYSUR_TOOL_ISSUED = '(SYSUR_AUTO_KEY FORM SYS_USERS WHERE USER_ID = 'EMP FIELD',
    RESERVATION_ID = '(NEXTGEN VAL)' """    
    #sysur_tool_sub = "SELECT SYSUR_AUTO_KEY FROM SYS_USERS WHERE SYSUR_AUTO_KEY = %s"%sysur_auto_key
    #stm_sub = "SELECT STM_AUTO_KEY FROM STOCK WHERE CTRL_NUMBER = %s AND CTRL_ID = %s"%(ctrl_number,ctrl_id)
    
    query = """select s.stm_auto_key,wo.si_number,sys.user_name,sr.stm_auto_key,
        wo.woo_auto_key
        from stock s
        left join stock_reservations sr on sr.stm_auto_key = s.stm_auto_key
        left join wo_operation wo on wo.woo_auto_key = sr.woo_auto_key
        left join sys_users sys on sys.sysur_auto_key = sr.sysur_auto_key
        where s.ctrl_number = %s and s.ctrl_id = %s"""%(ctrl_number,ctrl_id)
        
    stm = selection_dir(query,cr)
    stm = stm and stm[0] or []
    stm_auto_key = stm and stm[0]
    wo_number = stm and stm[1]
    sys_user = stm and stm[2]
    str_stm = stm and stm[3]
    
    if wo_task[-1] not in ['s','S','c','C']:
        query = """SELECT WOO_AUTO_KEY FROM WO_OPERATION
            WHERE SI_NUMBER = '%s'
            """%wo_task
        woo = selection_dir(query,cr)
        woo_auto_key = woo and woo[0] and woo[0][0] or None
    
    if str_stm:
        return 'Tool checked out to: %s on WO: %s'%(sys_user,wo_number),''
        
    if stm_auto_key:
        if not woo_auto_key:
            query = """INSERT INTO STOCK_RESERVATIONS 
                (STR_AUTO_KEY,STM_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY,ENTRY_DATE,SYSUR_TOOL_ISSUED,RESERVATION_ID)    
                VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s',%s,%s,TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss'),%s,G_STR_AUTO_KEY.NEXTVAL)
                """%(stm_auto_key,1,sysur_auto_key,date_now,sysur_auto_key)
            error = insertion_dir(query,cr)
        else:
            query = """INSERT INTO STOCK_RESERVATIONS 
                (STR_AUTO_KEY,STM_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY,ENTRY_DATE,SYSUR_TOOL_ISSUED,RESERVATION_ID,WOO_AUTO_KEY)    
                VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s',%s,%s,TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss'),%s,G_STR_AUTO_KEY.NEXTVAL,%s)
                """%(stm_auto_key,1,sysur_auto_key,date_now,sysur_auto_key,woo_auto_key)
            error = insertion_dir(query,cr)   
        if error:
            return error,''
    else:
        return 'Invalid tool.',''
    """2. INSERT INTO WO_TASK_TOOLS
]
    WWT_AUTO_KEY = '(NEXTGEN VAL)',
    WOT_AUTO_KEY = '(USER INPUT)' - 'S',
    PNM_AUTO_KEY = '(PNM_AUTO_KEY FROM STOCK TABLE FROM USER ENTERED CTRL#/ID COMBO)',
    QTY_NEEDED = '1',
    QTY_RESERVED = (THIS SHOULD BE LEFT ALONE AS I THINK IT WILL BE POPULATED VIA STEP 2 JUST LIKE STOCK ALLOCATION RESEVER TO THE WO_BOM['QTY_RESERVED']),
    QTY_CHECK_OUT = '0',
    DATE_CHECK_OUT = '(SYSTIME/DATE)'
    SYSUR_OUT = '(SYSUR_AUTO_KEY FORM SYS_USERS WHERE USER_ID = 'EMP FIELD'"""
   
    pnm_sub = "SELECT PNM_AUTO_KEY FROM STOCK WHERE CTRL_NUMBER = %s AND CTRL_ID = %s"%(ctrl_number,ctrl_id) 
    pnm = selection_dir(pnm_sub,cr)
    pnm_auto_key = pnm and pnm[0] and pnm[0][0] or None

    if pnm_auto_key:
        if wo_task[-1] in ['s','S','c','C']: 
            wot_auto_key = wo_task[:-1]        
        else:
            query = """SELECT WT.WOT_AUTO_KEY FROM WO_TASK WT WHERE  
                        WT.WOO_AUTO_KEY = (SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE UPPER('%s') = UPPER(SI_NUMBER))                
                        AND (WT.WOS_AUTO_KEY NOT IN 
                       (SELECT WOS_AUTO_KEY FROM WO_STATUS WHERE STATUS_TYPE IN ('Closed','Cancel'))
                       OR WT.WOS_AUTO_KEY IS NULL)
                       ORDER BY WT.WOT_AUTO_KEY
                    """%wo_task
            wot = selection_dir(query,cr)
            wot_auto_key = wot and wot[0] and wot[0][0] or None
        query = """INSERT INTO WO_TASK_TOOLS 
            (WTT_AUTO_KEY,WOT_AUTO_KEY,PNM_AUTO_KEY,QTY_NEEDED,QTY_RESERVED,QTY_CHECKED_OUT,QTY_CHECKED_IN,DATE_CHECK_OUT,SYSUR_OUT) 
            VALUES(G_WTT_AUTO_KEY.NEXTVAL,%s,'%s',%s,%s,%s,%s,TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss'),%s)
            """%(wot_auto_key,pnm_auto_key,1,1,1,0,date_now,kiosk_auto_key)        
        error = insertion_dir(query,cr)
        query = "SELECT WTT_AUTO_KEY FROM WO_TASK_TOOLS ORDER BY WTT_AUTO_KEY DESC"
        wtt = selection_dir(query,cr)
        wtt_auto_key = wtt and wtt[0] and wtt[0][0] or None
        """2. UPDATE RESERVATION"""
        query = """
        SELECT STR_AUTO_KEY FROM STOCK_RESERVATIONS
        WHERE ROWNUM <= 1 AND STM_AUTO_KEY = %s        
        ORDER BY STR_AUTO_KEY DESC
        """%stm_auto_key
        str_auto=selection_dir(query,cr)
        str_auto_key = str_auto and str_auto[0] and str_auto[0][0] or None
        if str_auto_key:
            query = """UPDATE STOCK_RESERVATIONS 
            SET WTT_AUTO_KEY = %s WHERE STR_AUTO_KEY = 
            (%s)"""%(wtt_auto_key,str_auto_key)
            error = updation_dir(query,cr)
    else:
        return 'Invalid tool.',''
    if not error or error == '{"recs": ""}':
        error = ''
        orcl_commit(con=con)
        aud_status = 'success'
        msg = "Successful check out."
    else: 
        aud_status = 'failure'       
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='tools')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        rec_input = (ctrl_id and 'Record with Ctrl#: '+ str(ctrl_number) + ' and Ctrl ID#: ' + str(ctrl_id)) or ''
        new_val = rec_input
        field_changed = "Tool checked out for: " + ctrl_number + ctrl_id + ', for task: ' + wo_task + ' '
        error += register_audit_trail(user_rec,field_changed,new_val,date_now,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'       
    return error,msg    
    
def get_cursor_con(quapi):
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    server_url = orcl_conn.url
    cr,con = None,None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    return cr,con
    
def get_res_issue(cr,ctrl_number,ctrl_id,wo_number,quantity):
    #            
    if wo_number[-1] in ['s','S']:
        query = """SELECT SR.STR_AUTO_KEY, STI.STI_AUTO_KEY, STI.QTY,
            S.STM_ORIGINAL, STI.WOT_AUTO_KEY, STI.WOB_AUTO_KEY, S.STM_AUTO_KEY,
            CASE WHEN WT.WOO_AUTO_KEY IS NULL THEN WOB.WOO_AUTO_KEY ELSE CASE WHEN 
            WT.WOO_AUTO_KEY IS NOT NULL THEN WT.WOO_AUTO_KEY ELSE SR.WOO_AUTO_KEY END END,
            S.PNM_AUTO_KEY,S.PCC_AUTO_KEY,S.SERIAL_NUMBER,SI.STM_AUTO_KEY
            FROM STOCK S
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
            LEFT JOIN STOCK_TI STI ON STI.STM_AUTO_KEY = S.STM_AUTO_KEY
            LEFT JOIN STOCK_TI SI ON SI.STI_ORIG = STI.STI_ORIG
            LEFT JOIN WO_TASK WT ON WT.WOT_AUTO_KEY = STI.WOT_AUTO_KEY 
            LEFT JOIN WO_BOM WOB ON WOB.WOT_AUTO_KEY = WT.WOT_AUTO_KEY             
            WHERE S.CTRL_NUMBER = %s 
            --AND SR.QTY_ISSUED <= %s
            AND S.CTRL_ID = %s
            AND STI.WOT_AUTO_KEY = %s
            ORDER BY STI.STI_AUTO_KEY DESC, SR.STR_AUTO_KEY DESC"""%(quantity,ctrl_number,ctrl_id,wo_number[:-1])
    else:
        query = """SELECT SR.STR_AUTO_KEY, STI.STI_AUTO_KEY, STI.QTY,
            S.STM_ORIGINAL, WOT.WOT_AUTO_KEY, WOB.WOB_AUTO_KEY, S.STM_AUTO_KEY,
            WO.WOO_AUTO_KEY,S.PNM_AUTO_KEY,S.PCC_AUTO_KEY,S.SERIAL_NUMBER,
            SI.STM_AUTO_KEY
            FROM STOCK S
            LEFT JOIN STOCK_TI STI ON STI.STM_AUTO_KEY = S.STM_AUTO_KEY
            LEFT JOIN STOCK_TI SI ON SI.STI_ORIG = STI.STI_ORIG
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY            
            LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = STI.WOB_AUTO_KEY        
            LEFT JOIN WO_TASK WOT ON WOT.WOT_AUTO_KEY = WOB.WOT_AUTO_KEY
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
            WHERE STI.QTY > 0 
            --AND SR.QTY_ISSUED <= %s
            AND S.CTRL_NUMBER = %s 
            AND S.CTRL_ID = %s            
            AND UPPER(WO.SI_NUMBER) = UPPER('%s')
            ORDER BY STI.STI_AUTO_KEY DESC NULLS LAST"""%(quantity,ctrl_number,ctrl_id,wo_number)
  
    stm = selection_dir(query,cr)
    return stm
    
    #UNISSUE MODE - change WO# to task.  User enters task and stock line and we return the issued qty just as we return reserved qty 
    #under issue mode.  return from assembly - no longer issued nor reserved.

def stock_qty_update(cr,quantity,stm_auto_key):
    query = """
        CREATE OR REPLACE PROCEDURE "WOB_TI_UPDATE"
        (QTY IN NUMBER,STM IN NUMBER)  AS
        V_CURSOR QC_SC_PKG.CURSOR_TYPE ;
        BEGIN                
            begin
            qc_trig_pkg.disable_triggers;
            UPDATE STOCK SET QTY_OH = QTY_OH + QTY, QTY_AVAILABLE = QTY_AVAILABLE + QTY WHERE STM_AUTO_KEY = STM;
            qc_trig_pkg.enable_triggers;
            end;
         END WOB_TI_UPDATE;"""         
    error = updation_dir(query,cr)    
    run_proc = """
        BEGIN
        WOB_TI_UPDATE(%s,%s);
        END;   
    """%(wob_auto_key,sti_auto_key)
    error = updation_dir(run_proc,cr)
    return error

    
def wob_ti_update(cr,wob_auto_key,sti_auto_key):
    query = """
        CREATE OR REPLACE PROCEDURE "WOB_TI_UPDATE"
        (WOB IN NUMBER,STI IN NUMBER)  AS
        V_CURSOR QC_SC_PKG.CURSOR_TYPE ;
        BEGIN                
            begin
            qc_trig_pkg.disable_triggers;
            UPDATE STOCK_TI SET WOB_AUTO_KEY=WOB,STI_ORIG=STI WHERE STI_AUTO_KEY = (SELECT STI_AUTO_KEY FROM STOCK_TI ORDER BY STI_AUTO_KEY DESC FETCH NEXT 1 ROWS ONLY);
            qc_trig_pkg.enable_triggers;
            end;
         END WOB_TI_UPDATE;"""         
    error = updation_dir(query,cr)   
    run_proc = """
        BEGIN
        WOB_TI_UPDATE(%s,%s);
        END;   
    """%(wob_auto_key,sti_auto_key)
    error = updation_dir(run_proc,cr)
    return error

@shared_task
def stock_unissue(quapi_id,session_id,sysur_auto_key,user_name,quantity,wo_number,ctrl_number,ctrl_id):

    error,msg,sti_orig,stm_auto_key = '','','',''

    from polls.models import QueryApi   

    quapi = QueryApi.objects.filter(id=quapi_id)

    quapi = quapi and quapi[0] or None

    cr,con = get_cursor_con(quapi)

    if not (cr and con):

        return 'Cannot connect to Oracle',msg,'' 

    right_now = datetime.now()

    entry_date = right_now.strftime('%m-%d-%Y %H:%M:%S')

    stm = get_res_issue(cr,ctrl_number,ctrl_id,wo_number,quantity)

    qty_issued = ''
    

    if not stm:

        return 'Stock not found.',msg,qty_issued

    qty_user = quantity and float(quantity) or 0

    if quantity and qty_user <= 0:

        return 'Qty must be greater than 0.',msg,qty_issued      

    for iss in stm:

        sti_auto_key = iss and iss[1] or ''

        qty_issued = iss and iss[2] or 0

        if sti_auto_key:  

            str_auto_key = iss and iss[0] or ''

            sti_auto_key = iss and iss[1] or ''           

            stm_original = iss and iss[3] or ''

            wot_auto_key = iss and iss[4] or ''

            wob_auto_key = iss and iss[5] or ''

            stm_auto_key = iss and iss[6] or ''

            woo_auto_key = iss and iss[7] or ''

            pnm_auto_key = iss and iss[8] or ''

            pcc_auto_key = iss and iss[9] or ''

            serial_number = iss and iss[10] or ''

            break

    if not stm_auto_key:

        return 'Stock not found.',msg,qty_issued

    if qty_user:

        if qty_issued > 0 and qty_issued < qty_user:

            error = 'Qty returned cannot be more than qty issued.'

            return error,msg,qty_issued

        query = "select gl_trans_link+1 from stock_ti order by gl_trans_link desc nulls last"

        gl_trans = selection_dir(query,cr)

        gl_trans_link = gl_trans and gl_trans[0] and gl_trans[0][0] or 0

        neg_qty = (-1)*qty_user

        #query = """INSERT INTO STOCK_TI

       #    (STM_AUTO_KEY,STM_PARENT,QTY,WOT_AUTO_KEY,WOB_AUTO_KEY,ENTRY_DATE,TRAN_DATE,TI_TYPE,SYSUR_AUTO_KEY,GL_TRANS_LINK,STI_ORIG)

        #    VALUES(%s,'%s',%s,%s,%s,TO_TIMESTAMP('%s','MM-DD-YYYY hh24:mi:ss'),

        #    TO_TIMESTAMP('%s','MM-DD-YYYY #hh24:mi:ss'),'I',%s,'%s','%s')"""%(stm_auto_key,stm_auto_key,neg_qty,wot_auto_key,wob_auto_key,entry_date,entry_date,sysur_auto_key,gl_trans_link,sti_auto_key)#           

        #error = insertion_dir(query,cr)

        if not sti_auto_key:

            return 'Stock line not issued to this work order.',msg,''

        query = """

            DECLARE

            CT qc_utl_pkg.cursor_type;

            P_STI NUMBER;

            P_QTY NUMBER;

            BEGIN

            CT:=qc_ic_pkg2.spi_issue_reverse(%s,%s);

            END;

        """%(sti_auto_key,quantity)       

        error = updation_dir(query,cr)

        if error not in ['{"recs": ""}','']:

            return error,msg,''

        #error = stock_turn_in(cr,wob_auto_key,woo_auto_key,pnm_auto_key,serial_number,qty_user,pcc_auto_key,stm_parent=stm_auto_key)

        #1. Disable triggers

        #2. Update the newly created STI record to include the wob_auto_key

        #query = """DELETE FROM STOCK_RESERVATIONS

        #    WHERE STI_AUTO_KEY = (SELECT STI_AUTO_KEY FROM STOCK_TI

        #    ORDER BY STI_AUTO_KEY DESC FETCH NEXT 1 ROWS ONLY)"""

        #error = updation_dir(query,cr)

        #if error != '{"recs": ""}':

        #    return error,msg,''

        #QTY_RESERVED = 0,

        #if error != '{"recs": ""}':

        #    return error,msg,''

        #error = wob_ti_update(cr,wob_auto_key,sti_auto_key)

        #error = wob_ti_update(cr,wob_auto_key,sti_auto_key)

        #if error != '{"recs": ""}':

        #    return error,msg,''

        #query = """UPDATE STOCK SET QTY_OH = QTY_OH + %s WHERE STM_AUTO_KEY = %s

        #    """%(quantity,stm_auto_key)

        #error = updation_dir(query,cr)           

        #update the audit trails with the correct users

        squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY =

            (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_name[:9],stm_auto_key)

        error = updation_dir(squery,cr)

        if error != '{"recs": ""}':

            return error,msg,''

 

        squery = """UPDATE AUDIT_TRAIL

            SET SYSUR_AUTO_KEY=%s

            WHERE

            SYSUR_AUTO_KEY = (select sysur_auto_key from sys_users where user_name='SYSDBA')

            AND SOURCE_TABLE IN ('STM','QUANTUM')

            AND STAMPTIME >= SYSDATE - 1"""%sysur_auto_key 

        error = updation_dir(squery,cr)

            

        if error != '{"recs": ""}':

            return error,msg,''

 

        query = """SELECT STI_AUTO_KEY FROM STOCK_TI WHERE

            STI_AUTO_KEY = (SELECT MAX(STI_AUTO_KEY) FROM STOCK_TI WHERE STM_AUTO_KEY = %s)

            """%(stm_auto_key)

        new_sti = selection_dir(query,cr)

        new_sti_key = new_sti and new_sti[0] and new_sti[0][0] or ''

       

        query = """

            begin

            qc_trig_pkg.disable_triggers;

            update stock_ti set sysur_auto_key = %s where sti_auto_key = %s;

            qc_trig_pkg.enable_triggers;

            end;

            """%(sysur_auto_key,new_sti_key)

        error = updation_dir(query,cr)

       

        if error != '{"recs": ""}':

            return error,msg,''

                   

    elif qty_issued > 0:

        #push the qty_issued back to the ui

        return error,msg,qty_issued

    elif not qty_issued or qty_issued == 0:

        return 'No stock issued.',msg,''      

    if not error or error == '{"recs": ""}':

        orcl_commit(con=con)

        aud_status = 'success'

        msg = 'Stock unissued.'

        error = ''

    else:

        aud_status = 'failure'

        msg = ''    

    from polls.models import MLApps as maps,QuantumUser as qu

    app_id = maps.objects.filter(code='stock-allocation')  

    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)

    user_rec = user_rec and user_rec[0] or None

    if user_rec:

        rec_input = ''

        if ctrl_id and ctrl_number:

            rec_input = (ctrl_id and 'Record with Ctrl#: '+ str(ctrl_number) + ' and Ctrl ID#: ' + str(ctrl_id)) or ''

        new_val = rec_input + ' Stock unissued'

        field_changed = "Stock unissued for: " + str(ctrl_number + ctrl_id) + ', with quantity: ' + str(quantity)

        field_changed += error

        now = right_now.strftime('%Y-%m-%d %H:%M:%S')

        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)

    else:

        error = 'Incorrect Quantum User ID.'        

    return error,msg,''

@shared_task
def issue_consumables(quapi_id,session_id,\
    sysur_auto_key,user_id,ctrl_number,\
    ctrl_id,wo_task,quantity,active_mode,\
    must_reserve):
    wob_auto_key,wot_auto_key,str_auto_key = '','',''
    error,msg,qty_res,qty_reserved,part_number,wot_sequence,si_number,new_wob_key = '','','','','','','',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg,'',''    
    #User enters task, enters stock line and then the app should populate
    #    the qty field with the qty_reserved from the str related to the WOB.  
    
    #ISSUE MODE: Check to see if the task is reserved via the wob_auto_key on STOCK_RESERVATIONS.     
    #Populate the quantity_reserved from the reservation in the quantity field on front-end.    
    #If not, we must reserve it with the user's quantity.
    #If the user enters a different qty than what we populate with, create a new bom with the full user-entered qty
    # and then issue for the qty reserved.    

    #we have to consider the qty_needed on the BoM
    #we find a reservation for a wob, then we just issue it.
    #if we don't find a reservation for that wob, we go through the reservation process and then issue what we just reserved.
    right_now = datetime.now()
    entry_date = right_now.strftime('%m-%d-%Y %H:%M:%S') 
    #if not must_reserve or must_reserve=='F':       
    #qty_reserved becomes the qty issued. 
   
    if not wo_task:
        return 'Must enter WO# or task.','',qty_res,must_reserve      
    if quantity:
        #ISSUE MODE: Reduce qty_oh, qty_available by qty_reserved from the reservation. 
        """
        
        If the STM that is entered by user is reserved to another 
        WOB then you get this error in consumable app. 
        this STM has QTY_RES to another WOO and WOB, 
        but QTY_AVAILABLE for that STM is > QTY user enters. 
        this should result in STR insertion for new WOB 
        under WOO entered and QTY user enters = QTY_RESERVED
        """
               
        if wo_task and ctrl_number:
            stm,error = get_stock_res(cr,wo_task,ctrl_number,ctrl_id)
            if not stm:
                return error,'',qty_res,must_reserve 
            stm_auto_key = stm and stm[0] and stm[0][0] or ''
            pnm_auto_key = stm and stm[0] and stm[0][2] or ''
            #part_number = stm and stm[0] and stm[0][4] or ''
            qty_oh = stm and stm[0] and stm[0][5] or 0
            res_found = stm and isinstance(stm[0],list) and len(stm[0]) > 6
            if res_found:
                return 'Reservation found. Use Stock Issue app.','',qty_res,must_reserve         
            else:
                #if wo_task[-1] in ['s','S']:
                    #query = """SELECT WOB_AUTO_KEY,QTY_NEEDED,
                    #QTY_RESERVED,QTY_ISSUED,WOT_AUTO_KEY,WOO_AUTO_KEY
                    #    WOO_AUTO_KEY FROM WO_BOM WHERE 
                    #    PNM_AUTO_KEY = %s AND 
                    #    WOT_AUTO_KEY = %s 
                    #    ORDER BY WOB_AUTO_KEY ASC"""%(pnm_auto_key,wo_task[:-1])
                #else:
                    #query = """SELECT WOB_AUTO_KEY,QTY_NEEDED,
                    #    QTY_RESERVED,QTY_ISSUED,WOT_AUTO_KEY,
                    #    WOO_AUTO_KEY FROM WO_BOM WHERE 
                    #    PNM_AUTO_KEY = %s AND WOO_AUTO_KEY = 
                    #    (SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER = '%s')
                    #    ORDER BY WOB_AUTO_KEY ASC"""%(pnm_auto_key,wo_task)                
                #wob = selection_dir(query,cr)
                wob_auto_key = None 
                #qty_needed = wob and wob[0] and wob[0][1] or 0
                #qty_reserved = wob and wob[0] and wob[0][2] or 0
                #qty_issued = wob and wob[0] and wob[0][3] or 0
                #wot_auto_key = wob and wob[0] and wob[0][4] or None
                #woo_auto_key = None
                if wob_auto_key:
                    """if qty_needed >= qty_reserved + qty_issued + float(quantity):
                        #Create a stock reservation for the wob
                        squery = """"""INSERT INTO STOCK_RESERVATIONS 
                        (STR_AUTO_KEY,STM_AUTO_KEY,WOB_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY) 
                        VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s)""""""%(stm_auto_key,wob_auto_key,quantity,sysur_auto_key)
                        error = insertion_dir(squery,cr)
                        if not error:
                            squery = """"""UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                                EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
                                (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG
                                WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')""""""%(sysur_auto_key,user_id[:9],stm_auto_key)
                            error = updation_dir(squery,cr)"""
                         
                    if woo_auto_key and wot_auto_key:
                        #no open need: qty_needed < qty_reserved + qty_issued + float(quantity)
                        #add WOB for that PN if that PN is not an "open need" & then reserve it.
                        today = datetime.now()
                        timestamp = today.strftime('%Y-%m-%d %H:%M:%S')
                        #need to join on part_condition_codes to get the cond_level        
                        cond_sub = "SELECT COND_LEVEL,PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE CONDITION_CODE = 'AR'"
                        pcc_data = selection_dir(cond_sub,cr)
                        pcc_auto_key = pcc_data and pcc_data[0] and pcc_data[0][1] or 1
                        cond_level = pcc_data and pcc_data[0] and pcc_data[0][0] or 0
                        
                        q_wob = """INSERT INTO WO_BOM (REQUISITION,WOO_AUTO_KEY,SYSUR_AUTO_KEY,WOT_AUTO_KEY,PNM_AUTO_KEY,
                            QTY_NEEDED,ACTIVITY,PCC_AUTO_KEY,COND_LEVEL,ENTRY_DATE) 
                            VALUES('T','%s','%s','%s','%s','%s','%s','%s','%s',
                            TO_TIMESTAMP('%s', 'yyyy-mm-dd hh24:mi:ss'))
                            """%(woo_auto_key,sysur_auto_key,wot_auto_key,pnm_auto_key,quantity,'Consumable',pcc_auto_key,cond_level,timestamp)
                        error = insertion_dir(q_wob,cr)
                        if not error:
                            #get new wob_auto_key
                            query = """SELECT WOB_AUTO_KEY FROM WO_BOM ORDER BY WOB_AUTO_KEY DESC FETCH NEXT 1 ROWS ONLY
                            """
                            wob = selection_dir(query,cr)
                            new_wob_key = wob and wob[0] and wob[0][0] or ''
                            #Create a stock reservation for the wob
                            squery = """INSERT INTO STOCK_RESERVATIONS 
                            (STR_AUTO_KEY,STM_AUTO_KEY,WOB_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY) 
                            VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s)"""%(stm_auto_key,new_wob_key,quantity,sysur_auto_key)
                            error = insertion_dir(squery,cr)
                            query = """
                                SELECT STR_AUTO_KEY FROM STOCK_RESERVATIONS
                                WHERE STM_AUTO_KEY = %s        
                                ORDER BY STR_AUTO_KEY DESC
                                """%stm_auto_key
                            str_auto=selection_dir(query,cr)
                            str_auto_key = str_auto and str_auto[0] and str_auto[0][0] or None
                else:
                    #no existing wob = no open need
                    #add WOB for that PN if that PN is not an "open need" & then reserve it.
                    #find the first open task for the woo
                    if wo_task[-1] not in ['s','S']:
                        query = """SELECT WOT.WOT_AUTO_KEY,WO.WOO_AUTO_KEY,WOS.STATUS_TYPE FROM WO_TASK WOT
                        LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
                        LEFT JOIN WO_STATUS WOS ON WOS.WOS_AUTO_KEY = WOT.WOS_AUTO_KEY
                        WHERE 
                        UPPER(WO.SI_NUMBER) = UPPER('%s')
                        AND (WOS.STATUS_TYPE NOT IN ('Closed','Cancel') OR WOS.STATUS_TYPE IS NULL)
                        ORDER BY WOT.WOT_AUTO_KEY
                        """%wo_task
                        # 
                        wot = selection_dir(query,cr)
                        
                        if not wot:
                            return 'No open tasks.','',qty_res,must_reserve
                        if wot:
                            wot_auto_key = wot and wot[0] and wot[0][0] or None
                            woo_auto_key = wot and wot[0] and wot[0][1] or None
                            today = datetime.now()
                            timestamp = today.strftime('%Y-%m-%d %H:%M:%S')
                            #need to join on part_condition_codes to get the cond_level
                            #setting PCC to 'INSPECTED' is unique to QTEC implementation
                            cond_sub = "SELECT COND_LEVEL,PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE CONDITION_CODE = 'INSPECTED'"
                            pcc_data = selection_dir(cond_sub,cr)
                            pcc_auto_key = pcc_data and pcc_data[0] and pcc_data[0][1] or 1
                            cond_level = pcc_data and pcc_data[0] and pcc_data[0][0] or 0
                            q_wob = """INSERT INTO WO_BOM (REQUISITION,WOO_AUTO_KEY,SYSUR_AUTO_KEY,WOT_AUTO_KEY,PNM_AUTO_KEY,
                                QTY_NEEDED,ACTIVITY,PCC_AUTO_KEY,COND_LEVEL,ENTRY_DATE) 
                                VALUES('T','%s','%s','%s','%s','%s','%s','%s','%s',
                                TO_TIMESTAMP('%s', 'yyyy-mm-dd hh24:mi:ss'))"""%(woo_auto_key,sysur_auto_key,wot_auto_key,pnm_auto_key,quantity,'Consumable',pcc_auto_key,cond_level,timestamp)
                            error = insertion_dir(q_wob,cr)
                            if not error:
                                #get new wob_auto_key
                                query = """SELECT WOB_AUTO_KEY FROM WO_BOM
                                    WHERE ROWNUM <= 1 ORDER BY WOB_AUTO_KEY DESC
                                """
                                wob = selection_dir(query,cr)
                                new_wob_key = wob and wob[0] and wob[0][0] or ''
                                #Create a stock reservation for the wob
                                squery = """INSERT INTO STOCK_RESERVATIONS 
                                (STR_AUTO_KEY,STM_AUTO_KEY,WOB_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY) 
                                VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s)"""%(stm_auto_key,new_wob_key,quantity,sysur_auto_key)
                                error = insertion_dir(squery,cr) 
                                query = """
                                    SELECT STR_AUTO_KEY FROM STOCK_RESERVATIONS
                                    WHERE STM_AUTO_KEY = %s        
                                    ORDER BY STR_AUTO_KEY DESC
                                    """%stm_auto_key
                                str_auto=selection_dir(query,cr)
                                str_auto_key = str_auto and str_auto[0] and str_auto[0][0] or None
                                #update the audit trails with the correct user
                                squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
                                    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE WOB_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],new_wob_key)
                                error = updation_dir(squery,cr)
                                squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
                                    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE WOB_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],new_wob_key)
                                error = updation_dir(squery,cr)                
                                if error != '{"recs": ""}':
                                    return error,msg,qty_res,must_reserve
                                
                                squery = """UPDATE AUDIT_TRAIL
                                    SET SYSUR_AUTO_KEY=%s 
                                    WHERE
                                    SYSUR_AUTO_KEY = (select sysur_auto_key from sys_users where user_name='SYSDBA') 
                                    AND SOURCE_TABLE IN ('STM','QUANTUM')
                                    AND STAMPTIME >= SYSDATE - 1"""%sysur_auto_key  
                                error = updation_dir(squery,cr)
                                if error != '{"recs": ""}':
                                    return error,msg,qty_res,must_reserve                                
                                orcl_commit(con=con)                                 
                
            if str_auto_key and wot_auto_key and not error or error == '{"recs": ""}':
                #query = """
                #    DECLARE 
                #    CT qc_utl_pkg.cursor_type; 
                #    V_STR NUMBER;
                #    V_WOT NUMBER;
                #    V_QTY NUMBER;
                #    V_STM NUMBER;
                #    V_SYS NUMBER;
                #    BEGIN
                #    CT:=qc_ic_pkg2.spi_issue_reservation('%s','%s','%s',NULL,'%s','%s','N');
                #    END;
                #"""%(str_auto_key,wot_auto_key,quantity,stm_auto_key,sysur_auto_key)

                #error = updation_dir(query,cr)
                #if error != '{"recs": ""}':
                #    return error,msg,qty_res,must_reserve
                #else:
                error,msg,qty_res,must_reserve = stock_issue(quapi_id,session_id,\
                    sysur_auto_key,user_id,ctrl_number,\
                    ctrl_id,wo_task,quantity,active_mode,\
                    must_reserve) 
                #return error,msg,qty_res,must_reserve    
                                                               
    return error,msg,qty_res,must_reserve
  
@shared_task
def stock_issue(quapi_id,session_id,\
    sysur_auto_key,user_id,ctrl_number,\
    ctrl_id,wo_task,quantity,active_mode,\
    must_reserve):
    wob_auto_key,wot_auto_key,str_auto_key = '','',''
    error,msg,qty_res,qty_reserved,part_number,wot_sequence,si_number = '','','','','','',''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg,'',''    
    #User enters task, enters stock line and then the app should populate
    #    the qty field with the qty_reserved from the str related to the WOB.  
    
    #ISSUE MODE: Check to see if the task is reserved via the wob_auto_key on STOCK_RESERVATIONS.     
    #Populate the quantity_reserved from the reservation in the quantity field on front-end.    
    #If not, we must reserve it with the user's quantity.
    #If the user enters a different qty than what we populate with, create a new bom with the full user-entered qty
    # and then issue for the qty reserved.    

    #we have to consider the qty_needed on the BoM
    #we find a reservation for a wob, then we just issue it.
    #if we don't find a reservation for that wob, we go through the reservation process and then issue what we just reserved.
    right_now = datetime.now()
    entry_date = right_now.strftime('%m-%d-%Y %H:%M:%S') 
    #if not must_reserve or must_reserve=='F':       
    #qty_reserved becomes the qty issued.    
    if not wo_task:
        return 'Must enter WO# or task.','',qty_res,must_reserve      
    if quantity:
        #ISSUE MODE: Reduce qty_oh, qty_available by qty_reserved from the reservation. 
        #if wo_task[:-1] in ['s','S']:
            #error,msg,qty_res = stock_reserve(quapi_id,session_id,sysur_auto_key,user_id,wo_task,quantity,active_mode,ctrl_number,ctrl_id)                                      
        if wo_task and ctrl_number:
            stm,error = get_stock_res(cr,wo_task,ctrl_number,ctrl_id)
            if not stm:
                return error,'',qty_res,must_reserve 
            stm_auto_key = stm and stm[0] and stm[0][0] or ''
            pnm_auto_key = stm and stm[0] and stm[0][2] or ''
            #part_number = stm and stm[0] and stm[0][4] or ''
            qty_oh = stm and stm[0] and stm[0][5] or 0
            res_found = stm and isinstance(stm[0],list) and len(stm[0]) > 6
            if not res_found:
                return 'No reservation found.','',qty_res,must_reserve
                #if not qty_oh or float(quantity) > qty_oh:
                    #return 'Qty not available.','',qty_res,must_reserve
            
            if stm and stm[0] and res_found:
                wob_auto_key = stm and stm[0] and stm[0][6] or ''
                wot_auto_key = stm and stm[0] and stm[0][7] or ''
                woo_auto_key = stm and stm[0] and stm[0][8] or ''
                wot_sequence = stm and stm[0] and stm[0][9] or ''
                si_number = stm and stm[0] and stm[0][10] or ''
                qty_reserved = stm and stm[0] and stm[0][5] or 0
                #stm_original = stm and stm[0] and stm[0][11] or ''
                str_auto_key = stm and stm[0] and stm[0][12] or ''
                #pcc_auto_key = stm and stm[0] and stm[0][13] or ''
                #serial_number = stm and stm[0] and stm[0][14] or ''
                qty_needed = stm and stm[0] and stm[0][15] or 0
                qty_issued = stm and stm[0] and stm[0][16] or 0
                qty_oh = stm and stm[0] and stm[0][17] or 0
                #wot_status_type = stm and stm[0] and stm[0][17] or 0
                
                #if wot_status_type in ('Closed','Cancel'):
                    #return 'Task is closed.','',qty_res,must_reserve      
                if qty_reserved and qty_reserved < float(quantity):
                    #qty entered by user should not be greater than what is already reserved.  
                    return 'Qty must not be greater than qty reserved.','',qty_res,must_reserve               
            #if qty_reserved and qty_reserved + float(quantity) >= qty_oh:
                #qty entered by user should not be greater than what is already reserved.  
                #return 'Qty not available.','',qty_res,must_reserve
            """
            1. add WOB for that PN if that PN is not an "open need" on the BOM (Non existent for that woo or WO_BOM['QTY_NEEDED'] =QTY_RESERVED+QTY_ISSUED
            2. IF WOB is found with "open need" then insert into Stock Reservations for that WOB
            3. if STM is already reserved to that WOB then just issue the qty user enters

            OPEN NEED: WO_BOM['QTY_NEEDED'] < QTY_RESERVED+QTY_ISSUED           
            """            
            if not res_found and 0:
                if wob_auto_key and wot_auto_key:
                    if qty_needed >= qty_reserved + qty_issued + float(quantity):
                        #Create a stock reservation for the wob
                        squery = """INSERT INTO STOCK_RESERVATIONS 
                        (STR_AUTO_KEY,STM_AUTO_KEY,WOB_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY) 
                        VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s)"""%(stm_auto_key,wob_auto_key,quantity,sysur_auto_key)
                        error = insertion_dir(squery,cr)
                        if not error:
                            squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                                EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
                                (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG
                                WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],stm_auto_key)
                            error = updation_dir(squery,cr)
                         
                    else:
                        #no open need: qty_needed < qty_reserved + qty_issued + float(quantity)
                        #add WOB for that PN if that PN is not an "open need" & then reserve it.
                        today = datetime.now()
                        timestamp = today.strftime('%Y-%m-%d %H:%M:%S')
                        #need to join on part_condition_codes to get the cond_level        
                        cond_sub = "SELECT COND_LEVEL,PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE CONDITION_CODE = 'AR'"
                        pcc_data = selection_dir(cond_sub,cr)
                        pcc_auto_key = pcc_data and pcc_data[0] and pcc_data[0][1] or None
                        cond_level = pcc_data and pcc_data[0] and pcc_data[0][0] or None
                        q_wob = """INSERT INTO WO_BOM (REQUISITION,WOO_AUTO_KEY,SYSUR_AUTO_KEY,WOT_AUTO_KEY,PNM_AUTO_KEY,
                            QTY_NEEDED,ACTIVITY,PCC_AUTO_KEY,COND_LEVEL,ENTRY_DATE) 
                            VALUES('T','%s','%s','%s','%s','%s','%s','%s','%s',
                            TO_TIMESTAMP('%s', 'yyyy-mm-dd hh24:mi:ss'))
                            """%(woo_auto_key,sysur_auto_key,wot_auto_key,pnm_auto_key,quantity,'Consumable',pcc_auto_key,cond_level,timestamp)
                        error = insertion_dir(q_wob,cr)
                        if not error:
                            #get new wob_auto_key
                            query = """SELECT WOB_AUTO_KEY FROM WO_BOM ORDER BY WOB_AUTO_KEY DESC FETCH NEXT 1 ROWS ONLY
                            """
                            wob = selection_dir(query,cr)
                            new_wob_key = wob and wob[0] and wob[0][0] or ''
                            #Create a stock reservation for the wob
                            squery = """INSERT INTO STOCK_RESERVATIONS 
                            (STR_AUTO_KEY,STM_AUTO_KEY,WOB_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY) 
                            VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s)"""%(stm_auto_key,new_wob_key,quantity,sysur_auto_key)
                            error = insertion_dir(squery,cr)
                            query = """
                                SELECT STR_AUTO_KEY FROM STOCK_RESERVATIONS
                                WHERE ROWNUM <= 1 AND STM_AUTO_KEY = %s        
                                ORDER BY STR_AUTO_KEY DESC
                                """%stm_auto_key
                            str_auto=selection_dir(query,cr)
                            str_auto_key = str_auto and str_auto[0] and str_auto[0][0] or None
                else:
                    #no existing wob = no open need
                    #add WOB for that PN if that PN is not an "open need" & then reserve it.
                    #find the first open task for the woo
                    if wo_task[-1] not in ['s','S']:
                        query = """SELECT WOT.WOT_AUTO_KEY,WO.WOO_AUTO_KEY,WOS.STATUS_TYPE FROM WO_TASK WOT
                        LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
                        LEFT JOIN WO_STATUS WOS ON WOS.WOS_AUTO_KEY = WOT.WOS_AUTO_KEY
                        WHERE 
                        UPPER(WO.SI_NUMBER) = UPPER('%s')
                        AND (WOS.STATUS_TYPE NOT IN ('Closed','Cancel') OR WOS.STATUS_TYPE IS NULL)
                        ORDER BY WOT.WOT_AUTO_KEY
                        """%wo_task
                        # 
                        wot = selection_dir(query,cr)
                        
                        if not wot:
                            return 'No open tasks.','',qty_res,must_reserve
                        if wot:
                            wot_auto_key = wot and wot[0] and wot[0][0] or None
                            woo_auto_key = wot and wot[0] and wot[0][1] or None
                            
                            today = datetime.now()
                            timestamp = today.strftime('%Y-%m-%d %H:%M:%S')
                            #need to join on part_condition_codes to get the cond_level        
                            cond_sub = "SELECT COND_LEVEL,PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE CONDITION_CODE = 'AR'"
                            pcc_data = selection_dir(cond_sub,cr)
                            pcc_auto_key = pcc_data and pcc_data[0] and pcc_data[0][1] or None
                            cond_level = pcc_data and pcc_data[0] and pcc_data[0][0] or None
                            q_wob = """INSERT INTO WO_BOM (REQUISITION,WOO_AUTO_KEY,SYSUR_AUTO_KEY,WOT_AUTO_KEY,PNM_AUTO_KEY,
                                QTY_NEEDED,ACTIVITY,PCC_AUTO_KEY,COND_LEVEL,ENTRY_DATE) 
                                VALUES('T','%s','%s','%s','%s','%s','%s','%s','%s',
                                TO_TIMESTAMP('%s', 'yyyy-mm-dd hh24:mi:ss'))"""%(woo_auto_key,sysur_auto_key,wot_auto_key,pnm_auto_key,quantity,'Consumable',pcc_auto_key,cond_level,timestamp)
                            error = insertion_dir(q_wob,cr)
                            if not error:
                                #get new wob_auto_key
                                query = """SELECT WOB_AUTO_KEY FROM WO_BOM
                                    WHERE ROWNUM <= 1 ORDER BY WOB_AUTO_KEY DESC
                                """
                                wob = selection_dir(query,cr)
                                new_wob_key = wob and wob[0] and wob[0][0] or ''
                                #Create a stock reservation for the wob
                                squery = """INSERT INTO STOCK_RESERVATIONS 
                                (STR_AUTO_KEY,STM_AUTO_KEY,WOB_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY) 
                                VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s)"""%(stm_auto_key,new_wob_key,quantity,sysur_auto_key)
                                error = insertion_dir(squery,cr) 
                                query = """
                                    SELECT STR_AUTO_KEY FROM STOCK_RESERVATIONS
                                    WHERE ROWNUM <= 1 AND STM_AUTO_KEY = %s        
                                    ORDER BY STR_AUTO_KEY DESC
                                    """%stm_auto_key
                                str_auto=selection_dir(query,cr)
                                str_auto_key = str_auto and str_auto[0] and str_auto[0][0] or None                  
                
            if str_auto_key and wot_auto_key and error in ['{"recs": ""}','']:
                query = """
                    DECLARE 
                    CT qc_utl_pkg.cursor_type; 
                    V_STR NUMBER;
                    V_WOT NUMBER;
                    V_QTY NUMBER;
                    V_STM NUMBER;
                    V_SYS NUMBER;
                    BEGIN
                    CT:=qc_ic_pkg2.spi_issue_reservation('%s','%s','%s',NULL,'%s','%s','N');
                    END;
                """%(str_auto_key,wot_auto_key,quantity,stm_auto_key,sysur_auto_key)
 
                error = updation_dir(query,cr)
                
                if error != '{"recs": ""}':
                    return error,msg,qty_res,must_reserve
               
                #update the audit trails with the correct users
                squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
                    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],stm_auto_key)
                error = updation_dir(squery,cr)
                squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
                    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],stm_auto_key)
                error = updation_dir(squery,cr)                
                if error != '{"recs": ""}':
                    return error,msg,qty_res,must_reserve
                              
                query = """UPDATE STOCK_TI SET SYSUR_AUTO_KEY = %s where STI_AUTO_KEY = 
                (SELECT MAX(STI_AUTO_KEY) FROM STOCK_TI)
                """%sysur_auto_key
                error = updation_dir(query,cr)
                
                if error != '{"recs": ""}':
                    return error,msg,qty_res,must_reserve
                    
                squery = """UPDATE AUDIT_TRAIL
                    SET SYSUR_AUTO_KEY=%s 
                    WHERE
                    SYSUR_AUTO_KEY = (select sysur_auto_key from sys_users where user_name='SYSDBA') 
                    AND SOURCE_TABLE IN ('STM','QUANTUM')
                    AND STAMPTIME >= SYSDATE - 1"""%sysur_auto_key  
                error = updation_dir(squery,cr)
                
                if error != '{"recs": ""}':
                    return error,msg,qty_res,must_reserve
                    
                #query = """UPDATE STOCK_AUDIT SET TRAN_TYPE = 'STOCK ISSUE' WHERE STA_AUTO_KEY = 
                #    (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')
                #"""%stm_auto_key
                #error = updation_dir(squery,cr)
                
                #if error != '{"recs": ""}':
                #    return error,msg,qty_res,must_reserve
                    
                msg += ' Successful stock issuance.'                      
    else:
        #if the user didn't enter a quantity, we must return the quantity that is reserved.  
        #if there isn't a reservation, we have to create the bom with the user's qty and reserve it.
        stm,error = get_stock_res(cr,wo_task,ctrl_number,ctrl_id)
        qty_reserved = stm and stm[0] and stm[0][5] or 0
        if qty_reserved:
            return error,msg,qty_reserved,'F'
        else:
            return 'No reservation found.',msg,'','T'
    orcl_commit(con=con)            
    if not error or error == '{"recs": ""}':
        aud_status = 'success'
        error = ''
    else: 
        aud_status = 'failure'       
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='stock-allocation')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec and 0:
        rec_input = (ctrl_id and 'Record with Ctrl#: '+ str(ctrl_number) + ' and Ctrl ID#: ' + str(ctrl_id)) or ''
        new_val = rec_input + ' Stock issued.'
        field_changed = "Stock issued for: " + ctrl_number + ctrl_id + ', with quantity: ' + quantity + ' for task: ' + wo_task + ' '
        #field_changed += ', PN: ' + str(part_number) + ',Task seq.: ' + str(wot_sequence) + ',WO#: ' + si_number
        field_changed += error
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)
    #else:
    #    error = 'Incorrect Quantum User ID.'        
    return error,msg,qty_res,must_reserve
    
def get_stock_res(cr,wo_task,ctrl_number,ctrl_id):
    error = ''  
    if wo_task[-1] in ['s','S']:
        query = """SELECT DISTINCT S.STM_AUTO_KEY,S.QTY_AVAILABLE,'',S.HISTORICAL_FLAG,
            '',SR.QTY_RESERVED,SR.WOB_AUTO_KEY,WOT.WOT_AUTO_KEY,WO.WOO_AUTO_KEY,WOT.SEQUENCE,
            WO.SI_NUMBER,S.STM_ORIGINAL,SR.STR_AUTO_KEY,S.PCC_AUTO_KEY,S.SERIAL_NUMBER,
            WOB.QTY_NEEDED,WOB.QTY_ISSUED,S.QTY_OH
            FROM STOCK S 
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY 
            LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY 
            LEFT JOIN WO_TASK WOT ON WOT.WOT_AUTO_KEY = WOB.WOT_AUTO_KEY
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOB.WOO_AUTO_KEY           
            WHERE S.CTRL_NUMBER = %s
            AND S.CTRL_ID = %s AND WOB.WOT_AUTO_KEY = %s
            """%(ctrl_number,ctrl_id,wo_task[:-1])
        stm = selection_dir(query,cr)              
    else:
        query = """SELECT DISTINCT S.STM_AUTO_KEY,S.QTY_AVAILABLE,'',S.HISTORICAL_FLAG,
            '',SR.QTY_RESERVED,SR.WOB_AUTO_KEY,WOT.WOT_AUTO_KEY,WO.WOO_AUTO_KEY,WOT.SEQUENCE,
            WO.SI_NUMBER,S.STM_ORIGINAL,SR.STR_AUTO_KEY,S.PCC_AUTO_KEY,S.SERIAL_NUMBER,
            WOB.QTY_NEEDED,WOB.QTY_ISSUED,S.QTY_OH
            FROM STOCK S
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
            LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
            LEFT JOIN WO_TASK WOT ON WOT.WOT_AUTO_KEY = WOB.WOT_AUTO_KEY
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
            WHERE 
            S.CTRL_NUMBER = %s
            AND S.CTRL_ID = %s AND UPPER(WO.SI_NUMBER) = UPPER('%s')"""%(ctrl_number,ctrl_id,wo_task)
        #query += " WHERE WO.WOO_AUTO_KEY  IS NOT NULL AND ROWNUM <= 10"
        #65734400001
        #JT81396-7-7
        stm = selection_dir(query,cr)
    if not stm:
        #check if the STM exists without the task
        query = """SELECT S.STM_AUTO_KEY,S.QTY_AVAILABLE,P.PNM_AUTO_KEY,
            S.HISTORICAL_FLAG,P.PN,S.QTY_OH            
            FROM STOCK S
            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
            WHERE S.CTRL_NUMBER = %s AND S.CTRL_ID = %s
        """%(ctrl_number,ctrl_id)
        stm = selection_dir(query,cr)
        if not stm:
           error = 'Stock not found.'  
    return stm,error  


@shared_task
def stock_reserve(quapi_id,session_id,sysur_auto_key,user_id,wo_task,quantity,active_mode,ctrl_number,ctrl_id):
    error,msg,split_bom,squery = '','',False,''
    if not quantity:
        return 'Must enter quantity.',msg,''
    if not wo_task:
        return 'Must enter task or SI Number.',msg,''
    if not ctrl_number:
        return 'Must enter stock.',msg,''
    from polls.models import QueryApi    
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    cr,con = get_cursor_con(quapi)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg,''
    if quantity:
        error,quantity = qty_to_float(quantity)
        if error:
            return error,msg,''
    #query = """SELECT S.STM_AUTO_KEY,S.QTY_AVAILABLE,S.PNM_AUTO_KEY,S.HISTORICAL_FLAG,
    #    P.PN FROM STOCK S
    #    LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY        
    #    WHERE S.CTRL_NUMBER = %s AND S.CTRL_ID = %s
    #    """%(ctrl_number,ctrl_id)
    #        --LEFT JOIN WO_OPERATION WOO ON WOO.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
    query = """SELECT S.STM_AUTO_KEY,S.QTY_AVAILABLE,S.PNM_AUTO_KEY,S.HISTORICAL_FLAG,
        P.PN,SR.WOB_AUTO_KEY,SR.QTY_RESERVED FROM STOCK S
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
        LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
        LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOB.WOO_AUTO_KEY
        LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY        
        WHERE S.CTRL_NUMBER = %s AND S.CTRL_ID = %s"""%(ctrl_number,ctrl_id)
    right_now = datetime.now()
    timestamp = right_now.strftime('%Y-%m-%d %H:%M:%S')     
    stm = selection_dir(query,cr)
    stm_auto_key = stm and stm[0] and stm[0][0] or None
    if not stm_auto_key:
        return 'Cannot find stock line.',msg,''       
    qty_available = stm and stm[0] and stm[0][1] or 0.0
    if qty_available <= 0 or qty_available < quantity:
        return 'Not enough stock available.',msg,''
    pnm_auto_key = stm and stm[0] and stm[0][2] or None
    if not pnm_auto_key:
        return 'Cannot find part.',msg,''
    historical_flag = stm and stm[0] and stm[0][3] or 'F'
    if active_mode == '1' and historical_flag == 'T':
        return 'Stock is historical.',msg,''  
    part_number = stm and stm[0] and stm[0][4] or ''   
    wob_auto_key = stm and stm[0] and stm[0][5] or ''
    qty_reserved = stm and stm[0] and stm[0][6] or 0.0    
    #if wob_auto_key and qty_reserved > 0.0:
    #    return 'Stock already reserved.',msg,''    
    #1. Add new stock reservation for stm   
    #2. Get wot, stm, wob (if doesn't exist, then create it, pcc data, timestamp and then create the reservation.)
    if not wob_auto_key:
        # or qty_reserved <= 0
        if wo_task[-1] in ['s','S']:    
            query = """SELECT WOB.WOB_AUTO_KEY FROM WO_BOM WOB      
                WHERE 
                WOB.PNM_AUTO_KEY = %s
                AND WOB.QTY_NEEDED - WOB.QTY_RESERVED + WOB.QTY_ISSUED >= %s
                AND WOB.WOT_AUTO_KEY = %s"""%(pnm_auto_key,quantity,wo_task[:-1])
        else:
            query = """SELECT WOB.WOB_AUTO_KEY FROM WO_BOM WOB
                LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOB.WOO_AUTO_KEY       
                WHERE 
                WOB.PNM_AUTO_KEY = %s
                AND WOB.QTY_NEEDED - WOB.QTY_RESERVED + WOB.QTY_ISSUED >= %s
                AND WO.SI_NUMBER = '%s'"""%(pnm_auto_key,quantity,wo_task)
        wob = selection_dir(query,cr)
        wob_auto_key = wob and wob[0] and wob[0][0] or None
    if not wob_auto_key:
        error = 'BoM not found.'   
    else: 
        squery = """INSERT INTO STOCK_RESERVATIONS 
            (STR_AUTO_KEY,STM_AUTO_KEY,WOB_AUTO_KEY,QTY_RESERVED,SYSUR_AUTO_KEY) 
            VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s',%s,%s)"""%(stm_auto_key,wob_auto_key,quantity,sysur_auto_key)
        error = insertion_dir(squery,cr)
        if not error:
            squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, 
                EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
                (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG
                WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],stm_auto_key)
            error = updation_dir(squery,cr)
    if error == '{"recs": ""}' or error == '':
        error=''
        msg = 'Successful reservation.'
    orcl_commit(con=con) 
    if active_mode == '1':
        if error == '':
            error=''
            aud_status = 'success'
        else: 
            aud_status = 'failure'       
        from polls.models import MLApps as maps,QuantumUser as qu
        app_id = maps.objects.filter(code='stock-allocation')   
        user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
        user_rec = user_rec and user_rec[0] or None
        if user_rec:
            rec_input = (ctrl_id and 'Record with Ctrl#: '+ str(ctrl_number) + ' and Ctrl ID#: ' + str(ctrl_id)) or ''
            new_val = rec_input
            mode = 'Reserved stock for '
            #add parts_master.pn, wo_task.sequence, wo_operation.si_number
            field_changed = mode + ctrl_number + ctrl_id + ', with quantity: ' + str(quantity) + ' for task/WO#: ' + wo_task + ' '
            field_changed += ', PN: ' + str(part_number)
            field_changed += error
            error += register_audit_trail(user_rec,field_changed,new_val,timestamp,app_id,quapi,status=aud_status) 
        else:
            error = 'Incorrect Quantum User ID.'
    return error,msg,quantity
   
@shared_task
def file_retrieval(quapi_id,wo_number,session_id):
    doc_data = []
    error,msg = '',''
    source_table = 'STOCK'
    from polls.models import QueryApi,OracleConnection as oc,Document
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    server_url = orcl_conn.url
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    #query for the image location/url/directory and the images by WO#
    query = """SELECT STM_AUTO_KEY FROM STOCK_TI WHERE WOB_AUTO_KEY IN 
    (SELECT WOB_AUTO_KEY FROM WO_BOM WHERE WOO_AUTO_KEY = (SELECT 
    WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER='%s'))
    AND QTY > 1 AND QTY_REVERSE = 0
    """%wo_number
    stms = selection_dir(query,cr)
    stm_list = construct_id_list(stms,id_pos=0)

    query = """select iml.source_pk,iml.source_table,iml.file_name,ims.server_ip,
        ims.server_port,ims.url_path,iml.image_key,iml.file_ext,iml.date_created 
        from image_list iml
        join image_server ims on ims.ims_auto_key = iml.ims_auto_key
        join image_codes imc on imc.imc_auto_key = iml.imc_auto_key
        where source_pk in %s
        and imc.regulatory_flag = 'T'
        and source_table='%s'"""%(stm_list,source_table)  
    #        
    file_data = selection_dir(query,cr)
    #store the image locations
    for row in file_data:    
        doc_data.append(Document(
        source_pk = row[0],
        source_table = row[1],
        server_url = row[3],
        server_port = row[4],
        url_path = row[5],
        file_name = row[2],
        file_key = row[6] or 0,
        file_extension = row[7],
        create_date = row[8],
        session_id = session_id,      
        ))
    if doc_data:     
        try:
            delete = Document.objects.filter(session_id=session_id).delete()
            rec = Document.objects.bulk_create(doc_data) or []    
        except Exception as exc:
            error += "Error, %s, with retrieving or creating files for WO#, %s"%(exc,wo_number) 
    else:
        error = 'No files found.'
        
    return error,msg

@shared_task
def stock_turn_in(cr,wob_auto_key,woo_auto_key,pnm_auto_key,serial_no,quantity,pcc_auto_key,stm_parent=None):
    #here we use the qc_ic_pkg.spi_wo_turn_in(WOB,v_woo,PNM,SERIAL,QTY,NULL,NULL,v_pcc)
    if not stm_parent:
        squery = """
        
            DECLARE 
            CT qc_utl_pkg.cursor_type; 
            v_stm NUMBER;
            v_pcc NUMBER;
            BEGIN
                select      pcc_auto_key
                into        v_pcc
                from        part_condition_codes
                where       condition_code = 'AR' ;
            v_stm := qc_ic_pkg.spi_wo_turn_in('%s', '%s', '%s', '%s', '%s',NULL,NULL,v_pcc);  
            END;"""%(wob_auto_key,woo_auto_key,pnm_auto_key,serial_no,quantity)
    else:
        squery = """
        
            DECLARE 
            CT qc_utl_pkg.cursor_type; 
            v_stm NUMBER;
            v_pcc NUMBER;
            BEGIN
                select      pcc_auto_key
                into        v_pcc
                from        part_condition_codes
                where       condition_code = 'AR' ;
            v_stm := qc_ic_pkg.spi_wo_turn_in('%s', '%s', '%s', '%s', '%s','%s',NULL,v_pcc);  
            END;"""%(wob_auto_key,woo_auto_key,pnm_auto_key,serial_no,quantity,stm_parent)       
    error = updation_dir(squery,cr)
    return error 
    
@shared_task
def create_pn(quapi_id,session_id,pn,desc,serialized,\
    sysur_auto_key,cr=None,con=None,serial_no='',lot_import=False):
                                                                   
                                             
    error,msg = '',''
    if not cr:
        from polls.models import QueryApi,OracleConnection as oc
        quapi = QueryApi.objects.filter(id=quapi_id)
        quapi = quapi and quapi[0] or None
        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
        orcl_conn = orcl_conn and orcl_conn[0] or None
        if orcl_conn:
            cr,con = orcl_connect(orcl_conn)
        if not (cr and con):
            return 'Cannot connect to Oracle',msg
                                                 
    if not desc:
        desc = pn
    #Validate PN + SERIAL_NUMBER combination does not already exist and qty_oh > 0
    if serial_no:
                 
        query = """SELECT P.PNM_AUTO_KEY,S.SERIAL_NUMBER FROM STOCK S
            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
            WHERE S.QTY_OH>0 AND P.PN='%s' AND S.SERIAL_NUMBER='%s'
            """%(pn,serial_no)
        recs = selection_dir(query,cr)
        if recs:
            return 'Part %s with serial number, %s, already exists.'%(),msg
    #manufacturer code:
    query = "SELECT MFG_AUTO_KEY FROM MANUFACTURER WHERE UPPER(DESCRIPTION) = 'MRO LIVE' OR UPPER(MFG_CODE) = 'MRO LIVE'"
    mfg_codes = selection_dir(query,cr)
    mfg_auto_key = mfg_codes and mfg_codes[0] and mfg_codes[0][0] or ''    
    if not mfg_auto_key:
        #create a new one with 'MRO LIVE' as description
        #return 'You must create a manufacturer, "MRO LIVE" in Quantum.',msg
        mfg_code = 'MRO LIVE'
        query = "INSERT INTO MANUFACTURER (MFG_AUTO_KEY,MFG_CODE,DESCRIPTION) VALUES(G_MFG_AUTO_KEY.NEXTVAL,'%s','%s')"%(mfg_code,mfg_code)
        error = insertion_dir(query,cr)
        if not error:
            query = "SELECT MFG_AUTO_KEY FROM MANUFACTURER WHERE DESCRIPTION = 'MRO LIVE'"
            mfg_codes = selection_dir(query,cr)
            mfg_auto_key = mfg_codes and mfg_codes[0] and mfg_codes[0][0] or ''
            if not mfg_auto_key:
                return 'Cannot create new manufacturer for "MRO LIVE"',msg
  
    query = """INSERT INTO PARTS_MASTER (REQ_VERIFICATION,PN,DESCRIPTION,SERIALIZED,MFG_AUTO_KEY,SYSUR_AUTO_KEY) 
        VALUES('T',SUBSTR('%s',0,40),SUBSTR('%s',0,50),'%s','%s','%s')"""%(pn,desc,serialized,mfg_auto_key,sysur_auto_key)
           
    error = insertion_dir(query,cr)
    if not lot_import:    
        orcl_commit(con=con)                  

    return error,msg
   
def check_serialization(quapi_id,cr,con,\
    session_id,sysur_auto_key,pn,user_desc,
    serial_no,quantity,lot_import):
    error,msg = '',''
    pnm_auto_key,activity = '',''
    
    query = """SELECT P.SERIALIZED,P.PNM_AUTO_KEY,P.ACTIVITY,OPM.DEFAULT_REPAIR,P.DESCRIPTION 
      FROM PARTS_MASTER P 
      LEFT JOIN OPERATION_MASTER OPM ON OPM.PNM_AUTO_KEY = P.PNM_AUTO_KEY
      WHERE UPPER(P.PN) = UPPER('%s') AND ROWNUM <= 1"""%pn
    serialized = selection_dir(query,cr)
    is_serialized = serialized and serialized[0] and serialized[0][0] or None
    is_serialized = (is_serialized == 'T' and 'T') or 'F'
    pnm_auto_key = serialized and serialized[0] and serialized[0][1] or None
    activity = serialized and serialized[0] and serialized[0][2] or ''
    description = serialized and serialized[0] and serialized[0][4] or user_desc     
    if activity != 'Repair':
        for ser in serialized:
            if ser[3] == 'T':
                activity = 'Repair'
                break
    if not pnm_auto_key:
        
        if not lot_import:
            #raise a pop-up to give the 
            #user the ability to create a new product.
            error = ''       
            return [error,msg,serial_no,pnm_auto_key,activity,description]
        else:
                                 
            if serial_no or quantity == 1:
                is_serialized = 'T'
            #1.	Create new PNM if does not exist
            error,msg = create_pn(quapi_id,session_id,\
                pn,user_desc,is_serialized,sysur_auto_key,\
                cr=cr,con=con,serial_no=serial_no,\
                lot_import = True)
            if not error:
                query = """SELECT PNM_AUTO_KEY FROM PARTS_MASTER 
                WHERE ROWNUM<=1 ORDER BY PNM_AUTO_KEY DESC
                """
                pnm = selection_dir(query,cr)
                pnm_auto_key = pnm and pnm[0] and pnm[0][0] or None    
    if is_serialized == 'T' and quantity > 1:
        #raise exception because the part is serialized but the quantity is greater than 1.
        error = 'Serialized parts cannot have quantity > 1.'
        #need a pop-up to read in a list
        return [error,msg,'',pnm_auto_key,activity,description]
    if is_serialized == 'F':
        
        serial_no = ''
    #1. check for serialized first:
    if is_serialized and is_serialized == 'T' and not serial_no:
        #raise exception because the part is serialized.
        if lot_import:
            serial_no = 'NSN'
        else:
            error = 'PN must have serial number.'
            return [error,msg,'',pnm_auto_key,activity,description]  
    return [error,msg,serial_no,pnm_auto_key,activity,description]      
      
@shared_task
def lot_teardown(quapi_id,sysur_auto_key,user_id,row,lot_import=False,line_count=0):
    show_modal,error,msg,err_stock,err_woo = '','','','',''
    qty_oh,si_num,new_wob_key,si_number,stock_line = 0,'','','',''
    lot_core_settings,sti_wob_auto_key,ic_udn_010 = '','',''
    container,consignment,ic_udl_005,ic_udn_006 = '','','',''
    unit_cost,condition,part_cert_num,notes = '','','',''
    tag_date,remarks,ic_udn_009,ic_udn_008,ic_udn_007,lot_apl_ro_cost='','','','','',''
    part_cert_by,lot_no,location,rec_date,update_clause = '','','','',''
    wot_data = []
    from polls.models import QueryApi,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    today = datetime.now()
    timestamp = today.strftime('%m/%d/%Y hh:mm:ss')
    datestamp = today.strftime('%m/%d/%Y')
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg,'' 
        
    wo_task = row[0]
    quantity = row[1]
    if not quantity:
        return 'Quantity is required.',msg,'' 
    try:
        quantity = quantity and float(row[1])
    except Exception as exc:
        return 'Quantity must be a number - %s.'%(exc),msg,''
    pn = row[2]
    #pn = pn.replace("'","''")
    if not pn:
        return 'Part number is required.',msg,'' 
    serial_no = row[3]
    session_id = row[4]
    notes = row[5]
    notes = notes.replace("'", r'')
    notes = notes.replace('"', r'')
    description = row[6]

    if wo_task:                                  
        woo_sub = """SELECT WO.WOO_AUTO_KEY,'',WO.SI_NUMBER,S.STM_AUTO_KEY,
                       S.GLA_AUTO_KEY,WO.GLA_AUTO_KEY,S.PNM_AUTO_KEY,WO.DPT_AUTO_KEY,
                       WO.CMP_AUTO_KEY,WO.SYSCM_AUTO_KEY,WO.KIT_QTY,WO.OPM_AUTO_KEY,WO.DUE_DATE,
                       WO.ATTENTION,WO.ECD_METHOD,WO.WWT_AUTO_KEY,WO.SVC_AUTO_KEY,
                       WO.NEW_WIP_ACCT,WO.BGS_DEFAULT,WO.CUR_AUTO_KEY,WO.GLA_LABOR,WO.GLA_MISC,
                       WO.LOT_CORE_SETTINGS,S.RECEIVER_NUMBER,WO.PCC_AUTO_KEY,WO.CNC_AUTO_KEY,
                       S.LOC_AUTO_KEY,S.STC_AUTO_KEY,S.CTS_AUTO_KEY,S.WHS_AUTO_KEY,S.STM_ORIGINAL,
                       WO.LOT_APL_RO_COST,WO.LOT_ALW_PRECOST,WO.LOT_REQ_INSPECTION,WO.LOT_COST_DELAYED,
                       S.ORIGINAL_PO_NUMBER,S.SERIAL_NUMBER,WO.WOO_AUTO_KEY,S.IC_UDF_008,S.CTRL_ID,
                       S.CTRL_NUMBER,PC.CONDITION_CODE,S.STOCK_LINE,L.LOCATION_CODE,S.MFG_LOT_NUM,
                       S.EXP_DATE,CNC.CONSIGNMENT_CODE,CMP.COMPANY_NAME,WO.WO_UDF_002,WO.WO_UDF_003
                               
                       FROM WO_OPERATION WO
                       LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
                       LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
                       LEFT JOIN WO_BOM WOB ON WOB.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
                       --LEFT JOIN STOCK_RESERVATIONS SRB ON SRB.WOB_AUTO_KEY = WOB.WOB_AUTO_KEY
                       LEFT JOIN PART_CONDITION_CODES PC ON PC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
                       LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
                       LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = WO.CNC_AUTO_KEY
                       LEFT JOIN COMPANIES CMP ON CMP.CMP_AUTO_KEY = WO.CMP_AUTO_KEY                       
                       WHERE UPPER(WO.SI_NUMBER) = UPPER('%s') AND WO.WO_TYPE = 'Lot' AND ROWNUM <= 1"""%(wo_task)
        wot_data = selection_dir(woo_sub,cr)      
                                                                                   
    stock_rec = wot_data and wot_data[0] and wot_data[0][3] or []
    if not stock_rec:
        woo_sub = """SELECT WO.WOO_AUTO_KEY,'',WO.SI_NUMBER,S.STM_AUTO_KEY,
                       S.GLA_AUTO_KEY,WO.GLA_AUTO_KEY,S.PNM_AUTO_KEY,WO.DPT_AUTO_KEY,
                       WO.CMP_AUTO_KEY,WO.SYSCM_AUTO_KEY,WO.KIT_QTY,WO.OPM_AUTO_KEY,WO.DUE_DATE,
                       WO.ATTENTION,WO.ECD_METHOD,WO.WWT_AUTO_KEY,WO.SVC_AUTO_KEY,
                       WO.NEW_WIP_ACCT,WO.BGS_DEFAULT,WO.CUR_AUTO_KEY,WO.GLA_LABOR,WO.GLA_MISC,
                       WO.LOT_CORE_SETTINGS,S.RECEIVER_NUMBER,WO.PCC_AUTO_KEY,WO.CNC_AUTO_KEY,
                       S.LOC_AUTO_KEY,S.STC_AUTO_KEY,S.CTS_AUTO_KEY,S.WHS_AUTO_KEY,S.STM_ORIGINAL,
                       WO.LOT_APL_RO_COST,WO.LOT_ALW_PRECOST,WO.LOT_REQ_INSPECTION,WO.LOT_COST_DELAYED,
                       S.ORIGINAL_PO_NUMBER,S.SERIAL_NUMBER,WO.WOO_AUTO_KEY,S.IC_UDF_008,S.CTRL_ID,
                       S.CTRL_NUMBER,PC.CONDITION_CODE,S.STOCK_LINE,L.LOCATION_CODE,S.MFG_LOT_NUM,
                       S.EXP_DATE,CNC.CONSIGNMENT_CODE,CMP.COMPANY_NAME,WO.WO_UDF_002,WO.WO_UDF_003
                               
                       FROM WO_OPERATION WO
                       LEFT JOIN RO_DETAiL ROD ON ROD.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
                       LEFT JOIN STOCK_RESERVATIONS SR ON SR.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY
                       LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
                       LEFT JOIN WO_BOM WOB ON WOB.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
                       --LEFT JOIN STOCK_RESERVATIONS SRB ON SRB.WOB_AUTO_KEY = WOB.WOB_AUTO_KEY
                       LEFT JOIN PART_CONDITION_CODES PC ON PC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
                       LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
                       LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = WO.CNC_AUTO_KEY
                       LEFT JOIN COMPANIES CMP ON CMP.CMP_AUTO_KEY = WO.CMP_AUTO_KEY                       
                       WHERE UPPER(WO.SI_NUMBER) = UPPER('%s') AND WO.WO_TYPE = 'Lot' AND ROWNUM <= 1"""%(wo_task)
        wot_data = selection_dir(woo_sub,cr)
    if not wot_data:
        error = 'No lot found.'
    else:
        woo_auto_key = wot_data[0][0] or ''
        wot_auto_key = wot_data[0][1] or ''
        si_number = wot_data[0][2] or '' 
        stm_parent_key = wot_data[0][3] or ''
        stm_auto_key = stm_parent_key
        stm_gla_key = wot_data[0][4] or ''
        woo_gla_key = wot_data[0][5] or ''
        dpt_auto_key = wot_data[0][7] or ''
        cmp_auto_key = wot_data[0][8] or ''
        syscm_auto_key = wot_data[0][9] or ''
        kit_qty = wot_data[0][10] or ''
        opm_auto_key = wot_data[0][11] or ''
        due_date = wot_data[0][12] or ''
        attention = wot_data[0][13] or ''
        ecd_method = wot_data[0][14] or ''
        wwt_auto_key = wot_data[0][15] or ''
        svc_auto_key = wot_data[0][16] or ''
        new_wip_acct = wot_data[0][17] or ''
        bgs_default = wot_data[0][18] or ''
        cur_auto_key = wot_data[0][19] or ''
        gla_labor = wot_data[0][20] or ''
        gla_misc = wot_data[0][21] or ''
        lot_core_settings = wot_data[0][22] or ''
        receiver_number = wot_data[0][23] or 'not_empty'                                                  
        pcc_auto_key = wot_data[0][24] or ''
        cnc_auto_key = wot_data[0][25] or ''
        loc_auto_key = wot_data[0][26] or ''                  
        stc_auto_key = wot_data[0][27] or ''
        cts_auto_key = wot_data[0][28] or ''
        whs_auto_key = wot_data[0][29] or ''
        stm_parent = wot_data[0][30] or ''
        lot_apl_ro_cost = wot_data[0][31] or ''
        lot_alw_precost = wot_data[0][32] or ''
        lot_req_inspection = wot_data[0][33] or ''
        lot_cost_delayed = wot_data[0][34] or ''
        original_po_number = wot_data[0][35] or ''
        serial_number = wot_data[0][36] or ''
        wob_auto_key = wot_data[0][37] or ''
        ex_esn = wot_data[0][38] or ''
        ctrl_id = wot_data[0][39] or ''
        ctrl_number = wot_data[0][40] or ''
        condition = wot_data[0][41] or ''
        stock_line = wot_data[0][42] or ''
        loc_code = wot_data[0][43] or ''
        mfg_lot_num = wot_data[0][44] or ''
        exp_date = wot_data[0][45] or ''
        consignment_code = wot_data[0][46] or ''
        customer = wot_data[0][47] or ''
        eng_model = wot_data[0][48] or ''
        ata_code = wot_data[0][49] or ''
                                  
        vals = check_serialization(quapi_id,cr,con,       
            session_id,sysur_auto_key,pn,description,
            serial_no,quantity,lot_import)
        error = vals[0]
        msg = vals[1]
        serial_no = vals[2]
        pnm_auto_key = vals[3]
        activity = vals[4]       
        description = vals[5] or description
        
        if lot_import:
            from datetime import date
            lot_no = row[10]
            unit_cost = row[11]
            condition = row[12]
            part_cert_num = row[13]
            notes = row[14]
            tag_date = row[15]
            if tag_date and isinstance(tag_date, date):
                tag_date = tag_date and tag_date.strftime('%m/%d/%Y') or ''
            else:
                tag_date = tag_date and len(tag_date) > 9 and tag_date[:10] or ''
            remarks = row[16]
            ic_udn_006 = row[17] or '' 
            ic_udn_007 = row[18] or ''       
            part_cert_by = row[20] or '' 
            rec_date = row[21]         
            if rec_date and isinstance(rec_date,date):
                rec_date = rec_date and rec_date.strftime('%m/%d/%Y') or ''
            else:
                rec_date = rec_date and len(rec_date) > 9 and rec_date[:10] or ''
            lot_alw_precost = row[22] or 'F'
            lot_apl_ro_cost = row[23] or 'F'
            ic_udn_008 = row[24] or ''             
            ic_udn_009 = row[25] or '' 
            ic_udn_010 = row[26] or ''
            hold_line = row[28] or False
            traceable_to = row[29] or ''
            obtained_from = row[30] or ''
            item_num = row[31] or ''
            #search for consignment and get cnc_auto_key
            query = """SELECT PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE UPPER(CONDITION_CODE) = UPPER('%s')"""%row[12]
            pcc = selection_dir(query,cr)
            pcc_auto_key = pcc and pcc[0] and pcc[0][0] or pcc_auto_key
            
            #search for condition and get pcc_auto_key
            query = """(SELECT CNC_AUTO_KEY FROM CONSIGNMENT_CODES WHERE UPPER(CONSIGNMENT_CODE) = UPPER('%s'))"""%row[9]
            cnc = selection_dir(query,cr)
            cnc_auto_key = cnc and cnc[0] and cnc[0][0] or cnc_auto_key
            
            #search for location and get loc_auto_key
            query = """(SELECT LOC_AUTO_KEY FROM LOCATION WHERE UPPER(LOCATION_CODE)=UPPER('%s'))"""%row[7]
            loc = selection_dir(query,cr)
            loc_auto_key = loc and loc[0] and loc[0][0] or loc_auto_key

            ex_esn = row[11]
            
            if original_po_number:
                query = """SELECT CMP_AUTO_KEY FROM PO_HEADER WHERE PO_NUMBER = '%s'"""%original_po_number
                poh = selection_dir(query,cr)
                cmp_auto_key = poh and poh[0] and poh[0][0] or cmp_auto_key
   
            if row[8]:
                #if it doesn't exist yet, generate the IC_UDL_005
                udl_query = """(SELECT UDL_AUTO_KEY FROM USER_DEFINED_LOOKUPS
                    WHERE UDL_COLUMN_NAME = '%s' AND UDL_CODE = '%s'
                    AND UDL_DESCRIPTION = '%s' AND SEQUENCE = 1)
                    """%('IC_UDL_005',row[8],row[8])
                udl = selection_dir(query,cr)
                if not udl:
                    ins_query = """INSERT INTO USER_DEFINED_LOOKUPS 
                        (UDL_AUTO_KEY,UDL_COLUMN_NAME,UDL_CODE,
                        UDL_DESCRIPTION,SEQUENCE,HISTORICAL_FLAG) 
                        VALUES(G_UDL_AUTO_KEY.NEXTVAL,'%s','%s','%s','%s','F')
                        """%('IC_UDL_005',row[8],row[8],1)   
                    udl_error = insertion_dir(ins_query,cr)
                    if udl_error != '{"recs": ""}':
                        udl_error = 'Line: ' + str(line_count) + '-:>' + udl_error
                        return udl_error,msg,'done'
                        
                if not udl and not error:
                    udl = selection_dir(query,cr)
                ic_udl_005 = udl and udl[0] and udl[0][0] or ''   
                        
            if part_cert_num:
                update_clause += ", PART_CERT_NUMBER = '%s'"%part_cert_num
                
            if part_cert_by:
                update_clause += ", IC_UDF_010 = '%s'"%part_cert_by
                
            if tag_date:
                update_clause += ", TAG_DATE = TO_DATE('%s','mm/dd/yyyy')"%tag_date
            
            if remarks:
                update_clause += ", REMARKS = '%s'"%remarks
                
            if ic_udl_005:
                update_clause += ", IC_UDN_005= '%s'"%ic_udl_005
                
            if ic_udn_006:
                update_clause += ", IC_UDF_020= '%s'"%ic_udn_006
                
            if ic_udn_007:
                update_clause += ", IC_UDN_007= '%s'"%ic_udn_007

            if ic_udn_008:
                update_clause += ", IC_UDN_008= '%s'"%ic_udn_008
            
            if ic_udn_009:
                update_clause += ", IC_UDN_009= '%s'"%ic_udn_009
            
            if ic_udn_010:
                update_clause += ", IC_UDN_010= '%s'"%ic_udn_010
            
            if rec_date:
                update_clause += ", REC_DATE = TO_DATE('%s','mm/dd/yyyy')"%rec_date
                
            if unit_cost:
                update_clause += ", IC_UDN_002= '%s'"%unit_cost               
                    
            if hold_line and hold_line in ['T','F','t','f']:
                update_clause += ", HOLD_LINE = UPPER('%s')"%hold_line
                
                             
               
                                                        
                                                         
               
            if traceable_to:
                                                                                                                                             
                update_clause += ", IC_UDF_007= '%s'"%traceable_to
                
            if obtained_from:
                                                                                                                                           
                update_clause += ", IC_UDF_005= '%s'"%obtained_from                
                
                                                        
                                                                                                                                                
                                         
                                                                             
            if item_num:
                update_clause += ", SDF_STM_001 = '%s'"%item_num

                                                        
                                                                                                                                         
                                
                                                                                                                                 
                                                      
                                                                                                              
                                                                            
                                                          
                                       
                     
                                 
        if not woo_auto_key:
            #raise exception because the part is serialized.
            error = 'Lot not found.'
            return error,msg,''           
        #cond_sub = "SELECT COND_LEVEL,PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE CONDITION_CODE = 'SV'"
        cond_sub = "SELECT COND_LEVEL,PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE PCC_AUTO_KEY = (SELECT PCC_MAIN_OUT FROM WO_CONTROL)"
        pcc_data = selection_dir(cond_sub,cr)
        pcc_out = pcc_data and pcc_data[0] and pcc_data[0][1] or ''
        cond_level = pcc_data and pcc_data[0] and pcc_data[0][0] or ''
        ins_qty = 0
        if pnm_auto_key:
            q_wob = """INSERT INTO WO_BOM (EXPECTED_SN,WOO_AUTO_KEY,SYSUR_AUTO_KEY,
            WOT_AUTO_KEY,PNM_AUTO_KEY,QTY_NEEDED,ACTIVITY,PCC_AUTO_KEY,
            COND_LEVEL,ENTRY_DATE,NOTES,QTY_TURN,QTY_RESERVED) 
            VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s',TO_DATE('%s', 'mm/dd/yyyy'),TO_CLOB('%s'),'%s','%s')
            """%(serial_no,woo_auto_key,sysur_auto_key,wot_auto_key,pnm_auto_key,quantity,'Work Order',pcc_auto_key or pcc_out,cond_level,datestamp,notes,ins_qty,ins_qty)
            wob_error = insertion_dir(q_wob,cr)
            if wob_error:
                wob_error = 'Line: ' + str(line_count) + '-:>' + wob_error
                return wob_error,msg,'done' 
    
        if not error and pnm_auto_key:
            query = """SELECT WOB_AUTO_KEY FROM WO_BOM 
            WHERE PNM_AUTO_KEY = '%s' 
            ORDER BY WOB_AUTO_KEY DESC
            """%(pnm_auto_key)
            wob = selection_dir(query,cr)
            wob_auto_key = wob and wob[0] and wob[0][0] or None
            cond_sub = """SELECT COND_LEVEL,PCC_AUTO_KEY 
                FROM PART_CONDITION_CODES WHERE 
                PCC_AUTO_KEY = (SELECT PCC_TURN_IN FROM WO_CONTROL)"""
            pcc_data = selection_dir(cond_sub,cr)
            pcc_in = pcc_data and pcc_data[0] and pcc_data[0][1] or None
            if wob_auto_key:
                error = stock_turn_in(cr,wob_auto_key,woo_auto_key,pnm_auto_key,serial_no,quantity,pcc_auto_key or pcc_in) 
            else:
                error = 'No BOM found for lot %s.'%wo_task
            if error != '{"recs": ""}':
                return error,msg,show_modal
            else:
                #find the new STM_AUTO_KEY that was created during the turn in
                #ctrl_id,ctrl_number,stock_line,description,si_number
                query = """SELECT S.STM_AUTO_KEY,S.CTRL_ID,S.CTRL_NUMBER,S.STOCK_LINE,P.DESCRIPTION,
                            S.QTY_OH,S.QTY_RESERVED,SR.STR_AUTO_KEY,S.SERIES_ID,S.SERIES_NUMBER,
                            STI.WOB_AUTO_KEY
                            FROM STOCK S
                            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY 
                            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY 
                            LEFT JOIN STOCK_TI STI ON STI.STM_AUTO_KEY = S.STM_AUTO_KEY                            
                            WHERE S.PNM_AUTO_KEY = %s 
                            ORDER BY S.STM_AUTO_KEY DESC"""%pnm_auto_key                                                                    
                                                                           
                stm = selection_dir(query,cr)
                stm_auto_key = stm and stm[0] and stm[0][0] or ''  
                ctrl_id = stm and stm[0] and stm[0][1] or 0 
                ctrl_number = stm and stm[0] and stm[0][2] or 0 
                stock_line = stm and stm[0] and stm[0][3] or ''
                description = description or (stm and stm[0] and stm[0][4]) or ''
                str_auto_key = stm and stm[0] and stm[0][7] or ''
                series_id = stm and stm[0] and stm[0][8] or 0 
                series_number = stm and stm[0] and stm[0][9] or 0
                sti_wob_auto_key = stm and stm[0] and stm[0][10] or 0
                record = [activity,quantity,exp_date]
                record += [serial_no,notes,woo_auto_key]
                record += [wo_task,pn,condition]
                record += [ctrl_id,ctrl_number,stock_line]
                record += [description,si_number,pnm_auto_key]
                record += [loc_code,mfg_lot_num,consignment_code]
                record += [stm_auto_key,customer,eng_model,ata_code]
                                                             
                error = synch_new_wob(session_id,record)
                if not ctrl_id:
                    ctrl_id = 1
                if stm_auto_key: 
                    if si_number:
                        if '-' in si_number:
                            wo_number = si_number.split('-')[0]                    
                            query = """SELECT S.SERIAL_NUMBER FROM STOCK S
                                LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
                                LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
                                WHERE UPPER(WO.SI_NUMBER) = UPPER('%s') AND ROWNUM<=1
                                """%wo_number
                            serial = selection_dir(query,cr)
                            serial_number = serial and serial[0] and serial[0][0] or serial_number
                    query = """UPDATE STOCK SET RECEIVER_NUMBER = '%s',PCC_AUTO_KEY = '%s',CNC_AUTO_KEY = '%s',LOC_AUTO_KEY = '%s',STC_AUTO_KEY = '%s',CMP_AUTO_KEY = '%s',CTS_AUTO_KEY = '%s',WHS_AUTO_KEY = '%s',SYSCM_AUTO_KEY = '%s',STM_ORIGINAL = '%s',LOT_APL_RO_COST = '%s',LOT_ALW_PRECOST = '%s',LOT_COST_DELAYED = '%s',ORIGINAL_PO_NUMBER = '%s',SYSUR_AUTO_KEY = '%s',PNM_AUTO_KEY = '%s',SERIAL_NUMBER = '%s',NOTES = TO_CLOB('%s'),OWNER = (SELECT COMPANY_NAME FROM SYS_COMPANIES WHERE SYSCM_AUTO_KEY = '%s'),INCIDENT_RELATED_FLAG = 'F',IC_UDF_008 = '%s',CTRL_ID = '%s',STM_LOT = '%s'%s WHERE STM_AUTO_KEY = %s"""%(\
                        receiver_number,pcc_auto_key,cnc_auto_key,loc_auto_key,stc_auto_key,\
                        cmp_auto_key,cts_auto_key,whs_auto_key,syscm_auto_key,stm_parent_key,\
                        lot_apl_ro_cost,lot_alw_precost,lot_cost_delayed,original_po_number,sysur_auto_key,\
                        pnm_auto_key,serial_no,notes,syscm_auto_key,ex_esn,ctrl_id,\
                        stm_parent_key,update_clause,stm_auto_key)

                    stk_error = updation_dir(query,cr)
                    if stk_error != '{"recs": ""}':
                        stk_error = 'Line: ' + str(line_count) + '-:>' + stk_error
                        return stk_error,msg,'done'                  
                query = """
                    CREATE OR REPLACE
                    PROCEDURE "TI_STOCK_UPDATE"
                    (QUSER IN NUMBER, STM IN NUMBER, QCODE IN VARCHAR2)  AS
                    V_CURSOR QC_SC_PKG.CURSOR_TYPE ;
                    BEGIN                
                        begin
                        qc_trig_pkg.disable_triggers;
                        UPDATE STOCK_TI SET STM_PARENT = %s WHERE STI_AUTO_KEY = (SELECT MAX(STI_AUTO_KEY) FROM STOCK_TI WHERE STM_AUTO_KEY = %s);
                        UPDATE SA_LOG SET SYSUR_AUTO_KEY = QUSER, EMPLOYEE_CODE = QCODE WHERE STA_AUTO_KEY = (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = STM AND EMPLOYEE_CODE = 'DBA');
                        qc_trig_pkg.enable_triggers;
                        end;
                     END TI_STOCK_UPDATE;"""%(stm_parent_key,stm_auto_key)
                                   
                sti_error = updation_dir(query,cr)
                if sti_error != '{"recs": ""}':
                    sti_error = 'Line: ' + str(line_count) + '-:>' + sti_error
                    return sti_error,msg,'done'  
                run_proc = """
                    BEGIN
                    TI_STOCK_UPDATE('%s',%s,'%s');
                    END;   
                """%(sysur_auto_key,stm_auto_key,user_id[:9])
                #error = updation_dir(run_proc,cr)                        
                #make a new reservation:
                update_sa = """DECLARE
                               BEGIN
                               QC_CNTX_PKG.create_sa(QC_CNTX_PKG.CV_WO_RES);
                               QC_CNTX_PKG.push_sa('%s','WOO_AUTO_KEY','%s','SI_NUMBER');
                               END;"""%(woo_auto_key,si_number)
                sa_error = updation_dir(update_sa,cr)
                if sa_error != '{"recs": ""}':
                    sa_error = 'Line: ' + str(line_count) + '-:>' + sa_error
                    return sa_error,msg,'done'                 
                record = [activity,quantity,exp_date]
                record += [serial_no,notes,woo_auto_key]
                record += [wo_task,pn,condition]
                record += [ctrl_id,ctrl_number,stock_line]
                record += [description,si_number,pnm_auto_key]
                record += [loc_code,mfg_lot_num,consignment_code]
                record += [stm_auto_key,customer,eng_model,ata_code]
                                                             
                synch_error = synch_new_wob(session_id,record)                                   
                """if stm_auto_key:
                    squery = """"""UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
                        (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')""""""%(sysur_auto_key,user_id[:9],stm_auto_key)
                    error = updation_dir(squery,cr)"""
                                                                                               
                squery = """UPDATE (SELECT SYSUR_AUTO_KEY FROM AUDIT_TRAIL WHERE SYSUR_AUTO_KEY = 1 AND SOURCE_TABLE IN ('WOB','STM','QUANTUM'))
                SET SYSUR_AUTO_KEY=%s"""%(sysur_auto_key)
                adt_error = updation_dir(squery,cr)
                if adt_error != '{"recs": ""}':
                    adt_error = 'Line: ' + str(line_count) + '-:>' + adt_error
                    return adt_error,msg,'done'                 
    aud_status = 'success'
    if error and error != '{"recs": ""}':               
        aud_status = 'failure'
    elif error == '{"recs": ""}' or error == '':
        error=''
        msg = 'Successful Teardown.'
    orcl_commit(con=con)
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='lot-teardown')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec and not lot_import:
        field_changed = 'quantity | part number | notes | serial number'
        if not error:      
            new_val = new_wob_key and 'New BoM with wob_auto_key: '
            new_val += str(new_wob_key) + ' and Sub-woo: '
            new_val += str(si_num) + ' have been created for parent WO/task#' 
            new_val += str(si_number) + '/' + str(wo_task)   
        else:
            new_val = error 
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)                                             
    return error,msg,show_modal
  
@shared_task
def teardown_rows(quapi_id,sysur_auto_key,user_id,row,\
    new_wo,is_default_repair,stm_serials=[],serial_notes=[]):
    error,msg,err_stock,err_woo,qty_oh,si_num,new_wob_key,si_number = '','','','',0,'','',''
    lot_core_settings,sti_wob_auto_key,new_woo_key,default_repair = '','','','F'
    esn,serials_print,part_info,condition_code = '','',[],''
    from polls.models import QueryApi,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    today = datetime.now()
    timestamp = today.strftime('%m/%d/%Y %H:%M:%S')
    datestamp = today.strftime('%m/%d/%Y')
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg,'',new_woo_key,default_repair,part_info 
    wo_task = row[0]
    if not (wo_task and wo_task[:-1].isdigit()) and wo_task[-1] in ['S','s'] :
        return 'WO Task is required and must be a number.',msg,'',new_woo_key,default_repair,part_info 
    quantity = row[1]
    if not quantity:
        return 'Quantity is required.',msg,'',new_woo_key,default_repair,part_info 
    try:
        quantity = quantity and float(row[1])
    except Exception as exc:
        return 'Quantity must be a number.',msg,'',new_woo_key,default_repair,part_info
   
    pn = row[2]
    pn = pn and pn.replace("'","''")       
    pn = pn and pn.strip() or ''  
    serial_no = row[3]
    session_id = row[4]
    notes = row[5]
    notes = notes.replace("'", r'')
    notes = notes.replace('"', r'')
    description = row[6]
    sn_quantity = quantity
    pnm_auto_key = row[7]
    condition_code = row[8]
    
    if not pn and not pnm_auto_key:
        return 'Part number is required.',msg,'',new_woo_key,default_repair,part_info 

    if pnm_auto_key:
        where_clause = "P.PNM_AUTO_KEY = %s"%pnm_auto_key
    else:
        where_clause = "UPPER(P.PN) = UPPER('%s')"%pn
        
    query = """SELECT P.SERIALIZED,P.PNM_AUTO_KEY,
      P.ACTIVITY,OPM.DEFAULT_REPAIR,P.DESCRIPTION,
      M.MFG_CODE,P.PN,OPM.OPM_AUTO_KEY
      FROM PARTS_MASTER P 
      LEFT JOIN OPERATION_MASTER OPM ON OPM.PNM_AUTO_KEY = P.PNM_AUTO_KEY
      JOIN MANUFACTURER M ON M.MFG_AUTO_KEY = P.MFG_AUTO_KEY
      WHERE %s AND P.ACTIVE_PART <> 'F'"""%where_clause
      
    stms = selection_dir(query,cr)
    pnms = []
    
    if len(stms) > 1:
        for stm in stms:
            
            pnm_auto_key = stm[1]
            if pnm_auto_key not in pnms:
                pnms.append(pnm_auto_key)
                part_desc = stm[4]
                mfg_code = stm[5]
                part_num = stm[6]
                part_info.append([pnm_auto_key,part_num,part_desc,mfg_code])
            
        return '','','show_pns',new_woo_key,default_repair,part_info
        
    is_serialized = stms and stms[0] and stms[0][0] or None
    is_serialized = (is_serialized == 'T' and 'T') or 'F'   
    pnm_auto_key = stms and stms[0] and stms[0][1] or None
    activity = stms and stms[0] and stms[0][2] or None
    default_repair = stms and stms[0] and stms[0][3] or False

    if activity != 'Repair':
        for ser in stms:
            if ser[3] == 'T':
                activity = 'Repair'
                break
                
    if not pnm_auto_key:
        #raise exception because the part was not found.     
        return '','','show_modal',new_woo_key,default_repair,part_info
        
    elif serial_no and is_serialized == 'F':
        return 'Part not serialized',msg,'',new_woo_key,default_repair,part_info  
        
    #if is_serialized == 'T' and quantity > 1 and len(stms) < quantity:
        #raise exception because the part is serialized but the quantity is greater than 1.
        #error = 'Serialized parts cannot have quantity > 1.'
        #need a pop-up to read in a list 
        #return error,msg,'',new_woo_key,default_repair,part_info
        
    if is_serialized == 'F':
        serial_no = ''
        
    if wo_task[-1] in ['s','S']:    
        woo_sub = """SELECT WO.WOO_AUTO_KEY,WT.WOT_AUTO_KEY,WO.SI_NUMBER,S.STM_AUTO_KEY,
                       S.GLA_AUTO_KEY,WO.GLA_AUTO_KEY,M.MODEL_NUMBER,WO.DPT_AUTO_KEY,
                       WO.CMP_AUTO_KEY,WO.SYSCM_AUTO_KEY,WO.KIT_QTY,WO.OPM_AUTO_KEY,WO.DUE_DATE,
                       WO.ATTENTION,WO.ECD_METHOD,WO.WWT_AUTO_KEY,WO.SVC_AUTO_KEY,
                       WO.NEW_WIP_ACCT,WO.BGS_DEFAULT,WO.CUR_AUTO_KEY,WO.GLA_LABOR,WO.GLA_MISC,
                       WO.LOT_CORE_SETTINGS,S.RECEIVER_NUMBER,WO.PCC_AUTO_KEY,WO.CNC_AUTO_KEY,
                       S.LOC_AUTO_KEY,S.STC_AUTO_KEY,S.CTS_AUTO_KEY,S.WHS_AUTO_KEY,S.STM_ORIGINAL,
                       WO.LOT_APL_RO_COST,WO.LOT_ALW_PRECOST,WO.LOT_REQ_INSPECTION,WO.LOT_COST_DELAYED,
                       S.ORIGINAL_PO_NUMBER,S.SERIAL_NUMBER,WO.COMPANY_REF_NUMBER,S.IC_UDF_008,S.EXP_DATE,
                       PCC.DESCRIPTION,L.LOCATION_CODE,S.MFG_LOT_NUM,CNC.CONSIGNMENT_CODE,
                       CMP.COMPANY_NAME,
                       WO.WO_UDF_001,WO.WO_UDF_002,WO.WO_UDF_003,WO.WO_UDF_004,S.QTY_OH
                       FROM WO_TASK WT
                       LEFT JOIN WO_BOM WOB ON WT.WOT_AUTO_KEY = WOB.WOT_AUTO_KEY
                       LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WT.WOO_AUTO_KEY
                       LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
                       LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
                       LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
                       LEFT JOIN MODEL M ON M.MDL_AUTO_KEY = P.MDL_AUTO_KEY
                       LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
                       LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
                       LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = WO.CNC_AUTO_KEY
                       JOIN COMPANIES CMP ON CMP.CMP_AUTO_KEY = WO.CMP_AUTO_KEY
                       WHERE WT.WOT_AUTO_KEY = '%s'
                       ORDER BY S.STM_AUTO_KEY DESC"""%(wo_task[:-1])
                       
    else:
        woo_sub = """SELECT WO.WOO_AUTO_KEY,'',WO.SI_NUMBER,S.STM_AUTO_KEY,
                       S.GLA_AUTO_KEY,WO.GLA_AUTO_KEY,M.MODEL_NUMBER,WO.DPT_AUTO_KEY,
                       WO.CMP_AUTO_KEY,WO.SYSCM_AUTO_KEY,WO.KIT_QTY,WO.OPM_AUTO_KEY,WO.DUE_DATE,
                       WO.ATTENTION,WO.ECD_METHOD,WO.WWT_AUTO_KEY,WO.SVC_AUTO_KEY,
                       WO.NEW_WIP_ACCT,WO.BGS_DEFAULT,WO.CUR_AUTO_KEY,WO.GLA_LABOR,WO.GLA_MISC,
                       WO.LOT_CORE_SETTINGS,S.RECEIVER_NUMBER,WO.PCC_AUTO_KEY,WO.CNC_AUTO_KEY,
                       S.LOC_AUTO_KEY,S.STC_AUTO_KEY,S.CTS_AUTO_KEY,S.WHS_AUTO_KEY,S.STM_ORIGINAL,
                       WO.LOT_APL_RO_COST,WO.LOT_ALW_PRECOST,WO.LOT_REQ_INSPECTION,WO.LOT_COST_DELAYED,
                       S.ORIGINAL_PO_NUMBER,S.SERIAL_NUMBER,WO.COMPANY_REF_NUMBER,S.IC_UDF_008,S.EXP_DATE,
                       PCC.DESCRIPTION,L.LOCATION_CODE,S.MFG_LOT_NUM,CNC.CONSIGNMENT_CODE,
                       CMP.COMPANY_NAME,
                       WO.WO_UDF_001,WO.WO_UDF_002,WO.WO_UDF_003,WO.WO_UDF_004,S.QTY_OH
                       FROM WO_OPERATION WO
                       LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
                       LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
                       LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
                       LEFT JOIN MODEL M ON M.MDL_AUTO_KEY = P.MDL_AUTO_KEY
                       LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
                       LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
                       LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = WO.CNC_AUTO_KEY
                       JOIN COMPANIES CMP ON CMP.CMP_AUTO_KEY = WO.CMP_AUTO_KEY
                       WHERE UPPER(WO.SI_NUMBER) = UPPER('%s')
                       AND WO.OPEN_FLAG = 'T'                       
                       ORDER BY S.STM_AUTO_KEY DESC"""%(wo_task)
                       # Lot WO's only - AND WO.WO_TYPE = 'Lot'
    wot_data = selection_dir(woo_sub,cr)
    if not wot_data:
        error = 'No records found.'
    else:
        woo_auto_key = wot_data[0][0] or ''
        wot_auto_key = wot_data[0][1] or ''
        si_number = wot_data[0][2] or '' 
        stm_auto_key = wot_data[0][3] or ''
        stm_gla_key = wot_data[0][4] or ''
        woo_gla_key = wot_data[0][5] or ''
        pnm_model = wot_data[0][6] or ''
        dpt_auto_key = wot_data[0][7] or ''
        cmp_auto_key = wot_data[0][8] or ''
        syscm_auto_key = wot_data[0][9] or ''
        kit_qty = wot_data[0][10] or ''
        opm_auto_key = wot_data[0][11] or ''
        due_date = wot_data[0][12] or ''
        attention = wot_data[0][13] or ''
        ecd_method = wot_data[0][14] or ''
        wwt_auto_key = wot_data[0][15] or ''
        svc_auto_key = wot_data[0][16] or ''
        new_wip_acct = wot_data[0][17] or ''
        bgs_default = wot_data[0][18] or ''
        cur_auto_key = wot_data[0][19] or ''
        gla_labor = wot_data[0][20] or ''
        gla_misc = wot_data[0][21] or ''
        lot_core_settings = wot_data[0][22] or ''
        receiver_number = wot_data[0][23] or ''
        pcc_auto_key = wot_data[0][24] or ''
        cnc_auto_key = wot_data[0][25] or ''    
        loc_auto_key = wot_data[0][26] or ''
        stc_auto_key = wot_data[0][27] or ''
        cts_auto_key = wot_data[0][28] or ''
        whs_auto_key = wot_data[0][29] or ''
        stm_parent = wot_data[0][30] or ''
        lot_apl_ro_cost = wot_data[0][31] or ''
        lot_alw_precost = wot_data[0][32] or ''
        lot_req_inspection = wot_data[0][33] or ''
        lot_cost_delayed = wot_data[0][34] or ''
        original_po_number = wot_data[0][35] or ''
        serial_number = wot_data[0][36] or ''
        company_ref_number = wot_data[0][37] or ''
        ex_esn = wot_data[0][38] or serial_number
        exp_date = wot_data[0][39] or ''
        #condition_code = wot_data[0][40] or ''
        loc_code = wot_data[0][41] or ''
        mfg_lot_num = wot_data[0][42] or ''
        consignment_code = wot_data[0][43] or ''
        customer = wot_data[0][44] or ''
        wo_udf_001 = wot_data[0][45] or ''#WO_UDF_001
        eng_model = wot_data[0][46] or ''#WO_UDF_002
        ata_code = wot_data[0][47] or ''#WO_UDF_003
        wo_udf_004 = wot_data[0][48] or ''#WO_UDF_004
        qty_oh = wot_data[0][49] or ''#S.QTY_OH
        
        if quantity > 1 and stms and is_serialized == 'T' and not stm_serials:
        
            #check for multiple stock lines (if serialized) 
            #and if found, send back message
            #to prompt user for serial numbers and then turn in each one            
            #create the stock objects and pass back prompt to user
            from polls.models import WOStatus
            stm_data = []
            count = 0
            while count < quantity:
                count += 1
                if count > quantity:
                    break
                stm_data.append(WOStatus(
                #stm_auto_key = stm[4] or 0,
                #serial_number = stm[5] or 'N/A',
                wo_number = wo_task,
                session_id = session_id,         
                ))
                
            if stm_data:     
                try:
                    delete = WOStatus.objects.filter(session_id=session_id).delete()
                    rec = WOStatus.objects.bulk_create(stm_data) or []    
                except Exception as exc:
                    error += "Error, %s, with creating stms locally. %s"%(exc,row[0])
                return error,msg,'show_serials',new_woo_key,default_repair,part_info 
                
        elif is_serialized and is_serialized == 'T' and not serial_no and not stm_serials:
            #raise exception because the part is serialized.
            error = 'PN must have serial number.'
            
            return error,msg,'',new_woo_key,default_repair,part_info 
        
        if not (stm_gla_key and woo_gla_key):        
            query = "SELECT GLA_AUTO_KEY FROM GL_DISTRIBUTION WHERE GLD_AUTO_KEY = (SELECT GLD_INV FROM SYS_COMPANIES WHERE SYSCM_AUTO_KEY = 1)"
            gla = selection_dir(query,cr)
            gla_auto_key = gla and gla[0] and gla[0][0] or None               
            if gla_auto_key:
                if stm_auto_key:
                    upd_qry = "UPDATE STOCK SET GLA_AUTO_KEY = '%s' WHERE STM_AUTO_KEY = '%s'"%(gla_auto_key,stm_auto_key)
                    err_stock = updation_dir(upd_qry,cr)
                    error = update_stock_audit_log(cr,sysur_auto_key,stm_auto_key,user_id,gla_auto_key)
                if woo_auto_key:
                    upd_qry = "UPDATE WO_OPERATION SET GLA_AUTO_KEY = '%s' WHERE WOO_AUTO_KEY = '%s'"%(gla_auto_key,woo_auto_key)
                    err_woo = updation_dir(upd_qry,cr)
                if stm_auto_key and err_stock != '{"recs": ""}' or woo_auto_key and err_woo != '{"recs": ""}':
                    #raise exception because the part is serialized.
                    error = 'Unable to assign GL Account.'
                    return error,msg,'',new_woo_key,default_repair,part_info
                else:
                    error = ''
        if not (wot_auto_key or woo_auto_key):
            #task or WO not found
            error = 'Task or lot not found.'
            return error,msg,'',new_woo_key,default_repair,part_info
            
                              
        cond_sub = """SELECT COND_LEVEL,PCC_AUTO_KEY,CONDITION_CODE
        FROM PART_CONDITION_CODES WHERE PCC_AUTO_KEY = 
        (SELECT PCC_MAIN_OUT FROM WO_CONTROL)"""
        pcc_data = selection_dir(cond_sub,cr)
        pcc_auto_key = pcc_data and pcc_data[0] and pcc_data[0][1] or 1
        pcc_main_out = pcc_auto_key
        cond_level = pcc_data and pcc_data[0] and pcc_data[0][0] or 0
        condition_out = pcc_data and pcc_data[0] and pcc_data[0][2] or ''
            
        if condition_code:
            cond_sub = """SELECT COND_LEVEL,PCC_AUTO_KEY,DESCRIPTION 
            FROM PART_CONDITION_CODES WHERE UPPER(CONDITION_CODE) = 
            UPPER('%s')"""%condition_code
            pcc_data = selection_dir(cond_sub,cr)
            pcc_auto_key = pcc_data and pcc_data[0] and pcc_data[0][1] or 1
            cond_level = pcc_data and pcc_data[0] and pcc_data[0][0] or 0
            condition_code = pcc_data and pcc_data[0] and pcc_data[0][2] or ''
            
        elif wot_data:
            condition_code = wot_data[0][40] or ''
            
        ins_qty = 0
        """
        need a change on the WOB Insert portion with a way to comment in or out depending on customer preference:

        lookup WOB for that WOO. if there is another PNM that matches something already in the WO BOM with 
        WOB['ACTIVITY'] = 'Work Order' then perform the turn-in (STI insert) from that WOB
        QTY_TURN == 0
        
        If not, then insert a new WOB.
        """  
        new_wob_key = None
        #FOR LANDING GEAR SHOPS
        """if wo_task[-1] in ['s','S']:
            where_c = "AND WOB.WOT_AUTO_KEY = '%s'"%wot_auto_key
        else:
            where_c = "AND WOB.WOO_AUTO_KEY = '%s'"%woo_auto_key"""
        #query = """SELECT WOB.WOB_AUTO_KEY FROM WO_BOM WOB 
        #    WHERE WOB.PNM_AUTO_KEY = %s 
        #    %s
        #    AND ACTIVITY = 'Work Order' AND QTY_TURN = 0
        #    ORDER BY WOB_AUTO_KEY DESC"""%(pnm_auto_key,where_c)
        #wobs = selection_dir(query,cr)
        #new_wob_key = wobs and wobs[0] and wobs[0][0] or None       
        #end landing gear shops
        if not new_wob_key:
            accumulation = new_wo and 'T' or 'F'
            bos_auto_key = """(SELECT BOS_AUTO_KEY FROM WO_BOM_STATUS
            WHERE BOM_STATUS_CODE = 'PROD')
            """
            
            if wot_auto_key:
                q_wob = """INSERT INTO WO_BOM (ACCUMULATION,WOO_AUTO_KEY,SYSUR_AUTO_KEY,
                WOT_AUTO_KEY,PNM_AUTO_KEY,QTY_NEEDED,ACTIVITY,PCC_AUTO_KEY,
                COND_LEVEL,ENTRY_DATE,NOTES,QTY_TURN,QTY_RESERVED,BOS_AUTO_KEY) 
                VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s',
                TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),TO_CLOB('%s'),'%s','%s',
                %s)
                """%(accumulation,woo_auto_key,sysur_auto_key,wot_auto_key,\
                pnm_auto_key,quantity,'Work Order',pcc_main_out,cond_level,\
                timestamp,notes,ins_qty,ins_qty,bos_auto_key)
            else:
                query = """SELECT WT.WOT_AUTO_KEY FROM WO_TASK WT WHERE  
                            WT.WOO_AUTO_KEY = %s                
                            AND (WT.WOS_AUTO_KEY NOT IN 
                           (SELECT WOS_AUTO_KEY FROM WO_STATUS 
                           WHERE STATUS_TYPE IN ('Closed','Cancel'))
                           OR WT.WOS_AUTO_KEY IS NULL)
                           ORDER BY WT.SEQUENCE
                        """%woo_auto_key
                wot = selection_dir(query,cr)
                wot_auto_key = wot and wot[0] and wot[0][0] or None
                if wot_auto_key:
                    q_wob = """INSERT INTO WO_BOM (ACCUMULATION,WOO_AUTO_KEY,
                    WOT_AUTO_KEY,SYSUR_AUTO_KEY,
                    PNM_AUTO_KEY,QTY_NEEDED,ACTIVITY,PCC_AUTO_KEY,COND_LEVEL,
                    ENTRY_DATE,NOTES,QTY_TURN,QTY_RESERVED,BOS_AUTO_KEY)
                    VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s',
                    TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),TO_CLOB('%s'),'%s','%s',
                    %s)
                    """%(accumulation,woo_auto_key,wot_auto_key,sysur_auto_key,\
                    pnm_auto_key,quantity,'Work Order',pcc_main_out,cond_level,\
                    timestamp,notes,ins_qty,ins_qty,bos_auto_key)
                else:
                    return 'Must have a task to create new BOM.',msg,'done',new_woo_key,default_repair,part_info                
            error = insertion_dir(q_wob,cr)
            if error:
                return error,'Enter serial numbers.','done',new_woo_key,default_repair,part_info
            query = "SELECT WOB_AUTO_KEY FROM WO_BOM WHERE WOO_AUTO_KEY = '%s' AND PNM_AUTO_KEY = '%s' AND PCC_AUTO_KEY = '%s' ORDER BY WOB_AUTO_KEY DESC"%(woo_auto_key,pnm_auto_key,pcc_main_out)
            new_wob = selection_dir(query,cr)
            new_wob_key = new_wob and new_wob[0] and new_wob[0][0] or None
            
        if not new_wob_key:
            error = 'BOM not created.'
       
        if not error and new_wob_key:
            #cond_sub = "SELECT COND_LEVEL,PCC_AUTO_KEY,DESCRIPTION FROM 
            #PART_CONDITION_CODES WHERE PCC_AUTO_KEY = 
            #(SELECT PCC_TURN_IN FROM WO_CONTROL)"
            #pcc_data = selection_dir(cond_sub,cr)
            #pcc_auto_key = pcc_data and pcc_data[0] and pcc_data[0][1] or 1
            #condition_code = pcc_data and pcc_data[0] and pcc_data[0][2] or ''
            
            if not stm_serials:
                error = stock_turn_in(cr,new_wob_key,woo_auto_key,pnm_auto_key,serial_no,quantity,pcc_auto_key)
                
            else:
                serial_no = ''
                count = 0
                for serial_num in stm_serials:
                    error = stock_turn_in(cr,new_wob_key,woo_auto_key,pnm_auto_key,serial_num,1,pcc_auto_key)
                    serial_no += serial_num + ';'
                    count += 1
                    
            if error == '{"recs": ""}':
                #find the new STM_AUTO_KEY that was created during the turn in
                #ctrl_id,ctrl_number,stock_line,description,si_number
                query = """SELECT S.STM_AUTO_KEY,S.CTRL_ID,S.CTRL_NUMBER,S.STOCK_LINE,P.DESCRIPTION,
                            S.QTY_OH,S.QTY_RESERVED,SR.STR_AUTO_KEY,S.SERIES_ID,S.SERIES_NUMBER,
                            STI.WOB_AUTO_KEY
                            FROM STOCK S
                            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY 
                            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY 
                            LEFT JOIN STOCK_TI STI ON STI.STM_AUTO_KEY = S.STM_AUTO_KEY                            
                            WHERE STI.WOB_AUTO_KEY = %s ORDER BY S.STM_AUTO_KEY DESC"""%(new_wob_key)         
                                                                           
                stock_recs = selection_dir(query,cr)
                count = 0
                
                for stm in stock_recs:
                
                    if stm_serials:
                        serial_number = stm_serials[count]
                    else:
                        serial_number = serial_no
                    
                    if serial_notes:
                        notes = serial_notes[count]
                        
                    count += 1
                    if count == 1:
                        from polls.models import WOStatus
                        delete = WOStatus.objects.filter(session_id=session_id).delete()
                    if count > quantity:
                        break
                    stm_auto_key = stm[0] or ''
                    ctrl_id = stm[1] or 1
                    ctrl_number = stm[2] or '' 
                    stock_line = stm[3] or ''
                    description = description or stm[4] or ''
                    str_auto_key = stm[7] or ''
                    series_id = stm[8] or 0
                    series_number = stm[9] or 0
                    sti_wob_auto_key = stm[10] or 0

                    if stm_auto_key:
                        """if si_number:
                            if '-' in si_number:
                                wo_number = si_number.split('-')[0]                    
                                query = """"""SELECT S.SERIAL_NUMBER,S.IC_UDF_008 FROM STOCK S
                                    LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
                                    LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
                                    WHERE UPPER(WO.SI_NUMBER) = UPPER('%s')
                                    """"""%wo_number
                                serial = selection_dir(query,cr)
                                serial_number = serial and serial[0] and serial[0][0] or serial_number
                                #esn = serial and serial[0] and serial[0][1] or ''"""
                                
                        query = """UPDATE STOCK SET 
                            RECEIVER_NUMBER = '%s',
                            PCC_AUTO_KEY = '%s',
                            CNC_AUTO_KEY = '%s',
                            LOC_AUTO_KEY = '%s',
                            STC_AUTO_KEY = '%s',
                            CMP_AUTO_KEY = '%s',
                            CTS_AUTO_KEY = '%s',
                            WHS_AUTO_KEY = '%s',
                            SYSCM_AUTO_KEY = '%s',
                            STM_ORIGINAL = '%s',
                            LOT_APL_RO_COST = '%s',
                            LOT_ALW_PRECOST = '%s',
                            LOT_COST_DELAYED = '%s',
                            ORIGINAL_PO_NUMBER = '%s',
                            SYSUR_AUTO_KEY = '%s',
                            PNM_AUTO_KEY = '%s',
                            --SERIAL_NUMBER = '',
                            NOTES = TO_CLOB('%s'),
                            OWNER = (SELECT COMPANY_NAME FROM SYS_COMPANIES WHERE SYSCM_AUTO_KEY = '%s'),
                            INCIDENT_RELATED_FLAG = 'F',
                            IC_UDF_008 = '%s',
                            CTRL_ID = %s
                            WHERE STM_AUTO_KEY = %s                         
                        """%(si_number,pcc_auto_key,cnc_auto_key,loc_auto_key,stc_auto_key,
                            cmp_auto_key,cts_auto_key,whs_auto_key,syscm_auto_key,stm_parent,
                            lot_apl_ro_cost,lot_alw_precost,lot_cost_delayed,original_po_number,sysur_auto_key,
                            pnm_auto_key,notes,syscm_auto_key,ex_esn,ctrl_id,stm_auto_key)                       
                        error = updation_dir(query,cr)                  
                    query = """
                        CREATE OR REPLACE
                        PROCEDURE "TI_STOCK_UPDATE"
                        (QUSER IN NUMBER, STM IN NUMBER, QCODE IN VARCHAR2)  AS
                        V_CURSOR QC_SC_PKG.CURSOR_TYPE ;
                        BEGIN                
                            begin
                            qc_trig_pkg.disable_triggers;
                            UPDATE SA_LOG SET SYSUR_AUTO_KEY = QUSER, EMPLOYEE_CODE = QCODE WHERE STA_AUTO_KEY = (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = STM AND EMPLOYEE_CODE = 'DBA');
                            qc_trig_pkg.enable_triggers;
                            end;
                         END TI_STOCK_UPDATE;""" 
                                         
                    #error = updation_dir(query,cr)
                    run_proc = """
                        BEGIN
                        TI_STOCK_UPDATE('%s',%s,'%s');
                        END;   
                    """%(sysur_auto_key,stm_auto_key,user_id[:9]) 
                    error = updation_dir(run_proc,cr)                        
                    #make a new reservation:
                    update_sa = """DECLARE
                                   BEGIN
                                   QC_CNTX_PKG.create_sa(QC_CNTX_PKG.CV_WO_RES);
                                   QC_CNTX_PKG.push_sa('%s','WOO_AUTO_KEY','%s','SI_NUMBER');
                                   END;"""%(woo_auto_key,si_number)
                    error = updation_dir(update_sa,cr) 
                                
                    #if stm_auto_key:
                    #query = "UPDATE WO_BOM SET QTY_RESERVED=%s,QTY_TURN=%s WHERE WOB_AUTO_KEY = %s"%(quantity,quantity,new_wob_key)
                    #error = updation_dir(query,cr)
                    if stm_serials:
                        sn_quantity = 1

                    if not new_wo:
                        if str_auto_key:
                            dquery = "DELETE FROM STOCK_RESERVATIONS WHERE STR_AUTO_KEY = %s"%str_auto_key
                            error = updation(dquery,cr)
                        squery = """INSERT INTO STOCK_RESERVATIONS (STR_AUTO_KEY,SYSUR_AUTO_KEY,STM_AUTO_KEY,WOB_AUTO_KEY,QTY_RESERVED) VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s','%s','%s',%s)"""%(sysur_auto_key,stm_auto_key,new_wob_key,sn_quantity)
                        error = insertion_dir(squery,cr)
                        squery = """UPDATE SA_LOG SET SYSUR_AUTO_KEY=%s, EMPLOYEE_CODE='%s' WHERE STA_AUTO_KEY = 
                        (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = %s AND EMPLOYEE_CODE = 'DBA')"""%(sysur_auto_key,user_id[:9],stm_auto_key)
                        error = updation_dir(squery,cr)
                                                                                               
                    squery = """UPDATE (SELECT SYSUR_AUTO_KEY FROM AUDIT_TRAIL WHERE SYSUR_AUTO_KEY = 1 AND SOURCE_TABLE IN ('WOB','STM','QUANTUM'))
                    SET SYSUR_AUTO_KEY=%s"""%(sysur_auto_key)

                    query = """SELECT SERIAL FROM 
                    VIEW_SPS_WO_OPERATION WHERE SI_NUMBER = '%s'
                    """%si_number
                    ser = selection_dir(query,cr)
                    ex_esn = ser and ser[0] and ser[0][0] or ''
      
                    error = updation_dir(squery,cr)                                                             
                    record = [activity,sn_quantity,exp_date]
                    record += [serial_number,notes,woo_auto_key]
                    record += [wo_task,pn,condition_code]
                    record += [ctrl_id,ctrl_number,stock_line]
                    record += [description,si_number,pnm_auto_key]
                    record += [loc_code,mfg_lot_num,consignment_code]
                    record += [stm_auto_key,customer,eng_model,ata_code]
                    record += [ex_esn,qty_oh,'']
                    error = synch_new_wob(session_id,record,del_existings=False)
                #error = synch_new_wob(session_id,activity,quantity,serial_no,notes,woo_auto_key,wo_task,pn,cond_level,ctrl_id,ctrl_number,stock_line,description,si_number,pnm_auto_key)       
                #if a task was entered by user with an 's' appended
                #Adam - 5/3/22 - disabled the creation of the new sub-woo and the delete on the reservation of the new bom.
                #if sti_wob_auto_key and new_wo:                
                if new_wo:
                    #create new sub-woo and reserve it to the stock line from above
                    #TODO: Resolve error for tasks that are part of sub-woos: 
                    #ORA-20685: Master Job can only be a Work Package, External or Non-Stock WO
                    last_qry ="""SELECT SI_NUMBER FROM WO_OPERATION WHERE SI_NUMBER 
                        LIKE '%s%s' AND SI_NUMBER <> '%s' 
                        ORDER BY WOO_AUTO_KEY DESC,ENTRY_DATE DESC"""%(si_number,'%',si_number)
                    #last_qry = """SELECT VW.WOO_AUTO_KEY,W.SI_NUMBER FROM VIEW_SPS_WO_OPERATION VW 
                    #    LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = VW.WOO_AUTO_KEY
                    #    WHERE VW.PARENT_WO = '%s'"""%si_number         
                    children = selection_dir(last_qry,cr)
                    #now we must sort the children by the after-dash part of their SI_NUMBERs
                    #ch_sorted = sorted(children,key=itemgetter(1))
                    child = children and children[0] and children[0][0] or None
                    #get the dashed part
                    child_dict = {}
                    si_num_split = child and '-' in child and child.split('-') or []               
                    dash_part = si_num_split and si_num_split[-1] or ''
                    if dash_part and dash_part.isnumeric():
                        child_dict[child] = int(dash_part) 
                    #for child in children:
                    #    si_num_split = '-' in child[0] and child[0].split('-') or []               
                    #    dash_part = si_num_split and si_num_split[-1] or None               y
                    #    if dash_part.isnumeric() and child[0] not in child_dict:
                    #        child_dict[child[0]] = int(dash_part)               
                    #max_sub = child_dict and max(child_list) or None                   
                    max_sub = child_dict and max(child_dict, key=lambda key: child_dict[key]) or None
                    if max_sub:
                        max_dash = child_dict[max_sub]
                        incrementer = str(int(max_dash) + 1)                    
                        #grab last part of the si_number and remove it
                        #add '-n' where n is an integer that is one more than the dash_part value
                        num_char = len(str(max_dash))
                        si_num = max_sub[:-num_char] + incrementer
                    else:
                        incrementer = '-001'
                        si_num = si_number + incrementer
                    #split it from the last dash and then take the integer part of that
                    if si_num:
                        query ="""INSERT INTO WO_OPERATION (STM_AUTO_KEY,WOB_AUTO_KEY,WOO_AUTO_KEY,OP_TYPE,WO_TYPE,DOC_NUMBER,ADMIN_TYPE,
                            SI_NUMBER,CMP_AUTO_KEY,WOO_PARENT,PNM_AUTO_KEY,PNM_MODIFY,SYSUR_AUTO_KEY,SYSCM_AUTO_KEY,
                            ENTRY_DATE,dpt_auto_key,kit_qty,opm_auto_key,due_date,attention,ecd_method,wwt_auto_key,
                            svc_auto_key,new_wip_acct,bgs_default,cur_auto_key,gla_labor,gla_misc,lot_core_settings) 
                            VALUES('%s','%s',G_WOO_AUTO_KEY.NEXTVAL,'W','Internal','%s','Standard WO','%s','%s','%s','%s','%s','%s','%s',TO_DATE('%s', 'mm/dd/yyyy'),'%s','%s','%s',TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss'),'%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"""%\
                            (stm_auto_key,new_wob_key,si_num,si_num,cmp_auto_key,'',pnm_auto_key,pnm_auto_key,sysur_auto_key,\
                            syscm_auto_key,datestamp,dpt_auto_key,quantity,opm_auto_key,due_date,attention,ecd_method,\
                            wwt_auto_key,svc_auto_key,new_wip_acct,bgs_default,cur_auto_key,gla_labor,gla_misc,lot_core_settings)
                        error = insertion_dir(query,cr)              
                        sel_qry = """SELECT WO.WOO_AUTO_KEY                     
                            FROM WO_OPERATION WO
                            WHERE WO.SI_NUMBER = '%s'
                            AND WO.SYSUR_AUTO_KEY = %s"""%(si_num,sysur_auto_key)
                        new_woo_key = selection_dir(sel_qry,cr)
                        new_woo_key = new_woo_key and new_woo_key[0] and new_woo_key[0][0] or None
                        #create the new woo record locally and save it + find the OPERATION TASKS associated with it
                        if new_woo_key and stm_auto_key:
                            
                            count = 0
                            for stm in stock_recs:
                                count += 1
                                if count > quantity:
                                    break
                                stm_auto_key = stm[0]
                                #query = "INSERT INTO VIEW_SPS_WO_OPERATION (WOO_AUTO_KEY,PARENT_WO,STM_AUTO_KEY) VALUES('%s','%s','%s')"%(new_woo_key,si_number,stm_auto_key)
                                #error = insertion_dir(query,cr)
                                #find and delete the reservation you just made. Then re-reserve it to the new sub-woo.                    
                                #dquery = "DELETE FROM STOCK_RESERVATIONS WHERE STM_AUTO_KEY = '%s' AND WOB_AUTO_KEY = '%s' AND QTY_RESERVED = %s"%(stm_auto_key,new_wob_key,quantity)                  
                                #error = updation_dir(dquery,cr)
                                
                                squery = """INSERT INTO STOCK_RESERVATIONS (STR_AUTO_KEY,STM_AUTO_KEY,WOO_AUTO_KEY,QTY_RESERVED) VALUES(G_STR_AUTO_KEY.NEXTVAL,'%s', '%s',%s)"""%(stm_auto_key,new_woo_key,sn_quantity)
                                error = insertion_dir(squery,cr)
                            """SELECT OEM_REPAIR FROM OPERATION_MASTER WHERE OPM_AUTO_KEY IN (SELECT OPM_AUTO_KEY FROM WO_REPAIRS WHERE WOO_AUTO_KEY IN (SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER = ))
                               OPERATION_MASTER - WO that is set up ahead of time (template)
                               OPERATION_TASKS - almost exact same tasks as WO_TASKS
                               OPM has OPTs - OPTs have everything we need to create a new WO_TASK
                                  For each new tasks, we insert the sub-woo and pnm from the sub-woo.
                               1. pnm_auto_key from WO
                               2. default_repair = 'T'
                               WOO.WWT_AUTO_KEY needs to populate from the OPM.WWT_AUTO_KEY for the default repair OPM we apply
                            """
                            query = """SELECT OPM.OPM_AUTO_KEY,OPM.OEM_REPAIR,OPM.OPERATION_ID,
                                OPM.EXPIRATION_DATE,WRV.VERSION_ID,OPM.DESCRIPTION,OPM.DEFAULT_REPAIR,
                                OPM.WWT_AUTO_KEY
                                FROM OPERATION_MASTER OPM 
                                JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = OPM.PNM_AUTO_KEY
                                LEFT JOIN WO_REPAIR_VERSION WRV ON WRV.WRV_AUTO_KEY = OPM.WRV_AUTO_KEY                               
                                WHERE P.PNM_AUTO_KEY = %s                           
                            """%(pnm_auto_key)
                            operations = selection_dir(query,cr)
                            opm = operations and operations[0]
                            opm_auto_key = opm and opm[0] or ''
                            oem_repair = opm and opm[1] or ''
                            operation_id = opm and opm[2] or ''
                            exp_date = opm and opm[3] or ''
                            version_id = opm and opm[4] or ''
                            op_desc = opm and opm[5] or ''
                            opm_default_repair = opm and opm[6] or ''
                            wwt_auto_key = opm and opm[7] or ''
                            if default_repair and wwt_auto_key:
                                query="""UPDATE WO_OPERATION 
                                SET WWT_AUTO_KEY='%s'
                                WHERE WOO_AUTO_KEY=%s
                                """%(wwt_auto_key,woo_auto_key)
                                error = updation_dir(query,cr)
                            if not default_repair:
                                #msg += 'Select task operation.'
                                from polls.models import Operation
                                opm_keys = []
                                for oper in operations:
                                    opm_auto_key = oper[0] or ''
                                    if opm_auto_key not in opm_keys:
                                        oem_repair = oper[1] or ''
                                        operation_id = oper[2] or ''
                                        exp_date = oper[3] or ''
                                        version_id = oper[4] or ''
                                        op_desc = oper[5] or ''
                                        opm_default_repair = oper[6]=='F' and False or oper[6]=='T' and True
                                        from polls.models import Operation
                                        operation = Operation.objects.create(
                                            session_id=session_id,
                                            opm_auto_key=opm_auto_key,
                                            operation_id=operation_id,
                                            exp_date=exp_date or None,
                                            version=version_id,
                                            part_number=pn,
                                            part_desc=description,
                                            op_desc=op_desc,
                                            default_repair=opm_default_repair,                                    
                                            )
                                        operation.save()
                                        opm_keys.append(opm_auto_key)
                            from polls.models import WOStatus
                            woo = WOStatus.objects.create(
                                session_id=session_id,
                                quantity = quantity,
                                serial_number = serial_number,
                                ctrl_id = ctrl_id,
                                ctrl_number = ctrl_number,
                                description = description,
                                part_number = pn,
                                woo_auto_key = new_woo_key,
                                si_number = si_num,
                                wo_number = si_num,
                                task_ref = oem_repair,
                                pnm_modify = pnm_model,
                                cust_ref_number = company_ref_number,
                                slug = ex_esn or serial_number,
                            )
                            woo.save()
                            #WORK REQUEST - need to get the WWT_AUTO_KEY from the WO_WORK_TYPE table.
                            #otherwise, we create the tasks based on the opm by feeding to this method                            
                            if default_repair and opm_auto_key:
                                
                                error,msg = create_opm_tasks(
                                    session_id,
                                    sysur_auto_key,
                                    quapi_id,
                                    opm_auto_key,
                                    new_woo_key,
                                    cursor=cr,
                                    datestamp=datestamp
                                    )
                                
                             
    aud_status = 'success'
    if error and error != '{"recs": ""}':               
        aud_status = 'failure'
    elif error == '{"recs": ""}':
        error=''
    orcl_commit(con=con)
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='teardown')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        field_changed = 'quantity | part number | notes | serial number'
        if not error:      
            new_val = new_wob_key and 'New BoM with wob_auto_key: '
            new_val += str(new_wob_key) + ' and Sub-woo: '
            new_val += str(si_num) + ' have been created for parent WO/task#' 
            new_val += str(si_number) + '/' + str(wo_task)   
        else:
            new_val = error 
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)        
    return error,msg,'done',new_woo_key,default_repair,part_info

@shared_task    
def create_opm_tasks(session_id,sysur_auto_key,quapi_id,opm_auto_key,new_woo_key,cursor=None,datestamp=None):
    if not cursor:
        from polls.models import QueryApi,OracleConnection as oc
        quapi = QueryApi.objects.filter(id=quapi_id)
        quapi = quapi and quapi[0] or None
        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
        orcl_conn = orcl_conn and orcl_conn[0] or None
        if orcl_conn:
            cursor,con = orcl_connect(orcl_conn)    
    if not datestamp:
        today = datetime.now()
        timestamp = today.strftime('%m/%d/%Y hh:mm:ss')
        datestamp = today.strftime('%m/%d/%Y')
    error,msg,wtm_auto_key,tasks = '','',None,[]
    #Lookup all OPT's that share that OPM 
    query = """SELECT OPT.OPT_AUTO_KEY,OPT.OPM_AUTO_KEY,
        OPT.WTM_AUTO_KEY,OPT.SQUAWK_DESC,OPT.SEQUENCE,
        OPT.LONG_DESCR,OPT.CLOSING_REQUIREMENTS,
        OPT.PSR_AUTO_KEY,WTM.DESCRIPTION
        FROM OPERATION_TASKS OPT
        JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = OPT.WTM_AUTO_KEY
        WHERE OPT.OPM_AUTO_KEY = %s"""%opm_auto_key
    opts = selection_dir(query,cursor)
    
    query = """SELECT MAX(WOT_AUTO_KEY) + 1 FROM WO_TASK"""
    wot_key = selection_dir(query,cursor)
    wot_auto_key = wot_key and wot_key[0] and wot_key[0][0] or 0
    wot_auto_key = int(wot_auto_key)
    wot_data = []
    from polls.models import WOTask
    for opt in opts:
        #create new task from each opt
        ref = 'Added from Teardown APP - MROLive'
        opt_auto_key = opt[0]
        opm_auto_key = opt[1]
        wtm_auto_key = opt[2]
        squawk_desc = opt[3]
        sequence = opt[4]
        long_descr = opt[5]
        closing_reqs = opt[6]
        psr_auto_key = opt[7]
        task_master_desc = opt[8]
        query = """INSERT INTO WO_TASK 
        (OPT_AUTO_KEY,OPM_AUTO_KEY,REF,WOO_AUTO_KEY,WTM_AUTO_KEY,
        SEQUENCE,TASK_START,PSR_AUTO_KEY,SYSUR_AUTO_KEY,
        SQUAWK_DESC,CLOSING_REQUIREMENTS)
        VALUES ('%s','%s','%s','%s','%s',%s,
        TO_DATE('%s', 'MM/DD/YYYY'),'%s','%s','%s',SUBSTR('%s',0,40))                             
        """%(opt_auto_key,opm_auto_key,ref,new_woo_key,wtm_auto_key,sequence,\
        datestamp,psr_auto_key,sysur_auto_key,squawk_desc,closing_reqs)
        error = insertion_dir(query,cursor)
        wot_auto_key += 1
        query = """
            SELECT WOT.SQUAWK_DESC, WOT.SEQUENCE, WTM.DESCRIPTION, WOT.WOT_AUTO_KEY
            FROM WO_TASK WOT
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
            LEFT JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WOT.WTM_AUTO_KEY                                    
            WHERE WTM.WTM_AUTO_KEY = %s
            AND WO.WOO_AUTO_KEY = %s                                  
            ORDER BY WOT.WOT_AUTO_KEY DESC
        """%(wtm_auto_key,new_woo_key)
        
        wot_data.append(WOTask(
            session_id=session_id,
            wot_sequence = sequence,
            task_master_desc = task_master_desc,
            wot_description = squawk_desc,
            wot_auto_key = wot_auto_key,                                      
        ))
    if wot_data:     
        try:
            delete = WOTask.objects.filter(session_id=session_id).delete()
            rec = WOTask.objects.bulk_create(wot_data) or []    
        except Exception as exc:
            error += "Error, %s, with creating tasks. %s"%(exc,row[0])
    else:
        error = 'No operations found and therefore no tasks were created.'
    return error,msg
    
def synch_new_wob(session_id,record,del_existings=True):
    from polls.models import WOStatus
    error = ''
    try:
        if del_existings:
            WOStatus.objects.filter(session_id=session_id).delete()
            
        exp_date = record[2] and len(record[2]) > 9 and record[2][:10] or None
        res = WOStatus(
            activity = record[0],
            quantity = record[1],
            exp_date = exp_date,
            serial_number = record[3] or 'N/A',
            notes = record[4],
            woo_auto_key = record[5],
            wo_task = record[6], 
            part_number = record[7],
            condition_code = record[8],
            ctrl_id = record[9],
            ctrl_number = record[10],
            stock_line = record[11],
            description = record[12],
            wo_number = record[13],
            si_number = record[13],
            pnm_auto_key = record[14],
            location_code = record[15],
            spn_code = record[16],
            consignment_code = record[17],
            stm_auto_key = record[18],
            customer = record[19],
            time_loc = record[20],#eng_model
            cart = record[21],
            slug = len(record) > 22 and record[22] or '',
            qty_oh = len(record) > 22 and record[23] or 0,
            status = len(record) > 22 and record[24] or '',
            session_id = session_id,            
            )    
        res.save()
    except Exception as exc:
        error += "Error with creating grid for the new WOB: %s"%(exc)       
    return error       
def template_bad_rows(bad_rows,session_id):
    from polls.models import WOStatus
    req_data,error = [],''
    for row in bad_rows:  
        opm_desc = row.get('Description','')
        #check for existing opm and if not create it
        operation_id = row.get('op_id','')
        opt_seq = row.get('Work Plan ID','')
        opt_squawk_desc = row.get('Source PN','')
        type = row.get('type','')
        if type == 'opm':
            opt_seq = ''
            opm_operation_id = str(row.get('Work Plan ID','')) + '_' + str(row.get('Source PN',''))            
        #opm_desc,opm_operation_id,wtm_desc,opt_seq,opt_squawk_desc,       
        req_data.append(WOStatus(
            wot_sequence = opt_seq,
            wo_task = opt_squawk_desc[:50],
            task_master_desc = opt_squawk_desc[:50],
            task_title = opt_squawk_desc[:50],
            task_ref = operation_id[:50],
            session_id = session_id,
            bulk_imp_error = row.get('error',''),  
            stock_line = row.get('line_count',0),            
        ))
    if req_data:     
        try:
            delete = WOStatus.objects.filter(session_id=session_id).delete()
            rec = WOStatus.objects.bulk_create(req_data) or []    
        except Exception as exc:
            error += "Error, %s, with creating bad rows for grid. %s"%(exc,row[0])
    return error
    
def req_bulk_import(session_id,user_id,recs):
    from polls.models import WOStatus
    req_data,error = [],''
    #SI_NUMBER, EMPLOYEE_CODE,SEQUENCE,TASK,PN,QTY_NEEDED, NOTES,ACTIVITY
    for row in recs:    
        req_data.append(WOStatus(
        wo_number = row[0],
        user_id = row[1],
        wot_sequence = row[2],
        wo_task = row[3],
        part_number = row[4],
        quantity = row[5],
        notes = row[6],
        wo_type = row[7],
        po_number = row[8],
        gate_qty = row[9],
        alt_avail = row[10],
        qty_reserved = row[11],
        session_id = session_id,      
        ))
    if req_data:     
        try:
            delete = WOStatus.objects.filter(session_id=session_id).delete()
            rec = WOStatus.objects.bulk_create(req_data) or []    
        except Exception as exc:
            error += "Error, %s, with creating parts request rows for grid for WO#, %s"%(exc,row[0])
    return error
    
@shared_task
def get_requests_view(quapi_id,session_id,user_id,wot_auto_key,wo_number,part_number):   
    error,msg = '',''  
    from polls.models import QueryApi
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    where_clause = ''
    if wot_auto_key:
        where_clause = " AND WOT_AUTO_KEY = %s"%wot_auto_key[:-1]
    if wo_number:
        where_clause += " AND UPPER(SI_NUMBER) = UPPER('%s')"%wo_number
    if part_number:
        where_clause += " AND UPPER(PN) = UPPER('%s')"%part_number       
    query = """
        SELECT SI_NUMBER,EMPLOYEE_CODE,SEQUENCE,TASK,
        PN,QTY_NEEDED,TO_CHAR(NOTES),ACTIVITY,
        PO_NUMBER,QTY_AVAILABLE,ALT_AVAIL,QTY_RESERVED FROM VIEW_WOBOM        
        WHERE ACTIVITY = 'Consumable'
        AND QTY_NEEDED >= QTY_ISSUED%s"""%where_clause
    recs = selection(query,quapi=quapi)
    if recs:
        error = req_bulk_import(session_id,user_id,recs)
    else:
        error = "No requests found."
    return error,msg
def str_to_int(element):
    #If you expect None to be passed:
    if element is None: 
        return False
    try:
        int(element)
        return int(element)
    except ValueError as error:
        return error
        
@shared_task
def wo_template_measures(quapi_id,sysur_auto_key,session_id):
    error,msg,fail_msg = '','',''  
    from polls.models import QueryApi,Document,OracleConnection as oc,QuantumUser as qu,MLApps as maps
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg,fail_msg,0 
    import_file = Document.objects.filter(session_id=session_id)
    import_file = import_file and import_file[0] or None
    file_path = import_file and os.path.join(import_file.docfile.path) or ''

    from openpyxl import load_workbook
    wb = load_workbook(filename = file_path)
    #sheet = wb['WO_TEMPLATE_TO_EXPORT'] 
    sheet = wb.worksheets[0]
    sheet_rows = sheet.iter_rows()
    row_list = []
    row_vals = [[v.value for v in row] for row in sheet_rows]
    col_headings = row_vals[0]
    count = 0
    for row in row_vals:
        #if it is the first row, then it is the headings
        #assign the headings as keys for each value in row
        if count == 0 or not row[0]:
            count += 1
            continue
        dict_row = {}
        col_count = 0
        for col in col_headings:
            dict_row[col] = row[col_count]
            col_count += 1       
        row_list.append(dict_row)
        count+=1
    if row_list: 
        line_count = 0
        bad_rows = []
        wtp_sub = "SELECT WTP_AUTO_KEY FROM WO_TASK_TYPE WHERE TASK_TYPE='MROL TASK'"
        wtp_auto_key = selection_dir(wtp_sub,cr)
        wtp_auto_key = wtp_auto_key and wtp_auto_key[0] and wtp_auto_key[0][0] or None
        if not wtp_auto_key:        
            wtp_ins = "INSERT INTO WO_TASK_TYPE (TASK_TYPE,DESCRIPTION) VALUES('MROL TASK','MROL TASK')"
            error = insertion_dir(wtp_ins,cr)
            if error:
                error += 'No task type found and problem adding it.'
                return error,msg,fail_msg
            wtp_sub = "SELECT WTP_AUTO_KEY FROM WO_TASK_TYPE WHERE TASK_TYPE='MROL TASK'"
            wtp_auto_key = selection_dir(wtp_sub,cr)
            wtp_auto_key = wtp_auto_key and wtp_auto_key[0] and wtp_auto_key[0][0] or None

        prev_op_id = ''
        prev_seq = 0
        op_id_count = 1
       
        for row in row_list:          
            bad_seq = False
            line_count += 1
            opm_desc = row.get('DESCRIPTION','')
            opm_operation_id = str(row.get('ATA','')) + '_' + str(row.get('MANUAL','')) + '-' + str(op_id_count)          
            wtm_desc = row.get('DEPARTMENT','')
            opt_seq = row.get('NO1','')
            import re
            opt_seq = re.sub(r'[^\d.]+', '', str(opt_seq))
            opt_seq = str_to_int(opt_seq)            
            if not opt_seq:
                opt_seq = 0
            
            if prev_op_id == opm_operation_id:
                if prev_seq > opt_seq:
                    op_id_count += 1
                    opm_operation_id = str(row.get('ATA','')) + '_' + str(row.get('MANUAL','')) + '-' + str(op_id_count) 
            #opm_operation_id += '-' + str(op_id_count)   
            prev_op_id = opm_operation_id
            prev_seq = opt_seq
            if not opt_seq:
                error += 'Sequence must be numeric and non-zero.'
                row['line_count'] = line_count
                bad_rows.append(row)
                continue
            opt_squawk_desc = str(row.get('WORKREQUIRED','')) + ' As per ' + str(row.get('SECTIONN',''))
            opt_squawk_desc += ' ' + str(row.get('TASKS','')) + ' ' + str(row.get('PROCEDURE','')) 
            opt_squawk_desc += ' ' + str(row.get('NOTES',''))
            if not opm_desc:
                row['error'] = 'Missing description.'
                row['line_count'] = line_count
                bad_rows.append(row)
                continue
            else:
                opm_key = get_existing_opm(opm_operation_id,cr)            
                #bad_seq = opm_key and check_bad_sequence(cr,opm_key,opt_seq) or False
                if 0 and bad_seq:
                    row['error'] = 'Task sequence, %s, already exists.'%opt_seq
                    row['line_count'] = line_count
                    bad_rows.append(row)
                    continue            
                if not opm_key:
                    opm_key = create_opm_meas(opm_desc,opm_operation_id,sysur_auto_key,cr,quapi)               
                if opm_key:
                    wtm_key = get_existing_wtm_meas(wtm_desc,cr)
                    wtm_key = wtm_key and wtm_key[0] and wtm_key[0][0] or None
                    #if wtm doesn't exist, then create it.
                    if not wtm_key:
                        wtm_key = create_wtm_meas(wtm_desc,wtp_auto_key,cr)
                        wtm_key = wtm_key and wtm_key[0] and wtm_key[0][0] or None
                    if wtm_key:
                        opt_key = create_opt_meas(wtm_key,opm_key,opt_seq,opt_squawk_desc,cr)
                        opt_key = opt_key and opt_key[0] and opt_key[0][0] or None  
                        if opt_key:
                            #create the task measures
                            #feed the entire set of task measures into the create method.
                            error = create_task_measures(row,opt_key,cr) 
                            if error:
                                row['error'] = error
                                row['line_count'] = line_count
                                bad_rows.append(row)                             
                        else:
                            error = "Problem creating operation task master."
                            row['error'] = error
                            bad_rows.append(row)
                            row['line_count'] = line_count   
                    else:
                        error = "Problem creating or finding task master."
                        row['error'] = error
                        bad_rows.append(row)
                        row['line_count'] = line_count                           
                else:
                    error = ' Problem creating a new or finding an existing operation master record.'
                    row['error'] = error
                    bad_rows.append(row)
                    row['line_count'] = line_count                    
        if len(bad_rows) < len(row_list):                    
            orcl_commit(con=con) 
            error = ''
            msg = 'Successfully imported ' + str(len(row_list) - len(bad_rows)) + ' rows. ' + str(len(bad_rows)) + ' rows rejected.'
        else:
            msg = 'Could not import any rows. Please see rejection reasons in grid.'
        if bad_rows:
            error = template_bad_rows(bad_rows,session_id)  
    #register audit trail record                
    aud_status = 'success'
    app_id = maps.objects.filter(code='wo-template-measures')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        field_changed = msg
        new_val = msg
        if error:             
            aud_status = 'failure'
            new_val = error
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status) 
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,msg,fail_msg
   
def insert_measures(measure,min_measure,max_measure,opt_key,cr):
    error = ''
    import re
    min_measure = re.sub(r'[^\d.]+', '', str(min_measure))
    if min_measure and min_measure.replace('.','',1).isdigit():
        min_measure = min_measure and float(min_measure) or 0
    else:
        error += ' Min, %s, is not a number.'%min_measure
        return error  
    max_measure = re.sub(r'[^\d.]+', '', str(max_measure))
    if max_measure and max_measure.replace('.','',1).isdigit():
        max_measure = max_measure and float(max_measure) or 0 
    else:
        error += ' Max, %s, is not a number.'%max_measure
        return error      
    query = """INSERT INTO TASK_MEASURE (TMS_AUTO_KEY,OPT_AUTO_KEY,MEASURE_DESC,MIN_VALUE,MAX_VALUE) VALUES(G_TMS_AUTO_KEY.NEXTVAL,%s,'%s',%s,%s)"""%(opt_key,measure[:255],min_measure,max_measure)
    error = insertion_dir(query,cr)
    return error
    
def create_task_measures(row,opt_key,cr): 
    #TODO - put check in there for mins and maxes that are non-numeric, like '____' 
    error = ''
    tm_measure1_desc = row.get('DIMENSION1','')
    tm_min1_value = row.get('MIN1','')
    tm_max1_value = row.get('MAX1','')   
    if tm_measure1_desc:
        error = insert_measures(tm_measure1_desc,tm_min1_value,tm_max1_value,opt_key,cr)
        tm_measure2_desc = row.get('DIMENSION2','')
        tm_min2_value = row.get('MIN2','')
        tm_max2_value = row.get('MAX2','')
        if tm_measure2_desc and tm_min2_value and tm_max2_value:
            error = insert_measures(tm_measure2_desc,tm_min2_value,tm_max2_value,opt_key,cr)
            tm_measure3_desc = row.get('DIMENSION3','')
            tm_min3_value = row.get('MIN3','')
            tm_max3_value = row.get('MAX3','')
            if tm_measure3_desc and tm_min3_value and tm_max3_value:
                error = insert_measures(tm_measure3_desc,tm_min3_value,tm_max3_value,opt_key,cr)
                tm_measure4_desc = row.get('DIMENSION4','')
                tm_min4_value = row.get('MIN4','')
                tm_max4_value = row.get('MAX4','')
                if tm_measure4_desc and tm_min4_value and tm_max4_value:
                    error = insert_measures(tm_measure4_desc,tm_min4_value,tm_max4_value,opt_key,cr)
                    tm_measure5_desc = row.get('DIMENSION5','')
                    tm_min5_value = row.get('MIN5','')
                    tm_max5_value = row.get('MAX5','')
                    if tm_measure5_desc and tm_min5_value and tm_max5_value:
                        error = insert_measures(tm_measure5_desc,tm_min5_value,tm_max5_value,opt_key,cr)
                        tm_measure6_desc = row.get('DIMENSION6','')
                        tm_min6_value = row.get('MIN6','')
                        tm_max6_value = row.get('MAX6','')                       
                        if tm_measure6_desc and tm_min6_value and tm_max6_value:
                            error = insert_measures(tm_measure6_desc,tm_min6_value,tm_max6_value,opt_key,cr)
                            tm_measure7_desc = row.get('DIMENSION7','')
                            tm_min7_value = row.get('MIN7','')
                            tm_max7_value = row.get('MAX7','')
                            if tm_measure7_desc and tm_min7_value and tm_max7_value:
                                error = insert_measures(tm_measure7_desc,tm_min7_value,tm_max7_value,opt_key,cr)
                                tm_measure8_desc = row.get('DIMENSION8','')
                                tm_min8_value = row.get('MIN8','')
                                tm_max8_value = row.get('MAX8','')
                                if tm_measure8_desc and tm_min8_value and tm_max8_value:
                                    error = insert_measures(tm_measure8_desc,tm_min8_value,tm_max8_value,opt_key,cr)
                                    tm_measure9_desc = row.get('DIMENSION9','')
                                    tm_min9_value = row.get('MIN9','')
                                    tm_max9_value = row.get('MAX9','')
                                    if tm_measure9_desc and tm_min9_value and tm_max9_value:
                                        error = insert_measures(tm_measure9_desc,tm_min9_value,tm_max9_value,opt_key,cr)                                    
                                        tm_measure10_desc = row.get('DIMENSION10','')
                                        tm_min10_value = row.get('MIN10','')
                                        tm_max10_value = row.get('MAX10','')
                                        if tm_measure10_desc and tm_min10_value and tm_max10_value:
                                            error = insert_measures(tm_measure10_desc,tm_min10_value,tm_max10_value,opt_key,cr)
                                            tm_measure11_desc = row.get('DIMENSION11','')
                                            tm_min11_value = row.get('MIN11','')
                                            tm_max11_value = row.get('MAX11','')
                                            if tm_measure11_desc and tm_min11_value and tm_max11_value:
                                                error = insert_measures(tm_measure11_desc,tm_min11_value,tm_max11_value,opt_key,cr)
                                                tm_measure12_desc = row.get('DIMENSION12','')
                                                tm_min12_value = row.get('MIN12','')
                                                tm_max12_value = row.get('MAX12','')
                                                if tm_measure12_desc and tm_min12_value and tm_max12_value:
                                                    error = insert_measures(tm_measure12_desc,tm_min12_value,tm_max12_value,opt_key,cr)
                                                    tm_measure13_desc = row.get('DIMENSION13','')
                                                    tm_min13_value = row.get('MIN13','')
                                                    tm_max13_value = row.get('MAX13','')
                                                    if tm_measure13_desc and tm_min13_value and tm_max13_value:
                                                        error = insert_measures(tm_measure13_desc,tm_min13_value,tm_max13_value,opt_key,cr)                                    
                                                        tm_measure14_desc = row.get('DIMENSION14','')
                                                        tm_min14_value = row.get('MIN14','')
                                                        tm_max14_value = row.get('MAX14','')
                                                        if tm_measure14_desc and tm_min14_value and tm_max14_value:
                                                            error = insert_measures(tm_measure14_desc,tm_min14_value,tm_max14_value,opt_key,cr)
    return error

def get_existing_opm_meas(opm_desc,cr):
    rec = []
    query = "SELECT OPM_AUTO_KEY FROM OPERATION_MASTER WHERE DESCRIPTION = '%s'"%opm_desc[:50]
    try:
        rec = selection_dir(query,cr)
    except Exception as e:
        print(e.args)
        print(e)
    return rec
    
def get_existing_wtm_meas(wtm_desc,cr):
    rec = []
    query = "SELECT WTM_AUTO_KEY FROM WO_TASK_MASTER WHERE DESCRIPTION = '%s'"%wtm_desc[:50]
    try:
        rec = selection_dir(query,cr)
    except Exception as e:
        print(e.args)
        print(e)
    return rec

def create_opm_meas(opm_desc,opm_operation_id,sysur_auto_key,cr,quapi):
    #NOTE: Template type fixed at W
    opm_key = None
    query = """INSERT INTO OPERATION_MASTER (OPM_AUTO_KEY,SYSUR_AUTO_KEY,OPERATION_ID,TEMPLATE_TYPE,DESCRIPTION) VALUES(G_OPM_AUTO_KEY.NEXTVAL,'%s',SUBSTR('%s',0,50),'W',SUBSTR('%s',0,50))"""\
        %(sysur_auto_key,opm_operation_id,opm_desc)
    error = insertion_dir(query,cr)
    error = orcl_commit(quapi=quapi)
    if not error:
        query = "SELECT OPM_AUTO_KEY FROM OPERATION_MASTER WHERE OPERATION_ID='%s'"%opm_operation_id
        query += " AND TEMPLATE_TYPE='W' AND DESCRIPTION='%s'"%opm_desc[:50]
        opm_key = selection_dir(query,cr)
        opm_key = opm_key and opm_key[0] and opm_key[0][0] or None
    return opm_key
    
def create_opt_meas(wtm_key,opm_key,opt_seq,opt_squawk_desc,cr):
    #NOTE: Template type fixed at 'W'
    #added insert for part_cost = 0.00
    opt_key = None
    query = """INSERT INTO OPERATION_TASKS (OPT_AUTO_KEY,OPM_AUTO_KEY,WTM_AUTO_KEY,SQUAWK_DESC,SEQUENCE,PARTS_PRICE,LONG_DESCR) VALUES(G_OPT_AUTO_KEY.NEXTVAL,'%s','%s',SUBSTR('%s',0,255),%s,0.00,SUBSTR('%s',0,255))"""\
        %(opm_key,wtm_key,opt_squawk_desc,opt_seq,opt_squawk_desc)  
    error = insertion_dir(query,cr)
    if not error:
        query = "SELECT OPT_AUTO_KEY FROM OPERATION_TASKS WHERE OPM_AUTO_KEY='%s'"%opm_key
        query += " AND WTM_AUTO_KEY = '%s'"%(wtm_key)
        opt_key = selection_dir(query,cr)
    return opt_key
    
def create_wtm_meas(wtm_desc,wtp_auto_key,cr):
    error,wtm_key = '',None
    query = """INSERT INTO WO_TASK_MASTER (WTM_AUTO_KEY,WTP_AUTO_KEY,DESCRIPTION,LONG_DESCR) 
    VALUES(G_WTM_AUTO_KEY.NEXTVAL,'%s',SUBSTR('%s',0,50),'%s')"""%(wtp_auto_key,wtm_desc[:50],wtm_desc[:255])
    error = insertion_dir(query,cr)
    if not error:
        query = "SELECT WTM_AUTO_KEY FROM WO_TASK_MASTER WHERE ROWNUM<=1 ORDER BY WTM_AUTO_KEY DESC"
        wtm_key = selection_dir(query,cr)
    return wtm_key
    
def model_create_entry(model,row,error,session_id):
    model(
        location_code = row['location_code'],
        location_name = row['location_name'],
        whs_code = row['warehouse_code'],
        whs_name = row['warehouse_name'],
        bulk_imp_error = error,
        session_id = session_id,
    )
    return model
    
@shared_task
def wo_template_create(quapi_id,sysur_auto_key,session_id): 
    error,msg,fail_msg,cr,got_there = '','','',None,False  
    good_rows = {}
    from polls.models import QueryApi,Document,OracleConnection as oc,MLApps as maps,QuantumUser as qu
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg,fail_msg,0 
    import_file = Document.objects.filter(session_id=session_id)
    import_file = import_file and import_file[0] or None
    file_path = import_file and os.path.join(import_file.docfile.path) or ''

    from openpyxl import load_workbook
    wb = load_workbook(filename = file_path)
    sheet = wb['Resultado'] 
    sheet_rows = sheet.iter_rows()
    row_list = []
    row_vals = [[v.value for v in row] for row in sheet_rows]
    col_headings = row_vals[0]
    
    count = 0
    for row in row_vals:
        #if it is the first row, then it is the headings
        #assign the headings as keys for each value in row
        if count == 0 or not row[0]:
            count += 1
            continue
        dict_row = {}
        col_count = 0
        for col in col_headings:
            dict_row[col] = row[col_count]
            col_count += 1       
        row_list.append(dict_row)
        count+=1
        #1. Update the fields mapping because the names have changed
            # Work Plan ID		Source PN (outgoing PN)	Material / Hardness	Detail Number (incoming PN)	CMM
            
        #2. Update the code to go row by row and search/create PNM's and OPM's
        #3. Code to detect when the OPTs start and do the search/create for OPT's as well.
        #4. Code to search/create PNMs but need this data first:
          #-must search/create MANUFACTURER['MFG_CODE']	
          #-must search/create APPLICATION_CODES['APPLICATION_CODE']    
    if row_list: 
        line_count = 0
        bad_rows = []
        wtp_sub = "SELECT WTP_AUTO_KEY FROM WO_TASK_TYPE WHERE TASK_TYPE='MROL TASK'"
        wtp_auto_key = selection_dir(wtp_sub,cr)
        wtp_auto_key = wtp_auto_key and wtp_auto_key[0] and wtp_auto_key[0][0] or None
        if not wtp_auto_key:        
            wtp_ins = "INSERT INTO WO_TASK_TYPE (TASK_TYPE,DESCRIPTION) VALUES('MROL TASK','MROL TASK')"
            error = insertion_dir(wtp_ins,cr)
            if error:
                error += 'No task type found and problem adding it.'
                return error,msg,fail_msg
            wtp_sub = "SELECT WTP_AUTO_KEY FROM WO_TASK_TYPE WHERE TASK_TYPE='MROL TASK'"
            wtp_auto_key = selection_dir(wtp_sub,cr)
            wtp_auto_key = wtp_auto_key and wtp_auto_key[0] and wtp_auto_key[0][0] or None 
        #loop through each row and process
        position = 'Initial'           
        for row in row_list:                                   
            bad_seq = False
            line_count += 1  
            work_plan = row.get('Work Plan ID','') 
            source_pn = row.get('Source PN','') 
            opm_desc = row.get('Description','')             
            if not work_plan:
                if line_count == 1:
                    trow['error'] = "No data to import." 
                    trow['line_count'] = line_count
                    break
                else:
                    trow['error'] = "No data to import." 
                    trow['line_count'] = line_count
                    bad_rows.append(trow)                
                    continue
            #if we are in the OPM part of the sheet:           
            if position != 'Task Code' and not got_there:                                                     
                opm_operation_id = str(work_plan) + '_' + str(source_pn)            
                if not opm_desc:
                    row['error'] = 'Missing operation description.'
                    row['line_count'] = line_count
                    row['type'] = 'opm'
                    row['op_id'] = opm_operation_id
                    bad_rows.append(row)
                    continue
                elif opm_desc == 'Task Code':
                    position = 'Task Code'
                    got_there = True
                    continue                              
                #check for existing opm and if not create it.                
                opm_key = get_existing_opm(opm_operation_id,cr)
                #if no opm_key was found, we have to create a new one.
                if not opm_key:
                    opm_key = create_opm(opm_desc,opm_operation_id,sysur_auto_key,cr,quapi,row) 
                if not opm_key:
                    error = ' Problem creating a new or finding an existing operation master record.'
                    row['error'] = error
                    row['line_count'] = line_count 
                    row['op_id'] = opm_operation_id
                    row['type'] = 'opm'
                    bad_rows.append(row)
                    continue
                else:    
                    task_pos = False                
                    line = 0                          
                    for trow in row_list:
                        line += 1        
                        opt_code = trow.get('Description','')
                        if not opt_code:
                            trow['error'] = 'Missing task description.'
                            trow['line_count'] = line
                            trow['type'] = 'opt'
                            trow['op_id'] = opm_operation_id
                            bad_rows.append(trow)
                            continue
                        if not task_pos and opt_code == 'Task Code':
                            task_pos = True
                            continue
                        elif task_pos == True:   
                            #Task Sequence	Task Code	Work Required
                            opt_seq = trow.get('Work Plan ID','')
                            opt_description = trow.get('Source PN','')   
                            if not opt_code:
                                trow['error'] = 'Missing task code.'
                                trow['line_count'] = line_count
                                trow['type'] = 'opt'
                                trow['op_id'] = opm_operation_id
                                bad_rows.append(trow)
                                continue                            
                            bad_seq = check_bad_sequence(cr,opm_key,opt_seq)                
                            if bad_seq: 
                                trow['error'] = 'Task sequence, %s, already exists.'%opt_seq
                                trow['line_count'] = line_count
                                trow['type'] = 'opt'
                                trow['op_id'] = opm_operation_id
                                bad_rows.append(trow)
                                continue                         
                            #seq is ok so add the tasks for each opm
                            wtm_key = get_existing_wtm(opt_description,cr)
                            wtm_key = wtm_key and wtm_key[0] and wtm_key[0][0] or None
                            #if wtm doesn't exist, then create it.
                            
                            if not wtm_key:
                                wtm_key = create_wtm(opt_description,wtp_auto_key,cr)
                                wtm_key = wtm_key and wtm_key[0] and wtm_key[0][0] or None
                            if wtm_key:
                                opt_key = create_opt(wtm_key,opm_key,opt_seq,opt_description,cr)
                                opt_key = opt_key and opt_key[0] and opt_key[0][0] or None
                            else:
                                trow['error'] = "Problem creating or finding wo task master." 
                                trow['line_count'] = line_count
                                trow['type'] = 'opt'
                                trow['op_id'] = opm_operation_id
                                bad_rows.append(trow)
                                continue 
                            if wtm_key and not opt_key:
                                trow['error'] = "Problem creating or finding operation task master." 
                                trow['line_count'] = line_count
                                trow['type'] = 'opt'
                                trow['op_id'] = opm_operation_id                                
                                bad_rows.append(trow)
                                continue 
                                
        if len(bad_rows) < len(row_list):                    
            orcl_commit(con=con) 
            error = ''
            msg = 'Successfully imported ' + str(len(row_list) - len(bad_rows)) + ' rows. ' + str(len(bad_rows)) + ' rows rejected.'
        else:
            msg = 'Could not import any rows. Please see rejection reasons in grid.'
        if bad_rows:
            error = template_bad_rows(bad_rows,session_id)    
    #register audit trail record                
    aud_status = 'success'
    app_id = maps.objects.filter(code='wo-template-import')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        field_changed = msg
        new_val = msg
        if error:             
            aud_status = 'failure'
            new_val = error
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,msg,fail_msg
    
def get_existing_opm(opm_operation_id,cr):
    rec=[]
    
    query = """SELECT OPM_AUTO_KEY FROM OPERATION_MASTER WHERE OPERATION_ID = SUBSTR('%s',0,50)"""%opm_operation_id
    try:
        rec = selection_dir(query,cr)
    except Exception as e:
        print(e.args)
        print(e)
    opm_key = rec and rec[0] and rec[0][0] or None
    return opm_key

def check_bad_sequence(cr,opm_key,opt_seq):
    bad_seq,rec = False,[]
    if opm_key:
        query = """SELECT SEQUENCE FROM OPERATION_TASKS WHERE SEQUENCE = %s AND OPM_AUTO_KEY = %s"""%(opt_seq,opm_key)
        try:
            rec = selection_dir(query,cr)
        except Exception as e:
            print(e.args)
            print(e)
        bad_seq = rec and rec[0] and rec[0][0] and True or False
    return bad_seq
    
def get_existing_wtm(wtm_desc,cr):
    rec = []
    query = "SELECT WTM_AUTO_KEY FROM WO_TASK_MASTER WHERE DESCRIPTION = SUBSTR('%s',0,50)"%wtm_desc
    try:
        rec = selection_dir(query,cr)
    except Exception as e:
        print(e.args)
        print(e)
    return rec
    
def get_create_codes(app_code,mfg_code,cr):
    #application code:
    error = ''
    query = "SELECT APC_AUTO_KEY FROM APPLICATION_CODES WHERE APPLICATION_CODE = '%s'"%app_code
    app_codes = selection_dir(query,cr)
    apc_auto_key = app_codes and app_codes[0] and app_codes[0][0] or ''
    if not apc_auto_key:
        #create a new app code
        query = "INSERT INTO APPLICATION_CODES (APPLICATION_CODE,DESCRIPTION) VALUES('%s','%s')"%(app_code,app_code)
        error = insertion_dir(query,cr)
    
    #manufacturer code:
    query = "SELECT MFG_AUTO_KEY FROM MANUFACTURER WHERE MFG_CODE = '%s'"%mfg_code
    mfg_codes = selection_dir(query,cr)
    mfg_auto_key = mfg_codes and mfg_codes[0] and mfg_codes[0][0] or ''

    if not mfg_auto_key:
        #create a new mfg code
        query = "INSERT INTO MANUFACTURER (MFG_CODE,DESCRIPTION) VALUES('%s','%s')"%(mfg_code,mfg_code)
        error = insertion_dir(query,cr)

    if not error:
        if not apc_auto_key:
            #get new app code
            query = "SELECT APC_AUTO_KEY FROM APPLICATION_CODES WHERE APPLICATION_CODE = '%s'"%app_code
            app_codes = selection_dir(query,cr)
            apc_auto_key = app_codes and app_codes[0] and app_codes[0][0] or '' 
            
        if not mfg_auto_key:
            #get new manufacturer code
            query = "SELECT MFG_AUTO_KEY FROM MANUFACTURER WHERE MFG_CODE = '%s'"%mfg_code
            mfg_codes = selection_dir(query,cr)
            mfg_auto_key = mfg_codes and mfg_codes[0] and mfg_codes[0][0] or ''       
    return apc_auto_key,mfg_auto_key
    
def create_opm(opm_desc,opm_operation_id,sysur_auto_key,cr,quapi,row):
    #NOTE: Template type fixed at W
    opm_key,pnm_code,apc_auto_key,mfg_auto_key = None,None,None,None
    pnm_ic_udf_003 = row.get('Material / Hardness','')
    incoming_pnm = row.get('Detail Number','')  
    outgoing_pnm = row.get('Source PN','') 
    pnm_description = row.get('Description','')
    pnm_serialized = row.get('Serialized (Y/N?)','F')
    pnm_serialized = 'Y' and 'T' or 'F'
    pnm_time_life = row.get('Life Limited (Y/N?)','F')
    pnm_time_life = 'Y' and 'T' or 'F'
    pnm_app_code = row.get('Aircraft / Type','')
    pnm_mfg_code = row.get('OEM','')
    #if app code or mfg code is not empty,
    if pnm_app_code or pnm_mfg_code:
        #then create the app and mfg codes if they don't already exist    
        apc_auto_key,mfg_auto_key = get_create_codes(pnm_app_code,pnm_mfg_code,cr)
    #find the pnm_keys that match the pns from the sheet    
    query = """SELECT PNM_AUTO_KEY,PN FROM PARTS_MASTER WHERE PN IN ('%s','%s')"""%(incoming_pnm,outgoing_pnm)
    pnm_keys = selection_dir(query,cr)
    incoming_pnm_key,outgoing_pnm_key=None,None 
    if pnm_keys:
        apc_auto_key,mfg_auto_key = get_create_codes(pnm_app_code,pnm_mfg_code,cr)    
    for pnm in pnm_keys:
        if pnm[1] == incoming_pnm:
            incoming_pnm_key = pnm[0]
        elif pnm[1] == outgoing_pnm:
            outgoing_pnm_key = pnm[0]
    #If we didn't find the incoming part, then we have to create it and get the auto_key       
    if not incoming_pnm_key and incoming_pnm:
        #need to create it
        pnm_code = incoming_pnm
        if pnm_app_code and pnm_mfg_code:
            query = """INSERT INTO PARTS_MASTER (PN,DESCRIPTION,SERIALIZED,TIME_LIFE,APC_AUTO_KEY,MFG_AUTO_KEY,IC_UDF_003) VALUES(SUBSTR('%s',0,40),'%s','%s','%s','%s','%s','%s')"""%(pnm_code,pnm_description,pnm_serialized,pnm_time_life,apc_auto_key,mfg_auto_key,pnm_ic_udf_003)
        if pnm_app_code and not pnm_mfg_code:
            query = """INSERT INTO PARTS_MASTER (PN,DESCRIPTION,SERIALIZED,TIME_LIFE,APC_AUTO_KEY,IC_UDF_003) VALUES(SUBSTR('%s',0,40),'%s','%s','%s','%s','%s')"""%(pnm_code,pnm_description,pnm_serialized,pnm_time_life,apc_auto_key,pnm_ic_udf_003)          
        if not pnm_app_code and pnm_mfg_code:
            query = """INSERT INTO PARTS_MASTER (PN,DESCRIPTION,SERIALIZED,TIME_LIFE,MFG_AUTO_KEY,IC_UDF_003) VALUES(SUBSTR('%s',0,40),'%s','%s','%s','%s','%s')"""%(pnm_code,pnm_description,pnm_serialized,pnm_time_life,mfg_auto_key,pnm_ic_udf_003)       
        if not pnm_app_code and not pnm_mfg_code:
            query = """INSERT INTO PARTS_MASTER (PN,DESCRIPTION,SERIALIZED,TIME_LIFE,IC_UDF_003) VALUES(SUBSTR('%s',0,40),'%s','%s','%s','%s')"""%(pnm_code,pnm_description,pnm_serialized,pnm_time_life,pnm_ic_udf_003)
        error = insertion_dir(query,cr)
        if not error:
            pnm_query = "SELECT PNM_AUTO_KEY FROM PARTS_MASTER WHERE PN=SUBSTR('%s',0,40)"%pnm_code
            pnm = selection_dir(pnm_query,cr)
            incoming_pnm_key = pnm and pnm[0] and pnm[0][0] or None
    #same for the outgoing_pnm_key:        
    if not outgoing_pnm_key and outgoing_pnm:
        query = """SELECT PNM_AUTO_KEY,PN FROM PARTS_MASTER WHERE PN = '%s'"""%outgoing_pnm
        pnm = selection_dir(query,cr)
        outgoing_pnm_key = pnm and pnm[0] and pnm[0][0] or None
        #if we didn't create it for the incoming pnm, we must create it now.
        if not outgoing_pnm_key:
            #need to create it
            pnm_code = outgoing_pnm
            if pnm_app_code and pnm_mfg_code:
                query = """INSERT INTO PARTS_MASTER (PN,DESCRIPTION,SERIALIZED,TIME_LIFE,APC_AUTO_KEY,MFG_AUTO_KEY,IC_UDF_003) VALUES(SUBSTR('%s',0,40),'%s','%s','%s','%s','%s','%s')"""%(pnm_code,pnm_description,pnm_serialized,pnm_time_life,apc_auto_key,mfg_auto_key,pnm_ic_udf_003)
            if pnm_app_code and not pnm_mfg_code:
                query = """INSERT INTO PARTS_MASTER (PN,DESCRIPTION,SERIALIZED,TIME_LIFE,APC_AUTO_KEY,IC_UDF_003) VALUES(SUBSTR('%s',0,40),'%s','%s','%s','%s','%s')"""%(pnm_code,pnm_description,pnm_serialized,pnm_time_life,apc_auto_key,pnm_ic_udf_003)          
            if not pnm_app_code and pnm_mfg_code:
                query = """INSERT INTO PARTS_MASTER (PN,DESCRIPTION,SERIALIZED,TIME_LIFE,MFG_AUTO_KEY,IC_UDF_003) VALUES(SUBSTR('%s',0,40),'%s','%s','%s','%s','%s')"""%(pnm_code,pnm_description,pnm_serialized,pnm_time_life,mfg_auto_key,pnm_ic_udf_003)       
            if not pnm_app_code and not pnm_mfg_code:
                query = """INSERT INTO PARTS_MASTER (PN,DESCRIPTION,SERIALIZED,TIME_LIFE,IC_UDF_003) VALUES(SUBSTR('%s',0,40),'%s','%s','%s','%s')"""%(pnm_code,pnm_description,pnm_serialized,pnm_time_life,pnm_ic_udf_003)
            error = insertion_dir(query,cr)
            if not error:
                pnm_query = "SELECT PNM_AUTO_KEY FROM PARTS_MASTER WHERE PN=SUBSTR('%s',0,40)"%pnm_code
                pnm = selection_dir(pnm_query,cr)
                outgoing_pnm_key = pnm and pnm[0] and pnm[0][0] or None
    #create the OPM       
    #query = """INSERT INTO OPERATION_MASTER (OPM_AUTO_KEY,SYSUR_AUTO_KEY,OPERATION_ID,TEMPLATE_TYPE,DESCRIPTION) VALUES(G_OPM_AUTO_KEY.NEXTVAL,'%s',SUBSTR('%s',0,50),'W',SUBSTR('%s',0,50))"""\
    #adding these values when creatin a new OPM
    opm_oem_repair = str(row.get('CMM','')) + '_' + str(row.get('Revision',''))+ '_' + str(row.get('Date',''))
    opm_block_13_8130 = str(row.get('Notes: (SBs y ADs)',''))
    query = """INSERT INTO OPERATION_MASTER (OPM_AUTO_KEY,OEM_REPAIR,BLOCK_13_8130,PNM_AUTO_KEY,PNM_MODIFY,SYSUR_AUTO_KEY,OPERATION_ID,TEMPLATE_TYPE,DESCRIPTION) VALUES(G_OPM_AUTO_KEY.NEXTVAL,'%s','%s','%s','%s','%s',SUBSTR('%s',0,50),'W',SUBSTR('%s',0,50))"""%(opm_oem_repair,opm_block_13_8130,incoming_pnm_key,outgoing_pnm_key,sysur_auto_key,opm_operation_id,opm_desc)
    error = insertion_dir(query,cr)
    if not error:
        query = "SELECT OPM_AUTO_KEY FROM OPERATION_MASTER WHERE OPERATION_ID=SUBSTR('%s',0,50)"%opm_operation_id
        query += " AND TEMPLATE_TYPE='W' AND DESCRIPTION='%s' ORDER BY OPM_AUTO_KEY DESC"%opm_desc[:50]
        opm_key = selection_dir(query,cr)
        opm_key = opm_key and opm_key[0] and opm_key[0][0] or None
    return opm_key
    
def create_opt(wtm_key,opm_key,opt_seq,opt_squawk_desc,cr):
    #NOTE: Template type fixed at 'W'
    #added insert for part_cost = 0.00
    opt_key = None
    query = """INSERT INTO OPERATION_TASKS (OPT_AUTO_KEY,OPM_AUTO_KEY,WTM_AUTO_KEY,SQUAWK_DESC,SEQUENCE,PARTS_PRICE,LONG_DESCR) VALUES(G_OPT_AUTO_KEY.NEXTVAL,'%s','%s',SUBSTR('%s',0,255),%s,0,'%s')"""\
        %(opm_key,wtm_key,opt_squawk_desc[:255],opt_seq,opt_squawk_desc[:255])   
    error = insertion_dir(query,cr)    
    if not error:
        query = "SELECT OPT_AUTO_KEY FROM OPERATION_TASKS WHERE OPM_AUTO_KEY='%s'"%opm_key
        query += " AND WTM_AUTO_KEY = '%s'"%(wtm_key)
        opt_key = selection_dir(query,cr)
    return opt_key
    
def create_wtm(wtm_desc,wtp_auto_key,cr):
    error,wtm_key = '',None
    query = """INSERT INTO WO_TASK_MASTER (WTM_AUTO_KEY,WTP_AUTO_KEY,DESCRIPTION,LONG_DESCR) 
    VALUES(G_WTM_AUTO_KEY.NEXTVAL,'%s',SUBSTR('%s',0,50),SUBSTR('%s',0,250))"""%(wtp_auto_key,wtm_desc,wtm_desc)
    error = insertion_dir(query,cr)
    if not error:
        query = "SELECT WTM_AUTO_KEY FROM WO_TASK_MASTER WHERE DESCRIPTION=SUBSTR('%s',0,50)"%wtm_desc
        wtm_key = selection_dir(query,cr)
    return wtm_key
          
def model_create_entry(model,row,error,session_id):
    model(
        location_code = row['location_code'],
        location_name = row['location_name'],
        whs_code = row['warehouse_code'],
        whs_name = row['warehouse_name'],
        bulk_imp_error = error,
        session_id = session_id,
    )
    return model

@shared_task
def loc_whs_bulk(quapi_id,session_id,sysur_auto_key,row_list):
    error,msg,fail_msg,bad_rows,loc_data,locations = '','','',0,[],[]
    from polls.models import OracleConnection as oc,QueryApi,WarehouseLocation,MLApps as maps,QuantumUser as qu
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.',fail_msg,bad_rows
    for row in row_list:
        error = ''       
        if not row['warehouse_code']:
            error = 'Must enter a warehouse code. '
            bad_rows += 1
            loc_data=prep_bulk(loc_data,row,error,session_id)         
            #continue
        elif not row['location_code']:
            error = 'Must enter a location code. '
            bad_rows += 1 
            loc_data=prep_bulk(loc_data,row,error,session_id)         
            #continue
        if row['location_code'] in locations:
            error = 'Duplicate location. '
            bad_rows += 1   
            loc_data=prep_bulk(loc_data,row,error,session_id)         
            #continue            
        #create an insert query statement for warehouse (if not found) and then
        #1. check warehouse and location tables to see if either warehouse or location already exists.
        loc_query = "SELECT LOC_AUTO_KEY FROM LOCATION WHERE LOCATION_CODE = '%s'"%row['location_code']
        loc_check = selection_dir(loc_query,cr) 
        new_loc_key = loc_check and loc_check[0] and loc_check[0][0] or None        
        loc_whs_query = """SELECT L.LOC_AUTO_KEY FROM LOCATION L
            LEFT JOIN WAREHOUSE_LOCATIONS WL ON WL.LOC_AUTO_KEY = L.LOC_AUTO_KEY 
            WHERE L.LOCATION_CODE = '%s' AND
            WL.WHS_AUTO_KEY = (SELECT WHS_AUTO_KEY FROM WAREHOUSE WHERE WAREHOUSE_CODE='%s')"""%(row['location_code'],row['warehouse_code'])
        loc_whs_check = selection_dir(loc_whs_query,cr)
        loc_whs_check = loc_whs_check and loc_whs_check[0] or None
        if loc_whs_check:
            bad_rows += 1
            error = 'This location already exists at this warehouse.'
            loc_data=prep_bulk(loc_data,row,error,session_id)         
            #continue
        if not error:
            whs_query = """
                SELECT WHS_AUTO_KEY FROM WAREHOUSE WHERE WAREHOUSE_CODE = '%s'
                """%(row['warehouse_code'])
            whs_key = selection_dir(whs_query,cr)
            whs_key = whs_key and whs_key[0] and whs_key[0][0] or None
            #2. if warehouse exists, then use the whs_auto_key for the location insert if the location doesn't already exist. 
            if not whs_key:
                insert_whs = """
                INSERT INTO WAREHOUSE
                (WAREHOUSE_CODE,DESCRIPTION)
                VALUES ('%s','%s')
                """%(row['warehouse_code'],row['warehouse_name'])
                error += insertion_dir(insert_whs,cr)
                whs_key = "(SELECT WHS_AUTO_KEY FROM WAREHOUSE WHERE WAREHOUSE_CODE = '%s')"%row['warehouse_code'] 
                if error:
                    bad_rows += 1
                    loc_data=prep_bulk(loc_data,row,error,session_id)
                    #continue                                               
            #3. Insert the location with its location_code and location_name
            if not loc_check:            
                insert_loc = """
                    INSERT INTO LOCATION
                    (LOCATION_CODE,DESCRIPTION)
                    VALUES('%s','%s')
                    """%(row['location_code'],row['location_name'])
                error += insertion_dir(insert_loc,cr)
                if error:
                    bad_rows += 1
                    loc_data=prep_bulk(loc_data,row,error,session_id)
                    #continue                                
                new_loc_key = "(SELECT LOC_AUTO_KEY FROM LOCATION WHERE LOCATION_CODE = '%s')"%row['location_code']
            #4. Insert the WAREHOUSE_LOCATIONS record with the new loc_auto_key and the whs_auto_key to tie them together.
            insert_lwh = """INSERT INTO WAREHOUSE_LOCATIONS 
                     (LOC_AUTO_KEY,WHS_AUTO_KEY) 
                     VALUES (%s,%s)        
            """%(new_loc_key,whs_key)
            error += insertion_dir(insert_lwh,cr)
            if error:
                bad_rows += 1
                loc_data=prep_bulk(loc_data,row,error,session_id) 
        if not error:
            con.commit()
            loc_data=prep_bulk(loc_data,row,msg,session_id)
    error = ''   
    if loc_data: 
        WarehouseLocation.objects.filter(session_id=session_id).delete()        
        try:    
            WarehouseLocation.objects.bulk_create(loc_data) or []    
        except Exception as exc:
            error += "Error with creating warehouse/locations locally for display in grid: %s"%(exc)
        msg = 'Successful import of ' + str(len(row_list)-bad_rows) + ' locations.'
        fail_msg = ' ' + str(bad_rows) + ' rows could not be imported as shown below in the grid. '
    #register audit trail record                
    aud_status = 'success'
    app_id = maps.objects.filter(code='loc-whs-import')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        field_changed = msg
        new_val = msg
        if error:             
            aud_status = 'failure'
            new_val = fail_msg + ' | ' + error
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status) 
    else:
        error = 'Incorrect Quantum User ID.'
    return error,fail_msg,bad_rows
    
def prep_bulk(loc_data,row,error,session_id):
    from polls.models import WarehouseLocation
    loc_data.append(WarehouseLocation(
        location_code = row['location_code'],
        location_name = row['location_name'],
        whs_code = row['warehouse_code'],
        whs_name = row['warehouse_name'],
        bulk_imp_error = error,
        session_id = session_id,
    )) 
    return loc_data
    
def task_bulk_create(quapi_id,session_id,sysur_auto_key,bad_rows,wo_number):
    from polls.models import WOStatus
    mbr_data,error = [],''
    for row in bad_rows:
        start_date = row.get('start_date',None)
        input_format = '%m/%d/%Y' 
        new_format = '%Y-%m-%d'        
        start_date = start_date and datetime.strptime(start_date,input_format) or ''
        start_date = start_date and start_date.strftime(new_format) or None 
        est_hours = row.get('skills_est_hours',0)
        if est_hours and est_hours.replace('.','',1).isdigit():
            est_hours = est_hours and float(est_hours) or 0
        else:
            est_hours = 0 
        wot_sequence = row.get('wot_sequence',0)      
        mbr_data.append(WOStatus(
        wo_number = wo_number,
        start_date = start_date,
        wot_sequence = wot_sequence,
        task_master_desc = row.get('task_desc',''),
        task_ref = row.get('task_ref',''),
        task_position = row.get('task_position',''),
        task_title = row.get('task_title',''),
        task_close_reqs = row.get('task_close_reqs',''),
        skill_desc = row.get('skill',''),
        wot_est_hours = est_hours,
        bulk_imp_error = row.get('error',''),
        session_id = session_id,      
        ))
    if mbr_data:            
        try:
            rec = WOStatus.objects.bulk_create(mbr_data) or []    
        except Exception as exc:
            error += "Error with creating unimportable tasks for WO#, %s, from bulk: %s"%(exc,wo_number)
    return error
    
@shared_task
def task_insertion(quapi_id,session_id,sysur_auto_key,task_list,wo_number):
    msg,bad_rows,imported_count,sequence,wot_auto_key = '',[],0,0,'' 
    error,msg,fail_msg,psr_key,bulk_error = '','','','',''
    from polls.models import OracleConnection as oc,QueryApi,WOStatus,QuantumUser as qu,MLApps as maps
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg,fail_msg,0,0  
            
    woo_sub = "SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE UPPER(SI_NUMBER) = UPPER('%s')"%wo_number
    woo_key = selection_dir(woo_sub,cr)
    woo_key = woo_key and woo_key[0] and woo_key[0][0] or '' 
    if not woo_key:
        bulk_error = 'WO does not exist.'
        return bulk_error,msg,fail_msg,0,0
    wtp_sub = "SELECT WTP_AUTO_KEY FROM WO_TASK_TYPE WHERE TASK_TYPE='MROL TASK'"
    wtp_auto_key = selection_dir(wtp_sub,cr)
    wtp_auto_key = wtp_auto_key and wtp_auto_key[0] and wtp_auto_key[0][0] or None
   
    if not wtp_auto_key:        
        wtp_ins = "INSERT INTO WO_TASK_TYPE (TASK_TYPE,DESCRIPTION) VALUES('MROL TASK','MROL TASK')"
        error = insertion_dir(wtp_ins,cr)
        if error:
            #bulk_error = 'No task type found and problem adding it.'
            return error,msg,fail_msg,0,0
        wtp_sub = "SELECT WTP_AUTO_KEY FROM WO_TASK_TYPE WHERE TASK_TYPE='MROL TASK'"
        wtp_auto_key = selection_dir(wtp_sub,cr)
        wtp_auto_key = wtp_auto_key and wtp_auto_key[0] and wtp_auto_key[0][0] or None
        
    for task in task_list:
        error = ''
        is_duplicate = False
        start_date = 'start_date' in task and task['start_date'] and format_start_date(task['start_date']) or None
        task['start_date'] = start_date
        skill = task.get('skill','')
        skill = "".join([x if ord(x) < 128 else ' ' for x in skill])
        skill = skill.replace("'", r"")
        wot_sequence = task.get('wot_sequence',0)
        task_desc = task.get('task_desc','')
        task_desc = "".join([x if ord(x) < 128 else ' ' for x in task_desc])
        task_desc = task_desc.replace("'", r"")
        task_ref = task.get('task_ref','')
        task_ref = "".join([x if ord(x) < 128 else ' ' for x in task_ref])
        task_ref = task_ref.replace("'", r"")
        task_position = task.get('task_position','')
        task_position = "".join([x if ord(x) < 128 else ' ' for x in task_position])
        task_position = task_position.replace("'", r"")
        task_title = task.get('task_title','')
        task_title = "".join([x if ord(x) < 128 else ' ' for x in task_title])
        task_title = task_title.replace("'", r"")
        task_close_reqs = task.get('task_close_reqs','')
        task_close_reqs = "".join([x if ord(x) < 128 else ' ' for x in task_close_reqs])
        task_close_reqs = task_close_reqs.replace("'", r"")
        skills_est_hours = task.get('skills_est_hours',0)
        wtm_key = task.get('wtm_auto_key',0)
        if wtm_key:
            query = "SELECT WTM_AUTO_KEY FROM WO_TASK_MASTER WHERE UPPER(DESCRIPTION) = UPPER('%s')"%wtm_key
            wtm = selection_dir(query,cr)
            wtm_key = wtm and wtm[0] and wtm[0][0] or None
 
        #uniqueness test - wot_sequence and task_title must be unique together
        #check for skill and if it doesn't exist, put the error message in there
        wok_sub = skill and "SELECT WOK_AUTO_KEY FROM WO_SKILLS WHERE DESCRIPTION = '%s' AND ROWNUM<=1"%skill or None 
        wok_key = wok_sub and selection_dir(wok_sub,cr) or ''    
        wok_key = wok_key and wok_key[0] and wok_key[0][0] or ''        
        if skill:
            if not wok_key:
                task['error'] = 'No skill found with that name.'
                bad_rows.append(task)
                continue                       
        #check if task_position actually exists
        #if not, add it.
        ins_pos = ''
        get_pos = task_position and "SELECT PSR_AUTO_KEY FROM POSITION_REF WHERE POS_REF='%s'"%task_position
        psr_key = get_pos and selection_dir(get_pos,cr) or ''
        psr_key = psr_key and psr_key[0] and psr_key[0][0] or ''
        if task_position and not psr_key:
            ins_pos = "INSERT INTO POSITION_REF (DESCRIPTION,POS_REF) VALUES('%s','%s')"%(task_position,task_position)
            error += insertion_dir(ins_pos,cr) 
            if error:
                task['error'] = error
                bad_rows.append(task)
                continue  
        if task_title and not wtm_key:                
            get_wtm = """
                SELECT WTM.WTM_AUTO_KEY,WT.SEQUENCE,WT.WOO_AUTO_KEY,WT.SQUAWK_DESC,WTM.DESCRIPTION FROM WO_TASK_MASTER WTM 
                LEFT JOIN WO_TASK WT ON WT.WTM_AUTO_KEY = WTM.WTM_AUTO_KEY
                WHERE WTM.DESCRIPTION = '%s'
                """%task_title 
            task_masters = get_wtm and selection_dir(get_wtm,cr) or []
            #check for duplicate records where wtm.description, wtm.woo_auto_key, and wot.ref exist as a trio
            for row in task_masters:
                if row[1] == int(wot_sequence) and row[2] == woo_key and row[4] == task_title:
                    #if the woo, wtm.description and wot.ref trio already exists, we flag as duplicate. 
                    task['error'] = 'Duplicate task_title/wot_sequence pair. Must be unique together per work order.'
                    bad_rows.append(task)
                    is_duplicate = True
                    break
            if is_duplicate:
                continue
            wtm_key = task_masters and task_masters[0] and task_masters[0][0] or None              
        if not wtm_key and task_title:   
            insert_wtm="""       
            INSERT INTO WO_TASK_MASTER (WTM_AUTO_KEY,WTP_AUTO_KEY,DESCRIPTION,LONG_DESCR,WOK_AUTO_KEY) 
            VALUES(G_WTM_AUTO_KEY.NEXTVAL,'%s',SUBSTR('%s',0,50),'%s','%s')
            """%(wtp_auto_key,task_title,task_title,wok_key)
            error = insertion_dir(insert_wtm,cr)
            if error:
                task['error'] = error
                bad_rows.append(task)
                continue
            wtm_sub = "SELECT WTM_AUTO_KEY FROM WO_TASK_MASTER WHERE DESCRIPTION='%s' AND WTP_AUTO_KEY=%s AND ROWNUM <=1"%(task_title,wtp_auto_key)
            wtm_key = selection_dir(wtm_sub,cr)     
            wtm_key = wtm_key and wtm_key[0] and wtm_key[0][0] or '' 
        if not psr_key and ins_pos:
            pos_sub = "SELECT PSR_AUTO_KEY FROM POSITION_REF WHERE POS_REF='%s' AND ROWNUM <=1 ORDER BY PSR_AUTO_KEY DESC"%task_position 
            psr_key = selection_dir(pos_sub,cr)
            psr_key = psr_key and psr_key[0] and psr_key[0][0] or ''           
        if 'start_date' in task and task['start_date']:
            if start_date:
                #if we have skills, insert them...
                if task.get('skills_est_hours','') != '' and skill != '':        
                    insert_task="""
                        INSERT INTO WO_TASK (SKILLS_EST_HOURS,REF,WOO_AUTO_KEY,WTM_AUTO_KEY,
                            SEQUENCE,TASK_START,PSR_AUTO_KEY,SYSUR_AUTO_KEY,SQUAWK_DESC,
                            DESCREPANCY_TEXT,CLOSING_REQUIREMENTS)
                        VALUES (%s,'%s','%s','%s',%s,TO_DATE('%s', 'MM/DD/YYYY'),'%s','%s','%s','%s',SUBSTR('%s',0,40))
                    """%(skills_est_hours,task_ref,woo_key,wtm_key,wot_sequence,start_date,psr_key,sysur_auto_key,task_desc,task_desc,task_close_reqs)
                #else, we don't insert skills...
                else:
                    insert_task="""
                        INSERT INTO WO_TASK (PRINT_TRAVELER_FLAG,REF,WOO_AUTO_KEY,WTM_AUTO_KEY,
                            SEQUENCE,TASK_START,PSR_AUTO_KEY,SYSUR_AUTO_KEY,SQUAWK_DESC,
                            DESCREPANCY_TEXT,CLOSING_REQUIREMENTS)
                        VALUES ('T','%s','%s','%s',%s,TO_DATE('%s', 'MM/DD/YYYY'),'%s','%s','%s','%s',SUBSTR('%s',0,40))
                    """%(task_ref,woo_key,wtm_key,wot_sequence,start_date,psr_key,sysur_auto_key,task_desc,task_desc,task_close_reqs)  
            else:          
                error = 'Invalid date format for task with title: %s and ref: %s.'%(task_title,task_ref)
                task['start_date'] = None
                task['error'] = error
                bad_rows.append(task)
                continue 
        else:
            #if we have skills, insert them...
            if skills_est_hours and skill:
                insert_task="""
                    INSERT INTO WO_TASK (SKILLS_EST_HOURS,REF,WOO_AUTO_KEY,WTM_AUTO_KEY,
                    SEQUENCE,PSR_AUTO_KEY,SYSUR_AUTO_KEY,SQUAWK_DESC,
                    DESCREPANCY_TEXT,CLOSING_REQUIREMENTS)
                    VALUES (%s,'%s','%s','%s',%s,'%s','%s','%s','%s',SUBSTR('%s',0,40))
                """%(skills_est_hours,task_ref,woo_key,wtm_key,wot_sequence,psr_key,sysur_auto_key,task_desc,task_desc,task_close_reqs)   
            #else, we don't insert start date nor skills...
            else:
                insert_task="""
                    INSERT INTO WO_TASK (REF,WOO_AUTO_KEY,WTM_AUTO_KEY,
                    SEQUENCE,PSR_AUTO_KEY,SYSUR_AUTO_KEY,SQUAWK_DESC,
                    DESCREPANCY_TEXT,CLOSING_REQUIREMENTS)
                    VALUES ('%s','%s','%s',%s,'%s','%s','%s','%s',SUBSTR('%s',0,40))
                """%(task_ref,woo_key,wtm_key,wot_sequence,psr_key,sysur_auto_key,task_desc,task_desc,task_close_reqs)                      
        if not error:
            error = insertion_dir(insert_task,cr)
            if error:
                task['error'] = error
                bad_rows.append(task)
                continue              
        if wtm_key and wok_key and skills_est_hours and skill:
            est_hours = skills_est_hours
            est_hours = est_hours and float(est_hours) or 0
            get_task = """
                SELECT WOT_AUTO_KEY 
                FROM WO_TASK 
                WHERE 
                WOO_AUTO_KEY = '%s' 
                AND WTM_AUTO_KEY = '%s'
                AND PSR_AUTO_KEY = '%s'
                AND SYSUR_AUTO_KEY = '%s'
                AND SQUAWK_DESC = '%s'"""%(woo_key,wtm_key,psr_key,sysur_auto_key,task_title)
            wot_key = selection_dir(get_task,cr)
            wot_key = wot_key and wot_key[0] and wot_key[0][0] or None 
            if wot_key:
                insert_task_skill = """
                    INSERT INTO WO_TASK_SKILLS (WOK_AUTO_KEY,WOT_AUTO_KEY,EST_HOURS) VALUES('%s','%s',%s)
                """%(wok_key,wot_key,est_hours)
                error = insertion_dir(insert_task_skill,cr)
        if error:
            task['error'] = error
            bad_rows.append(task)
            continue
        else:
            imported_count += 1
    if imported_count:
        msg = str(len(task_list) - len(bad_rows)) + ' task(s) have been successfully imported into Quantum.  '
        if bad_rows:
            msg += str(len(bad_rows)) + ' task(s) were not imported into Quantum (see grid below for the reasons).'
    else:
        fail_msg = 'None of the rows in the csv were able to be imported into Quantum.'
    WOStatus.objects.filter(session_id=session_id).delete()
    app_id = maps.objects.filter(code='jc-import') 
    if len(task_list) == 1:
        app_id = maps.objects.filter(code='non-routine') 
        q_wot = """SELECT MAX(WOT_AUTO_KEY) FROM WO_TASK"""
        wot = selection_dir(q_wot,cr)
        wot_auto_key = wot and wot[0] and wot[0][0] or ''
    if bad_rows:
        bulk_error = task_bulk_create(quapi,session_id,sysur_auto_key,bad_rows,wo_number)        
    if error == '{"recs": ""}':
        error = ''
    if not error:
        orcl_commit(con=con)    
        #register audit trail record                
        aud_status = 'success'
      
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        field_changed = msg
        new_val = msg
        if error:             
            aud_status = 'failure'
            new_val = fail_msg + ' | ' + error
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')   
        parameters = ['',str(wo_number),str(now)]
        parameters += [str(wot_sequence),str(task_desc)]
        parameters += ['','','','','','']
        parameters += ['','','','','','']
        if user_rec:
            new_val += error
            field_changed += error
            error += register_audit_trail(
            user_rec,
            field_changed,
            new_val,
            right_now,
            app_id,
            quapi,
            status=aud_status,
            reg_events=['Non-routine added'],
            parameters=parameters,
            sysur_auto_key=sysur_auto_key,
            ) 
    else:
        error = 'Incorrect Quantum User ID.'        
    return bulk_error,msg,fail_msg,len(bad_rows),wot_auto_key
   
def format_start_date(start_date,new_format=None):
    date_formats = ["%m-%d-%y","%m-%d-%Y","%Y-%m-%d","%y-%m-%d","%m/%d/%y"]
    date_formats += ["%m/%d/%Y","%Y/%m/%d","%y/%m/%d",'%d %B %Y','%d %B %y']
    date_formats += ['%d %b %Y','%d %b %y','%d %B, %Y','%d %B, %y','%d %b, %Y','%d %b, %y']
    date_formats += ['%b %d %Y','%b %d %y','%B %d, %Y','%B %d, %y','%b %d, %Y','%b %d, %y']
    date_formats += ['%B %d %Y','%B %d %y']
    date = dateparser.parse(start_date, date_formats=date_formats)
    if not new_format:
        new_format = '%m/%d/%Y'
    start_date = date and datetime.strftime(date,new_format) or None
    return start_date
  
def create_ros_bulk(quapi,session_id,ro_recs):
    from polls.models import WOStatus
    ro_data,error,msg = [],'',''
    for rec in ro_recs:
        ro_data.append(WOStatus(
            quapi_id=quapi,
            session_id=session_id,
            is_repair_order=True,
            wo_number = rec[0],
            vendor = rec[1],
            item_number = rec[2],
            part_number = rec[3],
            description = rec[4],
            serial_number = rec[5],
            next_dlv_date = rec[6] and rec[6][:10] or None,
            condition_code = rec[7],
            total_cost = rec[21],
            rod_auto_key = rec[9], 
            parts_cost = rec[10],
            labor_cost = rec[11],
            misc_cost = rec[12],
            approved_date = rec[13] and rec[13][:10] or None,
            quoted_date = rec[14] and rec[14][:10] or None, 
            notes = rec[15], 
            airway_bill = rec[16],
            pnm_modify = rec[17], 
            stm_auto_key = rec[18], 
            qty_reserved = rec[19], 
            si_number = rec[22],
            cust_ref_number = rec[23],
            quantity = rec[24],
            entry_date = rec[25] and rec[25][:10] or None, 
            arrival_date = rec[26] and rec[26][:10] or None,#SMH.SHIP_DATE
            rack = rec[27],#SMH.AIRWAY_BILL
            po_number = rec[28],#SMH.SM_NUMBER
            time_status = rec[29],#SMS.STATUS_CODE
            loc_validated_date = rec[30] or None,#ROD.RO_UDF_001 (QUOTE DATE from ROD)
            customer = rec[31],#SOH.CMP_AUTO_KEY
            wo_type = rec[32],#SMD.ROUTE_CODE
            priority = rec[33],#ROD.QTY_RESERVED*SOD.UNIT_PRICE (SO Value)
            due_date = rec[34] and rec[34][:10] or None,#SOD.DUE_DATE
            )
        )
    try:
        WOStatus.objects.bulk_create(ro_data) or None
    except Exception as err:
        logger.error("Error with creation of repair order locally. Message: '%s'",err.args)           
    return error,msg
    
def get_ro_details(cr,ro_id_list='',\
    ro_number=None,wo_number=None,\
    vendor=None,part_number=None,\
    entry_date=None,rst_auto_key=None,
    uda_status=None):
    add_where = ''
    if ro_id_list:
        add_where += "AND ROD.ROD_AUTO_KEY IN %s"%ro_id_list
    if ro_number:
        add_where += "AND REGEXP_LIKE (ROH.RO_NUMBER, '%s', 'i') "%ro_number
    if wo_number:
        add_where += """AND ROD.WOO_AUTO_KEY IN 
        (SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE REGEXP_LIKE 
        (SI_NUMBER, '%s', 'i'))"""%wo_number
    if rst_auto_key:
        if rst_auto_key == '0':
            add_where += "AND ROH.RST_AUTO_KEY IS NULL"
        else:
            add_where += "AND ROH.RST_AUTO_KEY = '%s'"%rst_auto_key
    if vendor:
        add_where += "AND REGEXP_LIKE (C.COMPANY_NAME, '%s', 'i') "%vendor
    if part_number:
        add_where += "AND REGEXP_LIKE (P.PN, '%s', 'i') "%part_number
    if entry_date:
        add_where += "AND ROD.ENTRY_DATE >= TO_DATE('%s','mm-dd-yyyy') "%entry_date    
    #QTY_REPAIR = 1 <>  would be an open item in the RO_DETAIL table. 
    #We only really care if the ro detail is closed.
   
    if uda_status:
        add_where += """AND (ROH.ROH_AUTO_KEY IN (SELECT AUTO_KEY FROM
        UDA_CHECKED WHERE ATTRIBUTE_VALUE = '%s' AND UDA_AUTO_KEY=64))
        """%(uda_status)
        
    query = """SELECT DISTINCT ROH.RO_NUMBER,C.COMPANY_NAME,ROD.ITEM_NUMBER,
        P.PN,P.DESCRIPTION,S.SERIAL_NUMBER,ROD.NEXT_DELIVERY_DATE,
        PCC.CONDITION_CODE,ROH.TOTAL_COST,ROD.ROD_AUTO_KEY,ROD.PARTS_COST,
        ROD.LABOR_COST,ROD.MISC_COST,ROD.MSG_QUOTE_APPR_DATE,ROD.MSG_QUOTE_REC_DATE,
        TO_CHAR(ROD.REPAIR_NOTES),ROD.MSG_AIRWAY_BILL,PT.PN,S.STM_AUTO_KEY,SR.QTY_RESERVED,
        WO.SI_NUMBER,(ROD.PARTS_COST + ROD.LABOR_COST+ ROD.MISC_COST)*ROD.QTY_REPAIR,
        SOH.COMPANY_REF_NUMBER,SOH.SO_NUMBER,ROD.QTY_REPAIR,ROD.ENTRY_DATE,
        SMH.SHIP_DATE,
        SMH.AIRWAY_BILL,
        SMH.SM_NUMBER,
        SMS.STATUS_CODE,
        ROD.RO_UDF_001,
        SOC.COMPANY_NAME,
        SMD.ROUTE_CODE,
        ROD.QTY_RESERVED*SOD.UNIT_PRICE,
        SOD.DUE_DATE
        FROM RO_DETAIL ROD
        LEFT JOIN RO_HEADER ROH ON ROH.ROH_AUTO_KEY = ROD.ROH_AUTO_KEY
        LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = ROH.CMP_AUTO_KEY
        LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = ROD.PNM_AUTO_KEY
        LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = ROD.PCC_AUTO_KEY
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY
        LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
        LEFT JOIN PARTS_MASTER PT ON PT.PNM_AUTO_KEY = ROD.PNM_MODIFY
        LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = ROD.WOO_AUTO_KEY
        LEFT JOIN SO_DETAIL SOD ON SOD.SOD_AUTO_KEY = ROD.ROD_AUTO_KEY
        LEFT JOIN SO_HEADER SOH ON SOH.SOH_AUTO_KEY = SOD.SOH_AUTO_KEY 
        LEFT JOIN COMPANIES SOC ON SOC.CMP_AUTO_KEY = SOH.CMP_AUTO_KEY        
        LEFT JOIN SM_DETAIL SMD ON SMD.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY
        LEFT JOIN SM_HEADER SMH ON SMH.SMH_AUTO_KEY = SMD.SMH_AUTO_KEY
        LEFT JOIN SM_STATUS SMS ON SMS.SMS_AUTO_KEY = SMH.SMH_AUTO_KEY        
        WHERE
        ((ROD.QTY_REPAIR <> ROD.QTY_REPAIRED + ROD.QTY_SCRAPPED) OR ROD.QTY_REPAIR = 0)
        AND S.HISTORICAL_FLAG = 'F'
        AND S.QTY_OH > 0
        AND SR.QTY_RESERVED > 0
        AND ROH.OPEN_FLAG = 'T'
        AND ROH.HISTORICAL_FLAG = 'F'
        AND ROH.NUMBER_OF_ITEMS > 0 
        """+ add_where

    recs = selection_dir(query,cr)
    return recs

@shared_task
def run_ro_mgmt(session_id,\
            wo_number=None,\
            part_number=None,\
            customer=None,\
            location=None,\
            condition_code=None,\
            socond_code=None,\
            user_id=None,\
            sysur_auto_key=None,\
            stock_label = None,\
            ctrl_id = None,\
            ctrl_number = None,\
            quapi_id = None,\
            clear_cart = None,\
            dj_user_id = None,\
            wos_auto_key = None,\
            stock_status = ''):
    from polls.models import WOStatus as wos_obj,QueryApi
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    if not wo_number and not ctrl_id and stock_label:
        ctrl_number = stock_label[:6]
        ctrl_id = stock_label[7:]
    msg,loc,synch,stat,trail,upd = '','','','','',''
    error,field_changed,new_val,audit_ok,fields_changed = '','','','',False
    stock_recs,updates,updated_stock_recs,strecs = [],{},[],[]
    if not session_id:
        return 'Invalid session.  Please login and try again.',msg 
    #1. send params to stock query method, get_stock_recs()
    #2. take stock lines and bulk create locally using add_wo_record()
    #3. pass back success msg or any error messages   
    stock_recs = get_ro_stock(stock_status=stock_status,wos_auto_key=wos_auto_key,inexact=True,ctrl_id=ctrl_id,ctrl_number=ctrl_number,location=location,customer=customer,wo_number=wo_number,part_number=part_number,quapi=quapi,cond_code=condition_code,socond_code=socond_code)
    if stock_recs:
        del_rows = wos_obj.objects.filter(session_id=session_id).delete() or None
        error,msg = add_ro_stock(session_id,quapi,user_id=user_id,stock_recs=stock_recs)        
    elif not error:
        error = 'No records found.'   
    return error,msg

#0.number,1.due_date,2.PN,3.part_desc,4.serial,5. cond_code,
#6. status,7. loc,8. whs,9.'',10.stm_auto_key,11.owner,
#12.stock_line,13.woo_auto_key,     
def add_ro_stock(session_id,quapi,user_id='',stock_recs=[]):
    from polls.models import WOStatus
    ro_data,error,msg = [],'',''
    for rec in stock_recs:
        #if rec[15] and rec[15] > 0 and not rec[16]:
        ro_data.append(WOStatus(
            quapi_id=quapi,
            session_id=session_id,
            is_repair_order=True,
            wo_number = rec[0],
            due_date = rec[1] and rec[1][:10] or None,
            part_number = rec[2],
            description = rec[3],
            serial_number = rec[4],
            condition_code = rec[5],
            status = rec[6],
            location_code = rec[7],
            wh_code = rec[8],
            customer = rec[9],
            stm_auto_key = rec[10],   
            stock_owner = rec[11],
            stock_line = rec[12], 
            woo_auto_key = rec[13] or 0,
            vendor_key = rec[23] or rec[14] or 0,
            quantity = rec[15] or 0,
            pnm_auto_key = rec[17] or 0,
            vendor = rec[22] or '',
            cond_level_gsix = rec[18] or '', 
            rack = rec[19],
            sub_wo_gate = rec[20] or rec[21],            
            )
        )
    try:
        WOStatus.objects.bulk_create(ro_data) or None
    except Exception as err:
        logger.error("Error with creation of stock records locally. Message: '%s'",err.args)           
    return error,msg
    
def get_ro_stock(stock_status='',wos_auto_key=None,inexact=True,ctrl_id=None,ctrl_number=None,location=None,customer=None,wo_number=None,part_number=None,quapi=None,cond_code=None,socond_code=None):
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    and_where = ''
    if ctrl_id and ctrl_number:
        and_where = "AND S.CTRL_ID = '%s' AND S.CTRL_NUMBER = '%s' "%(ctrl_id,ctrl_number)
    else:
        if wo_number:
            if not inexact:
                and_where = """AND (WO.SI_NUMBER = '%s' 
                OR WV.SI_NUMBER = '%s'
                OR W.SI_NUMBER = '%s'
                OR SOH.SO_NUMBER = '%s'
                OR STH.SO_NUMBER = '%s'
                ) 
                """%(wo_number,wo_number,wo_number,wo_number,wo_number)
            else:
                and_where = """AND (REGEXP_LIKE (WO.SI_NUMBER, '%s', 'i') 
                   OR REGEXP_LIKE (WV.SI_NUMBER, '%s', 'i') 
                   OR REGEXP_LIKE (W.SI_NUMBER, '%s', 'i') 
                   OR REGEXP_LIKE (SOH.SO_NUMBER, '%s', 'i')
                   OR REGEXP_LIKE (STH.SO_NUMBER, '%s', 'i'))
                   """%(wo_number,wo_number,wo_number,wo_number,wo_number)
        if cond_code:
            cond_reg = "REGEXP_LIKE (CONDITION_CODE, '%s', 'i')"%cond_code
            cond_code_where = "(SELECT PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE %s)"%cond_reg
            and_where += "AND (S.PCC_AUTO_KEY IN %s "%cond_code_where
            and_where += "OR ST.PCC_AUTO_KEY IN %s) "%cond_code_where
        if socond_code:
            socond_reg = "REGEXP_LIKE (CONDITION_CODE, '%s', 'i')"%socond_code
            socond_code_where = "(SELECT PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE %s)"%socond_reg
            and_where += "AND (SOD.PCC_AUTO_KEY IN %s "%socond_code_where
            and_where += "OR STD.PCC_AUTO_KEY IN %s) "%socond_code_where
        if customer:
            cust_where = "REGEXP_LIKE (COMPANY_NAME, '%s', 'i') "%customer
            cust_in = "SELECT CMP_AUTO_KEY FROM COMPANIES WHERE %s"%cust_where
            and_where += "AND (WO.CMP_AUTO_KEY IN (%s) OR WV.CMP_AUTO_KEY IN (%s) OR W.CMP_AUTO_KEY IN (%s)) "%(cust_in,cust_in,cust_in)
        if location:
            loc_reg = "REGEXP_LIKE (LOCATION_CODE, '%s', 'i') "%location
            loc_where = "(SELECT LOC_AUTO_KEY FROM LOCATION WHERE %s)"%loc_reg
            and_where += "AND (S.LOC_AUTO_KEY IN %s "%loc_where
            and_where += "OR ST.LOC_AUTO_KEY IN %s) "%loc_where
        if part_number:
            #pn_reg = "REGEXP_LIKE (PN, '%s', 'i') "%part_number
            #pn_where = "(SELECT PNM_AUTO_KEY FROM PARTS_MASTER WHERE %s)"%pn_reg
            #and_where += "AND (S.PNM_AUTO_KEY IN %s "%pn_where
            #and_where += "OR ST.PNM_AUTO_KEY IN %s) "%pn_where
            and_where += """AND (UPPER(P.PN) LIKE UPPER('%s%s') 
            OR UPPER(PT.PN) LIKE UPPER('%s%s')) """%(part_number,'%',part_number,'%')
        if wos_auto_key:
            if wos_auto_key == 'PENDING':
                and_where += """AND SOH.SOS_AUTO_KEY IS NULL 
                    AND STH.SOS_AUTO_KEY IS NULL AND (WO.WOS_AUTO_KEY IS NULL
                    AND W.WOS_AUTO_KEY IS NULL AND WV.WOS_AUTO_KEY IS NULL) """
            else:
                #rows - attribute name matches user selection
                and_where += """AND (RH.ROH_AUTO_KEY IN (SELECT AUTO_KEY FROM
                UDA_CHECKED WHERE ATTRIBUTE_VALUE = '%s' AND UDA_AUTO_KEY=64) OR
                RTH.ROH_AUTO_KEY IN (SELECT AUTO_KEY FROM
                UDA_CHECKED WHERE ATTRIBUTE_VALUE = '%s' AND UDA_AUTO_KEY=64))
                """%(wos_auto_key,wos_auto_key)
                #wos_sub = """(SELECT WOS_AUTO_KEY FROM WO_STATUS
                #    WHERE DESCRIPTION = '%s')"""%wos_auto_key[5:]
                #sos_sub = """(SELECT SOS_AUTO_KEY FROM SO_STATUS
                #    WHERE DESCRIPTION = '%s')"""%wos_auto_key[5:]
                """if wos_auto_key[:2] == 'SO':
                    and_where += """"""AND (SOH.SOS_AUTO_KEY = %s 
                    OR STH.SOS_AUTO_KEY = %s)""""""%(sos_sub,sos_sub)
                elif wos_auto_key[:2] == 'WO':
                    and_where += """"""AND
                    (WO.WOS_AUTO_KEY = %s 
                    OR W.WOS_AUTO_KEY = %s
                    OR WV.WOS_AUTO_KEY = %s) """"""%(wos_sub,wos_sub,wos_sub)"""
        if stock_status:
            #and_where += """AND (CASE WHEN S.STM_AUTO_KEY IS 
            #NOT NULL THEN S.IC_UDL_004 ELSE ST.IC_UDL_004 END) = %s"""%(stock_status)
            and_where += """AND (S.IC_UDL_004 = %s OR ST.IC_UDL_004 = %s) """%(stock_status,stock_status)
            
    order_by = " ORDER BY S.STM_AUTO_KEY DESC"
    query = """SELECT DISTINCT
        CASE WHEN SOH.SO_NUMBER IS NOT NULL THEN SOH.SO_NUMBER ELSE 
        (CASE WHEN STH.SO_NUMBER IS NOT NULL THEN STH.SO_NUMBER ELSE
        (CASE WHEN RH.RO_NUMBER IS NOT NULL THEN RH.RO_NUMBER ELSE
        (CASE WHEN RTH.RO_NUMBER IS NOT NULL THEN RTH.RO_NUMBER ELSE
        (CASE WHEN WO.SI_NUMBER IS NOT NULL THEN WO.SI_NUMBER ELSE 
        (CASE WHEN W.SI_NUMBER IS NOT NULL THEN W.SI_NUMBER ELSE
        (CASE WHEN WV.SI_NUMBER IS NOT NULL THEN WV.SI_NUMBER ELSE '' END)
        END) END) END) END) END) END,		
        WO.DUE_DATE,
        CASE WHEN PT.PN IS NOT NULL THEN PT.PN ELSE P.PN END,
        CASE WHEN PT.PN IS NOT NULL THEN PT.DESCRIPTION ELSE P.DESCRIPTION END,
         CASE WHEN ST.SERIAL_NUMBER IS NOT NULL THEN ST.SERIAL_NUMBER ELSE S.SERIAL_NUMBER END,
          CASE WHEN PTC.CONDITION_CODE IS NOT NULL THEN PTC.CONDITION_CODE ELSE PCC.CONDITION_CODE END,
           CASE WHEN SOS.SOS_AUTO_KEY IS NOT NULL THEN SOS.DESCRIPTION ELSE 
           CASE WHEN STS.SOS_AUTO_KEY IS NOT NULL THEN STS.DESCRIPTION ELSE
           CASE WHEN WOS.WOS_AUTO_KEY IS NOT NULL THEN WOS.DESCRIPTION ELSE
           CASE WHEN WS.WOS_AUTO_KEY IS NOT NULL THEN WS.DESCRIPTION ELSE
           CASE WHEN WVS.WOS_AUTO_KEY IS NOT NULL THEN WVS.DESCRIPTION END END END END END,
           CASE WHEN LT.LOCATION_CODE IS NOT NULL THEN LT.LOCATION_CODE ELSE L.LOCATION_CODE END,
           CASE WHEN WHT.WAREHOUSE_CODE IS NOT NULL THEN WHT.WAREHOUSE_CODE ELSE WH.WAREHOUSE_CODE END,
           CASE WHEN COH.COMPANY_NAME IS NOT NULL THEN COH.COMPANY_NAME ELSE 
           (CASE WHEN CTH.COMPANY_NAME IS NOT NULL THEN CTH.COMPANY_NAME ELSE
           (CASE WHEN C.COMPANY_NAME IS NOT NULL THEN C.COMPANY_NAME ELSE
           (CASE WHEN CO.COMPANY_NAME IS NOT NULL THEN CO.COMPANY_NAME ELSE
           (CASE WHEN CV.COMPANY_NAME IS NOT NULL THEN CV.COMPANY_NAME ELSE '' END) 
           END) END) END) END,
           CASE WHEN ST.STM_AUTO_KEY IS NOT NULL THEN ST.STM_AUTO_KEY ELSE S.STM_AUTO_KEY END,
           CASE WHEN ST.OWNER IS NOT NULL THEN ST.OWNER ELSE S.OWNER END,
           CASE WHEN ST.STOCK_LINE IS NOT NULL THEN ST.STOCK_LINE ELSE S.STOCK_LINE END,
           CASE WHEN WO.WOO_AUTO_KEY IS NOT NULL THEN WO.WOO_AUTO_KEY ELSE
           (CASE WHEN W.WOO_AUTO_KEY IS NULL THEN WV.WOO_AUTO_KEY ELSE 
           W.WOO_AUTO_KEY END) END,
           CASE WHEN CMTP.CMP_AUTO_KEY IS NOT NULL THEN CMTP.CMP_AUTO_KEY ELSE CMP.CMP_AUTO_KEY END,
           CASE WHEN ST.STM_AUTO_KEY IS NOT NULL THEN ST.QTY_OH ELSE S.QTY_OH END,
        CASE WHEN ST.STM_AUTO_KEY IS NOT NULL THEN STR.ROD_AUTO_KEY ELSE SR.ROD_AUTO_KEY END,
		CASE WHEN PT.PNM_AUTO_KEY IS NOT NULL THEN PT.PNM_AUTO_KEY ELSE P.PNM_AUTO_KEY END,
        CASE WHEN SOD.SOD_AUTO_KEY IS NOT NULL THEN PSD.CONDITION_CODE ELSE PSTD.CONDITION_CODE END,
        CASE WHEN ST.STM_AUTO_KEY IS NOT NULL THEN UDL.UDL_CODE 
        ELSE UDLT.UDL_CODE END,
        UDC.ATTRIBUTE_VALUE,
        UDCH.ATTRIBUTE_VALUE
        FROM STOCK S
        LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_004
        LEFT JOIN STOCK ST ON ST.STM_LOT = S.STM_AUTO_KEY
            LEFT JOIN (SELECT RD.PNM_AUTO_KEY,RH.CMP_AUTO_KEY,
                    ROW_NUMBER() OVER (PARTITION BY RD.PNM_AUTO_KEY ORDER BY RH.ENTRY_DATE DESC) RN
                FROM RO_DETAIL RD
                LEFT JOIN RO_HEADER RH
                    ON RH.ROH_AUTO_KEY = RD.ROH_AUTO_KEY) LASTCMP
                    ON LASTCMP.PNM_AUTO_KEY = S.PNM_AUTO_KEY
                    AND LASTCMP.RN = 1                
                LEFT JOIN COMPANIES CMP
                    ON CMP.CMP_AUTO_KEY = LASTCMP.CMP_AUTO_KEY
                LEFT JOIN (SELECT RTD.PNM_AUTO_KEY,RTH.CMP_AUTO_KEY,
                    ROW_NUMBER() OVER (PARTITION BY RTD.PNM_AUTO_KEY ORDER BY RTH.ENTRY_DATE DESC) RN
                FROM RO_DETAIL RTD
                LEFT JOIN RO_HEADER RTH
                    ON RTH.ROH_AUTO_KEY = RTD.ROH_AUTO_KEY) LASTCMTP
                    ON LASTCMTP.PNM_AUTO_KEY = ST.PNM_AUTO_KEY
                    AND LASTCMTP.RN = 1                
                LEFT JOIN COMPANIES CMTP
                    ON CMTP.CMP_AUTO_KEY = LASTCMTP.CMP_AUTO_KEY      
		LEFT JOIN USER_DEFINED_LOOKUPS UDLT ON UDLT.UDL_AUTO_KEY = ST.IC_UDL_004					
        LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY = S.WHS_AUTO_KEY
        LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
        LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
        LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
        LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
        LEFT JOIN WO_STATUS WOS ON WOS.WOS_AUTO_KEY = WO.WOS_AUTO_KEY
        LEFT JOIN COMPANIES CO ON CO.CMP_AUTO_KEY = WO.CMP_AUTO_KEY
        LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
        LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
        LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY
        LEFT JOIN SO_DETAIL SOD ON SOD.SOD_AUTO_KEY = SR.SOD_AUTO_KEY
        LEFT JOIN PART_CONDITION_CODES PSD ON PSD.PCC_AUTO_KEY = SOD.PCC_AUTO_KEY
        LEFT JOIN SO_HEADER SOH ON SOH.SOH_AUTO_KEY = SOD.SOH_AUTO_KEY
        LEFT JOIN COMPANIES COH ON COH.CMP_AUTO_KEY = SOH.CMP_AUTO_KEY
        LEFT JOIN SO_STATUS SOS ON SOS.SOS_AUTO_KEY = SOH.SOS_AUTO_KEY       
        LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = W.CMP_AUTO_KEY
        LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.STM_AUTO_KEY = S.STM_AUTO_KEY
        LEFT JOIN WO_OPERATION WV ON WV.WOO_AUTO_KEY = VW.WOO_AUTO_KEY
        LEFT JOIN WO_STATUS WVS ON WVS.WOS_AUTO_KEY = WV.WOS_AUTO_KEY        
        LEFT JOIN COMPANIES CV ON CV.CMP_AUTO_KEY = WV.CMP_AUTO_KEY
        LEFT JOIN LOCATION LT ON LT.LOC_AUTO_KEY = ST.LOC_AUTO_KEY
        LEFT JOIN WAREHOUSE WHT ON WHT.WHS_AUTO_KEY = ST.WHS_AUTO_KEY
        LEFT JOIN PARTS_MASTER PT ON PT.PNM_AUTO_KEY = ST.PNM_AUTO_KEY
        LEFT JOIN PART_CONDITION_CODES PTC ON PTC.PCC_AUTO_KEY = ST.PCC_AUTO_KEY
        LEFT JOIN STOCK_RESERVATIONS STR ON STR.STM_AUTO_KEY = ST.STM_AUTO_KEY
        LEFT JOIN SO_DETAIL STD ON STD.SOD_AUTO_KEY = STR.SOD_AUTO_KEY
        LEFT JOIN PART_CONDITION_CODES PSTD ON PSTD.PCC_AUTO_KEY = STD.PCC_AUTO_KEY
        LEFT JOIN SO_HEADER STH ON STH.SOH_AUTO_KEY = STD.SOH_AUTO_KEY
        LEFT JOIN COMPANIES CTH ON CTH.CMP_AUTO_KEY = STH.CMP_AUTO_KEY
        LEFT JOIN SO_STATUS STS ON STS.SOS_AUTO_KEY = STH.SOS_AUTO_KEY
        LEFT JOIN RO_DETAIL RD ON RD.ROD_AUTO_KEY = S.STM_AUTO_KEY
        LEFT JOIN RO_HEADER RH ON RH.ROH_AUTO_KEY = RD.ROH_AUTO_KEY
        LEFT JOIN RO_DETAIL RTD ON RTD.ROD_AUTO_KEY = ST.ROD_AUTO_KEY
        LEFT JOIN RO_HEADER RTH ON RTH.ROH_AUTO_KEY = RTD.ROH_AUTO_KEY
        LEFT JOIN UDA_CHECKED UDC ON UDC.AUTO_KEY = RH.ROH_AUTO_KEY
        LEFT JOIN UDA_CHECKED UDCH ON UDCH.AUTO_KEY = RTH.ROH_AUTO_KEY  
        WHERE S.QTY_OH > 0 %s
        AND (CASE WHEN SOD.SOD_AUTO_KEY IS NOT NULL THEN 
        (CASE WHEN S.PCC_AUTO_KEY <> SOD.PCC_AUTO_KEY THEN 1 ELSE
        0 END) ELSE (CASE WHEN STD.SOD_AUTO_KEY IS NOT NULL THEN 
        (CASE WHEN ST.PCC_AUTO_KEY <> STD.PCC_AUTO_KEY THEN 1 ELSE 
        0 END) ELSE 1 END) END)=1
        AND (CASE WHEN ST.STR_AUTO_KEY IS NOT NULL THEN 
        (CASE WHEN ST.QTY_RESERVED > 0 THEN 1 ELSE 0 END) ELSE
        (CASE WHEN STR.STR_AUTO_KEY IS NOT NULL THEN 
        (CASE WHEN STR.QTY_RESERVED > 0 THEN 1 ELSE 0 END) ELSE 1 END) END)=1
        AND (CASE WHEN ST.STR_AUTO_KEY IS NOT NULL THEN 
        (CASE WHEN ST.ROD_AUTO_KEY IS NULL THEN 1 ELSE 0 END) ELSE
        (CASE WHEN STR.STR_AUTO_KEY IS NOT NULL THEN 
        (CASE WHEN STR.ROD_AUTO_KEY IS NULL THEN 1 ELSE 0 END) ELSE 1 END) END)=1       
    """%(and_where)
    recs = selection_dir(query,cr)
    counter = 0
    stm_list = []
    ro_recs = []
    
    for stock in recs:
        pnm_auto_key = stock[17]
        stm_auto_key = stock[10]
        stock += ['','']
        
        if stm_auto_key in stm_list: 
            counter+=1        
            continue
            
        if part_number:
            if stock[2] != part_number:
                query = """SELECT PNM_AUTO_KEY,
                PN,DESCRIPTION FROM PARTS_MASTER
                WHERE UPPER(PN) = '%s'
                """%part_number
                part = selection_dir(query,cr)
                part = part and part[0]
                if part:
                    pnm_auto_key = part[0]
                    stock[2] = part[1]
                    stock[3] = part[2]
                    stock[17] = pnm_auto_key
                
        stm_list.append(stm_auto_key)
        
        if pnm_auto_key:
            query="""SELECT CMP.COMPANY_NAME,CMP.CMP_AUTO_KEY            
                FROM COMPANIES CMP,
                RO_HEADER ROH, 
                RO_DETAIL ROD, 
                PARTS_MASTER PNM
                WHERE CMP.CMP_AUTO_KEY = ROH.CMP_AUTO_KEY 
                AND ROH.ROH_AUTO_KEY = ROD.ROH_AUTO_KEY 
                AND ROD.PNM_AUTO_KEY = PNM.PNM_AUTO_KEY 
                AND PNM.PNM_AUTO_KEY = %s 
                ORDER BY ROD.ROD_AUTO_KEY DESC"""%pnm_auto_key
                
            last_vendor = selection_dir(query,cr)
            last_vendor = last_vendor and last_vendor[0]           
            #last_vendor = last_vendor and last_vendor[0] or ''
            stock[22] = last_vendor and last_vendor[0] or ''
            stock[23] = last_vendor and last_vendor[1] or 0
            
        ro_recs.append(stock)
        counter+=1
        
    return ro_recs
    
def get_stock_ros(cr,stm_auto_key=None,ro_number=None,stm_keys=[]):
    recs = []
    if stm_auto_key:
        where = 'S.STM_AUTO_KEY = %s'%stm_auto_key
        query = form_stock_ro(where)
        recs += selection_dir(query,cr)
    for skey in stm_keys:
        where = 'S.STM_AUTO_KEY IN %s'%skey        
        query = form_stock_ro(where)
        recs += selection_dir(query,cr)
    return recs
    
def form_stock_ro(where): 
    trial_left_joins = """
    LEFT JOIN (SELECT ROW_NUMBER() OVER (PARTITION BY ROD.PNM_AUTO_KEY ORDER BY ROH.ENTRY_DATE DESC) RON,
        ROD.PNM_AUTO_KEY,ROH.CMP_AUTO_KEY FROM RO_HEADER ROH 
        LEFT JOIN RO_DETAIL ROD
        ON ROH.ROH_AUTO_KEY = ROD.ROH_AUTO_KEY) LAST_V
        ON LAST_V.PNM_AUTO_KEY = S.PNM_AUTO_KEY
        AND LAST_V.RON = 1
        LEFT JOIN COMPANIES C ON LAST_V.CMP_AUTO_KEY = C.CMP_AUTO_KEY
        LEFT JOIN (
            SELECT SR.STM_AUTO_KEY FROM STOCK_RESERVATIONS SR
            WHERE ROD_AUTO_KEY IS NULL AND QTY_RESERVED > 0
            GROUP BY SR.STM_AUTO_KEY) STRO
            ON STRO.STM_AUTO_KEY = S.STM_AUTO_KEY
        LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = S.ROD_AUTO_KEY
        LEFT JOIN RO_HEADER ROH ON ROH.ROH_AUTO_KEY = ROD.ROH_AUTO_KEY
    """    
    left_joins="""left join         ( select          rd.pnm_auto_key, rh.cmp_auto_key,
                                    row_number() over (partition by rd.pnm_auto_key order by rh.entry_date desc) rn 
                    from            ro_detail rd 
                    left join       ro_header rh 
                    on              rh.roh_auto_key = rd.roh_auto_key
                  ) lastcmp
                on                lastcmp.pnm_auto_key = s.pnm_auto_key
                                  and lastcmp.rn = 1                    
                left join	        companies cmp 
                on                cmp.cmp_auto_key = lastcmp.cmp_auto_key 

                left join         ( select          r.stm_auto_key
                                    from            stock_reservations r
                                    where           rod_auto_key is not null
                                                    and qty_reserved > 0
                                    group by        r.stm_auto_key
                                  ) resro
                on                resro.stm_auto_key = s.stm_auto_key

                left join         ro_detail rd 
                on                rd.rod_auto_key = s.rod_auto_key
                left join         ro_header rh 
                on                rh.roh_auto_key = rd.roh_auto_key
    """
    query = """SELECT DISTINCT
        S.PNM_AUTO_KEY,
        S.PCC_AUTO_KEY,
        S.CMP_AUTO_KEY,
        S.QTY_OH,
        S.SERIAL_NUMBER,
        S.CMP_AUTO_KEY,
        CP.CMP_AUTO_KEY,
        S.CMP_AUTO_KEY,
        S.STM_AUTO_KEY,
        SR.STR_AUTO_KEY,
        S.DPT_AUTO_KEY,
        S.SYSUR_AUTO_KEY,
        S.SYSCM_AUTO_KEY,
        CASE WHEN WH.SHIP_ADDRESS1 IS NOT NULL THEN WH.SHIP_ADDRESS1 ELSE 
            (CASE WHEN DP.ADDRESS_LINE1 IS NOT NULL THEN DP.ADDRESS_LINE1 ELSE
                SCM.SHIP_ADDRESS1 END) END,
        CASE WHEN WH.SHIP_ADDRESS1 IS NOT NULL THEN WH.SHIP_ADDRESS1 ELSE 
            (CASE WHEN DP.ADDRESS_LINE1 IS NOT NULL THEN DP.ADDRESS_LINE2 ELSE
                SCM.SHIP_ADDRESS2 END) END,
        CASE WHEN WH.SHIP_ADDRESS1 IS NOT NULL THEN WH.SHIP_ADDRESS3 ELSE 
            (CASE WHEN DP.ADDRESS_LINE1 IS NULL THEN SCM.SHIP_ADDRESS3 END) END,
        CASE WHEN WH.SHIP_ADDRESS1 IS NOT NULL THEN WH.SHIP_ADDRESS4 ELSE 
            (CASE WHEN DP.ADDRESS_LINE1 IS NOT NULL THEN DP.CITY ||','|| DP.STATE ||','|| DP.ZIP_CODE ELSE
                 SCM.SHIP_CITY ||','|| SCM.SHIP_STATE ||','|| SCM.SHIP_ZIP_CODE END) END,
        CASE WHEN WH.SHIP_ADDRESS1 IS NOT NULL THEN WH.SHIP_ADDRESS5 ELSE SCM.SHIP_COUNTRY END,
        CASE WHEN WH.SHIP_ADDRESS1 IS NOT NULL THEN WH.SHIP_NAME ELSE 
            (CASE WHEN DP.ADDRESS_LINE1 IS NOT NULL THEN DP.DEPT_NAME ELSE
             SCM.SHIP_NAME END) END,
        CP.SHIP_ADDRESS1,
        CP.SHIP_ADDRESS2,
        CP.SHIP_ADDRESS3,
        CP.SHIP_CITY ||','|| CP.SHIP_STATE ||','|| CP.SHIP_ZIP_CODE,
        CP.SHIP_COUNTRY END,
        CP.SHIP_NAME END,
        CASE WHEN SR.WOB_AUTO_KEY IS NULL THEN VW.WOB_AUTO_KEY ELSE SR.WOB_AUTO_KEY END,
        CASE WHEN SR.SOD_AUTO_KEY IS NULL THEN VW.SOD_AUTO_KEY ELSE SR.SOD_AUTO_KEY END,
        CASE WHEN SR.WOO_AUTO_KEY IS NULL THEN VW.WOO_AUTO_KEY ELSE SR.WOO_AUTO_KEY END,
        TO_CHAR(S.NOTES)
        FROM STOCK S
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
        LEFT JOIN COMPANIES CP ON CP.CMP_AUTO_KEY = S.CMP_AUTO_KEY
        LEFT JOIN SYS_COMPANIES SCM ON SCM.SYSCM_AUTO_KEY = S.SYSCM_AUTO_KEY
        LEFT JOIN DEPARTMENT DP ON DP.DPT_AUTO_KEY = S.DPT_AUTO_KEY
        LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY = S.WHS_AUTO_KEY
        LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.STM_AUTO_KEY = S.STM_AUTO_KEY
        WHERE %s"""%(where)
    return query  

def insert_ro_header(cr,sysur_auto_key,vendor_key,right_now,srecs):
    ship_address1 = ''
    ship_address2 = ''
    ship_address3 = ''
    ship_address4 = ''
    ship_address5 = ''
    ship_name = '' 
    v_address1 = ''
    v_address2 = ''
    # v_address3 = ''
    v_address4 = ''
    v_address5 = ''
    v_name = '' 
    """GET THE SYS_COMPANIES SHIPPING ADDRESS"""
    query="""SELECT SYS.ADDRESS1, SYS.ADDRESS2, 
    SYS.ADDRESS3, SYS.CITY, SYS.STATE, SYS.ZIP_CODE,
    SYS.COMPANY_NAME FROM SYS_COMPANIES SYS
    WHERE SYS.SYSCM_AUTO_KEY = '%s'   
    """%srecs[12]
    rec = selection_dir(query,cr)        
    ship_add = rec and rec[0]
    if ship_add:
        ship_address1 = ship_add[0] or ''
        ship_address2 = ship_add[1] or ''
        ship_address3 = ship_add[2] or ''
        ship_address4 = ship_add[3] + ', ' + ship_add[4]
        ship_address4 += ' ' + ship_add[5]
        ship_address5 = ''
        ship_name = ship_add[6]
    """GET THE VENDOR ADDRESS"""
    query="""SELECT C.ADDRESS1, C.ADDRESS2, 
    C.ADDRESS3, C.CITY, C.STATE, C.ZIP_CODE,
    C.COMPANY_NAME FROM COMPANIES C
    WHERE C.CMP_AUTO_KEY = '%s'
    """%vendor_key
    rec = selection_dir(query,cr)
    v_add = rec and rec[0]  
    if v_add:
        v_address1 = v_add[0] or ''
        v_address2 = v_add[1] or ''
        v_address3 = v_add[2] or ''
        v_address4 = v_add[3] + ', ' + v_add[4]
        v_address4 += ' ' + v_add[5]
        v_address5 = ''
        v_name = v_add[6]
    
    error = ''
    rohs_created = 0
    #query = "SELECT RO_NUMBER FROM RO_HEADER ORDER BY ROH_AUTO_KEY DESC FETCH NEXT 1 ROWS ONLY"            
    #ro_num_recs = selection(query,cr=cr)
    #ro_number = ro_num_recs and ro_num_recs[0] and ro_num_recs[0][0] or ''
    #ro_number = ro_number and int(ro_number.split('-')[0]) or None
    #ro_number = ro_number and ro_number + 1 or 'None'
    #q1 = "SELECT RO_NUMBER FROM RO_HEADER ORDER BY ROH_AUTO_KEY DESC FETCH NEXT 1 ROWS ONLY"
    #ro_number = (TO_NUMBER(%s) + 1)%q1
    num_query = """select sl.last_number + 1, sl.sysnl_auto_key, sl.number_prefix
      from            sys_number_log sl
      left join       sys_number_log_codes lc
      on              lc.sysnlc_auto_key = sl.sysnlc_auto_key
      where           (lc.log_type_code = 'RO')"""
    sysnum = selection_dir(num_query,cr)    
    sysnum = sysnum and sysnum[0] or ''
    last_num_mas1 = sysnum and sysnum[0] or ''
    sysnl_auto_key = sysnum and sysnum[1] or 0
    number_prefix = sysnum and sysnum[2] or 'RO'
    if last_num_mas1:
        #adding new fields into insert 
        query = """INSERT INTO RO_HEADER (RO_NUMBER, CMP_AUTO_KEY,
          SYSUR_AUTO_KEY,ENTRY_DATE, SYSCM_AUTO_KEY, DPT_AUTO_KEY, OPEN_FLAG,
          ship_address1, ship_address2, ship_address3, ship_address4,
          ship_address5, ship_name,
          vendor_address1, vendor_address2, vendor_address3, 
          vendor_address4, vendor_address5, vendor_name) 
          VALUES('%s','%s','%s',TO_DATE('%s','MM-DD-YYYY'),'%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')
          """%(str(number_prefix) + str(last_num_mas1),vendor_key,sysur_auto_key,right_now[:10],
          srecs[12],srecs[10],'T',ship_address1,ship_address2,ship_address3,ship_address4,
          ship_address5,ship_name,v_address1,v_address2,v_address3,v_address4,v_address5,v_name)
        try:
            insertion_dir(query,cr)
        except cx_Oracle.DatabaseError as exc:
            error, = exc.args
            error = error.message
            logger.error("Error with inserting new RO: '%s'",error)               
    if (not error or error == '{"recs": ""}') and sysnl_auto_key:
        query ="""UPDATE sys_number_log snl 
          SET snl.last_number = '%s'
          WHERE snl.sysnl_auto_key = '%s'"""%(last_num_mas1,sysnl_auto_key)
        rohs_created = str(number_prefix) + str(last_num_mas1)
        error = updation_dir(query,cr)       
    return error,rohs_created
    
def insert_ro_detail(cr,stock_recs,right_now,sysur_auto_key,roh_auto_key=None): 
    """we do have an issue with RO Management though. When in stock search and creating an ro with a stm, 
        it creates an RO Detail with QTY_RESERVED = 0, this might be causing the receiving issues i'm having. 
        ROD['QTY_RESERVED'] = WOB/WOO/SOD['QTY_RESERVED'] where the STM came from or if STM['QTY_RESERVED'] = 0 
        THEN ROD['QTY_RESERVED'] = STM['QTY_AVAILABLE']
    """
    qty_reserved = 0
    query = stock_recs and """INSERT INTO RO_DETAIL 
    (PNM_AUTO_KEY,PCC_AUTO_KEY,ROH_AUTO_KEY,SYSUR_AUTO_KEY,
    ENTRY_DATE,QTY_REPAIR,QTY_REPAIRED,SERIAL_NUMBER,
    RO_TYPE,WOB_AUTO_KEY,SOD_AUTO_KEY,WOO_AUTO_KEY,QTY_RESERVED,
    REPAIR_NOTES) 
    VALUES('%s','%s','%s','%s',TO_DATE('%s','MM-DD-YYYY'),%s,0,'%s','Overhaul','%s','%s','%s','%s','%s')
    """%(stock_recs[0] or '',stock_recs[1] or '',roh_auto_key or stock_recs[2],sysur_auto_key or '',right_now[:10],stock_recs[3],stock_recs[4],stock_recs[25],stock_recs[26],stock_recs[27],qty_reserved,stock_recs[28])
    rods_created = 0
    error = insertion_dir(query,cr) 
    if error == '{"recs": ""}' or error == '':
        error = ''
        rods_created += 1
    #else:
    #   'No stock records to add.  Check your selection.'
    #now we must add the rod_auto_key to the reservation
    """ S.PNM_AUTO_KEY,0
        S.PCC_AUTO_KEY,1
        S.CMP_AUTO_KEY,2
        S.QTY_OH,3
        S.SERIAL_NUMBER,4
        S.CMP_AUTO_KEY,5
        CP.CMP_AUTO_KEY,6
        S.CMP_AUTO_KEY,7
        S.STM_AUTO_KEY,8
        SR.STR_AUTO_KEY,9
        S.DPT_AUTO_KEY,10
        S.SYSUR_AUTO_KEY,11
        S.SYSCM_AUTO_KEY,12"""
        
    query = """UPDATE STOCK SET IC_UDL_004 = NULL
        WHERE STM_AUTO_KEY = %s"""%stock_recs[8]
    error = updation_dir(query,cr)
    query = "select qty_reserved,qty_available,qty_oh from stock where stm_auto_key = '%s'"%stock_recs[8]
    stock_q = selection_dir(query,cr)
    qty_reserved = stock_q and stock_q[0] and stock_q[0][0] or 0
    if qty_reserved == 0:
        qty_reserved = stock_q and stock_q[0] and stock_q[0][1] or 0
    if qty_reserved == 0:
        qty_reserved = stock_q and stock_q[0] and stock_q[0][2] or 0
    str_auto_key = stock_recs and stock_recs[9] or None 
    query = """SELECT ROD_AUTO_KEY FROM RO_DETAIL WHERE  
        ROH_AUTO_KEY = %s AND SYSUR_AUTO_KEY = %s AND
        ENTRY_DATE = TO_DATE('%s','MM-DD-YYYY') ORDER BY ROD_AUTO_KEY DESC"""%(roh_auto_key,sysur_auto_key,right_now)
    rod = selection_dir(query,cr)
    rod_auto_key = rod and rod[0] and rod[0][0] or None
    if str_auto_key and qty_reserved > 0:
        query = "DELETE FROM STOCK_RESERVATIONS WHERE STR_AUTO_KEY = %s"%str_auto_key
        #query = "UPDATE STOCK_RESERVATIONS SET ROD_AUTO_KEY = %s WHERE STR_AUTO_KEY = %s"%(sub_query,str_auto_key)
        res = updation_dir(query,cr)           
    #else:
        #rod_new = "SELECT ROD_AUTO_KEY FROM RO_DETAIL WHERE SYSUR_AUTO_KEY"
    if rod_auto_key:
        new_res = """INSERT INTO STOCK_RESERVATIONS (STR_AUTO_KEY,STM_AUTO_KEY,ROD_AUTO_KEY,QTY_RESERVED) VALUES (G_STR_AUTO_KEY.NEXTVAL,%s,%s,%s)"""%(stock_recs[8],rod_auto_key,stock_recs[3])
        newrs = insertion_dir(new_res,cr)
        query = "UPDATE RO_DETAIL SET QTY_RESERVED = %s WHERE ROD_AUTO_KEY = %s"%(qty_reserved,rod_auto_key)
        res = updation_dir(query,cr)
    return rods_created
    
def get_ro_vendor(srec,session_id,wos_obj):
    moves = wos_obj.objects.filter(session_id=session_id,stm_auto_key=srec[8])
    last_vendor_key = None
    for move in moves:
        if move.vendor_key:
            last_vendor_key = move.vendor_key
            break
    return last_vendor_key
    
def get_new_roh(cr,vendor_key,right_now,sysur_auto_key=''):
    sysur_auto_key = sysur_auto_key and 'AND SYSUR_AUTO_KEY = %s '%sysur_auto_key or '' 
    entry_date = ''   
    if sysur_auto_key:
        entry_date = ''
    else:
        entry_date = "AND ENTRY_DATE=TO_DATE('%s','MM-DD-YYYY')"%right_now  
    query = """SELECT ROH_AUTO_KEY FROM RO_HEADER WHERE 
        CMP_AUTO_KEY = %s 
        %s
        %s
        ORDER BY ROH_AUTO_KEY 
        DESC"""%(vendor_key,entry_date,sysur_auto_key)
    roh = selection_dir(query,cr)
    roh_auto_key = roh and roh[0] or None
    roh_auto_key = roh_auto_key and roh_auto_key[0] or None
    return roh_auto_key
    
@shared_task
def add_new_ro(quapi_id,session_id,sysur_auto_key,ro_number=None,stm_keys=[],vendor=None,last_vendor=False,woo_keys=[]):
    error,msg,rods_created = '','',0
    from polls.models import WOStatus as wos_obj,QueryApi,MLApps as maps,QuantumUser as qu
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    #1. If not vendor, then get vendor from roh.last_vendor field
    #2. If vendor, then add the new ROH, RO_DETAIL
    #3. If ro_number entered in text field by user, 
    #then we look it up and add (create RO_DETAIL lines) under the RO_HEADER
    #Any selected stock lines get added to the new RO as detail lines.s
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.'      
    if stm_keys:
        stm_key_list = construct_akl(stm_keys)
        right_now = datetime.now()
        right_now = right_now.strftime('%m-%d-%Y') 
        if not vendor and last_vendor:
            #1. get last vendor from stock lines
            #2. add RO_HEADER row
            #3. Create a row for RO_DETAIL for each stock line 
            #get the last vendor from the stock line
            recs = []
            roh_msg = 'RO#(s) created: '
            count = 0
            
            vendor_rohs = {}
            ro_header_ids = []            
            #loop through and create one RO Header per last vendor and create RO_DETAIl lines for each stock line 
            
            for stm_key in stm_keys:                     
                srec = get_stock_ros(cr,stm_auto_key=stm_key)
                srec = srec and srec[0] or []
                vendor_key = get_ro_vendor(srec,session_id,wos_obj)
                #stock_recs[6] is the last vendor
                #create the RO_DETAIL line.
                #get the user from the form and lookup sysur_auto_key
                if not vendor_key:
                    return 'No vendor found.',''
                elif vendor_key not in vendor_rohs:
                    error,rohs_created = insert_ro_header(cr,sysur_auto_key,vendor_key,right_now,srec)
                    if roh_msg != 'RO#(s) created: ':
                        roh_msg += ', ' + str(rohs_created)
                    else:
                        roh_msg += str(rohs_created)
                    roh_auto_key = get_new_roh(cr,vendor_key,right_now,sysur_auto_key=sysur_auto_key)   
                    vendor_rohs[vendor_key] = [roh_auto_key]
                elif vendor_key in vendor_rohs:
                    roh_auto_key = get_new_roh(cr,vendor_key,right_now,sysur_auto_key=sysur_auto_key) 
                    vendor_rohs[vendor_key].append(roh_auto_key) 
                rods_created += insert_ro_detail(cr,srec,right_now,sysur_auto_key,roh_auto_key=roh_auto_key)                      
                count += 1                       
            msg = str(roh_msg) + ' along with one detail line each.'
        elif ro_number:
            #1. look up RO from RO# entered.
            ro_header = get_ro_header(cr,ro_number=ro_number)
            #2. Insert RO_DETAIL LINES (with detail line from stock line in question)
            #into existing (user-entered RO from step #1) 
            #Loop through the stock lines and add an RO_DETAIL for each of the stock lines.
            #for stm_auto_key in stm_keys:
            #error,rohs_created = insert_ro_header(cr,vendor_key,right_now,stock_recs)
            if ro_header:
                ro_number = ro_header[0] and ro_header[0][0] or None
                roh_auto_key = ro_header[0] and ro_header[0][1] or None   
                if roh_auto_key:
                    count = 0
                    for stm_key in stm_keys:
                        srec = get_stock_ros(cr,stm_auto_key=stm_key)
                        srec = srec and srec[0] or []
                        #vendor_key = srec and get_ro_vendor(srec,session_id,wos_obj) or None
                        #if not vendor_key:
                        #    return 'No vendor found.',''
                        #get vendor from stock record
                        rods_created += insert_ro_detail(cr,srec,right_now,sysur_auto_key,roh_auto_key=roh_auto_key)
                        count += 1
                    msg = str(rods_created) + ' detail line(s) created and added to RO# '+ str(ro_number) + '.'
            else:
                error = 'RO not found.'
                              
        elif vendor:
            #1. Simply create RO_HEADER FROM the vendor that the user entered in the pop-up
            #2. Insert RO_DETAIL LINES (with detail line from stock line in question)
            #     into existing (user-entered RO from step #1) 
            #Loop through the stock lines and add an RO_DETAIL to the 
            #vendors need to be created and accessed in dropdown on pop-up form
            #stock_recs = get_stock_ros(cr,stm_auto_key=stm_keys[0])
            #srec = stock_recs and stock_recs[0] or [] 
            #syscm_auto_key,dpt_auto_key
            #ship_address1, ship_address2, ship_address3, ship_address4,
            #ship_address5, ship_name,
            #vendor_address1, vendor_address2, vendor_address3, 
            #vendor_address4, vendor_address5, vendor_name
            #srecs[12],srecs[10],srecs[13],srecs[14],srecs[15],srecs[16],srecs[17],srecs[18],srecs[19],srecs[20],srecs[21],srecs[22],srecs[23],srecs[24]
            vendor_key = vendor
            if not vendor_key:
                return 'No vendor found.',''
            else:
                query = """
                SELECT NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,
                SYSCM_AUTO_KEY,SHIP_ADDRESS1,SHIP_ADDRESS2,SHIP_ADDRESS3,NULL,NULL,
                SHIP_NAME,ADDRESS1,ADDRESS2,ADDRESS3,NULL,NULL,COMPANY_NAME
                FROM COMPANIES WHERE CMP_AUTO_KEY = %s
                """%vendor_key
                vendor_info = selection_dir(query,cr)
                vrec = vendor_info and vendor_info[0] or [] 
            if vendor_key:
                error,rohs_created = insert_ro_header(cr,sysur_auto_key,vendor_key,right_now,vrec) or ''
                roh_auto_key = get_new_roh(cr,vendor_key,right_now,sysur_auto_key=sysur_auto_key)
                rods_created = 0
                if error == '{"recs": ""}':
                    error = ''                
                if not error and roh_auto_key:
                    count = 0
                    for stm_key in stm_keys:                     
                        srec = get_stock_ros(cr,stm_auto_key=stm_key)
                        srec = srec and srec[0] or []                          
                        #get vendor from stock record
                        rods_created += insert_ro_detail(cr,srec,right_now,sysur_auto_key,roh_auto_key=roh_auto_key)
                        count += 1
                    msg = 'RO#(s): ' + str(rohs_created) + ', created with ' + str(rods_created) + ' detail line(s).'                  
            else:
                error = "'%s' is not a vendor id in the database."%vendor            
        else:
            #get vendor in a pop-up
            return '','show_modal' 
    #register audit trail record
    if msg and msg != 'show_modal' and (not error or error == '{"recs": ""}'):
        orcl_commit(con=con)      
        error = ''        
        stock_lines_del = wos_obj.objects.filter(session_id=session_id,stm_auto_key__in = stm_keys)
        stock_lines_del = stock_lines_del and stock_lines_del.delete()
    aud_status = 'success'
    app_id = maps.objects.filter(code='repair-order-mgmt')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        field_changed = 'Added new RO.'
        new_val = 'Added new RO.'
        if error:             
            aud_status = 'failure'
            new_val = error
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,msg
 
def get_ro_vendor(srec,session_id,wos_obj):
    moves = wos_obj.objects.filter(session_id=session_id,stm_auto_key=srec[8])
    last_vendor_key = None
    for move in moves:
        if move.vendor_key:
            last_vendor_key = move.vendor_key
            break
    return last_vendor_key
    
def get_new_roh(cr,vendor_key,right_now,sysur_auto_key=''):
    sysur_auto_key = sysur_auto_key and 'AND SYSUR_AUTO_KEY = %s '%sysur_auto_key or '' 
    entry_date = ''   
    if sysur_auto_key:
        entry_date = ''
    else:
        entry_date = "AND ENTRY_DATE=TO_DATE('%s','MM-DD-YYYY')"%right_now  
    query = """SELECT ROH_AUTO_KEY FROM RO_HEADER WHERE 
        CMP_AUTO_KEY = %s 
        %s
        %s
        ORDER BY ROH_AUTO_KEY 
        DESC"""%(vendor_key,entry_date,sysur_auto_key)
    roh = selection_dir(query,cr)
    roh_auto_key = roh and roh[0] or None
    roh_auto_key = roh_auto_key and roh_auto_key[0] or None
    return roh_auto_key
      
def get_stored_vendor(vendor_id):
    #method that takes a user-entered vendor auto key and looks it up in local db
    from polls.models import Companies as comp
    recs = comp.objects.filter(cmp_auto_key = vendor_id)
    vendor = recs and recs[0] or []
    vendor = vendor and vendor.cmp_auto_key or None
    return vendor
    
def create_bulk_departments(recs,quapi_id,dj_user_id):
    from polls.models import Departments as depts
    dept_data = []
    error = ''
    for rec in recs:
        dept_data.append(depts(
                dpt_auto_key = rec[1],#0
                name = rec[0],#1 
                dj_user_id = dj_user_id,
                quapi_id = quapi_id,
        ))                   
    try:
        departments = dept_data and depts.objects.bulk_create(dept_data) or []    
    except Exception as exc:
        error = "Error with creating departments locally: %s"%exc 
    return error
    
@shared_task
def get_depts_n_sync(quapi_key,dj_user_id):
    error  = ''   
    from polls.models import Departments as dept,QueryApi as qa
    query = "SELECT DEPT_NAME,DPT_AUTO_KEY FROM DEPARTMENT"
    quapi = qa.objects.filter(id=quapi_key)   
    quapi_id = quapi and quapi[0] or None
    recs = quapi_id and selection(query,quapi=quapi_id) or []
    if recs:   
        departments = dept.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id)
        del_vendors = departments and departments.delete() or None
    else:
        return 'Error with creating departments locally.',''    
    error = recs and create_bulk_departments(recs,quapi_id,dj_user_id) or ''
    return error
    
    
def create_bulk_companies(recs,quapi_id,dj_user_id,is_acc_co=False,is_customer=False,is_vendor=False):
    from polls.models import Companies as comp
    comp_data = []
    error = ''
    for rec in recs:
        comp_data.append(comp(
                cmp_auto_key = rec[1],#0
                name = rec[0],#1 
                dj_user_id = dj_user_id,
                quapi_id = quapi_id,
                is_vendor = is_vendor,
                is_customer = is_customer,
                is_acc_co = is_acc_co,
        ))                   
    try:
        companies = comp_data and comp.objects.bulk_create(comp_data) or []    
    except Exception as exc:
        error = "Error with creating companies: %s"%exc 
    return error
 
@shared_task
def get_sys_companies_n_sync(quapi_key,dj_user_id):
    error  = ''
    #get all vendors
    #get quapi object first
    from polls.models import QueryApi as qa, Companies
    query = "SELECT COMPANY_NAME,SYSCM_AUTO_KEY FROM SYS_COMPANIES"
    quapi = qa.objects.filter(id=quapi_key)   
    quapi_id = quapi and quapi[0] or None
    recs = quapi_id and selection(query,quapi=quapi_id) or []
    if recs:   
        companies = Companies.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id,is_acc_co=True)
        del_vendors = companies and companies.delete() or None
    else:
        return 'Error with creating companies locally.',''    
    error = recs and create_bulk_companies(recs,quapi_id,dj_user_id,is_acc_co=True) or ''
    return error
    
@shared_task
def get_companies_n_sync(quapi_key,dj_user_id,is_vendor=False,is_customer=False,is_acc_co=False):
    error  = ''
    #get all vendors
    #get quapi object first
    where_cust,where_vend = '',''
    if is_vendor:
        where_vend = "VENDOR_FLAG = 'T' AND"
    if is_customer:
        where_cust = "CUSTOMER_FLAG = 'T' AND "   
    from polls.models import QueryApi as qa, Companies
    query = "SELECT COMPANY_NAME,CMP_AUTO_KEY FROM COMPANIES WHERE %s %s HISTORICAL = 'F'"%(where_vend,where_cust)
    quapi = qa.objects.filter(id=quapi_key)   
    quapi_id = quapi and quapi[0] or None
    recs = quapi_id and selection(query,quapi=quapi_id) or []
    if recs:   
        companies = Companies.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id)
        del_vendors = companies and companies.delete() or None
    else:
        return 'Error with creating companies locally.',''    
    error = recs and create_bulk_companies(recs,quapi_id,dj_user_id,is_acc_co=is_acc_co,is_customer=is_customer,is_vendor=is_vendor) or ''
    return error

def get_ro_header(cr,ro_number=None):   
    if ro_number:
        query = "SELECT RO_NUMBER,ROH_AUTO_KEY FROM RO_HEADER WHERE RO_NUMBER='%s'"%ro_number   
        recs = selection_dir(query,cr)
    return recs    

@shared_task
def run_ro_edit(session_id,quapi_id,\
    ro_number=None,wo_number=None,\
    vendor=None,part_number=None,\
    entry_date=None,rst_auto_key=None,uda_status=None):
    error,msg = '',''
    from polls.models import WOStatus as wos_obj,QueryApi,MLApps,QuantumUser as qu
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.'      
    ro_recs = get_ro_details(cr,ro_number=ro_number,\
        wo_number=wo_number,vendor=vendor,\
        part_number=part_number,\
        entry_date=entry_date,rst_auto_key=rst_auto_key,\
        uda_status=uda_status)
    if ro_recs:
        del_ros = wos_obj.objects.filter(session_id=session_id).delete() or None
        error,msg = create_ros_bulk(quapi,session_id,ro_recs)
    else:
        error = 'No records found.' 
    return error,msg
    
@shared_task
def split_ro(
    session_id,
    quapi_id,
    quantity,
    quantity_reserved,
    misc_cost=None,
    parts_cost=None,
    labor_cost=None,
    approved_date=None,
    quoted_date=None,
    next_dlv_date=None,
    notes='',
    ro_id_list=[],
    stm_auto_key=None,
    sysur_auto_key=None,
    ro_number = None,
    vendor = None,
    part_number = None,
    entry_date = None):
    error,msg = '',''
    from polls.models import WOStatus as wos_obj,QueryApi,MLApps as maps,QuantumUser as qu
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.'     
    upd_old = ''
    #first, we get the stock line for the ROD 
    #delete the reservation (if it exists)
    #Original ROD - adjust the qty_repair by subtracting off the user-entered qty
    #create a new ROD - user-entered qty = qty_repair 
    #re-reserve the original ROD and create reservation for the new one
    #TODO: 1. put in the qty_reserved here and for def insert_ro_detail(sum of all qty_reserved on stock_reservations for that rod_auto_key) 
    #2. 
    for rod_old in ro_id_list: 
        if quantity and (int(quantity) > int(quantity_reserved)):
            return 'Must enter a quantity <= ' + quantity_reserved + ' to split the RO Detail line.'
        next_item_qry = """SELECT MAX(ROD.ITEM_NUMBER)+1 FROM RO_DETAIL ROD          
            WHERE ROD.ROH_AUTO_KEY IN (SELECT ROH_AUTO_KEY FROM RO_DETAIL WHERE ROD_AUTO_KEY = %s) 
            AND ROWNUM <= 1"""%rod_old
        vals_query = """
            SELECT ROD.PNM_AUTO_KEY,ROD.ROH_AUTO_KEY,
            %s,ROD.ENTRY_DATE,%s,
            ROD.WOB_AUTO_KEY,ROD.PCC_AUTO_KEY,
            (%s),
            '%s','OEM','O','Overhaul','%s','%s','%s',TO_DATE('%s','mm/dd/yyyy'),TO_DATE('%s','mm/dd/yyyy'),TO_DATE('%s','mm/dd/yyyy')
            FROM RO_DETAIL ROD WHERE ROD.ROD_AUTO_KEY ='%s'
        """%(sysur_auto_key,quantity,next_item_qry,notes,labor_cost,misc_cost,parts_cost,quoted_date,approved_date,next_dlv_date,rod_old)
        #find the correct row
        insert_ro_query = """
            INSERT INTO RO_DETAIL (
            PNM_AUTO_KEY,
            ROH_AUTO_KEY,
            SYSUR_AUTO_KEY,
            ENTRY_DATE,
            QTY_REPAIR,
            WOB_AUTO_KEY,
            PCC_AUTO_KEY,
            ITEM_NUMBER,
            REPAIR_NOTES,
            RO_TYPE,
            ROUTE_CODE,
            ROUTE_DESC,
            LABOR_COST,
            MISC_COST,
            PARTS_COST,
            MSG_QUOTE_REC_DATE,
            MSG_QUOTE_APPR_DATE,
            NEXT_DELIVERY_DATE)
            %s"""%(vals_query)
        new_ro = insertion_dir(insert_ro_query,cr)
        if stm_auto_key and int(quantity) and int(quantity) > 0:
            """"""
            """stock_res = selection_dir(query,cr)
            if len(stock_res) > 1:
                return '','show_grid_modal'"""
            del_str = "DELETE FROM STOCK_RESERVATIONS WHERE ROD_AUTO_KEY=%s AND STM_AUTO_KEY = %s AND QTY_RESERVED >0"%(rod_old,stm_auto_key)
            dels = updation_dir(del_str,cr)
        else:
            return 'That RO Detail has no stock reservation. Choose another please.',''
        if quantity_reserved and quantity_reserved > 0:
            upd_old = "UPDATE RO_DETAIL SET QTY_REPAIR = QTY_REPAIR - %s WHERE ROD_AUTO_KEY = %s"%(quantity_reserved,rod_old)
            old = updation_dir(upd_old,cr)
            rod_new = "SELECT ROD_AUTO_KEY FROM RO_DETAIL WHERE ROWNUM<=1 ORDER BY ROD_AUTO_KEY DESC"
            rod_new = selection_dir(rod_new,cr)
            rod_new = rod_new and rod_new[0] and rod_new[0][0] or None
            
            new_res = """INSERT INTO STOCK_RESERVATIONS (STM_AUTO_KEY, ROD_AUTO_KEY, QTY_RESERVED) VALUES (%s,%s,%s)"""%(stm_auto_key,rod_new,quantity)
            newrs = insertion_dir(new_res,cr)
            """
            if rod_new:
                upd_new = "UPDATE RO_DETAIL SET QTY_REPAIR = QTY_REPAIR + %s WHERE ROD_AUTO_KEY = %s"%(quantity_reserved,rod_new)
                new = updation_dir(upd_new,cr)
                if new == None:
                    error = 'There was a problem and we could not update existing ro_detail record'"""
        #register audit trail record  
        #active_ros = wos_obj.objects.filter(session_id=session_id)
        #rod_complete = active_ros and active_ros.values_list('rod_auto_key',flat=True)
        #rod_complete = rod_complete and construct_akl(rod_complete) or [] 
        #ro_recs = []
        #for rods in rod_complete:
        #    ro_recs = get_ro_details(cr,ro_id_list=rods)
        ro_recs = get_ro_details(cr,ro_number=ro_number,vendor=vendor,part_number=part_number,entry_date=entry_date)
        if ro_recs:
            del_ros = wos_obj.objects.filter(session_id=session_id).delete() or None
            error,msg = create_ros_bulk(quapi,session_id,ro_recs)
        orcl_commit(con=con)
    aud_status = 'success'
    app_id = maps.objects.filter(code='repair-order-mgmt')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        field_changed = upd_old
        new_val = ''
        if error == '{"recs": ""}':
            error = ''    
        if not error:
            new_val = 'Successfully split 1 RO.' 
        else:             
            aud_status = 'failure'
            new_val = error 
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)     
    else:
        error = 'Incorrect Quantum User ID.'
    return error,msg + new_val
    
@shared_task
def ro_update_costs_dates(
    session_id,
    quapi_id,
    misc_cost=0,
    labor_cost=0,
    parts_cost=0,
    approved_date = None,
    quoted_date = None,
    next_dlv_date = None,
    ro_id_list = [],
    airway_bill=None,
    pnm_modify=None,
    cond_code=None,
    notes='',
    uda_status=None,
    header_notes='',
    sysur_auto_key=None,
    new_status=None,
    ro_number='',
    vendor='',
    part_number='',
    entry_date=None,
    wo_number='',
    receiver_instr='',
    est_quote_date=None,
    ship_date=None,
    ro_categ='',filters=[]):
    roh_update_set,ro_lists,error,msg,update_set,ro_ids,update_msg = '',[],'','','',[],''
    ro_recs = []    
    from polls.models import QueryApi,WOStatus,MLApps as maps,QuantumUser as qu   
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None 
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.' 
    #prepare query with parameters for updating RO Detail table
    query = "UPDATE RO_DETAIL SET "

    if ro_id_list:
        num_to_update = len(ro_id_list)
        ro_lists = construct_akl(ro_id_list) 
        if uda_status or header_notes:
            if header_notes:
                header_notes = header_notes.replace("'", "").\
                    replace("\r", "").replace("\n", ";").replace("\t", "")
                header_notes = header_notes.replace(";;",";")
                roh_update_set += "NOTES = NOTES || '; ' || TO_CLOB('%s')"%header_notes
                
            if uda_status:
                for ro_id in ro_id_list:
                
                    query = """SELECT ROH_AUTO_KEY FROM RO_DETAIL
                        WHERE ROD_AUTO_KEY = %s
                        """%ro_id
                    roh = selection_dir(query,cr)
                    roh_auto_key = roh and roh[0] and roh[0][0] or None
                    if roh_auto_key:
                        query = """SELECT AUTO_KEY,ATTRIBUTE_VALUE FROM UDA_CHECKED
                            WHERE AUTO_KEY = %s AND UDA_AUTO_KEY = 64
                            """%roh_auto_key
                        recs = selection_dir(query,cr)
                        uda = recs and recs[0] and recs[0][1] or None

                        if uda and uda != uda_status:
                            query = """UPDATE UDA_CHECKED SET ATTRIBUTE_VALUE = '%s'
                            WHERE AUTO_KEY = %s AND UDA_AUTO_KEY = 64
                            """%(uda_status,roh_auto_key)
                            error = updation_dir(query,cr)
                            
                        elif not uda:    
                            query = """INSERT INTO UDA_CHECKED (AUTO_KEY,ATTRIBUTE_VALUE,UDA_AUTO_KEY)
                                VALUES(%s,'%s',64)"""%(roh_auto_key,uda_status)
                            error = insertion_dir(query,cr)
                
            for rol in ro_lists:
                notes_query="""SELECT DISTINCT ROH_AUTO_KEY,TO_CHAR(NOTES) 
                    FROM RO_HEADER
                    WHERE ROH_AUTO_KEY IN (SELECT ROH_AUTO_KEY FROM RO_DETAIL 
                    WHERE ROD_AUTO_KEY IN %s)"""%rol
                rohs = selection_dir(notes_query,cr)
                for roh in rohs:
                    
                    roh_update_fin = roh_update_set and ',' or ''
                    if roh[1]:
                        roh_update_fin += "NOTES = NOTES || '; ' || TO_CLOB('%s')"%header_notes
                    else:
                        roh_update_fin += "NOTES = TO_CLOB('%s')"%header_notes
                    upd_roh = """UPDATE RO_HEADER SET %s 
                        WHERE ROH_AUTO_KEY = '%s'"""%(roh_update_set,roh[0]) 
                    error = updation_dir(upd_roh,cr)                        
        if airway_bill:
            update_set += update_set and ',' or ''
            update_set += "MSG_AIRWAY_BILL = '%s'"%airway_bill    
        if pnm_modify:
            pnmmod_where = "UPPER(PN) =  UPPER('%s')"%pnm_modify
            update_set += update_set and ',' or ''
            update_set += "PNM_MODIFY = (SELECT PNM_AUTO_KEY FROM PARTS_MASTER WHERE %s)"%pnmmod_where 
        if cond_code:
            cond_code_where = "UPPER(CONDITION_CODE) =  UPPER('%s')"%cond_code
            update_set += update_set and ',' or ''
            update_set += "PCC_AUTO_KEY = (SELECT PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE %s)"%cond_code_where         
        if misc_cost:
            update_set += update_set and ',' or ''
            update_set += 'MISC_COST = EXCHANGE_RATE*(%s),FOREIGN_MISC = EXCHANGE_RATE*(%s)'%(misc_cost,misc_cost)    
        if parts_cost:
            update_set += update_set and ',' or ''
            update_set += 'PARTS_COST = EXCHANGE_RATE*(%s),FOREIGN_PARTS = EXCHANGE_RATE*(%s)'%(parts_cost,parts_cost)
        if notes:
            notes = notes.replace("'", "").replace("\r", ";").replace("\n", ";").replace("\t", ";") 
            notes = notes.replace(";;",";")            
            update_set += update_set and ',' or ''
            update_set += "REPAIR_NOTES = REPAIR_NOTES || '; ' || TO_CLOB('%s')"%notes or '' 
        if labor_cost:
            update_set += update_set and ',' or ''
            update_set += 'LABOR_COST = EXCHANGE_RATE*(%s),FOREIGN_LABOR = EXCHANGE_RATE*(%s)'%(labor_cost,labor_cost) 
        if approved_date:
            update_set += update_set and ',' or ''
            update_set += "MSG_QUOTE_APPR_DATE = TO_DATE('%s', 'mm/dd/yyyy')"%approved_date    
        if quoted_date:
            update_set += update_set and ',' or ''
            update_set += "MSG_QUOTE_REC_DATE = TO_DATE('%s', 'mm/dd/yyyy')"%quoted_date  
        if next_dlv_date:
            update_set += update_set and ',' or ''
            update_set += "NEXT_DELIVERY_DATE = TO_DATE('%s', 'mm/dd/yyyy')"%next_dlv_date
        if est_quote_date:
            update_set += update_set and ',' or ''
            update_set += "RO_UDF_001 = TO_DATE('%s', 'mm/dd/yyyy')"%est_quote_date
        if ship_date:
            update_set += update_set and ',' or ''
            update_set += "COMMIT_SHIP_DATE = TO_DATE('%s', 'mm/dd/yyyy')"%ship_date
        if receiver_instr:
            update_set += update_set and ',' or ''
            update_set += "RECEIVER_INSTR = TO_CLOB('%s')"%receiver_instr
        if ro_categ:
            update_set += update_set and ',' or ''
            update_set += """RCT_AUTO_KEY = %s"""%ro_categ

        if labor_cost or parts_cost or misc_cost:
            for rod in ro_id_list:
                set_clause = ''
                update_rod = "UPDATE RO_DETAIL SET "
                query ="""select quoted_labor_home,quoted_parts_home,
                    quoted_misc_home from ro_detail where rod_auto_key=%s"""%rod

                rec = selection_dir(query,cr)
                rec = rec[0]
                qlh = rec[0] or 0
                qph = rec[1] or 0
                qmh = rec[2] or 0
                if rec:
                    if labor_cost and qlh <= 0:
                        set_clause += set_clause and ',' or ''
                        set_clause += """QUOTED_LABOR_HOME=EXCHANGE_RATE*(%s),
                        QUOTED_LABOR_FOREIGN=EXCHANGE_RATE*(%s)"""%(labor_cost,labor_cost)
                        
                    if parts_cost and qph <= 0:
                        set_clause += set_clause and ',' or ''
                        set_clause += """QUOTED_PARTS_HOME=EXCHANGE_RATE*(%s),
                        QUOTED_PARTS_FOREIGN=EXCHANGE_RATE*(%s)"""%(parts_cost,parts_cost)
                       
                    if misc_cost and qmh <= 0:
                        set_clause += set_clause and ',' or ''
                        set_clause += """QUOTED_MISC_HOME=EXCHANGE_RATE*(%s),
                        QUOTED_MISC_FOREIGN=EXCHANGE_RATE*(%s)"""%(misc_cost,misc_cost)
                        
                    where_clause = " WHERE ROD_AUTO_KEY = %s"%rod
                    
                    if set_clause:
                        query = update_rod + set_clause + where_clause                   
                        error = updation_dir(query,cr)
                        
        for rol in ro_lists:
            update_rod = "UPDATE RO_DETAIL SET "
            where_ids = " WHERE ROD_AUTO_KEY IN %s"%rol
            query = update_rod + update_set + where_ids
            
            if update_set:
                error = updation_dir(query,cr)
                
            elif not error:
                error = 'Need a value in the date, cost or the other fields to update.  Try again.'
            """if ro_lists and (update_set or roh_update_set or roh_update_fin) and error == '{"recs": ""}':
                update_msg = str(num_to_update) + ' records updated.'
                active_ros = WOStatus.objects.filter(session_id=session_id)
                rod_complete = active_ros and active_ros.values_list('rod_auto_key',flat=True)
                rod_complete = rod_complete and construct_akl(rod_complete) or [] 
                for rods in rod_complete:
                    ro_recs = get_ro_details(cr,ro_id_list=rods)
                if ro_recs:
                    del_ros = WOStatus.objects.filter(session_id=session_id).delete() or None
                    error,msg = create_ros_bulk(quapi,session_id,ro_recs)
                #else:
                #    error = 'Error creating records locally. Try again please.'"""
        orcl_commit(con=con)                
    #register audit trail record           
    aud_status = 'success'
    app_id = maps.objects.filter(code='repair-order-mgmt')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    
    if user_rec:
        field_changed = update_set + ' ' + roh_update_set
        if error == '{"recs": ""}' or not error:
            error = ''
            msg = 'Successful update.'
            
        if not error:       
            new_val = 'Repair order detail records updated: '
            new_val += ' for RO: %s'%ro_number
            new_val += ' RO Detail lines list: %s'%str(ro_lists)
        
        else:             
            aud_status = 'failure'
            new_val = error 
            field_changed = 'Nothing changed.'
            
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,msg
    
def update_stock_audit_log(cr,sysur_auto_key,stm_auto_key,user_id,gla_auto_key):
    query = """
        CREATE OR REPLACE PROCEDURE "GLA_STOCK_UPDATE"
        (QUSER IN NUMBER, STM IN NUMBER, QCODE IN VARCHAR2, GLA_KEY IN NUMBER)  AS
        v_query number;
        v_sysur number;
        v_pwd varchar2(150);
        V_CURSOR QC_SC_PKG.CURSOR_TYPE ;
        BEGIN                
            begin
            qc_trig_pkg.disable_triggers;
            UPDATE SA_LOG SET SYSUR_AUTO_KEY = QUSER, EMPLOYEE_CODE = QCODE WHERE STA_AUTO_KEY = (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = STM AND NEW_GLA_AUTO_KEY=GLA_KEY AND EMPLOYEE_CODE = 'DBA');
            qc_trig_pkg.enable_triggers;
            end;
         END GLA_STOCK_UPDATE;"""   
    error = updation_dir(query,cr)   
    run_proc = """
        BEGIN
        GLA_STOCK_UPDATE('%s',%s,'%s',%s);
        END;   
    """%(sysur_auto_key,stm_auto_key,user_id[:9],gla_auto_key) 
    error = updation_dir(run_proc,cr)
    return error
    
def update_trail(sysur_auto_key, woo_auto_key, new_status=None, new_location=None, stm_auto_key=None, quapi=None, is_rack=False, user_id = None, new_mgr=None, new_due_date=None, new_rank=None):
    #parameters: user_id - integer ID for PK to the SYS_USERS table
    #            woo_auto_key - Quantum db primary key for user input WO#
    #            new_status - user input status to which to update
    #            new_location - user input location (converted to database id for loc_auto_key in STM table)
    msg = ''
    q1,q2 = '',''
    if is_rack and new_status:
        descr = 'MRO Live update status - barcoding app'
        rack_auto_key = woo_auto_key
        woo_auto_keys ="""(SELECT ST.WOO_AUTO_KEY FROM STOCK_RESERVATIONS 
        ST WHERE ST.STM_AUTO_KEY IN (SELECT STM_AUTO_KEY FROM STOCK
        WHERE IC_UDL_005=%s))"""%rack_auto_key       
        query = """
            UPDATE AUDIT_TRAIL SET DESCR = '%s',SYSUR_AUTO_KEY=%s WHERE SOURCE_TABLE = 'WOO' AND SOURCE_AK IN %s AND SOURCE_FIELD = 'STATUS'
            AND NEW_VALUE = (SELECT DESCRIPTION FROM WO_STATUS WHERE WOS_AUTO_KEY = '%s')"""%(descr,sysur_auto_key, woo_auto_keys, new_status)
        msg = updation(query,quapi=quapi) 
    if not sysur_auto_key:
        return 'No user is logged in and therefore no update can be made to the audit trail table.'
    if new_rank:
        descr = 'MRO Live update rank - wo mgmt app'
        query = """
            UPDATE AUDIT_TRAIL SET DESCR = '%s',SYSUR_AUTO_KEY=%s WHERE SOURCE_TABLE = 'WOO' AND SOURCE_AK = %s AND SOURCE_FIELD = 'RANK'
            AND NEW_VALUE = %s"""%(descr,sysur_auto_key, woo_auto_key, new_rank)
        msg = updation(query,quapi=quapi)   
    if new_due_date:
        descr = 'MRO Live update due date - wo mgmt app'
        query = """
            UPDATE AUDIT_TRAIL SET DESCR = '%s',SYSUR_AUTO_KEY=%s WHERE SOURCE_TABLE = 'WOO' AND SOURCE_AK = %s AND SOURCE_FIELD = 'DUE_DATE'
            AND NEW_VALUE = TO_DATE('%s', 'mm-dd-yyyy')"""%(descr,sysur_auto_key, woo_auto_key, new_due_date)
        msg = updation(query,quapi=quapi)           
    if not is_rack and new_status:
        descr = 'MRO Live update status - barcoding app.'
        query = """
            UPDATE AUDIT_TRAIL SET DESCR = '%s',SYSUR_AUTO_KEY=%s WHERE SOURCE_TABLE = 'WOO' AND SOURCE_AK = %s AND SOURCE_FIELD = 'STATUS'
            AND NEW_VALUE = (SELECT DESCRIPTION FROM WO_STATUS WHERE WOS_AUTO_KEY = '%s')"""%(descr,sysur_auto_key, woo_auto_key, new_status)
        msg = updation(query,quapi=quapi)  
    return str(msg)
    
def check_if_same_status(wos_auto_key, woo_auto_key, user_id='', quapi=None):
    """
    Arguments: status - wos_auto_key supplied by user form input
            wo_number - SI_NUMBER supplied by user form input
            
    Returns: query results from function call to selection()     
    """
    update_ok = False
    query = """
        SELECT WOS_AUTO_KEY,SI_NUMBER FROM WO_OPERATION
            WHERE
            WOO_AUTO_KEY = %s"""%woo_auto_key
    results = selection(query, user_id=user_id,quapi=quapi)    
    if results and results[0] and results[0][0] != wos_auto_key:
        wos_auto_key = results[0][0]
        update_ok = True           
    return update_ok
    
def check_if_same_loc(location_code, woo_auto_key, user_id='',quapi=None):
    """
    Arguments: location - code supplied by user form input
            woo_auto_key - from a lookup of SI_NUMBER (wo_number) supplied by user form input           
    Returns: results from location       
    """
    update_ok = False
    stm_auto_key = ''
    #location code is what they enter/scan => from location - lookup location table to find location code
    query = """SELECT S.STM_AUTO_KEY, L.LOC_AUTO_KEY, L.LOCATION_CODE FROM WO_OPERATION W, WO_STATUS WS,STOCK_RESERVATIONS SR, STOCK S, LOCATION L WHERE W.WOS_AUTO_KEY = WS.WOS_AUTO_KEY AND W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY AND SR.STM_AUTO_KEY = S.STM_AUTO_KEY AND S.LOC_AUTO_KEY = L.LOC_AUTO_KEY AND W.WOO_AUTO_KEY = %s"""%(woo_auto_key)
    results = selection(query, user_id=user_id,quapi=quapi)
    if not results:
        query = """SELECT S.STM_AUTO_KEY, L.LOC_AUTO_KEY, L.LOCATION_CODE FROM WO_OPERATION W, STOCK_RESERVATIONS SR, STOCK S, LOCATION L WHERE W.WOS_AUTO_KEY IS NULL AND W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY AND SR.STM_AUTO_KEY = S.STM_AUTO_KEY AND S.LOC_AUTO_KEY = L.LOC_AUTO_KEY AND W.WOO_AUTO_KEY = %s"""%(woo_auto_key)
        results = selection(query, user_id=user_id,quapi=quapi)        
    result = results and results[0] or None
    stm_auto_key = result and result[0] or None
    if type(stm_auto_key) is list:
        stm_auto_key = stm_auto_key[0]
    if result and result[2] and result[2] != location_code:
        update_ok = True           
    return stm_auto_key,update_ok
    
def set_status(woo_auto_key, new_status_key, user_id='', quapi=None):
    recs = []
    error = ''
    timestamp = ''
    #update_status = woo_auto_key and new_status_key and check_if_same_status(new_status_key, woo_auto_key,user_id=user_id,quapi=quapi) or False 
    if user_id and woo_auto_key:
        query = "UPDATE WO_OPERATION SET WOS_AUTO_KEY = %s WHERE WOO_AUTO_KEY = %s AND WOS_AUTO_KEY <> %s OR WOS_AUTO_KEY IS NULL"%(new_status_key,woo_auto_key,new_status_key)
        error = updation(query,quapi=quapi)
    return str(error) 
    
def synch_record(wos_obj,user_id,woo_auto_key=None,wo_number='',new_status = None,new_location = None,update_stamp=None,quapi=None):
    if not user_id:
        msg = 'You must enter an employee ID first before updating any workorder statuses.'
        return msg 
    from polls.models import StatusSelection       
    msg = ''
    wos_status = '' 
    wos = wos_obj.objects.filter(is_racking=0, woo_auto_key=woo_auto_key, active=1, is_dashboard=0, user_id=user_id)
    right_now = datetime.now()
    right_now = right_now.strftime('%Y-%m-%d %H:%M:%S')       
    if new_status:
        stat_sel = StatusSelection.objects.filter(wos_auto_key=new_status)
        wos_status = stat_sel and stat_sel[0] and stat_sel[0].name or '-- No Status --'
        if woo_auto_key and wos:   
            wos.update(status = wos_status, update_stamp = update_stamp)          
    #if user didn't select a status and entered a WO that doesn't already
    #exist or didn't enter one at all, we create the WO locally.
    if not update_stamp: update_stamp = None
    if woo_auto_key and not wos:
        recs = get_wo_status(woo_auto_key, quapi=quapi, user_id=user_id)
        if recs:
            woo = recs[0]
            due_date = woo[2] and woo[2].strftime('%Y-%m-%d') or None
            status = str(woo[8]) + ' - ' + str(woo[1])
            try:
                wos = wos_obj.objects.create(
                    wo_number = wo_number,
                    supdate_msg = ' Successful update',
                    status = status,
                    due_date = due_date,
                    stock_line = woo[3],
                    part_number = woo[4],
                    description = woo[5],
                    serial_number = woo[6],
                    location_code = new_location or woo[7],
                    active = 1,
                    update_stamp = update_stamp,
                    user_id = user_id,
                    is_dashboard = 0,
                    woo_auto_key=woo_auto_key,
                )
                wos.save() 
            except Exception as error:
                logger.error("Error with creation of wo locally. Message: '%s'",error.args)
        else:
            msg = "The WO you entered, '%s', doesn't exist."%wo_number
    #if user entered both WO and status to update to and wos exists locally already:
    elif wos and new_status: 
        wos.update(status = wos_status or '-- no status --', update_stamp = update_stamp)        
    elif wos and woo_auto_key and not (new_status or new_location):
        msg = 'We have this WO in the list already.  You must do one of the following:  a. enter a valid WO# and a\
        status or b. select one or more WO\'s to remove or c. enter only a status to update all workorders on this list.'
    return msg
  
def update_loc(location_code, loc_auto_key, woo_auto_key, user_id, wos_obj,quapi=None):
    msg = ''
    stm_auto_key,update_ok = check_if_same_loc(location_code, woo_auto_key, user_id=user_id,quapi=quapi) 
    if update_ok and stm_auto_key and loc_auto_key: 
        #build the query that updates 'STOCK' with the new 
        #LOC_AUTO_KEY obtained by the output from the 
        #check_if_same_loc() call.    
        query = "UPDATE STOCK SET LOC_AUTO_KEY = %s WHERE STM_AUTO_KEY = %s"%(loc_auto_key,stm_auto_key)      
        msg = updation(query, user_id=user_id,quapi=quapi)
    woo_object = wos_obj.objects.filter(woo_auto_key=woo_auto_key, is_dashboard=0, is_racking=0, active=1, user_id=user_id)
    woo_object = woo_object and woo_object.update(location_code=location_code)           
    return msg
    
def get_woo_key(wo_number, user_id='',quapi=None):
    woo_auto_key = None
    query = "SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER = '%s' ORDER BY WOO_AUTO_KEY DESC"%wo_number
    woo_auto_key = selection(query, user_id=user_id, quapi=quapi)
    woo_auto_key = woo_auto_key and woo_auto_key[0] and woo_auto_key[0][0] or None
    return woo_auto_key
    
@shared_task
def get_loc_whs_cart_nsync_beta(quapi_key,dj_user_id,loc_ok=False,whs_ok=False,cart_ok=False,app=''):
    msg = ''
    error = ''
    joins,selects,where_clause = '','',''
    from polls.models import QueryApi as qa,Location as loc,Warehouse as whs,StockCart as scart    
    quapi = qa.objects.filter(id=quapi_key)
    quapi_id = quapi and quapi[0] or None  
    from polls.models import OracleConnection as oc
    orcl_conn = quapi_id and quapi_id.orcl_conn_id and oc.objects.filter(id=quapi_id.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle' 
    #dj_user_id = quapi and quapi.dj_user_id or None
    del_locs = loc_ok and dj_user_id and loc.objects.filter(quapi_id=quapi_key,dj_user_id=dj_user_id).delete() or None
    del_whss = whs_ok and dj_user_id and whs.objects.filter(quapi_id=quapi_key,dj_user_id=dj_user_id).delete() or None
    del_carts = cart_ok and dj_user_id and scart.objects.filter(quapi_id=quapi_key,dj_user_id=dj_user_id).delete() or None
    #begin creating query
    selects += " DISTINCT L.LOC_AUTO_KEY,L.LOCATION_CODE,WH.WHS_AUTO_KEY,WH.WAREHOUSE_CODE,WHL.LOC_AUTO_KEY,L.IQ_ENABLE"
    joins += """ LEFT JOIN WAREHOUSE_LOCATIONS WHL ON WHL.LOC_AUTO_KEY=L.LOC_AUTO_KEY 
    LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY = WHL.WHS_AUTO_KEY"""
    query = 'SELECT' + selects + ' FROM LOCATION L' + joins + where_clause
    recs = selection_dir(query,cr) or []
    joins = """ LEFT JOIN WAREHOUSE_LOCATIONS WHL ON WHL.WHS_AUTO_KEY=WH.WHS_AUTO_KEY 
    LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = WHL.LOC_AUTO_KEY"""
    where_clause = " WHERE WHL.WHS_AUTO_KEY IS NULL AND L.HISTORICAL <> 'T' AND WH.HISTORICAL <> 'T'"
    query = 'SELECT' + selects + ' FROM WAREHOUSE WH' + joins + where_clause
    sec_recs = selection_dir(query,cr)
    if sec_recs:
        recs += sec_recs
    if not recs:
        return 'Error with creating locations and/or warehouses locally.',''
    del_locs = whs.objects.filter(dj_user_id=dj_user_id,quapi_id = quapi_id).delete() 
    
    add_em = recs and create_bulk_locs(recs,quapi_id,dj_user_id,loc,whs,scart,loc_ok=loc_ok,whs_ok=whs_ok) or None
    if add_em:
        error = 'Error with creating locations, warehouses and carts locally. %s'%add_em
    #construct the query and get the cart records
    selects = " UDL.UDL_AUTO_KEY,UDL.UDL_CODE" 
    where_clause = " WHERE UDL.UDL_COLUMN_NAME='IC_UDL_005'"
    query = 'SELECT' + selects + ' FROM USER_DEFINED_LOOKUPS UDL' + where_clause
    recs = selection_dir(query,cr) or []
    if recs:  
        del_carts = scart.objects.filter(dj_user_id = dj_user_id,quapi_id = quapi_id).delete()     
        add_em = recs and create_bulk_locs(recs,quapi_id,dj_user_id,loc,whs,scart,cart_ok=cart_ok) or None
    return error,app

def create_bulk_locs(recs,quapi_id,dj_user_id,loc=None,whs=None,scart=None,loc_ok=False,whs_ok=False,cart_ok=False):
    loc_data,whs_data,cart_data = [],[],[]
    error = ''
    for rec in recs:
        if loc_ok and rec[0]:
            loc_data.append(loc(
                    loc_auto_key = rec[0] or 0,#0
                    location_code = rec[1] or 0,#1 
                    dj_user_id = dj_user_id or 0,
                    quapi_id = quapi_id,
                    iq_enable = (rec[5] == 'T' and True) or False
                    )
                )                   
        if whs_ok and rec[2]:
            whs_data.append(whs(
                    loc_auto_key = rec[4] or 0,#WHL.LOC_AUTO_KEY
                    whs_auto_key = rec[2] or 0,#2
                    warehouse_code = rec[3],#3
                    dj_user_id = dj_user_id or 0,
                    quapi_id = quapi_id,
                    )
                )
        if cart_ok and rec[0]:
            cart_data.append(scart(
                    udl_auto_key = rec[0] or 0,#4
                    udl_code = rec[1],#5
                    dj_user_id = dj_user_id or 0,
                    quapi_id = quapi_id,
                    )
                )
    try:
        locs = loc_data and loc.objects.bulk_create(loc_data) or []
        whss = whs_data and whs.objects.bulk_create(whs_data) or []
        carts = cart_data and scart.objects.bulk_create(cart_data) or []    
    except Exception as exc:
        error = "Error with creating the locations, carts or warehouses: %s"%exc 
    return error  

@shared_task 
def get_users_nsync_beta(quapi_key,dj_user_id,is_dashboard=0,app=''):
    msg,user_data = '',[]
    error = ''
    query = "SELECT SYSUR_AUTO_KEY,USER_NAME,EMPLOYEE_CODE,USER_ID FROM SYS_USERS WHERE ARCHIVED='F'"
    from polls.models import QuantumUser,QueryApi as qa
    quapi = qa.objects.filter(id=quapi_key)
    quapi_id = quapi and quapi[0] or None  
    from polls.models import OracleConnection as oc
    orcl_conn = quapi_id and quapi_id.orcl_conn_id and oc.objects.filter(id=quapi_id.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',app
    recs = quapi_id and selection_dir(query,cr) or []
    del_users = dj_user_id and QuantumUser.objects.filter(quapi_id=quapi_key,dj_user_id=dj_user_id).delete() or None
    
    for rec in recs:
        user_data.append(QuantumUser(
                user_auto_key = rec[0],
                user_name = rec[1],
                employee_code = rec[2],
                user_id=rec[3],
                dj_user_id = dj_user_id,
                quapi_id = quapi_id,
                )
            )
    try:
        new_user = user_data and QuantumUser.objects.bulk_create(user_data) or None
    except Exception as exc:
        error = "\r\Django - Error with creating the status: %s"%exc 
    return error,app
    
@shared_task 
def get_statuses_nsync_beta(quapi_key,dj_user_id,is_dashboard=0,app='',object_type=''):
    #TODO: must get the user that is currently logged in down in selection/insert/update methods 
    #user links us to the API creds we need to send an API request to run a query.
    #display WO Status Table in dropdown displaying WO_STATUS[‘DESCRIPTION’] 
    # and WO_STATUS[‘SEVERITY’], but only show Open WO_STATUS[‘STATUS_TYPE’])
    error,stat_data = '',[]
    from polls.models import StatusSelection as statsel,QueryApi
    quapi = QueryApi.objects.filter(id=quapi_key)
    quapi_id = quapi and quapi[0] or None  
    from polls.models import OracleConnection as oc
    orcl_conn = quapi_id and quapi_id.orcl_conn_id and oc.objects.filter(id=quapi_id.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',app

    if object_type and object_type == 'SO':
        query = "SELECT SOS_AUTO_KEY, SEQUENCE, 'SO - ' || DESCRIPTION FROM SO_STATUS WHERE STATUS_TYPE IN ('Open','Delay','Defer') ORDER BY SEQUENCE ASC"
    elif object_type and object_type == 'RO':
        query = "SELECT RST_AUTO_KEY, SEQUENCE, DESCRIPTION FROM RO_STATUS WHERE STATUS_TYPE IN ('Open','Delay','Defer') ORDER BY SEQUENCE ASC"
    else:
        query = "SELECT WOS_AUTO_KEY, SEVERITY, DESCRIPTION FROM WO_STATUS WHERE STATUS_TYPE IN ('Open','Delay','Defer') ORDER BY SEVERITY ASC"
        
    recs = selection_dir(query,cr)
    if app and app in ['ro-management','smd-management'] and object_type and object_type == 'SO':
        query = "SELECT WOS_AUTO_KEY, SEVERITY, 'WO - ' || DESCRIPTION FROM WO_STATUS WHERE STATUS_TYPE IN ('Open','Delay','Defer') ORDER BY SEVERITY ASC"
        recs += selection_dir(query,cr)
    del_stats = dj_user_id and statsel.objects.filter(quapi_id=quapi_key,is_dashboard=is_dashboard,dj_user_id=dj_user_id).delete() or None
    if is_dashboard == 1:
        wos_auto_key = 0
        severity = '0'
        name = 'PENDING'
        try:
            sstat = statsel.objects.create(quapi_id=quapi_id,dj_user_id = dj_user_id,is_dashboard = 1,wos_auto_key = wos_auto_key, severity = severity, name = name)
            sstat.save()
        except Exception as exc:
            error = "\r\nDjango - Error with creating the pending status: %s"%exc 
    for status in recs:
        wos_auto_key = status[0]
        severity = status[1]
        if object_type in ['RO','SO']:
            name = str(status[2])
        else:
            name = str(status[1]) + '-' + str(status[2])
        stat_data.append(statsel( 
            quapi_id=quapi_id,
            dj_user_id = dj_user_id,
            is_dashboard=is_dashboard,
            wos_auto_key = wos_auto_key,
            severity = severity,
            name = name,
            )
        )
    try:
        new_statii = stat_data and statsel.objects.bulk_create(stat_data) or None
    except Exception as exc:
        error = "\r\Django - Error with creating the statuses: %s"%exc 
    return error,app
    
@shared_task
def make_updates(
        sysur_auto_key=0,
        quapi_id=None,
        user_id=None,
        manual_ecd='',
        rank=None,
        manager=None,new_due_date=None,customer=None,status=None,search_mgr=None,due_date=None,wo_number=None,session_id=None,woo_ids=[]): 
    from dateutil.parser import parse
    woo_recs,commit_error = [],''
    from polls.models import WOStatus as wos_obj,QueryApi
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    pdue_date,sysur_auto_key,audit_ok = None,None,False
    error,msg,values_str, field_changed,ins_error,input_error = '','','','','','' 
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None     
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.',''     
    if not woo_ids:
        woo_recs = wos_obj.objects.filter(active=1, is_dashboard=1, user_id=user_id, session_id=session_id)
        woo_ids = woo_recs.values_list('woo_auto_key',flat=True)        
    if not woo_ids:
        return 'There are no active WOs to update.',msg 
    else:
        woo_lists = construct_akl(woo_ids) 
        if status:      
            #query = "SELECT SYSUR_AUTO_KEY FROM SYS_USERS WHERE REGEXP_LIKE(USER_ID, '%s', 'i')"%manager
            #user_mgr = quapi and selection(query,quapi=quapi) or None 
            #if not user_mgr:
                #input_error += "\r\nYou have entered a manager that doesn't exist: '%s'"%manager
                #return input_error,msg
            #sysur_auto_key = user_mgr and user_mgr[0] and user_mgr[0][0] or None 
            #prefix = ''
            field_changed += ' Status changed to '%status
            values_str += 'WOS_AUTO_KEY = %s.'%status  
        if manual_ecd:
            prefix = ''
            if values_str:
                prefix = ', '
            field_changed += prefix + ' MANUAL_ECD changed to %s.'%manual_ecd
            new_man_date = "TO_DATE('%s', 'mm/dd/yyyy')"%manual_ecd
            values_str += prefix + "MANUAL_ECD = %s"%new_man_date                      
        if rank:
            try:        
                int(rank)  
            except Exception as error:
                input_error += "\r\nOnly positive integers are allowed for rank." 
                return input_error,msg
            prefix = ''
            if values_str:
                prefix = ', '
            field_changed += prefix + ' Rank changed to %s.'%rank
            values_str += prefix + "RANK = '%s'"%rank   
        if manager:      
            query = "SELECT SYSUR_AUTO_KEY FROM SYS_USERS WHERE REGEXP_LIKE(USER_ID, '%s', 'i') OR REGEXP_LIKE(USER_NAME, '%s', 'i')"%manager
            user_mgr = quapi and selection(query,quapi=quapi) or None 
            if not user_mgr:
                input_error += "\r\nYou have entered a manager that doesn't exist: '%s'"%manager
                return input_error,msg
            sysur_auto_key = user_mgr and user_mgr[0] and user_mgr[0][0] or None 
            prefix = ''
            if values_str:
                prefix = ', '
            field_changed += prefix + ' Manager changed to %s.'%manager
            values_str += prefix + 'SYSUR_MANAGER = %s'%sysur_auto_key           
        if new_due_date:         
            upd_due_date = "TO_DATE('%s', 'mm/dd/yyyy')"%new_due_date              
            dd_update = False
            prefix = ''
            if values_str:
                prefix = ', '  
            field_changed += prefix + ' Due date changed to %s.'%new_due_date                
            values_str += prefix + 'DUE_DATE = %s'%upd_due_date
    for woo_list in woo_lists:
        where_clause = " WHERE WOO_AUTO_KEY IN %s"%woo_list        
        upd_query = "UPDATE WO_OPERATION SET %s"%values_str + where_clause       
        error = updation_dir(upd_query,cr)
    if error == '{"recs": ""}' or not error:
        error = ''
        commit_error = orcl_commit(con=con) 
    else:
        return error,''       
    active_woos = wos_obj.objects.filter(session_id=session_id)
    woo_complete = active_woos and active_woos.values_list('woo_auto_key',flat=True)
    woo_complete = woo_complete and construct_akl(woo_complete) or []  
    blow_away_old = active_woos and active_woos.delete() or None    
    for woo_list in woo_complete:
        #now get the updated woos
        #error = populate_wo_grid(session_id,cr,woo_list)
        error,msg = add_wo_record(session_id=session_id,is_dashboard=1,is_racking=0,wak_clause=woo_list,quapi=quapi,refresh=True,keep_recs=True) 
        update_error = update_audit_trail(user_id,woo_adt_list=woo_list,new_rank=rank,new_due_date=new_due_date,new_mgr=sysur_auto_key) 
        msg=' Successful update.'
    if field_changed and values_str:
        from polls.models import MLApps as maps,QuantumUser as qu
        app_id = maps.objects.filter(code='wo-management') 
        user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
        user_rec = user_rec and user_rec[0] or None
        field_changed = values_str + ' | wo number(s):= ' + str(woo_lists)
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')
        
        if user_rec:
            error = register_audit_trail(user_rec,field_changed,values_str,now,app_id,quapi)  
        if not error: 
            msg = ' Successful update.'
    return str(error) + str(input_error) + str(ins_error),str(msg + ' ' + field_changed)
    
def update_audit_trail(user_id, woo_adt_list = '', new_status=None, new_location=None, stm_auto_key=None, new_mgr=None, new_due_date=None, new_rank=None,quapi=None):
    #parameters: user_id - integer ID for PK to the SYS_USERS table
    #            wo_number - user input WO#
    #            new_status - user input status to which to update
    #            new_location - user input locastion (converted to database id for loc_auto_key in STM table)
    from polls.models import QuantumUser
    msg = ''
    user_rec = QuantumUser.objects.filter(user_id=user_id)
    user_key = user_rec and user_rec[0] and user_rec[0].user_auto_key or None
    q1,q2 = '',''
    if not user_key:
        return 'No user is logged in and therefore no update can be made to the audit trail table.'
    if new_rank:
        query = """UPDATE AUDIT_TRAIL SET SYSUR_AUTO_KEY=%s WHERE SOURCE_AK IN %s AND SOURCE_TABLE = 'WOO' AND SOURCE_FIELD = 'RANK' AND NEW_VALUE = %s"""%(user_key, woo_adt_list, new_rank)
        updation(query,user_id=user_id,quapi=quapi)  
    #if new_mgr:
    #    query = """
    #        UPDATE AUDIT_TRAIL SET SYSUR_AUTO_KEY=%s WHERE SOURCE_AK IN %s AND SOURCE_TABLE = 'WOO' AND SOURCE_FIELD = 'SYSUR_MANAGER' AND NEW_VALUE = (SELECT USER_ID FROM SYS_USERS WHERE SYSUR_AUTO_KEY = %s)"""%(user_key, woo_adt_list, new_mgr)
    #    updation(query,user_id=user_id,quapi=quapi)   
    if new_due_date:
        #UPDATE AUDIT_TRAIL SET SYSUR_AUTO_KEY=747 WHERE SOURCE_AK IN (1351,1696,1353,1354,1355,1356,1357,1358,1359,1360,1361,1362,1363,1364,1365,1366,1367,1368,893,2722,2723,2724,2768,1453,1454,1557,1558,1352) AND SOURCE_TABLE = 'WOO' AND SOURCE_FIELD = 'DUE_DATE' AND NEW_VALUE = TO_DATE('12-08-2019', 'mm-dd-yyyy')
        query = """UPDATE AUDIT_TRAIL SET SYSUR_AUTO_KEY=%s WHERE ADT_AUTO_KEY = (SELECT MAX(ADT_AUTO_KEY) FROM AUDIT_TRAIL WHERE SOURCE_AK IN %s AND SOURCE_TABLE = 'WOO' AND SOURCE_FIELD = 'DUE_DATE')"""%(user_key, woo_adt_list)
        updation(query,user_id=user_id,quapi=quapi)        
    if new_status:
        query = """UPDATE AUDIT_TRAIL SET SYSUR_AUTO_KEY=%s WHERE SOURCE_AK IN %s AND SOURCE_TABLE = 'WOO' AND SOURCE_FIELD = 'STATUS' AND NEW_VALUE = (SELECT DESCRIPTION FROM WO_STATUS WHERE WOS_AUTO_KEY = %s)"""%(user_key, woo_adt_list, new_status)
        updation(query,user_id=user_id,quapi=quapi) 
    return str(msg)
    
def get_stamptime(woo_auto_key=None, new_status=None, stm_auto_key=None, location_code = None, user_id='', quapi=None):
    res = ''
    if new_status:
        #get timestamp from audit trail for output in the results table
        q_stat = """SELECT STAMPTIME FROM AUDIT_TRAIL WHERE SOURCE_TABLE = 'WOO' AND SOURCE_AK = %s AND SOURCE_FIELD = 'STATUS' AND NEW_VALUE = (SELECT DESCRIPTION FROM WO_STATUS WHERE WOS_AUTO_KEY = '%s') ORDER BY STAMPTIME DESC"""%(woo_auto_key, new_status)
        res = quapi and selection(query=q_stat,user_id=user_id,quapi=quapi) or None
    if stm_auto_key and location_code:
        q_log = """SELECT BRI_AUTO_KEY,OLD_LOC_AUTO_KEY,NEW_LOC_AUTO_KEY,OLD_LOCATION_CODE,NEW_LOCATION_CODE,TIME_STAMP FROM SA_LOG WHERE NEW_LOCATION_CODE = '%s' AND STM_AUTO_KEY = %s ORDER BY STA_AUTO_KEY DESC NULLS LAST"""%(location_code,stm_auto_key)
        res = quapi and selection(query=q_log,user_id=user_id,quapi=quapi) or None
    return res
    
@shared_task   
def get_wo_family(wo_number,quapi_id,gate_number,session_id,woo_auto_key=None):
    error,msg = '',''
    quapi,recs = None,[]
    from polls.models import QueryApi
    if not quapi:
        quapi = QueryApi.objects.filter(id=quapi_id)
        quapi = quapi and quapi[0] or None
    recs = wo_number and get_toll_woos(is_detail=True,wo_number=wo_number,quapi=quapi) or []
    if not recs and woo_auto_key:
        recs = get_toll_woos(is_detail=True,wo_number=wo_number,quapi=quapi) or []
    if recs:
        error,msg = add_wo_record(session_id=session_id,is_toll=True,is_dashboard=1,is_racking=0,woo_recs=recs,quapi=quapi,sub_wo_gate=gate_number,is_detail=True,refresh=False) 
    else:
        error = 'Your Workorder has no sub-workorders'
    total_rows = len(recs)
    return error,total_rows

def get_wo_dashboard(wos_obj,wo_number=None,manager=None,location=None,warehouse=None,due_date=None,wak_clause=None,refresh=False,user_id='',quapi=None,exact_match=False):
    order_by = ' ORDER BY VTL.START_TIME NULLS LAST'
    where_status,where_clause,msg,fields_list = '','','',''
    recs,pending_recs = [],[]
    pending = False
    if wak_clause:
        where_clause += ' AND W.WOO_AUTO_KEY IN ' + wak_clause
    else:
        if manager:
            where_mgr = exact_match and "USER_ID = '%s'"%manager or ''
            where_mgr = not exact_match and "REGEXP_LIKE (USER_ID, '%s', 'i')"%manager or where_mgr
            query = "SELECT SYSUR_AUTO_KEY FROM SYS_USERS WHERE %s"%where_mgr
            user_mgr = selection(query, user_id=user_id,quapi=quapi) 
            if not user_mgr:
                msg+="\r\nYou have entered a manager that doesn't exist: '%s'"%manager
                return None
            sysur_auto_key = user_mgr and user_mgr[0] and user_mgr[0][0] or None 
            where_clause += " AND W.SYSUR_MANAGER = %s"%sysur_auto_key   
        #if manager:
        #    where_clause += " AND W.SYSUR_AUTO_KEY IN (SELECT SYSUR_AUTO_KEY FROM SYS_USERS WHERE REGEXP_LIKE (USER_ID, '%s', 'i'))"%manager
        if due_date:
            where_clause += " AND W.DUE_DATE <= TO_DATE('%s', 'mm-dd-yyyy')"%due_date           
        if wo_number:          
            where_in = """(SELECT WOO_AUTO_KEY FROM VIEW_SPS_WO_OPERATION WHERE PARENT_WO LIKE '%s%s%s')"""%('%',wo_number,'%')
            where_wo = exact_match and "W.SI_NUMBER = '%s'"%wo_number or ''
            where_wo = not exact_match and "(W.SI_NUMBER LIKE '%s%s%s'"%('%',wo_number,'%') or where_wo
            where_clause += " AND %s"%where_wo
            if not exact_match:
                where_clause += " OR W.WOO_AUTO_KEY IN %s)"%where_in
               
        if location:
            where_loc = exact_match and "L.LOCATION_CODE = '%s'"%location or ''
            if where_loc:
                where_clause += " AND %s"%where_loc
            else: 
                where_clause += " AND UPPER(L.LOCATION_CODE) LIKE UPPER('%s%s')"%(location,'%')
                #loc_clause = "WHERE LOC_AUTO_KEY IN (SELECT LOC_AUTO_KEY FROM LOCATION WHERE UPPER(LOCATION_CODE) LIKE UPPER('%s%s'))"%(location,'%')
                #where_loc = "(SELECT STM_AUTO_KEY FROM STOCK %s)"%loc_clause
                #where_clause += " AND VW.STM_AUTO_KEY IN %s"%where_loc                         
        if warehouse:
            #where_whs = exact_match and "WH.WAREHOUSE_CODE = '%s'"%warehouse or ''
            where_whs = "WH.WAREHOUSE_CODE = '%s'"%warehouse or ''
            if where_whs:
                where_clause += " AND %s"%where_whs
            else:
                where_clause += " AND UPPER(WH.WAREHOUSE_CODE) LIKE UPPER('%s%s')"%(warehouse,'%')           
                #whs_clause = "WHERE WHS_AUTO_KEY IN (SELECT WHS_AUTO_KEY FROM WAREHOUSE WHERE REGEXP_LIKE(WAREHOUSE_CODE, '%s', 'i'))"%warehouse
                #where_whs = "(SELECT STM_AUTO_KEY FROM STOCK %s)"%whs_clause
                #where_clause += " AND VW.STM_AUTO_KEY IN %s"%where_whs                               
        """if location:
            where_loc = exact_match and "LOCATION_CODE = '%s'"%location or ''
            where_loc = not exact_match and "REGEXP_LIKE (LOCATION_CODE, '%s', 'i')"%location or where_loc
            case_loc = "(SELECT LOC_AUTO_KEY FROM LOCATION WHERE %s)"%where_loc
            where_clause += " AND (CASE WHEN S.STM_AUTO_KEY IS NOT NULL AND L.LOC_AUTO_KEY IN %s THEN 1 ELSE"%case_loc
            where_clause += " (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL AND LD.LOC_AUTO_KEY IN %s THEN 1 ELSE"%case_loc
            where_clause += " (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL AND LW.LOC_AUTO_KEY IN %s THEN 1 ELSE"%case_loc
            where_clause += " (CASE WHEN SI.STM_AUTO_KEY IS NOT NULL AND LI.LOC_AUTO_KEY IN %s THEN 1 ELSE 0 END) END) END) END) = 1"%case_loc
        if warehouse:
            where_whs = exact_match and "WAREHOUSE_CODE = '%s'"%warehouse or ''
            where_whs = not exact_match and "REGEXP_LIKE (WAREHOUSE_CODE, '%s', 'i')"%warehouse or where_whs
            case_whs = "(SELECT WHS_AUTO_KEY FROM WAREHOUSE WHERE %s)"%where_whs
            where_clause += " AND (CASE WHEN S.STM_AUTO_KEY IS NOT NULL AND WH.WHS_AUTO_KEY IN %s THEN 1 ELSE"%case_whs
            where_clause += " (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL AND WHD.WHS_AUTO_KEY IN %s THEN 1 ELSE"%case_whs
            where_clause += " (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL AND WHW.WHS_AUTO_KEY IN %s THEN 1 ELSE"%case_whs
            where_clause += " (CASE WHEN SI.STM_AUTO_KEY IS NOT NULL AND WHI.WHS_AUTO_KEY IN %s THEN 1 ELSE 0 END) END) END) END) = 1"%case_whs"""         
    fields_list = """SELECT DISTINCT W.SI_NUMBER, 
    CASE WHEN WS.WOS_AUTO_KEY IS NULL OR WS.STATUS_TYPE = 'Closed' THEN 
        (CASE WHEN (SR.STM_AUTO_KEY IS NOT NULL OR SRT.STM_AUTO_KEY IS NOT NULL
            OR STW.STM_AUTO_KEY IS NOT NULL OR SI.STM_AUTO_KEY IS NOT NULL) THEN 'PENDING' ELSE 
                'AVAILABLE' END) ELSE WS.DESCRIPTION END,
    W.DUE_DATE, 
    CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.STOCK_LINE ELSE 
       (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.STOCK_LINE ELSE 
           (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.STOCK_LINE ELSE
               SI.STOCK_LINE END) 
                   END) END,
    CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN P.PN ELSE 
        (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN PW.PN ELSE 
            (CASE WHEN SD.PN IS NOT NULL THEN PD.PN ELSE
                PI.PN END) 
                   END) END,
    CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN P.DESCRIPTION ELSE 
        (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN PW.DESCRIPTION ELSE 
            (CASE WHEN PD.DESCRIPTION IS NOT NULL THEN PD.DESCRIPTION ELSE
                PI.DESCRIPTION END) 
                END) END,
    CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.SERIAL_NUMBER ELSE 
        (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.SERIAL_NUMBER ELSE 
            (CASE WHEN SD.SERIAL_NUMBER IS NOT NULL THEN SD.SERIAL_NUMBER ELSE
                SI.SERIAL_NUMBER END) 
                END) END,
    CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN L.LOCATION_CODE ELSE 
        (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN LW.LOCATION_CODE ELSE 
            (CASE WHEN LD.LOCATION_CODE IS NOT NULL THEN LD.LOCATION_CODE ELSE
                LI.LOCATION_CODE END) 
                END) END,
    WS.SEVERITY,W.ENTRY_DATE, W.COMPANY_REF_NUMBER, W.WOO_AUTO_KEY,
    W.RANK,WS.WOS_AUTO_KEY,
    CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.STM_AUTO_KEY ELSE 
        (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.STM_AUTO_KEY ELSE 
            (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.STM_AUTO_KEY ELSE
                SI.STM_AUTO_KEY END) 
                END) END,  
    VW.PARENT_WO,C.COMPANY_NAME,
    SU.USER_NAME,WT.WORK_TYPE,W.WOS_AUTO_KEY,
    CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN WH.WAREHOUSE_CODE ELSE 
        (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN WHW.WAREHOUSE_CODE ELSE 
            (CASE WHEN WHD.WAREHOUSE_CODE IS NOT NULL THEN WHD.WAREHOUSE_CODE ELSE
                WHI.WAREHOUSE_CODE END) 
                END) END,
    CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.LOC_VALIDATED ELSE 
        (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.LOC_VALIDATED ELSE 
            (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.LOC_VALIDATED ELSE
                SI.LOC_VALIDATED END) 
                END) END,
    CASE WHEN VTL.WOT_AUTO_KEY IS NOT NULL AND VTL.START_TIME IS NOT NULL
        AND VTL.STOP_TIME IS NULL THEN 'True' END,
    CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN ULS.UDL_CODE ELSE 
        (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN ULSW.UDL_CODE ELSE 
            (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN ULSD.UDL_CODE ELSE
                ULSI.UDL_CODE END)
                END) END
    """
    where_status = """ WHERE 
    (WS.WOS_AUTO_KEY IS NULL OR WS.STATUS_TYPE IN ('Open') OR WS.DESCRIPTION != 'SCRAPPED')"""
    where_stock = """ AND
    (CASE WHEN S.STM_AUTO_KEY IS NOT NULL AND S.QTY_RESERVED > 0 THEN 1 ELSE 
        (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL AND SW.QTY_RESERVED > 0 THEN 1 ELSE 
            (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL AND SD.QTY_RESERVED > 0 THEN 1 ELSE
                (CASE WHEN SI.STM_AUTO_KEY IS NOT NULL AND SI.QTY_RESERVED > 0 THEN 1 ELSE
                0 END) END) 
                END) END)=1     
    """
    query_joins = """ FROM WO_OPERATION W
         LEFT JOIN WO_WORK_TYPE WT ON WT.WWT_AUTO_KEY=W.WWT_AUTO_KEY
         LEFT JOIN SYS_USERS SU ON SU.SYSUR_AUTO_KEY=W.SYSUR_MANAGER 
         LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY
         LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY=W.WOO_AUTO_KEY
         LEFT JOIN STOCK S ON S.STM_AUTO_KEY=SR.STM_AUTO_KEY
         LEFT JOIN USER_DEFINED_LOOKUPS ULS ON ULS.UDL_AUTO_KEY = S.IC_UDL_005
         LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY=S.PNM_AUTO_KEY 
         LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY=S.WHS_AUTO_KEY 
         LEFT JOIN RO_DETAIL ROD ON ROD.WOO_AUTO_KEY = W.WOO_AUTO_KEY         
         LEFT JOIN STOCK_RESERVATIONS SRT ON SRT.ROD_AUTO_KEY=ROD.ROD_AUTO_KEY
         LEFT JOIN STOCK SD ON SD.STM_AUTO_KEY = SRT.STM_AUTO_KEY
         LEFT JOIN USER_DEFINED_LOOKUPS ULSD ON ULSD.UDL_AUTO_KEY = SD.IC_UDL_005
         LEFT JOIN LOCATION LD ON LD.LOC_AUTO_KEY = SD.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER PD ON PD.PNM_AUTO_KEY=SD.PNM_AUTO_KEY
         LEFT JOIN WAREHOUSE WHD ON WHD.WHS_AUTO_KEY=SD.WHS_AUTO_KEY
         LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = W.WOB_AUTO_KEY
         LEFT JOIN STOCK_RESERVATIONS STW ON STW.WOB_AUTO_KEY = WOB.WOB_AUTO_KEY
         LEFT JOIN STOCK SW ON SW.STM_AUTO_KEY=STW.STM_AUTO_KEY
         LEFT JOIN USER_DEFINED_LOOKUPS ULSW ON ULSW.UDL_AUTO_KEY = SW.IC_UDL_005
         LEFT JOIN LOCATION LW ON LW.LOC_AUTO_KEY = SW.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER PW ON PW.PNM_AUTO_KEY=SW.PNM_AUTO_KEY
         LEFT JOIN WAREHOUSE WHW ON WHW.WHS_AUTO_KEY=SW.WHS_AUTO_KEY
         LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY=W.CMP_AUTO_KEY 
         LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY 
         LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = VW.WOB_AUTO_KEY
         LEFT JOIN STOCK_TI STI ON STI.WOB_AUTO_KEY = W.WOB_AUTO_KEY
         LEFT JOIN STOCK SI ON SI.STM_AUTO_KEY = STI.STM_AUTO_KEY
         LEFT JOIN USER_DEFINED_LOOKUPS ULSI ON ULSI.UDL_AUTO_KEY = SI.IC_UDL_005
         LEFT JOIN LOCATION LI ON LI.LOC_AUTO_KEY = SI.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER PI ON PI.PNM_AUTO_KEY=SI.PNM_AUTO_KEY
         LEFT JOIN WAREHOUSE WHI ON WHI.WHS_AUTO_KEY=SI.WHS_AUTO_KEY
         LEFT JOIN WO_TASK WT ON WT.WOO_AUTO_KEY = W.WOO_AUTO_KEY
         LEFT JOIN VIEW_WO_TASK_LABOR VTL ON WT.WOT_AUTO_KEY = VTL.WOT_AUTO_KEY
         """ 
    
    query = fields_list + query_joins + where_status + where_stock + where_clause
    recs = selection(query, user_id=user_id, quapi=quapi) 
    """if not pending:
        where_status = "WHERE W.WOS_AUTO_KEY IS NULL"
        fields_list = "SELECT DISTINCT W.SI_NUMBER,W.WOS_AUTO_KEY,W.DUE_DATE,S.STOCK_LINE, P.PN, P.DESCRIPTION,S.SERIAL_NUMBER,L.LOCATION_CODE,W.WOS_AUTO_KEY,W.ENTRY_DATE,W.COMPANY_REF_NUMBER,W.WOO_AUTO_KEY,W.RANK,W.WOS_AUTO_KEY,S.STM_AUTO_KEY,VW.PARENT_WO,C.COMPANY_NAME,SU.USER_NAME,WT.WORK_TYPE,W.WOS_AUTO_KEY" 
        query = fields_list + query_joins + where_status + where_clause + order_by   
        pending_recs = selection(query, user_id=user_id, quapi=quapi) """    
    return recs
  
def get_wo_mgmt(wos_obj,is_dash=False,wo_number=None,part_number='',customer=None,status=None,manager=None,location=None,warehouse=None,due_date=None,wak_clause=None,refresh=False,user_id='',quapi=None,exact_match=False):
    if is_dash:
        return get_wo_dashboard(wos_obj,wo_number=wo_number,manager=manager,location=location,warehouse=warehouse,due_date=due_date,wak_clause=wak_clause,refresh=refresh,user_id=user_id,quapi=quapi,exact_match=exact_match)
    order_by = ' ORDER BY W.DUE_DATE ASC,W.RANK ASC'
    where_status,where_clause,msg,fields_list = '','','',''
    recs,pending_recs = [],[]
    pending = False
    if wak_clause:
        where_clause += ' AND W.WOO_AUTO_KEY IN ' + wak_clause
    else:
        if part_number:
            where_part = exact_match and "PN = '%s'"%part_number or ''
            if not where_part:
                where_part = "REGEXP_LIKE (PN, '%s', 'i')"%part_number
                part_clause = "WHERE PNM_AUTO_KEY IN (SELECT PNM_AUTO_KEY FROM PARTS_MASTER WHERE %s)"%where_part
                where_part = "(SELECT STM_AUTO_KEY FROM STOCK %s)"%part_clause
                where_clause += " AND VW.STM_AUTO_KEY IN %s"%where_part
                """case_part = "(SELECT PNM_AUTO_KEY FROM PARTS_MASTER WHERE %s)"%where_part
                where_clause += " AND (CASE WHEN S.STM_AUTO_KEY IS NOT NULL AND S.PNM_AUTO_KEY IN %s THEN 1 ELSE"%case_part
                where_clause += " (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL AND SD.PNM_AUTO_KEY IN %s THEN 1 ELSE"%case_part
                where_clause += " (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL AND SW.PNM_AUTO_KEY IN %s THEN 1 ELSE"%case_part
                where_clause += " (CASE WHEN SI.STM_AUTO_KEY IS NOT NULL AND SI.PNM_AUTO_KEY IN %s THEN 1 ELSE 0 END) END) END) END) = 1"%case_part"""
        if manager:
            where_mgr = exact_match and "USER_ID = '%s'"%manager or ''
            where_mgr = not exact_match and "REGEXP_LIKE (USER_ID, '%s', 'i') OR REGEXP_LIKE (USER_NAME, '%s', 'i')"%manager or where_mgr
            query = "SELECT SYSUR_AUTO_KEY FROM SYS_USERS WHERE %s"%where_mgr
            user_mgr = selection(query, user_id=user_id,quapi=quapi) 
            if not user_mgr:
                msg+="\r\nYou have entered a manager that doesn't exist: '%s'"%manager
                return None
            sysur_auto_key = user_mgr and user_mgr[0] and user_mgr[0][0] or None 
            where_clause += " AND W.SYSUR_MANAGER = %s"%sysur_auto_key   
        #if manager:
        #    where_clause += " AND W.SYSUR_AUTO_KEY IN (SELECT SYSUR_AUTO_KEY FROM SYS_USERS WHERE REGEXP_LIKE (USER_ID, '%s', 'i'))"%manager
        if due_date:
            where_clause += " AND W.DUE_DATE <= TO_DATE('%s', 'mm-dd-yyyy')"%due_date           
        if wo_number:       
            where_wo = exact_match and "W.SI_NUMBER = '%s'"%wo_number or ''          
            if not exact_match:
                where_wo = "W.SI_NUMBER LIKE '%s%s%s'"%('%',wo_number,'%') or where_wo
                where_in = """(SELECT WOO_AUTO_KEY FROM VIEW_SPS_WO_OPERATION WHERE PARENT_WO LIKE '%s%s%s')"""%('%',wo_number,'%')
                where_clause += " AND (%s"%where_wo
                where_clause += " OR W.WOO_AUTO_KEY IN %s)"%where_in  
            else:
                where_clause += " AND %s"%where_wo            
        if customer:
            where_cust = exact_match and " COMPANY_NAME = '%s'"%customer or ''
            where_cust = not exact_match and " REGEXP_LIKE (COMPANY_NAME, '%s', 'i')"%customer or where_cust
            where_clause += " AND W.CMP_AUTO_KEY IN (SELECT CMP_AUTO_KEY FROM COMPANIES WHERE%s)"%where_cust 
        if location:
            where_loc = exact_match and "L.LOCATION_CODE = '%s'"%location or ''
            if where_loc:
                where_clause += " AND %s"%where_loc
            else:
                where_clause += " AND UPPER(L.LOCATION_CODE) LIKE UPPER('%s%s')"%(location,'%')             
                #loc_clause = "WHERE LOC_AUTO_KEY IN (SELECT LOC_AUTO_KEY FROM LOCATION WHERE REGEXP_LIKE(LOCATION_CODE, '%s', 'i'))"%location
                #where_loc = "(SELECT STM_AUTO_KEY FROM STOCK %s)"%loc_clause
                #where_clause += " AND VW.STM_AUTO_KEY IN %s"%where_loc                         
                #where_clause += " REGEXP_LIKE (LD.LOCATION_CODE, '%s', 'i') OR"%location
                #where_clause += " REGEXP_LIKE (LW.LOCATION_CODE, '%s', 'i') OR"%location
                #where_clause += " REGEXP_LIKE (LI.LOCATION_CODE, '%s', 'i'))"%location
        if warehouse:
            where_whs = exact_match and "WH.WAREHOUSE_CODE = '%s'"%warehouse or ''
            if where_whs:
                where_clause += " AND %s"%where_whs
            else:
                where_clause += " AND UPPER(WH.WAREHOUSE_CODE) LIKE UPPER('%s%s')"%(warehouse,'%') 
                #whs_clause = "WHERE WHS_AUTO_KEY IN (SELECT WHS_AUTO_KEY FROM WAREHOUSE WHERE REGEXP_LIKE(WAREHOUSE_CODE, '%s', 'i'))"%warehouse
                #where_whs = "(SELECT STM_AUTO_KEY FROM STOCK %s)"%whs_clause
                #where_clause += " AND VW.STM_AUTO_KEY IN %s"%where_whs             
                """where_whs = "REGEXP_LIKE (WAREHOUSE_CODE, '%s', 'i')"%warehouse
                case_whs = "(SELECT WHS_AUTO_KEY FROM WAREHOUSE WHERE %s)"%where_whs
                where_clause += " AND (CASE WHEN S.STM_AUTO_KEY IS NOT NULL AND WH.WHS_AUTO_KEY IN %s THEN 1 ELSE"%case_whs
                where_clause += " (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL AND WHD.WHS_AUTO_KEY IN %s THEN 1 ELSE"%case_whs
                where_clause += " (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL AND WHW.WHS_AUTO_KEY IN %s THEN 1 ELSE"%case_whs
                where_clause += " (CASE WHEN SI.STM_AUTO_KEY IS NOT NULL AND WHI.WHS_AUTO_KEY IN %s THEN 1 ELSE 0 END) END) END) END) = 1"%case_whs"""       
        if status and status != '0':
            where_clause += " AND W.WOS_AUTO_KEY = %s"%int(status)
        elif status and status == '0':
            pending = True
    where_stock = """ AND
    (CASE WHEN S.STM_AUTO_KEY IS NOT NULL AND S.QTY_RESERVED > 0 THEN 1 ELSE 
        (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL AND SW.QTY_RESERVED > 0 THEN 1 ELSE 
            (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL AND SD.QTY_RESERVED > 0 THEN 1 ELSE
                (CASE WHEN SI.STM_AUTO_KEY IS NOT NULL AND SI.QTY_RESERVED > 0 THEN 1 ELSE
                0 END) END) 
                END) END)=1"""
    if pending:        
        where_status = """ WHERE W.WOS_AUTO_KEY IS NULL"""
        fields_list = """SELECT DISTINCT 
        W.SI_NUMBER,
        CASE WHEN WS.WOS_AUTO_KEY IS NULL OR WS.STATUS_TYPE = 'Closed' THEN 
            (CASE WHEN (SR.STM_AUTO_KEY IS NOT NULL OR SRT.STM_AUTO_KEY IS NOT NULL
                OR STW.STM_AUTO_KEY IS NOT NULL OR SI.STM_AUTO_KEY IS NOT NULL) THEN 'PENDING' ELSE 
                    'AVAILABLE' END) ELSE WS.DESCRIPTION END,
        W.DUE_DATE,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.STOCK_LINE ELSE 
           (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.STOCK_LINE ELSE 
               (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.STOCK_LINE ELSE
                   SI.STOCK_LINE END) 
                       END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN P.PN ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN PW.PN ELSE 
                (CASE WHEN SD.PN IS NOT NULL THEN PD.PN ELSE
                    PI.PN END) 
                       END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN P.DESCRIPTION ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN PW.DESCRIPTION ELSE 
                (CASE WHEN PD.DESCRIPTION IS NOT NULL THEN PD.DESCRIPTION ELSE
                    PI.DESCRIPTION END) 
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.SERIAL_NUMBER ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.SERIAL_NUMBER ELSE 
                (CASE WHEN SD.SERIAL_NUMBER IS NOT NULL THEN SD.SERIAL_NUMBER ELSE
                    SI.SERIAL_NUMBER END) 
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN L.LOCATION_CODE ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN LW.LOCATION_CODE ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN LD.LOCATION_CODE ELSE
                    LI.LOCATION_CODE END) 
                    END) END,
        W.WOS_AUTO_KEY,W.ENTRY_DATE,W.COMPANY_REF_NUMBER,
        W.WOO_AUTO_KEY,W.RANK,W.WOS_AUTO_KEY,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.STM_AUTO_KEY ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.STM_AUTO_KEY ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.STM_AUTO_KEY ELSE
                    SI.STM_AUTO_KEY END) 
                    END) END, 
        VW.PARENT_WO,C.COMPANY_NAME,SU.USER_NAME,
        WT.WORK_TYPE,W.WOS_AUTO_KEY,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN WH.WAREHOUSE_CODE ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN WHW.WAREHOUSE_CODE ELSE 
                (CASE WHEN WHD.WAREHOUSE_CODE IS NOT NULL THEN WHD.WAREHOUSE_CODE ELSE
                    WHI.WAREHOUSE_CODE END) 
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.LOC_VALIDATED ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.LOC_VALIDATED ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.LOC_VALIDATED ELSE
                    SI.LOC_VALIDATED END) 
                    END) END,
        CASE WHEN VTL.WOT_AUTO_KEY IS NOT NULL AND VTL.START_TIME IS NOT NULL
            AND VTL.STOP_TIME IS NULL THEN 'True' END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN ULS.UDL_CODE ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN ULSW.UDL_CODE ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN ULSD.UDL_CODE ELSE
                    ULSI.UDL_CODE END)
                    END) END,
        W.MANUAL_ECD,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.QTY_OH ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.QTY_OH ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.QTY_OH ELSE
                    SI.QTY_OH END) 
                    END) END,
        '','',''                    
        """
    else:
        fields_list = """SELECT DISTINCT W.SI_NUMBER, 
        CASE WHEN WS.WOS_AUTO_KEY IS NULL OR WS.STATUS_TYPE = 'Closed' THEN 
            (CASE WHEN (SR.STM_AUTO_KEY IS NOT NULL OR SRT.STM_AUTO_KEY IS NOT NULL
                OR STW.STM_AUTO_KEY IS NOT NULL OR SI.STM_AUTO_KEY IS NOT NULL) THEN 'PENDING' ELSE 
                    'AVAILABLE' END) ELSE WS.DESCRIPTION END,
        W.DUE_DATE, 
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.STOCK_LINE ELSE 
           (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.STOCK_LINE ELSE 
               (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.STOCK_LINE ELSE
                   SI.STOCK_LINE END) 
                       END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN P.PN ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN PW.PN ELSE 
                (CASE WHEN SD.PN IS NOT NULL THEN PD.PN ELSE
                    PI.PN END) 
                       END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN P.DESCRIPTION ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN PW.DESCRIPTION ELSE 
                (CASE WHEN PD.DESCRIPTION IS NOT NULL THEN PD.DESCRIPTION ELSE
                    PI.DESCRIPTION END) 
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.SERIAL_NUMBER ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.SERIAL_NUMBER ELSE 
                (CASE WHEN SD.SERIAL_NUMBER IS NOT NULL THEN SD.SERIAL_NUMBER ELSE
                    SI.SERIAL_NUMBER END) 
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN L.LOCATION_CODE ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN LW.LOCATION_CODE ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN LD.LOCATION_CODE ELSE
                    LI.LOCATION_CODE END) 
                    END) END,
        WS.SEVERITY,W.ENTRY_DATE, W.COMPANY_REF_NUMBER, W.WOO_AUTO_KEY,
        W.RANK,WS.WOS_AUTO_KEY,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.STM_AUTO_KEY ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.STM_AUTO_KEY ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.STM_AUTO_KEY ELSE
                    SI.STM_AUTO_KEY END) 
                    END) END,  
        VW.PARENT_WO,C.COMPANY_NAME,
        SU.USER_NAME,WT.WORK_TYPE,W.WOS_AUTO_KEY,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN WH.WAREHOUSE_CODE ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN WHW.WAREHOUSE_CODE ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN WHD.WAREHOUSE_CODE ELSE
                    WHI.WAREHOUSE_CODE END) 
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.LOC_VALIDATED ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.LOC_VALIDATED ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.LOC_VALIDATED ELSE
                    SI.LOC_VALIDATED END) 
                    END) END,
        CASE WHEN VTL.WOT_AUTO_KEY IS NOT NULL AND VTL.START_TIME IS NOT NULL
            AND VTL.STOP_TIME IS NULL THEN 'True' END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN ULS.UDL_CODE ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN ULSW.UDL_CODE ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN ULSD.UDL_CODE ELSE
                    ULSI.UDL_CODE END)
                    END) END,
        W.MANUAL_ECD,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.QTY_OH ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.QTY_OH ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.QTY_OH ELSE
                    SI.QTY_OH END) 
                    END) END,
        '','',''                    
        """
        where_status = """ WHERE 
        (WS.WOS_AUTO_KEY IS NULL OR (WS.STATUS_TYPE IN ('Open','Delay','Defer')))"""
        where_vtl = """ AND """
    query_joins = """ 
    FROM WO_OPERATION W
         LEFT JOIN WO_WORK_TYPE WT ON WT.WWT_AUTO_KEY=W.WWT_AUTO_KEY
         LEFT JOIN SYS_USERS SU ON SU.SYSUR_AUTO_KEY=W.SYSUR_MANAGER 
         LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY
         LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY=W.WOO_AUTO_KEY
         LEFT JOIN STOCK S ON S.STM_AUTO_KEY=SR.STM_AUTO_KEY 
         LEFT JOIN USER_DEFINED_LOOKUPS ULS ON ULS.UDL_AUTO_KEY = S.IC_UDL_005
         LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY=S.PNM_AUTO_KEY 
         LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY=S.WHS_AUTO_KEY 
         LEFT JOIN RO_DETAIL ROD ON ROD.WOO_AUTO_KEY = W.WOO_AUTO_KEY         
         LEFT JOIN STOCK_RESERVATIONS SRT ON SRT.ROD_AUTO_KEY=ROD.ROD_AUTO_KEY
         LEFT JOIN STOCK SD ON SD.STM_AUTO_KEY = SRT.STM_AUTO_KEY
         LEFT JOIN USER_DEFINED_LOOKUPS ULSD ON ULSD.UDL_AUTO_KEY = SD.IC_UDL_005
         LEFT JOIN LOCATION LD ON LD.LOC_AUTO_KEY = SD.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER PD ON PD.PNM_AUTO_KEY=SD.PNM_AUTO_KEY
         LEFT JOIN WAREHOUSE WHD ON WHD.WHS_AUTO_KEY=SD.WHS_AUTO_KEY
         LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = W.WOB_AUTO_KEY
         LEFT JOIN STOCK_RESERVATIONS STW ON STW.WOB_AUTO_KEY = WOB.WOB_AUTO_KEY
         LEFT JOIN STOCK SW ON SW.STM_AUTO_KEY=STW.STM_AUTO_KEY
         LEFT JOIN USER_DEFINED_LOOKUPS ULSW ON ULSW.UDL_AUTO_KEY = SW.IC_UDL_005
         LEFT JOIN LOCATION LW ON LW.LOC_AUTO_KEY = SW.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER PW ON PW.PNM_AUTO_KEY=SW.PNM_AUTO_KEY
         LEFT JOIN WAREHOUSE WHW ON WHW.WHS_AUTO_KEY=SW.WHS_AUTO_KEY
         LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY=W.CMP_AUTO_KEY 
         LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY
         LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = VW.WOB_AUTO_KEY
         LEFT JOIN STOCK_TI STI ON STI.WOB_AUTO_KEY = W.WOB_AUTO_KEY
         LEFT JOIN STOCK SI ON SI.STM_AUTO_KEY = STI.STM_AUTO_KEY
         LEFT JOIN USER_DEFINED_LOOKUPS ULSI ON ULSI.UDL_AUTO_KEY = SI.IC_UDL_005
         LEFT JOIN LOCATION LI ON LI.LOC_AUTO_KEY = SI.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER PI ON PI.PNM_AUTO_KEY=SI.PNM_AUTO_KEY
         LEFT JOIN WAREHOUSE WHI ON WHI.WHS_AUTO_KEY=SI.WHS_AUTO_KEY
         LEFT JOIN WO_TASK WT ON WT.WOO_AUTO_KEY = W.WOO_AUTO_KEY
         LEFT JOIN VIEW_WO_TASK_LABOR VTL ON WT.WOT_AUTO_KEY = VTL.WOT_AUTO_KEY"""
    query = fields_list + query_joins + where_status + where_stock + where_clause + order_by        
    recs = selection(query, user_id=user_id, quapi=quapi) 
    stm_keys = []
    final_recs = []
    for rec in recs:
        if rec[14] not in stm_keys:
            stm_keys.append(rec[14])
            final_recs.append(rec)        
    return final_recs
    
def get_toll_subs(parent_id=None,quapi=None,parent_ids=[]):     
    #fields_list = "SELECT STR.STR_AUTO_KEY,W.WOB_AUTO_KEY,S.WHS_AUTO_KEY,S.LOC_AUTO_KEY,WH.SEQUENCE FROM WO_OPERATION W "
    where_nulls = """ AND (S.HISTORICAL_FLAG = 'F' OR S.STM_AUTO_KEY IS NULL)
        AND ((SD.HISTORICAL_FLAG = 'F') OR SD.STM_AUTO_KEY IS NULL)
        AND ((SW.HISTORICAL_FLAG = 'F') OR SW.STM_AUTO_KEY IS NULL)
        AND ((SI.HISTORICAL_FLAG = 'F') OR SI.STM_AUTO_KEY IS NULL)
        AND (ROD.QTY_RESERVED > 0 OR ROD.ROD_AUTO_KEY IS NULL)"""
    where_clause = (parent_id and " WHERE WB.WOO_AUTO_KEY = %s"%parent_id) or (parent_ids and " WHERE WB.WOO_AUTO_KEY IN %s"%parent_ids) or ''
    child_query = """SELECT DISTINCT W.SI_NUMBER,
        CASE WHEN WS.DESCRIPTION = 'CLOSED' AND WOB.WOB_AUTO_KEY IS NOT NULL THEN 'Reserved' ELSE WS.DESCRIPTION END,
        W.DUE_DATE, 
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.STOCK_LINE ELSE 
           (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.STOCK_LINE ELSE 
               (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.STOCK_LINE ELSE
                   SI.STOCK_LINE END) 
                       END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN P.PN ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN PW.PN ELSE 
                (CASE WHEN SD.PN IS NOT NULL THEN PD.PN ELSE
                    PI.PN END) 
                       END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN P.DESCRIPTION ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN PW.DESCRIPTION ELSE 
                (CASE WHEN PD.DESCRIPTION IS NOT NULL THEN PD.DESCRIPTION ELSE
                    PI.DESCRIPTION END) 
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.SERIAL_NUMBER ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.SERIAL_NUMBER ELSE 
                (CASE WHEN SD.SERIAL_NUMBER IS NOT NULL THEN SD.SERIAL_NUMBER ELSE
                    SI.SERIAL_NUMBER END) 
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN L.LOCATION_CODE ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN LW.LOCATION_CODE ELSE 
                (CASE WHEN LD.LOCATION_CODE IS NOT NULL THEN LD.LOCATION_CODE ELSE
                    LI.LOCATION_CODE END) 
                    END) END,
        CASE WHEN (WS.DESCRIPTION <> 'CLOSED' OR WOB.WOB_AUTO_KEY IS NOT NULL) THEN WS.SEVERITY END,W.ENTRY_DATE,W.COMPANY_REF_NUMBER,W.WOO_AUTO_KEY,
        W.RANK,WS.WOS_AUTO_KEY,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.STM_AUTO_KEY ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.STM_AUTO_KEY ELSE 
                (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.STM_AUTO_KEY ELSE
                    SI.STM_AUTO_KEY END) 
                    END) END,                           
        VW.PARENT_WO,C.COMPANY_NAME,W.WOO_AUTO_KEY,WB.WOO_AUTO_KEY,W.WOS_AUTO_KEY,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN WH.SEQUENCE ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN WHW.SEQUENCE ELSE
                (CASE WHEN WHD.SEQUENCE IS NOT NULL THEN WHD.SEQUENCE ELSE
                    4 END)
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.QTY_OH ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.QTY_OH ELSE 
                (CASE WHEN SD.QTY_OH IS NOT NULL THEN SD.QTY_OH ELSE 
                    SI.QTY_OH END)
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN WH.WAREHOUSE_CODE ELSE 
            (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN WHW.WAREHOUSE_CODE ELSE 
                (CASE WHEN WHD.WAREHOUSE_CODE IS NOT NULL THEN WHD.WAREHOUSE_CODE ELSE
                    WHI.WAREHOUSE_CODE END) 
                    END) END,
        CASE WHEN S.STM_AUTO_KEY IS NOT NULL THEN S.LOC_VALIDATED ELSE 
            (CASE WHEN SD.STM_AUTO_KEY IS NOT NULL THEN SD.LOC_VALIDATED ELSE 
                (CASE WHEN SW.STM_AUTO_KEY IS NOT NULL THEN SW.LOC_VALIDATED ELSE
                    SI.LOC_VALIDATED END) 
                    END) END,
        '','','','','','',''
        FROM WO_OPERATION W
        LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY=W.WOO_AUTO_KEY
         LEFT JOIN STOCK S ON S.STM_AUTO_KEY=SR.STM_AUTO_KEY 
         LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY=S.PNM_AUTO_KEY 
         LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY=S.WHS_AUTO_KEY 
         LEFT JOIN RO_DETAIL ROD ON ROD.WOO_AUTO_KEY = W.WOO_AUTO_KEY         
         LEFT JOIN STOCK_RESERVATIONS SRT ON SRT.ROD_AUTO_KEY=ROD.ROD_AUTO_KEY
         LEFT JOIN STOCK SD ON SD.STM_AUTO_KEY = SRT.STM_AUTO_KEY
         LEFT JOIN LOCATION LD ON LD.LOC_AUTO_KEY = SD.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER PD ON PD.PNM_AUTO_KEY=SD.PNM_AUTO_KEY
         LEFT JOIN WAREHOUSE WHD ON WHD.WHS_AUTO_KEY=SD.WHS_AUTO_KEY
         LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = W.WOB_AUTO_KEY
         LEFT JOIN STOCK_RESERVATIONS STW ON STW.WOB_AUTO_KEY = WOB.WOB_AUTO_KEY
         LEFT JOIN STOCK SW ON SW.STM_AUTO_KEY=STW.STM_AUTO_KEY
         LEFT JOIN LOCATION LW ON LW.LOC_AUTO_KEY = SW.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER PW ON PW.PNM_AUTO_KEY=SW.PNM_AUTO_KEY
         LEFT JOIN WAREHOUSE WHW ON WHW.WHS_AUTO_KEY=SW.WHS_AUTO_KEY
         LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY=W.CMP_AUTO_KEY 
         LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY 
         LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = VW.WOB_AUTO_KEY
         LEFT JOIN STOCK_TI STI ON STI.WOB_AUTO_KEY = W.WOB_AUTO_KEY
         LEFT JOIN STOCK SI ON SI.STM_AUTO_KEY = STI.STM_AUTO_KEY
         LEFT JOIN LOCATION LI ON LI.LOC_AUTO_KEY = SI.LOC_AUTO_KEY
         LEFT JOIN PARTS_MASTER PI ON PI.PNM_AUTO_KEY=SI.PNM_AUTO_KEY
         LEFT JOIN WAREHOUSE WHI ON WHI.WHS_AUTO_KEY=SI.WHS_AUTO_KEY
        %s
    """%(where_clause)
    child_recs = selection(child_query, quapi=quapi)
    return child_recs

def get_toll_woos(wo_number=None,is_detail=False,gate_number='',customer=None,due_date=None,parent_ids=None,refresh=False,user_id='',quapi=None,session_id=None,exact_match=False):
    order_by = ' ORDER BY W.DUE_DATE ASC,W.RANK ASC'
    group_by = ' GROUP BY W.WOO_AUTO_KEY'
    where_status,where_clause,msg,fields_list,where_wo,error = '','','','','',''
    recs,pending_recs,final_recs,parent_recs,children,child_recs = [],[],[],[],[],[]
    pending=False
    query_joins = """
        FROM WO_OPERATION W 
        LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY 
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY=W.WOO_AUTO_KEY 
        LEFT JOIN STOCK S ON S.STM_AUTO_KEY=SR.STM_AUTO_KEY 
        LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY 
        LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY=S.PNM_AUTO_KEY 
        LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY=S.WHS_AUTO_KEY 
        LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY=W.CMP_AUTO_KEY 
        LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY"""
    where_nulls = """ 
        AND ((S.HISTORICAL_FLAG = 'F' AND S.QTY_OH > 0) OR S.STM_AUTO_KEY IS NULL)
        AND (SR.QTY_RESERVED > 0 OR SR.STR_AUTO_KEY IS NULL)"""
    if not parent_ids:   
        if due_date:
            where_clause += " WHERE W.DUE_DATE <= TO_DATE('%s', 'mm-dd-yyyy')"%due_date           
        if wo_number:          
            where_wo = " W.SI_NUMBER = '%s'"%wo_number
            if where_wo:
                where_clause += (where_clause and " AND %s"%where_wo) or " WHERE %s"%where_wo
        if is_detail and gate_number and gate_number!='0':
            where_clause += " AND WH.SEQUENCE=%s"%gate_number 
        if customer:
            where_cust = exact_match and " COMPANY_NAME = '%s'"%customer or ''
            where_cust = not exact_match and " REGEXP_LIKE (COMPANY_NAME, '%s', 'i')"%customer or where_cust
            where_clause += where_clause and " AND" or " WHERE"
            where_clause += " W.CMP_AUTO_KEY IN (SELECT CMP_AUTO_KEY FROM COMPANIES WHERE%s)"%where_cust           
        fields_list = """SELECT DISTINCT W.SI_NUMBER, WS.DESCRIPTION, W.DUE_DATE, 
            S.STOCK_LINE, P.PN, P.DESCRIPTION,S.SERIAL_NUMBER,L.LOCATION_CODE,
            WS.SEVERITY,W.ENTRY_DATE,W.COMPANY_REF_NUMBER,W.WOO_AUTO_KEY,
            W.RANK,WS.WOS_AUTO_KEY,S.STM_AUTO_KEY,VW.PARENT_WO,C.COMPANY_NAME,
            W.WOO_AUTO_KEY,W.WOO_AUTO_KEY,W.WOS_AUTO_KEY,WH.SEQUENCE,S.QTY_OH,
            WH.WAREHOUSE_CODE,S.LOC_VALIDATED,'','','','','','',''"""        
        where_nulls_parent = where_nulls + ' AND VW.PARENT_WO IS NULL'
        parent_query = fields_list + query_joins + where_clause + " AND W.SI_NUMBER LIKE '%s%s%s'"%('%',wo_number,'%') + where_nulls_parent   
        #should get parent and children (2 generations) and excludes 100057 from another family but related by one of his progeny through adoption of the WO to 100072 family.
        parent_recs = selection(parent_query, quapi=quapi)
        parent_ids = construct_id_list(parent_recs,11)        
    for parent in parent_recs:
        #error = create_wo_tasks(quapi,parent[17],session_id,woo_id = parent[17])
        children = get_toll_subs(parent_id=parent[17],quapi=quapi)
        child_recs = children
        count = 0
        gate1,gate2,gate3,gate4='','','',''
        gate1_qty,gate2_qty,gate3_qty,gate4_qty=0,0,0,0
        #recursively grabbing all sub-woos in heirarchy.
        while children and count < 77:
            parent_ids = construct_id_list(children,11)  
            #error = create_wo_tasks(quapi,parent[17],session_id,woo_id_list=parent_ids)            
            children = get_toll_subs(parent_ids=parent_ids,quapi=quapi)
            child_recs += children
            count += 1            
        for crec in child_recs: 
            crec+=[parent[17],gate1,gate2,gate3,gate4,gate1_qty,gate2_qty,gate3_qty,gate4_qty]
        #from itertools import groupby 
        #grp_by_parent = [list(g) for _, g in groupby(child_recs, lambda l: l[23])]          
        gate1,gate2,gate3,gate4,gate1_qty,gate2_qty,gate3_qty,gate4_qty = produce_gate_fractions(child_recs)                      
        #Since we are looping through parents, we know  this is the parent of the child we are working on, 
        #thus we add all of the gatge info obtained in the previous line.
        parent += [0,gate1,gate2,gate3,gate4,gate1_qty,gate2_qty,gate3_qty,gate4_qty]                  
    final_recs += parent_recs + child_recs   
    return final_recs
    
def get_list_ids(recs,id_pos=11): 
    parent_lists = []
    parent_ids = recs and '(' or ''
    count = 0
    for rec in recs:      
        parent_ids += str(rec[id_pos]) + ','
        if (count+1)%995 == 0 and recs[count+1]: 
            parent_ids = parent_ids[:-1]
            parent_ids += ')'              
            parent_lists.append(parent_ids)
            parent_ids = '(' + str(parent_ids[count+1]) + ','
        count += 1
    if parent_ids and parent_ids[-1] == ',':
        parent_ids = list(parent_ids)
        parent_ids[-1] = ')'
        parent_ids = "".join(parent_ids)
        parent_lists.append(parent_ids)
    return parent_lists
    
def construct_id_list(recs,id_pos=11): 
    parent_lists = []
    parent_ids = recs and '(' or ''
    count = 0
    for rec in recs:
        parent_ids += str(rec[id_pos]) + ','
        if (count+1)%495 == 0 and recs[count+1]: 
            parent_ids += ')'        
            parent_lists.append(parent_ids)
            parent_ids = '(' + str(parent_ids[count+1]) + ','
    if parent_ids and parent_ids[-1] == ',':
        parent_ids = list(parent_ids)#
        parent_ids[-1] = ')'
        parent_ids = "".join(parent_ids)
        count += 1
    return parent_ids
    
def produce_gate_fractions(recs):
    #takes list of lists of stock reservations and sums the QTY_OH of the stock line for each stock_reservation
    #must not add the stock quantities for the any recs that are parents (PARENT_WO (recs[15] is NULL)    
    gate1,gate2,gate3,gate4,total_count = 0,0,0,0,0
    #parent_pos = []
    #i=0
    #loops through child records and sums stock quantities for placement on ultimate parent (parent-less parent)
    for rec in recs:  
        #if not rec[15]:           
        #    parent_pos.append(i)
        #i+=1
        #if rec[15]:
        quantity = rec[21] and str(rec[21]).isnumeric() and int(rec[21]) or 0
        if rec[20] == 1:
            gate1 += quantity
        elif rec[20] == 2:
            gate2 += quantity
        elif rec[20] == 3:
            gate3 += quantity
        elif rec[20] == 4:
            gate4 += quantity
        total_count += quantity
    gate1_qty,gate2_qty,gate3_qty,gate4_qty = gate1,gate2,gate3,gate4
    gate1= gate1 and total_count and str(gate1) + ' / ' + str(total_count) or None
    gate2= gate2 and total_count and str(gate2) + ' / ' + str(total_count) or None  
    gate3= gate3 and total_count and str(gate3) + ' / ' + str(total_count) or None 
    gate4= gate4 and total_count and str(gate4) + ' / ' + str(total_count) or None
    return gate1,gate2,gate3,gate4,gate1_qty,gate2_qty,gate3_qty,gate4_qty

def construct_text(values, id_pos = 11):
    
    woo_id_list = values and '(' + '"' + str(values[0]) + '"' or ''
    woo_lists = []
    if len(values) == 1:
        woo_id_list += ')'
        woo_lists = [woo_id_list]
        return woo_lists
    count = 1
    if values and len(values) > 1:
        for wak in values[1:]:
            woo_id_list += ',' + '"' + str(wak) + '"'
            #if we get to the 496th WOO, then we close out the string and will begin with another element to start the next list of 496.
            if (count+1)%495 == 0:                     
                woo_id_list += ')'           
                woo_lists.append(woo_id_list)
                if len(values) > 495:
                    woo_id_list = '(' + '"' + str(values[count+1]) + '"'
            count += 1 
        woo_id_list += ')'
        woo_lists.append(woo_id_list)
    return woo_lists
 
def construct_akl(woo_ids, id_pos = 11):
    
    woo_id_list = woo_ids and '(' + str(woo_ids[0]) or ''
    woo_lists = []
    if len(woo_ids) == 1:
        woo_id_list += ')'
        woo_lists = [woo_id_list]
        return woo_lists
    count = 1
    if woo_ids and len(woo_ids) > 1:
        for wak in woo_ids[1:]:
            woo_id_list += ',' + str(wak)
            #if we get to the 496th WOO, then we close out the string and will begin with another element to start the next list of 496.
            if (count+1)%495 == 0:                     
                woo_id_list += ')'           
                woo_lists.append(woo_id_list)
                if len(woo_ids) > 495:
                    woo_id_list = '(' + str(woo_ids[count+1])
            count += 1 
        woo_id_list += ')'
        woo_lists.append(woo_id_list) 
    return woo_lists
    
def get_child_boms(wob_auto_key,woo_auto_key,quapi):
    query = "SELECT * FROM WO_BOM WHERE WOO_AUTO_KEY ='%s' AND ACTIVITY <> 'ASSY' AND WOB_AUTO_KEY <> '%s'"%(woo_auto_key,wob_auto_key)
    return selection(query,quapi=quapi) 
    
def unique_items(the_list,position=0):
    found = set()
    for item in sort_list:
        if item[position] not in found:
            yield item
            found.add(item[position])  
@shared_task
def bom_schedule(woo_ids,quapi_id,user_id,session_id,run_count):
    parent_woos,error,msg,woo_lists,update_vals,field_changed,due_date,default_offset = [],'','',[],[],'',None,0
    from polls.models import QueryApi,WOStatus
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None  
    active_woos = WOStatus.objects.filter(session_id=session_id)
    woo_complete = active_woos and active_woos.values_list('woo_auto_key',flat=True)
    woo_complete = woo_complete and construct_akl(woo_complete) or [] 
    uda = ['RANK','BOMNEEDDATE']
    uda_codes = '(\'' + uda[0] + '\',\'' + uda[1] + '\')' 
    woo_aks = []
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    con = False
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.','' 
    if not woo_ids and active_woos:
        woo_ids = active_woos
        woo_ids = woo_ids.values_list('woo_auto_key',flat=True)         
    if woo_ids:
        if len(woo_ids) > 1:
            woo_lists = construct_akl(woo_ids)
        elif len(woo_ids) == 1:
            woo_lists = ['(' + str(woo_ids[0]) + ')']
    updates,bom_updates,records = {},{},[]    
    for list in woo_lists: 
        query = """SELECT DISTINCT W.WOO_AUTO_KEY,W.RANK,
                 CASE WHEN PWO.DUE_DATE IS NOT NULL THEN PWO.DUE_DATE ELSE W.DUE_DATE END,
                 W.SI_NUMBER,
                 WB.PNM_AUTO_KEY,WB.WOB_AUTO_KEY,VW.PARENT_WO,WB.WOO_AUTO_KEY,WB.NEED_DATE,VWB.WOO_AUTO_KEY
                 FROM WO_OPERATION W                 
                 LEFT JOIN WO_BOM WB ON W.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
                 LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY
                 LEFT JOIN VIEW_SPS_WO_OPERATION VWB ON VWB.WOB_AUTO_KEY = WB.WOB_AUTO_KEY 
                 LEFT JOIN WO_OPERATION PWO ON PWO.SI_NUMBER = VW.PARENT_WO                  
                 WHERE (W.WOO_AUTO_KEY IN 
                 (SELECT VW.WOO_AUTO_KEY FROM VIEW_SPS_WO_OPERATION VW WHERE 
                 VW.PARENT_WO IN (SELECT SI_NUMBER FROM WO_OPERATION WHERE WOO_AUTO_KEY IN %s))
                 OR W.WOO_AUTO_KEY IN %s) OR W.SI_NUMBER LIKE 
                 (SELECT SI_NUMBER FROM WO_OPERATION WHERE WOO_AUTO_KEY IN %s) || '%s'"""%(list,list,list,'%')
        woo_recs = selection_dir(query,cr)
        parent_due_date = woo_recs and woo_recs[0] and woo_recs[0][2] and woo_recs[0][2][:10] or None
        woo_updates = []
        wob_updates = []        
        for rec in woo_recs:
            wob_auto_key = rec[5]
            woo_auto_key = rec[9]
            if woo_auto_key not in woo_updates and not wob_auto_key in wob_updates:
                woo_updates.append(woo_auto_key)
                wob_updates.append(wob_auto_key)
                due_date = rec[2] and rec[2][:10] or parent_due_date            
                sub_q = """(SELECT DISTINCT S.PNM_AUTO_KEY FROM WO_OPERATION W 
                    LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY = W.WOO_AUTO_KEY
                    LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
                    WHERE W.WOO_AUTO_KEY = %s)"""%(rec[0])    
                q_default = """SELECT UDC.ATTRIBUTE_VALUE FROM USER_DEFINED_ATTRIBUTES UDA 
                    INNER JOIN UDA_CHECKED UDC ON UDC.UDA_AUTO_KEY = UDA.UDA_AUTO_KEY 
                    AND UDA.UDA_CODE = 'DEFAULTBOMOFFSET' AND UDC.AUTO_KEY IN %s"""%sub_q       
                default_offset = selection_dir(q_default,cr) 
                default_offset = default_offset and default_offset[0] or None
                default_offset = default_offset and default_offset[0] or 0 
                #update the bom with bom_id = rec[5] with default need date value
                #W.WOO_AUTO_KEY,W.RANK,W.DUE_DATE,W.SI_NUMBER,
                #WB.PNM_AUTO_KEY,WB.WOB_AUTO_KEY,VW.PARENT_WO,
                #WB.WOO_AUTO_KEY,WB.NEED_DATE,VWB.WOO_AUTO_KEY
                #if default_offset:            
                if due_date and wob_auto_key:          
                    where = "WHERE WOB_AUTO_KEY=%s"%wob_auto_key
                    set_vals = "NEED_DATE = TO_DATE('%s', 'yyyy-mm-dd') - %s"%(due_date,default_offset)         
                    query = "UPDATE WO_BOM SET %s %s"%(set_vals,where)
                    error = updation_dir(query,cr)
                if woo_auto_key and due_date:
                    where_clause = "WHERE WOO_AUTO_KEY = %s"%woo_auto_key
                    q1 = "UPDATE WO_OPERATION SET DUE_DATE = TO_DATE('%s', 'yyyy-mm-dd') - %s %s"%(due_date,default_offset,where_clause)
                    error = updation_dir(q1,cr)
        query = """
        SELECT UBC.ATTRIBUTE_VALUE,W.WOO_AUTO_KEY,W.RANK,W.DUE_DATE,W.SI_NUMBER,UBC.AUTO_KEY,
        WB.PNM_AUTO_KEY,UBA.UDA_CODE,WB.WOB_AUTO_KEY,VW.PARENT_WO,WB.WOO_AUTO_KEY,WB.NEED_DATE,VWB.WOO_AUTO_KEY
             FROM WO_OPERATION W       
             LEFT JOIN WO_BOM WB ON W.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
             LEFT JOIN UDA_CHECKED UBC ON UBC.AUTO_KEY = WB.PNM_AUTO_KEY
             LEFT JOIN USER_DEFINED_ATTRIBUTES UBA ON UBA.UDA_AUTO_KEY = UBC.UDA_AUTO_KEY
             LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY
             LEFT JOIN VIEW_SPS_WO_OPERATION VWB ON VWB.WOB_AUTO_KEY = WB.WOB_AUTO_KEY           
             WHERE UBA.UDA_CODE = 'BOMNEEDDATE' AND (W.WOO_AUTO_KEY IN 
             (SELECT VW.WOO_AUTO_KEY FROM VIEW_SPS_WO_OPERATION VW WHERE 
             VW.PARENT_WO IN (SELECT SI_NUMBER FROM WO_OPERATION WHERE WOO_AUTO_KEY IN %s))
             OR W.WOO_AUTO_KEY IN %s)"""%(list,list)
        records += selection_dir(query,cr)
        query = """
        SELECT UBC.ATTRIBUTE_VALUE,W.WOO_AUTO_KEY,W.RANK,W.DUE_DATE,W.SI_NUMBER,UBC.AUTO_KEY,
        WB.PNM_AUTO_KEY,UBA.UDA_CODE,WB.WOB_AUTO_KEY,VW.PARENT_WO,WB.WOO_AUTO_KEY,WB.NEED_DATE,VWB.WOO_AUTO_KEY
             FROM WO_OPERATION W       
             LEFT JOIN WO_BOM WB ON W.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
             LEFT JOIN UDA_CHECKED UBC ON UBC.AUTO_KEY = WB.PNM_AUTO_KEY
             LEFT JOIN USER_DEFINED_ATTRIBUTES UBA ON UBA.UDA_AUTO_KEY = UBC.UDA_AUTO_KEY
             LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY
             LEFT JOIN VIEW_SPS_WO_OPERATION VWB ON VWB.WOB_AUTO_KEY = WB.WOB_AUTO_KEY
             WHERE UBA.UDA_CODE = 'RANK'
             AND (W.WOO_AUTO_KEY IN 
             (SELECT VW.WOO_AUTO_KEY FROM VIEW_SPS_WO_OPERATION VW WHERE 
             VW.PARENT_WO IN (SELECT SI_NUMBER FROM WO_OPERATION WHERE WOO_AUTO_KEY IN %s))
             OR W.WOO_AUTO_KEY IN %s)"""%(list,list)
        records += selection_dir(query,cr)
        #query = """
        """SELECT UBC.ATTRIBUTE_VALUE,W.WOO_AUTO_KEY,W.RANK,W.DUE_DATE,W.SI_NUMBER,UBC.AUTO_KEY,
        WB.PNM_AUTO_KEY,'DEFAULTBOMOFFSET',WB.WOB_AUTO_KEY,VW.PARENT_WO,WB.WOO_AUTO_KEY,WB.NEED_DATE,VWB.WOO_AUTO_KEY
             FROM WO_OPERATION W       
             LEFT JOIN WO_BOM WB ON W.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
             LEFT JOIN UDA_CHECKED UBC ON UBC.AUTO_KEY = WB.PNM_AUTO_KEY
             LEFT JOIN USER_DEFINED_ATTRIBUTES UBA ON UBA.UDA_AUTO_KEY = UBC.UDA_AUTO_KEY
             LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY
             LEFT JOIN VIEW_SPS_WO_OPERATION VWB ON VWB.WOB_AUTO_KEY = WB.WOB_AUTO_KEY
             WHERE UBA.UDA_AUTO_KEY IS NULL
             AND (W.WOO_AUTO_KEY IN 
             (SELECT VW.WOO_AUTO_KEY FROM VIEW_SPS_WO_OPERATION VW WHERE 
             VW.PARENT_WO IN (SELECT SI_NUMBER FROM WO_OPERATION WHERE WOO_AUTO_KEY IN %s))"""
             #OR W.WOO_AUTO_KEY IN %s)"""%(list,list)
        #records += selection_dir(query,cr)
    for vals in records:   
        woo_auto_key = vals[12] or vals[1]
        due_date = vals[3] and vals[3][:10] or None 
        #if the bom has a sub-wo, then we update the date on that WO too:       
        if vals[7] == 'RANK' and vals[0]:          
            where = "WHERE WOO_AUTO_KEY=%s"""%(woo_auto_key)
            set_vals = "RANK = '%s'"%vals[0]
            query = "UPDATE WO_OPERATION SET %s %s"%(set_vals,where)
            #error = updation(query,quapi=quapi)
            error = updation_dir(query,cr)                         
        elif vals[7] == 'BOMNEEDDATE' and vals[8]:
            #we need the WOB_AUTO b_KEY for the bom we're about to update
            #we need the WOO tied to the WOB (if there is one)
            #down below, we are updating due_date on WO's and we must get the due_date from the parent_wo rather than using the BoM's           
            offset = vals[0]
            if due_date and vals[12] and offset:
                where_clause = "WHERE WOO_AUTO_KEY = %s"%vals[12]
                q1 = "UPDATE WO_OPERATION SET DUE_DATE = TO_DATE('%s', 'yyyy-mm-dd') - %s %s"%(due_date,offset,where_clause)
                error = updation_dir(q1,cr)               
            if due_date and vals[8] and vals[0]:
                where = "WHERE WOB_AUTO_KEY=%s"""%vals[8]
                set_vals = "NEED_DATE = TO_DATE('%s', 'yyyy-mm-dd') - %s"%(due_date,vals[0])         
                query = "UPDATE WO_BOM SET %s %s"%(set_vals,where)
                #error = updation(query,quapi=quapi) 
                error = updation_dir(query,cr) 
            elif not due_date and vals[0] and vals[4] and vals[9] and vals[8]:
                #No due date on WO so we take the due date of the parent if it exists.
                query = """SELECT W.DUE_DATE,W.WOO_AUTO_KEY,WO.DUE_DATE,WO.WOO_AUTO_KEY FROM WO_OPERATION W
                            LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY
                            LEFT JOIN WO_OPERATION WO ON WO.SI_NUMBER = VW.PARENT_WO
                            WHERE W.SI_NUMBER = '%s'"""%vals[9]
                res = selection_dir(query,cr) 
                res = res and res[0] or None
                due_date = res and res[0] or None
                due_date = due_date and due_date[:10] or None
                woo_auto_key = res and res[1] or None
                if not due_date and res and vals[8]:                   
                    due_date = res and res[2] or None
                    due_date = due_date and due_date[:10] or None
                    if not due_date:
                        error = 'WO, %s, has no due date and neither does the next highest WO (parent).'%vals[4] 
                if due_date:               
                    where = "WHERE WOB_AUTO_KEY=%s"""%vals[8]
                    set_vals = "NEED_DATE = TO_DATE('%s', 'yyyy-mm-dd') - %s"%(due_date,vals[0])         
                    query = "UPDATE WO_BOM SET %s %s"%(set_vals,where)
                    error = updation_dir(query,cr) 

    if error and error != '{"recs": ""}':
        return error,msg
    error = ''
    msg = ' Successful BoM Schedule update.'  
    blow_away_old = active_woos and active_woos.delete() or None
    for woo_list in woo_complete:
        #now get the updated woos into the db                
        error,msg = add_wo_record(session_id=session_id,is_dashboard=1,is_racking=0,user_id=user_id,wak_clause=woo_list,quapi=quapi,refresh=True,keep_recs=True) 
    orcl_commit(con=con)
    if run_count < 5:
        run_count += 1
        error,msg = bom_schedule(woo_ids,quapi_id,user_id,session_id,run_count)       
    return error,msg
   
def populate_wo_grid(session_id,cr,woo_list):
    error = '' 
    query="""SELECT DISTINCT VW.PARENT_WO,W.SI_NUMBER,P.PN,P.DESCRIPTION,WS.DESCRIPTION,
        W.DUE_DATE,W.RANK,SU.USER_NAME,UDL.UDL_CODE,L.LOCATION_CODE,S.LOC_VALIDATED,
        S.SERIAL_NUMBER,C.COMPANY_NAME,
        CASE WHEN VTL.WOT_AUTO_KEY IS NOT NULL AND VTL.START_TIME IS NOT NULL
            AND VTL.STOP_TIME IS NULL THEN 'True' END
        FROM WO_OPERATION W
        LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.PARENT_WO = W.SI_NUMBER
        LEFT JOIN STOCK_RESERVATIONS SR ON SR.WOO_AUTO_KEY = W.WOO_AUTO_KEY
        LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
        LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
        LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY
        LEFT JOIN SYS_USERS SU ON SU.SYSUR_AUTO_KEY = W.SYSUR_MANAGER
        LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_005
        LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
        LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = W.CMP_AUTO_KEY 
        LEFT JOIN WO_TASK WT ON WT.WOO_AUTO_KEY = W.WOO_AUTO_KEY
        LEFT JOIN VIEW_WO_TASK_LABOR VTL ON WT.WOT_AUTO_KEY = VTL.WOT_AUTO_KEY 
        WHERE W.WOO_AUTO_KEY IN %s       
    """%woo_list
    recs = selection_dir(query,cr)
    if recs:
        msg = create_wo_mgmt_bulk(session_id,recs)
        if not msg:
            error = 'There was a problem creating stock lines'
    return error

def create_wo_mgmt_bulk(session_id,recs):
    objects = []
    from polls.models import WOStatus
    for parent_wo,wo_number,part_number,description,status,\
        due_date,rank,manager,rack,location_code,loc_validated_date,\
        serial_number,customer,notes,wh_code,stm_auto_key,stock_line,\
        timestamp in recs:
        objects.append(WOStatus( 
            parent_wo=parent_wo,
            wo_number=wo_number,
            part_number=part_number,
            description=description,
            status=status,
            due_date=due_date and due_date[:10] or None,
            rank=rank,
            int_rank = rank or 0,
            manager=manager,
            rack=rack,
            location_code=location_code,
            loc_validated_date=loc_validated_date or None,
            serial_number=serial_number,
            customer=customer,
            notes=notes,
            wh_code=wh_code,
            stm_auto_key=stm_auto_key,
            stock_line=stock_line,
            session_id = session_id,            
        ))
    msg = WOStatus.objects.bulk_create(objects)    
    return msg
    
def create_tasks_bulk(task_recs):
    objects = []
    #wot_auto_key,wot_sequence,
    #wot_description,wot_status,status_type,wot_labor_hours,
    #wot_labor_last,wot_technician,woo_auto_key,
    #task_master_desc,ult_parent_woo,session_id
    """
    """    
    from polls.models import WOTask as task_obj
    for task in task_recs:
        objects.append(task_obj(  
            wot_auto_key = task[0],#WT.WOT_AUTO_KEY
            wot_sequence = task[1] or 0,#WT.SEQUENCE
            wot_description = task[2] or task[9],# WT.SQUAWK_DESC
            wot_status = str(task[3]),#status
            wot_labor_hours = isinstance(task[5], float) and task[5] or 0,
            wot_labor_last = task[6] or '',
            wot_technician = task[7],
            woo_auto_key = task[8] or 0,
            task_master_desc = task[9] and task[9][:20] or '',
            skills_est_hours = task[11] or 0,
            ult_parent_woo = task[12] or 0,
            session_id = task[13],
            status_type = str(task[4]), #WOS.STATUS_TYPE         
        ))
    msg = task_obj.objects.bulk_create(objects)
    return msg

@shared_task   
def create_wo_tasks(quapi_id,ult_parent_woo,session_id,woo_id_list=[],woo_id=None,wo_number=''):
    #this method takes the parent woos and sub-woos in a tuple-looking string to pass
    #into method that runs sql query
    #creates wo_task data in bulk
    #returns error message or empty string if no error
    error = ''
    recs = []
    from polls.models import QueryApi,WOTask
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    
    if woo_id_list or woo_id or wo_number:
        recs = get_wo_tasks(quapi,wak_list = woo_id_list,woo_id=woo_id,wo_number=wo_number)
    else:
        #get all WTM Descriptions and send back
        recs = get_all_tasks(quapi)
        
    existing_tasks = WOTask.objects.filter(session_id=session_id)
    del_existings = existing_tasks.delete()
    
    if recs:
        for rec in recs:
            """if len(rec) > 4 and rec[5]:
                rec[5] = float(rec[5])
            else:
                rec[5] = 0.00"""
            rec += [ult_parent_woo,session_id,wo_number]
        res = create_tasks_bulk(recs)
        if not res:
            error = 'There was a problem with creating the tasks locally. Try again please.'
    return error
    
def get_all_tasks(quapi):
    query="""SELECT WTM.WTM_AUTO_KEY, WTM.AUTO_SEQ_NUM, WTM.LONG_DESCR,
        WTM.DESCRIPTION, 
        WTM.DESCRIPTION,
        0, 
        '', 
        '',WTM.WTM_AUTO_KEY,
        WTM.DESCRIPTION,WTM.AUTO_SEQ_NUM,
        0        
        FROM WO_TASK_MASTER WTM
        WHERE WTM.WO = 'T'
        AND WTM.MO = 'T'
        AND WTM.WP = 'T'
        ORDER BY WTM.AUTO_SEQ_NUM, WTM.DESCRIPTION"""
    recs = selection(query,quapi=quapi)
    return recs

def get_wo_tasks(quapi,wak_list=[],woo_id=None,wo_number=''):

    query = """
        SELECT WT.WOT_AUTO_KEY, WT.SEQUENCE, VTS.DESCRIPTION,
        WOS.SEVERITY || ' - ' || CASE WHEN WO.WOS_AUTO_KEY IS NOT NULL 
        THEN WOS.DESCRIPTION
        ELSE 'PENDING' END, 
        WOS.STATUS_TYPE,
        VTS.LABOR_HOURS, 
        CASE WHEN VTL.START_TIME <
        VTL.STOP_TIME THEN VTL.STOP_TIME ELSE VTL.START_TIME END, 
        SU.FIRST_NAME || ' ' || SU.LAST_NAME,WO.WOO_AUTO_KEY,
        WTM.DESCRIPTION,WTM.AUTO_SEQ_NUM,
        WT.SKILLS_EST_HOURS
        FROM WO_TASK_MASTER WTM
        LEFT JOIN WO_TASK WT ON WTM.WTM_AUTO_KEY = WT.WTM_AUTO_KEY
        LEFT JOIN WO_OPERATION WO ON WT.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
        LEFT JOIN VIEW_WO_STATUS_TASK VTS ON WT.WOT_AUTO_KEY = VTS.WOT_AUTO_KEY
        LEFT JOIN WO_STATUS WOS ON WOS.WOS_AUTO_KEY = WT.WOS_AUTO_KEY
        LEFT JOIN VIEW_WO_TASK_LABOR VTL ON WT.WOT_AUTO_KEY = VTL.WOT_AUTO_KEY
        LEFT JOIN SYS_USERS SU ON VTL.SYSUR_AUTO_KEY = SU.SYSUR_AUTO_KEY
        """

    query="""SELECT WT.WOT_AUTO_KEY, WT.SEQUENCE, VTS.DESCRIPTION,
    WOS.SEVERITY || ' - ' || CASE WHEN WO.WOS_AUTO_KEY IS NOT NULL 
    THEN WOS.DESCRIPTION
    ELSE 'PENDING' END, 
    WOS.STATUS_TYPE,
    VTS.LABOR_HOURS, 
    CASE WHEN VTL.START_TIME <
       VTL.STOP_TIME THEN VTL.STOP_TIME ELSE VTL.START_TIME END, 
    SU.FIRST_NAME || ' ' || SU.LAST_NAME,WO.WOO_AUTO_KEY,
    WTM.DESCRIPTION,WTM.AUTO_SEQ_NUM,WT.SKILLS_EST_HOURS
    FROM WO_TASK WT
    LEFT JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WT.WTM_AUTO_KEY
       LEFT JOIN WO_OPERATION WO ON WT.WOO_AUTO_KEY = WO.WOO_AUTO_KEY
       LEFT JOIN VIEW_WO_STATUS_TASK VTS ON WT.WOT_AUTO_KEY = VTS.WOT_AUTO_KEY
       LEFT JOIN WO_STATUS WOS ON WOS.WOS_AUTO_KEY = WT.WOS_AUTO_KEY
       LEFT JOIN VIEW_WO_TASK_LABOR VTL ON WT.WOT_AUTO_KEY = VTL.WOT_AUTO_KEY
       LEFT JOIN SYS_USERS SU ON VTL.SYSUR_AUTO_KEY = SU.SYSUR_AUTO_KEY"""
       
    if woo_id:
        query+="""
        WHERE WO.WOO_AUTO_KEY = %s
        """%woo_id
    elif wo_number:        
        query+="""
        WHERE UPPER(WO.SI_NUMBER) = UPPER('%s')
        """%wo_number
    elif wak_list:
        query+="""
        WHERE WO.WOO_AUTO_KEY IN %s
        """%wak_list   
    query += """ AND (WOS.STATUS_TYPE 
        IS NULL OR WOS.STATUS_TYPE 
        NOT IN ('Closed', 'Cancel'))"""
    #query += """ AND WTM.WO = 'T'
    #    AND WTM.MO = 'T'
    #    AND WTM.WP = 'T'
    #   """
    
    query += """ ORDER BY WT.SEQUENCE DESC NULLS LAST
        """
    recs = selection(query,quapi=quapi)
    return recs
     
def bom_get_subs(parent_ids=None,parent_id=None,cr=None):
    query = ""
    if parent_id:
        query = """SELECT VW.WOO_AUTO_KEY,W.DUE_DATE,WB.WOB_AUTO_KEY
                    FROM WO_OPERATION W 
                    LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY 
                    LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = VW.WOB_AUTO_KEY 
                    WHERE WB.WOO_AUTO_KEY = %s"""%parent_id
    if parent_ids:
        query = """SELECT VW.WOO_AUTO_KEY,W.DUE_DATE,WB.WOB_AUTO_KEY
                    FROM WO_OPERATION W 
                    LEFT JOIN VIEW_SPS_WO_OPERATION VW ON VW.WOO_AUTO_KEY = W.WOO_AUTO_KEY 
                    LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = VW.WOB_AUTO_KEY 
                    WHERE WB.WOO_AUTO_KEY IN %s"""%parent_ids                   
    subs = query and selection_dir(query,cr) or []
    return subs
    
def bom_sub_woos(parent_ids=None,parent_id=None,cr=None):            
    count = 0
    #recursively grabbing all sub-woos in heirarchy.
    children = bom_get_subs(parent_id=parent_id,cr=cr)
    sub_woos = children
    while children and count < 77:
        parent_ids = construct_id_list(children,0)          
        children = bom_get_subs(parent_ids=parent_ids,cr=cr)
        sub_woos += children
        count += 1                
    return sub_woos

@shared_task
def add_wo_record(is_mgmt=False,part_number='',is_dashboard=1,\
    is_toll=0,is_racking=0,keep_recs=False,user_id=None,quapi_id=0,\
    quapi=None,customer=None,status=None,location=None,warehouse=None,\
    wo_number=None,due_date=None,manager=None,new_manager=None,\
    new_rank=None,new_due_date=None,refresh=False,session_id=None,\
    wak_clause=None,woo_recs=[],exact_match=False,sub_wo_gate=None,\
    is_detail=0,is_dash=False,lbh_auto_key=0):
    msg,error,parent_wo,ult_parent,vendor,has_labor,poss_dupes = '','','',0,'',False,{}
    wos_status,woo_data,wh_code,condition_code,stock_owner,loc_validated_date = '',[],'','','',None
    gate1,gate2,gate3,gate4,qty_at_gate=0,0,0,0,0
    gate1_qty,gate2_qty,gate3_qty,gate4_qty=0,0,0,0
    from polls.models import WOStatus as wos_obj,QueryApi
    if not quapi:
        quapi = QueryApi.objects.filter(id=quapi_id)
        quapi = quapi and quapi[0] or None    
    right_now = datetime.now()   
    #if user didn't select a status and entered a WO that doesn't already
    #exist or didn't enter one at all, we create the WO locally.       
    recs,active_woos,woo_records = [],[],[]
    due_date_var,timestamp,time_loc,sysur_mgr,wo_type = '','','','',''
    if not status or status == 'False' and status != '0': status = ''
    #periodic or triggered refresh of active_woos
    if refresh and wak_clause:
        recs = get_wo_mgmt(wos_obj,wak_clause=wak_clause,refresh=True,user_id=user_id,quapi=quapi,exact_match=exact_match,is_dash=is_dash) 
    elif woo_recs and not refresh:
        recs = woo_recs
        active_woos = (is_toll and is_detail and wos_obj.objects.filter(is_detail=1,active=1, is_dashboard=1, session_id=session_id, is_racking=0)) or []       
    elif not (woo_recs or refresh):
        if is_toll:
            recs = get_toll_woos(wo_number=wo_number,gate_number=sub_wo_gate,due_date=due_date,customer=customer,quapi=quapi,session_id=session_id,exact_match=exact_match)            
        else:         
            recs = get_wo_mgmt(wos_obj,wo_number=wo_number,part_number=part_number,status=status,customer=customer,due_date=due_date,manager=manager,location=location,warehouse=warehouse,user_id=user_id,quapi=quapi,exact_match=exact_match,is_dash=is_dash)
        active_woos = wos_obj.objects.filter(is_detail=is_detail, active=1, is_dashboard=1, session_id=session_id, is_racking=0)        
    if not active_woos:
        if (part_number or customer or status or due_date or wo_number or manager) and keep_recs:
            recs = get_wo_mgmt(wos_obj,wo_number=wo_number,part_number=part_number,status=status,customer=customer,due_date=due_date,manager=manager,user_id=user_id,quapi=quapi,exact_match=exact_match,is_dash=is_dash)
        active_woos = wos_obj.objects.filter(active=1, session_id=session_id, is_dashboard=0, is_racking=1)        
    woos_deleted = not keep_recs and active_woos and active_woos.delete() or None       
    if not recs and not refresh:
        error += 'No records found.'
        return error,msg
    if recs and recs[0] and len(recs[0]) != 0 and recs[0][0] and isinstance(recs[0][0], list):
        recs = recs[0]
    quantity = 0
    
    for woo in recs:
        format = '%Y-%m-%d %H:%M:%S' 
        if is_dash or is_dashboard or is_toll or is_mgmt:
            if woo[0] in poss_dupes:               
                if poss_dupes[woo[0]] == 'True':
                    continue
                else:
                    for wo_rec in woo_records:
                        if wo_rec == woo or (is_dash and wo_rec[0] == woo[0]):
                            woo_records.remove(wo_rec)
                            break
            customer = woo[16]
            vendor = len(woo) >=29 and str(woo[28]) or ''
            loc_validated_date = woo[21] and datetime.strptime(woo[21],format) or None
            quantity = len(woo)>=41 and str(woo[40]).isnumeric() and woo[40] or 0
            if is_dash or is_dashboard:
                if len(woo) > 22:
                    has_labor = woo[22] or None 
                #if not is_mgmt:                    
                    poss_dupes[woo[0]] = has_labor
        else:
            customer = woo[23]  
                   
        due_date = woo[2] and woo[2][:10] or None
        if is_racking:
            status = str(woo[1])
            condition_code = str(woo[26])
            stock_owner = str(woo[27])
            customer = woo[23]
            quantity = len(woo) > 29 and woo[29] or 0
        else:
            status = str(woo[8]) + '-' + str(woo[1])          
        if is_toll:
            
            loc_validated_date = woo[23] and datetime.strptime(woo[23],format) or None
            is_detail = woo[15] and True or False
            qty_at_gate = woo[21]            
            wh_code = woo[22]
            ult_parent = woo[24]
            gate1 = woo[25]
            gate2 = woo[26]
            gate3 = woo[27]
            gate4 = woo[28] 
            gate1_qty = woo[29]
            gate2_qty = woo[30]
            gate3_qty = woo[31]
            gate4_qty = woo[32]
            if sub_wo_gate and is_detail:
                if sub_wo_gate=='1':
                    qty_at_gate = gate1_qty
                elif sub_wo_gate=='2':
                    qty_at_gate = gate2_qty
                elif sub_wo_gate=='3':
                    qty_at_gate = gate3_qty
                elif sub_wo_gate=='4':
                    qty_at_gate = gate4_qty
                else:
                    qty_at_gate = 0 
            """elif not sub_wo_gate and sub_wo_gate != '0' and is_detail:                
                qty_at_gate = woo[21]  
                if gate1_qty:
                    sub_wo_gate = '1'             
                elif gate2_qty:
                    sub_wo_gate = '2'
                elif gate3_qty:
                    sub_wo_gate = '3'                    
                elif gate4_qty:
                    sub_wo_gate = '4'    
                else:
                    sub_wo_gate = None """                 
        #Get time of last update to this status
        si_number = woo[0]
        woo_auto_key = woo[11]
        stm_auto_key = woo[14]
        if woo_auto_key and is_dashboard and not is_toll:
            wh_code = woo[20]
            format = '%Y-%m-%dT%H:%M:%S'           
            timestamp = woo[19] and not int(woo[19]) or None
            timestamp = timestamp and datetime.strptime(timestamp,format) or None
            if timestamp:
                delta = timestamp and (right_now - timestamp)
                days, hours, minutes = delta.days, delta.seconds // 3600, delta.seconds // 60 % 60
                timestamp =  str(days) + 'D:' + str(hours) + 'H'
            #get time of last update to this location           
            #time_loc = get_stamptime(woo_auto_key = woo_auto_key,stm_auto_key = stm_auto_key, location_code = woo[7],user_id=user_id,quapi=quapi)
            """time_loc = woo[20] or None
            time_loc = time_loc and time_loc[0] or None
            if time_loc:
                time_loc = datetime.strptime(time_loc,format)
                if type(time_loc) is datetime:
                    delta = time_loc and (right_now - time_loc)
                    days, hours, minutes = delta.days, delta.seconds // 3600, delta.seconds // 60 % 60
                    #'Time elapsed since last status update:
                    time_loc =  str(days) + 'D:' + str(hours) + 'H'"""
        
        if not si_number and not woo_auto_key and stm_auto_key and is_racking:
            #str_auto_key = woo[20]
            #then get the woo from the bom
            wh_code = woo[20]
            si_number = woo[24]
            status = (si_number and 'Reserved') or 'Available'
        if not si_number and woo[25] and is_racking:
            status = 'Reserved'
            si_number = woo[25]
        
        woo_data = [
            si_number,#SI Number [0]
            status or ' - PENDING',#status with sev and desc[1],[8]
            timestamp or None,
            due_date or None,
            due_date_var,
            woo[3],#stock_line [3]
            woo[4],#part_number [4]
            woo[5],#part_desc [5]
            woo[6],#serial_no [6]
            woo[7],#9loc_code  [7] wak
            time_loc or None,#10 
            customer,#11 #C.COMPANY_NAME []
            woo[9][:10] or None,#12 entry_date [9]
            woo[10],#13 company_ref [10]
            woo_auto_key or 0,#14 woo_auto_key [11]
            woo[18],#WT.WORK_TYPE (Manager)
            woo[17],#SU.USER_NAME (WO Work Type)
            woo[12],#17 rank [12]
            woo[13] or 0,#18 wos_auto_key
            stm_auto_key or 0,#19 stm_auto_key
            1,#20 active
            is_dashboard,#21
            user_id or '',#22 
            session_id or '',#23
            is_racking,#24
            woo[15] or '',#25 #VWO.PARENT_WO FROM WO_OPERATION
            gate1_qty > 0 and gate1 or '',
            gate2_qty > 0 and gate2 or '',
            gate3_qty > 0 and gate3 or '',
            gate4_qty > 0 and gate4 or '',
            gate1_qty > 0 and gate1_qty or None,
            gate2_qty > 0 and gate2_qty or None,
            gate3_qty > 0 and gate3_qty or None,
            gate4_qty > 0 and gate4_qty or None,
            is_toll and woo[20] or 0,
            quapi,
            is_detail or 0,
            qty_at_gate or 0,#qty at gate
            ult_parent or 0,
            condition_code or '',
            stock_owner or '',
            loc_validated_date,
            has_labor and "green-check.png" or 'blank.png',#notes field
            woo[12] or 0,#rank
            len(woo)>= 40 and woo[39] and isinstance(woo[39],str)\
            and len(woo[39])>=11 and woo[39][:10] or None,#manual_ecd
            quantity,#quantity
            len(woo)>= 40 and woo[41],#remarks
            len(woo)>= 40 and woo[42],#cons code
            len(woo)>= 40 and woo[43],#uom_code
        ]
        if is_racking:
            #WH.WAREHOUSE_CODE,UDL.UDL_CODE,S.CTRL_ID,S.CTRL_NUMBER
            woo_data = woo_data and woo_data + woo[15:19] or []
        else:
            woo_data = woo_data and woo_data + [wh_code,woo[23],0,0] or []
        woo_records.append(woo_data)

    create_recs = woo_records and create_in_bulk(woo_records,wos_obj) or False
    if not create_recs:
        error = 'Problem creating the data locally in create_in_bulk'
    if not error and not recs:
        error = 'No records found.'
    return error,msg
    
def create_in_bulk(woo_recs, wos_obj):
    objects = []
    for wo_number,status,time_status,due_date,due_date_var,stock_line,\
        part_number,description,serial_number,location_code,time_loc,\
        customer,entry_date,cust_ref_number,woo_auto_key,wo_type,manager,\
        rank,wos_auto_key,stm_auto_key,active,is_dashboard,user_id,session_id,\
        is_racking,parent_wo,gate_1_text,gate_2_text,gate_3_text,gate_4_text,\
        gate_1_qty,gate_2_qty,gate_3_qty,gate_4_qty,sub_wo_gate,quapi,is_detail,\
        gate_qty,parent_auto_key,condition_code,stock_owner,loc_validated_date,\
        has_labor,int_rank,exp_date,quantity,remarks,consignment,uom_code,\
        wh_code,rack,ctrl_id,ctrl_number in woo_recs:
        objects.append(wos_obj(
                wo_number = wo_number,#0
                status = status,#1
                time_status = time_status,#2
                due_date = due_date,#3
                due_date_var = due_date_var,#4
                stock_line = stock_line,#5
                part_number = part_number,#6
                description = description,#7
                serial_number = serial_number,#8
                location_code = location_code,#9
                time_loc = time_loc,#10
                customer = customer,#11
                entry_date = entry_date,#12
                cust_ref_number = cust_ref_number,#13
                woo_auto_key = woo_auto_key,#14
                wo_type = wo_type,#15
                manager = manager,#16
                rank = rank,#17
                wos_auto_key=wos_auto_key,#18
                stm_auto_key=stm_auto_key,#19
                active = active,#20
                is_dashboard = is_dashboard,#21
                user_id = user_id or '',#22
                session_id = session_id or '',#23 
                is_racking = is_racking,#24
                parent_wo = parent_wo,#25
                gate_1_text=gate_1_text,#26
                gate_2_text=gate_2_text,#27
                gate_3_text=gate_3_text,#28
                gate_4_text=gate_4_text,#29
                gate_1_qty=gate_1_qty,#30
                gate_2_qty=gate_2_qty,#31
                gate_3_qty=gate_3_qty,#32
                gate_4_qty=gate_4_qty,#33
                sub_wo_gate=sub_wo_gate,#34
                is_detail = is_detail,#35
                gate_qty = gate_qty,#36
                parent_auto_key = parent_auto_key,#37
                condition_code = condition_code,#38
                stock_owner=stock_owner,#39
                loc_validated_date=loc_validated_date,#40
                wh_code = wh_code,#41
                rack = rack,#42
                ctrl_id = ctrl_id,#43
                ctrl_number = ctrl_number,#44
                quapi_id = quapi,#45
                notes = has_labor,#46
                int_rank = int_rank,#s
                exp_date=exp_date,#48
                quantity=quantity,#49
                remarks=remarks,
                consignment_code=consignment,
                uom_code=uom_code,
        ))
    wos_obj.objects.bulk_create(objects)
    return True
    
def get_wo_status(woo_auto_key,user_id='',quapi=None):  
    #assuming nothing about stock but that status is set on WO 
    query = """
        SELECT W.SI_NUMBER, WS.DESCRIPTION, W.DUE_DATE, S.STOCK_LINE, P.PN, P.DESCRIPTION, 
          S.SERIAL_NUMBER, L.LOCATION_CODE, WS.SEVERITY FROM WO_OPERATION W
          LEFT JOIN STOCK_RESERVATIONS SR ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
          LEFT JOIN STOCK S ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
          LEFT JOIN PARTS_MASTER P ON S.PNM_AUTO_KEY = P.PNM_AUTO_KEY
          LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
          LEFT JOIN WO_STATUS WS ON W.WOS_AUTO_KEY=WS.WOS_AUTO_KEY
        WHERE
          (W.WOS_AUTO_KEY IS NULL 
          OR WS.STATUS_TYPE IN ('Open','Delay','Defer'))
          AND W.WOO_AUTO_KEY = %s"""%woo_auto_key   
    return selection(query, user_id=user_id, quapi=quapi) 
#============================================================================QUANTUM QUERY METHODS===================================================================    
#====================================================GENERAL HELPER/QUERY QUANTUM METHODS================================================

import cx_Oracle
def jsonicize(recs):    
    from django.http import JsonResponse
    #responseData = {
    #    'recs':recs,
    #}
    return JsonResponse(recs,safe=False)
    
def orcl_connect(schema):
    cr=None
    con=None
    if schema:
        try:
            #import pyodbc
            #con_str = 'DSN=QCTL;pwd=quantum'
            #con = pyodbc.connect(con_str)
            #pool = cx_Oracle.SessionPool(schema.schema, schema.db_user, schema.host + ':' + str(schema.port) + '/' + schema.sid, min=1, max=3, increment=1, encoding="UTF-8")
            #con = pool.acquire()
            if schema.db_pwd and schema.db_pwd != 'none':
                con = cx_Oracle.connect(user=schema.db_user,\
                    password=schema.db_pwd,\
                    dsn=schema.host+':'+str(schema.port)+'/'+schema.sid)
            else:
                connstr = schema.schema + '/' + schema.db_user + '@' + schema.host + ':' + str(schema.port) + '/' + schema.sid
                con = cx_Oracle.connect(connstr)
            """from django.contrib.auth.models import User
            from django.contrib.sessions.models import Session
            from django.utils import timezone
            active_sessions = Session.objects.filter(expire_date__gte=timezone.now())
            #active_sessions = Session.objects.all()
            num_sessions = len(active_sessions)
            if num_sessions > 2:
                pool = cx_Oracle.SessionPool(schema.schema, schema.db_user, schema.host + ':' + str(schema.port) + '/' + schema.sid, min=2, max=3, increment=1, encoding="UTF-8")
                con = pool.acquire()"""
            #print('successful connection to Oracle')
        except Exception as exc:
            error, = exc.args
    cr = con and con.cursor() or None
    if not con:
        return False,False
    return cr,con
    
def insertion_dir(query,cr,quapi=None):
    recs = []
    msg = ''
    if not cr and quapi:
        
        from polls.models import OracleConnection as oc
        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
        orcl_conn = orcl_conn and orcl_conn[0] or None
        if orcl_conn:
            cr,con = orcl_connect(orcl_conn)
        if not (cr and con):
            return 'Cannot connect to Oracle.'
    try:
        cr.execute(query)      
    except Exception as exc:
        error, = exc.args
        msg = error.message
        print("Oracle-Error-Code:", error.code)
        print("Oracle-Error-Message:", msg)
    return str(msg)
       
def selection_dir(query,cr):
    recs = []
    res = []
    all_res = []  
    try:
        cr.execute(query)
        recs = cr.fetchall()        
    except cx_Oracle.DatabaseError as exc:
        error, = exc.args
        print("Oracle-Error-Code:", error.code)
        print("Oracle-Error-Message:", error.message)
    if recs:       
        for count,rec in enumerate(recs):
            res = ['' if (field == None or field == 'None' or field == 'Null' or field == 'null' or field=='none') else field for field in rec]
            all_res.append(res)
    import json
    all_res = json.dumps(all_res,default=str)
    import ast
    recs = ast.literal_eval(all_res)
    return recs
    
def updation_dir(query,cr):   
    msg = ''
    try:
        cr.execute(query)                  
    except cx_Oracle.DatabaseError as exc:
        error, = exc.args
        msg = " Oracle-Error-Message:" + str(error.message) 
    msg = not msg and '{"recs": ""}' or msg
    return str(msg)   
      
def updation(query,user_id='',quapi=None):
    msg = ''  
    if query and quapi:
        from polls.models import OracleConnection as oc
        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
        orcl_conn = orcl_conn and orcl_conn[0] or None
        if orcl_conn:
            cr,con = orcl_connect(orcl_conn)
        if not (cr and con):
            return 'Cannot connect to Oracle' 
        else:
            msg = updation_dir(query,cr) 
    return msg
    
def insertion(query,user_id='',quapi=None):
    recs = []
    message = ''
    new_auto_key = None
    if query and quapi:
        #must be that it is an API call so we reroute it to the API
        url = quapi.conn_str
        url = url and url + 'insertion' or None
        params = {
            'query': query,
            'user_id': user_id,
            'type': 'insertion',
            'schema': quapi.orcl_conn_id or 0,
        }
        response = requests.post(url,json=params,params=params)
        success = (response.status_code == 200)  # 200 =  Success
        json =  success and response.json() or {}
        records = json and json['recs'] or [[]]
        for row in records:
            new_auto_key = row and row[0] or None
            break
    return response and response.text or ''
                       
def selection(query,table_name='',quapi=None,user_id=None):
    recs = []
    if query and quapi:
        from polls.models import OracleConnection as oc
        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
        orcl_conn = orcl_conn and orcl_conn[0] or None
        if orcl_conn:
            cr,con = orcl_connect(orcl_conn)
        if not (cr and con):
            return 'Cannot connect to Oracle' 
        else:
            recs = selection_dir(query,cr)              
    return recs

def orcl_commit(user_id='',quapi=None, con=None):
    #commit the database updates
    msg = ''
    if quapi and not con:
        from polls.models import OracleConnection as oc
        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
        orcl_conn = orcl_conn and orcl_conn[0] or None
        cr,con = orcl_connect(orcl_conn)
    if not con:
        return 'Cannot connect to Oracle' 
    try:    
        con.commit()
        con.close()
    except cx_Oracle.DatabaseError as exc:
        error, = exc.args
        msg = error.message
        logger.error("Oracle-Error-Code: '%s'",error.code)
        logger.error("Oracle-Error-Message: '%s'",msg)          
    return msg
    
    
def benchmark_test(benchmark,query,quapi):
    msg = ''
    quapi = QueryApi.objects.filter(id=1)
    query = "UPDATE STOCK SET LOC_AUTO_KEY=1 WHERE STM_AUTO_KEY=2605"
    benchmark(updation(query,quapi=quapi))
    return msg
    
#========================================================PI UPDATES CODE================================================================================
def lookup_stm_auto_key(ctrl_id,ctrl_number,stm=None,user_id='',quapi=None):
    #where_clause = "WHERE HISTORICAL_FLAG='F' AND QTY_OH>0"
    ##stm,qty_oh,syscm,pcc,cnc,loc_auto_key,whs_auto_key,stc,dpt,str,qty_reserved
    if not stm:
        query = """SELECT STM_AUTO_KEY,QTY_OH,SYSCM_AUTO_KEY,
            PCC_AUTO_KEY,CNC_AUTO_KEY,LOC_AUTO_KEY,WHS_AUTO_KEY,
            STC_AUTO_KEY,DPT_AUTO_KEY,PNM_AUTO_KEY,CTRL_NUMBER,CTRL_ID
            FROM STOCK WHERE CTRL_ID = %s AND CTRL_NUMBER = %s
            """%(ctrl_id,ctrl_number)
    else:
        query = """SELECT STM_AUTO_KEY,QTY_OH,SYSCM_AUTO_KEY,
            PCC_AUTO_KEY,CNC_AUTO_KEY,LOC_AUTO_KEY,WHS_AUTO_KEY,
            STC_AUTO_KEY,DPT_AUTO_KEY,PNM_AUTO_KEY,CTRL_NUMBER,CTRL_ID
            FROM STOCK WHERE STM_AUTO_KEY = '%s'
            """%(stm)
        
    res = selection(query, user_id=user_id, quapi=quapi)
    return res                       
    
def lookup_batch(batch_no,stm_auto_key=None,user_id='',quapi=None):
    locations = []
    loc_from,loc_to = '','' 
    pid_auto_key,pid = None,None
    #get the pih_auto_key from pi_header for that batch
    query = "SELECT PIH_AUTO_KEY,LOC_FROM,LOC_TO FROM PI_HEADER WHERE PI_NUMBER = '%s'"%batch_no
    pi_header = selection(query, user_id=user_id,quapi=quapi) 
    pih = pi_header and pi_header[0] or None
    pih_auto_key = pih and pih[0] or None
    #get the pid_auto_key from PI_DETAIL
    if pih_auto_key and stm_auto_key:
        query = "SELECT PID_AUTO_KEY,LOC_AUTO_KEY FROM PI_DETAIL WHERE PIH_AUTO_KEY= %s AND STM_AUTO_KEY = %s"%(pih_auto_key,stm_auto_key)
        pi_detail = selection(query, user_id=user_id, quapi=quapi)
        pid = pi_detail and pi_detail[0] or None
        pid_auto_key = pid and pid[0] or None
    return pid,pid_auto_key,pih,pih_auto_key
    
def get_loc_code(loc_auto_key,user_id='',quapi=None):
    query = "SELECT LOCATION_CODE FROM LOCATION WHERE LOC_AUTO_KEY=%s"%loc_auto_key
    recs = selection(query, user_id=user_id, quapi=quapi)
    return recs

def insert_detail_record(loc_auto_key,pih_auto_key,stock_rec,sysur_auto_key,qty_oh,new_qty,ctrl_id,ctrl_number,user_id='',cr=None):
    stm_auto_key = stock_rec and stock_rec[0] or ''
    pnm_auto_key = stock_rec and stock_rec[9] or ''
    whs_auto_key = stock_rec and stock_rec[6] or ''
    if not (pnm_auto_key and stm_auto_key):
        error = 'Something is wrong with the stock.'
    else:
        #insert a new  PID_AUTO_KEY into PI_DETAIL with the LOC_AUTO_KEY = where that LOC_AUTO_KEY = (user input for LOCATION_CODE)
        query = """INSERT INTO PI_DETAIL (STM_AUTO_KEY,PID_AUTO_KEY,SYSUR_AUTO_KEY,QTY,QTY_FOUND,
            CTRL_ID,CTRL_NUMBER,PIH_AUTO_KEY,PNM_AUTO_KEY,WHS_AUTO_KEY,LOC_AUTO_KEY) 
            VALUES ('%s',G_PID_AUTO_KEY.NEXTVAL,%s,'%s','%s','%s','%s',%s,'%s','%s','%s')"""%(stm_auto_key or '',sysur_auto_key,qty_oh,int(new_qty),ctrl_id,ctrl_number,pih_auto_key,pnm_auto_key or '',whs_auto_key or '',loc_auto_key or '')                
        error = insertion_dir(query,cr)  
    return error
    
def qty_to_float(text):
    import re
    error,new_float = '',''
    
    match = re.findall(r'(\d+(?:\.\d+)?)', text)
    if not match:
        error = 'Quantity must be a number.'
    else:
        try:
            new_float = float(text)
        except ValueError as error:
            return str(error),new_float
    return error,new_float
    
def loc_stock_transfer(cr,srec):
    error = ''
    """ stm,qty_oh,syscm,pcc,cnc,loc_auto_key,whs_auto_key,stc,dpt,str,qty_reserved
        params.append(rec[14])#0.stm_auto_key
        params.append(qty_input!=None and qty_input or rec[29])#1.qty_oh
        params.append(syscm_auto_key or rec[30] or 1)#2.syscm_auto_key
        params.append(rec[31] or '')#3.pcc_auto_key                       
        params.append(rec[32] or '')#4.cnc_auto_key
        params.append(loc_auto_key or rec[21] or '')#5.loc_auto_key 
        params.append((valid_wh and whs_auto_key) or rec[22] or '')#6.whs_auto_key
        params.append(rec[33] or '')#7.stc_auto_key
        params.append(dpt_auto_key or rec[34] or '')#8.dpt_auto_key
        params.append(rec[20] or '')#str_auto_key  
        params.append(rec[35])#qty_reserved
        params.append(rec[36] or '')#sod_auto_key
        params.append(rec[37] or '')#rod_auto_key
        params.append(rec[38] or '')#wob_auto_key
        params.append(rec[39] or '')#pod_auto_key
        params.append(rec[11] or '')#woo_auto_key 
    """
    #stm,qty_oh,syscm,pcc,cnc,loc_auto_key,whs_auto_key,stc,dpt,str,qty_reserved
    stm = srec[0]
    qty_oh = srec[1]
    syscm = srec[2]
    pcc = srec[3]
    cnc = srec[4]
    loc = srec[5]
    whs = srec[6]
    stc = srec[7]
    dpt = srec[8]
    #stra = srec[9]
    #qty_reserved = srec[10]
    squery = """DECLARE CT qc_utl_pkg.cursor_type; BEGIN CT := 
    QC_STOCK_PKG.SPI_STOCK_TRANSFER('%s', %s, '%s', '%s', '%s', '%s', '%s', '%s', '%s'); close CT; END;
    """%(stm,qty_oh,syscm,pcc,cnc,loc,whs,stc,dpt)
    try:
        updation_dir(squery,cr)
    except cx_Oracle.DatabaseError as exc:
        error, = exc.args
        msg = error.message
        logger.error("Error with stock transfer: '%s'",msg)    
    return error
@shared_task
def make_pi_updates(session_id,\
    batch_no,ctrl_id,ctrl_number,\
    new_qty,stock_label,user_id,\
    sysur_auto_key,quapi_id=None,\
    loc_input='',location_key=0,\
    stm=''):
    audit_ok,reent,pid_auto_key,pih,aud_status = True,False,None,None,'failure'
    qty,msg,error,new_auto_key,pih_auto_key,audit_ok,set_clause = '','','','','',False,''   
    from polls.models import PILogs,QueryApi
    if new_qty:
        error,new_qty = qty_to_float(new_qty)
        if error:
            return error,msg
    else:
        return 'Must enter qty.',msg
    if stock_label and not stock_label.isnumeric():
        error+="Stock label must be a number.  "
        reent = '1' 
        return error,msg        
    if batch_no and not batch_no.isalnum():
        error+="Batch must be alpha-numeric."
        return error,msg
    quapi = quapi_id and QueryApi.objects.filter(id=quapi_id) or None
    quapi = quapi and quapi[0] or None
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    cr,con = orcl_connect(orcl_conn)    
    if not cr and con:
        error = 'Not connected to Quantum.'
        return error,msg        
                                                                                      
    if not ctrl_id and stock_label:
        ctrl_number = stock_label[:6]
        ctrl_id = stock_label[7:]
    stock_rec = lookup_stm_auto_key(\
        ctrl_id=ctrl_id,\
        ctrl_number=ctrl_number,\
        stm=stm,\
        user_id=user_id,quapi=quapi)     
    if not stock_rec and (ctrl_number and ctrl_id):
        ctrl_number = ctrl_number + '0'
        stock_rec = lookup_stm_auto_key(ctrl_id=ctrl_id,ctrl_number=ctrl_number,user_id=user_id, quapi=quapi) 
        if not stock_rec and (ctrl_number and ctrl_id):
            ctrl_number = ctrl_number + '0'
            stock_rec = lookup_stm_auto_key(ctrl_id=ctrl_id,ctrl_number=ctrl_number,user_id=user_id,quapi=quapi)         
        if not stock_rec:               
            error = 'Stock line, %s, doesn\'t exist.'%stock_label    
    stock_rec = stock_rec and stock_rec[0] or None
    #stm,qty_oh,syscm,pcc,cnc,loc_auto_key,whs_auto_key,stc,dpt,str,qty_reserved
    stm_auto_key = stock_rec and stock_rec[0] or ''
    if not stm_auto_key:
        error = 'Stock record not found.'
        return error,''
    qty_oh = stock_rec and stock_rec[1] or 0 
    syscm_auto_key = stock_rec and stock_rec[2] or ''
    pcc_auto_key = stock_rec and stock_rec[3] or ''
    cnc_auto_key = stock_rec and stock_rec[4] or ''
    loc_auto_key = stock_rec and stock_rec[5] or ''
    whs_auto_key = stock_rec and stock_rec[6] or ''
    stc_auto_key = stock_rec and stock_rec[7] or ''
    dpt_auto_key = stock_rec and stock_rec[8] or ''
    pnm_auto_key = stock_rec and stock_rec[9] or ''
    ctrl_number = stock_rec and stock_rec[10] or ''
    ctrl_id = stock_rec and stock_rec[11] or ''
    str_auto_key,qty_reserved = '',0   
    #need to check the pnm_auto_key to see if the part is serialized
    
    qty = new_qty and int(new_qty) or 0
    if qty and pnm_auto_key and qty > 1 or qty < -1:
        query = "SELECT SERIALIZED FROM PARTS_MASTER WHERE PNM_AUTO_KEY=%s AND SERIALIZED='T'"%pnm_auto_key
        serialized = selection_dir(query,cr)
        if serialized:
            error = 'Serialized parts cannot have quantity > 1 for label, %s.'%stock_label
    whs_auto_key = stock_rec and stock_rec[2] or None
    
    if not location_key and loc_input:
        query="""SELECT LOC_AUTO_KEY FROM LOCATION
            WHERE UPPER(LOCATION_CODE) = UPPER('%s')"""%loc_input
        loc = selection_dir(query,cr)
        loc_auto_key = loc and loc[0] and loc[0][0] or ''
        location_key = loc_auto_key
                                   
    if batch_no and not error:    
        pid,pid_auto_key,pih,pih_auto_key = lookup_batch(batch_no,stm_auto_key,user_id=user_id,quapi=quapi)
        #loc_auto_key = pid and pid[1] or None
        #if not loc_auto_key:
        #    loc_auto_key = stock_rec and stock_rec[3] or None
        #location_code = loc_auto_key and get_loc_code(loc_auto_key,user_id=user_id,quapi=quapi) or None
        #location_code = location_code and location_code[0] and location_code[0][0] or None
    right_now = datetime.now()
    now = right_now.strftime('%Y-%m-%d %H:%M:%S')

    if pid_auto_key and not error:
        #TODO: need to update stm with new location 
        if stm_auto_key:   
            #update the pi_detail table with the stm_auto_key    
            where_clause = 'WHERE PID_AUTO_KEY = %s'%pid_auto_key           
            #now update QTY_FOUND and location in PI_DETAIL
            if loc_auto_key:
                set_clause = ", LOC_AUTO_KEY = %s"%loc_auto_key
                                        
                                   
                                        
            query = """UPDATE PI_DETAIL 
                SET QTY_FOUND = %s,
                SYSUR_AUTO_KEY = %s%s %s
                """%(int(new_qty),sysur_auto_key,set_clause,where_clause)   
            error = updation_dir(query,cr)
            #find the correct warehoue for the location and set them both on STOCK table.
            if location_key:
                query = """SELECT WHS_AUTO_KEY FROM WAREHOUSE_LOCATIONS 
                    WHERE LOC_AUTO_KEY = %s"""%location_key
                whs = selection_dir(query,cr)
                whs_auto_key = whs and whs[0] and whs[0][0] or ''
                #stm,qty_oh,syscm,pcc,cnc,loc_auto_key,whs_auto_key,stc,dpt,str,qty_reserved
                stock_rec = [stm_auto_key,qty_oh,syscm_auto_key]
                stock_rec += [pcc_auto_key,cnc_auto_key,loc_auto_key]
                stock_rec += [whs_auto_key,stc_auto_key,dpt_auto_key]
                stock_rec += [str_auto_key,qty_reserved]
                
                error = loc_stock_transfer(cr,stock_rec)                           
                            
                query = """UPDATE STOCK
                        SET LOC_VALIDATED = TO_TIMESTAMP('%s', 'YYYY-MM-DD HH24:MI:SS'),
                        LOC_AUTO_KEY = %s
                        WHERE STM_AUTO_KEY = %s
                        """%(now,loc_auto_key,stm_auto_key) 
                error = updation_dir(query,cr)    
                query = """
                    CREATE OR REPLACE PROCEDURE "PI_STOCK_UPDATE"
                    (QUSER IN NUMBER, STM IN NUMBER, QCODE IN VARCHAR2, LOC_KEY IN NUMBER)  AS
                    v_query number;
                    v_sysur number;
                    v_pwd varchar2(150);
                    V_CURSOR QC_SC_PKG.CURSOR_TYPE ;
                    BEGIN                 
                        begin
                        qc_trig_pkg.disable_triggers;
                        UPDATE SA_LOG SET SYSUR_AUTO_KEY = QUSER, 
                            EMPLOYEE_CODE = QCODE 
                          WHERE STA_AUTO_KEY = (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = STM AND NEW_LOC_AUTO_KEY = LOC_KEY AND EMPLOYEE_CODE = 'DBA');
                        qc_trig_pkg.enable_triggers;
                        end;
                     END PI_STOCK_UPDATE;"""
                
                error = updation_dir(query,cr)   
                run_proc = """
                    BEGIN
                    PI_STOCK_UPDATE('%s',%s,'%s',%s);
                    END;   
                """%(sysur_auto_key,stm_auto_key,user_id[:9],location_key)             
                error = updation_dir(run_proc,cr)            
        #update the pi_detail table with the stm_auto_key    
        where_clause = 'WHERE PID_AUTO_KEY = %s'%pid_auto_key           
        #now update QTY_FOUND in PI_DETAIL
        query = "UPDATE PI_DETAIL SET QTY_FOUND = %s, SYSUR_AUTO_KEY = %s %s"%(int(new_qty),sysur_auto_key,where_clause)   
        error = updation_dir(query,cr)                                                                                                               
        aud_sub = """SELECT ADT_AUTO_KEY FROM AUDIT_TRAIL 
        WHERE ROWNUM < = 1 AND SOURCE_TABLE = 'PIH'
        ORDER BY ADT_AUTO_KEY DESC"""
        adt = selection_dir(aud_sub,cr)
        adt_auto_key = adt and adt[0] and adt[0][0] or 0
        if adt_auto_key:
            query = """UPDATE AUDIT_TRAIL 
                 SET SYSUR_AUTO_KEY=%s 
                 WHERE ADT_AUTO_KEY = %s
                 """%(sysur_auto_key,adt_auto_key)
            error = updation_dir(query,cr)
        orcl_commit(con=con)
        audit_ok = True
        pi_log = PILogs.objects.create(
            stock_label = stock_label,
            batch_no = batch_no,
            location_code = loc_input, 
            quantity = int(new_qty),
            ctrl_number = ctrl_number,
            ctrl_id = ctrl_id,
            stm_auto_key = stm_auto_key,
            active = 1,
            user_id = user_id,
            session_id = session_id,
        )
        pi_log.save()
        msg = ' Successful update. '  

    elif not pid_auto_key and pih_auto_key and not error:  
        
        #TODO: need to update stm with new location
        if stm_auto_key: 
            error = insert_detail_record(loc_auto_key,pih_auto_key,stock_rec,sysur_auto_key,qty_oh,new_qty,ctrl_id,ctrl_number,user_id=user_id,cr=cr)         
            if location_key:
                #find the correct warehoue for the location and set them both on STOCK table.
                query = """SELECT WHS_AUTO_KEY FROM WAREHOUSE_LOCATIONS 
                    WHERE LOC_AUTO_KEY = %s"""%location_key
                whs = selection_dir(query,cr)
                whs_auto_key = whs and whs[0] and whs[0][0] or 0
                #stock_rec = [stm_auto_key,qty_oh,syscm_auto_key,location_key,whs_auto_key]
                #stm,qty_oh,syscm,pcc,cnc,loc_auto_key,whs_auto_key,stc,dpt,str,qty_reserved
                stock_rec = [stm_auto_key,qty_oh,syscm_auto_key]
                stock_rec += [pcc_auto_key,cnc_auto_key,location_key]
                stock_rec += [whs_auto_key,stc_auto_key,dpt_auto_key]
                stock_rec += [str_auto_key,qty_reserved]
                error = loc_stock_transfer(cr,stock_rec)
                query = """UPDATE STOCK
                        SET LOC_VALIDATED = TO_TIMESTAMP('%s', 'YYYY-MM-DD HH24:MI:SS'),
                        LOC_AUTO_KEY = %s
                        WHERE STM_AUTO_KEY = %s
                        """%(now,loc_auto_key,stm_auto_key) 
                error = updation_dir(query,cr)            
            query = """
                CREATE OR REPLACE PROCEDURE "PI_STOCK_UPDATE"
                (QUSER IN NUMBER, STM IN NUMBER, QCODE IN VARCHAR2, LOC_KEY IN NUMBER)  AS
                v_query number;
                v_sysur number;
                v_pwd varchar2(150);
                V_CURSOR QC_SC_PKG.CURSOR_TYPE ;
                BEGIN                 
                    begin
                    qc_trig_pkg.disable_triggers;
                    UPDATE SA_LOG SET SYSUR_AUTO_KEY = QUSER, EMPLOYEE_CODE = QCODE WHERE STA_AUTO_KEY = (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = STM AND EMPLOYEE_CODE = 'DBA');
                    qc_trig_pkg.enable_triggers;
                    end;
                 END PI_STOCK_UPDATE;"""          
            error = updation_dir(query,cr)   
            run_proc = """
                BEGIN
                PI_STOCK_UPDATE('%s',%s,'%s',%s);
                END;   
            """%(sysur_auto_key,stm_auto_key,user_id[:9],location_key)            

        aud_sub = """SELECT ADT_AUTO_KEY FROM AUDIT_TRAIL 
            WHERE ROWNUM <= 1 AND SOURCE_TABLE = 'PIH'
            ORDER BY ADT_AUTO_KEY DESC"""
        adt = selection_dir(aud_sub,cr)
        adt_auto_key = adt and adt[0] and adt[0][0] or 0
        if adt_auto_key:
            query = """UPDATE AUDIT_TRAIL SET SYSUR_AUTO_KEY=%s WHERE ADT_AUTO_KEY = %s"""%(sysur_auto_key,adt_auto_key)
            error = updation_dir(query,cr)
        #if error != '{"recs": ""}':
        orcl_commit(con=con)
        audit_ok=True
        pi_log = PILogs.objects.create(
            stock_label = stock_label,
            batch_no = batch_no,
            location_code = loc_input, 
            quantity = int(new_qty),
            ctrl_number = ctrl_number,
            ctrl_id = ctrl_id,
            stm_auto_key = stm_auto_key,
            active = 1,
            user_id = user_id,
            session_id = session_id,
        )
        pi_log.save() 
        msg = ' Successful update. '        
    #if not pi_detail records with that stm_auto_key, then we have to prompt the user for a location
    #elif not pid_auto_key and pih:
        #return 'show_modal',''      
    #now we log what we have just done by creating the PILogs object with all of the data.    
    elif not error:
        error ='No batch exists.'
    if error == '{"recs": ""}':
        error=''
        aud_status = 'success'
    else: 
        aud_status = 'failure'         
    from polls.models import MLApps as maps,QuantumUser as qu
    app_id = maps.objects.filter(code='physical-inventory')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        rec_input = (ctrl_id and 'Record with Ctrl#: '+ str(ctrl_number) + ' and Ctrl ID#: ' + str(ctrl_id)) or stock_label or ''
        new_val = rec_input and (rec_input + ', Batch: ' + str(batch_no) + ', Location Code: ' + str(loc_input) + ', Qty: ' + str(new_qty)) or ''
        if not new_val:
            new_val = 'Inventory detail record created for stock label: '+ str(stock_label) + ', batch: ' + str(batch_no) + ',' +  'location: ' + str(loc_input) + ' new qty: ' + str(new_qty) or ''
        field_changed = 'quantity | ' + error
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')
        error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status) 
        if error == '{"recs": ""}':
            error=''
    else:
        error = 'Incorrect Quantum User ID.'
    return error,msg
    
def verify_loc_input(locations,loc_auto_key):
    res = loc_auto_key in (item for loc_list in locations for item in loc_list) 
    return res
   
#============================================================================Racking=====================================================================
def check_if_same_whs(whs_code, woo_auto_key, user_id='', quapi=None):
    """
    Arguments: location - code supplied by user form input
            woo_auto_key - from a lookup of SI_NUMBER (wo_number) supplied by user form input           
    Returns: results from location       
    """
    update_ok = False
    stm_auto_key = ''
    #location code is what they enter/scan => 
    #from location - lookup location table to find location code
    query = """
    SELECT S.STM_AUTO_KEY, WH.WHS_AUTO_KEY, WH.WAREHOUSE_CODE 
        FROM WO_OPERATION W, WO_STATUS WS,STOCK_RESERVATIONS SR, STOCK S, WAREHOUSE WH 
        WHERE
            W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY 
            AND SR.STM_AUTO_KEY = S.STM_AUTO_KEY 
            AND S.WHS_AUTO_KEY = WH.WHS_AUTO_KEY 
            AND W.WOO_AUTO_KEY = %s"""%(woo_auto_key)
    results = selection(query, user_id=user_id, quapi=quapi)      
    result = results and results[0] or None
    stm_auto_key = result and result[0] or None
    if type(stm_auto_key) is list:
        stm_auto_key = stm_auto_key[0]
    if result and result[2] and result[2] != whs_code:
        update_ok = True           
    return stm_auto_key,update_ok        
    
def update_wh(warehouse, whs_auto_key, woo_auto_key, user_id, wos_obj, rack_auto_key=None, quapi=None):  
    stm_auto_key,update_ok = check_if_same_whs(warehouse, woo_auto_key, user_id, quapi=quapi)
    msg = ''
    if stm_auto_key and update_ok and whs_auto_key:
        where_clause = rack_auto_key and ' AND IC_UDL_005 = %s'%rack_auto_key or ''
        query = "UPDATE STOCK SET whs_auto_key = %s WHERE stm_auto_key = %s%s"%(whs_auto_key,stm_auto_key,where_clause)
        msg = updation(query, user_id=user_id, quapi=quapi)
    else: 
        msg += 'Not able to update the warehouse.'
    #if not msg and not rack_auto_key:
    #    woo_object = wos_obj.objects.filter(woo_auto_key=woo_auto_key, is_racking=0, is_dashboard=0, active=1, user_id=user_id)
    #    woo_object = woo_object and woo_object.update(wh_code=warehouse)    
    return msg
    
def get_whs_from_loc(loc_auto_key, user_id, quapi=None):
    query = "SELECT WHS_AUTO_KEY FROM WAREHOUSE_LOCATIONS WHERE LOC_AUTO_KEY=%s"%loc_auto_key
    res = selection(query, user_id=user_id, quapi=quapi)
    return res
    
def check_if_valid_whs_beta(whs_auto_key=None,rack_auto_key=None,loc_auto_key=None,stm_auto_key=None,existing_loc_key=None,existing_whs_key=None,dj_user_id='',quapi=None):
    whs_needs_update,valid_wh,msg,loc_whs_keys,loc_whs_key,error = False,False,'',[],None,''
    #check to see if we must update the warehouse because the location belongs to a different warehouse
    #checking to see if location change to NDT 1 updates warehouse to GATE 1"
    #if the user didn't enter a whs but did enter a loc,
    #then we must check that the user entered a loc has either only one warehouse (case _1)
    
    if existing_loc_key:
        query = """SELECT WHS_AUTO_KEY,LOC_AUTO_KEY FROM
            WAREHOUSE_LOCATIONS WHERE LOC_AUTO_KEY = %s
            """%(existing_loc_key)
        loc_whs_keys = selection(query,quapi=quapi)  
    elif loc_auto_key and not whs_auto_key:
        query = """SELECT WHS_AUTO_KEY,LOC_AUTO_KEY FROM
            WAREHOUSE_LOCATIONS WHERE LOC_AUTO_KEY = %s
            """%(loc_auto_key)
        loc_whs_keys = selection(query,quapi=quapi)
    elif whs_auto_key and not loc_auto_key:
        query = """SELECT WHS_AUTO_KEY,LOC_AUTO_KEY FROM
            WAREHOUSE_LOCATIONS WHERE WHS_AUTO_KEY = %s
            """%(whs_auto_key)
        loc_whs_keys = selection(query,quapi=quapi)
    elif whs_auto_key and loc_auto_key:
        query = """SELECT WHS_AUTO_KEY,LOC_AUTO_KEY FROM
            WAREHOUSE_LOCATIONS WHERE WHS_AUTO_KEY = %s
            """%(whs_auto_key)
        loc_whs_keys = selection(query,quapi=quapi)
        if not loc_whs_keys:
            query = """SELECT WHS_AUTO_KEY,LOC_AUTO_KEY FROM
                WAREHOUSE_LOCATIONS WHERE LOC_AUTO_KEY = %s
                """%(loc_auto_key)
            loc_whs_keys = selection(query,quapi=quapi)
            if not loc_whs_keys:
                error = 'Location not valid.'
                return msg,error,valid_wh,loc_whs_key,whs_needs_update
            else:
                error = 'Warehouse not valid.'     
    if not whs_auto_key and loc_auto_key and loc_whs_keys:
        #must make sure we can update to this warehouse.
        #we can update if there is only one. 
        if len(loc_whs_keys) == 1:
            loc_whs_key = loc_whs_keys[0] and loc_whs_keys[0][0] or None
            valid_wh = True
            whs_needs_update = existing_whs_key!=loc_whs_key or False     
        else:
            error = not valid_wh and 'Multiple warehouses found.' or '' 
    #user entered warehouse and loc
    if whs_auto_key and loc_auto_key and loc_whs_keys:  
        #if there is only one warehouse and it doesn't match, we cannot update, otherwise we do
        if len(loc_whs_keys) == 1:
            loc_whs_key = loc_whs_keys[0] and loc_whs_keys[0][0] or None
            if whs_auto_key and loc_whs_key and loc_whs_key != whs_auto_key:
                error = 'Try again. Warehouse not valid for the location.' 
            else:
                valid_wh = True 
                loc_whs_key = whs_auto_key 
                whs_needs_update = True  
                msg += "  Warehouse updated." 
        #if there is more than 1 warehouse, we just simply look for a match with the user-entered and if we are good, we update                
        elif len(loc_whs_keys) > 1:
            for loc_whs in loc_whs_keys:
                if loc_whs[0] == whs_auto_key:
                    valid_wh = True
                    loc_whs_key = whs_auto_key
                    msg = "  Warehouse updated." 
                    break
                                
        if not valid_wh and not msg:
            msg = 'Warehouse not valid for the location.' 
    # user entered both loc and whs and as long as they correspond in the rel table in Quantum, we updated both
    elif whs_auto_key and loc_auto_key and not loc_whs_keys:
        valid_wh = True
        msg += "  Warehouse updated."     
    #user entered just a whs      
    elif whs_auto_key and existing_loc_key and not loc_auto_key and loc_whs_keys:
        #must make sure we can update to this warehouse.
        for loc_whs in loc_whs_keys:
            if loc_whs[0] == whs_auto_key:
                valid_wh = True
                loc_whs_key = whs_auto_key
                break                                        
    return msg,error,valid_wh,loc_whs_key,whs_needs_update

def update_location_from_rack(sysur_auto_key,user_id,rack_auto_key=None, mode=None,loc_auto_key=None, whs_auto_key=None, iq_enable=False,dj_user_id='',quapi=None):
    #update all locations for stock lines that are on the rack the users scanned in (Transer Cart mode)
    update_wh,loc_whs_key,valid_whs_key,whs_msg,msg,error = False,'',None,'','',''
    set_rack_null = iq_enable == False and ", IC_UDL_005 = NULL" or ''    
    #if the loc_whs_key doesn't match the actual warehouse key on stock line, then change the warehouse
    #on stock to match that of the location's warehouse    
    #or if the stock line has a warehouse that is not tied to the location, we must change it                       
    query = """SELECT S.STM_AUTO_KEY,S.QTY_OH,S.SYSCM_AUTO_KEY,S.PCC_AUTO_KEY, 
               S.CNC_AUTO_KEY,S.LOC_AUTO_KEY,S.WHS_AUTO_KEY,S.STC_AUTO_KEY,S.DPT_AUTO_KEY,
               SR.STR_AUTO_KEY,SR.QTY_RESERVED,SR.SOD_AUTO_KEY,SR.ROD_AUTO_KEY,SR.WOB_AUTO_KEY,
               SR.POD_AUTO_KEY,SR.WOO_AUTO_KEY
               FROM STOCK S 
               LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
               WHERE IC_UDL_005 = %s"""%rack_auto_key
    stock_recs = selection(query,quapi=quapi)    
    for strec in stock_recs:
        params = []
        params.append(strec[0])#stm_auto_key
        params.append(strec[1])#qty_oh
        params.append(strec[2])#syscm_auto_key
        params.append(strec[3])#pcc_auto_key                       
        params.append(strec[4])#cnc_auto_key
        params.append(strec[5])#loc_auto_key 
        params.append(strec[6])#whs_auto_key
        params.append(strec[7])#stc_auto_key
        params.append(strec[8])#dpt_auto_key
        params.append(strec[9])#str_auto_key  
        params.append(strec[10])#qty_reserved 
        params.append(strec[11])#sod_auto_key
        params.append(strec[12])#rod_auto_key
        params.append(strec[13])#wob_auto_key
        params.append(strec[14])#pod_auto_key
        params.append(strec[15])#woo_auto_key 
        if loc_auto_key: 
            #query = """UPDATE STOCK SET LOC_AUTO_KEY = %s%s WHERE IC_UDL_005 = %s AND (LOC_AUTO_KEY <> %s OR LOC_AUTO_KEY IS NULL) AND HISTORICAL_FLAG = 'F'"""%(loc_auto_key,set_rack_null,rack_auto_key,loc_auto_key)
            #msg += updation(query,quapi=quapi) 
            whs_msg,error,valid_wh,valid_whs_key,whs_needs_update = check_if_valid_whs_beta(rack_auto_key=rack_auto_key,loc_auto_key=loc_auto_key,whs_auto_key=whs_auto_key,dj_user_id=dj_user_id,quapi=quapi) 
            params[5] = loc_auto_key
            msg = qry_stock_transfer(sysur_auto_key,user_id,params,quapi) 
            if valid_whs_key:
                #query = """UPDATE STOCK SET WHS_AUTO_KEY = %s WHERE IC_UDL_005 = %s AND (WHS_AUTO_KEY <> %s OR WHS_AUTO_KEY IS NULL) AND HISTORICAL_FLAG = 'F'"""%(valid_whs_key,rack_auto_key,valid_whs_key)
                #msg += updation(query,quapi=quapi) 
                params[6] = valid_whs_key
                msg = qry_stock_transfer(sysur_auto_key,user_id,params,quapi)
                update_wh = not msg and True or False
            #if user didn't enter the warehouse, then we have to check every record to see if we need to 
            #update the warehouse to one that is valid for the location        
            else:
                #query = "SELECT WHS_AUTO_KEY,STM_AUTO_KEY FROM STOCK WHERE IC_UDL_005 = %s"%rack_auto_key
                #res = selection(query,quapi=quapi) 
                whs_key = strec[6]
                if whs_key:
                    whs_msg,error,valid_wh,valid_whs_key,whs_needs_update = check_if_valid_whs_beta(whs_key,dj_user_id=dj_user_id,loc_auto_key=loc_auto_key,quapi=quapi) or False 
                if not valid_wh and valid_whs_key:
                    params[6] = valid_whs_key
                    msg = qry_stock_transfer(sysur_auto_key,user_id,params,quapi)
                    update_wh = not msg and True or False
        elif not loc_auto_key and whs_auto_key and rack_auto_key:
            #check that the warehouse is a valid one for each stock line on the cart  
            loc_key = strec[5]
            if loc_key:
                whs_msg,error,valid_wh,valid_whs_key,whs_needs_update = check_if_valid_whs_beta(whs_auto_key=whs_auto_key,dj_user_id=dj_user_id,loc_auto_key=loc_key,quapi=quapi) or False 
                if not whs_msg and valid_wh and valid_whs_key:
                    params[6]=whs_auto_key
                    msg = qry_stock_transfer(sysur_auto_key,user_id,params,quapi)
                    #query = "UPDATE STOCK SET WHS_AUTO_KEY = %s WHERE STM_AUTO_KEY = %s"%(whs_auto_key,stm_key)
                    #msg += updation(query,quapi=quapi)
                    update_wh = not msg and True or False
                else:
                    msg += "Warehouse not updated."                    
    return msg + whs_msg,error,update_wh
    
def get_locs_from_wh(whs_auto_key, user_id='',quapi=None):
    query = "SELECT LOC_AUTO_KEY FROM WAREHOUSE_LOCATIONS WHERE WHS_AUTO_KEY=%s"%whs_auto_key
    res = selection(query, user_id=user_id, quapi=quapi)
    return res
    
def update_whs_from_rack(rack_auto_key, whs_auto_key, user_id='', quapi=None):
    #update all stock line's warehouses that are on the rack the users scanned in
    valid_loc_keys = get_locs_from_wh(whs_auto_key)
    loc_key_list = valid_loc_keys and valid_loc_keys[0] and valid_loc_keys[0][0] and '(' + str(valid_loc_keys[0][0]) or '('
    count = 1
    for loc in valid_loc_keys[1:]:
        loc_key_list += ',' + str(loc[0])
        if (count+1)%995 == 0 and loc_key_list[count+1]:
            loc_key_list += ') OR LOC_AUTO_KEY IN (' + str(loc_key_list[count+1])
        count += 1              
    loc_key_list += ')'
    query = """UPDATE STOCK SET WHS_AUTO_KEY = %s 
                    WHERE IC_UDL_005 = %s
                   AND (WHS_AUTO_KEY <> %s OR WHS_AUTO_KEY IS NULL)
                   AND HISTORICAL_FLAG = 'F'
                   AND LOC_AUTO_KEY IN %s"""%(whs_auto_key,rack_auto_key,whs_auto_key,loc_key_list)
    msg = updation(query, user_id=user_id, quapi=quapi)  
    return msg
    
def assign_new_wh(stm_auto_key, whs_auto_key, user_id='',quapi=None):
    query = "UPDATE STOCK SET WHS_AUTO_KEY = %s WHERE STM_AUTO_KEY = %s"%(whs_auto_key,stm_auto_key)
    msg = updation(query, user_id=user_id, quapi=quapi)
    return msg
        
def assign_rack_new_wh(rack_code, whs_auto_key, user_id='',quapi=None):
    query = "UPDATE STOCK SET WHS_AUTO_KEY = %s WHERE STM_AUTO_KEY IN (SELECT STM_AUTO_KEY FROM STOCK WHERE IC_UDL_005 = (SELECT UDL_AUTO_KEY FROM USER_DEFINED_LOOKUPS WHERE UDL_CODE = '%s'))"%(whs_auto_key,rack_code)
    msg = updation(query, user_id=user_id, quapi=quapi)
    return msg
    
def get_warehouse_location(location_code,user_id='',quapi=None):
    #from location code, we will look up and get the warehouse from the rel table
    query = "SELECT WHS_AUTO_KEY FROM WAREHOUSE_LOCATIONS WHERE LOC_AUTO_KEY = (SELECT LOC_AUTO_KEY FROM LOCATION WHERE LOCATION_CODE = %s)"%location_code
    res = selection(query, user_id=user_id, quapi=quapi)
    return res
    
def get_stock_recs(ctrl_number=None,ctrl_id=None,stm_auto_key=None,woo_auto_key=None,wo_number=None,user_id=None,quapi=None):
    and_where=''
    if woo_auto_key:
        and_where = "AND W.WOO_AUTO_KEY = %s"%woo_auto_key
    if stm_auto_key:
        and_where = "AND S.STM_AUTO_KEY = %s"%stm_auto_key      
    elif ctrl_id and ctrl_number:
        and_where = "AND S.CTRL_ID = %s AND S.CTRL_NUMBER = %s"%(ctrl_id,ctrl_number)
    elif wo_number:
        and_where = "AND W.SI_NUMBER = '%s'"%wo_number
    query = """
        SELECT DISTINCT W.SI_NUMBER,SR.WOO_AUTO_KEY,W.WOS_AUTO_KEY,S.STM_AUTO_KEY,
            SR.STR_AUTO_KEY,S.LOC_AUTO_KEY,S.WHS_AUTO_KEY 
            FROM STOCK S 
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY 
            LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY 
            LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY 
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WB.WOO_AUTO_KEY 
            LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = SR.ROD_AUTO_KEY 
            LEFT JOIN RO_HEADER ROH ON ROH.ROH_AUTO_KEY = ROD.ROH_AUTO_KEY 
        WHERE S.HISTORICAL_FLAG = 'F' AND S.QTY_OH > 0 %s"""%and_where
    res = selection(query,user_id=user_id,quapi=quapi)
    return res
def get_stms_data(stm_keys,cr):  
    res = []
    and_where,order_by='',''
        
    if stm_keys:
        stm_keys = construct_akl(stm_keys)
        for keys in stm_keys:
            and_where += "AND S.STM_AUTO_KEY in %s "%keys
            
    query = """
        SELECT DISTINCT W.SI_NUMBER, WS.DESCRIPTION, W.DUE_DATE, S.STOCK_LINE, P.PN, P.DESCRIPTION, 
          S.SERIAL_NUMBER,L.LOCATION_CODE,WS.SEVERITY,W.ENTRY_DATE,W.COMPANY_REF_NUMBER,
          W.WOO_AUTO_KEY,W.RANK,WS.WOS_AUTO_KEY,S.STM_AUTO_KEY,WH.WAREHOUSE_CODE,
          UDL.UDL_CODE,S.CTRL_ID,S.CTRL_NUMBER,S.IC_UDL_005,SR.STR_AUTO_KEY,S.LOC_AUTO_KEY,
          S.WHS_AUTO_KEY,C.COMPANY_NAME,WO.SI_NUMBER,WOR.SI_NUMBER,PCC.CONDITION_CODE,S.OWNER,
          CO.COMPANY_NAME,S.QTY_OH,S.SYSCM_AUTO_KEY,S.PCC_AUTO_KEY,
          S.CNC_AUTO_KEY,S.STC_AUTO_KEY,S.DPT_AUTO_KEY,SR.QTY_RESERVED,SR.SOD_AUTO_KEY,SR.ROD_AUTO_KEY,
          SR.WOB_AUTO_KEY,SR.POD_AUTO_KEY,P.SERIALIZED,S.REMARKS,CNC.CONSIGNMENT_CODE,UC.UOM_CODE
          FROM STOCK S
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
            LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY 
            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
            LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
            LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY = S.WHS_AUTO_KEY
            LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY 
            LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_005
            LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = W.CMP_AUTO_KEY
            LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
            LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = SR.ROD_AUTO_KEY
            LEFT JOIN WO_OPERATION WOR ON WOR.WOO_AUTO_KEY = ROD.WOO_AUTO_KEY
            LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
            LEFT JOIN COMPANIES CO ON CO.CMP_AUTO_KEY = S.CMP_AUTO_KEY
            LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
            LEFT JOIN UOM_CODES UC ON UC.UOM_AUTO_KEY = P.UOM_AUTO_KEY
        WHERE
          (SR.STR_AUTO_KEY IN (SELECT STR.STR_AUTO_KEY FROM STOCK_RESERVATIONS STR 
          WHERE STR.STM_AUTO_KEY=SR.STM_AUTO_KEY AND STR.QTY_RESERVED > 0 AND SR.QTY_INVOICED > 0 ) 
          OR SR.STR_AUTO_KEY IS NULL)
          AND S.HISTORICAL_FLAG = 'F'
          AND S.QTY_OH > 0 %s"""%and_where + order_by
    res = selection_dir(query,cr)
    
    if not res:
        query = """
        SELECT DISTINCT W.SI_NUMBER, WS.DESCRIPTION, W.DUE_DATE, S.STOCK_LINE, P.PN, P.DESCRIPTION, 
          S.SERIAL_NUMBER,L.LOCATION_CODE,WS.SEVERITY,W.ENTRY_DATE,W.COMPANY_REF_NUMBER,
          SR.WOO_AUTO_KEY,W.RANK,WS.WOS_AUTO_KEY,S.STM_AUTO_KEY,WH.WAREHOUSE_CODE,
          UDL.UDL_CODE,S.CTRL_ID,S.CTRL_NUMBER,S.IC_UDL_005,SR.STR_AUTO_KEY,S.LOC_AUTO_KEY,
          S.WHS_AUTO_KEY,C.COMPANY_NAME,WO.SI_NUMBER,WOR.SI_NUMBER,PCC.CONDITION_CODE,S.OWNER,
          CO.COMPANY_NAME,S.QTY_OH,S.SYSCM_AUTO_KEY,S.PCC_AUTO_KEY,
          S.CNC_AUTO_KEY,S.STC_AUTO_KEY,S.DPT_AUTO_KEY,SR.QTY_RESERVED,SR.SOD_AUTO_KEY,SR.ROD_AUTO_KEY,
          SR.WOB_AUTO_KEY,SR.POD_AUTO_KEY,P.SERIALIZED,S.REMARKS,CNC.CONSIGNMENT_CODE,UC.UOM_CODE
          FROM STOCK S
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
            LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY 
            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
            LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
            LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY = S.WHS_AUTO_KEY
            LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY 
            LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_005
            LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = W.CMP_AUTO_KEY
            LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
            LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = SR.ROD_AUTO_KEY
            LEFT JOIN WO_OPERATION WOR ON WOR.WOO_AUTO_KEY = ROD.WOO_AUTO_KEY
            LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
            LEFT JOIN COMPANIES CO ON CO.CMP_AUTO_KEY = S.CMP_AUTO_KEY
            LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
            LEFT JOIN UOM_CODES UC ON UC.UOM_AUTO_KEY = P.UOM_AUTO_KEY
            WHERE
              (SR.QTY_RESERVED > 0 AND SR.QTY_INVOICED > 0 OR SR.STR_AUTO_KEY IS NULL)
              AND S.HISTORICAL_FLAG = 'F'
              AND S.QTY_OH > 0 %s"""%and_where + order_by
        res = selection_dir(query,cr)

    stm_keys = []
    final_recs = []
    for rec in res:
        if rec[14] not in stm_keys:
            stm_keys.append(rec[14])
            final_recs.append(rec)
            
    return final_recs                     
 
def get_wos_from_rack_beta(quapi=None,cmp_auto_key=None,\
    syscm_auto_key=None,dpt_auto_key=None,consignment=None,\
    acc_co=None,inexact=False,is_rod=False,customer=None,\
    location=None,cond_code=None,wos_auto_key=None,\
    rack_auto_key=None,whs_auto_key=None,loc_auto_key=None,\
    woo_auto_key=None,stm_auto_key=None,ctrl_id=None,\
    ctrl_number=None,wo_number=None,user_id='',stm_keys=[],\
    cr=None):  
    res = []
    and_where,order_by='',''

    orcl_conn = None
    if not cr:
        from polls.models import OracleConnection as oc
        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
        orcl_conn = orcl_conn and orcl_conn[0] or None 
        if orcl_conn:
            cr,con = orcl_connect(orcl_conn)
        if not (cr and con):
            return 'Cannot connect to Oracle.' 
        
    if woo_auto_key:
        and_where = "AND W.WOO_AUTO_KEY = %s "%woo_auto_key
    elif stm_auto_key:
        and_where = "AND S.STM_AUTO_KEY = %s "%stm_auto_key      
    elif ctrl_id and ctrl_number:
        and_where = "AND S.CTRL_ID = %s AND S.CTRL_NUMBER = %s "%(ctrl_id,ctrl_number)
    elif wo_number and wo_number[0] not in ['c','C']:
        if not inexact:
            and_where = "AND W.SI_NUMBER = '%s' OR S.STM_AUTO_KEY = %s "%(wo_number,wo_number)
        else:
            and_where = "AND REGEXP_LIKE (W.SI_NUMBER, '%s', 'i') OR S.STM_AUTO_KEY = %s "%(wo_number,wo_number)
    
    elif wo_number and wo_number[0] in ['C','c']:
        and_where = "AND S.STM_AUTO_KEY = %s "%wo_number[1:]
    elif stm_keys:
        stm_keys = construct_akl(stm_keys)
        for keys in stm_keys:
            and_where += "AND S.STM_AUTO_KEY in %s "%keys
    else:
        if consignment:
            cons_where = "UPPER(CONSIGNMENT_CODE) = UPPER('%s')"%consignment
            and_where += "AND S.CNC_AUTO_KEY = (SELECT CNC_AUTO_KEY FROM CONSIGNMENT_CODES WHERE %s) "%cons_where
        if rack_auto_key:
            and_where += "AND S.IC_UDL_005 IS NOT NULL AND S.IC_UDL_005 = %s "%rack_auto_key
        if whs_auto_key:
            and_where += "AND S.WHS_AUTO_KEY IS NOT NULL AND S.WHS_AUTO_KEY = %s "%whs_auto_key
        if loc_auto_key:
            and_where += "AND S.LOC_AUTO_KEY IS NOT NULL AND S.LOC_AUTO_KEY = %s "%loc_auto_key 
        if cmp_auto_key:
            and_where += "AND S.CMP_AUTO_KEY IS NOT NULL AND S.CMP_AUTO_KEY = %s "%cmp_auto_key
        if syscm_auto_key:
            and_where += "AND S.SYSCM_AUTO_KEY IS NOT NULL AND S.SYSCM_AUTO_KEY = %s "%syscm_auto_key
        if dpt_auto_key:
            and_where += "AND S.DPT_AUTO_KEY IS NOT NULL AND S.DPT_AUTO_KEY = %s "%dpt_auto_key
        if cond_code:
            cond_code_where = "UPPER(CONDITION_CODE) =  UPPER('%s') "%cond_code
            and_where += "AND S.PCC_AUTO_KEY IN (SELECT PCC_AUTO_KEY FROM PART_CONDITION_CODES WHERE %s) "%cond_code_where         
        if customer:
            cust_where = "REGEXP_LIKE (COMPANY_NAME, '%s', 'i') "%customer
            and_where += "AND WO.CMP_AUTO_KEY IN (SELECT CMP_AUTO_KEY FROM COMPANIES WHERE %s) "%cust_where
        if location:
            loc_where = "UPPER(LOCATION_CODE) = UPPER('%s') "%location
            and_where += "AND S.LOC_AUTO_KEY IN (SELECT LOC_AUTO_KEY FROM LOCATION WHERE %s) "%loc_where
        if wos_auto_key:
            and_where += "AND WS.WOS_AUTO_KEY = %s "%wos_auto_key
                     
    query = """
        SELECT DISTINCT W.SI_NUMBER, WS.DESCRIPTION, W.DUE_DATE, S.STOCK_LINE, P.PN, P.DESCRIPTION, 
          S.SERIAL_NUMBER,L.LOCATION_CODE,WS.SEVERITY,W.ENTRY_DATE,W.COMPANY_REF_NUMBER,
          W.WOO_AUTO_KEY,W.RANK,WS.WOS_AUTO_KEY,S.STM_AUTO_KEY,WH.WAREHOUSE_CODE,
          UDL.UDL_CODE,S.CTRL_ID,S.CTRL_NUMBER,S.IC_UDL_005,SR.STR_AUTO_KEY,S.LOC_AUTO_KEY,
          S.WHS_AUTO_KEY,C.COMPANY_NAME,WO.SI_NUMBER,WOR.SI_NUMBER,PCC.CONDITION_CODE,S.OWNER,
          CO.COMPANY_NAME,S.QTY_OH,S.SYSCM_AUTO_KEY,S.PCC_AUTO_KEY,
          S.CNC_AUTO_KEY,S.STC_AUTO_KEY,S.DPT_AUTO_KEY,SR.QTY_RESERVED,SR.SOD_AUTO_KEY,SR.ROD_AUTO_KEY,
          SR.WOB_AUTO_KEY,SR.POD_AUTO_KEY,P.SERIALIZED,S.REMARKS,CNC.CONSIGNMENT_CODE,UC.UOM_CODE
          FROM STOCK S
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
            LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY 
            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
            LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
            LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY = S.WHS_AUTO_KEY
            LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY 
            LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_005
            LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = W.CMP_AUTO_KEY
            LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
            LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = SR.ROD_AUTO_KEY
            LEFT JOIN WO_OPERATION WOR ON WOR.WOO_AUTO_KEY = ROD.WOO_AUTO_KEY
            LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
            LEFT JOIN COMPANIES CO ON CO.CMP_AUTO_KEY = S.CMP_AUTO_KEY
            LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
            LEFT JOIN UOM_CODES UC ON UC.UOM_AUTO_KEY = P.UOM_AUTO_KEY
        WHERE
          (SR.STR_AUTO_KEY IN (SELECT STR.STR_AUTO_KEY FROM STOCK_RESERVATIONS STR 
          WHERE STR.STM_AUTO_KEY=SR.STM_AUTO_KEY) 
          OR SR.STR_AUTO_KEY IS NULL)
          AND S.HISTORICAL_FLAG = 'F'
          AND S.QTY_OH > 0 %s"""%and_where + order_by
    res = selection_dir(query,cr)

    if not res:
        query = """
        SELECT DISTINCT W.SI_NUMBER, WS.DESCRIPTION, W.DUE_DATE, S.STOCK_LINE, P.PN, P.DESCRIPTION, 
          S.SERIAL_NUMBER,L.LOCATION_CODE,WS.SEVERITY,W.ENTRY_DATE,W.COMPANY_REF_NUMBER,
          SR.WOO_AUTO_KEY,W.RANK,WS.WOS_AUTO_KEY,S.STM_AUTO_KEY,WH.WAREHOUSE_CODE,
          UDL.UDL_CODE,S.CTRL_ID,S.CTRL_NUMBER,S.IC_UDL_005,SR.STR_AUTO_KEY,S.LOC_AUTO_KEY,
          S.WHS_AUTO_KEY,C.COMPANY_NAME,WO.SI_NUMBER,WOR.SI_NUMBER,PCC.CONDITION_CODE,S.OWNER,
          CO.COMPANY_NAME,S.QTY_OH,S.SYSCM_AUTO_KEY,S.PCC_AUTO_KEY,
          S.CNC_AUTO_KEY,S.STC_AUTO_KEY,S.DPT_AUTO_KEY,SR.QTY_RESERVED,SR.SOD_AUTO_KEY,SR.ROD_AUTO_KEY,
          SR.WOB_AUTO_KEY,SR.POD_AUTO_KEY,P.SERIALIZED,S.REMARKS,CNC.CONSIGNMENT_CODE,UC.UOM_CODE
          FROM STOCK S
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
            LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY 
            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
            LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
            LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY = S.WHS_AUTO_KEY
            LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY 
            LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_005
            LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = W.CMP_AUTO_KEY
            LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
            LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = SR.ROD_AUTO_KEY
            LEFT JOIN WO_OPERATION WOR ON WOR.WOO_AUTO_KEY = ROD.WOO_AUTO_KEY
            LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
            LEFT JOIN COMPANIES CO ON CO.CMP_AUTO_KEY = S.CMP_AUTO_KEY
            LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
            LEFT JOIN UOM_CODES UC ON UC.UOM_AUTO_KEY = P.UOM_AUTO_KEY
            WHERE
              S.HISTORICAL_FLAG = 'F'
              AND S.QTY_OH > 0 %s"""%and_where + order_by
        res = selection_dir(query,cr)
        
    if not res and wo_number:
        and_where = "AND W.SI_NUMBER = '%s'"%wo_number
        query = """
        SELECT DISTINCT W.SI_NUMBER, WS.DESCRIPTION, W.DUE_DATE, S.STOCK_LINE, P.PN, P.DESCRIPTION, 
          S.SERIAL_NUMBER,L.LOCATION_CODE,WS.SEVERITY,W.ENTRY_DATE,W.COMPANY_REF_NUMBER,
          W.WOO_AUTO_KEY,W.RANK,WS.WOS_AUTO_KEY,S.STM_AUTO_KEY,WH.WAREHOUSE_CODE,
          UDL.UDL_CODE,S.CTRL_ID,S.CTRL_NUMBER,S.IC_UDL_005,SR.STR_AUTO_KEY,S.LOC_AUTO_KEY,
          S.WHS_AUTO_KEY,C.COMPANY_NAME,WO.SI_NUMBER,ROD.ROD_AUTO_KEY,PCC.CONDITION_CODE,S.OWNER,
          CO.COMPANY_NAME,S.QTY_OH,S.SYSCM_AUTO_KEY,S.PCC_AUTO_KEY,
          S.CNC_AUTO_KEY,S.STC_AUTO_KEY,S.DPT_AUTO_KEY,SR.QTY_RESERVED,SR.SOD_AUTO_KEY,SR.ROD_AUTO_KEY,
          SR.WOB_AUTO_KEY,SR.POD_AUTO_KEY,P.SERIALIZED,S.REMARKS,CNC.CONSIGNMENT_CODE,UC.UOM_CODE
          FROM WO_OPERATION W
            LEFT JOIN RO_DETAIL ROD ON ROD.WOO_AUTO_KEY = W.WOO_AUTO_KEY
            LEFT JOIN STOCK_RESERVATIONS SR ON SR.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY
            LEFT JOIN STOCK S ON S.STM_AUTO_KEY = SR.STM_AUTO_KEY
            LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
            LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
            LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY = S.WHS_AUTO_KEY
            LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY 
            LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_005
            LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = W.CMP_AUTO_KEY
            LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = SR.WOB_AUTO_KEY
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
            LEFT JOIN PART_CONDITION_CODES PCC ON PCC.PCC_AUTO_KEY = S.PCC_AUTO_KEY
            LEFT JOIN COMPANIES CO ON CO.CMP_AUTO_KEY = S.CMP_AUTO_KEY
            LEFT JOIN CONSIGNMENT_CODES CNC ON CNC.CNC_AUTO_KEY = S.CNC_AUTO_KEY
            LEFT JOIN UOM_CODES UC ON UC.UOM_AUTO_KEY = P.UOM_AUTO_KEY
        WHERE
          (SR.STR_AUTO_KEY IN (SELECT STR.STR_AUTO_KEY FROM STOCK_RESERVATIONS STR 
          WHERE STR.STM_AUTO_KEY=SR.STM_AUTO_KEY) 
          OR SR.STR_AUTO_KEY IS NULL)
          AND S.HISTORICAL_FLAG = 'F'
          AND S.QTY_OH > 0 %s"""%and_where + order_by
        res = selection_dir(query,cr) 
    
    stm_keys = []
    final_recs = []
    for rec in res:
        if rec[14] not in stm_keys:
            stm_keys.append(rec[14])
            final_recs.append(rec)
            
    return final_recs
    
"""App retrieves a list of stock lines with that UDL_AUTO_KEY in IC_UDL_005 of the stock table
    - User begins scanning SI_NUMBERS or CTRL # + CTRL ID application finds the STM_AUTO_KEY and this assigns or updates the stock lines’ “Rack” (IC_UDL_005). [method done]
    - Now all stock lines scanned are assign that “Rack” (UDL_AUTO_KEY in STOCK tables’ IC_UDL_005) [DEF]
    - Entering “RACK” and updating “LOCATION” should update LOC_AUTO_KEY on all STM_AUTO_KEYs assigned to that rack (unless that STM_AUTO_KEY is already assigned to that LOC_AUTO_KEY).[DEF]
"""

def create_rc_detail(entry_date,rch_auto_key,rcs_auto_key,pnm_auto_key,\
    order_type,order_num,cust_ref_number,loc_auto_key,whs_auto_key,\
    header_auto_key,priority,airway_bill,cr,tracking):
    error,msg,rc_number = '','',''
    #get the part data to create lines
    def_qty,quantity = 0,0 
    detail_rows,ro_rec,po_rec,so_rec = [],[],[],[] 
    
    """
     ALL RECEIPTS!!!	
        a.	If tracking number is found for ROD, POD, or SOD then only insert items that match 
            i.	(not all where qty_ordered < qty_received)
        b.	if tracking  number is found on ROH, POH, or SOH then we insert all ROD, POD, SOD where qty_ordered < qty_received
    """    
    if order_type == 'PO':
        #match the detail line tracking#
        tracking = """AND (UPPER(POD.AIRWAY_BILL) = UPPER('%s') 
            OR UPPER(POD.TRACKING_NUMBER) = UPPER('%s'))
            """%(airway_bill or order_num,airway_bill or order_num)
        from_clause =  """SELECT POD.ALT_PNM_AUTO_KEY,POD.POD_AUTO_KEY,
            POD.QTY_ORDERED,POD.QTY_REC FROM PO_DETAIL POD 
            WHERE POD.POH_AUTO_KEY = %s %s"""%(header_auto_key,tracking) 
        detail_rows = selection_dir(from_clause,cr)
        
        if not detail_rows:
            #match the header auto key or tracking
            from_clause =  """SELECT POD.ALT_PNM_AUTO_KEY,POD.POD_AUTO_KEY,
                POD.QTY_ORDERED,POD.QTY_REC FROM PO_DETAIL POD
                WHERE POD.QTY_ORDERED > POD.QTY_REC
                AND POD.POH_AUTO_KEY = '%s'"""%(header_auto_key)          
            detail_rows = selection_dir(from_clause,cr)  
            
    elif order_type == 'SO':

        #match the detail line tracking#
        tracking = "AND UPPER(E.CORE_TRACKING_NUMBER) = UPPER('%s')"%(airway_bill or order_num)       
        from_clause =  """SELECT SOD.PNM_AUTO_KEY,SOD.SOD_AUTO_KEY,POD.QTY_ORDERED,
            POD.QTY_REC,SOD.QTY_ORDERED,POD.QTY_REC FROM SO_DETAIL SOD
            LEFT JOIN SO_HEADER SOH ON SOH.SOH_AUTO_KEY = SOD.SOH_AUTO_KEY
            LEFT JOIN PURCHASE_SALES PS ON PS.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
            LEFT JOIN PO_DETAIL POD ON POD.POD_AUTO_KEY = PS.POD_AUTO_KEY
            LEFT JOIN EXCHANGE E ON E.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
            WHERE SOH.SOH_AUTO_KEY = '%s' %s"""%(header_auto_key,tracking)
        detail_rows = selection_dir(from_clause,cr)

        if not detail_rows:
            #match the header auto key or tracking    
            from_clause =  """SELECT SOD.PNM_AUTO_KEY,SOD.SOD_AUTO_KEY,POD.QTY_ORDERED,
                POD.QTY_REC,SOD.QTY_ORDERED,POD.QTY_REC FROM SO_DETAIL SOD
                LEFT JOIN SO_HEADER SOH ON SOH.SOH_AUTO_KEY = SOD.SOH_AUTO_KEY
                LEFT JOIN PURCHASE_SALES PS ON PS.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
                LEFT JOIN PO_DETAIL POD ON POD.POD_AUTO_KEY = PS.POD_AUTO_KEY
                WHERE SOH.SOH_AUTO_KEY = '%s'
                AND SOD.QTY_ORDERED > SOD.QTY_SHIP
                """%(header_auto_key)
            detail_rows = selection_dir(from_clause,cr) 
            
    elif order_type == 'RO':

        tracking = """AND (UPPER(ROD.MSG_AIRWAY_BILL) = UPPER('%s') 
            OR UPPER(ROD.TRACKING_NUMBER) = UPPER('%s'))"""%(airway_bill or order_num,airway_bill or order_num)
        from_clause =  """SELECT ROD.PNM_MODIFY,ROD.ROD_AUTO_KEY,POD.QTY_ORDERED,
                POD.QTY_REC,ROD.QTY_REPAIR,ROD.QTY_REPAIRED,ROD.QTY_RESERVED FROM RO_DETAIL ROD
                LEFT JOIN SO_DETAIL SOD ON SOD.SOD_AUTO_KEY = ROD.SOD_AUTO_KEY 
                LEFT JOIN PURCHASE_SALES PS ON PS.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
                LEFT JOIN PO_DETAIL POD ON POD.POD_AUTO_KEY = PS.POD_AUTO_KEY 
                WHERE ROD.ROH_AUTO_KEY ='%s' %s"""%(header_auto_key,tracking)
        detail_rows = selection_dir(from_clause,cr)

        if not detail_rows:
            #match the header auto key or tracking                
            from_clause =  """SELECT ROD.PNM_MODIFY,ROD.ROD_AUTO_KEY,POD.QTY_ORDERED,
                POD.QTY_REC,ROD.QTY_REPAIR,ROD.QTY_REPAIRED,ROD.QTY_RESERVED FROM RO_DETAIL ROD
                LEFT JOIN SO_DETAIL SOD ON SOD.SOD_AUTO_KEY = ROD.SOD_AUTO_KEY 
                LEFT JOIN PURCHASE_SALES PS ON PS.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
                LEFT JOIN PO_DETAIL POD ON POD.POD_AUTO_KEY = PS.POD_AUTO_KEY 
                WHERE ROD.ROH_AUTO_KEY ='%s'
                AND ROD.QTY_REPAIRED < ROD.QTY_REPAIR
                """%(header_auto_key)
            detail_rows = selection_dir(from_clause,cr)
                                         
   
    lines_created = 0
    for row in detail_rows:   
                                                                     
        if order_type != 'RO':
            q_ordered,q_received = row[2],row[3]
            if q_ordered == '' or q_ordered == 0:
                #need qty from RO or SO?
                q_ordered = row[4] and int(row[4]) or 0  
            if q_received == '':
                q_received = row[5] and int(row[5]) or 0
            quantity = q_ordered - q_received
        pod_auto_key,sod_auto_key,rod_auto_key = '','',''  
        
        if order_type == 'PO':
            pod_auto_key = row[1]
            where_clause = "RCD.POD_AUTO_KEY = %s"%pod_auto_key
            query = """
			SELECT POD.PCC_AUTO_KEY,RCC.LOC_AUTO_KEY,
			    RCC.WHS_AUTO_KEY,POD.CNC_AUTO_KEY,
				RCC.IFC_AUTO_KEY,RCC.STC_AUTO_KEY,
				POH.CMP_AUTO_KEY,POD.UNIT_COST,
				POD.QTY_ORDERED - POD.QTY_REC,
                POD.RECEIVER_INSTR,
                (select max(ctrl_number) from rc_detail) + 1,
                (select max(series_number) from rc_detail) + 1
				FROM PO_DETAIL POD
				JOIN PO_HEADER POH ON POH.POH_AUTO_KEY = POD.POH_AUTO_KEY
				JOIN RC_CONTROL RCC ON RCC.RCC_AUTO_KEY = 1
                WHERE POD.POD_AUTO_KEY = %s
            """%pod_auto_key
            recs = selection_dir(query,cr)
            if not recs:
                error += 'PO Detail line not found.'
            po_rec = recs and recs[0] or []
        elif order_type == 'SO':
            sod_auto_key = row[1]
            where_clause = "RCD.SOD_AUTO_KEY = %s"%sod_auto_key
            query = """
			SELECT SOD.PCC_AUTO_KEY,RCC.LOC_AUTO_KEY,
			    RCC.WHS_AUTO_KEY,SOD.CNC_AUTO_KEY,
				RCC.IFC_AUTO_KEY,RCC.STC_AUTO_KEY,
				SOH.CMP_AUTO_KEY,SOD.UNIT_COST,
				SOD.QTY_ORDERED - SOD.QTY_DELIVERED,
                SOD.RECEIVER_INSTR,
                (select max(ctrl_number) from rc_detail) + 1,
                (select max(series_number) from rc_detail) + 1
				FROM SO_DETAIL SOD
				JOIN SO_HEADER SOH ON SOH.SOH_AUTO_KEY = SOD.SOH_AUTO_KEY
				JOIN RC_CONTROL RCC ON RCC.RCC_AUTO_KEY = 1
                WHERE SOD.SOD_AUTO_KEY = %s
            """%sod_auto_key
            recs = selection_dir(query,cr)
            if not recs:
                error += 'SO Detail line not found. '
            so_rec = recs and recs[0] or []
        elif order_type == 'RO':
            rod_auto_key = row[1]
            query = """SELECT S.PCC_AUTO_KEY,S.WHS_AUTO_KEY,S.LOC_AUTO_KEY,
                ROD.RECEIVER_INSTR,S.CTS_AUTO_KEY,ROH.CMP_AUTO_KEY,S.CNC_AUTO_KEY,
                S.STC_AUTO_KEY,S.IFC_AUTO_KEY,STR.STR_AUTO_KEY,S.REMARKS,S.OWNER,
                S.IC_UDF_005,S.IC_UDF_006,S.IC_UDF_007,S.IC_UDF_008,S.IC_UDF_009,
                S.IC_UDF_010,S.UNIT_COST,
                (select max(series_number) from rc_detail) + 1,
                S.SERIES_ID,
                (select max(ctrl_number) from rc_detail) + 1,
                S.LOT_APL_RO_COST,S.LOT_ALW_PRECOST,S.ORIGINAL_PO_NUMBER,S.VISIBLE_MKT,
                POD.QTY_ORDERED
                FROM STOCK S
                LEFT JOIN STOCK_RESERVATIONS STR ON STR.STM_AUTO_KEY = S.STM_AUTO_KEY
                LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = STR.ROD_AUTO_KEY
                LEFT JOIN RO_HEADER ROH ON ROH.ROH_AUTO_KEY = ROD.ROH_AUTO_KEY
                LEFT JOIN SO_DETAIL SOD ON SOD.SOD_AUTO_KEY = ROD.SOD_AUTO_KEY 
                LEFT JOIN PURCHASE_SALES PS ON PS.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
                LEFT JOIN PO_DETAIL POD ON POD.POD_AUTO_KEY = PS.POD_AUTO_KEY
                WHERE ROD.ROD_AUTO_KEY = %s"""%rod_auto_key
            recs = selection_dir(query,cr)
            if not recs:
                error += 'RO Detail line not found. '
            ro_rec = recs and recs[0] or []
            where_clause = "RCD.ROD_AUTO_KEY = %s"%rod_auto_key

        query = """SELECT RCH.RC_NUMBER,L.LOCATION_CODE 
            FROM RC_HEADER RCH, RC_DETAIL RCD, LOCATION L
            WHERE %s AND RCD.RCH_AUTO_KEY = RCH.RCH_AUTO_KEY
            AND L.LOC_AUTO_KEY = RCD.LOC_AUTO_KEY AND ROWNUM<=1
            ORDER BY RCH.RCH_AUTO_KEY DESC"""%where_clause
        rc_data = selection_dir(query,cr)
        rc_number = rc_data and rc_data[0] and rc_data[0][0] or None
        location = rc_data and rc_data[0] and rc_data[0][1] or None
        if rc_number:
            msg += '%s line already on receiver: %s, at location: %s, priority: %s.  '\
            %(order_type,rc_number,location,priority)
            continue
        if error:
            return error,msg,rc_number
        lines_created += 1
        pnm_auto_key = row[0]
        #PN, PN2, SERIALIZED get from the related PNM_AUTO_KEY. then MFG_CODE, 
        #MFG_CODE2 related to the MFG_AUTO_KEY for that PNM. MANUFACTURER table

        qry = """SELECT P.PN,P.SERIALIZED,M.MFG_CODE FROM PARTS_MASTER P
            LEFT JOIN MANUFACTURER M ON M.MFG_AUTO_KEY = P.MFG_AUTO_KEY
            WHERE P.PNM_AUTO_KEY = '%s'"""%pnm_auto_key
        recs = selection_dir(qry,cr)
        part = recs and recs[0] or None
        pn = part and part[0] or None
        serialized = part and part[1] or None
        #serialized = 'T'
        mfg_code = part and part[2] or None


        """
            •	Status
            •	Disposition
            •	Condition
            •	Default Location
            •	Default Warehouse
            •	Default Consignment
            •	ILS Flag
            •	Default Stock Category
            •	Company Field
            •	Airway Bill
            •	Ctrl #
            •	Ctrl ID
            •	Series #
            •	Series ID
            •	Unit Cost
            •	Qty Original
        """

        if order_type == 'RO' and ro_rec: 
            pcc_auto_key = ro_rec[0] 
            whs_auto_key = ro_rec[1]
            loc_auto_key = ro_rec[2]
            receiver_instr = ro_rec[3]
            cts_auto_key = ro_rec[4]
            cmp_auto_key = ro_rec[5]
            cnc_auto_key = ro_rec[6]
            stc_auto_key = ro_rec[7]
            ifc_auto_key = ro_rec[8]
            str_auto_key = ro_rec[9]
            remarks = ro_rec[10]
            owner = ro_rec[11]
            ic_udf_005 = ro_rec[12]
            ic_udf_006 = ro_rec[13]
            ic_udf_007 = ro_rec[14]
            ic_udf_008 = ro_rec[15]
            ic_udf_009 = ro_rec[16]
            ic_udf_010 = ro_rec[17]
            unit_cost = ro_rec[18]
            series_number = ro_rec[19]
            series_id = ro_rec[20]
            ctrl_number = ro_rec[21]
            rec_type = 'ROR'
            lot_apl_ro_cost = ro_rec[22]
            lot_alw_precost = ro_rec[23]
            original_po_number = ro_rec[24]
            visible_mkt = ro_rec[25]
            qty_orig_po = ro_rec[26]  
            #insert the rc_detail line with the correct rch_auto_key,pnm_auto_key,quantity and mfg_code
            ctrl_id = lines_created
            rc_ctrl_number = ctrl_number
            rc_ctrl_id = ctrl_id             
            if serialized == 'T':
                rc_ctrl_number = ''
                rc_ctrl_id = ''                
            query = """INSERT INTO RC_DETAIL (  
                UNIT_PRICE,PCC_AUTO_KEY,WHS_AUTO_KEY,LOC_AUTO_KEY,
                RECEIVER_INSTR,ITEM_NUMBER,CTS_AUTO_KEY,CMP_AUTO_KEY,CNC_AUTO_KEY,
                STC_AUTO_KEY,IFC_AUTO_KEY,STR_AUTO_KEY,REMARKS,OWNER,IC_UDF_005,IC_UDF_006,IC_UDF_007,
                IC_UDF_008,IC_UDF_009,IC_UDF_010,UNIT_COST,SERIES_NUMBER,SERIES_ID,CTRL_NUMBER,CTRL_ID,
                ENTRY_DATE,REC_TYPE,LOT_APL_RO_COST,LOT_ALW_PRECOST,ORIGINAL_PO_NUMBER,VISIBLE_MKT,QTY_ORIG_PO,
                ALT_PNM_AUTO_KEY,RCD_AUTO_KEY,RCS_AUTO_KEY,RCH_AUTO_KEY,
                POD_AUTO_KEY,SOD_AUTO_KEY,ROD_AUTO_KEY,PNM_AUTO_KEY,PN,PN2,SERIALIZED,QTY_APPR,
                QTY_DENIED,QTY_HIST_DENIED,QTY_HIST_APPR,QTY,MFG_CODE,MFG2,AIRWAY_BILL) 
                VALUES(0,'%s','%s','%s','%s','%s','%s','%s','%s','%s','%s',
                '%s','%s','%s','%s','%s','%s','%s','%s','%s','%s',
                '%s','%s','%s','%s',TO_DATE('%s','mm-dd-yyyy'),'%s','%s','%s','%s','%s','%s',
                %s,G_RCD_AUTO_KEY.NEXTVAL,'%s','%s','%s','%s','%s','%s','%s','%s','%s','%s',
                0,0,0,'%s','%s','%s','%s')
                """%(pcc_auto_key,whs_auto_key,loc_auto_key,receiver_instr,lines_created,cts_auto_key,\
                cmp_auto_key,cnc_auto_key,stc_auto_key,ifc_auto_key,str_auto_key,remarks,\
                owner,ic_udf_005,ic_udf_006,ic_udf_007,ic_udf_008,ic_udf_009,ic_udf_010,\
                unit_cost,series_number,series_id,rc_ctrl_number,rc_ctrl_id,entry_date,rec_type,lot_apl_ro_cost,\
                lot_alw_precost,original_po_number,visible_mkt,qty_orig_po,pnm_auto_key,rcs_auto_key,\
                rch_auto_key,pod_auto_key,sod_auto_key,rod_auto_key,pnm_auto_key,pn,pn,serialized,quantity,\
                quantity,mfg_code,mfg_code,airway_bill)    
            error = insertion_dir(query,cr)
            if serialized == 'T':
                query = "SELECT MAX(RCD_AUTO_KEY) FROM RC_DETAIL"
                rcd = selection_dir(query,cr)
                          
                                    
                rcd_auto_key = rcd and rcd[0] and rcd[0][0] or ''
                query = "SELECT RCS_APPR,RDC_APPR FROM RC_CONTROL WHERE RCC_AUTO_KEY = 1"
                rcs = selection_dir(query,cr)
                rcs = rcs and rcs[0] or []
                rcs_auto_key = rcs and rcs[0]
                rdc_auto_key = rcs and rcs[1]
                
                query="""SELECT S.NOTES,S.STM_AUTO_KEY,S.SERIAL_NUMBER,
                        S.MFG_DATE,S.EXP_DATE,S.TAGGED_BY,S.TAG_DATE,S.TAG_NUMBER,
                        (SELECT MAX(CTRL_NUMBER) FROM RC_SERIAL) + 1
                        FROM STOCK S,STOCK_RESERVATIONS STR, 
                        RO_DETAIL ROD, RC_DETAIL RCD, RC_HEADER RCH             
                    WHERE 
                        S.STM_AUTO_KEY = STR.STM_AUTO_KEY
                        AND STR.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY
                        AND ROD.ROD_AUTO_KEY = RCD.ROD_AUTO_KEY
                        AND RCD.RCH_AUTO_KEY = RCH.RCH_AUTO_KEY
                        AND RCD.RCD_AUTO_KEY = %s
                    """%rcd_auto_key  
                stk_rec = selection_dir(query,cr)
                stk_rec = stk_rec and stk_rec[0] or []
                notes = stk_rec and stk_rec[0] or ''
                stm_original = stk_rec and stk_rec[1] or ''
                serial_number = stk_rec and stk_rec[2] or ''           
                mfg_date= stk_rec and stk_rec[3] or ''
                exp_date = stk_rec and stk_rec[4] or ''
                tagged_by = stk_rec and stk_rec[5] or ''
                tag_date = stk_rec and stk_rec[6] or ''
                tag_number = stk_rec and stk_rec[7] or ''
                ctrl_number = stk_rec and stk_rec[8] or ctrl_number
                query = """INSERT INTO RC_SERIAL (RCS_AUTO_KEY,RECEIVED,SIGN_OFF_FLAG,
                    RCL_AUTO_KEY,CTRL_NUMBER,CTRL_ID,NOTES,RCD_AUTO_KEY,
                    RDC_AUTO_KEY,STM_ORIGINAL,SERIAL_NUMBER,
                    SERIES_NUMBER,VISIBLE_MKT,MFG_DATE,EXP_DATE,SERIES_ID,
                    TAGGED_BY,TAG_DATE,TAG_NUMBER) 
                                                                                   
                    VALUES('%s','F','F',G_RCL_AUTO_KEY.NEXTVAL,
                                                            
                    '%s','%s',TO_CLOB('%s'),'%s','%s','%s','%s','%s','%s','%s','%s','%s',
                    '%s','%s','%s')
                                                                        
                    """%(rcs_auto_key,ctrl_number,ctrl_id,notes,rcd_auto_key,\
                    rdc_auto_key,stm_original,serial_number,series_number,\
                    visible_mkt,mfg_date,exp_date,series_id,tagged_by,tag_date,tag_number)
                                                                      
                                               
                error = insertion_dir(query,cr)
                    
        elif po_rec or so_rec:
            rec_type = ''
            if po_rec:
                rec_type = 'POP'
            pcc_auto_key = po_rec and po_rec[0] or so_rec and so_rec[0] or ''
            loc_auto_key = po_rec and po_rec[1] or so_rec and so_rec[1] or ''
            whs_auto_key = po_rec and po_rec[2] or so_rec and so_rec[2] or ''
            cnc_auto_key = po_rec and po_rec[3] or so_rec and so_rec[3] or ''
            ifc_auto_key = po_rec and po_rec[4] or so_rec and so_rec[4] or ''
            stc_auto_key = po_rec and po_rec[5] or so_rec and so_rec[5] or ''
            cmp_auto_key = po_rec and po_rec[6] or so_rec and so_rec[6] or ''
            unit_cost = po_rec and po_rec[7] or so_rec and so_rec[7] or ''
            qty_orig = po_rec and po_rec[8] or so_rec and so_rec[8] or ''
            receiver_instr = po_rec and po_rec[9] or so_rec and so_rec[9] or '' 
            ctrl_number = po_rec and po_rec[10] or so_rec and so_rec[10] or ''
            series_number = po_rec and po_rec[11] or so_rec and so_rec[11] or '' 
            ctrl_id = lines_created
            rc_ctrl_number = ctrl_number
            rc_ctrl_id = ctrl_id             
            if serialized == 'T':
                rc_ctrl_number = ''
                rc_ctrl_id = ''              
            query = """INSERT INTO RC_DETAIL (SERIES_NUMBER,CTRL_NUMBER,CTRL_ID,REC_TYPE,SERIES_ID,UNIT_PRICE,ITEM_NUMBER,ALT_PNM_AUTO_KEY,ENTRY_DATE,RCD_AUTO_KEY,RCS_AUTO_KEY,RCH_AUTO_KEY,
                POD_AUTO_KEY,SOD_AUTO_KEY,ROD_AUTO_KEY,PNM_AUTO_KEY,
                PN,PN2,SERIALIZED,QTY_APPR,QTY_DENIED,QTY_HIST_DENIED,
                QTY_HIST_APPR,QTY,MFG_CODE,MFG2,PCC_AUTO_KEY,LOC_AUTO_KEY,
                WHS_AUTO_KEY,CNC_AUTO_KEY,IFC_AUTO_KEY,STC_AUTO_KEY,
                CMP_AUTO_KEY,UNIT_COST,QTY_ORIG_PO,RECEIVER_INSTR,AIRWAY_BILL) 
                VALUES('%s','%s','%s','%s',0,0,%s,'%s',TO_DATE('%s','mm-dd-yyyy'),G_RCD_AUTO_KEY.NEXTVAL,
                '%s','%s','%s','%s','%s','%s','%s','%s',
                '%s','%s',0,0,0,'%s','%s','%s','%s','%s','%s','%s','%s',
                '%s','%s','%s','%s','%s','%s')"""%(series_number,rc_ctrl_number,rc_ctrl_id,rec_type,\
                lines_created,pnm_auto_key,entry_date,rcs_auto_key,\
                rch_auto_key,pod_auto_key,sod_auto_key,rod_auto_key,pnm_auto_key,\
                pn,pn,serialized,quantity,quantity,mfg_code,mfg_code,\
                pcc_auto_key,loc_auto_key,whs_auto_key,cnc_auto_key,\
                ifc_auto_key,stc_auto_key,cmp_auto_key,unit_cost,qty_orig,\
                receiver_instr,airway_bill)
            error = insertion_dir(query,cr)
            query = "SELECT MAX(RCD_AUTO_KEY) FROM RC_DETAIL"
            rcd = selection_dir(query,cr)
            rcd_auto_key = rcd and rcd[0] and rcd[0][0] or ''
            if serialized == 'T':
                """
                1.	Insert a new RCS_AUTO_KEY into RC_SERIAL to match 
                    total quantity pending receipt (RC_DETAIL[‘QTY’]
                    a.	If RC_DETAIL[‘QTY’] = 3 then there should be 
                        3 new RCS_AUTO_KEYs inserted to RC_SERIAL 
                        with all the same RCD_AUTO_KEY
                2.	There is no STM_ORIGINAL to pull expected data 
                    from so all is null except for RC_STATUS and 
                    RC_DISP_CODES which will both come from RC_CONTROL
                """
                notes = 'RC Detail line created from MRO Live Dock Receiving'
                count = 0

                while count < quantity:
                    query="""SELECT MAX(CTRL_NUMBER) FROM RC_SERIAL
                        """
                    rec = selection_dir(query,cr)
                    ctrl_number = rec and rec[0] and rec[0][0] +1 or ctrl_number
                    count += 1
                    query = """INSERT INTO RC_SERIAL 
                        (RCS_AUTO_KEY,CTRL_NUMBER,CTRL_ID,RECEIVED,SIGN_OFF_FLAG,
                        RCL_AUTO_KEY,RDC_AUTO_KEY,RCD_AUTO_KEY) 
                        VALUES(1,'%s','%s','F','F',G_RCL_AUTO_KEY.NEXTVAL,1,'%s')
                        """%(ctrl_number,count,rcd_auto_key)
                    error = insertion_dir(query,cr)               
    if not lines_created:
        error += ' No receiver lines created.'
    else:
        msg += ' %s receiver lines were created.'%lines_created
    if not detail_rows:
        error += ' No order lines found.'
    return error,msg,rc_number

@shared_task
def create_rc(quapi_id,session_id,airway_bill,order_num,order_type,\
    arrival_date,location,syscm_auto_key,cmp_auto_key,\
    dpt_auto_key,next_num,cust_ref_number,pnm_auto_key,\
    priority,soh_auto_key,roh_auto_key,poh_auto_key,\
    sysur_auto_key,tracking,sysnl_auto_key):
        
    msg,order_type2,existing_rc,sysnl,sysnl_log = '','','',[],[]
    location_code = ''
    from polls.models import QueryApi,OracleConnection as oc,MLApps as maps,QuantumUser as qu
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg,location_code,existing_rc
        
    #today = datetime.now()
    #timestamp = today.strftime('%m/%d/%Y %H:%M:%S')
    #today = today.strftime('%m/%d/%Y')
    
    query = "SELECT SYSTIMESTAMP FROM DUAL"
    today = selection_dir(query,cr)
    today = today and today[0] and today[0][0] 
    timestamp = today and today[:19]
    today = today and today[:10]

    #date_format = '%Y-%m-%d %H:%M:%S'
    #timestamp = timestamp and datetime.strptime(timestamp,date_format)
    #date_format = '%Y-%m-%d'
    #today = today and datetime.strptime(today,date_format)

    if not today and timestamp:
        return 'Problem with database time.',msg,location_code,existing_rc                                                                
        
    #get the loc_auto_key from RC_CONTROL
    loc_auto_key = ''
    query = """SELECT RCC.LOC_AUTO_KEY,RCC.WHS_AUTO_KEY,
        RCC.RCS_AUTO_KEY,L.LOCATION_CODE FROM RC_CONTROL RCC
        LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = RCC.LOC_AUTO_KEY
        """
    loc = selection_dir(query,cr)
    loc_auto_key = loc[0] and loc[0][0] or ''
    whs_auto_key = loc[0] and loc[0][1] or ''
    rcs_auto_key = loc[0] and loc[0][2] or ''
    location_code = loc[0] and loc[0][3] or ''        
    if order_type == 'PO':
        order_type2 = 'SO'
    elif order_type == 'RO':
        order_type2 = 'WO'
    elif order_type == 'SO':
        order_type2 = 'PO'
        
 
    query = """INSERT INTO RC_HEADER 
        (SYSUR_AUTO_KEY,COMPANY_NAME,SOH_AUTO_KEY,ROH_AUTO_KEY,POH_AUTO_KEY,AIRWAY_BILL,OPEN_FLAG,RC_NUMBER,ORDER_NUMBER1,ORDER_NUMBER2,ORDER_TYPE1,ORDER_TYPE2,
        SYSCM_AUTO_KEY,CMP_AUTO_KEY,DPT_AUTO_KEY,ENTRY_DATE,ARRIVAL_DATE,LOC_AUTO_KEY) 
        VALUES('%s',(SELECT COMPANY_NAME FROM COMPANIES WHERE CMP_AUTO_KEY = %s),'%s','%s','%s','%s','T','%s','%s','%s','%s','%s','%s','%s','%s',TO_TIMESTAMP('%s','yyyy-mm-dd hh24:mi:ss'),TO_DATE('%s','yyyy-mm-dd'),'%s')
        """%(sysur_auto_key,cmp_auto_key,soh_auto_key,roh_auto_key,poh_auto_key,airway_bill,next_num,order_num,cust_ref_number,order_type,order_type2,syscm_auto_key or '',cmp_auto_key or '',dpt_auto_key or '',timestamp,today,loc_auto_key)
    error = insertion_dir(query,cr)
           
    query = """SELECT RCH_AUTO_KEY FROM RC_HEADER 
        WHERE
        ARRIVAL_DATE = TO_DATE('%s','yyyy-mm-dd')
        AND RC_NUMBER = '%s'
        ORDER BY RCH_AUTO_KEY DESC
    """%(today,next_num)
    rch = selection_dir(query,cr)
    rch_auto_key = rch and rch[0] and rch[0][0] or None

    if rch_auto_key:
        query = """UPDATE RC_HEADER 
            SET SYSUR_AUTO_KEY = %s 
            WHERE RCH_AUTO_KEY = %s
        """%(sysur_auto_key,rch_auto_key)
        
        error = updation_dir(query,cr)
    
    if error in ['{"recs": ""}','']:
        
        if rch:
            msg = 'Successful creation of RC#: ' + str(next_num) + '.'
   
            if not sysnl_auto_key:
                
                query = """SELECT SYSNL_AUTO_KEY FROM SYS_NUMBER_LOG SN
                      LEFT JOIN SYS_NUMBER_LOG_CODES SNC 
                      ON SNC.SYSNLC_AUTO_KEY = SN.SYSNLC_AUTO_KEY
                      WHERE SNC.LOG_TYPE_CODE = 'RC'"""
                sysnl = selection_dir(query,cr)
                sysnl_auto_key = sysnl and sysnl[0] and sysnl[0][0] or None
                import re 
                next_num = re.sub("[^0-9]", "", next_num)
                if sysnl_auto_key and next_num:
                    query ="""UPDATE sys_number_log snl 
                      SET snl.last_number = '%s'
                      WHERE snl.sysnl_auto_key = %s"""%(next_num,sysnl_auto_key)               
                    error = updation_dir(query,cr)                                       
        else:
            error = 'There was a problem so no receiver was created.'
    if error == '{"recs": ""}' or error == '':
        orcl_commit(con=con)
        error = ''        
    #register audit trail record                
    app_id = maps.objects.filter(code='dock-receiving')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        field_changed = 'Successful creation of RC#: ' + str(next_num) + '.'
        if not error:
            new_val = 'Successful creation of RC#: ' + str(next_num) + '.'
            aud_status = 'success'
        else:             
            aud_status = 'failure'
            new_val = 'Failed to create RC#: ' + str(next_num) + '.'
            field_changed = 'Nothing changed.'
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')  
        error = register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'        
    return error,msg,location_code,existing_rc
       
@shared_task
def update_sysnl(quapi_id,session_id,sysnl_auto_key,next_num):
    
    error,msg = '',''
    from polls.models import QueryApi,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    import re    
    next_num = re.sub("[^0-9]", "", next_num)
        
    query ="""UPDATE sys_number_log 
        SET last_number = '%s'
        WHERE sysnl_auto_key = %s"""%(next_num,sysnl_auto_key)
            
    error = updation_dir(query,cr) 
    
    if error == '{"recs": ""}' or error == '':
        orcl_commit(con=con)
        error = ''

    return error,msg                        
    
@shared_task
def get_receivers(quapi_id,session_id,closed=False,from_date=None,to_date=None):
    error,msg = '',''
    from polls.models import QueryApi,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
        
    open_flag = "RCH.OPEN_FLAG = 'T'"
    if closed:
        open_flag = "RCH.OPEN_FLAG = 'F'"
    
    date_from = ''    
    if from_date:
        date_from = " AND RCH.CLOSED_DATE >= TO_DATE('%s','MM-DD-YYYY')"%from_date
        
    date_to = ''  
    if to_date:
        date_to = " AND RCH.CLOSED_DATE <= TO_DATE('%s','MM-DD-YYYY')"%to_date

    query = """SELECT
            RCH.RCH_AUTO_KEY,
            CASE WHEN SOH.SOH_AUTO_KEY IS NOT NULL THEN SOH.SO_NUMBER ELSE
            (CASE WHEN W.WOO_AUTO_KEY IS NOT NULL THEN W.SI_NUMBER ELSE
            (CASE WHEN WR.WOO_AUTO_KEY IS NOT NULL THEN WR.SI_NUMBER ELSE WP.SI_NUMBER END) END) END,
            CASE WHEN PW.NEED_DATE IS NOT NULL THEN PW.NEED_DATE ELSE
            (CASE WHEN WOB.NEED_DATE IS NOT NULL THEN WOB.NEED_DATE ELSE
            (CASE WHEN W.DUE_DATE IS NOT NULL THEN W.DUE_DATE ELSE
            (CASE WHEN WR.DUE_DATE IS NOT NULL THEN WR.DUE_DATE ELSE
            (CASE WHEN WP.DUE_DATE IS NOT NULL THEN WP.DUE_DATE ELSE SOH.DUE_DATE 
            END) END) END) END) END,
            RCH.RC_NUMBER,
            RCH.ARRIVAL_DATE,
            L.LOCATION_CODE,
            RCD.QTY,
            0,'',
            CASE WHEN CS.CMP_AUTO_KEY IS NOT NULL THEN CS.COMPANY_NAME ELSE (
            CASE WHEN CE.CMP_AUTO_KEY IS NOT NULL THEN CE.COMPANY_NAME ELSE (
            CASE WHEN CP.CMP_AUTO_KEY IS NOT NULL THEN CP.COMPANY_NAME ELSE
            (CASE WHEN CH.CMP_AUTO_KEY IS NOT NULL THEN CH.COMPANY_NAME END)
            END) END) END,
            RCH.ENTRY_DATE,
            --CASE WHEN POH.PO_NUMBER IS NOT NULL THEN POH.PO_NUMBER ELSE
            --    (CASE WHEN ROH.RO_NUMBER IS NOT NULL THEN ROH.RO_NUMBER ELSE
            --    (CASE WHEN SOH.SO_NUMBER IS NOT NULL THEN SOH.SO_NUMBER ELSE
            --    (CASE WHEN SHE.SO_NUMBER IS NOT NULL THEN SHE.SO_NUMBER ELSE
            --      SH.SO_NUMBER END) END) END) END,
            RCH.ORDER_NUMBER1,
            CASE WHEN SVP.SHIP_VIA_CODE IS NOT NULL THEN SVP.SHIP_VIA_CODE ELSE
            (CASE WHEN SVR.SHIP_VIA_CODE IS NOT NULL THEN SVR.SHIP_VIA_CODE ELSE
            SVS.SHIP_VIA_CODE END) END,
            RCH.COMPANY_NAME,
            SOR.FIRST_NAME || ' ' || SOR.LAST_NAME,
            SRR.FIRST_NAME || ' ' || SRR.LAST_NAME,
            RCH.AIRWAY_BILL,
            TO_CHAR(RCH.NOTES),
            RCH.CLOSED_DATE
            FROM RC_HEADER RCH
            LEFT JOIN RC_DETAIL RCD ON RCD.RCH_AUTO_KEY = RCH.RCH_AUTO_KEY
            LEFT JOIN SO_DETAIL SOD ON SOD.SOD_AUTO_KEY = RCD.SOD_AUTO_KEY
            LEFT JOIN SO_HEADER SOH ON SOH.SOH_AUTO_KEY = SOD.SOH_AUTO_KEY
            LEFT JOIN COMPANIES CS ON CS.CMP_AUTO_KEY = SOH.CMP_AUTO_KEY
            LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = RCD.ROD_AUTO_KEY
            LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = ROD.WOB_AUTO_KEY
            LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = WOB.WOO_AUTO_KEY
            LEFT JOIN WO_OPERATION WR ON WR.WOO_AUTO_KEY = ROD.WOO_AUTO_KEY
            LEFT JOIN COMPANIES C ON C.CMP_AUTO_KEY = W.CMP_AUTO_KEY
            LEFT JOIN RO_HEADER ROH ON ROH.ROH_AUTO_KEY = ROD.ROH_AUTO_KEY
            LEFT JOIN PO_DETAIL POD ON POD.POD_AUTO_KEY = RCD.POD_AUTO_KEY
            LEFT JOIN PO_HEADER POH ON POH.POH_AUTO_KEY = POD.POH_AUTO_KEY
            LEFT JOIN PURCHASE_SALES PS ON PS.POD_AUTO_KEY = POD.POD_AUTO_KEY
            LEFT JOIN PURCHASE_WO PWO ON PWO.POD_AUTO_KEY = POD.POD_AUTO_KEY
            LEFT JOIN WO_BOM PW ON PW.WOB_AUTO_KEY = PWO.WOB_AUTO_KEY
            LEFT JOIN WO_OPERATION WP ON WP.WOO_AUTO_KEY = PW.WOO_AUTO_KEY
            LEFT JOIN COMPANIES CP ON CP.CMP_AUTO_KEY = WP.CMP_AUTO_KEY
            LEFT JOIN SO_DETAIL SDE ON SDE.SOD_AUTO_KEY = PS.SOD_AUTO_KEY
            LEFT JOIN SO_HEADER SHE ON SHE.SOH_AUTO_KEY = SDE.SOH_AUTO_KEY
            LEFT JOIN COMPANIES CE ON CE.CMP_AUTO_KEY = SHE.CMP_AUTO_KEY
            LEFT JOIN SO_DETAIL SD ON SD.SOD_AUTO_KEY = ROD.SOD_AUTO_KEY  
            LEFT JOIN SO_HEADER SH ON SH.SOH_AUTO_KEY = SD.SOH_AUTO_KEY
            LEFT JOIN COMPANIES CH ON CH.CMP_AUTO_KEY = SH.CMP_AUTO_KEY
            LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = RCH.LOC_AUTO_KEY
            LEFT JOIN PO_HEADER PO ON PO.POH_AUTO_KEY = RCH.POH_AUTO_KEY
            LEFT JOIN RO_HEADER RO ON RO.ROH_AUTO_KEY = RCH.ROH_AUTO_KEY
            LEFT JOIN SO_HEADER SO ON SO.SOH_AUTO_KEY = RCH.SOH_AUTO_KEY
            LEFT JOIN SYS_USERS SOR ON SOR.SYSUR_AUTO_KEY = SHE.SYSUR_AUTO_KEY
            LEFT JOIN SYS_USERS SRR ON SRR.SYSUR_AUTO_KEY = RCH.SYSUR_AUTO_KEY
            LEFT JOIN SHIP_VIA_CODES SVH ON SVH.SVC_AUTO_KEY = POH.SVC_AUTO_KEY
            LEFT JOIN SHIP_VIA_CODES SVP ON SVP.SVC_AUTO_KEY = PO.SVC_AUTO_KEY
            LEFT JOIN SHIP_VIA_CODES SVR ON SVR.SVC_AUTO_KEY = RO.SVC_AUTO_KEY
            LEFT JOIN SHIP_VIA_CODES SVS ON SVS.SVC_AUTO_KEY = SO.SVC_AUTO_KEY            
            WHERE %s%s%s                    
            ORDER BY WOB.NEED_DATE NULLS LAST,PW.NEED_DATE NULLS LAST,
            PW.NEED_DATE NULLS LAST,
            SD.DUE_DATE NULLS LAST,SDE.DUE_DATE NULLS LAST,
            SOD.DUE_DATE NULLS LAST,CS.COMPANY_NAME NULLS LAST,
            CP.COMPANY_NAME NULLS LAST,CE.COMPANY_NAME NULLS LAST,
            CH.COMPANY_NAME NULLS LAST,SHE.SO_NUMBER NULLS LAST,
            SH.SO_NUMBER NULLS LAST,SOH.SO_NUMBER NULLS LAST
    """%(open_flag,date_from,date_to)
    #now create each receiver locally, summing up the qty's
    
    receivers = selection_dir(query,cr)
    rc_data = []
    rc_numbers = []
    date_format = '%Y-%m-%d %H:%M:%S' 
    due_date = ''
    priority = ''
    
    for rec in receivers:
        if rc_data:
            if len(rc_data[-1]) > 8:
                rc_data[-1][8] = priority
            else:
                rc_data[-1].append(priority)
            if due_date:
                rc_data[-1][2] = due_date
        if rec[0] not in rc_numbers:
            #initialize variables to store date min data
            due_days_list = []
            min_due_days = 0
            priority = 4
            due_date = ''
            #start the count for quantity sum of lines
            rec.append(0)                 
            rc_data.append(rec)
            rc_numbers.append(rec[0])    
        if rec[2] != '':                   
            due_days_list.append(datetime.strptime(rec[2],date_format) - datetime.now())             
            min_due_days = min(due_days_list) 
            if min_due_days:
                min_due_days = min_due_days.days          
                calc_priority = (min_due_days < 4 and 1) or (min_due_days < 11 and 2) or (min_due_days >= 10 and 3)                
                priority = min_due_days and calc_priority or ''
                due_date = rec[2]
        #counting quantities for each line
        if rec[6] != '' and rec[6] != 0:
            rc_data[-1][7] += 1        
    if rc_data:
        from polls.models import WOStatus
        todel = WOStatus.objects.filter(session_id=session_id).delete()       
        error = create_rc_headers(rc_data,session_id)          
        
    return error,msg
    
def create_rc_headers(rc_data,session_id):
    from polls.models import WOStatus
    rc_objects,error,msg = [],'',''
    """supplier order#,
                CASE WHEN SVP.SHIP_VIA_CODE IS NOT NULL THEN SVP.SHIP_VIA_CODE ELSE
            (CASE WHEN SVR.SHIP_VIA_CODE IS NOT NULL THEN SVR.SHIP_VIA_CODE ELSE
            SVS.SHIP_VIA_CODE END) END,
            RCH.COMPANY_NAME,
            SOR.FIRST_NAME || ' ' || SOR.LAST_NAME as Created By,
            SRR.FIRST_NAME || ' ' || SRR.LAST_NAME as Received By,
            RCH.AIRWAY_BILL,
            TO_CHAR(RCH.NOTE)
    """    
    for rec in rc_data:   
            
        rc_objects.append(WOStatus( 
        location_code=rec[5],
        wo_number=rec[1],
        arrival_date= rec[4] and rec[4][:10] or None,
        session_id = session_id,
        next_num = rec[3],
        due_date = rec[2] and rec[2][:10] or None,
        gate_qty = rec[7] or 0,  
        priority = len(rec)>8 and rec[8] or 4,
        customer = len(rec)>9 and rec[9] or '',
        loc_validated_date = rec[10] or None,#rch.entry_date
        si_number = rec[11],#supplier order#
        supdate_msg = rec[12],#ship_via
        account_company = rec[13],#rch.company     
        stock_owner = rec[14],#sale user - first name/last name
        user_id = rec[15],#received by user name
        airway_bill = rec[16],#airway_bill
        notes = rec[17],#notes
        closed_date = rec[18] or None,  #rch.closed_date     
        ))
    
    try:
        WOStatus.objects.bulk_create(rc_objects)
    except Exception as exc:
        error = exc.args       
    return error
    
@shared_task
def get_soropos(quapi_id,label,session_id,next_num=0):
    error,msg,roh_auto_key,poh_auto_key='','','',''
    from polls.models import QueryApi,WOStatus,OracleConnection as oc
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle',msg
    soropos = query_soropos(cr,session_id,label,next_num=next_num)
    rch_auto_key = soropos and soropos[0] and soropos[0][0] or None
    soh_auto_key = soropos and soropos[0] and soropos[0][11] or None
    detail_type = soropos and soropos[0] and soropos[0][1] or None
    if detail_type == 'RO':
        roh_auto_key = soropos and soropos[0] and soropos[0][8] or None
    elif detail_type == 'PO':
        poh_auto_key = soropos and soropos[0] and soropos[0][8] or None
    if soropos and not error:
        priority,error = assign_priority(cr,soh_auto_key,roh_auto_key,poh_auto_key,detail_type,soropos)
        todel = WOStatus.objects.filter(session_id=session_id).delete() 
        create_rcs_bulk(soropos,priority,session_id,cr)        
    else:
        error = 'No records found.'       
    return error,msg
     
def get_due_dates(cr,soh_auto_key,roh_auto_key,poh_auto_key,detail_type): 
    if detail_type == 'SO':
        clause = """AND (SCD.SOH_AUTO_KEY = %s
            OR SCDO.SOH_AUTO_KEY = %s
            OR SDO.SOH_AUTO_KEY = %s
            OR SOD.SOH_AUTO_KEY = %s 
            OR SDE.SOH_AUTO_KEY = %s 
            OR SD.SOH_AUTO_KEY = %s)
            """%(soh_auto_key,soh_auto_key,soh_auto_key,soh_auto_key,soh_auto_key,soh_auto_key)
    elif detail_type == 'PO': 
        clause = """AND (POD.POH_AUTO_KEY = %s 
        OR PO.POH_AUTO_KEY = %s 
        OR PCH.POH_AUTO_KEY = %s)
        """%(poh_auto_key,poh_auto_key,poh_auto_key)
    else:
        clause = """AND (ROD.ROH_AUTO_KEY = %s 
            OR RCDO.ROH_AUTO_KEY = %s)"""%(roh_auto_key,roh_auto_key)
    date_query = """
        SELECT  
            CASE WHEN WO.DUE_DATE IS NOT NULL THEN WO.DUE_DATE ELSE 
            (CASE WHEN WB.NEED_DATE IS NOT NULL THEN WB.NEED_DATE ELSE
            (CASE WHEN WNB.DUE_DATE IS NOT NULL THEN WNB.DUE_DATE ELSE
            (CASE WHEN WB.NEED_DATE IS NOT NULL THEN WB.NEED_DATE ELSE
            (CASE WHEN PWOB.NEED_DATE IS NOT NULL THEN PWOB.NEED_DATE ELSE
            (CASE WHEN SCDO.DUE_DATE IS NOT NULL THEN SCDO.DUE_DATE ELSE
            (CASE WHEN SCD.DUE_DATE IS NOT NULL THEN SCD.DUE_DATE ELSE
            (CASE WHEN SDO.DUE_DATE IS NOT NULL THEN SDO.DUE_DATE ELSE
            SD.DUE_DATE END) END) END) END) END) END) END) END
            FROM RC_HEADER RCH 
            LEFT JOIN RC_DETAIL RCD ON RCD.RCH_AUTO_KEY = RCH.RCH_AUTO_KEY
            LEFT JOIN SO_DETAIL SOD ON SOD.SOD_AUTO_KEY = RCD.SOD_AUTO_KEY
            LEFT JOIN SO_HEADER SOH ON SOH.SOH_AUTO_KEY = SOD.SOH_AUTO_KEY
            LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = RCD.ROD_AUTO_KEY
            LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = ROD.WOB_AUTO_KEY
            LEFT JOIN RO_HEADER ROH ON ROH.ROH_AUTO_KEY = ROD.ROH_AUTO_KEY
            LEFT JOIN PO_DETAIL POD ON POD.POD_AUTO_KEY = RCD.POD_AUTO_KEY
            LEFT JOIN PO_HEADER POH ON POH.POH_AUTO_KEY = POD.POH_AUTO_KEY
            LEFT JOIN PURCHASE_SALES PS ON PS.POD_AUTO_KEY = POD.POD_AUTO_KEY
            LEFT JOIN PURCHASE_WO PWO ON PWO.POD_AUTO_KEY = POD.POD_AUTO_KEY
            LEFT JOIN WO_BOM PWOB ON PWOB.WOB_AUTO_KEY = PWO.WOB_AUTO_KEY
            LEFT JOIN SO_DETAIL SDE ON SDE.SOD_AUTO_KEY = PS.SOD_AUTO_KEY
            LEFT JOIN SO_HEADER SHE ON SHE.SOH_AUTO_KEY = SDE.SOH_AUTO_KEY
            LEFT JOIN SO_DETAIL SD ON SD.SOD_AUTO_KEY = ROD.SOD_AUTO_KEY  
            LEFT JOIN SO_HEADER SH ON SH.SOH_AUTO_KEY = SD.SOH_AUTO_KEY
            LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = RCH.LOC_AUTO_KEY
            LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = RCH.WOO_AUTO_KEY
            LEFT JOIN PO_HEADER PO ON PO.POH_AUTO_KEY = RCH.POH_AUTO_KEY
            LEFT JOIN PO_DETAIL PD ON PD.POH_AUTO_KEY = PO.POH_AUTO_KEY
            LEFT JOIN PURCHASE_WO PUWO ON PUWO.POD_AUTO_KEY = PD.POD_AUTO_KEY
            LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = PUWO.WOB_AUTO_KEY
            LEFT JOIN WO_OPERATION WNB ON WNB.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
            LEFT JOIN PURCHASE_SALES PSO ON PSO.POD_AUTO_KEY = PD.POD_AUTO_KEY
            LEFT JOIN SO_DETAIL SDO ON SDO.SOD_AUTO_KEY = PSO.SOD_AUTO_KEY
            LEFT JOIN PO_HEADER PCH ON PCH.POH_AUTO_KEY = RCH.POH_AUTO_KEY
            LEFT JOIN PO_DETAIL PCD ON PCD.POH_AUTO_KEY = PCH.POH_AUTO_KEY
            LEFT JOIN PURCHASE_SALES PCS ON PCS.POD_AUTO_KEY = PCD.POD_AUTO_KEY
            LEFT JOIN SO_DETAIL SCD ON SCD.SOD_AUTO_KEY = PCS.SOD_AUTO_KEY
            LEFT JOIN RO_HEADER RCO ON RCO.ROH_AUTO_KEY = RCH.ROH_AUTO_KEY
            LEFT JOIN RO_DETAIL RCDO ON RCDO.ROH_AUTO_KEY = RCO.ROH_AUTO_KEY
            LEFT JOIN RO_HEADER RCHO ON RCHO.ROH_AUTO_KEY = RCDO.ROH_AUTO_KEY
            LEFT JOIN SO_HEADER SCO ON SCO.SOH_AUTO_KEY = RCH.SOH_AUTO_KEY
            LEFT JOIN SO_DETAIL SCDO ON SCDO.SOH_AUTO_KEY = SCO.SOH_AUTO_KEY            
            WHERE RCH.OPEN_FLAG = 'T'
            %s            
        ORDER BY RCH.RCH_AUTO_KEY DESC 
    """%clause     
    recs = selection_dir(date_query,cr)
    return recs
  
def assign_priority(cr,soh_auto_key,roh_auto_key,poh_auto_key,detail_type,soropos):
    #doc_auto_key = soropos and soropos[0][1] or None
    #if not doc_auto_key:
    #    return 0,'There was a problem finding the document.'
    #need to check SOH
    #need to find the WOB_AUTO_KEY from the PO/RO associated with the bom and get WOB.DUE_DATE
    #probably a rel table -  
    #recs = get_due_dates(cr,soh_auto_key,roh_auto_key,poh_auto_key,detail_type,soropos) 
    rc_data,rc_numbers,error = [],[],''
    date_format = '%Y-%m-%d %H:%M:%S' 
    due_date = ''
    priority = 4  
    srp = soropos and soropos[0]
    if srp and srp[13]:
        due_date = srp[13]              
        if rc_data:
            if len(rc_data[-1]) > 8:
                rc_data[-1][8] = priority
            else:
                rc_data[-1].append(priority)
            if due_date:
                rc_data[-1][2] = due_date
        if due_date not in rc_numbers:
            #initialize variables to store date min data
            due_days_list = []
            min_due_days = 0
            priority = ''
            #start the count for quantity sum of lines
            srp.append(0)                 
            #rc_data.append(srp)
            #rc_numbers.append(due_date)      
        if due_date != '': 
            due_days_list.append(datetime.strptime(due_date,date_format) - datetime.now())             
            min_due_days = min(due_days_list) 
            if min_due_days:
                min_due_days = min_due_days.days                
                priority = (min_due_days and ((min_due_days < 4 and 1) or (min_due_days < 11 and 2) or (min_due_days >= 10 and 3))) or ''
                due_date = due_date                                 
    return priority,error

def query_soropos(cr,session_id,label,next_num=0):
    detail_type,last_num,sysnl = '',0,[]
    ro_recs,po_recs,so_recs = [],[] ,[]
    
    if not next_num:
        query = """SELECT SN.SYSNL_AUTO_KEY,SN.DESCRIPTION,
              CASE WHEN SN.NUMBER_PREFIX IS NOT NULL THEN 
              SN.NUMBER_PREFIX || TO_CHAR(SN.LAST_NUMBER + 1)
              ELSE TO_CHAR(SN.LAST_NUMBER + 1) END,
              SN.LAST_NUMBER          
              FROM SYS_NUMBER_LOG SN
              LEFT JOIN SYS_NUMBER_LOG_CODES SNC 
              ON SNC.SYSNLC_AUTO_KEY = SN.SYSNLC_AUTO_KEY
              WHERE SNC.LOG_TYPE_CODE = 'RC'"""
              
        
        sysnl = selection_dir(query,cr)
        sysnl_auto_key = sysnl and sysnl[0][0]
        
        if len(sysnl) == 1:
            next_num = sysnl and sysnl[0][2]
    #import re 
    #next_num = re.sub("[^0-9]", "", next_num)

    if len(sysnl) > 1:
        #create the sysnl locally
        from polls.models import NumberLog

        nl_data = []
        sequence = 3
        for nl in sysnl:
            if nl[1].upper() == 'REGULAR RECEIPT':
                sequence = 0
            elif nl[1].upper() == 'LOT RECEIVING':
                sequence = 1
            elif nl[1].upper() == 'NON STOCK/SUPPLIES':
                sequence = 2
            nl_data.append(NumberLog(
                sysnl_auto_key = nl[0],
                description = nl[1],
                name = nl[1],
                next_number = nl[2],
                session_id = session_id,
                sequence = sequence,
                ))
        try:
            NumberLog.objects.all().delete()
            NumberLog.objects.bulk_create(nl_data)
        except Exception as exc:
            error = exc.args

    if not next_num and len(sysnl) == 1:
        query = """
            SELECT CASE WHEN NUMBER_PREFIX IS NOT NULL THEN NUMBER_PREFIX || TO_CHAR(ABS(LAST_NUMBER)+1) ELSE
                        TO_CHAR(ABS(LAST_NUMBER) + 1) END
                          FROM SYS_NUMBER_LOG SN
                          LEFT JOIN SYS_NUMBER_LOG_CODES SNC ON SNC.SYSNLC_AUTO_KEY = SN.SYSNLC_AUTO_KEY
                          WHERE     
                          SNC.LOG_TYPE_CODE = 'RC'
            """
        last_num = selection_dir(query,cr)
        last_num = last_num and last_num[0] and last_num[0][0] or None
    
    else:
        last_num = next_num
        
    label = label.strip()
  
    if label:       
        #airway = "SELECT CORE_TRACKING_NUMBER FROM EXCHANGE WHERE SOD_AUTO_KEY = SOD.SOD_AUTO_KEY"
        
        query="""SELECT DISTINCT 
                '%s',
                'SO',
                SOH.SO_NUMBER,
                E.CORE_TRACKING_NUMBER,
                SOH.SYSCM_AUTO_KEY,
                SOH.DPT_AUTO_KEY,
                CMP.COMPANY_NAME,
                CMP.CMP_AUTO_KEY,
                SOH.SOH_AUTO_KEY,
                SOD.PNM_AUTO_KEY,
                SOH.COMPANY_REF_NUMBER,
                SOH.SOH_AUTO_KEY,
                SOH.SO_NUMBER,
                SOD.DUE_DATE,
                CP.COMPANY_NAME,
                E.CORE_TRACKING_NUMBER,
                'T',
                WO.SI_NUMBER
                FROM SO_HEADER SOH            
                LEFT JOIN SO_DETAIL SOD ON SOD.SOH_AUTO_KEY = SOH.SOH_AUTO_KEY
                LEFT JOIN STOCK_RESERVATIONS ST ON ST.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
                LEFT JOIN STOCK S ON S.STM_AUTO_KEY = ST.STM_AUTO_KEY
                LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = SOD.ROD_AUTO_KEY
                LEFT JOIN STOCK_RESERVATIONS STR ON STR.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY
                LEFT JOIN STOCK SR ON SR.STM_AUTO_KEY = STR.STM_AUTO_KEY
                LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = S.STM_LOT
                LEFT JOIN SYS_COMPANIES SYSCM ON SOH.SYSCM_AUTO_KEY = SYSCM.SYSCM_AUTO_KEY
                LEFT JOIN EXCHANGE E ON E.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
                LEFT JOIN COMPANIES CMP ON CMP.CMP_AUTO_KEY = SOH.CMP_AUTO_KEY
                LEFT JOIN PURCHASE_SALES PS ON PS.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
                LEFT JOIN PO_DETAIL POD ON POD.POD_AUTO_KEY = PS.POD_AUTO_KEY
                LEFT JOIN PO_HEADER PO ON PO.POH_AUTO_KEY = POD.POH_AUTO_KEY
                LEFT JOIN COMPANIES CP ON CP.CMP_AUTO_KEY = PO.CMP_AUTO_KEY
                WHERE (UPPER(SOH.SO_NUMBER) = UPPER('%s')
                OR UPPER(E.CORE_TRACKING_NUMBER) = UPPER('%s'))
                --AND (STR.STR_AUTO_KEY IS NULL OR ROD.QTY_RESERVED <= 0)
                --AND (ST.STR_AUTO_KEY IS NULL OR ROD.QTY_RESERVED <= 0)
                --AND ROD.QTY_RESERVED < ROD.QTY_REPAIR
                --AND S.QTY_OH > 0 AND SR.QTY_OH > 0
                AND SOD.CLOSED_UPDATE = 'F'
                """%(last_num,label,label)
        ro_recs,po_recs = [],[]    
        so_recs = selection_dir(query,cr)
        #so_rec = so_recs and so_recs[0]
        #open_flag = so_rec and so_rec[-1]
        #soh_auto_key = so_rec and so_rec[8]
        #if not open_flag and soh_auto_key:
            #query="""SELECT COUNT(*) FROM RC_HEADER WHERE OPEN_FLAG = 'T' AND (ROH_AUTO_KEY IN (SELECT ROH_AUTO_KEY FROM RO_HEADER #WHERE OPEN_FLAG = 'F') OR POH_AUTO_KEY IN (SELECT POH_AUTO_KEY FROM PO_HEADER WHERE OPEN_FLAG = 'F'))
            #"""
        if not so_recs:                  
            query="""SELECT DISTINCT
                    '%s',
                    'RO',
                    ROH.RO_NUMBER,
                    CASE WHEN ROD.MSG_AIRWAY_BILL IS NOT NULL 
                    THEN ROD.MSG_AIRWAY_BILL ELSE ROD.TRACKING_NUMBER END,
                    ROH.SYSCM_AUTO_KEY,
                    ROH.DPT_AUTO_KEY,
                    CASE WHEN CW.CMP_AUTO_KEY IS NOT NULL THEN CW.COMPANY_NAME ELSE
                    (CASE WHEN CWO.CMP_AUTO_KEY IS NOT NULL THEN CWO.COMPANY_NAME ELSE 
                    (CASE WHEN CWN.CMP_AUTO_KEY IS NOT NULL THEN CWN.COMPANY_NAME ELSE
                    CMP.COMPANY_NAME END) END) END,
                    CR.CMP_AUTO_KEY,
                    ROH.ROH_AUTO_KEY,
                    ROD.PNM_AUTO_KEY,
                    CASE WHEN ROH.COMPANY_REF_NUMBER IS NOT NULL THEN SOH.COMPANY_REF_NUMBER ELSE ROH.COMPANY_REF_NUMBER END,
                    SOH.SOH_AUTO_KEY,
                    CASE WHEN SOH.SOH_AUTO_KEY IS NOT NULL THEN SOH.SO_NUMBER ELSE
                    (CASE WHEN WO.WOO_AUTO_KEY IS NOT NULL THEN WO.SI_NUMBER ELSE
                    (CASE WHEN WON.WOO_AUTO_KEY IS NOT NULL THEN WON.SI_NUMBER ELSE WOO.SI_NUMBER END) END) END,
                    CASE WHEN WOB.NEED_DATE IS NOT NULL THEN WOB.NEED_DATE ELSE
                    (CASE WHEN WOO.DUE_DATE IS NOT NULL THEN WOO.DUE_DATE ELSE
                    (CASE WHEN WO.DUE_DATE IS NOT NULL THEN WO.DUE_DATE ELSE
                    (CASE WHEN WON.DUE_DATE IS NOT NULL THEN WON.DUE_DATE ELSE SOH.DUE_DATE 
                    END) END) END) END,
                    CP.COMPANY_NAME,
                    ROH.TRACKING_NUMBER,
                    ROH.OPEN_FLAG,
                    WOR.SI_NUMBER              
                    FROM RO_HEADER ROH    
                    LEFT JOIN COMPANIES CR ON CR.CMP_AUTO_KEY = ROH.CMP_AUTO_KEY                    
                    LEFT JOIN RO_DETAIL ROD ON ROD.ROH_AUTO_KEY = ROH.ROH_AUTO_KEY
                    LEFT JOIN STOCK_RESERVATIONS ST ON ST.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY
                    LEFT JOIN STOCK S ON S.STM_AUTO_KEY = ST.STM_AUTO_KEY
                    LEFT JOIN WO_OPERATION WOR ON WOR.WOO_AUTO_KEY = S.STM_LOT
                    LEFT JOIN SYS_COMPANIES SYSCM ON ROH.SYSCM_AUTO_KEY = SYSCM.SYSCM_AUTO_KEY
                    LEFT JOIN SO_DETAIL SOD ON SOD.SOD_AUTO_KEY = ROD.SOD_AUTO_KEY
                    LEFT JOIN SO_HEADER SOH ON SOH.SOH_AUTO_KEY = SOD.SOH_AUTO_KEY
                    LEFT JOIN COMPANIES CMP ON CMP.CMP_AUTO_KEY = SOH.CMP_AUTO_KEY
                    LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = ROD.WOO_AUTO_KEY
                    LEFT JOIN COMPANIES CW ON CW.CMP_AUTO_KEY = WO.CMP_AUTO_KEY
                    LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = ROD.WOB_AUTO_KEY
                    LEFT JOIN WO_OPERATION WOO ON WOO.WOO_AUTO_KEY = WOB.WOO_AUTO_KEY
                    LEFT JOIN COMPANIES CWO ON CWO.CMP_AUTO_KEY = WOO.CMP_AUTO_KEY
                    LEFT JOIN WO_TASK WOT ON WOT.WOT_AUTO_KEY = ROD.WOT_AUTO_KEY
                    LEFT JOIN WO_OPERATION WON ON WON.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
                    LEFT JOIN COMPANIES CWN ON CWN.CMP_AUTO_KEY = WON.CMP_AUTO_KEY
                    LEFT JOIN PURCHASE_SALES PS ON PS.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
                    LEFT JOIN PO_DETAIL POD ON POD.POD_AUTO_KEY = PS.POD_AUTO_KEY
                    LEFT JOIN PO_HEADER PO ON PO.POH_AUTO_KEY = POD.POH_AUTO_KEY
                    LEFT JOIN COMPANIES CP ON CP.CMP_AUTO_KEY = PO.CMP_AUTO_KEY
                    WHERE 
                    (UPPER(ROH.RO_NUMBER)=UPPER('%s')
                    OR UPPER(ROD.MSG_AIRWAY_BILL)= UPPER('%s')
                    OR UPPER(ROD.TRACKING_NUMBER) = UPPER('%s')
                    OR UPPER(ROH.TRACKING_NUMBER) = UPPER('%s'))
                    --AND (ST.STR_AUTO_KEY IS NULL OR ROD.QTY_RESERVED <= 0)
                    --AND ROD.QTY_RESERVED < ROD.QTY_REPAIR
                    AND S.QTY_OH > 0
            """%(last_num,label,label,label,label)
            ro_recs = selection_dir(query,cr)
        if not (so_recs or ro_recs):       
         
            query="""SELECT DISTINCT 
                    '%s',
                   'PO',
                    POH.PO_NUMBER,
                    CASE WHEN POD.AIRWAY_BILL IS NOT NULL 
                    THEN POD.AIRWAY_BILL ELSE POD.TRACKING_NUMBER END,
                    POH.SYSCM_AUTO_KEY,
                    POH.DPT_AUTO_KEY,
                    CASE WHEN CW.CMP_AUTO_KEY IS NOT NULL 
                    THEN CW.COMPANY_NAME ELSE
                    (CASE WHEN COW.CMP_AUTO_KEY IS NOT NULL THEN 
                    COW.COMPANY_NAME ELSE 
                    (CASE WHEN CMP.CMP_AUTO_KEY IS NOT NULL
                    THEN CMP.COMPANY_NAME ELSE                  
                    (CASE WHEN CRO.COMPANY_NAME IS NOT NULL
                    THEN CRO.COMPANY_NAME ELSE                    
                    CMR.COMPANY_NAME END) END) END) END,
                    CP.CMP_AUTO_KEY,
                    POH.POH_AUTO_KEY,
                    POD.PNM_AUTO_KEY,
                    CASE WHEN POH.COMPANY_REF_NUMBER IS NULL THEN SOH.COMPANY_REF_NUMBER ELSE POH.COMPANY_REF_NUMBER END,
                    SOH.SOH_AUTO_KEY,
                    CASE WHEN SOH.SOH_AUTO_KEY IS NOT NULL THEN SOH.SO_NUMBER ELSE
                    WO.SI_NUMBER END,
                    CASE WHEN WB.NEED_DATE IS NOT NULL THEN WB.NEED_DATE ELSE
                    (CASE WHEN WOB.NEED_DATE IS NOT NULL THEN WOB.NEED_DATE ELSE
                    (CASE WHEN W.DUE_DATE IS NOT NULL THEN W.DUE_DATE ELSE
                    (CASE WHEN WO.DUE_DATE IS NOT NULL THEN WO.DUE_DATE ELSE
                    SOH.DUE_DATE END) END) END) END,
                    CP.COMPANY_NAME,
                    POH.TRACKING_NUMBER,
                    POH.OPEN_FLAG,
                    CASE WHEN WOP.SI_NUMBER IS NOT NULL THEN WOP.SI_NUMBER END                   
                    FROM PO_HEADER POH
                    LEFT JOIN COMPANIES CP ON CP.CMP_AUTO_KEY = POH.CMP_AUTO_KEY
                    LEFT JOIN PO_DETAIL POD ON POD.POH_AUTO_KEY = POH.POH_AUTO_KEY
                    LEFT JOIN STOCK_RESERVATIONS ST ON ST.POD_AUTO_KEY = POD.POD_AUTO_KEY
                    LEFT JOIN STOCK S ON S.STM_AUTO_KEY = ST.STM_AUTO_KEY
                    LEFT JOIN WO_OPERATION WOP ON WOP.WOO_AUTO_KEY = S.STM_LOT
                    LEFT JOIN SYS_COMPANIES SYSCM ON POH.SYSCM_AUTO_KEY = SYSCM.SYSCM_AUTO_KEY                                                                                                                                                                         
                    LEFT JOIN PURCHASE_SALES PS ON PS.POD_AUTO_KEY = POD.POD_AUTO_KEY
                    LEFT JOIN SO_DETAIL SOD ON SOD.SOD_AUTO_KEY = PS.SOD_AUTO_KEY
                    LEFT JOIN SO_HEADER SOH ON SOH.SOH_AUTO_KEY = SOD.SOH_AUTO_KEY
                    LEFT JOIN COMPANIES CMP ON CMP.CMP_AUTO_KEY = SOH.CMP_AUTO_KEY
                    LEFT JOIN RO_DETAIL ROD ON ROD.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY
                    LEFT JOIN STOCK_RESERVATIONS STR ON STR.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY
                    LEFT JOIN WO_BOM WB ON WB.WOB_AUTO_KEY = ROD.WOB_AUTO_KEY
                    LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = WB.WOO_AUTO_KEY
                    LEFT JOIN COMPANIES CW ON CW.CMP_AUTO_KEY = W.CMP_AUTO_KEY
                    LEFT JOIN RO_HEADER ROH ON ROH.ROH_AUTO_KEY = ROD.ROH_AUTO_KEY
                    LEFT JOIN COMPANIES CRO ON CRO.CMP_AUTO_KEY = ROH.CMP_AUTO_KEY
                    LEFT JOIN RC_DETAIL RCD ON RCD.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY
                    LEFT JOIN RC_HEADER RCH ON RCH.RCH_AUTO_KEY = RCD.RCH_AUTO_KEY
                    LEFT JOIN COMPANIES CMR ON CMR.CMP_AUTO_KEY = RCH.CMP_AUTO_KEY
                    LEFT JOIN PURCHASE_WO PWO ON PWO.POD_AUTO_KEY = POD.POD_AUTO_KEY                    
                    LEFT JOIN WO_BOM WOB ON WOB.WOB_AUTO_KEY = PWO.WOB_AUTO_KEY
                    LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOB.WOO_AUTO_KEY
                    LEFT JOIN COMPANIES COW ON COW.CMP_AUTO_KEY = WO.CMP_AUTO_KEY
                    WHERE 
                    (UPPER(POH.PO_NUMBER)=UPPER('%s')
                    OR UPPER(POD.AIRWAY_BILL)=UPPER('%s')
                    OR UPPER(POD.TRACKING_NUMBER) = UPPER('%s')
                    OR UPPER(POH.TRACKING_NUMBER) = UPPER('%s'))
                    --AND (STR.STR_AUTO_KEY IS NULL OR ROD.QTY_RESERVED <= 0)
                    --AND (ST.STR_AUTO_KEY IS NULL OR ROD.QTY_RESERVED <= 0)
                    --AND ROD.QTY_RESERVED < ROD.QTY_REPAIR
                    AND S.QTY_OH > 0 AND SR.QTY_OH > 0
            """%(last_num,label,label,label,label)
            po_recs = selection_dir(query,cr)            
    return so_recs + ro_recs + po_recs

def get_customer_info(order_num,cr):
    error = ''
    query = """SELECT CMP.COMPANY_NAME, ROH.CMP_AUTO_KEY, ROH.RO_NUMBER, ROH.DUE_DATE
        FROM  RC_HEADER RCH, RC_DETAIL RCD, RO_DETAIL ROD, RO_HEADER ROH, COMPANIES CMP
        WHERE RCH.RCH_AUTO_KEY = RCD.RCD_AUTO_KEY
        AND RCD.ROD_AUTO_KEY = ROD.ROD_AUTO_KEY
        AND ROD.ROH_AUTO_KEY = ROH.ROH_AUTO_KEY
        AND RCH.OPEN_FLAG = 'T'
        AND RCH.CMP_AUTO_KEY IS NULL
        AND ROH.RO_NUMBER='%s'
        AND CMP.CMP_AUTO_KEY = ROH.CMP_AUTO_KEY

        UNION     

        SELECT CMP.COMPANY_NAME, POH.CMP_AUTO_KEY, POH.PO_NUMBER, RCH.ARRIVAL_DATE
        FROM  RC_HEADER RCH, RC_DETAIL RCD, PO_DETAIL POD, PO_HEADER POH, COMPANIES CMP
        WHERE RCH.RCH_AUTO_KEY = RCD.RCD_AUTO_KEY
        AND RCD.POD_AUTO_KEY = POD.POD_AUTO_KEY 
        AND POD.POH_AUTO_KEY = POH.POH_AUTO_KEY 
        AND RCH.OPEN_FLAG = 'T'
        AND RCH.CMP_AUTO_KEY IS NULL
        AND POH.PO_NUMBER='%s'
        AND CMP.CMP_AUTO_KEY = POH.CMP_AUTO_KEY        
        
        UNION  

        SELECT CMP.COMPANY_NAME, SOH.CMP_AUTO_KEY, SOH.SO_NUMBER, SOH.DUE_DATE
        FROM  RC_HEADER RCH, RC_DETAIL RCD, SO_DETAIL SOD, 
        SO_HEADER SOH, COMPANIES CMP
        WHERE RCH.RCH_AUTO_KEY = RCD.RCD_AUTO_KEY
        AND RCD.SOD_AUTO_KEY = SOD.SOD_AUTO_KEY 
        AND SOD.SOH_AUTO_KEY = SOH.SOH_AUTO_KEY 
        AND RCH.OPEN_FLAG = 'T'
        AND RCH.CMP_AUTO_KEY IS NULL
        AND SOH.SO_NUMBER='%s'
        AND CMP.CMP_AUTO_KEY = SOH.CMP_AUTO_KEY          
        """%(order_num,order_num,order_num)
    cmp = selection_dir(query,cr)
    cmp = cmp and cmp[0] or []
    return cmp

def create_rcs_bulk(soropos,priority,session_id,cr):
    from polls.models import WOStatus
    rc_data,error,msg = [],'',''
    today = datetime.now()
    today = today.strftime('%Y-%m-%d')  
    for rec in soropos:
        si_number = rec[12]   
        customer = rec[6]
        due_date = rec[13] and rec[13][:10] or None 
        cmp_auto_key = rec[7]        
        if not customer:
            order_num = rec[2]
            cmp = get_customer_info(order_num,cr)
            if cmp:            
                customer = cmp[0] or ''
                cmp_auto_key = cmp[1] or 0
                si_number = not rec[12] and cmp[2] or rec[12]
                due_date = not rec[13] and cmp[3] and cmp[3][:10] or due_date
                     
        rc_data.append(WOStatus(
            wo_number=rec[2],
            dpt_auto_key=rec[5] or 0,
            syscm_auto_key=rec[4] or 0,
            cmp_auto_key = cmp_auto_key or 0,
            parent_auto_key = rec[8] or 0,
            wo_type= rec[1],
            customer = customer,
            session_id = session_id,
            next_num = rec[0],
            airway_bill = rec[3],
            priority = priority,
            arrival_date = today,
            pnm_modify = rec[9],
            cust_ref_number = rec[10],
            si_number = si_number or '',
            due_date = due_date or None,
            vendor = rec[14],
            spn_code = rec[15],#header tracking
            is_toll = rec[16] == 'T' and True or False,#open_flag
            stock_owner = rec[17]#lot#
            
        ))
    msg = WOStatus.objects.bulk_create(rc_data)    
    return msg

def get_woo_from_stock(stm_auto_key,user_id='',quapi=None):
    query = """SELECT W.WOO_AUTO_KEY,W.SI_NUMBER FROM WO_OPERATION W
          LEFT JOIN STOCK_RESERVATIONS SR ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY
          LEFT JOIN STOCK S ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY
                  WHERE S.STM_AUTO_KEY = %s
                  AND S.HISTORICAL_FLAG = 'F'"""%stm_auto_key
    res = selection(query, user_id=user_id,quapi=quapi)
    return res
    
def get_records(ctrl_id=None,ctrl_number=None,wo_number=None,woo_auto_key=None,rack_auto_key=None,user_id='',quapi=None):
    where_clause = ''
    if ctrl_id and ctrl_number:
        where_clause = "WHERE S.CTRL_ID = %s AND S.CTRL_NUMBER = %s"%(ctrl_id,ctrl_number)
    elif wo_number:
        where_clause = "WHERE W.SI_NUMBER = '%s'"%wo_number
    elif woo_auto_key:
        where_clause = "WHERE SR.WOO_AUTO_KEY = %s"%woo_auto_key
    elif rack_auto_key:
        where_clause = "WHERE S.IC_UDL_005 = %s"%rack_auto_key    
    query = """SELECT W.SI_NUMBER,P.PN, P.DESCRIPTION,S.SERIAL_NUMBER,L.LOCATION_CODE,WH.WAREHOUSE_CODE,
    UDL.UDL_CODE,S.CTRL_ID,S.CTRL_NUMBER,W.WOO_AUTO_KEY,S.STM_AUTO_KEY,S.IC_UDL_005,
    S.LOC_AUTO_KEY,S.WHS_AUTO_KEY,W.WOS_AUTO_KEY FROM STOCK S 
    LEFT JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY = S.STM_AUTO_KEY 
    LEFT JOIN WO_OPERATION W ON W.WOO_AUTO_KEY = SR.WOO_AUTO_KEY 
    LEFT JOIN PARTS_MASTER P ON P.PNM_AUTO_KEY = S.PNM_AUTO_KEY
    LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = S.LOC_AUTO_KEY
    LEFT JOIN WAREHOUSE WH ON WH.WHS_AUTO_KEY = S.WHS_AUTO_KEY
    LEFT JOIN WO_STATUS WS ON WS.WOS_AUTO_KEY = W.WOS_AUTO_KEY
    LEFT JOIN USER_DEFINED_LOOKUPS UDL ON UDL.UDL_AUTO_KEY = S.IC_UDL_005 %s"""%where_clause       
    res = selection(query, user_id=user_id,quapi=quapi)
    return res 

def create_records(stock_recs, wos_obj, is_racking=1):
    objects = []
    for wo_number,part_number,description,serial_number,location_code,wh_code,rack,ctrl_id,ctrl_number,woo_auto_key,stm_auto_key,ic_udl_oofive,active,is_dashboard,user_id,is_racking in stock_recs:
        objects.append(wos_obj(
                wo_number = wo_number,#0 - W.SI_NUMBER
                part_number = part_number,#1- P.PN
                description = description,#2- P.DESCRIPTION
                serial_number = serial_number,#3- S.SERIAL_NUMBER
                location_code = location_code,#4- L.LOCATION_CODE
                wh_code = wh_code,#5 -WH.WAREHOUSE_CODE
                rack = rack,#6- S.IC_UDL_005
                ctrl_id = ctrl_id,#7- S.CTRL_ID
                ctrl_number = ctrl_number,#8-S.CTRL_NUMBER
                woo_auto_key = woo_auto_key,#9-W.WOO_AUTO_KEY
                stm_auto_key = stm_auto_key,#10-S.STM_AUTO_KEY
                active = active,
                is_dashboard = is_dashboard,
                user_id = user_id,  
                is_racking = is_racking,                
        ))
    msg = wos_obj.objects.bulk_create(objects)
    return msg
    
def update_stock_rack_beta(sysur_auto_key,user_id,stm_keys,\
    stock_recs=[],dpt_auto_key=None,cond_code=None,\
    consignment=None,syscm_auto_key=None,\
    qty_input=None,rack_auto_key=None,\
    whs_auto_key=None,warehouse_code=None,\
    loc_auto_key=None,location_code=None,cart_code='',\
    existing_loc_key=None,iq_enable=False,\
    dj_user_id='',quapi=None,cr=None,con=None):
    set_where,set_rack,set_loc,set_wh,valid_wh,valid_whs_key = '','','','',False,None
    where_clause,whrack,whloc,whware,set_clause,msg,error='','','','','','',''
    
    if not (cr and con):
        from polls.models import OracleConnection as oc
        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
        orcl_conn = orcl_conn and orcl_conn[0] or None 
        if orcl_conn:
            cr,con = orcl_connect(orcl_conn)
        if not (cr and con):
            return msg,'Cannot connect to Oracle.',valid_whs_key,valid_wh
      
    if stm_keys:
        stm_key = stm_keys and stm_keys[0] or None
        stm_keys = construct_akl(stm_keys)
        query = "SELECT SYSTIMESTAMP FROM DUAL"
        today = selection_dir(query,cr)
        today = today and today[0] and today[0][0] and today[0][0][:18]
        date_format = '%Y-%m-%d %H:%M:%S'
        today = datetime.strptime(today,date_format)
        server_time = today.strftime("%m/%d/%Y %H:%M:%S")
        for keys in stm_keys: 
            where_clause = ' WHERE STM_AUTO_KEY IN %s'%keys 
            #if loc and cart in update mode, then           
            if loc_auto_key:        
                if not rack_auto_key:
                    if iq_enable:
                        set_clause = "SET IC_UDL_005 = NULL,IC_UDL_005_TXT = NULL" 
                        msg = "Cart set to null."
                        where_clause += " AND IC_UDL_005 IS NOT NULL"
                else:
                    if iq_enable and 0:
                        error = 'Cannot update stationary location & cart.'
                    else:
               
                        set_clause = "SET IC_UDL_005 = %s,IC_UDL_005_TXT = '%s'"%(rack_auto_key,cart_code) 
                 
                if not error:
                    #if the user only enters location, then we just check the existing warehouse
                    #set_clause += set_clause and ",loc_auto_key=%s"%loc_auto_key or "SET loc_auto_key=%s"%loc_auto_key
                    #where_clause += " AND (loc_auto_key <> %s OR loc_auto_key IS NULL)"%loc_auto_key
                    msg,error,valid_wh,valid_whs_key,whs_needs_update = check_if_valid_whs_beta(whs_auto_key=whs_auto_key,loc_auto_key=loc_auto_key,dj_user_id=dj_user_id,quapi=quapi)                
                    if valid_whs_key:
                        whs_auto_key = valid_whs_key
            elif rack_auto_key:
                #check to make sure that the stock line's woo is rep'ed on the cart.              
                set_clause = "SET IC_UDL_005 = %s,IC_UDL_005_TXT = '%s'"%(rack_auto_key,cart_code) 
            if not loc_auto_key and warehouse_code:
                from polls.models import Warehouse as whs
                whs_auto_key = whs.objects.filter(warehouse_code=warehouse_code) 
                whs_auto_key = whs_auto_key and whs_auto_key[0] and whs_auto_key[0].whs_auto_key 
                #wos.objects.filter(stm_auto_key = keys)    
                query = "SELECT LOC_AUTO_KEY FROM STOCK WHERE STM_AUTO_KEY IN %s"%keys
                recs = selection(query,quapi=quapi)
                existing_loc_key = recs and recs[0] and recs[0][0] or None 
                if existing_loc_key:                
                    msg,error,valid_wh,valid_whs_key,whs_needs_update = check_if_valid_whs_beta(existing_loc_key=existing_loc_key,whs_auto_key=whs_auto_key,dj_user_id=dj_user_id,quapi=quapi) 
                if not valid_wh:
                    error += "Invalid warehouse for the existing location." 
            if valid_wh and valid_whs_key and whs_auto_key:
                msg='Successful update.'
                #set_clause += set_clause and ",whs_auto_key=%s"%whs_auto_key or "SET whs_auto_key=%s"%whs_auto_key
                #where_clause += " AND (whs_auto_key <> %s OR whs_auto_key IS NULL)"%whs_auto_key
            elif not loc_auto_key and warehouse_code:
                error += "Invalid warehouse for this location."
            if stock_recs and not error:             
                if set_clause and where_clause:               
                    query = "UPDATE STOCK " + set_clause + where_clause
                    error += updation_dir(query,cr)  
                    query = """UPDATE STOCK SET 
                    LOC_VALIDATED = TO_TIMESTAMP('%s','MM-DD-YYYY HH24:MI:SS')
                    WHERE STM_AUTO_KEY IN %s"""%(server_time,keys)
                    error = updation_dir(query,cr)                                         
                    query = """
                        CREATE OR REPLACE PROCEDURE "CSPI_STOCK_UPDATE"
                        (QUSER IN NUMBER, STM IN NUMBER, QCODE IN VARCHAR2)  AS
                        v_query number;
                        v_sysur number;
                        v_pwd varchar2(150);
                        V_CURSOR QC_SC_PKG.CURSOR_TYPE ;
                        BEGIN                 
                            begin
                            qc_trig_pkg.disable_triggers;
                            UPDATE SA_LOG SET SYSUR_AUTO_KEY = QUSER, EMPLOYEE_CODE = QCODE WHERE STA_AUTO_KEY = (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = STM AND EMPLOYEE_CODE = 'DBA');
                            qc_trig_pkg.enable_triggers;
                            end;
                         END CSPI_STOCK_UPDATE;"""   
                    error = updation_dir(query,cr)
                    run_proc = """
                        BEGIN
                        CSPI_STOCK_UPDATE('%s',%s,'%s');
                        END;   
                    """%(sysur_auto_key,stm_key,user_id[:9])
                    
                #if loc_auto_key or whs_auto_key or consignment or cond_code or dpt_auto_key: 
                if loc_auto_key and existing_loc_key != loc_auto_key:                
                    for rec in stock_recs:
                        params=[]
                        params.append(rec[14])#stm_auto_key
                        params.append(qty_input!=None and qty_input or rec[29])#qty_oh
                        params.append(syscm_auto_key or rec[30] or 1)#syscm_auto_key
                        params.append(rec[31] or '')#pcc_auto_key                       
                        params.append(rec[32] or '')#cnc_auto_key
                        params.append(loc_auto_key or rec[21] or '')#loc_auto_key 
                        params.append((valid_wh and whs_auto_key) or rec[22] or '')#whs_auto_key
                        params.append(rec[33] or '')#stc_auto_key
                        params.append(dpt_auto_key or rec[34] or '')#dpt_auto_key
                        params.append(rec[20] or '')#str_auto_key  
                        params.append(rec[35])#qty_reserved
                        params.append(rec[36] or '')#sod_auto_key
                        params.append(rec[37] or '')#rod_auto_key
                        params.append(rec[38] or '')#wob_auto_key
                        params.append(rec[39] or '')#pod_auto_key
                        params.append(rec[11] or '')#woo_auto_key                        
                        error = qry_stock_transfer(sysur_auto_key,user_id,params,quapi,cr=cr,con=con,recs=stock_recs,new_whs_code=warehouse_code,new_loc_code=location_code)
                if consignment:
                    query = consignment and """UPDATE STOCK SET CNC_AUTO_KEY = 
                    (SELECT CNC_AUTO_KEY FROM CONSIGNMENT_CODES WHERE UPPER(CONSIGNMENT_CODE) = UPPER('%s'))
                    %s AND STM_AUTO_KEY = %s"""%(consignment,where_clause,rec[14])
                    cnc_auto_key = query and updation(query,quapi=quapi) or None
                    #cnc_auto_key = cnc_auto_key and cnc_auto_key[0] and cnc_auto_key[0][0] or None
                if cond_code:
                    query = """UPDATE STOCK SET PCC_AUTO_KEY = (SELECT PCC_AUTO_KEY FROM PART_CONDITION_CODES 
                    WHERE UPPER(CONDITION_CODE) = UPPER('%s')) %s AND STM_AUTO_KEY = %s
                    """%(cond_code,where_clause,rec[14])
                    pcc_auto_key = query and updation(query,quapi=quapi) or None
                    #pcc_auto_key = pcc_auto_key and pcc_auto_key[0] and pcc_auto_key[0][0] or None
                if server_time and sysur_auto_key:
                    query = """SELECT SMD.QTY_RESERVED,
                        STR.QTY_SHIP                    
                        FROM STOCK_RESERVATIONS STR
                        JOIN SM_DETAIL SMD ON 
                        SMD.SMD_AUTO_KEY=STR.SMD_AUTO_KEY
                        WHERE STR.STM_AUTO_KEY IN %s
                        """%keys
                    smd_qty = selection_dir(query,cr)
                    smd_qty_res = smd_qty and smd_qty[0] and smd_qty[0][0]
                    smd_qty_ship = smd_qty and smd_qty[0] and smd_qty[0][1]
                    if smd_qty and smd_qty_res and smd_qty_ship:
                        if smd_qty_res > 0 and smd_qty_ship > 0:
                            query="""UPDATE STOCK_RESERVATIONS STR
                               SET STR.PKG_QTY_SCANNED = %s, 
                               STR.SYSUR_AUTO_KEY_SCAN = %s, 
                               STR.DATE_SCAN = TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss')
                               WHERE STR.STM_AUTO_KEY IN %s
                               """%(smd_qty_res,sysur_auto_key,server_time,keys) 
                            error = updation_dir(query,cr)
                orcl_commit(con=con)
                msg = 'Successful update.'
    if error == '{"recs": ""}':
        error = ''
                
    return msg,error,valid_whs_key,valid_wh


##########################UNICAL VERSION###################
"""def update_stock_rack_beta(sysur_auto_key,user_id,stm_keys,\
    stock_recs=[],dpt_auto_key=None,cond_code=None,\
    consignment=None,syscm_auto_key=None,\
    qty_input=None,rack_auto_key=None,\
    whs_auto_key=None,warehouse_code=None,\
    loc_auto_key=None,location_code=None,cart_code='',\
    existing_loc_key=None,iq_enable=False,\
    dj_user_id='',quapi=None,cr=None,con=None): 
    set_where,set_rack,set_loc,set_wh,valid_wh,valid_whs_key = '','','','',False,None
    where_clause,whrack,whloc,whware,set_clause,msg,error='','','','','','',''
    
    if not (cr and con):
        from polls.models import OracleConnection as oc
        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
        orcl_conn = orcl_conn and orcl_conn[0] or None 
        if orcl_conn:
            cr,con = orcl_connect(orcl_conn)
        if not (cr and con):
            return msg,'Cannot connect to Oracle.',valid_whs_key,valid_wh
       
    if stm_keys:
        stm_key = stm_keys and stm_keys[0] or None
        stm_keys = construct_akl(stm_keys)
        query = "SELECT SYSTIMESTAMP FROM DUAL"
        today = selection_dir(query,cr)
        today = today and today[0] and today[0][0] and today[0][0][:18]
        date_format = '%Y-%m-%d %H:%M:%S'
        today = datetime.strptime(today,date_format)
        server_time = today.strftime("%m/%d/%Y %H:%M:%S")
        for keys in stm_keys: 
            where_clause = ' WHERE STM_AUTO_KEY IN %s'%keys 
            #if loc and cart in update mode, then           
            if loc_auto_key:        
                if not rack_auto_key:
                    if iq_enable:
                        set_clause = "SET IC_UDL_005 = NULL,IC_UDL_005_TXT = NULL" 
                        msg = "Cart set to null."
                        where_clause += " AND IC_UDL_005 IS NOT NULL"
                else:
                    if iq_enable and 0:
                        error = 'Cannot update stationary location & cart.'
                    else:
               
                        set_clause = "SET IC_UDL_005 = %s,IC_UDL_005_TXT = '%s'"%(rack_auto_key,cart_code) 
                 
                #if not error:
                    #if the user only enters location, then we just check the existing warehouse
                    #set_clause += set_clause and ",loc_auto_key=%s"%loc_auto_key or "SET loc_auto_key=%s"%loc_auto_key
                    #where_clause += " AND (loc_auto_key <> %s OR loc_auto_key IS NULL)"%loc_auto_key
                    #msg,error,valid_wh,valid_whs_key,whs_needs_update = check_if_valid_whs_beta(whs_auto_key=whs_auto_key,loc_auto_key=loc_auto_key,dj_user_id=dj_user_id,quapi=quapi)                
                    #if valid_whs_key:
                        #whs_auto_key = valid_whs_key
            elif rack_auto_key:
                #check to make sure that the stock line's woo is rep'ed on the cart.              
                set_clause = "SET IC_UDL_005 = %s,IC_UDL_005_TXT = '%s'"%(rack_auto_key,cart_code) 
            #if not loc_auto_key and warehouse_code:
                #from polls.models import Warehouse as whs
                #whs_auto_key = whs.objects.filter(warehouse_code=warehouse_code) 
                #whs_auto_key = whs_auto_key and whs_auto_key[0] and whs_auto_key[0].whs_auto_key 
                #wos.objects.filter(stm_auto_key = keys)    
            query = "SELECT LOC_AUTO_KEY FROM STOCK WHERE STM_AUTO_KEY IN %s"%keys
            recs = selection(query,quapi=quapi)
            existing_loc_key = recs and recs[0] and recs[0][0] or None 
                #if existing_loc_key:                
                    #msg,error,valid_wh,valid_whs_key,whs_needs_update = #check_if_valid_whs_beta(existing_loc_key=existing_loc_key,whs_auto_key=whs_auto_key,dj_user_id=dj_user_id,quapi=quapi) 
                #if not valid_wh:
                    #error += "Invalid warehouse for the existing location." 
            #if valid_wh and valid_whs_key and whs_auto_key:
                #msg='Successful update.'
                #set_clause += set_clause and ",whs_auto_key=%s"%whs_auto_key or "SET whs_auto_key=%s"%whs_auto_key
                #where_clause += " AND (whs_auto_key <> %s OR whs_auto_key IS NULL)"%whs_auto_key
            #elif not loc_auto_key and warehouse_code:
                #error += "Invalid warehouse for this location."
            if stock_recs and not error:             
                if set_clause and where_clause:               
                    query = "UPDATE STOCK " + set_clause + where_clause
                    error += updation_dir(query,cr) 
                    user_time = "Tranferred by user %s at %s"%(user_id,server_time)                    
                    query = """"""UPDATE STOCK SET 
                    LOC_VALIDATED = TO_TIMESTAMP('%s','MM-DD-YYYY HH24:MI:SS'),NOTES='%s'
                    WHERE STM_AUTO_KEY IN %s
                    """"""%(server_time,user_time,keys)
                    error = updation_dir(query,cr)                                         
                    query = """"""
                        CREATE OR REPLACE PROCEDURE "CSPI_STOCK_UPDATE"
                        (QUSER IN NUMBER, STM IN NUMBER, QCODE IN VARCHAR2)  AS
                        v_query number;
                        v_sysur number;
                        v_pwd varchar2(150);
                        V_CURSOR QC_SC_PKG.CURSOR_TYPE ;
                        BEGIN                 
                            begin
                            qc_trig_pkg.disable_triggers;
                            UPDATE SA_LOG SET SYSUR_AUTO_KEY = QUSER, EMPLOYEE_CODE = QCODE WHERE STA_AUTO_KEY = (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = STM AND EMPLOYEE_CODE = 'DBA');
                            qc_trig_pkg.enable_triggers;
                            end;
                         END CSPI_STOCK_UPDATE;""""""   
                    error = updation_dir(query,cr)
                    run_proc = """"""
                        BEGIN
                        CSPI_STOCK_UPDATE('%s',%s,'%s');
                        END;   
                    """"""%(sysur_auto_key,stm_key,user_id[:9])
                    
                #if loc_auto_key or whs_auto_key or consignment or cond_code or dpt_auto_key: 
                if loc_auto_key and existing_loc_key != loc_auto_key:                
                    for rec in stock_recs:
                        params=[]
                        params.append(rec[14])#stm_auto_key
                        params.append(qty_input!=None and qty_input or rec[29])#qty_oh
                        params.append(syscm_auto_key or rec[30] or 1)#syscm_auto_key
                        params.append(rec[31] or '')#pcc_auto_key                       
                        params.append(rec[32] or '')#cnc_auto_key
                        params.append(loc_auto_key or rec[21] or '')#loc_auto_key 
                        params.append((valid_wh and whs_auto_key) or rec[22] or '')#whs_auto_key
                        params.append(rec[33] or '')#stc_auto_key
                        params.append(dpt_auto_key or rec[34] or '')#dpt_auto_key
                        params.append(rec[20] or '')#str_auto_key  
                        params.append(rec[35])#qty_reserved
                        params.append(rec[36] or '')#sod_auto_key
                        params.append(rec[37] or '')#rod_auto_key
                        params.append(rec[38] or '')#wob_auto_key
                        params.append(rec[39] or '')#pod_auto_key
                        params.append(rec[11] or '')#woo_auto_key                        
                        error = qry_stock_transfer(sysur_auto_key,user_id,params,quapi,cr=cr,con=con,recs=stock_recs,new_whs_code=warehouse_code,new_loc_code=location_code)
                if consignment:
                    query = consignment and """"""UPDATE STOCK SET CNC_AUTO_KEY = 
                    (SELECT CNC_AUTO_KEY FROM CONSIGNMENT_CODES WHERE UPPER(CONSIGNMENT_CODE) = UPPER('%s'))
                    %s AND STM_AUTO_KEY = %s""""""%(consignment,where_clause,rec[14])
                    cnc_auto_key = query and updation(query,quapi=quapi) or None
                    #cnc_auto_key = cnc_auto_key and cnc_auto_key[0] and cnc_auto_key[0][0] or None
                if cond_code:
                    query = """"""UPDATE STOCK SET PCC_AUTO_KEY = (SELECT PCC_AUTO_KEY FROM PART_CONDITION_CODES 
                    WHERE UPPER(CONDITION_CODE) = UPPER('%s')) %s AND STM_AUTO_KEY = %s
                    """"""%(cond_code,where_clause,rec[14])
                    pcc_auto_key = query and updation(query,quapi=quapi) or None
                    #pcc_auto_key = pcc_auto_key and pcc_auto_key[0] and pcc_auto_key[0][0] or None
                if server_time and sysur_auto_key:
                    query = """"""SELECT SMD.QTY_RESERVED,
                        STR.QTY_SHIP                    
                        FROM STOCK_RESERVATIONS STR
                        LEFT JOIN SM_DETAIL SMD ON 
                        SMD.SMD_AUTO_KEY=STR.SMD_AUTO_KEY
                        WHERE STR.STM_AUTO_KEY IN %s
                        """"""%keys
                    smd_qty = selection_dir(query,cr)
                    smd_qty_res = smd_qty and smd_qty[0] and smd_qty[0][0]
                    smd_qty_ship = smd_qty and smd_qty[0] and smd_qty[0][1]
                    if smd_qty and smd_qty_res and smd_qty_ship:
                        if smd_qty_res > 0 and smd_qty_ship > 0:
                            query=""""""UPDATE STOCK_RESERVATIONS STR
                               SET STR.PKG_QTY_SCANNED = %s, 
                               STR.SYSUR_AUTO_KEY_SCAN = %s, 
                               STR.DATE_SCAN = TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss')
                               WHERE STR.STM_AUTO_KEY IN %s
                               """"""%(smd_qty_res,sysur_auto_key,server_time,keys) 
                            error = updation_dir(query,cr)
                orcl_commit(con=con)
                msg = 'Successful update.'
                               
    if error == '{"recs": ""}':
        error = ''
                
    return msg,error,valid_whs_key,valid_wh"""

@shared_task
def get_wtls(quapi_id,sysur_auto_key,session_id='',date_from=None,date_to=None):
    error,msg,date_clause = '','',''
    from polls.models import OracleConnection as oc, QueryApi as qa, QuantumUser as qu
    quapi = qa.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.','' 
    if date_from:
        if date_to:
            fdate = datetime.strptime(date_from,'%m/%d/%Y')
            tdate = datetime.strptime(date_to,'%m/%d/%Y')
            if fdate > tdate:
                error = 'Date from must be less than or equal to date to.'
        if user_clause:
            date_clause += ' AND'
        date_clause += " TO_DATE('%s','mm/dd/yyyy') <= WTL.ENTRY_DATE"%(date_from)
    if not error:    
        if date_to:
            if user_clause or date_clause:
                date_clause += ' AND'
            date_clause += " TO_DATE('%s','mm/dd/yyyy') >= WTL.ENTRY_DATE"%(date_to)       
    query = """
        SELECT WTL.WOT_AUTO_KEY,WTM.DESCRIPTION,WOT.SEQUENCE,WO.SI_NUMBER,WTL.START_TIME FROM WO_TASK_LABOR WTL
        LEFT JOIN WO_TASK WOT ON WOT.WOT_AUTO_KEY = WTL.WOT_AUTO_KEY
        LEFT JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WOT.WTM_AUTO_KEY
        LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
        WHERE WTL.SYSUR_AUTO_KEY = %s
        AND WTL.STOP_TIME IS NULL ORDER BY WTL.WTL_AUTO_KEY
        %s
    """%(sysur_auto_key,date_clause)
    wo_tasks = selection_dir(query,cr)
    from polls.models import TaskLabor
    wtl_data = []
    
    for wtl in wo_tasks:
        wtl_data.append(TaskLabor(
        wot_auto_key = wtl[0],
        task_desc = wtl[1],
        sequence = wtl[2],
        wo_number = wtl[3],
        start_time = wtl[4],
        sysur_auto_key = sysur_auto_key,
        session_id = session_id,
        )
    )
    try:
        del_wtls = TaskLabor.objects.filter(sysur_auto_key=sysur_auto_key).delete()
        TaskLabor.objects.bulk_create(wtl_data) or None
    except Exception as err:
        logger.error("Error with user batches. Message: '%s'",err.args)    
    return error,msg   
    
@shared_task
def get_batches(quapi_id,sysur_auto_key,session_id,per_user=False):
    error,msg = '',''
    from polls.models import OracleConnection as oc, QueryApi as qa, QuantumUser as qu
    quapi = qa.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.',''

    if per_user:        
        query = """SELECT BATCH_ID,START_TIME,
            STOP_TIME,DESCRIPTION,SYSUR_AUTO_KEY FROM 
            LABOR_BATCH_HEADER 
            WHERE ROWNUM <= 10 AND SYSUR_AUTO_KEY = %s
            ORDER BY LBH_AUTO_KEY DESC"""%sysur_auto_key
    else:
        query = """SELECT BATCH_ID,START_TIME,
            STOP_TIME,DESCRIPTION,SYSUR_AUTO_KEY FROM 
            LABOR_BATCH_HEADER 
            WHERE ROWNUM <= 10
            ORDER BY LBH_AUTO_KEY DESC"""      
            
    batches = selection_dir(query,cr)
    from polls.models import LaborBatch as lb
    batch_data = []
    for batch in batches:
        batch_data.append(lb(
        batch_id = batch[0],
        description = batch[3],
        start_time = batch[1] or None,
        stop_time = batch[2] or None,
        sysur_auto_key = batch[4] or 0,
        session_id = session_id,
        )
    )
    try:
        #del_lbs = lb.objects.filter(session_id=session_id).delete()
        del_lbs = lb.objects.all().delete()
        lb.objects.bulk_create(batch_data) or None
    except Exception as err:
        logger.error("Error with user batches. Message: '%s'",err.args)    
    return error,msg
    
    
#task to create a new batch
@shared_task
def create_batch(quapi_id,session_id,sysur_auto_key,batch_no,start_batch=False):
    batch_no = str(batch_no[:4]) + '_'
    if not description:
        description = batch_no or 'new'
    if batch_no == '_':
        batch_no = description + batch_no
    if start_batch:
        query = """
            INSERT INTO LABOR_BATCH_HEADER 
            (LBH_AUTO_KEY,SYSUR_AUTO_KEY,BATCH_ID,DESCRIPTION,START_TIME,SVR_START_TIME) 
            VALUES(G_LBH_AUTO_KEY.NEXTVAL,%s,G_LBH_AUTO_KEY.NEXTVAL,'%s',TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'))
        """%(sysur_auto_key,description,right_now,right_now)
    else:
        query = """
            INSERT INTO LABOR_BATCH_HEADER 
            (LBH_AUTO_KEY,SYSUR_AUTO_KEY,BATCH_ID,DESCRIPTION) 
            VALUES(G_LBH_AUTO_KEY.NEXTVAL,%s,G_LBH_AUTO_KEY.NEXTVAL,'%s')
        """%(sysur_auto_key,description)
    error = insertion_dir(query,cr)                    
    
                                 
            
                                                                                

                                                                    
                                                      
                    
                                                                           
                                       
                                                                                   
                                                       
                                   
                                                                  
                                   
                                                               
                                                   
                                          
                                                  
                                                                                   
                                  
                                                             
                                                                      
                                                                     
                                                                   
                                                                    
                                                                      
                                                                                      
              
                                         
                                                                   
                                               
                                                                                                      
                                          
                                      
                                                                                                 
                                                   
                 
                                        
                        
                                                                         
                                          
                       
                                           
                       
                                         
                   
                   
                                           
                                                                                         
                                                                                                                                                               
                              
         
                   
                                           
                                                               
                                                                         
                                        
                                   
                   
                                                
                    
                                                              
                                                             
                 
                   

                        
                         
                                               
                                                                                                                                                                                      
                          
                                                                  
                                                         
                                     
             
                                           
                     
                                                          
                         
                           
                                                  
                                                                 
                                                        
                                             
                                               
                                     
                                                                                    
                              

                      

#parse the batch number
def parse_batch_no(batch_no):
    
    input_code = batch_no.partition('_')
    batch_name = input_code and input_code[0] or ''
    lbh_auto_key = input_code and len(input_code) > 2 and input_code[2] or ''
    
    if '_' in lbh_auto_key:
        lbh_auto_key = lbh_auto_key.partition('_')
        lbh_auto_key = lbh_auto_key and lbh_auto_key[2]
    return lbh_auto_key,batch_name
    
def gen_batch_id(cr,recall_batch_id):
    #1. get all batches like batch_id - XXX
    recall_batch_id = recall_batch_id.split('-')[0]
    query = """SELECT BATCH_ID FROM 
        LABOR_BATCH_HEADER 
        WHERE BATCH_ID LIKE '%s%s' AND BATCH_ID <> '%s'
        """%(recall_batch_id,'%',recall_batch_id)
    batches = selection_dir(query,cr)
    batch_dict,new_batch_id,max_sub = {},recall_batch_id,None
    for batch_id in batches:
        batch_id = batch_id[0]
        batch_split = '-' in batch_id and batch_id.split('-') or None               
        dash_part = batch_split and batch_split[-1] or None
        if dash_part and dash_part.isnumeric():
            batch_dict[batch_id] = int(dash_part)        
    max_sub = batch_dict and max(batch_dict, key=lambda key: batch_dict[key]) or None   
    if max_sub:
        max_dash = batch_dict[max_sub]
        incrementer = str(int(max_dash) + 1)                    
        #grab last part of the si_number and remove it
        #add '-n' where n is an integer that is one more than the dash_part value
        num_char = len(str(max_dash))
        new_batch_id = max_sub[:-num_char] + incrementer
    else:
        new_batch_id = recall_batch_id + '-1'    
    return new_batch_id
       
@shared_task
def batch_labor(quapi_id,session_id,sysur_auto_key,\
    wot_auto_key,batch_no = '',active_batch='',\
    start_batch=False,start_tasks=False,create_batch=False,\
    description='',wot_ids=[],user_name=''):
    error,msg,new_val,field_changed,lbh_auto_key,woo_wots='','','','','',{}
    cart_code,batch_lines = '',''
    from polls.models import OracleConnection as oc, QueryApi as qa, MLApps as maps, QuantumUser as qu
    quapi = qa.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.',batch_no,description,''
    query = "SELECT SYSTIMESTAMP FROM DUAL"
    today = selection_dir(query,cr)
    today = today and today[0] and today[0][0] and today[0][0][:18]
    date_format = '%Y-%m-%d %H:%M:%S'
    #replace the next single line of code with following 2 lines
    #need to add the timedelta = 2 hours for MTU in batching
    #from datetime import timedelta
    #today = datetime.strptime(today,date_format) + timedelta(hours=2)
    today = datetime.strptime(today,date_format)
    right_now = today.strftime("%m/%d/%Y %H:%M:%S")
    if wot_auto_key and not wot_auto_key[-1].isdigit() and not start_batch and not create_batch:
        #if not wot_auto_key[-1].isdigit():
        wot_auto_key = wot_auto_key[:-1]
        query="""SELECT WT.WOT_AUTO_KEY, WOS.STATUS_TYPE,
            WT.SEQUENCE,WTM.DESCRIPTION        
            FROM WO_TASK WT
            LEFT JOIN WO_STATUS WOS ON WT.WOS_AUTO_KEY = WOS.WOS_AUTO_KEY
            LEFT JOIN WO_TASK_MASTER WTM ON WT.WTM_AUTO_KEY = WTM.WTM_AUTO_KEY        
            WHERE WT.WOT_AUTO_KEY = %s"""%wot_auto_key
        rec = selection_dir(query,cr)
        if not rec:
            error = "Task %s not found."%wot_auto_key
        wo_status_type = rec and rec[0] and rec[0][1]
        wot_seq = rec and rec[0] and rec[0][2]
        wot_desc = rec and rec[0] and rec[0][3]
        if wo_status_type == 'Closed':
            error = "Task %s - %s is closed"%(wot_seq,wot_desc)
        if error:
            return error,batch_no,description,''           
    #if we passed active_batch, then we must look it up and get all of the tasks associated with it
    #so we can create a new batch with all of the lbd_auto_key's from the previous branch.
    #Inputs:
    #   User enters/selects - 278_CART 01
    #   User enters CART 01
    #   User enters 278
    batch_id = active_batch and active_batch[:9] or description and description[:9]
    if not create_batch and not active_batch:
        active_batch = batch_no
    if not batch_no and batch_id or create_batch:
        batch_id = gen_batch_id(cr,batch_id)          
    #input_code = active_batch.partition('_')
    #cart_code = input_code and input_code[0] or ''
    #lbh_auto_key = input_code and len(input_code) > 2 and input_code[2] or ''

    #if '_' in lbh_auto_key:
    #    lbh_auto_key = lbh_auto_key.partition('_')
    #    lbh_auto_key = lbh_auto_key and lbh_auto_key[2]
        
    # and not (create_batch or start_batch)
    if active_batch and not wot_ids:
        lbh_sub = """SELECT LBH_AUTO_KEY FROM LABOR_BATCH_HEADER
        WHERE UPPER(BATCH_ID)=UPPER('%s')"""%active_batch
        
        query="""SELECT WOT_AUTO_KEY FROM LABOR_BATCH_DETAIL
        WHERE LBH_AUTO_KEY = (%s)
        """%lbh_sub
        wot_ids = selection_dir(query,cr)
        #wot_ids= [wot for sublist in wot_ids for wot in sublist]
    
    if not batch_no or create_batch:
        #if active_batch is numeric but no lbh_auto_key | or not lbh_auto_key.isnumeric() 
        if not create_batch and not lbh_auto_key and active_batch and not active_batch.isnumeric():
            find_batch = """SELECT LBH_AUTO_KEY FROM LABOR_BATCH_HEADER LBH 
                WHERE UPPER(BATCH_ID) LIKE UPPER('%s%s') 
                AND STOP_TIME IS NULL"""%('%',active_batch)
            lbh_cart = selection_dir(find_batch,cr)
            lbh_auto_key = lbh_cart and lbh_cart[0] and lbh_cart[0][0] or '' 
        if create_batch or not lbh_auto_key:
            # or not lbh_auto_key.isnumeric()            
            batch_no = str(active_batch[:4])
            if not description:
                description = active_batch
            if start_batch:
                query = """
                    INSERT INTO LABOR_BATCH_HEADER 
                    (LBH_AUTO_KEY,SYSUR_AUTO_KEY,BATCH_ID,DESCRIPTION,START_TIME,SVR_START_TIME) 
                    VALUES(G_LBH_AUTO_KEY.NEXTVAL,%s,'%s','%s',TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'))
                """%(sysur_auto_key,batch_id,description,right_now,right_now)
            else:
                query = """
                    INSERT INTO LABOR_BATCH_HEADER 
                    (LBH_AUTO_KEY,SYSUR_AUTO_KEY,BATCH_ID,DESCRIPTION) 
                    VALUES(G_LBH_AUTO_KEY.NEXTVAL,%s,'%s','%s')
                """%(sysur_auto_key,batch_id,description)
            error = insertion_dir(query,cr)
            if error:
                return error,batch_id,description,''               
          
        elif not batch_no and active_batch and cart_code:
            description = cart_code
            cart_code = cart_code + '_'
            query = """
                    INSERT INTO LABOR_BATCH_HEADER 
                    (LBH_AUTO_KEY,SYSUR_AUTO_KEY,BATCH_ID,DESCRIPTION,START_TIME,SVR_START_TIME) 
                    VALUES(G_LBH_AUTO_KEY.NEXTVAL,%s,'%s',
                    '%s',TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'))
                """%(sysur_auto_key,batch_id,description,right_now,right_now)      
            error = insertion_dir(query,cr)
        if description:
            
            batch_id_q = "SELECT BATCH_ID,LBH_AUTO_KEY FROM LABOR_BATCH_HEADER WHERE ROWNUM<=1 ORDER BY LBH_AUTO_KEY DESC"
            batch_id = selection_dir(batch_id_q,cr)  
            batch_no = batch_id and batch_id[0] and batch_id[0][0] or None  
            lbh_auto_key = batch_id and batch_id[0] and batch_id[0][1] or None   
            msg = "Batch ID: \'%s\' created."%(batch_no)
            
            if start_batch and wot_ids:
                for wot in wot_ids:
                    query = """
                        INSERT INTO LABOR_BATCH_DETAIL
                            (LBD_AUTO_KEY,LBH_AUTO_KEY,WOT_AUTO_KEY) 
                        VALUES(G_LBD_AUTO_KEY.NEXTVAL,%s,%s)
                        """%(lbh_auto_key,wot)
                    error = insertion_dir(query,cr)                    
                    query = "SELECT MAX(LBD_AUTO_KEY) FROM LABOR_BATCH_DETAIL"                 
                    lbd = selection_dir(query,cr)
                    lbd_auto_key = lbd and lbd[0] or lbd[0][0] or ''
                    error,task_msg = add_wo_labor(quapi_id,session_id,\
                    sysur_auto_key,lbd_auto_key=lbd_auto_key,\
                    wot_auto_key=wot+'s',start_batch=True,\
                    today='',user_name=user_name)
                if wot_ids:
                    orcl_commit(con=con)
                    return error,batch_no or active_batch,description,msg
    elif not create_batch and batch_no and not lbh_auto_key:   
        batch_header = ''
        #if active_batch and not active_batch.isnumeric() or (batch_no and not batch_no.isnumeric()):
        if not batch_no.isnumeric():
            query = """
                SELECT LBH_AUTO_KEY FROM LABOR_BATCH_HEADER WHERE UPPER(BATCH_ID) LIKE UPPER('%s%s') 
                AND STOP_TIME IS NULL"""%('%',batch_no)
            batch_header = selection_dir(query,cr)
        else:
            query = """
                SELECT LBH_AUTO_KEY FROM LABOR_BATCH_HEADER WHERE BATCH_ID = '%s'
            """%batch_no
            batch_header = selection_dir(query,cr)
        lbh_auto_key = batch_header and batch_header[0] and batch_header[0][0] or None
        if not lbh_auto_key:
            error = 'No batch found.'           
    if not start_batch and not create_batch and active_batch and (not wot_auto_key or start_tasks):
        if active_batch and description:
            batch_lines = """SELECT DISTINCT LBD.WOT_AUTO_KEY,LBD.WOT_AUTO_KEY,LBD.LBD_AUTO_KEY
                FROM LABOR_BATCH_DETAIL LBD
                LEFT JOIN LABOR_BATCH_HEADER LBH ON LBH.LBH_AUTO_KEY = LBD.LBH_AUTO_KEY
                WHERE LBH.BATCH_ID = '%s'"""%(active_batch)
        elif batch_no and lbh_auto_key:
            batch_lines = """SELECT LBD.WOT_AUTO_KEY,WOT.WOT_AUTO_KEY,LBD.LBD_AUTO_KEY 
            FROM LABOR_BATCH_DETAIL LBD
            JOIN WO_TASK WOT ON WOT.WOT_AUTO_KEY = LBD.WOT_AUTO_KEY
            WHERE LBD.LBH_AUTO_KEY = %s"""%lbh_auto_key
        elif not start_tasks and not batch_no:
            if not cart_code: cart_code = active_batch
            batch_lines = """SELECT WOTW.WOT_AUTO_KEY,WOT.WOT_AUTO_KEY,WO.WOO_AUTO_KEY 
            from stock_reservations sr
            left join stock s on s.stm_auto_key = sr.stm_auto_key
            left join wo_operation wo on wo.woo_auto_key = sr.woo_auto_key
            left join wo_task wotw on wotw.woo_auto_key = wo.woo_auto_key
            left join wo_bom wob on wob.woo_auto_key = wo.woo_auto_key 
            left join wo_task wot on wot.wot_auto_key = wob.wot_auto_key
            left join wo_status wos on wos.wos_auto_key = wot.wos_auto_key
            left join user_defined_lookups udl on udl.udl_auto_key = s.ic_udl_005         
            WHERE 
            UPPER(udl.udl_code) = UPPER('%s')
            and (wos.status_type = 'Open' or wos.wos_auto_key is null) 
            order by wot.sequence"""%(cart_code)
        if batch_lines:
            wot_keys = selection_dir(batch_lines,cr)
        if not wot_keys:
            error = 'No open tasks in batch: \'%s\'.'%active_batch
            return error,batch_no or active_batch,description,msg
        woo_wots = {}
        add_wot = False
        for wot in wot_keys:
            if len(wot) > 1 and wot[1] not in woo_wots:                
                woo_wots[wot[1]] = wot[0] or wot[1]
                add_wot = True
            if start_tasks and (add_wot or active_batch.isnumeric()):
                error,msg = add_wo_labor(quapi_id,session_id,\
                sysur_auto_key,wot_auto_key=str(wot[0] or wot[1])+'s',\
                start_batch=True,today='',user_name=user_name,
                lbd_auto_key=wot[2])               
            elif lbh_auto_key:
                query = """SELECT LBD_AUTO_KEY FROM LABOR_BATCH_DETAIL
                WHERE LBH_AUTO_KEY = %s AND WOT_AUTO_KEY = %s
                """%(lbh_auto_key,wot[0])
                dupe = selection_dir(query,cr)
                if dupe:
                    error = 'Task %s already in batch'%wot[0]
                if not error:
                    query = """
                        INSERT INTO LABOR_BATCH_DETAIL
                            (LBD_AUTO_KEY,LBH_AUTO_KEY,WOT_AUTO_KEY) 
                        VALUES(G_LBD_AUTO_KEY.NEXTVAL,%s,%s)
                        """%(lbh_auto_key,wot[0])
                    error = insertion_dir(query,cr)
        #if not active_batch.isnumeric():
            #error,msg = add_wo_labor(quapi_id,session_id,sysur_auto_key,wot_auto_key=active_batch+'s',start_batch=start_batch)
        if not error and not description:
            msg = "Batch \'%s\' recalled on %s."%(description or active_batch,right_now)         
    if lbh_auto_key and wot_auto_key and not start_tasks and not start_batch and not create_batch:
        if not wot_auto_key.isdigit():            
            wot_auto_key = wot_auto_key[:-1]
            #We are going to update each task labor entry for the woos on this cart.
            query = """select wot.wot_auto_key
               from stock_reservations sr
               join stock s on s.stm_auto_key = sr.stm_auto_key
               join wo_operation wo on wo.woo_auto_key = sr.woo_auto_key
               join wo_task wot on wot.woo_auto_key = wo.woo_auto_key
               join wo_task_master wtm on wtm.wtm_auto_key = wot.wtm_auto_key
               join user_defined_lookups udl on udl.udl_auto_key = s.ic_udl_005 
               join sys_users sys on sys.sysur_auto_key = wot.sysur_auto_key
               join wo_skills wok on wok.wok_auto_key = sys.wok_auto_key
               where UPPER(udl.udl_code) = UPPER('%s')"""%(wot_auto_key)
            cart_recs = selection_dir(query,cr)
            for wot in cart_recs:
                query = """SELECT LBD_AUTO_KEY FROM LABOR_BATCH_DETAIL
                WHERE LBH_AUTO_KEY = %s AND WOT_AUTO_KEY = %s
                """%(lbh_auto_key,wot_auto_key)

                #dupe = selection_dir(query,cr)
                #if dupe:
                #    error = 'Task %s already in batch'%wot_auto_key
                if not error:
                    query = """INSERT INTO LABOR_BATCH_DETAIL
                        (LBD_AUTO_KEY,LBH_AUTO_KEY,WOT_AUTO_KEY) 
                    VALUES(G_LBD_AUTO_KEY.NEXTVAL,%s,%s)
                    """%(lbh_auto_key,wot[0])
                    error = insertion_dir(query,cr)
        elif not start_batch:
            query = """SELECT LBD_AUTO_KEY FROM LABOR_BATCH_DETAIL
            WHERE LBH_AUTO_KEY = %s AND WOT_AUTO_KEY = %s
            """%(lbh_auto_key,wot_auto_key)
            #dupe = selection_dir(query,cr)
            #if dupe:
            #    error = 'Task %s already in batch'%wot_auto_key
            if not error:                 
                query = """INSERT INTO LABOR_BATCH_DETAIL
                    (LBD_AUTO_KEY,LBH_AUTO_KEY,WOT_AUTO_KEY) 
                VALUES(G_LBD_AUTO_KEY.NEXTVAL,%s,%s)
                """%(lbh_auto_key,wot_auto_key)
                error = insertion_dir(query,cr)
        if not error:
            msg = "Task %s added to batch."%wot_auto_key
    elif lbh_auto_key and start_tasks:
        #find all WTL's in the batch and set the start time for all
        if start_batch:
            query = """
                UPDATE LABOR_BATCH_HEADER SET START_TIME = TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),
                SVR_START_TIME = TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss')
                WHERE LBH_AUTO_KEY = %s AND SYSUR_AUTO_KEY = %s AND START_TIME IS NULL
            """%(right_now,right_now,lbh_auto_key,sysur_auto_key)
            error = updation_dir(query,cr)
            if not woo_wots:
                query = """SELECT LBD.WOT_AUTO_KEY, LBD.LBD_AUTO_KEY FROM LABOR_BATCH_DETAIL LBD
                           WHERE LBD.LBH_AUTO_KEY = %s
                """%lbh_auto_key
                open_wots = selection_dir(query,cr)
                for wot in open_wots:
                    wot_auto_key = str(wot[0]) + 's'
                    lbd_auto_key = wot[1]
                    error,wtl_msg = add_wo_labor(quapi_id,
                        session_id,sysur_auto_key,
                        lbd_auto_key=lbd_auto_key,
                        wot_auto_key=wot_auto_key,
                        today='',user_name=user_name,
                        start_batch=True)
        elif cart_code and not description:
            error,msg = add_wo_labor(quapi_id,session_id,
            sysur_auto_key,wot_auto_key=batch_no+'c',
            start_batch=True,today='',user_name=user_name)         
        msg = "Batch \'%s\' started on %s."%(batch_no or description,right_now)
        if error == '{"recs": ""}': 
            error = ""
    orcl_commit(con=con)
    if error == '{"recs": ""}' or not error:       
        error = ''
    #if not error and wot_auto_key and not active_batch and start_batch:
        #wot_auto_key = wot_auto_key + 's'
        #error,msg = add_wo_labor(quapi_id,session_id,sysur_auto_key,wot_auto_key=wot_auto_key,user_name=user_name)
        #if error == '{"recs": ""}': 
        #    error = ""
    #register audit trail record  
    #batch_no = not batch_no.isnumeric() and str(lbh_auto_key) + batch_no or batch_no   
    if create_batch:    
        aud_status = 'success'
        #now = today.strftime(date_format)
        app_id = maps.objects.filter(code='labor-tracking')   
        user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
        user_rec = user_rec and user_rec[0] or None
        if user_rec:
            if error:             
                aud_status = 'failure'
                new_val = error
                field_changed = 'Nothing changed.'
                batch_no = ''
            else:
                new_val = batch_no and 'Added new labor batch with batch_id: %s for task: %s'%(batch_no,wot_auto_key) or ''
                field_changed = new_val
            register_audit_trail(user_rec,field_changed,new_val,right_now,app_id,quapi,status=aud_status)
        else:
            error = 'Incorrect Quantum User ID.'        
    return error,batch_no or active_batch,description,msg

@shared_task
def stop_labor(quapi_id,session_id,sysur_auto_key,batch_no,wot_auto_key,yes_complete,user_name=''):
    tc_error,tc_msg,error,msg,new_val,field_changed,description='','','','','','',''
    from polls.models import OracleConnection as oc, QueryApi as qa, MLApps as maps, QuantumUser as qu
    quapi = qa.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.',''              
    #today = datetime.now()   
    #right_now = today.strftime("%m/%d/%Y %H:%M:%S")
    query = "SELECT SYSTIMESTAMP FROM DUAL"
    today = selection_dir(query,cr)
    today = today and today[0] and today[0][0] and today[0][0][:18]
    audit_date = today
    date_format = '%Y-%m-%d %H:%M:%S'
    #replace the next single line of code with following 2 lines
    #need to add the timedelta = 2 hours for MTU in batching
    #from datetime import timedelta
    #today = datetime.strptime(today,date_format) + timedelta(hours=2)
    today = datetime.strptime(today,date_format)
    right_now = today.strftime("%m/%d/%Y %H:%M:%S")
    query = """
        UPDATE LABOR_BATCH_HEADER SET STOP_TIME = TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),
        SVR_STOP_TIME = TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss')
        WHERE BATCH_ID LIKE '%s%s' AND SYSUR_AUTO_KEY = %s AND STOP_TIME IS NULL
    """%(right_now,right_now,'%',batch_no,sysur_auto_key)
    error = updation_dir(query,cr)

    if error == '{"recs": ""}':
        #wot_auto_key = wot_auto_key + 'c'
        query = "SELECT LBH_AUTO_KEY,DESCRIPTION FROM LABOR_BATCH_HEADER WHERE BATCH_ID LIKE '%s%s'"%('%',batch_no)
        batch_header = selection_dir(query,cr)
        lbh_auto_key = batch_header and batch_header[0] and batch_header[0][0] or None
        description = batch_header and batch_header[0] and batch_header[0][1]
        batch_lines = """SELECT DISTINCT LBD.WOT_AUTO_KEY,LBH.START_TIME 
            FROM LABOR_BATCH_DETAIL LBD
            LEFT JOIN LABOR_BATCH_HEADER LBH ON LBH.LBH_AUTO_KEY = LBD.LBH_AUTO_KEY
            WHERE LBH.LBH_AUTO_KEY = %s AND LBH.SYSUR_AUTO_KEY = %s"""%(lbh_auto_key,sysur_auto_key)
        wot_keys = selection_dir(batch_lines,cr)
        hours = 0
        task_count = len(wot_keys)
        
        for wot in wot_keys:
            start_time = wot[1] or None
            if not start_time:
                error += 'Start time not entered.'
            start_time = start_time and datetime.strptime(start_time,date_format) or None
            delta = start_time and today - start_time or False
            hours += delta and delta.days*24 + delta.seconds/3600 or 0              
        total_hours = task_count and hours/task_count or hours        
                                        
        for wot in wot_keys:
            error,msg = add_wo_labor(quapi_id,session_id,sysur_auto_key,wot_auto_key=str(wot[0])+'c',yes_complete=yes_complete,total_hours=total_hours,today='',user_name=user_name)
            tc_error += error
            tc_msg += msg              
            
    if error == '{"recs": ""}':
        error = ''
          
    orcl_commit(con=con)
    #register audit trail record            
    aud_status = 'success'
    #right_now = datetime.now()
    #time_now = right_now.strftime(date_format)
    app_id = maps.objects.filter(code='labor-management')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    field_changed += ' changed.'
    new_val += ' val.'                                   
    if user_rec:
        if error:             
            aud_status = 'failure'
            new_val = error
            field_changed = 'Nothing changed.'
        else:
            new_val = batch_no and 'Set stop time for labor batch with batch_id: %s'%(batch_no) or ''
            msg = " Batch %s stopped on %s."%(batch_no,right_now)  
        register_audit_trail(user_rec,field_changed,new_val,audit_date,app_id,quapi,status=aud_status)
    else:
        error = 'Incorrect Quantum User ID.'
    return error + tc_error,msg + tc_msg
    
@shared_task
def create_labor_line(user_name,session_id,wo_number,sequence,task_desc,hours,start_time,stop_time,right_now):
    
    from polls.models import TaskLabor
    task_labor_del = TaskLabor.objects.filter(sequence = sequence,task_desc = task_desc).delete()
    task_labor = TaskLabor.objects.create(
        session_id = session_id,
        wo_number = wo_number,
        sequence = sequence,
        task_desc = task_desc, 
        hours = hours,
        user_name = user_name,
        entry_date = right_now and right_now[:10] or None,
        start_time = start_time or None,
        stop_time = stop_time or None,
    )
    error = task_labor.save()
    error = error and str(error) or ''
    return error
                 
             
@shared_task 
def add_wo_labor(quapi_id,session_id,sysur_auto_key,\
    wot_auto_key='',yes_complete='F',total_hours=0,\
    start_batch=False,today='',user_name='',lbd_auto_key=0):
    error,msg,where,field_changed,recs,wola_recs,task_master_desc,wot_sequence = '','','','',[],[],'',0
    #query and join on wo_task, wo_skills to get the records for that particular wo_number and cart combo
    from polls.models import OracleConnection as oc, QueryApi as qa, MLApps as maps, QuantumUser as qu
    quapi = qa.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.',''
    #query = "select sysur_auto_key from sys_users where user_id = '%s'"%user_id
    #user = selection_dir(query,cr)
    #sysur_auto_key = user and user[0] and user[0][0] or None
    if not sysur_auto_key:
        error = "User not found."
        return error,msg
                               
    m_format = '%m/%d/%Y %H:%M:%S'
    date_format = '%Y-%m-%d %H:%M:%S'
                      
    #query = "SELECT SYSTIMESTAMP FROM DUAL"
    #today = selection_dir(query,cr)
    #today = today and today[0] and today[0][0] and today[0][0][:18]
    
    if not today: 
        today = datetime.now()               
    else:
        today = datetime.strptime(today,m_format)
    today = today - timedelta(hours=1)
    right_now = today.strftime(m_format)  
    time_now = today.strftime(date_format)
         
    if wot_auto_key:
        """if wot_auto_key[-1] in ['s','S','c','C']:
            query = "SELECT WOT_AUTO_KEY FROM WO_TASK WHERE WOT_AUTO_KEY = %s"%wot_auto_key[:-1]
            wot = selection_dir(query,cr)
            if not wot:
                error = "Task not found."
                return error,msg"""            
        wos_sub = 'select wos_complete from wo_control where rownum <= 1'
        wos = selection_dir(wos_sub,cr)
        wos_auto_key = wos and wos[0] and wos[0][0]  
                
        if wot_auto_key[-1] in ['c','C']:
            wot_auto_key = wot_auto_key[:-1]
            if wot_auto_key.isdigit():
                field_changed = 'Updated task labor entry for task %s. '%wot_auto_key
                #then we will be finding the the WO_LABOR entry by task and stamps the stoptime
                query = """SELECT WTL.WTL_AUTO_KEY,WTL.START_TIME,WO.SI_NUMBER,
                    WOT.SEQUENCE,WTM.DESCRIPTION,WOT.WOS_AUTO_KEY,WOT.WOT_AUTO_KEY FROM WO_TASK_LABOR WTL
                    LEFT JOIN WO_TASK WOT ON WOT.WOT_AUTO_KEY = WTL.WOT_AUTO_KEY
                    LEFT JOIN WO_OPERATION WO ON WO.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
                    LEFT JOIN WO_TASK_MASTER WTM ON WTM.WTM_AUTO_KEY = WOT.WTM_AUTO_KEY
                    WHERE WTL.WTL_AUTO_KEY = (SELECT MAX(WTL_AUTO_KEY) FROM WO_TASK_LABOR 
                    WHERE WOT_AUTO_KEY = %s AND SYSUR_AUTO_KEY = %s AND STOP_TIME IS NULL)"""%(wot_auto_key,sysur_auto_key)
                recs = selection_dir(query,cr)
                #if not recs:
                    #error = "No open task labor entry."  
                #loop through all open WO_LABOR entries for task and user at hand
                #and get start_time and update stop time for t
                wos_task_key = wos_auto_key               
                for rec in recs:   
                    wtl_auto_key = rec and rec[0] or None
                    wos_task_key = rec and rec[5] or None
                    indiv_wot_key = rec and rec[6] or None
                    if not wtl_auto_key:
                        error += 'Task \'%s\' not found.'%indiv_wot_key
                    if wos_task_key == wos_auto_key:
                        return 'Task \'%s\' is already closed.'%indiv_wot_key,msg
                    start_time = rec and rec[1] or None 
                    if not start_time:
                        error = 'Start time not entered.'
                    start_time = start_time and datetime.strptime(start_time,date_format) or None
                    if not total_hours: 
                        delta = start_time and today - start_time or False
                        hours = delta and delta.days*24 + delta.seconds/3600 or 0  
                    else:
                        hours = total_hours                        
                    query = """UPDATE WO_TASK_LABOR SET 
                        STOP_TIME = TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),
                        HOURS = %s,
                        HOURS_BILLABLE = %s                        
                        WHERE WTL_AUTO_KEY = %s"""%(right_now,hours,hours,wtl_auto_key)
                    error = updation_dir(query,cr)

                    if not error or error == '{"recs": ""}':
                        msg += 'Labor stopped. WO #: ' + rec[2]
                        msg += ' - Seq: ' + str(rec[3]) 
                        msg += '-' + str(rec[4]) 
                        msg += '. '
                        error = create_labor_line(user_name,session_id,rec[2],rec[3],rec[4],hours,rec and rec[1] or None,time_now,time_now)   

                if yes_complete == 'T':
                    """
                    a.	Need to check for some things against a WOT 
                        before allowing it to be closed when the 
                        user is prompted “Task Completed”
                        i.	Are there any WWT entries where qty_check_out <> qty_check_in
                            1.	If so display “Tools currently checked out to task”
                        ii.	Are there any WTL entries from a different sysur_auto_key with no stop_time
                        iii.	Are there any wob with qty_needed <> qty_issued

                    """
                    wos_auto_key = wos_task_key                    
                    query = """SELECT WTT.WTT_AUTO_KEY
                        FROM WO_TASK WT
                        LEFT JOIN WO_TASK_TOOLS WTT ON WTT.WOT_AUTO_KEY = WT.WOT_AUTO_KEY
                        WHERE (WTT.QTY_CHECKED_OUT <> WTT.QTY_CHECKED_IN OR WTT.WOT_AUTO_KEY IS NULL)
                        AND WT.WOT_AUTO_KEY = %s
                    """%(wot_auto_key)
                    wtt = selection_dir(query,cr)
                    wtt = wtt and wtt[0]
                    wtt_auto_key = wtt and wtt[0] or ''
                    if wtt_auto_key:
                        error += 'Tools currently checked out to task: %s.'%wot_auto_key
                        
                    query = """SELECT WTL.WTL_AUTO_KEY,WT.WOS_AUTO_KEY
                        FROM WO_TASK WT
                        LEFT JOIN WO_TASK_LABOR WTL ON WTL.WOT_AUTO_KEY = WT.WOT_AUTO_KEY 
                        WHERE (WTL.STOP_TIME IS NULL OR WTL.WOT_AUTO_KEY IS NULL) 
                        AND WT.WOT_AUTO_KEY = %s AND WTL.SYSUR_AUTO_KEY = %s
                    """%(wot_auto_key,sysur_auto_key)
                    wot = selection_dir(query,cr)
                    wot = wot and wot[0]
                    wtl_auto_key = wot and wot[0] or ''
                    if wtl_auto_key:
                        error += 'There is another open labor for this task: %s.'%wot_auto_key

                    query = """SELECT WOB.WOB_AUTO_KEY,WOB.QTY_RESERVED
                        FROM WO_TASK WT
                        LEFT JOIN WO_BOM WOB ON WOB.WOT_AUTO_KEY = WT.WOT_AUTO_KEY 
                        WHERE (WOB.QTY_NEEDED <> WOB.QTY_ISSUED OR WOB.WOT_AUTO_KEY IS NULL)
                        AND WT.WOT_AUTO_KEY = %s
                    """%(wot_auto_key)
                    wot = selection_dir(query,cr)
                    wot = wot and wot[0]
                    wob_auto_key = wot and wot[0] or ''                   
                    if wob_auto_key:
                        error += 'There is an open BOM for this task: %s.'%wot_auto_key                     
                    #Check if WOB[‘QTY_RESERVED’] > 0  
                    qty_reserved = wot and wot[1] or 0                                      
                    if qty_reserved > 0:
                        error += 'The qty reserved must not be greater than 0: %s.'%wot_auto_key
                      
                                
                    if not (wob_auto_key or wtt_auto_key or qty_reserved or wtl_auto_key):                    
                        sub_woc = "(SELECT WOS_COMPLETE FROM WO_CONTROL)"   
                       
                        query = """INSERT INTO WO_TASK_STATUS (SYSUR_AUTO_KEY,
                            WOT_AUTO_KEY,WOS_AUTO_KEY,ENTRY_DATE,SYSTEM_DATE,WOS_PREVIOUS)                               
                            VALUES(%s,%s,%s,TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),TO_DATE('%s','mm/dd/yyyy'),'%s')
                            """%(sysur_auto_key,wot_auto_key,sub_woc,right_now,right_now[:10],wos_task_key)
                        error += insertion_dir(query,cr)
                        if not error:
                            msg += "  Task(s) closed."
                        #update the task status to that of 'closed.'
                        query = """UPDATE WO_TASK SET WOS_AUTO_KEY = (SELECT WOS_COMPLETE FROM WO_CONTROL)
                            WHERE WOT_AUTO_KEY = %s"""%wot_auto_key
                            
                        upd_error = updation_dir(query,cr)
                        if upd_error != '{"recs": ""}':
                            error += upd_error
                        #check signoff status on the task and the status of the task
                        query = """SELECT WT.SYSUR_SIGN_OFF,WT.SYSUR_SIGN_OFF2,WT.WOT_AUTO_KEY
                            FROM WO_TASK WT
                            WHERE WT.WOT_AUTO_KEY = %s
                        """%(wot_auto_key)
                        wot = selection_dir(query,cr)
                        wot = wot and wot[0]
                        wot_auto = wot and wot[2]
                        if not wot_auto:
                            return 'No task found for: %s.'%wot_auto_key,''
                            
                        signoff1 = wot and wot[0]
                        signoff2 = wot and wot[1] 

                        if signoff1 and signoff2:
                            error += 'Task already signed off: %s.'%wot_auto_key
                        query = """SELECT CAN_SIGN_OFF1,
                            CAN_SIGN_OFF2
                            FROM SYS_USERS
                            WHERE SYSUR_AUTO_KEY = %s
                        """%sysur_auto_key
                        
                        wot = selection_dir(query,cr)
                        wot = wot and wot[0]
                        can_signoff1 = wot and wot[0]
                        can_signoff2 = wot and wot[1]

                        if can_signoff1 == 'F' and can_signoff2 == 'F':
                            error += 'User not certified for task sign off.  Task ID: %s.'%wot_auto_key
                       
                        if can_signoff1 == 'T':
                            if not signoff1:
                                set_qry = """UPDATE WO_TASK SET SYSUR_SIGN_OFF = %s,
                                SIGN_OFF_DATE = TO_DATE('%s','mm/dd/yyyy')            
                                WHERE WOT_AUTO_KEY = %s"""%(sysur_auto_key,right_now[:10],wot_auto_key) 
                                upd_error = updation_dir(set_qry,cr)
                                if upd_error != '{"recs": ""}':
                                    error += upd_error

                        if can_signoff2 == 'T':
                            if not signoff2:
                                set_qry = """UPDATE WO_TASK SET SYSUR_SIGN_OFF2 = %s, 
                                SIGN_OFF_DATE2 = TO_DATE('%s','mm/dd/yyyy')             
                                WHERE WOT_AUTO_KEY = %s"""%(sysur_auto_key,right_now[:10],wot_auto_key) 
                                upd_error = updation_dir(set_qry,cr) 
                                if upd_error != '{"recs": ""}':
                                    error += upd_error    
                                    
        elif wot_auto_key[-1] in ['s','S']:
            hours = 0
            sysur_entry = sysur_auto_key
            wot_auto_key = wot_auto_key[:-1]
            if wot_auto_key.isdigit():
                field_changed = 'Inserted new task labor entry for task %s. '%wot_auto_key
                query = """SELECT WK.BURDEN_RATE, WK.BILLING_RATE, 
                    SU.WOK_AUTO_KEY, 'dummy', WK.FIXED_OVERHEAD
                    FROM WO_SKILLS WK 
                    LEFT JOIN SYS_USERS SU ON SU.WOK_AUTO_KEY = WK.WOK_AUTO_KEY                 
                    WHERE SU.SYSUR_AUTO_KEY = %s"""%sysur_auto_key
                recs = selection_dir(query,cr)
                rec = recs and recs[0] or None
                burden_rate = rec and rec[0] or 0
                billing_rate = rec and rec[1] or 0
                wok_auto_key = rec and rec[2] or None
                overhead = rec and rec[4] or 0
                if not start_batch:
                    query = """SELECT START_TIME,WOT_AUTO_KEY FROM WO_TASK_LABOR
                        WHERE SYSUR_AUTO_KEY = %s AND
                        STOP_TIME IS NULL AND START_TIME IS NOT NULL
                        """%sysur_auto_key
                    rec = selection_dir(query,cr)
                    rec = rec and rec[0]
                    start_time = rec and rec[0]
                    wot_started = rec and rec[1]
                    if start_time and wot_started:
                        error =  'Labor for task, %s, already started for user.'%wot_started
                        return error,msg  
                query = """SELECT WO.SI_NUMBER,WOT.SEQUENCE,WOK.DESCRIPTION,
                   WTM.DESCRIPTION,UDL.UDL_CODE,WOT.WOT_AUTO_KEY,WOT.WOS_AUTO_KEY                                    
                   FROM WO_TASK WOT
                   left join wo_operation wo on wo.woo_auto_key = WOT.woo_auto_key
                   LEFT JOIN stock_reservations SR ON SR.WOO_AUTO_KEY = WOT.WOO_AUTO_KEY
                   left join stock s on s.stm_auto_key = sr.stm_auto_key
                   left join wo_task_master wtm on wtm.wtm_auto_key = wot.wtm_auto_key
                   left join user_defined_lookups udl on udl.udl_auto_key = s.ic_udl_005 
                   left join sys_users sys on sys.sysur_auto_key = wot.sysur_auto_key
                   left join wo_skills wok on wok.wok_auto_key = sys.wok_auto_key
                   WHERE WOT.WOT_AUTO_KEY = %s"""%wot_auto_key 
                stock_recs = selection_dir(query,cr)
                stock_rec = stock_recs and stock_recs[0] or None
                wola_recs = rec and stock_rec and [rec + stock_rec] or []
                wot_key = stock_rec and stock_rec[5] or None                      
                wos_task_key = stock_rec and stock_rec[6] or None
                if not wot_key:
                    error =  'Task \'%s\' not found.'%wot_auto_key
                    return error,msg
                if wos_task_key == wos_auto_key:
                    error =  'Task \'%s\' is already closed.'%wot_auto_key
                    return error,msg
                
                if wok_auto_key:
                    #query = """INSERT INTO WO_TASK_LABOR 
                    #(LBD_AUTO_KEY,WTL_AUTO_KEY,WOK_AUTO_KEY,HOURS,BILLING_RATE,BURDEN_RATE,SYSUR_ENTRY,WOT_AUTO_KEY,SYSUR_AUTO_KEY,START_TIME,FIXED_OVERHEAD) 
                    #VALUES (%s,G_WTL_AUTO_KEY.NEXTVAL,%s,%s,%s,%s,'%s',%s, %s,TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),'%s')"""%(lbd_auto_key,wok_auto_key,hours,billing_rate,burden_rate,sysur_entry,wot_auto_key,sysur_auto_key,right_now,overhead)
                    query="""SELECT P.PNM_AUTO_KEY, WOK.WOK_AUTO_KEY
                        FROM PARTS_MASTER P, WO_SKILLS WOK
                        WHERE P.PNM_AUTO_KEY = WOK.PNM_AUTO_KEY
                        AND WOK.WOK_AUTO_KEY = '%s'
                    """%wok_auto_key
                    
                    pnm_skills = selection_dir(query,cr)
                    pnm_auto_key = pnm_skills and pnm_skills[0] and pnm_skills[0][0]

                    if not (billing_rate and pnm_auto_key):
                        return 'Check skills configuration with billing rate and part.',msg
                        
                    query = """INSERT INTO WO_TASK_LABOR 
                    (WTL_AUTO_KEY,PNM_AUTO_KEY,WOK_AUTO_KEY,HOURS,BILLING_RATE,BURDEN_RATE,SYSUR_ENTRY,WOT_AUTO_KEY,SYSUR_AUTO_KEY,START_TIME,FIXED_OVERHEAD) 
                                                                                                                           
                    VALUES (G_WTL_AUTO_KEY.NEXTVAL,%s,%s,%s,%s,%s,'%s',%s, %s,TO_TIMESTAMP('%s','mm/dd/yyyy hh24:mi:ss'),'%s')
                    """%(pnm_auto_key,wok_auto_key,hours,billing_rate,burden_rate,sysur_entry,wot_auto_key,sysur_auto_key,right_now,overhead)

                    error = insertion_dir(query,cr)
                    if error:
                        return error,msg
                    
                    latest_wtl = "SELECT MAX(WTL_AUTO_KEY) FROM WO_TASK_LABOR"
                    upd_query = """UPDATE WO_TASK_LABOR SET LBD_AUTO_KEY=%s WHERE WTL_AUTO_KEY=(%s)"""%(lbd_auto_key,latest_wtl)
                    upd_error = updation_dir(upd_query,cr)
                    if not error:
                        if stock_rec:
                            msg = 'WO #: ' + stock_rec[0]
                            msg += ' | Seq: ' + str(stock_rec[1])
                            msg += ' | ' + stock_rec[3]
                            msg += ' started %s.'%right_now    
                        else:                        
                            msg = 'Task %s | '%wot_auto_key 
                            msg += 'Seq: ' + str(wot_sequence) 
                            msg += ' | ' + task_master_desc
                            msg += ' started %s.'%right_now
                        error += create_labor_line(user_name,session_id,stock_rec[0],stock_rec[1] or wot_sequence,stock_rec[3] or task_master_desc,0,time_now,None,time_now)
                        query = """SELECT SYSUR_AUTO_KEY FROM SYS_USERS WHERE EMPLOYEE_CODE='DBA'"""
                        dba_user = selection_dir(query,cr)
                        dba_user_key = dba_user and dba_user[0] and dba_user[0][0] or None
                        if dba_user_key:
                            query = """SELECT WTU_AUTO_KEY FROM WO_TASK_STATUS 
                                WHERE WOT_AUTO_KEY = %s AND ROWNUM<=1 AND SYSUR_AUTO_KEY = %s
                                ORDER BY WTU_AUTO_KEY DESC
                            """%(wot_auto_key,dba_user_key)
                            wtu = selection_dir(query,cr)
                            wtu_auto_key = wtu and wtu[0] and wtu[0][0] or None
                            if wtu_auto_key:
                                query = """UPDATE WO_TASK_STATUS SET SYSUR_AUTO_KEY = %s 
                                    WHERE WTU_AUTO_KEY = %s
                                    """%(sysur_auto_key,wtu_auto_key)
                                error = updation_dir(query,cr)
                        if yes_complete == 'T':                  
                            query = """UPDATE WO_TASK SET
                                WOS_AUTO_KEY = %s 
                                WHERE WOT_AUTO_KEY = %s"""%(wos_auto_key,wot_auto_key)
                            error = updation_dir(query,cr)
                else:
                    error = 'Employee skill entry not found.'
                    return error,msg
            elif 0:
                #checking if this is a cart and that cart is associated with open labor entries.   
                #We are going to update each task labor entry for the woos on this cart.
                query = """select wot.wot_auto_key,
                   wok.wok_auto_key,wok.billing_rate,wok.burden_rate,
                   wo.si_number,wot.sequence,wok.description,
                   wtm.description,wot.wos_auto_key,wot.wot_auto_key
                   from stock_reservations sr
                   join stock s on s.stm_auto_key = sr.stm_auto_key
                   join wo_operation wo on wo.woo_auto_key = sr.woo_auto_key
                   join wo_task wot on wot.woo_auto_key = wo.woo_auto_key
                   join wo_task_master wtm on wtm.wtm_auto_key = wot.wtm_auto_key
                   join user_defined_lookups udl on udl.udl_auto_key = s.ic_udl_005 
                   join sys_users sys on sys.sysur_auto_key = wot.sysur_auto_key
                   join wo_skills wok on wok.wok_auto_key = sys.wok_auto_key
                   where UPPER(udl.udl_code) = UPPER('%s')"""%wot_auto_key
                recs = selection_dir(query,cr)
                wola_recs = recs
                for wtl_id in recs:
                    wot_key = wtl_id[0]
                    wok_auto_key = wtl_id[1]
                    billing_rate = wtl_id[2]
                    burden_rate = wtl_id[3]                      
                    wos_task_key = wtl_id[8] or None
                    indiv_wot_key = wtl_id[9]
                    if not wot_key:
                        error =  'Task \'%s\' not found.'%wot_auto_key
                        #return error,msg
                    if wos_task_key == wos_auto_key:
                        error =  'Task \'%s\' is already closed.'%indiv_wot_key
                        #return error,msg 
                    if not error:                        
                        query = """INSERT INTO WO_TASK_LABOR 
                           (lbd_auto_key,WTL_AUTO_KEY,WOK_AUTO_KEY,HOURS,BILLING_RATE,BURDEN_RATE,SYSUR_ENTRY,WOT_AUTO_KEY,SYSUR_AUTO_KEY,START_TIME,FIXED_OVERHEAD) 
                           VALUES ('%s',G_WTL_AUTO_KEY.NEXTVAL,%s,%s,%s,%s,'%s',%s, %s,TO_TIMESTAMP('%s','%s','mm/dd/yyyy hh24:mi:ss'),'%s')
                           """%(lbd_auto_key,wok_auto_key,hours,billing_rate,burden_rate,sysur_entry,wot_key,sysur_auto_key,right_now,overhead)
                        error = insertion_dir(query,cr)
                        query = """UPDATE WO_TASK_STATUS SET SYSUR_AUTO_KEY = %s 
                                WHERE WTU_AUTO_KEY = (SELECT WTU_AUTO_KEY FROM 
                                WO_TASK_STATUS WHERE ROWNUM <= 1 AND WOT_AUTO_KEY = %s AND SYSUR_AUTO_KEY = 
                            (SELECT SYSUR_AUTO_KEY FROM SYS_USERS WHERE EMPLOYEE_CODE='DBA'))
                                """%(sysur_auto_key,wot_key)
                        error = updation_dir(query,cr)
                if recs and error in ['','{"recs": ""}']:
                    msg = 'Tasks on cart %s started on %s.'%(wot_auto_key,right_now)
                elif error and error != '{"recs": ""}':
                    return error,msg                   
                elif not recs:
                    error = "No task(s) found."
                    return error,msg
            if wola_recs:
                #insert the new records locally and get the data from the above select
                from polls.models import WOStatus as wos_obj
                error = wo_labor_create(wola_recs,session_id,wos_obj)            
        else:
            error = 'Must append either "c" or "s"'        
    if error == '{"recs": ""}':
        error = '' 
    orcl_commit(con=con)
    #register audit trail record             
    aud_status = 'success'
    time_now = today.strftime(date_format)
    app_id = maps.objects.filter(code='labor-management')   
    user_rec = qu.objects.filter(user_auto_key=sysur_auto_key)
    user_rec = user_rec and user_rec[0] or None
    if user_rec:
        new_val = field_changed + '.'
        if error:             
            aud_status = 'failure'
            new_val = error
            field_changed += ' Nothing changed.' 
        register_audit_trail(user_rec,field_changed,new_val,time_now,app_id,quapi,status=aud_status) 
    else:
        error = 'Incorrect Quantum User Id.'    
    return error,msg
  
@shared_task
def wo_task_status(quapi_id,wot_auto_key,sysur_auto_key):
    error,msg = '',''
    from polls.models import OracleConnection as oc, QueryApi as qa
    quapi = qa.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.',''
    query = """UPDATE WO_TASK_STATUS SET SYSUR_AUTO_KEY = %s 
        WHERE WTU_AUTO_KEY = (SELECT WTU_AUTO_KEY FROM WO_TASK_STATUS 
        WHERE WOT_AUTO_KEY = %s AND ROWNUM <= 1 AND SYSUR_AUTO_KEY = 
        (SELECT SYSUR_AUTO_KEY FROM SYS_USERS WHERE EMPLOYEE_CODE='DBA'))
        """%(sysur_auto_key,wot_auto_key[:-1])
    error = updation_dir(query,cr)
    if error == '{"recs": ""}':
        error = ''
        orcl_commit(con=con)  
    return error,msg
    
def wo_labor_create(wola_recs,session_id,wos_obj):
    error,msg = '',''
    wola_data,error,msg = [],'',''
    for rec in wola_recs:
        wola_data.append(wos_obj(
            session_id=session_id,
            wo_number = rec[4],
            wot_sequence = rec[5],
            skill_desc = rec[6],
            task_master_desc = rec[7],
            rack = rec[8],
            )
        )
    try:
        wos_obj.objects.bulk_create(wola_data) or None
    except Exception as err:
        logger.error("Error with creation of repair order locally. Message: '%s'",err.args) 
    return error

def qry_stock_transfer(sysur_auto_key,user_id,params,quapi,recs=[],new_whs_code='',new_loc_code='',cr=None,con=None):
    orcl_conn = None
    if not cr:
        from polls.models import OracleConnection as oc
        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
        orcl_conn = orcl_conn and orcl_conn[0] or None 
        if orcl_conn:
            cr,con = orcl_connect(orcl_conn)
        if not (cr and con):
            return 'Cannot connect to Oracle.'       
    msg = ''
    #if not reserved, don't delete anything
    """if params[9]:
        query = "DELETE FROM STOCK_RESERVATIONS WHERE (WOB_AUTO_KEY is not null or woo_auto_key is not null or sod_auto_key is not null) AND STR_AUTO_KEY=%s"%params[9]
        try:
            updation_dir(query,cr)
        except cx_Oracle.DatabaseError as exc:
            error, = exc.args
            msg = error.message
            logger.error("Error with unreserving stock: '%s'",msg) """  

    squery = """DECLARE CT qc_utl_pkg.cursor_type; BEGIN CT := 
    QC_STOCK_PKG.SPI_STOCK_TRANSFER('%s', %s, '%s', '%s', '%s', '%s', '%s', '%s', '%s'); close CT; END;"""%(params[0],params[1],params[2],params[3],params[4],params[5],params[6],params[7],params[8])  
    #squery = "DECLARE CT qc_utl_pkg.cursor_type; BEGIN CT := QC_STOCK_PKG.SPI_STOCK_TRANSFER('325300', 1, '1', '1', '1', '1410', '23', '', ''); close CT; END;"
    #squery = "DECLARE CT qc_utl_pkg.cursor_type; BEGIN CT := QC_STOCK_PKG.SPI_STOCK_TRANSFER('325300', 1.0, '1', '1', '1', '2977', '23', '', ''); close CT; END;"
    
    error = updation_dir(squery,cr)
    """except cx_Oracle.DatabaseError as exc:
        error, = exc.args
        msg = error.message
        logger.error("Error with stock transfer: '%s'",msg)"""
    query = """
        CREATE OR REPLACE PROCEDURE "SPI_STOCK_UPDATE"
        (QUSER IN NUMBER, STM IN NUMBER, QCODE IN VARCHAR2)  AS
        v_query number;
        v_sysur number;
        v_pwd varchar2(150);
        V_CURSOR QC_SC_PKG.CURSOR_TYPE ;
        BEGIN                 
            begin
            qc_trig_pkg.disable_triggers;
            UPDATE SA_LOG SET SYSUR_AUTO_KEY = QUSER, EMPLOYEE_CODE = QCODE WHERE STA_AUTO_KEY = (SELECT MAX(STA_AUTO_KEY) FROM SA_LOG WHERE STM_AUTO_KEY = STM AND EMPLOYEE_CODE = 'DBA');
            qc_trig_pkg.enable_triggers;
            end;
         END SPI_STOCK_UPDATE;"""   
    error = updation_dir(query,cr)
    run_proc = """
        BEGIN
        SPI_STOCK_UPDATE('%s',%s,'%s');
        END;   
    """%(sysur_auto_key,params[0],user_id[:9])
    error = updation_dir(run_proc,cr) 
    error = updation_dir(run_proc,cr)
    error = updation_dir(run_proc,cr)    
    
    #TODO - fix problem with re-reservation for TRANSFER mode.           
    if error == '{"recs": ""}':
        error = ''
    if not error and orcl_conn:
        orcl_commit(con=con)
        
    return error
    
def register_audit_trail(user_rec,field_changed,\
    new_val,right_now,ml_apps_id,quapi,\
    status='success',reg_events=[],\
    parameters=[],sysur_auto_key=0):
    error,event_error,mail_error,audit_error='','','',''
    wo_number,pn,task,cond_code,quantity='','','','',''
    ml_apps_id = ml_apps_id and ml_apps_id[0] or None
    app_id = ml_apps_id and ml_apps_id.id or None                                         
    #quapi = quapi and quapi.id                                      
    description = 'Audit trail entry for user=%s,'%user_rec.user_id
    description += ' field(s): %s, %s, '%(field_changed,new_val)
    description += ' app: %s'%(ml_apps_id and ml_apps_id.name or '')
    #from portal.audit_tasks import reg_audit_trail
    from polls.models import AuditTrail as adt,\
    EventManager as evt, MailMail as mail,\
    EventNotification as evt_note,\
    MailGroup as mail_grp
    body = field_changed
    new_log = adt.objects.create(
        field_changed = field_changed,
        description = description,
        new_val = new_val + '.',
        create_date = right_now,
        ml_apps_id = ml_apps_id,
        user_id = user_rec.user_id,
        quapi_id = quapi,
        status = status,
    )
    new_log.save()                                               
    
    if status=='success':                         
        
        for reg_event in reg_events:
                
            #find the mail_group_id and event_id
            #['Non-routine added','Parts Request','BoM Update'] 
            #['Activity to Part Repair','Changed BOM status to CSM']
            #['BOM status to CSM','BOM status to ENG']
            #['BOM status to INV',]
            event_id = evt.objects.filter(
                name = reg_event,
                ml_apps_id = ml_apps_id,
                ) 
                        
            event_id = event_id and event_id[0]
            if not event_id:
                return 'No event found for %s.'%reg_event
                
            mail_group_id = event_id and event_id.mail_group_id
            #if not event_id:
            #    error = 'Event not found. '
                
            adt_id = new_log and new_log.id
            adt_id = adt_id and adt.objects.filter(id=adt_id)
            audit_trail_id = adt_id and adt_id[0] or None
            from polls.models import UserProfile as up
            user_pro = up.objects.filter(sysur_auto_key=sysur_auto_key)
            user_pro = user_pro and user_pro[0]
            user_key = user_pro.user and user_pro.user.id
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user_id = User.objects.filter(id=user_key)
            user_id = user_id and user_id[0]
            
            if parameters and event_id:
                unver_part = parameters[0]
                wo_number = parameters[1]
                entry_date = parameters[2]
                sequence = parameters[3]
                task = parameters[4]
                pn = parameters[5]
                description = parameters[6]
                cond_code = parameters[7]
                orig_qty = parameters[8]
                quantity = parameters[9]
                qty_html = 0                
                notes = parameters[10]
                list_price = parameters[11]
                emp_name = user_id.first_name + ' ' + user_id.last_name
                hold_line = parameters[12]
                reject_part = parameters[13]
                repair_part = parameters[14]
                csm_status = parameters[15]
                eng_status = parameters[16]
                unver_part_html,wo_number_html,entry_date_html,seq_html,task_html,pn_html='','','','','',''
                desc_html,cond_code_html,orig_qty_html='','',''
                qty_html,emp_name_html,notes_html,list_price_html,hold_line_html='','','','',''
                reject_part_html,repair_part_html,csm_status_html,eng_status_html='','','',''
                if unver_part:
                    unver_part_html = """
                       <tr>
                        <td>
                        </td>
                        <td>Unverified Part
                        </td>
                      </tr>
                    """
                if wo_number:
                    wo_number_html = """
                      <tr>
                        <td>WO#
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%wo_number                  
                if entry_date:
                    entry_date_html = """
                      <tr>
                        <td>DATE
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%entry_date
                if sequence:
                    seq_html = """
                      <tr>
                        <td>SEQ
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%sequence
                if task:
                    task_html = """
                      <tr>
                        <td>TASK
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%task
                if pn:
                   pn_html = """
                       <tr>
                        <td>PN
                        </td>
                        <td>%s
                        </td>
                      </tr>
                   """%pn
                if description:
                   desc_html = """
                       <tr>
                        <td>DESC
                        </td>
                        <td>%s
                        </td>
                      </tr>
                   """%description
                if cond_code:
                    cond_code_html = """
                      <tr>
                        <td>COND
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%cond_code
                if orig_qty:
                    orig_qty_html = """
                      <tr>
                        <td>ORIG QTY
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%orig_qty
                if quantity:
                    qty_html = """
                      <tr>
                        <td>QTY
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%quantity
                if emp_name:
                    emp_name_html = """
                      <tr>
                        <td>EMP
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%emp_name
                if notes:
                    notes_html = """
                      <tr>
                        <td>NOTES
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%notes
                if list_price:
                    list_price_html = """
                      <tr>
                        <td>LIST PRICE
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%list_price
                if hold_line:
                    hold_line_html = """
                      <tr>
                        <td>HOLD
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%hold_line
                    
                if reject_part:
                    reject_part_html = """
                      <tr>
                        <td>
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    """%reject_part
                    
                """if repair_part:
                    repair_part_html = 
                      <tr>
                        <td>REPAIR PART
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    %repair_part
                if csm_status:
                    csm_status_html = 
                      <tr>
                        <td>
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    %csm_status  
                if eng_status:
                    eng_status_html = 
                      <tr>
                        <td>
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    %eng_status"""
                event_note ="""
                      <tr>
                        <td>
                        </td>
                        <td>%s
                        </td>
                      </tr>
                    %reg_event"""               
                body_html = """
                    <table style="border 1px solid; margin:5px;">
                      %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s
                    </table>
                """%(unver_part_html,wo_number_html,entry_date_html,seq_html,task_html,pn_html,\
                desc_html,cond_code_html,orig_qty_html,\
                qty_html,emp_name_html,notes_html,list_price_html,hold_line_html,\
                reject_part_html,repair_part_html,csm_status_html,eng_status_html,\
                )
                
                #register the notification and set up the email
                new_event = evt_note.objects.create(
                    create_date = right_now,
                    name = description,
                    trigger_sql = '',
                    quantum_table = '',
                    event_id = event_id,
                    audit_trail_id = audit_trail_id,
                    wo_number = wo_number,
                    pn = pn,
                    task = task,
                    emp_name = user_id.first_name + ' ' + user_id.last_name,
                    cond_code = cond_code,
                    quantity = quantity or 0,
                    user_id = user_id,
                )
                new_event.save()
                
                #register the notification and set up the email
                new_mail = mail.objects.create(
                    create_date = right_now,
                    subject = 'New ' + reg_event,
                    body = body,
                    body_html = body_html,
                    to_emails = mail_group_id.to_emails,  
                    cc_field = mail_group_id.cc_field,
                    from_email = mail_group_id.from_email,
                    mail_group_id = mail_group_id,
                    event_note_id = new_event, 
                    status = 'draft',
                )
                
                new_mail.save()
            
    return error
     
def clear_cart_records(rack_auto_key=None,stm_auto_key=None,quapi=None):
    error,where = '',''
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.'
    if rack_auto_key:
        where = "WHERE IC_UDL_005 = '%s'"%rack_auto_key
    elif stm_auto_key:
        where = "WHERE stm_auto_key = %s"%stm_auto_key
    query = where and "UPDATE STOCK SET IC_UDL_005 = NULL %s"%where or ''
    error = updation_dir(query,cr)
    if error == '{"recs": ""}':
        error = '' 
        error = orcl_commit(con=con)
    return error
    
def set_rack_status(rack_auto_key,new_status,user_id=None,quapi=None):      
    error,where,where_rack = '','',''
    where_rack = "WHERE S.IC_UDL_005 = %s"%rack_auto_key
    #1. look up woo_auto_key based on rack's stock line reservation
    #2. if no reservation with a woo, then we go to the next one
    #3. If we have a woo, we update its status
    where = """WHERE WOO_AUTO_KEY IN 
    (SELECT SR.WOO_AUTO_KEY FROM STOCK_RESERVATIONS SR
    WHERE SR.STM_AUTO_KEY IN (SELECT STM_AUTO_KEY FROM STOCK S
    %s))"""%(where_rack)
    query = "UPDATE WO_OPERATION SET WOS_AUTO_KEY = %s %s"%(new_status,where)
    error = updation(query,quapi=quapi)   
    return error
    
def get_rch_data(quapi,rch_auto_key=None,rc_number=None):
    query = """SELECT DISTINCT RC.RC_NUMBER,
    C.COMPANY_NAME,RC.ORDER_TYPE1,RC.ORDER_NUMBER1,
    RC.DATE_CREATED,L.LOCATION_CODE,RC.AIRWAY_BILL,
    SU.USER_NAME,SC.COMPANY_NAME,RC.RCH_AUTO_KEY
    FROM RC_HEADER RC
    LEFT JOIN COMPANIES C ON RC.CMP_AUTO_KEY = C.CMP_AUTO_KEY
    LEFT JOIN SYS_COMPANIES SC ON RC.SYSCM_AUTO_KEY = SC.SYSCM_AUTO_KEY
    LEFT JOIN SYS_USERS SU ON RC.SYSUR_AUTO_KEY = SU.SYSUR_AUTO_KEY
    LEFT JOIN LOCATION L ON L.LOC_AUTO_KEY = RC.LOC_AUTO_KEY
    WHERE RC.OPEN_FLAG = 'T' """
    if rch_auto_key:   
        where = "AND RC.RCH_AUTO_KEY = %s"%rch_auto_key
    elif rc_number:
        where = "AND RC.RC_NUMBER = '%s'"%rc_number
    recs = selection(query+where,quapi=quapi)
    return recs
    
def create_rch(rch_recs,session_id,location_code):
    from polls.models import StockReceiver as srec
    msg = ''
    if rch_recs:
        bye = srec.objects.filter(session_id=session_id).delete()
    try:
        recs = srec.objects.create(
            rc_number = rch_recs[0],
            company_name = rch_recs[1],
            order_type = rch_recs[2],
            order_number = rch_recs[3],
            create_date = rch_recs[4][:10] or None,
            location_code = location_code,
            airway_bill = rch_recs[6],
            user_name = rch_recs[7],
            account_company = rch_recs[8],
            session_id = session_id,
            rch_auto_key = rch_recs[9],
        )
        recs.save()
    except Exception as error:
        logger.error("Error with creation of RCH locally. Message: '%s'",error.args)
    return msg
    
def update_rc_loc(quapi,rch_auto_key,loc_auto_key):
    error,where = '',''
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.'   
    query = """UPDATE RC_HEADER SET LOC_AUTO_KEY = '%s' WHERE OPEN_FLAG = 'T' AND RCH_AUTO_KEY = '%s'"""%(loc_auto_key,rch_auto_key)
    error = updation_dir(query,cr)
    if error == '{"recs": ""}':
        error = ''
        error = orcl_commit(con=con)    
    return error
    
def update_rc_cart(quapi,rch_auto_key,rack_auto_key,stm_auto_key=None):
    from polls.models import OracleConnection as oc
    orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
    orcl_conn = orcl_conn and orcl_conn[0] or None 
    if orcl_conn:
        cr,con = orcl_connect(orcl_conn)
    if not (cr and con):
        return 'Cannot connect to Oracle.'
    query = """SELECT S.STM_AUTO_KEY FROM STOCK S
        JOIN STOCK_RESERVATIONS SR ON SR.STM_AUTO_KEY=S.STM_AUTO_KEY
        JOIN RC_DETAIL RCD ON RCD.STR_AUTO_KEY = SR.STR_AUTO_KEY
        LEFT JOIN SO_DETAIL SOD ON SOD.SOD_AUTO_KEY = SR.SOD_AUTO_KEY
        LEFT JOIN RO_DETAIL ROD ON ROD.ROD_AUTO_KEY = SR.ROD_AUTO_KEY
        LEFT JOIN PO_DETAIL POD ON POD.POD_AUTO_KEY = SR.POD_AUTO_KEY
        WHERE RCH_AUTO_KEY = %s
        """%rch_auto_key
    stm = selection_dir(query,cr)
    stm_auto_key = stm and stm[0] and stm[0][0]
    if stm_auto_key:
        query = """UPDATE STOCK SET IC_UDL_005 = %s
            WHERE STM_AUTO_KEY = %s
            """%(rack_auto_key,stm_auto_key)
        error = updation_dir(query,cr)
    else:
        error = 'No reservation found. No update to RC.'
    if error == '{"recs": ""}':
        error = ''
        error = orcl_commit(con=con) 
    return error
    
@shared_task
def run_racking_beta(session_id,cond_code=None,syscm_auto_key=None,consignment=None,quantity=None,dpt_auto_key=None,clear_cart=False,stock_label=None,quapi_id=None,mode=0,rack_auto_key=None,lookup_recs = 0,whs_auto_key=None,loc_auto_key=None,new_status='',wo_number='',user_id='',dj_user_id=0,sysur_auto_key=None,woo_key_list=[],ctrl_number=None,ctrl_id=None,woo_auto_key=None,rack='',location='',warehouse='',new_status_name='',iq_enable=False): 
    #initialize variables
    is_rch,rch_auto_key,stm_auto_key,stock_rec,ic_udl_oofive,aud_status = False,None,None,[],None,'failure'
    msg,loc,synch,stat,trail,upd,error,field_changed,new_val,audit_ok,fields_changed,has_error = '','','','','','','','','',False,'',False
    stock_recs,updates,updated_stock_recs,strecs,update_ok,new_whs_code = [],{},[],[],False,''  
    from polls.models import WOStatus as wos_obj,QueryApi
    quapi = QueryApi.objects.filter(id=quapi_id)
    quapi = quapi and quapi[0] or None
    #get woo_auto_key from query and then pass it to all of these queries instead of SI_NUMBER as the key.
    if stock_label and quapi:
        #query= "SELECT RCH_AUTO_KEY FROM RC_HEADER WHERE RC_NUMBER='%s'"%stock_label
        #res = selection(query,quapi=quapi)
        rch_recs = get_rch_data(quapi,rc_number = stock_label)
        rch_auto_key = rch_recs and rch_recs[0] and rch_recs[0][9] or None
        if not rch_auto_key:
            query = "SELECT WOO_AUTO_KEY FROM WO_OPERATION WHERE SI_NUMBER='%s'"%stock_label
            res = selection(query,quapi=quapi)
            woo_auto_key = res and res[0] and res[0][0] or None
    if not (rch_auto_key or woo_auto_key) and not ctrl_id and stock_label:
        ctrl_number = stock_label[:6]
        ctrl_id = stock_label[7:]
    if not session_id:
        return 'Invalid session .',msg 
    if not mode or (mode in ['1','2','3'] and not stock_label) or (mode == '1' and lookup_recs not in [0,'0','False',False]):
        active_woos = wos_obj.objects.filter(active=1, is_dashboard=0, user_id=user_id, is_racking=1)    
        woos_updated = active_woos and active_woos.delete() or None        
    if not error:   
        right_now = datetime.now()
        now = right_now.strftime('%Y-%m-%d %H:%M:%S')
        nowly = right_now.strftime('%m-%d-%Y %H:%M:%S') 
        #Mode 1 - Assign stock records entered by user to rack      
        if mode == '1':  #Update mode
            if rch_auto_key:
                is_rch = True
                rch_recs = rch_recs and rch_recs[0] or []
                if whs_auto_key:
                    error = 'Can only update location and cart for receiver.'
                elif loc_auto_key:
                    #0. Create the update query and run it to change the location and commit.               
                    error = update_rc_loc(quapi,rch_auto_key,loc_auto_key)
                    #1. get the data you need to create the grid line to display to the user
                    #2. create the rch record locally and then pass back a flag to the view method (is_rch= True)
                    # the view will then pass that flag to the template and the template will display the rch grid instead of the typical stock line grid.
                                                 
                    audit_ok = True
                    update_status = error and 'Failed' or 'Successful'
                    new_val = "%s update - RC: %s, location: %s."%(update_status,stock_label,location)
                    msg = new_val
                    field_changed = 'Location'
                elif rack_auto_key:
                    
                    error = update_rc_cart(quapi,rch_auto_key,rack_auto_key)
                    audit_ok = True
                    update_status = error and 'Failed' or 'Successful'
                    new_val = "%s update - RC: %s, cart: %s."%(update_status,stock_label,location)
                    msg = new_val
                    field_changed = 'Cart'
                else:
                    error = 'Must have a location or cart to update the receiver.'
            elif not rch_auto_key and stock_label:
                stm_keys = []            
                strecs = get_wos_from_rack_beta(ctrl_id=ctrl_id,ctrl_number=ctrl_number,woo_auto_key=woo_auto_key,wo_number=stock_label,quapi=quapi) 
                if not strecs:
                    if ctrl_id and ctrl_number:
                        #lookup again by adding a zero to the end of the ctrl_number
                        ctrl_number = ctrl_number + '0'
                        strecs = get_wos_from_rack_beta(ctrl_id=ctrl_id,ctrl_number=ctrl_number,quapi=quapi) 
                if not strecs:
                    strecs = get_wos_from_rack_beta(wo_number=stock_label,quapi=quapi)
                if not strecs:
                    error+= 'No records exist for: %s'%stock_label
                #W.SI_NUMBER_0,SR.WOO_AUTO_KEY_1,W.WOS_AUTO_KEY_2,S.STM_AUTO_KEY_3,SR.STR_AUTO_KEY_4,S.LOC_AUTO_KEY_5,S.WHS_AUTO_KEY_6                
                elif new_status:
                    i=0
                    for stock_rec in strecs: 
                        stm_keys.append(stock_rec[14])                    
                        if stock_rec[13] == new_status: 
                            error= 'The record, \"%s\", is already in this status. '%stock_label
                        else:                            
                            stat = set_status(stock_rec[11],new_status,user_id=user_id,quapi=quapi)               
                            trail = update_trail(sysur_auto_key,stock_rec[11],new_status=new_status,user_id=user_id,quapi=quapi)                           
                            field_changed += ', status'
                            audit_ok = True
                            new_val += str(new_status) 
                            strecs[i][1] = new_status_name#1-description
                            strecs[i][13] = new_status#13wos_auto_key
                            i+=1                       
            if not stock_label and not (cond_code or syscm_auto_key or consignment or quantity or dpt_auto_key or rack_auto_key or loc_auto_key or whs_auto_key or new_status):
                error += ' Enter a record to assign it to a cart, location or warehouse. '
            if not stock_label and new_status:
                error+= ' Enter a record number to assign a new status.' 
            if strecs and not error:
                
                for srec in strecs:
                    if srec and not srec[6] and srec[40] == 'T':
                        return 'Part must have a serial number.','','F'    
                #get all stm_auto_key's in a list so we can bulk update
                if rack_auto_key or loc_auto_key or whs_auto_key or dpt_auto_key or cond_code or consignment or quantity:  
                    if not stm_keys:
                        stm_keys = [stc[14] for stc in strecs]                         
                    msg,error,valid_whs_key,update_whs = update_stock_rack_beta(
                        sysur_auto_key,
                        user_id,
                        stm_keys=stm_keys,
                        stock_recs=strecs,
                        dpt_auto_key=dpt_auto_key,
                        cond_code=cond_code,
                        consignment=consignment,
                        syscm_auto_key=syscm_auto_key,
                        qty_input=quantity,
                        rack_auto_key=rack_auto_key,
                        loc_auto_key=loc_auto_key,
                        whs_auto_key=whs_auto_key,
                        warehouse_code=warehouse,
                        location_code=location,
                        cart_code=rack,               
                        iq_enable=iq_enable,
                        dj_user_id=dj_user_id,
                        quapi=quapi                    
                    )
                    audit_ok = True
                    """i=0
                       
                    for stock_rec in strecs:
                        if not error and (not msg or msg=='Successful Update.') and rack_auto_key:
                            strecs[i][16] = rack
                        if iq_enable:
                            strecs[i][16] = '' 
                        if not error and (not msg or msg=='Successful Update.') and loc_auto_key:
                            strecs[i][7] = location
                        if not error and (not msg or msg=='Successful Update.'):
                            if update_whs and warehouse:
                               strecs[i][15] = warehouse
                            elif valid_whs_key:
                                from polls.models import Warehouse
                                whss = Warehouse.objects.filter(
                                    whs_auto_key=valid_whs_key,
                                    dj_user_id=dj_user_id,
                                    quapi_id=quapi_id
                                )
                                whss = whss and whss[0] or None
                                new_whs_code = whss and whss.warehouse_code or ''
                                strecs[i][15] = new_whs_code
                        i+=1"""
                    new_val = 'User entry:' + str(stock_label) + 'Rack: %s | Loc: %s | Whs: %s | '%(rack,location,warehouse) + msg + ' | ' + error
                if (not error or error == '{"recs": ""}') and not msg:
                    error = ''
                    strecs = get_wos_from_rack_beta(wo_number=stock_label,quapi=quapi)
                    error,sync_msg = add_wo_record(session_id=session_id,is_dashboard=0,is_racking=1,keep_recs=True,woo_recs=strecs,quapi=quapi)  
                    msg += ' Successful update of record #: %s. '%stock_label      
            elif not error and (rack_auto_key or loc_auto_key or whs_auto_key or cond_code or syscm_auto_key or consignment or dpt_auto_key) and not rch_auto_key:                  
                stock_recs = get_wos_from_rack_beta(syscm_auto_key=syscm_auto_key,cond_code=cond_code,consignment=consignment,dpt_auto_key=dpt_auto_key,whs_auto_key=whs_auto_key,loc_auto_key=loc_auto_key,rack_auto_key=rack_auto_key,quapi=quapi)           
                #use method from WO Mgmt to bring in WO's in bulk
                if stock_recs:
                    error,msg = add_wo_record(session_id=session_id,is_dashboard=0,is_racking=1,woo_recs=stock_recs,quapi=quapi)   
                elif not error:
                    error = 'No records found.'                 
                audit_ok = False 
        #Mode 2 - Move stock on rack to another location or rack as long as the stock line is not historical.            
        elif mode == '2': #Move stock to another warehouse, location or cart as long as the stock line is not historical. 
            #if location and iq_enable = 'T', then throw an exception that tells them to go to update and do it one WO at a time          
            #if rack_auto_key and loc_auto_key and iq_enable == True:
                #error+= 'Cannot transfer entire cart to stationary location.'
                #error = ''                
            if not rack_auto_key:
                error+= 'Enter a valid cart code.'
            if rack_auto_key and not error:
                if whs_auto_key or loc_auto_key:                   
                    msg,wh_updated,update_wh = update_location_from_rack(sysur_auto_key,user_id,mode=mode,rack_auto_key=rack_auto_key,loc_auto_key=loc_auto_key,whs_auto_key=whs_auto_key,iq_enable=iq_enable,dj_user_id=dj_user_id,quapi=quapi)
                    update_ok = not error and True or False
                    new_val = wh_updated and warehouse or '' + ' | ' + str(location)
                if new_status:        
                    error += set_rack_status(rack_auto_key,new_status,quapi=quapi)  
                    trail = update_trail(sysur_auto_key,rack_auto_key,new_status=new_status,is_rack=True,quapi=quapi) 
                    update_ok = not error and True or False
                    new_val += 'Rack ' + rack + ' | ' + ' WO status updated to ' + str(new_status)
                if not update_ok:
                    audit_ok = False             
                audit_ok = update_ok                                
                #wos will be a list of lists of data to create the WO's 
                stock_recs = get_wos_from_rack_beta(rack_auto_key=rack_auto_key,quapi=quapi)
                #use method from WO Mgmt to bring in WO's in bulk
                if not stock_recs:
                    error = 'There are no records on the cart.'
                else:
                    error,msg = add_wo_record(session_id=session_id,is_dashboard=0,is_racking=1,woo_recs=stock_recs,quapi=quapi) 
            msg = update_ok and not error and msg + ' Successful update.' or ''             
        #Mode 3 - Validate Location and Rack               
        elif mode == '3': #validate mode
            #if not stock_label:
            #    return error,'Enter a record to validate.'
            #when we have a list of WO's and they are scanned with no rack entered, they will be removed from the list.
            #match with existing record in SQLite
            #first check that it is on the list
           
            if clear_cart and rack and rack_auto_key:
                
                error = clear_cart_records(rack_auto_key=rack_auto_key,quapi=quapi)               
                if not error:
                    wos_obj.objects.all().filter(user_id=user_id,is_dashboard=0,active=1,is_racking=1,rack=rack).delete()
                    audit_ok = True
                    msg = '%s cleared.'%rack
                else:
                    audit_ok = True
                    field_changed = error
                    msg = '%s not cleared.'%rack
                    
            elif stock_label:
                #TODO - need to let the user know either they've already validated or that it isn't on the cart/location/warehouse.
                #1 - we already have the lookup to see if it is in the grid.
                #2. need a quick lookup to see if it is on the cart/location/warehouse.
                woo_to_val = None
                if wo_number[0] in ['c','C']:
                    woo_to_val = wos_obj.objects.filter(session_id=session_id,stm_auto_key=wo_number[1:]) 
                elif not woo_to_val and (ctrl_id and ctrl_id.isdigit() and ctrl_number and ctrl_number.isdigit()):
                    woo_to_val = wos_obj.objects.filter(session_id=session_id,ctrl_id=int(ctrl_id),ctrl_number=int(ctrl_number))      
                else:
                    woo_to_val = wos_obj.objects.filter(session_id=session_id,wo_number = stock_label)
                if woo_to_val:               
                    stm_auto_key = woo_to_val and woo_to_val[0] or None
                    stm_auto_key = stm_auto_key and stm_auto_key.stm_auto_key or None                    
                    stock_rec = stm_auto_key and get_wos_from_rack_beta(stm_auto_key=stm_auto_key,user_id=user_id,quapi=quapi) or None                   
                    if not stock_rec:
                        error+= 'No workorders nor stock lines exist for: %s'%stock_label
                    else:
                        stm_auto_key = stock_rec[0][14]
                        ic_udl_oofive = stock_rec[0][19]
                        location_key = stock_rec[0][21]
                        warehouse_key = stock_rec[0][22]
                        status_key = stock_rec[0][13]
                        woo_auto_key = stock_rec[0][11]
                    if rack_auto_key:
                        if ic_udl_oofive != rack_auto_key:
                            error+= 'This record is not on this rack: %s'%rack 
                    elif loc_auto_key:
                        if loc_auto_key != location_key:
                            error+= 'This record is not in this location: %s'%location
                    elif whs_auto_key: 
                        if whs_auto_key != warehouse_key:
                            error+= 'This record is not in warehouse: %s'%warehouse
                    else:
                        error+= 'You must input a warehouse, location or cart and a record number/scan to validate.'              
                    if stm_auto_key and not error:
                        #update stock table loc_validated with timestamp
                        ##******TODO*******FIND WAY TO GET THE SYSTEM TIMEZONE AND SUBTRACT NUMBER OF HOURS*********
                        from polls.models import QueryApi,OracleConnection as oc
                        quapi = QueryApi.objects.filter(id=quapi_id)
                        quapi = quapi and quapi[0] or None
                        orcl_conn = quapi and quapi.orcl_conn_id and oc.objects.filter(id=quapi.orcl_conn_id) or None
                        orcl_conn = orcl_conn and orcl_conn[0] or None
                        if orcl_conn:
                            cr,con = orcl_connect(orcl_conn)
                        if not (cr and con):
                            return 'Cannot connect to Oracle',msg,False
                        query = "SELECT SYSTIMESTAMP FROM DUAL"
                        today = selection_dir(query,cr)
                        today = today and today[0] and today[0][0] and today[0][0][:18]
                        date_format = '%Y-%m-%d %H:%M:%S'
                        today = datetime.strptime(today,date_format)
                        server_time = today.strftime("%m/%d/%Y %H:%M:%S")                        
                        query = """UPDATE STOCK SET LOC_VALIDATED = TO_TIMESTAMP('%s', 'MM-DD-YYYY HH24:MI:SS')
                                   WHERE STM_AUTO_KEY = %s"""%(server_time,stm_auto_key)
                        error = updation_dir(query,cr)
                        error = orcl_commit(con=con) 
                if not woo_to_val:
                    error+= "Record not found: %s."%stock_label
                if not error or error in ['{"recs": "no errors"}','{"recs": ""}']:             
                    woo_to_val.delete()
                    msg = 'Record Validated.'
                    error = ''                
            #wos will be a list of lists of data to create the WO's
            elif rack_auto_key or loc_auto_key or whs_auto_key and not stock_label:
                #wos will be a list of lists of data to create the WO's
                stock_recs = get_wos_from_rack_beta(rack_auto_key=rack_auto_key,loc_auto_key=loc_auto_key,whs_auto_key=whs_auto_key,user_id=user_id,quapi=quapi)   
                #use method from WO Mgmt to bring in WO's in bulk
                if stock_recs:
                    error,msg = add_wo_record(session_id=session_id,is_dashboard=0,is_racking=1,woo_recs=stock_recs,quapi=quapi) 
                else:
                    field_changed = create_fc_string(type='filter',rack=rack,location=location,warehouse=warehouse)
                    error = 'No records found.'
            else:
                error = 'Enter a cart, location or warehouse to get records to validate or a record you wish to validate against your record set.'
    #must register an audit trail record locally - ADT
    #app_id = MLApps.objects.filter(name='Barcoding')
    if audit_ok:
        if error == '{"recs": ""}' or not error:
            error = ''
            aud_status = 'success'
        else: 
            aud_status = 'failure'  
        from polls.models import MLApps as map,QuantumUser as qu
        maps = map.objects.all()
        app_id = maps.filter(code='barcoding')
        user_rec = qu.objects.filter(user_auto_key = sysur_auto_key)
        user_rec = user_rec and user_rec[0] or None
        stm_keys = (stock_recs and set(rec[14] for rec in stock_recs)) or (updated_stock_recs and set(rec[14] for rec in updated_stock_recs)) or (strecs and set(rec[14] for rec in strecs)) or []
        stm_key_list = list(stm_keys)
        woo_keys = (stock_recs and set(rec[0] for rec in stock_recs)) or (updated_stock_recs and set(rec[0] for rec in updated_stock_recs)) or (strecs and set(rec[0] for rec in strecs)) or []
        woo_num_list = list(woo_keys)
        field_changed = new_val + error
        if aud_status == 'success':
            field_changed = not rch_auto_key and ('User entered: ' + str(stock_label) + ' updated stm_auto_key: ' + str(stm_key_list) + ', WOs: ' + str(woo_num_list) + ' | ' + field_changed + ' | ' + error) or field_changed
        if user_rec and (stock_recs or strecs or is_rch):
            error += register_audit_trail(user_rec,field_changed,new_val,now,app_id,quapi,status=aud_status)
    if error == '{"recs": "no errors"}':             
        error = ''  
    return str(error),str(msg),is_rch
    
def create_fc_string(type='filter',rack=None,location=None,warehouse=None,stock_label=None):
    field_changed = ''
    if type == 'filter':
        field_changed = 'No records found for Cart: ' + str(rack)
        field_changed += str(location) and ', Location: ' + str(location)
        field_changed += str(warehouse) and ', Warehouse: ' + str(warehouse)
    if stock_label and type=='update':
        field_changed = 'Problem exists with your entry for record: %s'%stock_label
    return field_changed