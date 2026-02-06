#!/usr/bin/env python
"""MRO Live Demo Version 1.01.01.02
Trailing last to numbers:
Patch .01 - 7.09.25 Fixes labor mgmt and labor tracking apps
Patch .02 - 7.16.25 Fixes to Labor Tracking batching functionality."""
# -*- coding: utf8 -*-
# encoding=utf8
from portal.forms import WODashboardForm,PIUpdateForm
from polls.models import MoTemplate,PILogs,WOStatus
from polls.models import QueryApi,StatusSelection,WOTask
from polls.models import TaskLabor,Operation
from polls.models import QuantumUser,AppModes,AuditTrail
from polls.models import MLApps,UserAppPerms,Companies
from polls.models import Location,Warehouse,StockCart
from polls.models import UserQuapiRel,UserProfile
from polls.models import UserGroupProfile,StockReceiver 
from polls.models import ColumnSettings,Departments
from polls.models import WarehouseLocation,Document,Sale,Consignments
from polls.models import MailGroup,EventNotification,EventManager,MailMail
from polls.models import ShipVia,Priority
from django.http import Http404
from django.views.generic import TemplateView
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse
from django.core.signing import Signer
import os
import csv
import re
import sys
import importlib
import itertools
import math
import logging                              
logger = logging.getLogger(__name__)
from datetime import datetime, timedelta
from dateutil.parser import parse
from django.contrib.auth import logout
from django.core.exceptions import ValidationError
from django.conf import settings
from django.db.models import F
from django.http import JsonResponse
FILE_PATH = settings.MEDIA_URL  
BASE_PATH = '/home/ubuntu/mo_template'


def get_doc_file(request):
    # Implement your authentication and permission checks here
    #if not request.user.is_authenticated:
    #    raise Http404("Not authorized to access this file.")

    #lookup image key and identify the image via unique identifier
    from django.http import FileResponse
    file_path = ''
    req_get = request.GET
    
    file_hash = req_get.get('Key')    
    filename = Document.objects.filter(file_hash = file_hash)
    
    if filename:
        filename = filename and filename[0] and filename[0].file_name or ''
        file_path = os.path.join(settings.MEDIA_ROOT, '', filename)
        
    if not os.path.exists(file_path):
        raise Http404("File not found.")

    return FileResponse(open(file_path, 'rb'), content_type='application/octet-stream')

def report_tmpl_import(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec,fail_ms = 'user not set',None,''
    val_dict,form = {},{} 
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id 
    val_dict['apps'] = MLApps.objects.all()
        
    if request.method == 'GET':
        form = WODashboardForm() 
        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post, request.FILES)    
        active_user = 'active_user' in req_post and req_post['active_user'] or ''  
        filter_user = 'filter_user' in req_post and req_post['filter_user'] or ''        
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
          
        val_dict.update({
            'msg': msg,
            'active_user': active_user or user_id,
            'filter_user': filter_user or active_user or user_id,
            'user_id': user_id or filter_user or active_user,
            'user_name': user_name,            
            'session_id': session_id,                    
        })

        if request.FILES:               
            error,msg = import_report_temp(request,quapi_id)
        #export_list = export_json_tmpl('CTS Basic Label')    
            
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/report_tmpl_import.html', val_dict)
    
def export_json_tmpl(tmpl_name):
    
    #1. search for template
    #2. get all template lines
    #3. loop through and create 
    #the list of data elements in dictionaries
    
    from polls.models import ReportTmpl as rept,\
        ReportTmplDetail as repline
    
    report_tmpl = rept.objects.filter(name=tmpl_name)
    report_tmpl = report_tmpl and report_tmpl[0]
    rep_lines = repline.objects.filter(rep_tmpl=report_tmpl)   
    export_list = []

    for line in rep_lines:
        
        db_table = line.db_table
        db_field = line.db_field
        db_total = ''
        
        if db_field and db_table:
            db_total = str(db_table) + '.' + str(db_field)
        
        img_path = line.img_path or ''
        fixed_text = line.fixed_text or ''
        
        inner_dict = {
                "type": line.data_type,
                "x": line.xcoord,
                "y": line.ycoord,
                "width": line.width,
                "height": line.height,
                "text": fixed_text,
                "font_size": line.font_size,
                "url": img_path,
                "color": [
                    line.rgb_red,
                    line.rgb_green,
                    line.rgb_blue,
                ],
                "db_field": db_total,
                "formatting": [
                    line.font_bold and 1 or 0,
                    line.font_ital and 1 or 0,
                    line.font_udrl and 1 or 0,
                ],
            }
        
        export_list.append(inner_dict)     
        
    print(export_list)
    return export_list

def import_report_temp(request,quapi_id): 
    error,msg = '',''
    detail_data,report_data,report_meta = [],[],[]

    from polls.models import ReportTmpl,ReportTmplDetail
    req_post = request.POST
    req_files = request.FILES
    canv_height = req_post.get('canv_height',1)
    canv_width = req_post.get('canv_width',1)
    page_width = req_post.get('page_width',1)
    page_height = req_post.get('page_height',1)
    name = req_post.get('name','new report')
    code = req_post.get('code','new-report')
    font_type = req_post.get('font','Helvetica')
    left_margin = req_post.get('left_margin',0)
    text_chunk = req_post.get('text_chunk',0)
    textarea_chunk =req_post.get('textarea_chunk',0)
    app_id = req_post.get('app_id',1)
    app_id = MLApps.objects.filter(id = app_id)
    app_id = app_id and app_id[0] or None
    up_file = req_files['loc_whs_file']
    file_name = up_file.name
    file_name = file_name.replace(' ','_')
    up_file = up_file and Document(docfile=up_file) or None
    up_file.save()
    file_path = BASE_PATH + FILE_PATH + file_name
    
    if not (code and name and page_height and page_width and app_id):
        return 'All fields must be filled in.',msg
        
    report_meta.append(ReportTmpl(
        canv_height = float(canv_height),
        canv_width = float(canv_width),
        page_width = float(page_width),
        page_height = float(page_height),
        name = name,
        code = code,
        app_id = app_id,
        font_type = font_type,
        left_margin = float(left_margin),
        text_chunk = float(text_chunk),
        textarea_chunk = float(textarea_chunk),
    ))
        
    ReportTmpl.objects.bulk_create(report_meta)
    rep_tmpl = ReportTmpl.objects.filter(
        canv_height = float(canv_height),
        canv_width = float(canv_width),
        page_width = float(page_width),
        page_height = float(page_height),
        name = name,
        code = code,
        app_id = app_id,                 
        )
    rep_tmpl = rep_tmpl and rep_tmpl[0]     

    with open(file_path, 'r') as file:
        data = json.load(file)
        
        for row in data:
            error = ''
            data_type = row.get('type','text')           
            xcoord = row.get('x',0)
            ycoord = row.get('y',0)
            x2coord = row.get('x2',0)
            y2coord = row.get('y2',0)
            if not (xcoord and ycoord) and row.get('x1',0):
                xcoord = row.get('x1',0)
                ycoord = row.get('y1',0)
            #image, box, barcode width
            width = row.get('width',3)
            height = row.get('height',3)
            fixed_text = row.get('text','')
            font_type = row.get('font','')
            img_path = row.get('url','')
            if img_path == 'https://example.com/image.jpg':
                img_path = ''
            #font size, color, decorators                
            font_size = row.get('font_size',12)
            font_color = row.get('fontcolor',[0,0,0])
            font_dec = row.get('formatting', [0,0,0])
            db_field = row.get('db_field','')
            db_table,rgb_red,rgb_green,rgb_blue = '','','',''
                
            if db_field:
                db_vals = db_field.split('.',1)
                db_field = db_vals[0]
                if len(db_vals)>1:
                    db_table = db_field
                    db_field = db_vals[1]
                
            if rep_tmpl:
                detail_data.append(ReportTmplDetail(
                    data_type = data_type,
                    xcoord = xcoord,
                    ycoord = ycoord,
                    x2coord = x2coord,
                    y2coord = y2coord,
                    height = height,
                    width = width,
                    fixed_text = fixed_text,
                    font_size = font_size,
                    rgb_red = font_color[0],
                    rgb_green = font_color[1],
                    rgb_blue = font_color[2],
                    font_bold = font_dec[0],
                    font_udrl = font_dec[1],
                    font_ital = font_dec[2],
                    db_field = db_field,
                    db_table = db_table,
                    rep_tmpl = rep_tmpl,
                    img_path = img_path,
                ))
        
        try:        
            tmpls = ReportTmplDetail.objects.bulk_create(detail_data)
        except Exception as e:
            error = e
        msg = 'Successfully added report template.'
    if error:
        msg = ''
    return error,msg

def create_barcodes_pymu(report_vals,app_code,tmpl_code,barcode_txt='',x_delta=-0.85):
    from reportlab.graphics.barcode import code39,code93
    from reportlab.lib.pagesizes import A4,letter,landscape
    from reportlab.lib.units import mm,inch
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import Color,black,red,green   
    #set the filepath and logo_path for the pdf label file.
    file_path = "/home/ubuntu/mo_template/uploads/"
    file_path += str(app_code) + "_" + str(tmpl_code) + ".pdf" 
    img_path = "/home/ubuntu/mo_template/static/logo.jpg"
    model_recs = report_vals['stock_recs']
    db_val = None
    x_delta = x_delta*mm
    #lookup report template based on app and tmpl_code  
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    #courier new fonts directory
    font_path = 'C:\\Users\\Adam\\Envs\\mro_portal\\mo_template\\static\\fonts\\Courier New'
    #reg_path = os.path.join(font_path, 'cour.ttf') 
    path_exists = os.path.exists(font_path)

    if path_exists:
        reg_path = os.path.join(font_path, 'cour.ttf') 
        pdfmetrics.registerFont(TTFont('CourierNew', reg_path))
        #bold font
        bold_path = os.path.join(font_path, 'courbd.ttf') 
        pdfmetrics.registerFont(TTFont('CourierNew-Bold', bold_path))
        # Register the bold-italic font
        boldi_path = os.path.join(font_path, 'courbi.ttf') 
        pdfmetrics.registerFont(TTFont('CourierNew-BoldItalic', boldi_path))
        # Register the italic font
        italic_path = os.path.join(font_path, 'couri.ttf')
        pdfmetrics.registerFont(TTFont('CourierNew-Italic', italic_path))
        from reportlab.pdfbase.pdfmetrics import registerFontFamily

        registerFontFamily('CourierNew', 
                           normal='CourierNew', 
                           bold='CourierNew-Bold', 
                           italic='CourierNew-Italic', 
                           boldItalic='CourierNew-BoldItalic')
    
    if app_code:
        app_id = MLApps.objects.filter(code = app_code)
        app_id = app_id and app_id[0] or None
       
        if app_id:                                           
            
            from polls.models import ReportTmpl,ReportTmplDetail            
            rep_tmpl = ReportTmpl.objects.filter(
                code = tmpl_code,
                app_id = app_id,                
                )
                           
                
            rep_tmpl = rep_tmpl and rep_tmpl[0]
            if not rep_tmpl:
                return 'No report template found.'
                                                   
            pwidth = rep_tmpl.page_width
            pheight = rep_tmpl.page_height
            rep_details = ReportTmplDetail.objects.filter(rep_tmpl=rep_tmpl)
            num_reports = len(model_recs)
            #define report starting (x,y) and page/canvas size
            x = 1 * mm                
            y = pheight * num_reports * inch
            pagesize = (pwidth * inch, pheight * num_reports * inch)
            c = canvas.Canvas(file_path, pagesize=pagesize)
            c.setPageSize((pwidth * inch,pheight * num_reports * inch))
            fonttype = rep_tmpl.font_type
            c.setFont(fonttype,12)
            textarea_chunk_size = rep_tmpl.textarea_chunk_size
            text_chunk_size = rep_tmpl.text_chunk_size
    
            for rec in model_recs:
                
                for elbox in rep_details:
                    y_val = 0
                    x_pixels = elbox.xcoord
                    y_pixels = elbox.ycoord
                    x2_pixels = elbox.x2coord
                    y2_pixels = elbox.y2coord
                    dpi = 150
                    fonttype = rep_tmpl.font_type
                    c.setFont(fonttype,12)
                    #convert x and y + height/width from px to mm
                    x_coord = x_pixels * 25.4 / dpi
                    y_coord = y_pixels * 25.4 / dpi
                    y_coord = pheight * 25.4 - y_coord
                    x2_coord = x2_pixels * 25.4 / dpi
                    y2_coord = y2_pixels * 25.4 / dpi
                    y2_coord = pheight * 25.4 - y2_coord
                    width = elbox.width
                    width = int(width)
                    height = elbox.height
                    height = int(height)
                    db_field = elbox.db_field
                    db_table = elbox.db_table
                    fixed_text = elbox.fixed_text
                    fontsize = elbox.font_size
                    img_path = elbox.img_path
                    rgb_red = elbox.rgb_red / 256
                    rgb_blue = elbox.rgb_blue / 256
                    rgb_green = elbox.rgb_green / 256
                    box_color = Color(rgb_red,rgb_green,rgb_blue)
                    #font decorators
                    font_bold = elbox.font_bold
                    font_udrl = elbox.font_udrl
                    font_ital = elbox.font_ital
                    
                    if elbox.data_type == 'checkbox':
                        
                        x_coord = x_coord + x_delta
                        
                        mm_width = 17 * 25.4 / dpi
                        mm_height = 17 * 25.4 / dpi
                        delta = 3
                        yval = (y_coord - mm_height + delta)

                        if mm_width and mm_height:
                            c.rect(x_coord * mm, yval * mm,(mm_width  + delta/10) * mm,(mm_height  + delta/10) * mm,fill=0, stroke=1)

                    if elbox.data_type in ['box','image'] and not img_path:
                        
                        x_coord = x_coord + x_delta
                        mm_width = width * 25.4 / dpi
                        mm_height = height * 25.4 / dpi
                        delta = 0
                        yval = (y_coord - mm_height + delta)

                        if rgb_red > 0.00 or rgb_blue > 0.00 or rgb_green > 0.00:
                            c.setFillColorRGB(rgb_red,rgb_green,rgb_blue)
                            
                        if mm_width and mm_height:
                            c.rect(x_coord * mm, yval * mm,(mm_width  + delta/10) * mm,(mm_height  + delta/10) * mm, fill=1, stroke=1)

                    if elbox.data_type in ['line']:
                            
                        delta = 3.5
                        dpi = 150
                        #convert x and y + height/width from px to mm
                        x1_coord = x_pixels * 25.4 / dpi
                        y1_coord = y_pixels * 25.4 / dpi
                        y1_coord = pheight * 25.4 - y1_coord + delta
                        x2_coord = x2_pixels * 25.4 / dpi
                        y2_coord = y2_pixels * 25.4 / dpi
                        y2_coord = pheight * 25.4 - y2_coord + delta

                        if rgb_red > 0.00 or rgb_blue > 0.00 or rgb_green > 0.00:
                            c.setFillColorRGB(rgb_red,rgb_green,rgb_blue)
                        
                        c.line((x1_coord - delta)* mm, y1_coord * mm, (x2_coord - delta)* mm, y2_coord * mm)
                    
                    if elbox.data_type in ['text','textarea']:
                        
                        x_coord = x_coord + x_delta
              
                        if font_bold or fontsize:
                            
                            if font_bold:
                                if not font_ital:
                                    fonttype += '-Bold'
                                  
                                else:
                                    fonttype += '-BoldItalic'
                            
                            elif font_ital:
                                fonttype += '-Italic'
                                                                
                            if font_udrl:
                                underline = True
                                
                            c.setFont(fonttype,fontsize)
                      
                        if rgb_red > 0.00 or rgb_blue > 0.00 or rgb_green > 0.00:
                            c.setFillColorRGB(rgb_red,rgb_green,rgb_blue)

                        db_val = db_field and getattr(rec, db_field)
                        line_height = 2
                        yval = y_coord
                        decrem = line_height * fontsize * 25.4 / dpi
                        text_lines = fixed_text and fixed_text.split('\n') or []
                        num_lines = len(text_lines)
                        textobject = c.beginText()
                        
                        if db_val:
                            
                            line_text = str(db_val)                         
                            
                            if elbox.data_type not in ['textarea']:
                                line_text = line_text[:text_chunk_size] 
                                textobject.setTextOrigin(x_coord * mm, yval * mm,)
                                textobject.textLine(line_text)
                                c.drawText(textobject)
                                yval = yval - decrem
                                    
                            else:
                                chunk_size = textarea_chunk_size
                                import textwrap
                                chunkies = textwrap.wrap(line_text, textarea_chunk_size)
                                
                                for chk in chunkies:
                                    
                                    textobject.setTextOrigin(x_coord * mm, yval * mm,)
                                    textobject.textLine(chk)
                                    c.drawText(textobject)
                                    yval = yval - decrem
                            
                        else:   
                            
                            for line in text_lines:
                                                    
                                line_text = str(line)            
                                #c.drawString(x_coord * mm, yval * mm, str(line))
                                textobject.setTextOrigin(x_coord * mm, yval * mm,)
                                textobject.textLine(line_text)
                                c.drawText(textobject)
                                #c.setFillColor(black)
                                yval = yval - decrem
                        
                        c.setFillColor(black)
                        
                    c.setFont(fonttype,fontsize)
                                                    
                    if elbox.data_type == 'barcode':
                         
                        x_coord = x_coord + x_delta
                        #create the barcode element
                        #barcode_text = fixed_text
                        
                        #if db_field:
                        #    barcode_text = getattr(rec, db_field)
                            
                        #barcode_text = barcode_text and str(barcode_text) or fixed_text
                        barcode_text = barcode_txt
                        mm_height = height * 25.4 / dpi
                        yval = y_coord - mm_height
                        height = height / dpi
                        
                        mm_width = width * 25.4 / dpi
                        #(11 dots * 25.4 mm/inch / 150 dots/inch) = mm_width
                        sfactor = 11 * 25.4 / dpi
                        #1pt = 1/72 inch = 0.3528 mm
                        pt_mm_factor = 0.3528
                        width = mm_width / (sfactor * len(barcode_text) + 40) * pt_mm_factor * mm
                        #width = width / dpi / 2 #width in px divided by the dpi constant
                        barcode39Std = code39.Extended39(barcode_text,\
                            barHeight=height * inch, barWidth=width, stop=1, checksum=0) 
                        code = barcode39Std
                        code.drawOn(c, (x_coord -15) * mm, yval * mm)

                    if elbox.data_type == 'image' and img_path:
                        
                        x_coord = x_coord + x_delta
                        mm_height = height * 25.4 / dpi
                        mm_width = width * 25.4 / dpi
                        y_val = y_coord - mm_height
                        
                        c.drawImage(img_path, x_coord * mm, y_val * mm,\
                            width = mm_width * mm,height = mm_height * mm, mask='auto')
                        
                y -= pheight * mm
                
            error = c.save() 
    return error      
    
#++++++++++++===================User Input Code============================++++++++++++++++++
def user_input(request,quapi_id=None):

    filter_user_id,active_app,msg,error,user_name,show_grid = '','','','','',False
    val_dict,results,req_post,form,date_from,date_to,options_col,session_id = {},[],{},{},'','',[],''
    sel_status,aud_status,page_size = False,False,25
    right_now = datetime.now()
    date_to_input = right_now + timedelta(days=1)
    date_to_input = date_to_input.strftime('%m/%d/%Y')
    date_from_input = right_now - timedelta(days=45)
    date_from_input = date_from_input.strftime('%m/%d/%Y')
    from polls.models import UserInput as Uinp,StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user) or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    audit_apps = user_apps.filter(audit_ok = True)
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        format = '%Y-%m-%d %H:%M:%S'
        new_format = '%m/%d/%Y'
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        filter_user_id = 'user_id' in req_post and req_post['user_id'] or None
        date_from_input = 'date_from' in req_post and req_post['date_from'] or None
        date_from = date_from_input and datetime.strptime(date_from_input, new_format) or None
        date_from = date_from and datetime.strftime(date_from, format) or None
        date_to_input = 'date_to' in req_post and req_post['date_to'] or None
        date_to = date_to_input and datetime.strptime(date_to_input, new_format) or None
        date_to = date_to and datetime.strftime(date_to, format) or None
        active_app = 'app_selector' in req_post and req_post['app_selector'] or None
        user_name = req_post.get('user_name','')
        user_id = filter_user_id and QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=filter_user_id) or None
        user_id = user_id and user_id[0] or None
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        user_name = user_id and user_id.user_name or ''
        show_grid = True
        #aud_status = 'status' in req_post and req_post['status'] or []
        #options_col,page_size = get_options(req_post,session_id)            
        #if not user_id:
        #    error = 'Invalid user.'
        if not active_app:
            error = 'Please enter an application.' 
        if not (date_from and date_to):
            error = 'Please enter a date from and date to.'        
    else:
        form = WODashboardForm(initial={'date_from': date_from_input, 'date_to': date_to_input})
    app_id = active_app and UserAppPerms.objects.filter(id=active_app) or None
    app_id = app_id and app_id[0] and app_id[0].ml_apps_id

    if app_id and app_id.id == 403:
        app_id = MLApps.objects.filter(code__icontains='Labor')
        app_id = [app.id for app in app_id]
  
    if req_post and app_id:  
               
        if user_name:
            if date_from and date_to:
                results = Uinp.objects.filter(ml_apps_id=app_id,user_name__iexact=user_name,timestamp__gte=date_from,timestamp__lte=date_to) or []  
            elif date_from and not date_to:
                results = Uinp.objects.filter(ml_apps_id=app_id,user_name__iexact=user_name,timestamp__gte=date_from) or []           
            elif not date_from and date_to: 
                results = Uinp.objects.filter(ml_apps_id=app_id,user_name__iexact=user_name,timestamp__lte=date_to) or [] 
            else: 
                results = Uinp.objects.filter(ml_apps_id=app_id,user_id__iexact=filter_user_id) or []             
        elif date_from and not date_to:
            results = Uinp.objects.filter(ml_apps_id=app_id,timestamp__gte=date_from) or []           
        elif not date_from and date_to: 
            results = Uinp.objects.filter(ml_apps_id=app_id,timestamp__lte=date_to) or [] 
        elif date_from and date_to:
            results = Uinp.objects.filter(ml_apps_id=app_id,timestamp__gte=date_from,timestamp__lte=date_to) or [] 
        else: 
            results = Uinp.objects.filter(quapi_id=quapi_id,ml_apps_id__in=app_id) or []               
    if not results and req_post and app_id:
        error = 'No user input records match your search.'
    val_dict.update({
        'app_set': user_apps,
        'audit_apps': audit_apps,
        'user_name': user_name,
        'user_id': filter_user_id,
        'user': user,
        'msg': msg,
        'error': error,
        'date_from': date_from_input,
        'date_to': date_to_input,
        'active_app': active_app and int(active_app) or None,
        'msg': msg,
        'quapi_id': quapi_id,
        'show_grid': show_grid, 
        'total_rows': len(results),
        'form': form,
        #'aud_status': aud_status,
        'user': user,
        'options_col': options_col,
        'page_size': page_size,
        'session_id': session_id,
        })
    return render(request, 'mrolive/user_input.html', val_dict) 

def event_notifications(request,quapi_id):
    error,msg,set_error,save_error = '','','',''
    total_rows = 0
    user = request and request.user or None
    val_dict = {}
    val_dict['quapi_id'] = quapi_id
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    #if request.method == 'GET':
    form = WODashboardForm()
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='event-manager')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)

    val_dict['mail_groups'] = MailGroup.objects.all()
    #look up all users and return them in val_dict
    from django.contrib.auth import get_user_model
    User = get_user_model()
    val_dict['users'] = User.objects.all()   
    val_dict['apps'] = MLApps.objects.all()
    
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        name = req_post.get('name','')
        wo_number = req_post.get('wo_number','')
        pn = req_post.get('pn','')
        task = req_post.get('task','')
        emp_name = req_post.get('emp_name','')
        cond_code = req_post.get('cond_code','')
        quantity = req_post.get('quantity','')
        user_id = req_post.get('user_id','')
        mail_group_id = req_post.get('mail_group_id','')
        
        val_dict.update({
            'wo_number': wo_number,
            'pn': pn,
            'cond_code': cond_code,
            'task': task,
            'sel_user': user_id,
            'sel_group': mail_group_id,            
        })
        #search for event_notifications record
        # Here you list all your filter names
        filter_names = (
            'wo_number',\
            'pn','task',\
            'emp_name',\
            'cond_code',\
            'user_id',\
            'mail_group_id',
        )
        filter_dict={}
        records = EventNotification.objects.all()
        
        for param in filter_names:
            if param in req_post and req_post[param]:
                filter_dict[param] = req_post[param]  
               
        queryset = records.filter(**filter_dict)
        total_rows = len(queryset)
        
        if not total_rows:
            error = 'No records found.'
        
    val_dict.update({
        'total_rows': total_rows,
        'error': error,
        'user': user,
        'form': form,
        'msg': msg,
    }) 
    
    return render(request, 'registration/event_notifications.html', val_dict)  

def mail_groups(request,quapi_id):
    error,msg,set_error,save_error = '','','',''
    user = request and request.user or None
    val_dict = {}
    val_dict['quapi_id'] = quapi_id
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    #if request.method == 'GET':
    form = WODashboardForm() 
    #look up all users and return them in val_dict for the grid
    from django.contrib.auth import get_user_model
    User = get_user_model()
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='event-manager')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    #mail_groups = MailGroup.objects.filter(active=True)
    mail_groups = MailGroup.objects.all()
    total_rows = len(mail_groups)
    val_dict['total_rows'] = total_rows
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        name = req_post.get('name','')
        to_emails = req_post.get('to_emails','')
        from_email = req_post.get('from_email','')
        cc_field = req_post.get('cc_field','')
        text_fields = req_post.get('text_fields','')
        is_email = req_post.get('is_send_email','')
        is_deactivate = req_post.get('is_deactivate','')
        sel_groups = []
        if 'sel_groups[]' in req_post:
            sel_groups = req_post.getlist('sel_groups[]')
            
        if sel_groups:
        
            #if button = 'deactivate
            if is_deactivate == '1':
            
                groups = MailGroup.objects.filter(id__in = sel_groups)
                for grp in groups:
                    activate = 'green-check.png'
                    if grp.active==activate:
                        activate = 'blank.png'
                    grp.active=activate
                    grp.save()

        else:
            sel_groups = []
            
        if is_email == '1':
                #from portal.tasks import send_event_notes
                error,msg = send_event_notes(sel_groups)
        else:
            #create new event manager record
            
            new_mg = MailGroup.objects.create(
                to_emails=to_emails,
                from_email=from_email,
                cc_field=cc_field,
                text_fields=text_fields,
            )
            error = new_mg.save()
            msg = 'Successfully added mail group.'
		   
    val_dict.update({
        'error': error,
        'user': user,
        'form': form,
        'msg': msg,
    })  
    return render(request, 'registration/mail_groups.html', val_dict)
       
def send_event_notes(sel_groups=[]):
    error,msg = '',''
    from django.conf import settings
    from django.core.mail import EmailMessage
    #from django.core.mail import send_mail
    import ast
    email = None
    
    mail_groups =  MailGroup.objects.all()
    if sel_groups:
        mail_groups = mail_groups.filter(id__in=sel_groups)        
    
    events =  EventManager.objects.all()    
    today = datetime.now()
    today = today.strftime('%Y-%m-%d %H:%M:%S')
    
    #loop over all events        
    for evtm in events:
        #get all unsent emails that are in this event
        emails = MailMail.objects.filter(
            status__in=['draft','failed',''],
            event_note_id__event_id=evtm,
        )
        
        body,body_html = '',''
        count = 1
        
        #loop over each email not in 'Success' state
        for email in emails:

            header = '\r\n New %s Notification: %s.............................'%(count,email.subject)
            body = header + '................................................................\r\n'
            body += email.body
           
            if email.body_html:
            
                body_html += """
                <table>
                  <tr>
                    <td>%s
                    </td>
                  </tr>
                </table>
                %s
                """%(header,email.body_html)
                
                count += 1
                
        if email and body_html:
            recipient_list = email.to_emails.split(",")
            bcc_list = email.cc_field.split(",")
            reply_to_list = email.from_email.split(",")

            mail = EmailMessage(
                email.subject,
                body_html,
                email.from_email,
                recipient_list,
                bcc_list,
                reply_to=reply_to_list,
            )
            mail.content_subtype = 'html'
            mail.send()
            msg = 'Message(s) sent!'

            for email in emails:
                email.status = 'sent'
                email.date_sent = today
                email.save()

    return error,msg                                    
      
def event_manager(request,quapi_id):
    error,msg,set_error,save_error = '','','',''
    user = request and request.user or None
    val_dict = {}
    val_dict['quapi_id'] = quapi_id
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    #if request.method == 'GET':
    form = WODashboardForm() 
    #look up all users and return them in val_dict for the grid
    from django.contrib.auth import get_user_model
    User = get_user_model()
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='event-manager')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)

    val_dict['mail_groups'] = MailGroup.objects.all()
    val_dict['apps'] = MLApps.objects.all()
    
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        name = req_post.get('name','')
        trigger_sql = req_post.get('trigger_sql','')
        quantum_table = req_post.get('quantum_table','')
        ml_apps_id = req_post.get('ml_apps_id','')
        mail_group_id = req_post.get('mail_group_id','')
        
        #create new event manager record
        from polls.models import EventManager as em,MailGroup as mg
        if ml_apps_id and name and mail_group_id:
            
            ml_apps = MLApps.objects.filter(id=int(ml_apps_id))
            ml_apps = ml_apps and ml_apps[0]
            mail_group = mg.objects.filter(id=int(mail_group_id))
            mail_group = mail_group and mail_group[0]
            new_em = em.objects.create(
                name=name,
                trigger_sql=trigger_sql,
                quantum_table=quantum_table,
                ml_apps_id=ml_apps,
                mail_group_id=mail_group,
            )
            error = new_em.save()
            msg = 'Successfully added event.'
        
    val_dict.update({
        'error': error,
        'user': user,
        'form': form,
        'msg': msg,
    })  
    return render(request, 'registration/event_manager.html', val_dict)  

def stock_lookup(request,quapi_id,wo_number='',part_number=''):
    error,msg='',''
    val_dict = {}
    user = request and request.user or None
    username = user and user.username or ''
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict) 
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='stock-lookup')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    if request.method == 'GET':
        form = WODashboardForm()
        val_dict['part_number'] = part_number
        val_dict['wo_number'] = wo_number
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        stock_search = req_post.get('stock_search','')  
        stock_print = req_post.get('stock_print','') 
        label = req_post.get('label','')  
        wo_number = req_post.get('wo_number','')  
        part_number = req_post.get('part_number','')          
        user_name = req_post.get('user_name','')                
        session_id = req_post.get('session_id','')
        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','')
        stm_sels = []
        if 'woos_list[]' in req_post:
            stm_sels = req_post.getlist('woos_list[]')            
            
        if stock_search == '1':
            stm_auto_key = req_post.get('label','')
            if stm_auto_key and stm_auto_key[0] in ['c','C']:
                stm_auto_key = stm_auto_key[1:]
            parameters = [stm_auto_key,wo_number,part_number,user_name]
            from portal.tasks import get_stock_lookup
            res = get_stock_lookup.delay(quapi_id,session_id,parameters)            
            error,msg = res.get()
            stms = WOStatus.objects.filter(session_id = session_id)            
            val_dict.update({
                'total_rows': len(stms),
                'session_id': session_id,
                'error': error,
            }) 
            
        elif stock_print == '1':            
            stms = WOStatus.objects.filter(session_id = session_id,\
                stm_auto_key__in = stm_sels)

            printed_date = datetime.now()
            printed_date = printed_date.strftime('%m/%d/%Y %H:%M:%S') 
            val_dict.update({
                'stock_recs': stms,
                'printed_by': username,
                'printed_date': printed_date,
                'lookup': '1',
                })

            printset = app_allow and app_allow[0] and app_allow[0].printset_id
            auth_key = printset and printset.printnode_auth_key
            if printset and auth_key:
                error = create_barcodes(val_dict) 
                if not error:                    
                    error = printnode_pdf(printset,auth_key)
                    val_dict['error'] = error
                    if not error:
                        val_dict['msg'] = 'Barcode label printed.' 
            else:
                return render(request, 'mrolive/plain_barcode_mro.html', val_dict)
        
    val_dict['form'] = form
    return render(request, 'mrolive/stock_lookup.html', val_dict)

def task_management(request,quapi_id):
    error,msg,tasks,task_list = '','','',[]
    user = request and request.user or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')    
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict) 
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='task-management')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')
    task_statuses = StatusSelection.objects.filter(session_id='1234567').order_by('name')  
    val_dict['task_statuses'] = task_statuses    
    if request.method == 'GET':
        form = WODashboardForm()
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        wo_number = req_post.get('wo_number','')
        filter_wo = req_post.get('filter_wo','')
        si_number = req_post.get('si_number','')
        status = req_post.get('task_status','')
        status_type = req_post.get('status_type','')
        action_taken = req_post.get('action_taken','')
        work_required = req_post.get('work_required','')
        description = req_post.get('description','')
        sequence = req_post.get('sequence','')                 
        sel_rows = req_post.get('sel_rows','') or 0
        is_search = req_post.get('is_search','0')  
        is_update = req_post.get('is_update','0') 
        launch_update = req_post.get('launch_update','0')         
        session_id = 'session_id' in req_post and req_post['session_id'] or '' 
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        from portal.tasks import get_task_statuses
        res = get_task_statuses.delay(quapi_id,session_id)
        res.get()      
        task_statuses = StatusSelection.objects.filter(session_id=session_id,is_dashboard = False)
        status_types = StatusSelection.objects.filter(session_id=session_id,is_dashboard = True)
    
        val_dict.update({
            #'sequence': sequence,
            #'status': status,
            #'description': description,             
            #'sequence': sequence,            
            #'action_taken': action_taken,
            #'work_required': work_required,
            'session_id': session_id,
            'wo_number': wo_number,
            'si_number': si_number,
            'sel_rows': sel_rows,
            'form': form,
            'session_id': session_id,
            'task_statuses': task_statuses,
            'status_types': status_types,
            })  

        if 'task_sels[]' in req_post:
            wot_list = req_post.getlist('task_sels[]')
            
        elif 'wots_list[]' in req_post:
            wot_list = req_post.getlist('wots_list[]') 
            
        if is_search == '1': 
            if not wo_number:
                val_dict['error'] = 'Must input a WO#.'
                return render(request, 'mrolive/task_management.html', val_dict)   
            from portal.tasks import get_task_mgmt
            parameters=[status,sequence,description,work_required,action_taken,status_type,wo_number]
            res = get_task_mgmt.delay(quapi_id,session_id,parameters)            
            error,msg = res.get()  
            tasks = WOTask.objects.filter(session_id = session_id)
            if not tasks:
                error = 'No tasks found.'
            val_dict.update({
                'tasks': tasks,
                'total_rows': len(tasks),
                'error': error,
                'task_statuses': task_statuses,                
            })
        if is_update == '1' or launch_update == '1':   
            #   if status or status_type or sequence or description or work_required or action_taken:
            #if task_selector:
            #    wo_task = task_selector.split("-",1)
            #    wo_task = wo_task and wo_task[1] or ''
            #    wo_task = wo_task and wo_task.replace(" ","",1) or ''
            status_msg = ''
            parameters=[status,sequence,description,work_required,action_taken,status_type,'']
            from portal.tasks import get_task_mgmt,update_task_mgmt
            
            if wot_list:
                if is_update == '1':
                    if any(p.strip() for p in parameters):
                        res = update_task_mgmt.delay(quapi_id,session_id,\
                           sysur_auto_key,parameters,wot_list)
                        error,status_msg = res.get()
                    parameters = ['','','','','','','']                    
                    res = get_task_mgmt.delay(quapi_id,session_id,\
                        parameters,wot_list=wot_list)                     
                    task_error,msg = res.get()                    
                    tasks = WOTask.objects.filter(session_id = session_id,\
                        wot_auto_key__in=wot_list)
                    task = tasks and tasks[0]
                    esn = task and task.esn or ''
                    customer = task and task.customer or ''
                    eng_model = task and task.eng_model or ''
                    ac_reg = task and task.ac_reg or '' 
                    ac_model = task and task.ac_model or '' 
                    ac_sn = task and task.ac_sn or ''
                    part_description = task and task.part_description or ''
                    val_dict.update({
                        'wo_number': wo_number or filter_wo,
                        'si_number': wo_number or filter_wo,
                        'esn': esn,
                        'description': part_description,
                        'customer': customer,
                        'eng_model': eng_model, 
                        'ac_reg': ac_reg, 
                        'ac_model': ac_model, 
                        'ac_sn': ac_sn,
                        'task_list': tasks,
                        'total_rows': len(tasks),
                        'error': error,
                        'msg': msg or status_msg, 
                    })
                    if error:
                        return render(request, 'mrolive/task_management.html', val_dict)                      
                    return render(request, 'mrolive/task_mgmt_traveller_mtu.html', val_dict)
                    
                elif launch_update == '1':
                    #launch update pop-up
                    val_dict['launch_update'] = 'T'
                    val_dict['wot_list'] = wot_list
                    return render(request, 'mrolive/task_management.html', val_dict)
            if 'wots_list[]' not in req_post and is_search != '1':
                error = 'Select at least one row to update.' 
                
    val_dict['form'] = form
    return render(request, 'mrolive/task_management.html', val_dict)
    
def user_create(quapi_id,user_id,sysur_auto_key,req_files,session_id):
    import_file,error,msg,fail_msg = None,'','','' 
    up_file = 'loc_whs_file' in req_files or None
    if up_file:
        up_file = req_files['loc_whs_file']
        file_name = up_file.name
        file_name = file_name.replace(" ","")
        up_file.name = file_name
    if up_file:
        import_file = Document(session_id=session_id,docfile=up_file)           
    if import_file:
        import_file.save()
        from portal.tasks import user_create
        res = user_create.delay(quapi_id,user_id,sysur_auto_key,session_id)
        error,msg,fail_msg = res.get()
    return error,msg,fail_msg
    
def user_import(request,quapi_id):
    error,msg,val_dict = '','',{}
    template_path = 'mrolive/user_import.html'
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='user-import')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':
        form = WODashboardForm()          
    if request.method == 'POST':
        req_post = request.POST
        import_file = request.FILES
        form = WODashboardForm(req_post)
        session_id=req_post.get('session_id')
        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','')       
        val_dict.update({
            'msg': msg,
            'user_name': user_name,            
            'session_id': session_id,                    
        })
        if request.FILES: 
            error,msg,fail_msg = user_create(quapi_id,user_id,sysur_auto_key,import_file,session_id)
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, template_path, val_dict)
    
def labor_mgmt_detail(request,user_name='',wo_number='',conn_key=0,session_id=''): 
    template_path = 'mrolive/labor_mgmt_detail.html'
    val_dict,wtl_list={},[]
    total_rows,total_hours,error,msg,active_mode,dash_error = 0,0,'','','',''
    session_id,error,msg,date_from,date_to,user_id,wo_number = '','','','','','',''
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    reg_user_id = user and user.is_authenticated and user.id or None
    if not reg_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')
    val_dict['user_id'] = user and user.username or ''
    val_dict['quapi_id'] = conn_key

    if request.method == 'GET':
        
        from portal.tasks import labor_dashboard
        if wo_number:
            res = labor_dashboard.delay(conn_key,session_id,wo_number=wo_number,is_detail=True)
            labor_recs = TaskLabor.objects.filter(session_id = session_id,wo_number=wo_number)
        elif user_name:
            res = labor_dashboard.delay(conn_key,session_id,user_id=user_name,is_detail=True)
            labor_recs = TaskLabor.objects.filter(session_id = session_id,user_name=user_name)
        error,msg,total_hours = res.get()
        
        val_dict['total_rows'] = len(labor_recs)
        val_dict['error'] = error
    val_dict['session_id'] = session_id
    val_dict['user_name'] = user_name
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    alloc_app = MLApps.objects.filter(code="labor-management")
    alloc_app = alloc_app and alloc_app[0] or None
    modes = alloc_app and get_modes(alloc_app) or []
    if not modes:
        return redirect('/login/')
        
    val_dict['modes'] = modes
    
    if request.method == 'POST':
        #this is the submit to update/add/edit users labor entries
        req_post = request.POST
        form = WODashboardForm(req_post)
        lsearch = req_post.get('labor_search','')
        lupdate = req_post.get('labor_update','')
        update_user = req_post.get('update_user','')
        wo_number = req_post.get('wo_number','')
        user_id = req_post.get('user_id','')
        date_start = req_post.get('date_start','')
        date_stop = req_post.get('date_stop','')
        active_mode = req_post.get('active_mode','')
        mod_type = req_post.get('mode_selector',active_mode)
        update_submitted = req_post.get('update_submitted','')
        is_update = req_post.get('is_update','0')
        update_start = req_post.get('update_start','')
        update_stop = req_post.get('update_stop','')
        update_skill= req_post.get('update_skill','')
        filter_user = req_post.get('filter_user','')
        filter_stop = req_post.get('filter_stop','')
        filter_start = req_post.get('filter_skill','')
        wot_auto_key = req_post.get('update_task','')
        session_id = req_post.get('session_id',session_id)
        update_session = req_post.get('update_session',session_id)
        if not (session_id or update_session):
            session_id = req_post.get('csrfmiddlewaretoken','')
        val_dict['session_id'] = session_id or update_session
        val_dict['user_name'] = user_id or user_name or ''
        if 'wtl_list[]' in req_post:
            wtl_list = req_post.getlist('wtl_list[]')  
        elif 'wtl_sels[]' in req_post:
            wtl_list = req_post.getlist('wtl_sels[]')            
  
        if lsearch:
            from portal.tasks import labor_dashboard
            res = labor_dashboard.delay(conn_key,
                session_id,user_id=user_id,wo_number=wo_number,
                is_detail=True,date_from=date_start,date_to=date_stop)
            error,msg,total_hours = res.get()

        elif lupdate or update_submitted or is_update=='1':      
            if mod_type.upper() in ['ADD','EDIT']:
                if not update_submitted:
                    if mod_type.upper()=='EDIT':
                    
                        val_dict.update({'labor_count':len(wtl_list)})
                        if not wtl_list:
                            error = 'No rows selected.'
                            
                        from portal.tasks import get_wo_skills
                        res = get_wo_skills.delay(conn_key,session_id)
                        error,msg = res.get()
                        from polls.models import TaskSkills as ts
                        skills = ts.objects.all()
                        val_dict['skill_list'] = skills     
                        
                    if not error:
                        if wtl_list:
                            sel_recs = TaskLabor.objects.filter(\
                                session_id=session_id,\
                                wtl_auto_key=wtl_list[0]) 
                            rec = sel_recs and sel_recs[0]
                                
                            if len(wtl_list) == 1:
                                #start_time = rec.start_time\
                                    #and datetime.strftime(rec.start_time,'%m/%d/%Y %H:%M:%S')
                                start_time = rec.start_time\
                                    and rec.start_time - timedelta(hours=5) 
                                start_time = start_time and\
                                    datetime.strftime(start_time,'%m/%d/%Y %I:%M %p') or ''
                                stop_time = rec.stop_time\
                                    and rec.stop_time - timedelta(hours=5) 
                                stop_time = stop_time and\
                                    datetime.strftime(stop_time,'%m/%d/%Y %I:%M %p') or ''
                                #stop_time = rec.stop_time\
                                    #and datetime.strftime(rec.stop_time,'%m/%d/%Y %H:%M:%S')
                                val_dict.update({
                                    'date_start': start_time,
                                    'date_stop': stop_time,
                                    'task_user': rec.user_name,
                                    })
                            
                        val_dict.update({
                            'show_modal': 'T',
                            'active_mode': mod_type.upper(),
                            'wtl_list': wtl_list,
                            'user_id': user_id,
                        })                       
                                                    
                elif wot_auto_key or update_user or update_start or update_stop or update_skill:
                    if mod_type.upper()=='EDIT':
                        if not wtl_list:
                            error = 'No rows selected.'
                    else:
                        val_dict['active_mode'] = 'ADD'
                      
                    if not error:
                        from portal.tasks import labor_modify
                        res = labor_modify.delay(conn_key,\
                        session_id,sysur_auto_key,wtl_list,\
                        mod_type,user_change=update_user,\
                        user_name=user_name,wot_auto_key=wot_auto_key,\
                        date_start=update_start,date_stop=update_stop,\
                        wo_skill=update_skill)
                        error,msg = res.get()
                        if not error and mod_type.upper()=='EDIT' or error:
                            from portal.tasks import labor_dashboard
                            res = labor_dashboard.delay(conn_key,\
                            session_id,user_id=user_name,\
                            date_from='',date_to='',is_detail=True)
                            dash_error,dash_msg,total_hours = res.get()
                        val_dict['wtl_sels'] = [int(x) for x in wtl_list]
                        
            if mod_type.upper()=='EDIT' and not wtl_list:
                error = 'No rows selected.'
                
        labor_recs = []
        if wtl_list and 'wtl_sels' in val_dict and mod_type.upper()=='EDIT':                
            labor_recs = TaskLabor.objects.filter(session_id=session_id,wtl_auto_key__in=wtl_list)
            
        elif mod_type.upper()=='ADD' and wot_auto_key:
            session_id = 'ay8nNoi80920KHOI:jgals82'
            labor_recs = TaskLabor.objects.filter(session_id=session_id)
           
        else:
            labor_recs = TaskLabor.objects.filter(session_id=session_id)
            
        total_rows = len(labor_recs)
        val_dict['total_rows'] = total_rows
        
    val_dict.update({
        'msg': msg,
        'error': error + dash_error,
        'wo_number': wo_number,
        'date_to':date_to,
        'date_from':date_from,
        'user_id':user_id,
        'session_id': session_id,
        'conn_key': conn_key,
    })
    
    #labor-management-detail 
    val_dict['form'] = WODashboardForm()
    return render(request, template_path, val_dict)

def labor_management(request,quapi_id=None):
    template_path = 'mrolive/labor_management.html'
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    session_id,error,msg,date_from,date_to,user_id = '','','','','',''
    wo_number = ''
    val_dict = {}
    #val_dict = {'date_to':'','date_from':'','user_id':'','quapi_id': quapi_id}
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    val_dict['sysur_auto_key'] = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    val_dict['user_name'] = user_name
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')        
    val_dict['emp_vals'] = quapi_id and QuantumUser.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id).distinct() or ''
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    user_permitted = user_apps.filter(ml_apps_id__code="labor-management")

    if not user_permitted:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')         
    if request.method == 'GET':
        session_id = 'nay2nHGad7ngjau4289h'
        val_dict['session_id'] = session_id
        form = WODashboardForm()
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        user_id = req_post.get('user_id','')
        date_from = req_post.get('date_from','')
        date_to = req_post.get('date_to','')
        session_id = req_post.get('session_id','')
        wo_number = req_post.get('wo_number','')
        
        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','')
        val_dict['session_id'] = session_id
        if wo_number:
            #portal/labor-mgmt-detail/#=user_name#/1/#=session_id#\">#=user_name#
            #from portal.tasks import labor_dashboard
            #res = labor_dashboard.delay(conn_key,
            #    session_id,wo_number=wo_number)
            #error,msg,total_hours = res.get()
            mgmt_detail_link = '/portal/labor-mgmt-detail/%s/%s/%s/%s'\
                %(user_name,quapi_id,wo_number,session_id)
            return redirect(mgmt_detail_link)
            
        user_rec = None
        if user_id:        
            user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id) or None
        if not user_rec and (user_id or user_name or user_logged):
            user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_name__iexact=user_id) or None
        if not user_rec and user_id:
            user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,employee_code__iexact=user_id) or None
        user_rec = user_rec and user_rec[0] or None
        if user_id and not user_rec:
            error = 'User not found.'
        user_key = user_rec and user_rec.user_auto_key or 0
        from portal.tasks import labor_dashboard
        res = labor_dashboard.delay(quapi_id,session_id,user_id=user_id,date_from=date_from,date_to=date_to,is_mgmt=True)
        error,msg,total_hours = res.get()
        labor_recs = TaskLabor.objects.filter(session_id = session_id)
        val_dict['total_rows'] = len(labor_recs)
        val_dict['total_hours'] = total_hours
        val_dict['user_name'] = user_id.upper()                               
    form = WODashboardForm(val_dict)
    val_dict['form'] = form 
    val_dict['user'] = user
    val_dict.update({
        'msg': msg,
        'error': error,
        'wo_number': wo_number,
        'date_to':date_to,
        'date_from':date_from,
        'user_id':user_id,
        'quapi_id': quapi_id,
        'session_id': session_id,
        'conn_key': quapi_id,
    })
    #labor-management
    return render(request, template_path, val_dict)   
   
def labor_dashboard(request,quapi_id=None):
    template_path = 'mrolive/labor_dashboard.html'
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    session_id,error,msg,date_from,date_to,user_id = '','','','','',''
    wo_number = ''
    val_dict,error,msg = {},'',''
    val_dict['quapi_id'] = quapi_id 
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    val_dict['user_name'] = user_name
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')        
    val_dict['emp_vals'] = quapi_id and QuantumUser.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id).distinct() or ''
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    user_permitted = user_apps.filter(ml_apps_id__code="labor-dashboard")
    if not user_permitted:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    if request.method == 'GET':
        #from portal.tasks import labor_dashboard
        #res = labor_dashboard.delay(quapi_id,\
        #    '25hau9agh')
        #error,msg,total_hours = res.get()
        form = WODashboardForm()
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        user_id = req_post.get('user_id','')
        date_from = req_post.get('date_from','')
        date_to = req_post.get('date_to','')
        wo_number = req_post.get('wo_number','')
        session_id = '25hau9agh'
        
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        val_dict['session_id'] = session_id      
        user_rec = None
        
        if user_id:        
            user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id) or None
        if not user_rec and (user_id or user_name or user_logged):
            user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_name__iexact=user_id) or None
        if not user_rec and user_id:
            user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,employee_code__iexact=user_id) or None
        user_rec = user_rec and user_rec[0] or None
        user_key = user_rec and user_rec.user_auto_key or 0
        
        if not date_to:
            val_dict['error'] = 'Must enter start date.'
                                                                              
        if date_from or date_to or wo_number or user_id:
            from portal.tasks import labor_dashboard
            res = labor_dashboard.delay(quapi_id,\
                session_id,user_id=user_id,\
                date_from=date_from,date_to=date_to,\
                is_dashboard=True,wo_number=wo_number)
            error,msg,total_hours = res.get()
            val_dict['total_hours'] = total_hours 
            
        labor_recs = TaskLabor.objects.filter(session_id = session_id)
        val_dict['total_labor_rows'] = len(labor_recs)                                                           
        val_dict['user_name'] = user_id
                                                 
    form = WODashboardForm(val_dict)
    val_dict['form'] = form 
    val_dict['user'] = user
    
    val_dict.update({
        'msg': msg,
        'error': error,
        'wo_number': wo_number,
        'date_to':date_to,
        'date_from':date_from,
        'user_id':user_id,
        'quapi_id': quapi_id,
        'session_id': session_id,
        'conn_key': quapi_id,
    })
    return render(request, template_path, val_dict)   
    
def consumables(request,quapi_id=None):     
    location,user_id,user_logged,update,warehouse,rack,new_rack,rerack,rack_user = '','','','','','','','',''
    wo_number,user_error,stat_error,loc_error = '','','',''
    val_dict,form,updated_woos,all_woos,woo_num_list,woo_key_list = {},{},[],[],[],[]           
    msg,loc_msg,stat_msg,error,lookup_recs,clear_cart = '','','','',False,False
    loc_key,whs_key,cart_key,new_status_name=None,None,None,''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    val_dict['quapi_id'] = quapi_id 
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    val_dict['user_name'] = user_name
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')        
    val_dict['emp_vals'] = quapi_id and QuantumUser.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id).distinct() or ''
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    alloc_app = MLApps.objects.filter(name="Consumables")
    alloc_app = alloc_app and alloc_app[0] or None
    modes = alloc_app and get_modes(alloc_app) or []
    if not modes:
        return redirect('/login/') 
    val_dict['modes'] = modes
    if request.method == 'GET':
        if not reg_user_id or not dj_user_id:
            val_dict['error'] = 'Access denied.'
            return redirect('/login/')  
        #from portal.tasks import get_users_nsync_beta
        form = WODashboardForm()   
        #res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=0,app='consumables')
        #user_error,app = res.get() 
        val_dict['active_mode'] = '3'
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        quantity = 'quantity' in req_post and req_post['quantity'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        if not dj_user_id:
            dj_user_id = 'dj_user_id' in req_post and req_post['dj_user_id'] or ''#dj admin user id
        user_id = 'user_id' in req_post and req_post['user_id'] or ''#sysur_auto_key
        user_logged = 'user_logged' in req_post and req_post['user_logged'] or ''
        rack_user = 'rack_user' in req_post and req_post['rack_user'] or '' 
        user_id = user_id or user_logged or rack_user or ''
        #lookup user_id in the database to make sure we can authenticate
        user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id)
        user_rec = user_rec and user_rec[0] or None
        clear_form = 'clear_form' in req_post and req_post['clear_form'] or False       
        lookup_recs = 'lookup_recs' in req_post and req_post['lookup_recs'] or False           
        location = 'location' in req_post and req_post['location'] or '' 
        rack = 'rack' in req_post and req_post['rack'] or '' 
        new_rack = 'new_rack' in req_post and req_post['new_rack'] or '' 
        warehouse = 'warehouse' in req_post and req_post['warehouse'] or ''
        active_mode = 'mode_selector' in req_post and req_post['mode_selector'] or ''
        sel_mode = 'sel_mode' in req_post and req_post['sel_mode'] or ''  
        if not (sel_mode or active_mode):
            val_dict['error'] = 'Must select a mode.'
            render(request, 'mrolive/stock_reserve.html', val_dict) 
        cart_code = 'cart_code' in req_post and req_post['cart_code'] or ''         
        new_status = 'new_status' in req_post and req_post['new_status'] or ''     
        show_status = 'show_status' in req_post and req_post['show_status'] or ''
        show_user = 'show_user' in req_post and req_post['show_user'] or ''
        show_all = 'show_all' in req_post and req_post['show_all'] or ''
        clear_cart = 'ccart_form' in req_post and True or False
        label = 'label' in req_post and req_post['label'] or ''
        wo_task = 'wo_task' in req_post and req_post['wo_task'] or ''
        do_status = sel_mode and (sel_mode == '2' or sel_mode == '1') or False
        do_user = sel_mode or False       
        do_all = user_id or user_logged or rack_user or False  
        options_col,page_size = get_options(req_post,session_id)            
        val_dict.update({
            'wo_number': '',
            'all_woos': updated_woos,
            'msg': msg,
            'warehouse': warehouse,
            'location': location,
            'dj_user_id': dj_user_id,
            'user_id': user_id or (user_rec and user_rec.user_id) or user_logged or rack_user,
            'user_name': user_name,
            'rack': rack or new_rack,
            'user_logged': user_logged or user_id,
            'rack_user': user_id,
            'new_rack': rack,
            'modes': modes,
            'active_mode': active_mode or sel_mode or '',
            'sel_mode': sel_mode or active_mode or '',
            'cart_code': cart_code or rack or '',
            'label': '',
            'new_status': new_status and int(new_status) or None,
            'lookup_recs': lookup_recs,
            'show_status': show_status or do_status,
            'do_status': do_status or show_status,
            'show_user': show_user or do_user,
            'do_user': do_status or show_status,
            'show_all': show_all and show_all!='0' or do_all,
            'do_all': do_all or show_all,
            'session_id': session_id,
            'sel_rows': sel_rows,
            'total_rows': total_rows,
            'quantity': '',
            'form': form,  
            'options_col': options_col,
            'page_size': page_size,             
            })  
       
        #if user submitted clear list form by pressing button
        if clear_form and req_post['clear_form']=='1':       
            WOStatus.objects.filter(user_id=user_logged,is_dashboard=0,active=1,is_racking=1).delete()
            val_dict['all_woos'] = []            
            val_dict['msg'] = '' 
            val_dict['active_mode'] = sel_mode
            val_dict['show_all'] = 1  
            form = WODashboardForm(val_dict)
            val_dict['form'] = form                         
            return render(request, 'mrolive/stock_reserve.html', val_dict)              
        wo_number = wo_number or label or ''
        ctrl_number,ctrl_id = '',''        
        if wo_number and len(wo_number) > 6:
            ctrl_number = wo_number[:6]               
            ctrl_id = wo_number[7:]
        #create the demo wostatus objec
        #TODO:
        #   1. Add a new field for task + s
        #   2. Lookup task
        #   3. Find the ctrl#/id and then add a new reservation to the task for the stock move?
        #   4. Find the lowest wob_auto_key bom from the pnm (from the stm entered by user) and then 
        """
            
        """
        app_mode = sel_mode or active_mode or ''
        
        if app_mode == '1':
            from portal.tasks import stock_reserve
            res = stock_reserve.delay(quapi_id,session_id,sysur_auto_key,user_id,wo_task,quantity,active_mode,ctrl_number,ctrl_id)
            error,msg,qty_res = res.get()
        elif app_mode == '3':
            qty_res = 'quantity' in req_post and req_post['quantity']
            must_reserve = not qty_res and 'T' or 'F'
            must_reserve = 'must_reserve' in req_post and req_post['must_reserve']
            from portal.tasks import issue_consumables
            res = issue_consumables.delay(quapi_id,session_id,sysur_auto_key,user_name,ctrl_number,ctrl_id,wo_task,quantity,app_mode,must_reserve)
            error,msg,qty_res,must_reserve = res.get()  
            #val_dict['must_reserve'] = qty_res or 'T'
            #if qty_res:
            val_dict['quantity'] = ''
            val_dict['label'] = ''
            val_dict['wo_task'] = wo_task
            val_dict['must_reserve'] = must_reserve
        elif app_mode == '4':
            from portal.tasks import stock_unissue
            res = stock_unissue.delay(quapi_id,session_id,sysur_auto_key,user_name,quantity,wo_task,ctrl_number,ctrl_id)
            error,msg,quantity = res.get()
            val_dict['quantity'] = '' 
            val_dict['label'] = ''
            val_dict['wo_task'] = wo_task            
        updated_woos = WOStatus.objects.filter(session_id=session_id)
        val_dict['all_woos'] = updated_woos
        val_dict['total_rows'] = str(len(updated_woos))
        val_dict['msg'] = msg   
        val_dict['error'] = error + user_error + stat_error + loc_error       
        if not wo_number and lookup_recs not in [1,'1']:
            val_dict['lookup_recs'] = 1
        elif wo_number and lookup_recs not in [0,'0']:
            val_dict['lookup_recs'] = 0            
    form = WODashboardForm(val_dict)
    val_dict['form'] = form  
    val_dict['user'] = user 
    val_dict['app_type'] = 'consumables'    
    return render(request, 'mrolive/stock_reserve.html', val_dict)
    
def shipping_dashboard(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec,fail_ms = 'user not set',None,''
    val_dict,form = {},{} 
    error,msg,loc_msg,stat_msg = '','','',''
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None    
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='smd-dashboard')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':
        form = WODashboardForm() 
    val_dict['modes'] = app_id and get_modes(app_id) or []        
    val_dict['status_vals'] = dj_user_id and StatusSelection.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id) or []        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        sm_number=req_post.get('wo_number','')
        so_number=req_post.get('so_number','')
        ro_number=req_post.get('ro_number','')
        po_number=req_post.get('po_number','')
        due_date=req_post.get('due_date','')        
        location=req_post.get('location','')              
        user_id = req_post.get('user_id','')
        session_id = req_post.get('session_id','')
        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','') 
            
        val_dict.update({
            'msg': msg,
            'user_id': user_id,
            'user_name': user_name,            
            'session_id': session_id, 
            'sel_rows': 0, 
            'location': location,
            'po_number': po_number,
            'ro_number': ro_number,
            'so_number': so_number, 
            'sm_number': sm_number,
            'due_date': due_date,            
        })
        
        sm_create = False
        if not sm_create:
            if sm_number or location or due_date or sm_number or so_number or ro_number or po_number:         
                filter_list = [location,due_date,so_number,ro_number,po_number,sm_number]
                from portal.tasks import search_shipping
                res = search_shipping.delay(quapi_id,session_id,sysur_auto_key,filter_list,is_dashboard=1)
                error,msg = res.get()
                all_woos = WOStatus.objects.filter(session_id=session_id)
                val_dict['all_woos'] = all_woos
                val_dict['total_rows'] = len(all_woos)
            else:
                error = "Must have selection in filter."            
        else:
            from portal.tasks import create_shipping
            res = create_shipping.delay(quapi_id,session_id,sysur_auto_key,filter_list)
            error,msg = res.get()
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/shipping_dashboard.html', val_dict)   

def shipping_mgmt(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec,fail_ms = 'user not set',None,''
    val_dict,form = {},{} 
    error,msg,loc_msg,stat_msg = '','','',''
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None    
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='smd-management')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    val_dict['quapi_id'] = quapi_id
    from portal.tasks import get_shipping_status,get_ship_vias,get_priorities
    from portal.tasks import get_users_nsync_beta
    session_id = 'anSI823(2$3%234MLK8'    
    res = get_ship_vias.delay(quapi_id,session_id,app='Shipping Mgmt')
    svia_error,app = res.get() 
    ship_vias = ShipVia.objects.filter(session_id=session_id) 
    res = get_priorities.delay(quapi_id,session_id,app='Shipping Mgmt')
    pri_error,app = res.get()  
    priorities = Priority.objects.filter(session_id=session_id) 
    res = get_shipping_status.delay(quapi_id,session_id,app='Shipping Mgmt')
    stat_error,app = res.get()  
    statuses = StatusSelection.objects.filter(session_id=session_id) 
    res = get_users_nsync_beta.delay(quapi_id,dj_user_id)
    users = QuantumUser.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id)
    
    val_dict.update({
        'status_vals':statuses,
        'users': users,
        'ship_via_codes':ship_vias,
        'priorities':priorities,
    })
    
    if request.method == 'GET':
        form = WODashboardForm()
        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post) 
        customer=req_post.get('customer','')
        order=req_post.get('order','')
        entry_date=req_post.get('entry_date','')
        status=req_post.get('status','')
        part_number=req_post.get('part_number','')
        description=req_post.get('description','')
        user_id=req_post.get('user_id','')
        ship_via=req_post.get('ship_via','')
        priority=req_post.get('priority','')
        location=req_post.get('location','')
        whs=req_post.get('warehouse','') 
        launch_update = req_post.get('launch_update','')
        is_search = req_post.get('search_stock','')
        is_update = req_post.get('is_update','')  
        user_update = req_post.get('user_update','')
        tote = req_post.get('tote','')
        notes = req_post.get('notes','')        
        session_id=req_post.get('session_id','')
        
        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','')
            
        val_dict.update({
            'msg': msg,
            'customer': customer,
            'order': order,
            'entry_date': entry_date,
            'status': status,
            'part_number': part_number,
            'user_id': user_id,
            'description': description,
            'ship_via': ship_via,
            'priority': priority,
            'location': location,
            'whs': whs, 
            'user_name': user_name,            
            'session_id': session_id, 
            'sel_rows': 0,            
        })

        filter_list = [customer,order,entry_date,status,part_number]
        filter_list += [description,ship_via,priority,location,whs,user_id]
        
        smds_list = []
        if 'smd_sels[]' in req_post:
            smds_list = req_post.getlist('smd_sels[]')
            
        elif 'woos_list[]' in req_post:
            smds_list = req_post.getlist('woos_list[]')  
            val_dict['smds_list'] = smds_list
            
        
        if is_search == '1':

            if any(f.strip() for f in filter_list):         
                
                from portal.tasks import search_shipping
                res = search_shipping.delay(quapi_id,session_id,\
                    sysur_auto_key,filter_list)
                error,msg = res.get()
                all_smds = WOStatus.objects.filter(session_id=session_id)
                val_dict['total_rows'] = len(all_smds)
                
            else:
                error = 'Please enter a value into at least one filter.'
                
        elif launch_update == '1':
            val_dict['session_id'] = session_id
            val_dict['launch_update'] = 'T'
            val_dict['smds_list'] = smds_list 
            all_smds = WOStatus.objects.filter(session_id=session_id)
            val_dict['total_rows'] = len(all_smds)   
            
        elif is_update == '1':
        
            if not smds_list:
                error = 'Select grid rows to update'
                return render(request, 'mrolive/shipping_edit.html', val_dict) 
       
            from portal.tasks import update_shipping,search_shipping
            res = update_shipping.delay(quapi_id,session_id,\
                sysur_auto_key,user_update,smds_list,tote,notes)
            error,msg = res.get()
            
            if not error:

                from portal.tasks import search_shipping
                res = search_shipping.delay(quapi_id,session_id,\
                    sysur_auto_key,filter_list=[],smds_list=smds_list)
                error,msg = res.get()
                
                updated_smds = WOStatus.objects.filter(
                    session_id=session_id,
                    )

                scan_time = datetime.now()
                scan_time = scan_time.strftime('%m/%d/%Y %H:%M:%S')
                
                val_dict.update({
                    'smds':updated_smds,
                    'user_id': user_update,
                    'scan_time': scan_time,
                })
                
                return render(request, 'mrolive/shipping_label.html', val_dict)                 
            
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/shipping_edit.html', val_dict)  
    
def shipping_edit(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec,fail_ms = 'user not set',None,''
    val_dict,form = {},{} 
    error,msg,loc_msg,stat_msg = '','','',''
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None    
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='smd-edit')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':
        form = WODashboardForm() 
        from portal.tasks import get_statuses_nsync_beta
        res = get_statuses_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=1,app='smd-management',object_type='SO')
        stat_error,app = res.get()
    val_dict['status_vals'] = dj_user_id and StatusSelection.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id) or []        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post) 
        customer=req_post.get('customer','')
        order=req_post.get('order','')
        due_date=req_post.get('due_date','')
        status=req_post.get('status','')
        part_number=req_post.get('part_number','')        
        active_user = 'active_user' in req_post and req_post['active_user'] or ''  
        filter_user = 'filter_user' in req_post and req_post['filter_user'] or ''        
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''        
        val_dict.update({
            'msg': msg,
            'active_user': active_user or user_id,
            'filter_user': filter_user or active_user or user_id,
            'user_id': user_id or filter_user or active_user,
            'user_name': user_name,            
            'session_id': session_id, 
            'sel_rows': 0,            
        })
        #Company	Cust. Order	Due Date	Status	Part Number	Description	Qty	Seial	Loc
        sm_create = False
        if not sm_create:
            if customer or order or due_date or status or part_number:         
                filter_list = [customer,order,due_date,status,part_number]
                from portal.tasks import search_shipping
                res = search_shipping.delay(quapi_id,session_id,sysur_auto_key,filter_list)
                                                                      
                error,msg = res.get()
                all_woos = WOStatus.objects.filter(session_id=session_id)
                val_dict['all_woos'] = all_woos
                val_dict['total_rows'] = len(all_woos)
                
        else:
            from portal.tasks import create_shipping
            res = create_shipping.delay(quapi_id,session_id,sysur_auto_key,filter_list)
            error,msg = res.get()
                                       
                                                                   
                
                                 
                                        
                                           
                                           
                  
                
                                                                                                                                                
            
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/shipping_edit.html', val_dict)                                                 

def lot_import(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec,fail_ms = 'user not set',None,''
    val_dict,form = {},{} 
    error,msg,loc_msg,stat_msg = '','','',''
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='lot-import')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':
        form = WODashboardForm()          
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post, request.FILES)    
        active_user = 'active_user' in req_post and req_post['active_user'] or ''  
        filter_user = 'filter_user' in req_post and req_post['filter_user'] or ''        
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''        
        val_dict.update({
            'msg': msg,
            'active_user': active_user or user_id,
            'filter_user': filter_user or active_user or user_id,
            'user_id': user_id or filter_user or active_user,
            'user_name': user_name,            
            'session_id': session_id,                    
        })

        if request.FILES:         
            lot_error,msg,fail_msg = lot_create(quapi_id,user_name,sysur_auto_key,request.FILES,session_id)
            val_dict['fail_msg'] = fail_msg
            grid_rows = WOStatus.objects.filter(session_id = session_id)
            val_dict['total_rows'] = 'F'
            if fail_msg:
                val_dict['total_rows'] = len(grid_rows)
            
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/lot_import.html', val_dict)  
  
def lot_create(quapi_id,user_id,sysur_auto_key,req_files,session_id):
    import_file,error,msg,fail_msg,file_name = None,'','','','' 
    up_file = 'loc_whs_file' in req_files or None
    if up_file:
        up_file = req_files['loc_whs_file']
        file_name = up_file.name
        file_name = file_name.replace(" ","")
        up_file.name = file_name
    if up_file:
        import_file = Document(file_name=file_name,session_id=session_id,docfile=up_file)           
    if import_file:
        import_file.save()
                                                                                  
        from portal.tasks import lot_create
        res = lot_create.apply_async(
            queue='import',
            priority=1,
            args=[quapi_id,user_id,sysur_auto_key,session_id,file_name],
            )
        error,msg,fail_msg = res.get()
    return error,msg,fail_msg
 
def lot_teardown(request,quapi_id=None):
    new_status,location,filter_status,show_msg,task_list = '','','','',[]
    user_rec,fail_ms,msg = None,'',''
    val_dict,form = {},{}
    total_rows = 0    
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')     
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict)
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='lot-teardown')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    val_dict['user_apps'] = user_apps
    val_dict['quapi_id'] = quapi_id  
    val_dict['user_name'] = user_name
    val_dict['lot_teardown'] = 't'
    session_id = 'Aioewnaisghaie2ri8lkjas'
    from portal.tasks import get_stock_status,get_conditions
    res = get_stock_status.delay(quapi_id,session_id)
    error = res.get()        
    val_dict['statuses'] = StatusSelection.objects.filter(session_id=session_id)   
    res = get_conditions.delay(quapi_id,session_id)
    error = res.get()
    from polls.models import PartConditions    
    val_dict['conditions'] = PartConditions.objects.filter(session_id=session_id)
    
    if request.method == 'GET':
        form = WODashboardForm()         

    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)   
        val_dict['form'] = form
        wo_task = req_post.get('wo_task','')
        quantity = req_post.get('quantity','')        
        serial_number = req_post.get('serial_number','')
        part_number = req_post.get('part_number','')
        format_type = req_post.get('type','')
        auto_click = req_post.get('auto_click','')
        notes = req_post.get('notes','')
        session_id = req_post.get('session_id','')
        show_msg = req_post.get('show_modal','')
        modal_pn = req_post.get('modal_pn','')
        modal_desc = req_post.get('modal_desc','')
        description = req_post.get('description','')
        stock_status = req_post.get('stock_status','')
        modal_serialized = req_post.get('modal_serialized','')
        upd_status = req_post.get('upd_status','')
        upd_notes = req_post.get('notes','')        
        upd_qty = req_post.get('upd_qty','')
        upd_cond_code = req_post.get('upd_cond_code','')        
        upd_sn = req_post.get('upd_sn','') 
        upd_traceable = req_post.get('upd_traceable','') 
        upd_tag_date = req_post.get('upd_tag','') 
        upd_mfg_date = req_post.get('upd_mfg','') 
        upd_tagged = req_post.get('upd_tagged','')
        upd_certified = req_post.get('upd_certified','') 
        upd_obtained = req_post.get('upd_obtained','') 
        upd_hold = req_post.get('upd_hold',False)
        upd_mfctr = req_post.get('upd_mfctr','')
        upd_ctry_origin = req_post.get('upd_ctry_origin','')
        upd_tag_type = req_post.get('upd_tag_type','')
        upd_tsn_csn = req_post.get('upd_tsn_csn','')
        upd_tso_cso = req_post.get('upd_tso_cso','')
        upd_exp_date = req_post.get('upd_exp_date','')
        upd_insp_date = req_post.get('upd_insp_date','')
        upd_pn = req_post.get('upd_pn','')
        upd_desc = req_post.get('upd_desc','')
        upd_cons = req_post.get('upd_cons','')
        upd_loc = req_post.get('upd_loc','')
        upd_should_be = req_post.get('upd_should_be','')
        upd_trac = req_post.get('upd_trac','')
        upd_alt_pn = req_post.get('upd_alt_pn','')
        upd_remarks = req_post.get('upd_remarks','')
        upd_cure_date = req_post.get('upd_cure_date','')
        app_code = req_post.get('app_code','')
        session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        row = []
        row.append(wo_task)
        row.append(quantity)
        row.append(part_number)
        row.append(serial_number)
        row.append(session_id)
        row.append(notes)
        row.append(modal_desc)
        row.append(description)
        row.append(stock_status)
        row.append(upd_cond_code)        
        row.append(upd_traceable) 
        row.append(upd_tag_date) 
        row.append(upd_mfg_date) 
        row.append(upd_tagged)
        row.append(upd_certified) 
        row.append(upd_obtained) 
        row.append(upd_hold)
        row.append(upd_mfctr)
        row.append(upd_ctry_origin)
        row.append(upd_tag_type)
        row.append(upd_tsn_csn)
        row.append(upd_tso_cso)
        row.append(upd_exp_date)
        row.append(upd_insp_date)
        row.append(upd_cons)
        row.append(upd_loc)
        row.append(upd_should_be)
        row.append(upd_trac)
        row.append(upd_alt_pn)
        row.append(upd_remarks)
        row.append(upd_cure_date)        
        if wo_task and show_msg in ['','done']:          
            error,msg,show_msg = create_lot_teardown(quapi_id,sysur_auto_key,user_name,row)
            if show_msg != 'show_modal':
                show_msg = 'done' 
                auto_click = ''
                
        elif wo_task and show_msg == 'got_data':
            #create the new PN with this data using a new task (method) called create_pn
            from portal.tasks import create_pn          
            res = create_pn.delay(quapi_id,session_id,modal_pn,modal_desc,modal_serialized,sysur_auto_key)
            error,msg = res.get()
            row[2] = modal_pn
            error,msg,show_msg = create_lot_teardown(quapi_id,sysur_auto_key,user_name,row)
            show_msg = 'done'
            auto_click = ''
            val_dict['wo_task'] = ''
            val_dict['error'] = error
            val_dict['msg'] = msg
            
        elif not wo_task:
            error = "WO# is required."
            val_dict.update({
                'error': error,           
                'session_id': session_id,
                'part_number': part_number,
                'quantity': quantity,
                'wo_task': wo_task,
                'serial_number': serial_number,
                'notes': notes,
                'stock_status': stock_status,
                'description': description,
                'user_name': user_name,  
                'auto_click': 'done',                               
            }) 
            return render(request, 'mrolive/teardown.html', val_dict)
           
        updated_woos = WOStatus.objects.filter(session_id=session_id)
        record = updated_woos and updated_woos[0] or None
        ctrl_id = record and record.ctrl_id or 0
        ctrl_number = record and record.ctrl_number or 0
        stm = record and record.stm_auto_key or ''
        if not stm or 1:
            element = '0' + str(ctrl_number) 
            element += '00000' + str(ctrl_id)
        else:
            element = 'C' + str(stm)
        pn = record and record.part_number or ''
        pnm_auto_key = record and record.pnm_auto_key or ''
        activity = record and record.activity or ''
        description = record and record.description or ''
        desc_last = ''
        if len(description) > 10:
            if len(description) > 35:
                desc_last = description[10:35]
            else:
                desc_last = description[10:]
            description = description[:10]           
        serial_no = record and record.serial_number or ''
        qty = record and record.quantity or 0
        wo_number = record and record.wo_number or ''
        repair = record and activity and activity == 'Repair' or ''
        stock_line = record and record.stock_line or ''
        exp_date = record and record.exp_date or ''
        condition_code = record and record.condition_code or ''
        consignment_code = record and record.consignment_code or ''
        loc_code = record and record.location_code or ''
        mfg_lot_num = record and record.spn_code or ''
        eng_model = record and record.int_rank or ''
        customer = record and record.customer or '' 
        status = record and record.status or ''        
        #get the task list to display on the Teardown label

        if pnm_auto_key:
            from portal.tasks import get_part_attributes
            res = get_part_attributes.delay(quapi_id,sysur_auto_key,session_id,pnm_auto_key,group='IN HOUSE',create_anew=False)  
            error,part_msg = res.get()
        if not error:
            from polls.models import UserDefAtts as uda
            task_list = uda.objects.filter(session_id=session_id).order_by('att_seq')
        printed_date = datetime.now()
        printed_date = printed_date.strftime('%m/%d/%Y %H:%M:%S') 

        val_dict.update({
            'element': element,
            'ctrl_id': ctrl_id,
            'ctrl_number': ctrl_number,
            'pn': pn,
            'stm_auto_key': stm,
            'part_number': pn,
            'description': description,
            'desc_last': desc_last,
            'serial_no': serial_no,
            'serial_number': serial_no,
            'quantity': qty or quantity,
            'wo_number': wo_number,  
            'si_number': wo_number,
            'type': format_type, 
            'auto_click': auto_click, 
            'task_list': task_list, 
            'repair': repair,
            'stock_line': stock_line,
            'exp_date': exp_date,
            'condition_code': condition_code,
            'consignment_code': consignment_code,
            'loc_code': loc_code,
            'mfg_lot_num': mfg_lot_num,
            'eng_model': eng_model,
            'customer': customer,
            'printed_by': user_name,
            'printed_date': printed_date,
            'status': status,
            'stock_recs': updated_woos,
            'lotdown': 1,
        }) 

        if msg == 'Successful Teardown.':    
            part_number = ''
            serial_number = ''
            notes = ''
            quantity = ''
            printset = app_allow and app_allow[0] and app_allow[0].printset_id
            auth_key = printset and printset.printnode_auth_key
            stock_recs = WOStatus.objects.filter(session_id=session_id)
            val_dict['stock_recs'] = stock_recs
            
            if 0:
                create_barcodes(element,ctrl_id,ctrl_number,pn,description,serial_no,qty,wo_number,task_list,repair)
                              
                #val_dict['error'] = printnode_pdf([printset],auth_key)
                val_dict['msg'] = 'Successful teardown. Barcode label printed.'                
            
            val_dict['app_code'] = 'lot-teardown'               

            if 1:            
                #return render(request, 'mrolive/teardown.html', val_dict)
                return render(request, 'mrolive/plain_barcode_mro.html', val_dict)
            else:
                if not error:
                    if not repair:
                        return render(request, 'mrolive/plain_barcode_unical.html', val_dict)
                    else:
                        return render(request, 'mrolive/report_inspection_mro.html', val_dict)               
                else:
                    val_dict['error'] = error
                    return render(request, 'mrolive/teardown.html', val_dict)
                    
        if show_msg in ['show_modal']:
            modal_pn = part_number
        val_dict.update({         
            'session_id': session_id, 
            'quantity': quantity,
            'part_number': part_number,
            'wo_task': wo_task,
            'serial_number': serial_number,
            'notes': notes,  
            'show_modal': show_msg,
            'modal_pn': modal_pn,
            'user_name': user_name,
            
        })            
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/teardown.html', val_dict) 
  
def create_expiry(request):
    val_dict = {}
    if request.method == 'GET':
        form = WODashboardForm()        
    if request.method == 'POST':
        req_post = request.POST
        expiry_date = req_post.get('expiry_date',None)
        expiry_date = '2026/12/31 00:00:00'
        form = WODashboardForm(req_post)     
        signer = Signer(salt='transubstantiation')
        enc_date = signer.sign(expiry_date)
        exp_date = signer.unsign(enc_date)      
    return render(request, 'mrolive/create_expiry.html', val_dict)     
    
def update_expiry(user,quapi_id,sysur_auto_key,req_files): 
    error,msg = '','' 
    from portal.tasks import update_expiry
    up_file = 'file' in req_files and req_files['file'] or None
    save_file = up_file and Document(docfile=up_file) or None
    save_file.save()
    user_groups = user.groups.values_list('id',flat=True)
    group = user_groups and user_groups[0] or None
    groups = UserGroupProfile.objects.filter(quantum_cmp_key__isnull=False)
    lic_group = groups and groups[0] or None
    if not lic_group:
        error = 'User must belong to a company group.'
        return error,msg	    
    quantum_cmp_key = lic_group.quantum_cmp_key
    conn_string = lic_group.conn_string
    private_key = lic_group.private_key
    if up_file:
        file_path = '/home/ubuntu/mo_template/media/' + up_file.name
        file_path = file_path.replace(' ','_') 
        with open(file_path, mode='r') as txt_file:        
            for line in txt_file:
                signer = Signer(salt='transubstantiation')                
                expiry_date = signer.unsign(line)
                res = update_expiry.delay(quapi_id,conn_string,quantum_cmp_key,expiry_date)
                error,msg = res.get()
                break 
    else:
        error = 'No file found.'       
    return error,msg
    
def expiry_set(request,quapi_id):
    error,msg,form,val_dict = '','',{},{}
    new_status,location,filter_status = '','',''
    user_id,user_rec = 'user not set',None
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request and request.user or None  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    if not reg_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id        
    if request.method == 'GET':
        form = WODashboardForm()        
    if request.method == 'POST':
        req_post = request.POST
        req_files = request.FILES
        form = WODashboardForm(req_post, req_files)     
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''          
        val_dict.update({
            'msg': msg,
            'session_id': session_id,                      
        })
        if req_files:         
            error,msg = update_expiry(user,quapi_id,sysur_auto_key,req_files)
            
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/expiry_set.html', val_dict) 
    
    
def exchange_portal(request,quapi_id):
    error,msg,warning_msg = '','',''
    user = request and request.user or None
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')   
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict)
    form = WODashboardForm()  
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='exchange-portal')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    if request.method == 'GET':
        form = WODashboardForm()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        is_search = req_post.get('is_search','0')
        wo_number = req_post.get('wo_number','')
        pn = req_post.get('pn','')
        description = req_post.get('description','')
        session_id = req_post.get('session_id','')
        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','')
        val_dict.update({
            'session_id': session_id,
            'wo_number': wo_number,       
        })
        keeper_list = 'keepers_list[]' in req_post and req_post.getlist('keepers_list[]') or []
        shipper_list = 'shippers_list[]' in req_post and req_post.getlist('shippers_list[]') or []
        
        if wo_number and not (keeper_list and shipper_list):
            from portal.tasks import get_exchange_stock
            res = get_exchange_stock.delay(quapi_id,session_id,wo_number)
            error,msg = res.get()          
            keeper_stms = WOStatus.objects.filter(session_id=session_id,wo_type='KEEPER') or []
            pn = keeper_stms and keeper_stms[0] and keeper_stms[0].part_number or ''
            description = keeper_stms and keeper_stms[0] and keeper_stms[0].description or ''
            shipper_stms = WOStatus.objects.filter(session_id=session_id,wo_type='SHIPPER') or []
            val_dict.update({
                'error': error,
                'msg': msg,
                'total_keepers': len(keeper_stms),
                'total_shippers': len(shipper_stms),
                'sel_keepers': 0,
                'sel_shippers': 0,
                'pn': pn,
                'description': description,
            })                   
        elif keeper_list and shipper_list:
            if len(keeper_list) == 1 and len(shipper_list)== 1:
                from portal.tasks import exchange_stock 
                res = exchange_stock.delay(quapi_id,session_id,sysur_auto_key,keeper_list,shipper_list,wo_number)
                error,msg = res.get()
            else:
                error = 'Select one and only one row in each table.'            
            keeper_stms = WOStatus.objects.filter(session_id=session_id,wo_type='KEEPER') or []
            shipper_stms = WOStatus.objects.filter(session_id=session_id,wo_type='SHIPPER') or []
            val_dict.update({
                'error': error,
                'msg': msg,
                'total_keepers': len(keeper_stms),
                'total_shippers': len(shipper_stms),
                'sel_keepers': 0,
                'sel_shippers': 0,
                'pn': pn,
                'description': description,
            }) 
        else:
            error = 'Must select rows to exchange.'       
    val_dict['form'] = form
    return render(request, 'mrolive/exchange_portal.html', val_dict)

def get_stock_uom(request):
    req_post = request.POST
    stock_label = req_post.get('stock_label','')
    quapi_id = req_post.get('quapi_id','')

    if stock_label:
        try:
            from portal.tasks import get_stock_uom
            import random
            import string
            session_id = random.choices(string.ascii_lowercase)
            res = get_stock_uom.delay(quapi_id,session_id,stock_label)
            error,uom_code = res.get()
        except Exception:           
            return JsonResponse('error retrieving uom.')
        return JsonResponse({'uom_code':uom_code}, safe = False)

def get_parts_ajax(request):
    req_post = request.POST
    part_char = req_post.get('part_char','')
    quapi_id = req_post.get('quapi_id','')

    if part_char:
        try:
            from polls.models import PartNumbers
            from portal.tasks import get_part_numbers
            import random
            import string
            session_id = random.choices(string.ascii_lowercase)
            res = get_part_numbers.delay(quapi_id,session_id,part_char=part_char)
            error = res.get()
            parts = PartNumbers.objects.filter(session_id=session_id)
            #parts = parts and parts.values_list('part_number',flat=True) or []
            #parts = list(parts)
        except Exception:           
            return JsonResponse('error retrieving parts.')
        return JsonResponse(list(parts.values('part_number')), safe = False)

def e_signoff(request,quapi_id):
    error,msg,warning_msg = '','',''
    user = request and request.user or None
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')   
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict)
    form = WODashboardForm()  
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='e-signoff')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    if request.method == 'GET':
        form = WODashboardForm()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        is_search = req_post.get('is_search','0')
        active_task = req_post.get('task_selector','')
        wo_number = req_post.get('wo_number','')
        session_id = req_post.get('session_id','')
        description = req_post.get('description','')
        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','')
        val_dict.update({
            'session_id': session_id,
            'wo_number': wo_number, 
            'description': description, 
            'active_task': active_task            
        })
        if wo_number and is_search == '1':
            from portal.tasks import get_signoff_tasks
            if wo_number[-1] in ['s','S','c','C']:
                wot_auto_key = wo_number[:-1]
                res = get_signoff_tasks.delay(quapi_id,session_id,wot_auto_key=wot_auto_key)
                error,msg = res.get()
            else:
                res = get_signoff_tasks.delay(quapi_id,session_id,wo_number=wo_number)
                error,msg = res.get()                
            wtms = WOTask.objects.filter(session_id=session_id).order_by('wot_sequence','wot_auto_key') or []
            val_dict.update({
                'wtms': wtms,
                'error': error,
                'msg': msg,
                'total_rows': len(wtms),
                'sel_rows': 0,
            })                    
        elif is_search != '1':
            from portal.tasks import e_signoff   
            wot_id_list = []
            if 'woos_list[]' in req_post and is_search != '1':
                wot_id_list = req_post.getlist('woos_list[]')
            if active_task and not wot_id_list:
                res = e_signoff.delay(quapi_id,session_id,sysur_auto_key,active_task) 
                error,msg = res.get()                
            for wot in wot_id_list:   
                res = e_signoff.delay(quapi_id,session_id,sysur_auto_key,wot)
                signoff_error,signoff_msg = res.get()
                msg += signoff_msg
                error += signoff_error
            wtms = WOTask.objects.filter(session_id=session_id).order_by('wot_sequence','wot_auto_key')
            val_dict.update({
                'wtms': wtms,
                'error': error,
                'msg': msg,
                'warning_msg': warning_msg,
                'total_rows': len(wtms),
                'sel_rows': 0,
            })           
    val_dict['form'] = form
    return render(request, 'mrolive/e-signoff.html', val_dict)
    
def lot_inspection(request,quapi_id,wo_number=''):
    error,msg,update_msg,boms = '','','',''
    user = request and request.user or None
    username = user and user.username or ''
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict) 
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='lot-inspection')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  

    from portal.tasks import get_stock_status,get_consignment_codes
    session_id = 'annW8NK23R(*vN;ARERKTNALSD;KAJ'
    res = get_stock_status.delay(quapi_id,session_id)
    error = res.get()        
    val_dict['statuses'] = StatusSelection.objects.filter(session_id=session_id)
    res = get_consignment_codes.delay(quapi_id,session_id)
    error = res.get()      
    val_dict['cons_codes'] = Consignments.objects.filter(session_id=session_id)        
    if request.method == 'GET':
        form = WODashboardForm()
        val_dict['wo_number'] = wo_number        
    
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        show_modal = ''
        pn = req_post.get('pn','')
        qty_need = req_post.get('quantity','')
        part_number = req_post.get('part_number','')
        description = req_post.get('description','')
        wo_number = req_post.get('wo_number','')
        filter_wo = req_post.get('filter_wo','')
        update_pn = req_post.get('update_pn','')
        notes = req_post.get('notes','')  
        consignment = req_post.get('consignment','')        
        stock_status = req_post.get('stock_status','') or req_post.get('filter_status','')
        serial_number = req_post.get('serial_number','')
        upd_status = req_post.get('upd_status','')
        upd_notes = req_post.get('upd_notes','')        
        upd_qty = req_post.get('upd_qty','')
        upd_cond_code = req_post.get('upd_cond_code','')        
        upd_sn = req_post.get('upd_sn','') 
        upd_traceable = req_post.get('upd_traceable','') 
        upd_tag_date = req_post.get('upd_tag','') 
        upd_mfg_date = req_post.get('upd_mfg','') 
        upd_tagged = req_post.get('upd_tagged','')
        upd_certified = req_post.get('upd_certified','') 
        upd_obtained = req_post.get('upd_obtained','') 
        upd_hold = req_post.get('upd_hold',False)
        upd_mfctr = req_post.get('upd_mfctr','')
        upd_ctry_origin = req_post.get('upd_ctry_origin','')
        upd_tag_type = req_post.get('upd_tag_type','')
        upd_tsn_csn = req_post.get('upd_tsn_csn','')
        upd_tso_cso = req_post.get('upd_tso_cso','')
        upd_exp_date = req_post.get('upd_exp_date','')
        upd_insp_date = req_post.get('upd_insp_date','')
        upd_pn = req_post.get('upd_pn','')
        upd_desc = req_post.get('upd_desc','')
        
        upd_cons = req_post.get('upd_cons','')
        upd_loc = req_post.get('upd_loc','')
        upd_should_be = req_post.get('upd_should_be','')
        upd_trac = req_post.get('upd_trac','')
        upd_alt_pn = req_post.get('upd_alt_pn','')
        upd_remarks = req_post.get('upd_remarks','')
        upd_cure_date = req_post.get('upd_cure_date','')
        update_msg = req_post.get('update_msg','')
        
        #sel_rows = req_post.get('sel_rows','') or 0
        is_search = req_post.get('is_search','0')  
        is_update = req_post.get('is_update','0')
        is_accept = req_post.get('is_accept','0')
        yes_print = req_post.get('yes_print','0')
        no_print = req_post.get('no_print','0')
        session_id = req_post.get('session_id','')
        filter_session = req_post.get('filter_session','')  
        already_printed = req_post.get('already_printed','0')
        
        if not (session_id or filter_session):
            session_id = req_post.get('csrfmiddlewaretoken','') or ''
        
        elif filter_session:
            session_id = filter_session
        
        val_dict.update({
            'part_number': part_number,
            'description': description,
            'serial_number': serial_number,            
            'qty_need': qty_need,            
            'session_id': session_id,
            'wo_number': wo_number,
            #'sel_rows': sel_rows,
            'form': form,
            })
  
        req_files = request.FILES
        file_name,import_file,file_hash,file_ext = '','','',''
        up_file = 'img_file' in req_files or None

        if up_file:
            up_file = req_files['img_file']
            file_name = up_file.name
            file_name = file_name.replace(" ","")
            up_file.name = file_name
            import hashlib
            file_hash = hashlib.sha256(file_name.encode('utf-8')).hexdigest()
            file_hash = file_hash[:50]
            file_ext = file_name.split('.')
            file_ext = file_ext and '.' + file_ext[1] or '.jpg'
            
        if up_file:
            import_file = up_file and\
            Document(session_id=session_id,docfile=up_file,\
            file_name=file_name,file_hash=file_hash,file_extension=file_ext) or None
            
        if import_file:
            import_file.save()
        
        if 'lot_list[]' in req_post:
            lot_list = req_post.getlist('lot_list[]')

        elif 'lot_sels[]' in req_post:
            lot_list = req_post.getlist('lot_sels[]')
       
        if is_search == '1' and (wo_number or stock_status\
            or part_number or description or serial_number\
            or consignment): 
            from portal.tasks import get_lots
            filters = [wo_number,part_number,description,serial_number,stock_status,consignment]
            res = get_lots.delay(quapi_id,session_id,filters)            
            error,msg = res.get() 
            lots = WOStatus.objects.filter(session_id = session_id)
                
            if not lots:
                error = 'No stock found.'
                
            val_dict.update({
                'total_rows': len(lots),
                'error': error,               
            })
            
        elif is_accept == '1':
            """1.	Update STOCK[‘QTY_REC_FROM_LOT’] = WO_BOM[‘QTY_NEEDED’]
               2.	print barcode label for stock line(s) selected
            """
            if 'lot_list[]' in req_post:               
                parameters = ['',upd_notes,upd_qty,upd_cond_code,upd_sn]
                filters = [wo_number,part_number,description,serial_number,stock_status,consignment]
                from portal.tasks import update_accept_lots
                res = update_accept_lots.delay(quapi_id,username,session_id,\
                    sysur_auto_key,lot_list,filters,parameters,\
                    accept_only=True)
                error,msg = res.get()
            else:
                error = 'Must select grid row(s).'
            lots = WOStatus.objects.filter(session_id = session_id)
            val_dict.update({
                'show_modal': not error and 'label' or '',
                'error': error,
                'msg': msg,
                'total_rows': len(lots),
                'lot_list': lot_list,
                })
            
        elif 'lot_list[]' in req_post and is_update == '1':
            
            if 'lot_list[]' in req_post:
                
                from portal.tasks import get_conditions,get_companies,\
                    get_certs,get_stock_status,get_tag_types,\
                    get_consignment_codes,get_locations
                
                res = get_consignment_codes.delay(quapi_id,session_id)
                error = res.get()
                res = get_conditions.delay(quapi_id,session_id)
                error = res.get()
                res = get_companies.delay(quapi_id,session_id)
                error = res.get()
                res = get_certs.delay(quapi_id,session_id)
                error = res.get()
                res = get_stock_status.delay(quapi_id,session_id)
                error = res.get()  
                res = get_tag_types.delay(quapi_id,session_id)
                error = res.get()            
                
                from polls.models import PartConditions as pcc,\
                   Companies as cmp, Departments as dept, StatusSelection as stat,\
                   StockCart as stk_cart, Consignments as cons
                
                cons_codes = cons.objects.filter(session_id=session_id).order_by('code')
                conditions = pcc.objects.filter(session_id=session_id).order_by('condition_code')
                companies = cmp.objects.filter(session_id=session_id).order_by('name')
                cert_codes = dept.objects.filter(session_id=session_id).order_by('name')
                statuses = stat.objects.filter(session_id=session_id).order_by('name')
                tag_types = stk_cart.objects.filter(session_id=session_id).order_by('name')
                
                lots = WOStatus.objects.filter(session_id = session_id,stm_auto_key__in=lot_list)
                full_lots = WOStatus.objects.filter(session_id = session_id)
                prod_pn,prod_desc,prod_cons,sl_list = '','','',''
                count = 0
                for lot in lots:
                    if count == 0:
                        prod_pn = lot.part_number
                        prod_desc = lot.description
                        prod_cons = lot.consignment_code
                        
                    else:                    
                        if prod_pn != lot.part_number:
                            prod_pn = 'Multiple'
                        
                        if prod_desc != lot.description:
                            prod_desc = 'Multiple'
                            
                        if prod_cons != lot.consignment_code:
                            prod_cons = 'Multiple'
                            
                    sl_list += lot.stock_line + ' | '
                    count += 1
                    
                val_dict.update({
                    'show_modal': 'T',
                    'lot_list': lot_list,
                    'sl_list': sl_list,
                    'wo_number': wo_number,
                    'part_number': part_number,
                    'description': description,
                    'stock_status': stock_status,
                    'serial_number': serial_number,
                    'conditions': conditions,
                    'companies': companies,
                    'cert_codes': cert_codes,
                    'cons_codes': cons_codes,
                    'locations': [],
                    'tag_types': tag_types,
                    'statuses': statuses,
                    'prod_pn': prod_pn,
                    'prod_desc': prod_desc,
                    'prod_cons': prod_cons,
                    })
            else:
                error = 'Must select grid row(s).'
            val_dict.update({
                'error': error,
                'msg': msg,
                'total_rows': len(full_lots),
                })
                
        elif 'lot_sels[]' in req_post and lot_list and 'yes_print' not in req_post:
            
            if file_name:
                
                #call the fxn that pushes the image via sftp and creates the Quantum record
                from portal.tasks import import_part_image
                res = import_part_image.delay(quapi_id,session_id,lot_list,file_name,file_ext,file_hash)
                error,msg = res.get()
                
                if error:
                    val_dict['error'] = error    
                
            if upd_cond_code or upd_qty or upd_sn or upd_notes\
                or upd_traceable or upd_tag_date or upd_mfg_date\
                or upd_tagged or upd_obtained or upd_status\
                or upd_insp_date or upd_loc\
                or upd_should_be or upd_trac or upd_alt_pn or upd_remarks\
                or upd_cure_date or upd_exp_date or upd_hold or upd_tso_cso\
                or upd_ctry_origin or upd_tag_type or upd_mfctr:
                from portal.tasks import update_lots
                parameters = [upd_status,upd_notes,upd_qty,upd_cond_code,upd_sn]
                parameters += [upd_traceable,upd_tag_date,upd_mfg_date,upd_tagged]
                parameters += [upd_obtained,upd_hold,upd_mfctr,upd_ctry_origin,upd_tag_type]
                parameters += [upd_tsn_csn,upd_tso_cso,upd_insp_date]
                parameters += [upd_pn,upd_desc,upd_cons,upd_loc,upd_should_be]
                parameters += [upd_trac,upd_alt_pn,upd_remarks,upd_cure_date]
                parameters += [upd_exp_date,]
                wo_number = req_post.get('filter_wo','')
                part_number = req_post.get('filter_pn','')
                description = req_post.get('filter_desc','')
                serial_number = req_post.get('filter_serial','')
                stock_status = req_post.get('filter_status','')
                consignment = req_post.get('filter_cons','')
                filters = [wo_number,part_number,description,serial_number,stock_status,consignment] 
                
                res = update_lots.delay(quapi_id,username,session_id,\
                    sysur_auto_key,lot_list,parameters,filters)
                
                error,msg = res.get()  
            
            if not error:
                show_modal = 'label'                
                
            lots = WOStatus.objects.filter(session_id = session_id)           
            val_dict.update({
                'show_modal': show_modal,
                'total_rows': len(lots),
                'lot_list': lot_list,
                'error': error,
                'msg': msg,
                'update_pn': upd_alt_pn and '1' or '0',                
                'wo_number': wo_number,
                'part_number': part_number,
                'description': description,                
                'qty_need': qty_need,
                'stock_status': stock_status,                
                'notes': notes,                
            }) 
                
        elif ('lot_sels[]' not in req_post or 'lot_list[]' not in req_post):
            if is_update == '1' or is_accept == '1':
                val_dict['error'] = 'Select rows in the grid to accept or update.' 
                lots = WOStatus.objects.filter(session_id = session_id)             
                val_dict.update({
                    'total_rows': len(lots),
                    'wo_number': filter_wo,
                })
                
        if already_printed != 'T' and 'yes_print' in req_post and req_post.get('yes_print') == '1' and lot_list:
            pnm_update = update_pn=='1' and True or False

            if pnm_update:
                stms = WOStatus.objects.filter(\
                    session_id = session_id,\
                    is_dashboard = pnm_update)
            else:
                stms = WOStatus.objects.filter(\
                    session_id = session_id,\
                    stm_auto_key__in = lot_list)
                
            printed_date = datetime.now()
            printed_date = printed_date.strftime('%m/%d/%Y %H:%M:%S')  
            
            val_dict.update({
                'already_printed': 'T',
                'printed_date': printed_date,
                'stock_recs': stms,
                'wo_number': filter_wo,
                'si_number': filter_wo,
                'lot_inspection': '1',
                'inspector': user.first_name + ' ' + user.last_name,
                'printed_by': user.first_name + ' ' + user.last_name,
                })
                
            return render(request, 'mrolive/plain_barcode_mro.html', val_dict)
                
    val_dict['msg'] = update_msg or msg 
    val_dict['form'] = form
    return render(request, 'mrolive/lot_inspection.html', val_dict)    
  
def lot_management(request,quapi_id):
    error,boms = '',''
    user = request and request.user or None
    username = user and user.username or ''
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict) 
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='lot-management')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')   
    #if request.method == 'GET':
    form = WODashboardForm()
    session_id = 'annW8NK23R(*vN;ARERKTNALSD;KAJ'
    from portal.tasks import get_stock_status,get_consignment_codes
    res = get_stock_status.delay(quapi_id,session_id)
    error = res.get()        
    val_dict['statuses'] = StatusSelection.objects.filter(session_id=session_id)      
    res = get_consignment_codes.delay(quapi_id,session_id)        
    error = res.get()
    from polls.models import Consignments as cons
    val_dict['cons_codes'] = cons.objects.filter(session_id=session_id)

    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        pn = req_post.get('pn','')
        qty_need = req_post.get('quantity','')
        part_number = req_post.get('part_number','')
        description = req_post.get('description','')
        wo_number = req_post.get('wo_number','')
        notes = req_post.get('notes','') 
        consignment = req_post.get('consignment','')
        cons_sel = req_post.get('cons_selector','')           
        stock_status = req_post.get('stock_status','') 
        serial_number = req_post.get('serial_number','')
        instr = req_post.get('instr','')
        sel_rows = req_post.get('sel_rows','') or 0
        is_search = req_post.get('is_search','0')  
        is_update = req_post.get('is_update','0') 
        certifs = request.POST.getlist('cert',[])
        certifs = ', '.join(certifs)    
        upd_notes = "%s | %s.  %s"%(instr,certifs,req_post.get('upd_notes',''))
        upd_status = req_post.get('upd_status','')        
        session_id = 'session_id' in req_post and req_post['session_id'] or '' 
        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','') or ''
            
                                 
                                                                                                                               
                         
        val_dict.update({
            'part_number': part_number,
            'description': description,
            'serial_number': serial_number,            
            'qty_need': qty_need,                      
            'session_id': session_id,
            'wo_number': wo_number,
            'consignement': consignment or cons_sel,
            'sel_rows': sel_rows,
            'form': form,
            'session_id': session_id,
            'stock_status': stock_status,
              
            })
            
        if 'lot_list[]' in req_post:
            lot_list = req_post.getlist('lot_list[]')  
        elif 'lot_sels[]' in req_post:
            lot_list = req_post.getlist('lot_sels[]')
          
        if is_search == '1' and (wo_number or stock_status or part_number or description or serial_number or cons_sel): 
            from portal.tasks import get_lots
            filters = [wo_number,part_number,description,serial_number,stock_status,cons_sel]
            res = get_lots.delay(quapi_id,session_id,filters)           
            error,msg = res.get() 
            lots = WOStatus.objects.filter(session_id = session_id)
            if not lots:
                error = 'No lots found.'
            val_dict.update({
                'total_rows': len(lots),
                'error': error,              
            })                     
        elif 'lot_list[]' in req_post and is_update == '1':
            count = 0
            full_lots = WOStatus.objects.filter(session_id = session_id)
            lots = WOStatus.objects.filter(session_id = session_id,stm_auto_key__in=lot_list)
            prod_pn,prod_desc,prod_cons,sl_list = '','','',''
            
            for lot in lots:
            
                if count == 0:
                    prod_pn = lot.part_number
                    prod_desc = lot.description
                    prod_cons = lot.consignment_code
                    
                else:                    
                    if prod_pn != lot.part_number:
                        prod_pn = 'Multiple'
                    
                    if prod_desc != lot.description:
                        prod_desc = 'Multiple'
                        
                    if prod_cons != lot.consignment_code:
                        prod_cons = 'Multiple'
                        
                sl_list += lot.serial_number + ' | '
                count += 1
                
            val_dict.update({
                'total_rows': len(full_lots),
                'show_modal': 'T',
                'sl_list': sl_list,
                'lot_list': lot_list,
                'wo_number': wo_number,
                'part_number': part_number,
                'description': description,
                'stock_status': stock_status,
                'serial_number': serial_number,
                'consignment': consignment,
                'prod_pn': prod_pn,
                'prod_desc': prod_desc,
                'prod_cons': prod_cons,
                })
            
        elif 'lot_sels[]' in req_post and is_search != '1':
            if upd_status or upd_notes:
                from portal.tasks import update_lots,get_lots
                parameters = [upd_status,upd_notes,'','',''] 
                wo_number = req_post.get('filter_wo','')
                part_number = req_post.get('filter_pn','')
                description = req_post.get('filter_desc','')
                serial_number = req_post.get('filter_serial','')
                stock_status = req_post.get('filter_status','')
                consignement = req_post.get('filter_cons','') 
                filters = [wo_number,part_number,description,serial_number,stock_status,consignement]      

                res = update_lots.delay(quapi_id,username,session_id,sysur_auto_key,lot_list,parameters,filters,is_mgmt=True)
                error,msg = res.get()   
                res = get_lots.delay(quapi_id,session_id,filters)            
                error,msg = res.get()                 
                lots = WOStatus.objects.filter(session_id = session_id)             
                val_dict.update({
                    'total_rows': len(lots),
                    'error': error,
                    'msg': msg,                
                    'wo_number': wo_number,
                    'part_number': part_number,
                    'description': description,                
                    'qty_need': qty_need,                              
                    'notes': notes,  
                    'stock_status': stock_status,
                    'sel_rows': 0,                    
                })         
    val_dict['form'] = form
    return render(request, 'mrolive/lot_management.html', val_dict)    
  
def bom_management(request,quapi_id):
    error,boms = '',''
    user = request and request.user or None
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict) 
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='bom-management')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')
    bom_statuses = StatusSelection.objects.filter(session_id='1234567').order_by('name')  
    val_dict['bom_statuses'] = bom_statuses    
    if request.method == 'GET':
        form = WODashboardForm()
        #session_id = '1234567'
        #from portal.tasks import get_activities_conditions_tasks
        #res = get_activities_conditions_tasks.delay(quapi_id,session_id)
        #error,msg = res.get()          
        #from polls.models import Activities, PartConditions            
        #activities = Activities.objects.filter(session_id = session_id).order_by('activity') 
        #conditions = PartConditions.objects.filter(session_id = session_id).order_by('condition_code')
        #wo_tasks = WOTask.objects.filter(session_id = session_id).order_by('task_master_desc','wot_sequence')

        #val_dict.update({
        #    'error': error,
        #    'conditions': conditions,
        #    'activities': activities,
        #    'wo_tasks': wo_tasks,         
        #})        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        pn = req_post.get('pn','')
        qty_need = req_post.get('quantity','')
        part_number = req_post.get('part_number','')
        description = req_post.get('description','')
        stm_auto_key = req_post.get('stm_auto_key','')
        condition = req_post.get('condition_code','')     
        activity = req_post.get('activity','')                
        wo_task = req_post.get('wo_task','')
        wo_number = req_post.get('wo_number','')
        si_number = req_post.get('si_number','')        
        act_sel = req_post.get('act_selector','')
        cond_sel = req_post.get('cond_selector','')
        task_sel = req_post.get('task_sel','')
        task_selector = req_post.get('task_selector','')
        bom_status = req_post.get('bom_status','') 
        notes = req_post.get('notes','')
        sel_rows = req_post.get('sel_rows','') or 0
        is_search = req_post.get('is_search','0') 
        launch_update = req_post.get('launch_update','0')
        is_update = req_post.get('is_update','0')        
        modal_update = req_post.get('modal_update','0')         
        session_id = req_post.get('session_id','')  
        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','')
            
        from portal.tasks import get_bom_statuses
        res = get_bom_statuses.delay(quapi_id,session_id)
        res.get()      
        bom_statuses = StatusSelection.objects.filter(session_id=session_id).order_by('name')
        from portal.tasks import get_activities_conditions_tasks
        res = get_activities_conditions_tasks.delay(quapi_id,session_id,si_number = wo_number or si_number)
        error,msg = res.get()          
        from polls.models import Activities, PartConditions            
        activities = Activities.objects.filter(session_id = session_id).order_by('activity') 
        conditions = PartConditions.objects.filter(session_id = session_id).order_by('condition_code')
        wo_tasks = WOTask.objects.filter(session_id = session_id,si_number__iexact=wo_number).order_by('wot_sequence')
        
        val_dict.update({
            'part_number': part_number,
            'pn': '',
            'description': description,             
            'qty_need': qty_need,            
            'stm_auto_key': stm_auto_key,
            'session_id': session_id,
            'wo_number': wo_number,
            'wo_task': wo_task,
            'si_number': si_number,
            'sel_rows': sel_rows,
            'form': form,
            'task_sel': '',
            'act_sel': '',
            'cond_sel': '',
            'notes': '',
            'quantity': '',
            'bom_statuses': bom_statuses,
            'wo_tasks': wo_tasks,
            'bom_status': bom_status,
            'conditions': conditions,
            'activities': activities,
            'wo_tasks': wo_tasks,
            })    
        
        wob_id_list = []
        if 'wobs_list[]' in req_post and is_search != '1':
            wob_id_list = req_post.getlist('wobs_list[]')  
            
        elif 'wob_sels[]' in req_post:
            wob_id_list = req_post.getlist('wob_sels[]')
            
        if launch_update == '1':
            #launch update pop-up
            all_boms = WOStatus.objects.filter(
                session_id = session_id)   
               
            boms = all_boms.filter(
                wob_auto_key__in=wob_id_list)
            prod_pn,prod_desc,prod_qty = '','',''
            prod_cond,prod_act,prod_notes = '','',''
            count = 0
            
            for bom in boms:
            
                if count == 0:
                    prod_pn = bom.part_number
                    prod_desc = bom.description
                    prod_qty = bom.qty_needed
                    prod_cond = bom.condition_code
                    prod_act = bom.activity
                    prod_notes = bom.notes
                    prod_task = bom.wot_sequence + ' - ' + bom.task_master_desc
                    prod_stat = bom.wos_auto_key
                       
                else: 
                
                    if prod_pn != bom.part_number:
                        prod_pn = ''
                    
                    if prod_desc != bom.description:
                        prod_desc = ''
                        
                    if prod_qty != bom.qty_needed:
                        prod_qty = ''
                        
                    if prod_cond != bom.condition_code:
                        prod_cond = ''
                        
                    if prod_act != bom.activity:
                        prod_act = ''
                        
                    if prod_notes != bom.notes:
                        prod_notes = ''
                        
                    if prod_task != bom.wot_sequence + ' - ' + bom.task_master_desc:
                        prod_task = ''
                        
                    if prod_stat != bom.wos_auto_key:
                        prod_stat = ''

                count += 1
                
            val_dict.update({
                'part_number': prod_pn,
                'description': prod_desc,
                'qty_needed': prod_qty,
                'condition': prod_cond,
                'activity': prod_act,
                'notes': prod_notes,
                'task_master_desc': prod_task,
                'wos_auto_key': prod_stat,
                'launch_update': 'T',
                'bom_list': wob_id_list,
                'boms': boms,
                'total_rows': len(all_boms),
                'sel_rows': len(boms),
                })
                
            return render(request, 'mrolive/bom_management.html', val_dict)        
                    
        if wo_number and not (bom_status or task_selector\
            or part_number or description or cond_sel\
            or act_sel or qty_need): 
            from portal.tasks import get_wo_bom           
            res = get_wo_bom.delay(quapi_id,session_id,wo_number,task_selector,part_number,description,cond_sel,act_sel,bom_status)            
            error,msg = res.get()  
            boms = WOStatus.objects.filter(session_id = session_id)
            if not boms: 
                error = 'No BoMs found.'
            val_dict.update({
                'boms': boms,
                'total_rows': len(boms),
                'error': error,
                'conditions': conditions,
                'activities': activities,
                'wo_tasks': wo_tasks,
                'bom_statuses': bom_statuses,                 
            }) 
            
        elif is_search == '1': 
        
            from portal.tasks import get_wo_bom        
            res = get_wo_bom.delay(quapi_id,session_id,wo_number,\
                task_selector,part_number,description,cond_sel,act_sel,bom_status)                     
            error,msg = res.get() 
            boms = WOStatus.objects.filter(session_id = session_id)
            if not boms: 
                error = 'No BoMs found.'
                
            val_dict.update({
                'boms': boms,
                'total_rows': len(boms),
                'error': error,
                'wo_number': wo_number,
                'conditions': conditions,
                'activities': activities,
                'wo_tasks': wo_tasks,
                'bom_statuses': bom_statuses, 
                'task_sel': task_sel,
                'act_sel': act_sel,
                'cond_sel': cond_sel,
                'notes': notes,
                'quantity': qty_need,
                'part_number': part_number,
                'description': description,             
                'qty_need': qty_need,                 
            })
            
        elif modal_update:
            #if task_selector:
            #    wo_task = task_selector.split("-",1)
            #    wo_task = wo_task and wo_task[1] or ''
            #    wo_task = wo_task and wo_task.replace(" ","",1) or ''
            if not (bom_status or task_selector or part_number\
                or description or cond_sel or act_sel\
                or qty_need or notes):
                error = 'Must enter values in an update field.'

            elif not wob_id_list and is_search != '1':
                error = 'Select at least one row to update.'
            
            else:            
                from portal.tasks import update_wo_bom                                
                res = update_wo_bom.delay(quapi_id,session_id,wo_number,wob_id_list,\
                    part_number,description,cond_sel,act_sel,qty_need,task_sel,notes,bom_status,\
                    sysur_auto_key,task_selector)
                error,msg = res.get()         
                  
            boms = WOStatus.objects.filter(\
                session_id = session_id,
                wob_auto_key__in=wob_id_list)  
                
            val_dict.update({
                'boms': boms,
                'total_rows': len(boms),
                'error': error,
                'msg': msg,                
                'wo_number': wo_number,
                'part_number': part_number,
                'pn': pn,
                'description': description,                
                'qty_need': qty_need,                               
                'conditions': conditions,
                'activities': activities,
                'wo_tasks': wo_tasks, 
                'bom_statuses': bom_statuses,
                'notes': notes,                
            }) 
        else:
            val_dict['error'] = 'No values entered for update.' 
            
    val_dict['form'] = form
    return render(request, 'mrolive/bom_management.html', val_dict)

def part_management(request,quapi_id):
    error = ''
    user = request and request.user or None
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict)
    form = WODashboardForm()  
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='part-management')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    if request.method == 'GET':
        form = WODashboardForm()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        active_att = req_post.get('att_code','')
        pnm_auto_key = req_post.get('pnm_auto_key','')
        pn = req_post.get('pn','')
        part_number = req_post.get('part_number','')
        description = req_post.get('description','')
        serial_no = req_post.get('serial_no','')
        qty = req_post.get('qty','')
        stm_auto_key = req_post.get('stm_auto_key','')
        condition = req_post.get('condition_code','')     
        serial_number = req_post.get('serial_number','')
        consignment = req_post.get('consignment','')               
        wo_number = req_post.get('wo_number','')
        si_number = req_post.get('si_number','')
        ctrl_number = req_post.get('ctrl_number','')
        ctrl_id = req_post.get('ctrl_id','')
        is_search = req_post.get('is_search','0')
        if not ctrl_number and wo_number and len(wo_number) > 6:
            ctrl_number = wo_number[:6]              
            ctrl_id = wo_number[7:]
        session_id = 'session_id' in req_post and req_post['session_id'] or '' 
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        val_dict.update({
            'part_number': part_number,
            'pnm_auto_key': pnm_auto_key,
            'description': description,
            'pn': pn,
            'qty': qty,
            'condition_code': condition,             
            'serial_number': serial_number,
            'consignment': consignment,
            'stm_auto_key': stm_auto_key,
            'session_id': session_id,
            'wo_number': wo_number,
            'si_number': si_number,
            'ctrl_id': ctrl_id,
            'ctrl_number': ctrl_number,
            'active_att': active_att,
            'serial_no': serial_no,
            })
        if stm_auto_key and not (is_search == '1' and wo_number):               
            stock_id_list = req_post.getlist('woos_list[]')                       
            if len(stock_id_list) > 1:
                val_dict['error'] = 'Can only update one part at a time.'
                return render(request, 'mrolive/part_management.html', val_dict)                 
            from polls.models import UserDefAtts
            atts = UserDefAtts.objects.filter(session_id = session_id)
            att_recs = []

            for att_rec in atts:
                att_val = att_rec.att_name in req_post and req_post[att_rec.att_name] or ''
                if att_val:                                           
                    att = []
                    att.append(att_val)
                    att.append(pnm_auto_key)
                    att.append('pnm')
                    att.append(att_rec.att_name)
                    att.append(part_number or pn)
                    att_recs.append(att)
            if att_recs:
                from portal.tasks import update_attributes
                res = update_attributes.delay(quapi_id,session_id,sysur_auto_key,att_recs)
                error,msg = res.get()
            val_dict['error'] = error
            barcode_value = ctrl_number + '00000' + ctrl_id
            val_dict['element'] = barcode_value
            from portal.tasks import get_part_attributes
            res = get_part_attributes.delay(quapi_id,sysur_auto_key,session_id,pnm_auto_key,stype='one',group='IN HOUSE')
            error,msg = res.get()
            new_atts = UserDefAtts.objects.filter(att_type = 'IN HOUSE',session_id = session_id).order_by('att_seq')
            val_dict['task_list'] = new_atts
            val_dict['app_code'] = 'part-management'   
            if new_atts[0].att_value == new_atts[0].att_name:
                val_dict['task_list'] = []
            printset = app_allow and app_allow[0] and app_allow[0].printset_id
            auth_key = printset and printset.printnode_auth_key
            if printset and auth_key:               
                error = create_barcodes(barcode_value,ctrl_id,ctrl_number,part_number,description,serial_no,qty,si_number,new_atts,'')
                if error:                               
                    val_dict['error'] = error
                else:
                    error = printnode_pdf(printset,auth_key)
                    val_dict['msg'] = 'Attributes Updated. Label printed.'
            else:          
                return render(request, 'mrolive/plain_barcode_mro.html', val_dict)
        elif is_search == '1' and wo_number:
            val_dict.update({
                'wo_number': wo_number,
                'session_id': session_id,
                })               
            from portal.tasks import get_inspection
            res = get_inspection.delay(quapi_id,session_id,wo_number,ctrl_number,ctrl_id)
            error,msg = res.get()
            val_dict['error'] = error
            if not error:
                stock_recs = WOStatus.objects.filter(session_id = session_id)
                stock_rec = stock_recs and stock_recs[0] or None               
                if stock_rec:
                    pnm_auto_key = stock_rec.pnm_auto_key
                    val_dict.update({
                        'total_rows': len(stock_recs),
                        'si_number': stock_rec.si_number,
                        'wo_number': stock_rec.wo_number,
                        'part_number': stock_rec.part_number,
                        'pn': stock_rec.part_number,
                        'description': stock_rec.description,
                        'condition_code': stock_rec.condition_code,      
                        'serial_no': stock_rec.serial_number,
                        'serial_number': stock_rec.serial_number,
                        'consignment': stock_rec.consignment_code,
                        'stm_auto_key': stock_rec.stm_auto_key,
                        'qty': stock_rec.qty_oh,
                        'ctrl_number': stock_rec.ctrl_number,
                        'ctrl_id': stock_rec.ctrl_id,
                        'pnm_auto_key': pnm_auto_key,
                        })                                      
                from portal.tasks import get_part_attributes
                res = get_part_attributes.delay(quapi_id,sysur_auto_key,session_id,pnm_auto_key,stype='all',group='IN HOUSE')
                error,msg = res.get()
                from polls.models import UserDefAtts
                sched_atts = UserDefAtts.objects.filter(att_type = 'SCHEDULING',session_id = session_id).order_by('att_seq')
                route_atts = UserDefAtts.objects.filter(att_type = 'IN HOUSE',session_id = session_id).order_by('att_seq')
                val_dict['sched_atts'] = sched_atts
                val_dict['route_atts'] = route_atts                
    val_dict['form'] = form  
    return render(request, 'mrolive/part_management.html', val_dict) 

def inspection(request,quapi_id,si_number=''):
    error,new_woo_key = '',None
    user = request and request.user or None
    user_id = user and user.username or 'No Username'
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    val_dict = {}

    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')   
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict)
    form = WODashboardForm()  
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='inspection')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    if request.method == 'GET':
        form = WODashboardForm() 
        val_dict.update({
            'wo_number': si_number,
        })          
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        bag_no = req_post.get('spn_code','')
        active_att = req_post.get('att_code','')
        pnm_auto_key = req_post.get('pnm_auto_key','')
        pn = req_post.get('pn','')
        part_number = req_post.get('part_number','')
        description = req_post.get('description','')
        serial_no = req_post.get('serial_no','')
        notes = req_post.get('notes','')
        qty = req_post.get('qty','')
        stm_auto_key = req_post.get('stm_auto_key','')
        str_auto_key = req_post.get('str_auto_key','')
        wob_auto_key = req_post.get('wob_auto_key','')
        condition = req_post.get('cond_selector','')
        conditions = req_post.get('conditions','')        
        condition_code = req_post.get('condition_code','')         
        serial_number = req_post.get('serial_number','')           
        wo_number = req_post.get('wo_number','')
        si_number = req_post.get('si_number','')
        single_stm = req_post.get('single_stm','')
        #wo_number and wo_number.isdigit():
        new_wo = req_post.get('new_wo','') or req_post.get('mod_new_wo','')
        update_insp = req_post.get('update_insp','')
        is_update = req_post.get('is_update','')
        is_search = req_post.get('is_search','')
        quantity = req_post.get('quantity','')
        val_dict['new_wo'] = new_wo and 'checked' or ''
        val_dict['mod_new_wo'] = new_wo and 'checked' or '' 
        is_default_repair = app_allow and app_allow[0] and app_allow[0].default_repair
        opm_list = req_post.getlist('opm_list[]')
        ctrl_number,ctrl_id = '',''
        #if not ctrl_number and wo_number and wo_number.isdigit():
        if wo_number and wo_number.isdigit():

            if len(wo_number) >= 12:
            
                entry = wo_number.rsplit("00000", 1) 
                
                if len(entry) > 1:                
                    ctrl_number = entry[0]
                    ctrl_id = entry[1] 
                
                if not ctrl_number:
                
                    if len(wo_number) > 12:
                        ctrl_number = wo_number[:7]              
                        ctrl_id = wo_number[8:]
                        
                    elif len(wo_number) == 12:
                        ctrl_number = wo_number[:6]
                        ctrl_id = wo_number[7:]                
        
        session_id = req_post.get('session_id','') or req_post.get('update_session','')      
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        val_dict.update({
            'part_number': part_number or pn,
            'pnm_auto_key': pnm_auto_key,
            'description': description,
            'pn': pn or part_number, 
            'qty': qty,
            'quantity': quantity or qty,
            'condition': condition_code, 
            'conditions': conditions,             
            'single_stm': single_stm,
            'notes': notes,
            'serial_number': serial_number,
            'str_auto_key': str_auto_key,
            'wob_auto_key': wob_auto_key,
            'session_id': session_id,
            'wo_number': wo_number or si_number,
            'si_number': si_number,
            'ctrl_id': '',
            'ctrl_number': '',
            'active_att': active_att,
            'serial_no': serial_no,
            'printed_by': user.first_name + ' ' + user.last_name,
            'sel_rows': 0,
            'stm_auto_key': stm_auto_key,
            })
  
        if is_update == '1':
            val_dict['show_modal'] = 'T'
            from portal.tasks import get_activities_conditions_tasks
            res = get_activities_conditions_tasks.delay(
                quapi_id,session_id,si_number=wo_number
            )
            error,msg = res.get()            
            from polls.models import PartConditions
            val_dict['conditions'] = PartConditions.objects.filter(
                session_id = session_id
            )            
                    
        elif update_insp and (part_number or condition or serial_number or notes):
            stm_list = [] 
            if 'woos_list[]' in req_post and req_post.getlist('woos_list[]'):
                stm_list = req_post.getlist('woos_list[]')
                stock_recs = WOStatus.objects.filter(session_id = session_id,id__in = stm_list)                     
            if not stm_list:
                stock_recs = WOStatus.objects.filter(session_id = session_id) 
                stock_recs = stock_recs and stock_recs[0] and [stock_recs[0]] or []                                            
            import random
            import string
            for stock_rec in stock_recs:                
                if part_number == stock_rec.part_number:
                    part_number = ''
                if serial_number == stock_rec.serial_number:
                    serial_number = ''
                if condition == stock_rec.condition_code:
                    condition = ''
                if notes == stock_rec.notes:
                    notes = ''
                if str(stock_rec.qty_oh) == str(quantity):
                    quantity = 0
                  
                if (part_number or condition or serial_number or notes or quantity):
                    session_id = ''.join(random.choices(string.ascii_uppercase + string.digits, k=9))                
                    from portal.tasks import update_inspection
                    res = update_inspection.delay(user_id,quapi_id,\
                    session_id,sysur_auto_key,stock_rec.stm_auto_key,\
                    wo_number,stock_rec.ctrl_number,stock_rec.ctrl_id,\
                    bag_no=bag_no,part_number=part_number,condition=condition,\
                    serial_number=serial_number,\
                    qty_oh=stock_rec.qty_oh,notes=notes,del_previous=True,quantity=quantity)
                    error,msg = res.get()

                if error:
                
                    val_dict.update({
                        'wo_number': wo_number,
                        'session_id': session_id, 
                        'wo_task': si_number or wo_number,                  
                        'error': error,
                        'total_rows': len(stock_recs)
                    })
                    return render(request, 'mrolive/inspection.html', val_dict)
                    
            stock_recs = WOStatus.objects.filter(session_id = session_id)
            stock_rec = stock_recs and stock_recs[0] or None
            printed_date = datetime.now()
            printed_date = printed_date.strftime('%m/%d/%Y %H:%M:%S')

            if stock_recs:
                app_code='inspection'
                tmpl_code='inspection-label'
                barcode_txt =  '0' + stock_rec.ctrl_number + '00000' + stock_rec.ctrl_id

                val_dict.update({
                    'element': barcode_txt,
                    'stock_recs': stock_recs,
                    'wo_number': '',
                    'si_number': wo_number,
                    'ctrl_id': '',
                    'ctrl_number': '',
                    'printed_date': printed_date,
                    'inspector': user.first_name + ' ' + user.last_name,
                    'printed_by': user.first_name + ' ' + user.last_name,
                    'session_id': session_id, 
                    'inspection': 'T',
                    'wo_task': si_number or wo_number,
                    'records': stock_recs,  
                    'app_code': app_code,
                    'tmpl_code': tmpl_code,                    
                })

                #printset = app_allow and app_allow[0] and app_allow[0].printset_id
                #auth_key = printset and printset.printnode_auth_key
                                           
                #error = create_barcodes_pymu(val_dict,app_code,tmpl_code,barcode_txt=barcode_txt,x_delta=5.5)
                
                if 0:
                    #error = printnode_pdf(printset,auth_key) 
                    val_dict['error'] = error
                    val_dict['msg'] = 'Successful Inspection. Label printed.'
                    
                #elif user_id == 'JBAXT':                
                    
                #    val_dict['error'] = error
                    
                #    if not error:
                #        val_dict['msg'] = 'Succesful Update. 8130-3 report created.'                    
                                
                val_dict['error'] = error   
                #return render(request, 'mrolive/inspection.html', val_dict)  
                if not new_wo:                
                    return render(request, 'mrolive/plain_barcode_mro.html', val_dict)
                else:
                    return render(request, 'mrolive/teardown_traveller_mro_header.html', val_dict) 
                
            else:
                return render(request, 'mrolive/plain_barcode_mro.html', val_dict)
                
            if new_wo:
                from portal.tasks import create_new_wo
                res = create_new_wo.delay(quapi_id,session_id,user_id,sysur_auto_key,stm_auto_key,wo_number,ctrl_number,ctrl_id,is_default_repair) 
                error,msg,new_woo_key = res.get()
                val_dict['new_woo_key'] = new_woo_key
                
            if not is_default_repair and not opm_list:            
                updated_opms = Operation.objects.filter(session_id=session_id)            
                val_dict.update({        
                    'total_rows': str(len(updated_opms))
                })
            if new_woo_key and is_default_repair:
                #call the task method and pass the opm selected by the user
                if opm_list:
                    opm_auto_key = opm_list[0]
                    from portal.tasks import create_opm_tasks
                    res = create_opm_tasks.delay(session_id,sysur_auto_key,quapi_id,opm_auto_key,new_woo_key) 
                    error,opm_msg = res.get() 
                    msg += opm_msg
                    val_dict.update({
                        'error': error,
                        'msg': msg,
                        })
                    #return render(request, 'mrolive/inspection.html', val_dict)                    
                else:
                    val_dict['error'] = 'No operation selected.'
                    return render(request, 'mrolive/inspection.html', val_dict)  
            printset = app_allow and app_allow[0] and app_allow[0].printset_id
            if stock_recs and printset and printset.printnode_auth_key:
                if not error:
                    val_dict['msg'] = 'Inspection performed. Label printed.'
                else:
                    val_dict['error'] = error
       
        elif is_search == '1' and wo_number:
                   
            val_dict.update({
                'wo_number': wo_number,
                'session_id': session_id,
                }) 
                            
            from portal.tasks import get_inspection             
            res = get_inspection.delay(quapi_id,session_id,\
                wo_number,ctrl_number,ctrl_id,del_previous=True)
            error,msg,hold_line = res.get()
            val_dict['error'] = error
            if not error:
                stock_rec = WOStatus.objects.filter(session_id = session_id)
                val_dict['total_rows'] = len(stock_rec)  
                stock_rec = stock_rec and stock_rec[0] or None               
                if stock_rec:
                    pnm_auto_key = stock_rec.pnm_auto_key  
                    val_dict.update({
                        'si_number': stock_rec.si_number,
                        'part_number': stock_rec.part_number,
                        'pn': stock_rec.part_number,
                        'description': stock_rec.description,
                        'condition_code': stock_rec.condition_code, 
                        'cond_level': stock_rec.cond_level,                        
                        'notes': stock_rec.notes,                                              
                        'serial_number': stock_rec.serial_number,
                        'parent_serial': stock_rec.slug,
                        'stm_auto_key': stock_rec.stm_auto_key,
                        'str_auto_key': stock_rec.str_auto_key,
                        'wob_auto_key': stock_rec.wob_auto_key,
                        'qty': stock_rec.qty_oh,
                        'ctrl_number': stock_rec.ctrl_number,
                        'ctrl_id': stock_rec.ctrl_id,
                        'pnm_auto_key': pnm_auto_key,
                        'rec_date': stock_rec.arrival_date,
                        'po_number': stock_rec.po_number,
                        'ro_number': stock_rec.ro_number,
                        'spn_code': stock_rec.spn_code,
                        })
            else:
                val_dict['wo_number'] = ''
        
    val_dict['form'] = form  
    return render(request, 'mrolive/inspection.html', val_dict) 
 
def email_send(request,quapi_id):
    user = request and request.user or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    form = WODashboardForm() 
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='users')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post) 
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        email = req_post.get('email','')
        message = req_post.get('override','')
        subject = req_post.get('label','') 
        attach = request.FILES['file']        
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        try:
            from django.core.mail import EmailMessage
            mail = EmailMessage(subject, message, settings.EMAIL_HOST_USER, [email])
            mail.attach(attach.name, attach.read(), attach.content_type)
            mail.send() 
            val_dict['msg'] = 'Message sent!'
        except:
             val_dict['error'] = 'Either your file is too large or corrupt'
    val_dict['form'] = form  
    return render(request, 'mrolive/email_send.html', val_dict) 
    
def email_attachment(quapi_id,session_id,sysur_auto_key,req_files):
    error,success_msg,fail_msg = '','','' 
    up_file = 'file' in req_files and req_files['file'] or None
    att_save = up_file and Document(docfile=up_file) or None
   
    if att_save:
        att_save.save()     
        file_path = '/home/ubuntu/mo_template/media/' + up_file.name
        file_path = file_path.replace(' ','_')               
        #now that we no longer need the new file, delete it!
        os.remove(file_path)
    return error,success_msg,fail_msg  
  
def wo_task_create(request,quapi_id,wo_number=''):
    wot_auto_key = ''
    user = request and request.user or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    form = WODashboardForm() 
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='non-routine')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    from portal.tasks import create_wo_tasks
    session_id = '1234567' 
    res = create_wo_tasks.delay(quapi_id,1,session_id)
    error = res.get()        
    wtms = WOTask.objects.filter(session_id=session_id).order_by('wot_sequence') or []
    #new_sequence = wtms and wtms[0].wot_sequence and int(wtms[0].wot_sequence) + 1
    val_dict.update({
        'wtms': wtms,
        #'sequence': new_sequence,
        #'wot_sequence': new_sequence,
        })
    if request.method == 'GET':
        val_dict['wo_number'] = wo_number
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        val_dict['form'] = form         
        session_id = 'session_id' in req_post and req_post['session_id'] or ''       
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        wo_number = req_post.get('wo_number','')
        customer = req_post.get('customer','')
        part_number = req_post.get('part_number','')
        description = req_post.get('description','')
        serial_no = req_post.get('serial_no','')
        si_number = req_post.get('si_number','')
        task_desc = req_post.get('override','')
        sequence = req_post.get('sequence','')
        wot_sequence = req_post.get('wot_sequence','')
        wtm_auto_key = req_post.get('description_selector','')
        show_task_entry = req_post.get('show_task_entry','')
        is_search = req_post.get('is_search','0')
       
        if wo_number and is_search == '1':
        
            from portal.tasks import create_wo_tasks
            res = create_wo_tasks.delay(quapi_id,1,session_id,wo_number=wo_number)
            error = res.get()           
            wots = WOTask.objects.filter(session_id=session_id).order_by('-wot_sequence') or []
            
            wot_sequence = wots and wots[0].wot_sequence and int(wots[0].wot_sequence) + 1 or 1
            val_dict.update({
                'show_task_entry': 'T',
                'wot_sequence': wot_sequence or sequence,
                'sequence': wot_sequence or sequence,
                'error': error,
                'session_id': session_id,
                'wo_number': wo_number,
                'serial_no': serial_no,
                'customer': customer,
                'part_number': part_number,
                'description': description,                
            })           
        elif task_desc and wtm_auto_key and sequence and is_search != '1':
            from portal.tasks import task_insertion,get_task_mgmt
            task = {}
            today = datetime.now()
            today = today.strftime('%m/%d/%Y')
            task['start_date'] = today
            task['skill'] = ''
            task['task_desc'] = task_desc
            task['wot_sequence'] = sequence or wot_sequence
            task['wtm_auto_key'] = wtm_auto_key
            task['task_position'] = ''
            task['task_title'] = ''
            task['task_close_reqs'] = ''
            task['skills_est_hours'] = 0
            task_list = [task]
            res = task_insertion.delay(quapi_id,session_id,sysur_auto_key,task_list,wo_number)   
            error,msg,fail_msg,bad_rows,wot_auto_key = res.get()
            if error:
                val_dict['error'] = error  
                val_dict['form'] = form                                        
                return render(request, 'mrolive/wo_task_create.html', val_dict)
            else:
                msg = 'Task: %s, with sequence %s created.'%(task_desc,task['wot_sequence'])
                val_dict['msg'] = msg 
            parameters = ['','','','','','','']                    
            res = get_task_mgmt.delay(quapi_id,session_id,\
                parameters,wot_list=[wot_auto_key])                     
            task_error,task_msg = res.get()            
            tasks = WOTask.objects.filter(session_id = session_id,\
                wot_auto_key__in=[wot_auto_key])

            task = tasks and tasks[0]
            esn = task and task.esn or ''
            customer = task and task.customer or ''
            esn = task and task.esn or ''
            customer = task and task.customer or ''
            eng_model = task and task.eng_model or ''
            ac_reg = task and task.ac_reg or '' 
            ac_model = task and task.ac_model or '' 
            ac_sn = task and task.ac_sn or ''
            part_description = task and task.part_description or ''

            val_dict.update({
                'wo_number': wo_number,
                'si_number': wo_number,
                'esn': esn,
                'description': part_description,
                'customer': customer,
                'eng_model': eng_model,
                'ac_reg': ac_reg, 
                'ac_model': ac_model, 
                'ac_sn': ac_sn,
                'task_list': [task],
                'error': error + task_error,
                'msg': msg + task_msg,
                'session_id': session_id,
                'override': '',
                'show_task_entry': show_task_entry,
                'app_name': 'non-routine',
                'quapi_id': quapi_id,
            })            
            return render(request, 'mrolive/non_routine_traveller_mtu.html', val_dict)  
            
        elif wo_number and show_task_entry:
            error = 'Must enter task, description and sequence.'
            val_dict.update({
                'show_task_entry': show_task_entry,
                'wot_sequence': wot_sequence or sequence,
                'sequence': wot_sequence or sequence,
                'session_id': session_id,
                'wo_number': wo_number,
                'description_selector': wtm_auto_key,
                'override': task_desc,                
                'error': error,
            })
                
    val_dict['form'] = form
    return render(request, 'mrolive/wo_task_create.html', val_dict)    
    
def part_attributes(request,quapi_id):
    user = request and request.user or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    form = WODashboardForm() 
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='users')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post) 
        session_id = 'session_id' in req_post and req_post['session_id'] or ''       
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        if request.FILES:        
            error,msg,fail_msg = attributes_create(quapi_id,session_id,sysur_auto_key,request.FILES)  
            val_dict['error'] = error
            val_dict['msg'] = msg
            val_dict['fail_msg'] = fail_msg            
    val_dict['form'] = form  
    return render(request, 'mrolive/part_attributes.html', val_dict) 
    
def attributes_create(quapi_id,session_id,sysur_auto_key,req_files):
    error,success_msg,fail_msg,key_msg = '','','','' 
    up_file = 'file' in req_files and req_files['file'] or None
    file_save = up_file and Document(docfile=up_file) or None
   
    if file_save:
        file_save.save()     
        file_path = '/home/ubuntu/mo_template/media/' + up_file.name
        file_path = file_path.replace(' ','_')
        from portal.tasks import synch_attributes
        #import csv
        from openpyxl import load_workbook
        wb = load_workbook(filename = file_path)
        sheet = wb.worksheets[0]
        sheet_rows = sheet.iter_rows()
        row_list = []        
        row_vals = [[v.value for v in row] for row in sheet_rows]
        found_headings = False
        from portal.tasks import get_att_keys
        for row in row_vals: 
            if row[0] == 'PN':
                found_headings = True
                #get the headings and check to make sure they exist as attribute names
                res = get_att_keys.delay(quapi_id,row)
                att_key_list,key_msg = res.get()
                #att_key_list = att_keys.split(",")
                continue          
            if not found_headings:
                continue
            else:                               
                res = synch_attributes.delay(quapi_id,sysur_auto_key,session_id,row,att_key_list)
                error,msg,imp_msg = res.get()
                fail_msg += imp_msg
                success_msg += msg
        if not found_headings:
            error = "First column heading must be \'PN\'."            
        #now, if there is a csv that was saved, delete it!
        os.remove(file_path)
    return error,success_msg,key_msg + fail_msg
    
def print_settings(request,quapi_id):
    user = request and request.user or None
    val_dict = {}
    val_dict['quapi_id'] = quapi_id
    from django.contrib.auth import get_user_model
    User = get_user_model()
    user_id = user and user.id
    user_name = User.objects.filter(id=user_id)[0]    
    val_dict['user_name'] = user_name
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    form = WODashboardForm()
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    val_dict['user_apps'] = UserAppPerms.objects.filter(user=user_id,ml_apps_id__print_enabled=True) or []
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='users')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id) 
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')       
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        printer_name = request.POST.get('printer_name','')
        computer_name = request.POST.get('computer_name','')
        auth_key = request.POST.get('auth_key','')
        #take each app_assign and assign the printer_id
        if printer_name and auth_key:
            from polls.models import PrintSetting
            pset = PrintSetting(
                printer_name = printer_name,
                computer_name = computer_name,
                printnode_auth_key = auth_key,
            )
            error = pset.save()
            val_dict['error'] = error
            if not error:
                val_dict['msg'] = 'Successful update.'
    val_dict['form'] = form
    return render(request, 'registration/print_settings.html', val_dict)           

def user_prints(request,user_id,quapi_id):
    user = request and request.user or None
    val_dict = {}
    val_dict['quapi_id'] = quapi_id
    from django.contrib.auth import get_user_model
    User = get_user_model()
    user_name = User.objects.filter(id=user_id)
    user_name= user_name and user_name[0] or ''                                         
    val_dict['user_name'] = user_name
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    form = WODashboardForm()
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    val_dict['user_apps'] = UserAppPerms.objects.filter(user=user_id,ml_apps_id__print_enabled=True) or []
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='users')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id) 
    #assigned_apps = uap.objects.filter(user=user_id)
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    #if request.method == 'GET':       
    from polls.models import PrintSetting
    print_settings = PrintSetting.objects.all() 
    val_dict['print_settings'] = print_settings        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        printer_assigns = request.POST.getlist('printer_assigns')
        app_assigns = request.POST.getlist('app_assigns')
        #take each app_assign and assign the printer_id
        if printer_assigns:
            from polls.models import PrintSetting
            for uap_id in app_assigns:
                uap = UserAppPerms.objects.get(id=uap_id)
                uap.printset_id = PrintSetting.objects.get(id=printer_assigns[0])
                error = uap.save()
                val_dict['error'] = error
            if not error:
                val_dict['msg'] = 'Successful update.'
    return render(request, 'registration/user_prints.html', val_dict)           

def user_connections(request,user_id,quapi_id):
    user = request and request.user or None
    val_dict = {}
    val_dict['quapi_id'] = quapi_id
    from django.contrib.auth import get_user_model
    User = get_user_model()
    user_name = User.objects.filter(id=user_id)[0]    
    val_dict['user_name'] = user_name
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    form = WODashboardForm()
    uqp = UserQuapiRel
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='users')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id) 
    #assigned_apps = uap.objects.filter(user=user_id)
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    if request.method == 'GET':       
        rem_conns = uqp.objects.filter(user=user_id)
        ass_conns = []
        for qconn in QueryApi.objects.all():
            user_conn = uqp.objects.filter(quapi_id=qconn.id,user=user_id)
            if not user_conn:
                ass_conns.append(qconn)
        val_dict['ass_conn'] = ass_conns
        val_dict['rem_conn'] = rem_conns        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        ass_user_id = request.POST.getlist('ass_user_id')
        ass_conn = request.POST.getlist('ass_conn')
        rem_conn = request.POST.getlist('rem_conn')
        if not (ass_conn or rem_conn):
            val_dict['error'] = 'Select connections to assign or remove.'
        #lookup sysur_auto_key from existing Quantum users?
        for conn in ass_conn:
            #create the user permissions record
            #save it
            #update the apps_assigned list with the apps newly assigned
            admin_uid = User.objects.filter(id=user_id)[0]            
            q_id = QueryApi.objects.filter(id=conn)[0]           
            user_quapi = uqp(
                dj_username=admin_uid.username + ' - ' + q_id.name,
                quapi_id=q_id,
                user=admin_uid,
                dj_user_id=user_id,
            )
            user_quapi.save()
        if rem_conn:
            #create the user permissions record
            #save it
            #update the apps_assigned list with the apps newly assigned           
            dels = uqp.objects.filter(id__in=rem_conn).delete()                                  
        rem_conns = uqp.objects.filter(user=user_id)
        ass_conns = []
        for user_conn in QueryApi.objects.all():
            user_quapi = uqp.objects.filter(quapi_id=user_conn.id,user=user_id)
            if not user_quapi:
                ass_conns.append(user_conn)
        val_dict['ass_conn'] = ass_conns
        val_dict['rem_conn'] = rem_conns
    return render(request, 'registration/user_connections.html', val_dict)           
                            
def user_apps(request,user_id,quapi_id):
    user = request and request.user or None
    val_dict = {}
    val_dict['quapi_id'] = quapi_id
    from django.contrib.auth import get_user_model
    User = get_user_model()
    user_name = user_id and User.objects.filter(id=user_id) or None
    user_name = user_name and user_name[0] and user_name[0].username    
    val_dict['user_name'] = user_name
    if not user_name:
        val_dict['error'] = 'No user name found.'
        render(request, 'registration/user_apps.html', val_dict)     
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    form = WODashboardForm()
    from polls.models import UserAppPerms as uap
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and uap.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='users')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    #assigned_apps = uap.objects.filter(user=user_id)
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    if request.method == 'GET':       
        apps_to_remove = uap.objects.filter(user=user_id)
        apps_to_assign = []
        for ml_app in MLApps.objects.all():
            user_app_perm = uap.objects.filter(ml_apps_id=ml_app,user=user_id)
            if not user_app_perm:
                apps_to_assign.append(ml_app)
        val_dict['apps_to_assign'] = apps_to_assign
        val_dict['apps_to_remove'] = apps_to_remove        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        apps_to_assign = req_post.getlist('apps_to_assign')
        apps_to_remove = req_post.getlist('apps_to_remove')           
        if not (apps_to_assign or apps_to_remove):
            val_dict['error'] = 'Select app(s) to assign or remove.'
        for app in apps_to_assign:
            #create the user permissions record
            #save it
            #update the apps_assigned list with the apps newly assigned
            from django.contrib.auth import get_user_model
            User = get_user_model() 
            admin_uid = User.objects.filter(id=user_id)[0]          
            ml_apps_id = MLApps.objects.filter(id=app)[0]           
            user_perm = uap(
               audit_ok=True,
               dj_username=admin_uid.username + ' - ' + ml_apps_id.name,
               ml_apps_id=ml_apps_id,
               user=admin_uid,
               dj_user_id=user_id)
            user_perm.save()
        if apps_to_remove:
            #create the user permissions record
            #save it
            #update the apps_assigned list with the apps newly assigned           
            dels = uap.objects.filter(id__in=apps_to_remove).delete()                                  
        apps_to_remove = uap.objects.filter(user=user_id)
        apps_to_assign = []
        for ml_app in MLApps.objects.all():
            user_app_perm = uap.objects.filter(ml_apps_id=ml_app,user=user_id)
            if not user_app_perm:
                apps_to_assign.append(ml_app)
        val_dict.update(
        {
            'apps_to_assign': apps_to_assign,
            'apps_to_remove': apps_to_remove,
        })
    return render(request, 'registration/user_apps.html', val_dict)           
                              
def user_profiles(request,quapi_id):
    error,msg,set_error,save_error = '','','',''
    user = request and request.user or None
    val_dict = {}
    val_dict['quapi_id'] = quapi_id
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    #if request.method == 'GET':
    form = WODashboardForm() 
    #look up all users and return them in val_dict for the grid
    from django.contrib.auth import get_user_model
    User = get_user_model()
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='users')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    users = UserProfile.objects.all()
    total_rows = len(users)

    for user_profile in users:
        apps = UserAppPerms.objects.filter(user=user_profile.user)
        num_apps = apps and len(apps) or 0
        user_profile.num_apps = num_apps 
        user_profile.save()
        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        first_name = 'first_name' in req_post and req_post['first_name']
        last_name = 'last_name' in req_post and req_post['last_name']
        username = 'username' in req_post and req_post['username']
        user_id = 'user_id' in req_post and req_post['user_id']
        email = 'email' in req_post and req_post['email']
        password = 'password' in req_post and req_post['password']
        password_reentry = 'password_reentry' in req_post and req_post['password_reentry']
        kiosk = req_post.get('assign_kiosk','0')
        val_dict.update({
            'first_name':first_name,
            'last_name':last_name,
            'username': username,
            'email': email,   
            'user_id': user_id, 
            'form': form, 
            'assign_kiosk': kiosk,            
        })
        if username and (first_name or last_name or email or password or kiosk):          
            user_profile = UserProfile.objects.filter(user_name=username) or None   
            user_profile = user_profile and user_profile[0] or None
            user_change = user_profile and user_profile.user or None            
               
            if user_change:
                if first_name:
                    user_change.first_name = first_name
                if last_name:
                    user_change.last_name = last_name
                if username:
                    user_change.username = username
                if email:
                    user_change.email = email
                if password:
                    if password != password_reentry:
                        error = 'Passwords must match.'
                    # calculating the length
                    verify = len(password) < 8
                    if verify:
                        error += ' Must be at least 8 characters.'
                    # searching for digits
                    verify = re.search(r"\d", password) is None
                    if verify:
                        error += ' Must contain at least one number.'
                    # searching for uppercase
                    verify = re.search(r"[A-Z]", password) is None
                    if verify:
                        error += ' Must have at least one uppercase letter.'
                    # searching for lowercase
                    verify = re.search(r"[a-z]", password) is None
                    if verify:
                        error += ' Must have at least one lowercase letter.'
                    # searching for symbols
                    verify = re.search(r"[ !#$%&'()*+,-./[\\\]^_`{|}~"+r'"]', password) is None
                    if verify:
                        error += ' Must have at least one special character (symbol).'
                        
                # overall result               
                if not error:
                    if password:
                        set_error = user_change.set_password(password)
                    user_change.save() 
                    user_profile.user_name = user_change.username
                    user_profile.first_name = user_change.first_name
                    user_profile.last_name = user_change.last_name
                    user_profile.email = user_change.email
                    user_profile.user_key = user_change.id
                    is_kiosk = False
                    kiosk_check = 'blank.png'
                    if kiosk == '1':
                        #find the user_profile record and assign Labor Management
                        labor = MLApps.objects.filter(name='Labor Tracking')
                        kiosk_check = 'green-check.png'                                                                         
                        user_profile.kiosk_apps.set(labor)  
                        val_dict['assign_kiosk'] = 'checked'
                    else:
                        val_dict['assign_kiosk'] = '0'
                        non_labor = MLApps.objects.filter(code='barcoding')
                        user_profile.kiosk_apps.set(non_labor) 
                    user_profile.is_kiosk = is_kiosk
                    user_profile.kiosk_check = kiosk_check
                    user_profile.save()
                    
                    if not error:                    
                        msg = 'User updated successfully.'
		   
    val_dict.update({
        'total_rows': total_rows,
        'error': error,
        'user': user,
        'form': form,
        'msg': msg,
    })  
    return render(request, 'registration/user_profiles.html', val_dict)  
    
def pass_reset(request,quapi_id):
    error,msg,set_error,save_error = '','','',''
    user = request and request.user
    username = user and user.username
    val_dict = {}
    val_dict['quapi_id'] = quapi_id
    dj_user_id = user and user.id
    val_dict['dj_user_id'] = dj_user_id
    val_dict['username'] = username
    if not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    form = WODashboardForm() 
    #look up all users and return them in val_dict for the grid
    from django.contrib.auth import get_user_model
    User = get_user_model() 
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    #app_id = MLApps.objects.filter(code='users')[0]
    #app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    #if not (reg_user_id and app_allow):
        #val_dict['error'] = 'Access denied.'
        #return redirect('/login/')    
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        username = 'username' in req_post and req_post['username']
        user_id = 'user_id' in req_post and req_post['user_id']
        password = 'password' in req_post and req_post['password']
        password_reentry = 'password_reentry' in req_post and req_post['password_reentry']
        val_dict.update({
            'username': username, 
            'user_id': user_id, 
            'form': form,           
        })
        if user_id and password:
            user_change = User.objects.filter(id=user_id)
            user_change = user_change and user_change[0] or None
            if password:
                if password != password_reentry:
                    error = 'Passwords must match.'
                    #return render(request, 'registration/pass_reset.html', val_dict)
                # calculating the length
                verify = len(password) < 8
                if verify:
                    error += ' Must be at least 8 characters.'
                # searching for digits
                verify = re.search(r"\d", password) is None
                if verify:
                    error += ' Must contain at least one number.'
                # searching for uppercase
                verify = re.search(r"[A-Z]", password) is None
                if verify:
                    error += ' Must have at least one uppercase letter.'
                # searching for lowercase
                verify = re.search(r"[a-z]", password) is None
                if verify:
                    error += ' Must have at least one lowercase letter.'
                # searching for symbols
                verify = re.search(r"[ !#$%&'()*+,-./[\\\]^_`{|}~"+r'"]', password) is None
                if verify:
                    error += ' Must have at least one special character (symbol).'
                # overall result               
                if not error and user_change:
                    set_error = user_change.set_password(password)
                    user_change.save()
            if not error:                    
                save_error = user_change.save()
                msg = 'Password changed successfully.'          
    val_dict.update({
        'error': error,
        'user': user,
        'form': form,
        'msg': msg,
    }) 
    return render(request, 'registration/pass_reset.html', val_dict) 
       
def shop_mgmt_dashboard(request,quapi_id,loc_text):
    user = request and request.user or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='shop-management')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    val_dict['quapi_id'] = quapi_id  
    val_dict['user_name'] = user_name  
    if request.method == 'GET':
        form = WODashboardForm()
        session_id = request.session.session_key       
        from portal.tasks import get_loc_stocks
        res = get_loc_stocks.apply_async(
            queue='refresh', 
            priority=1,
            args=[quapi_id,session_id,loc_text],
            )
        error,msg = res.get()
        all_woos = WOStatus.objects.filter(session_id = session_id)
        total_rows = len(all_woos)
        val_dict.update({
            'quapi_id': quapi_id,  
            'session_id': session_id, 
            'all_woos': all_woos, 
            'total_rows': total_rows,            
        })
    val_dict['form'] = form                
    return render(request, 'mrolive/shop_mgmt_dashboard.html', val_dict)  
            
def shop_management(request,quapi_id):
    user = request and request.user or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='shop-management')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    val_dict['quapi_id'] = quapi_id  
    val_dict['user_name'] = user_name  
    if request.method == 'GET':
        form = WODashboardForm()         
        """loc_list = ['INCOMING INSP','AWT APPROVAL','ENGINEERING']
        loc_list += ['CYLINDER BLOCK','MANUAL MACHINES','JIG GRINDER']
        loc_list += ['CNC', 'WELDING', 'FPI']
        loc_list += ['COLD SPRAY','GRINDER','BLENDING']
        loc_list += ['OUTSIDE SERVICE','FINAL INSP','SHIPPING']"""
        loc_list = ['DISASSEMBLY','CLEANING','NDT']
        loc_list += ['INSPECTION','MACHINE SHOP','ASSEMBLY']
        loc_list += ['PAINT','FINAL']        
        from portal.tasks import get_loc_counts
        res = get_loc_counts.apply_async(
            queue='refresh', 
            priority=1, 
            args=[quapi_id,loc_list],
            )
        error,counts = res.get()
        itr = 0
        loc_counts = {}
        for count in counts:
            loc_counts[loc_list[itr]] = count
            itr += 1
        val_dict.update({
            'loc_list': loc_list,
            'loc_counts': loc_counts,
            'quapi_id': quapi_id,         
        })
    val_dict['form'] = form                
    return render(request, 'mrolive/shop-management.html', val_dict)   
    
def wo_order_clause(request,quapi_id):
    error = ''
    user = request and request.user or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='wo-order-clause')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    val_dict['quapi_id'] = quapi_id  
    val_dict['user_name'] = user_name
    if request.method == 'GET':
        form = WODashboardForm()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        override_selector = 'override_selector' in req_post and req_post['override_selector'] or ''
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0
        override = 'override' in req_post and req_post['override'] or ''
        update_session_id = 'update_session_id' in req_post and req_post['update_session_id'] or ''
        if not session_id and not update_session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
            update_session_id = session_id
        woo_auto_key = ''
        woo_id_list = []
        val_dict.update({
            'session_id': session_id or update_session_id,
            'update_session_id': update_session_id or session_id,
            'wo_number': wo_number, 
            'total_rows': total_rows,
            'sel_rows': sel_rows,            
        })
        #if 'woos_list[]' in req_post:
        if override_selector and override:
            """woo_id_list = req_post.getlist('woos_list[]') 
            if len(woo_id_list) > 1:
                val_dict['form'] = form  
                val_dict['error'] = "Select one WO for update."    
                return render(request, 'mrolive/wo_order_clause.html', val_dict)
            """
            woo = WOStatus.objects.filter(id=override_selector) 
            woo = woo and woo[0]           
            woo_auto_key = woo and woo.woo_auto_key or None
            sequence = woo and woo.wot_sequence or None           
            if woo_auto_key:            
                from portal.tasks import set_overrides             
                res = set_overrides.delay(quapi_id,session_id,woo_auto_key,override,sequence,sysur_auto_key)
                error,msg = res.get()                     
        elif wo_number and not woo_auto_key:
            from portal.tasks import get_overrides        
            res = get_overrides.delay(quapi_id,session_id,wo_number)
            error,msg = res.get()
        #get all overrides 
        requests = WOStatus.objects.filter(session_id=session_id).order_by('wot_sequence')
        val_dict['requests'] = requests            
    val_dict['form'] = form
    val_dict['error'] = error    
    return render(request, 'mrolive/wo_order_clause.html', val_dict) 

def file_retrieval(request,quapi_id):
    user = request and request.user or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='file-retrieval')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    val_dict['quapi_id'] = quapi_id  
    val_dict['user_name'] = user_name    
    if request.method == 'GET':
        form = WODashboardForm()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)   
         
        wo_number = req_post and req_post['wo_number'] or ''
        session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        from zipfile import ZipFile
        from os.path import basename
        path = 'C:/QuantumDocImages'
        file_list = os.listdir(path)
        right_now = datetime.now()
        right_now = right_now.strftime('%m-%d-%Y %H:%M:%S')
        
        #file1 = 'C:/Users/mrolive/Documents/mo_template/uploads/barcode.pdf'
        #file2 = 'C:/media/test.txt'
        #with ZipFile('C:/Users/mrolive/Documents/mo_template/uploads/sample_test.zip', 'w') as zip_obj:
           # Add multiple files to the zip
        from portal.tasks import file_retrieval
        import urllib.request as urllib
        from io import StringIO
        #response = HttpResponse(f.getvalue(), content_type="application/zip")
        #response['Content-Disposition'] = 'attachment; filename=foobar.zip'
        res = file_retrieval.delay(quapi_id,wo_number,session_id) 
        error,msg = res.get()
        val_dict.update({
            'error': error,
            'msg': msg,
            })
        zip_obj = None
        
        if not error:
            download_to = 'C:/QuantumDocImages/'        
            zip_obj = ZipFile(download_to + wo_number + '_' + right_now + '_files.zip', 'w')
            
            for q_file in file_list:
                
                file_stamp = os.path.getmtime(path + '/' + q_file)   
                file_date = datetime.fromtimestamp(file_stamp)
                start_date = file_date - timedelta(seconds = 15)                
                start_date = start_date.strftime('%Y-%m-%d %H:%M:%S')
                end_date = file_date + timedelta(seconds = 15)
                end_date = end_date.strftime('%Y-%m-%d %H:%M:%S')                
                file_match = Document.objects.filter(session_id=session_id,create_date__range=[start_date,end_date])
                file_match = file_match and file_match[0] or None
                
                if file_match:              
                    zip_obj.write(path + '/' + q_file,q_file)
                    
            if zip_obj:
                val_dict['msg'] = 'Successful download'

    val_dict['form'] = form                
    return render(request, 'mrolive/file_zip.html', val_dict)    
    
def file_insert(request,quapi_id):
    user = request and request.user or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='file-import')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
        
    val_dict['quapi_id'] = quapi_id  
    val_dict['user_name'] = user_name    
    if request.method == 'GET':
        form = WODashboardForm()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)   
         
        wo_number = req_post and req_post['wo_number'] or ''
        session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        from zipfile import ZipFile
        from os.path import basename
        path = 'C:/QuantumDocImages'
        file_list = os.listdir(path)
        right_now = datetime.now()
        right_now = right_now.strftime('%m-%d-%Y %H:%M:%S')

        from portal.tasks import file_import
        #need uploader
        res = file_insert.delay(quapi_id,session_id) 
        error,msg = res.get()
        val_dict.update({
            'error': error,
            'msg': msg,
            })
        zip_obj = None
        
        if not error:
            download_to = 'C:/QuantumDocImages/'        

                    
            if zip_obj:
                val_dict['msg'] = 'Successful Upload'

    val_dict['form'] = form                
    return render(request, 'mrolive/file_import.html', val_dict)
    
def create_insp_barcodes(code_vals):
    """
    Condition: Factory New
    CTRL #:10063       CTRL ID: 1
    Serial: 100634_2   Part: 5003159
    Desc: PL SA P
    Qty: 1.0           EX ESN: Trace2
    Stock Line: 4      WO: 010063000001
    NOTES: 
    """
    from reportlab.graphics.barcode import code39
    from reportlab.lib.pagesizes import A6,landscape
    from reportlab.lib.units import mm,inch
    from reportlab.pdfgen import canvas         
    barcode39Std = code39.Standard39(code_vals['element'], barHeight=75, barWidth=1.43, stop=1, checksum=0)  
    #set the filepath and logo_path for the pdf label file.
    file_path = "/home/ubuntu/mo_template/uploads/barcode.pdf" 
    logo_path = "/home/ubuntu/mo_template/static/bischoff_logo.jpg" 
    c = canvas.Canvas(file_path, pagesize=A6)
    c.setPageSize((105*mm,153*mm))
    #c.rotate(90)
    code = barcode39Std
    x = 5 * mm
    y = 125 * mm
    c.setFont("Courier",14)
    for srec in code_vals['stock_recs']:
        code.drawOn(c, -2*mm, y)
        y = y - 5 * mm
        #c.drawString(32*mm + x, y, code_vals['element'])       
        if srec.cond_level_gsix == 'T':
            c.setFont("Courier",20)
            c.drawString(19*mm + x, y - 4 * mm,'SERVICEABLE')
            c.setFont("Courier",14)
        else:
            c.setFont("Courier",20)
            c.drawString(18*mm + x, y - 4 * mm,'UNSERVICEABLE')	
            c.setFont("Courier",14)            
        c.drawString(x, y - 12 * mm, 'CTRL#: ' + str(srec.ctrl_number))
        c.drawString(65 * mm, y - 12 * mm,'CTRL ID: ' + str(srec.ctrl_id)) 
        c.drawString(x, y - 18 * mm, 'Serial#: ' + srec.serial_number)   
        c.drawString(x, y - 24 * mm, 'Part: ' + srec.part_number)
        c.drawString(x, y - 30 * mm, 'Desc: ' + srec.description[:27])       
        c.drawString(x, y - 36 * mm, 'Condition: ' + srec.condition_code) 
        c.drawString(x, y - 42 * mm, 'Qty: ' + str(srec.qty_oh)) 
        c.drawString(65 * mm, y - 42 * mm, 'RO:' + srec.ro_number) 
        c.drawString(x, y - 48 * mm, 'EX ESN: ' + srec.slug)  
        c.drawString(65 * mm, y - 48 * mm, 'PO#: ' + srec.po_number) 
        c.drawString(x, y - 54 * mm, 'SL #:' + srec.stock_line) 
        c.drawString(65 * mm, y - 54 * mm, 'WO#: ' + srec.wo_number)
        c.drawString(x, y - 60 * mm, 'Rec. Date: ' + srec.arrival_date.strftime('%m-%d-%Y')) 
        c.drawString(65 * mm, y - 60 * mm, 'Cons: ' + srec.consignment_code)        
        c.rect(x-0.5*mm,y - 85 * mm, 99*mm, 23*mm, fill=0, stroke=1)
        
        if srec.cond_level_zero == 'T':
            c.drawString(x + 1*mm, y - 67 * mm, 'Notes:')
        else:
            c.drawString(x + 1*mm, y - 67 * mm, 'Reason:') 
        c.drawString(x + 2*mm, y - 76 * mm, srec.notes)  
        c.rect(63*mm, y - 113* mm, 40.5*mm, 27*mm, fill=0, stroke=1)
        c.drawImage(logo_path,x-1*mm, y-112*mm,width=153,height=65,mask='auto')        
        c.drawString(69.5*mm, y - 117 * mm, code_vals['inspector'])
        #c.drawString(x, y - 92 * mm, code_vals['inspector'])
        c.drawString(x+1.5*mm, y - 117 * mm, 'BA-LG-007 Rev 11/22')
        y -= 120 * mm
    error = c.save() 
    return error
    
def create_barcodes_srsg(barcode_value,ctrl_id,ctrl_number,part_no,desc,serial,qty,wo_number,task_list,repair):
    from reportlab.graphics.barcode import code39
    from reportlab.lib.pagesizes import A6,landscape
    from reportlab.lib.units import mm,inch
    from reportlab.pdfgen import canvas  
    from reportlab.lib.colors import Color,black    
    filepath = "/home/ubuntu/mo_template/uploads/barcode.pdf"   
    c = canvas.Canvas(filepath, pagesize=A6)
    c.setPageSize(landscape((105*mm,153*mm)))
    c.rotate(90)
    barcode39Std = code39.Standard39(barcode_value, humanReadable=True, barHeight=75, barWidth=1.20, stop=1, checksum=0)  
    code = barcode39Std 
    y = -32 * mm
    x = 5 * mm
    code.drawOn(c, -2*mm, y)
    #image = '/home/ubuntu/mo_template/static/bischoff_logo.jpg'
    #c.drawImage(image, 43*mm, y, width=162,height=70,mask='auto')
    y = y - 6 * mm    
    descrip = desc[:22]
    # move the origin up and to the left
    #c.translate(inch,inch)
    # define a large font
    c.setFont("Courier", 14)
    c.drawString(x + 28*mm, y, barcode_value)   
    c.drawString(x, y - 6 * mm, 'CTRL#: ' + str(ctrl_number))
    c.drawString(42 * mm, y - 6 * mm,'CTRL ID: ' + str(ctrl_id))      
    c.drawString(x, y - 12 * mm, 'Part: ' + str(part_no))  
    c.drawString(x, y -18 * mm, 'Desc: ' + descrip)
    #if desc[22:]:   
        #c.drawString(x, y - 24 * mm, desc[22:])     
    c.drawString(x, y - 24 * mm, 'Serial#: ' + str(serial))
    c.drawString(x, y - 30 * mm, 'Qty: ' + str(qty)) 
    c.drawString(28 * mm, y - 30 * mm,' WO#: ' + str(wo_number))
    if repair:
        c.drawString(x, y - 42 * mm, 'Send for Outside Service') 
        c.rect(x-0.5*mm,y-61*mm,97*mm,24*mm,fill=0,stroke=1)   
    else:
        z = 42 * mm
        num = 1
        for task in task_list:
            att_val = str(task.att_value)
            att_val = att_val[:19] + (att_val[19:] and '..')
            c.drawString(x + 1*mm, y - z +2*mm, str(task.att_seq) + '.') 
            c.drawString(x + 5*mm, y - z +2*mm, str(task.att_name))
            c.drawString(x + 6*mm, y - z -4*mm, att_val)             
            c.rect(x + 86*mm, y - z -4* mm, 6*mm, 6*mm, fill=0, stroke=1)  
            if (num % 2) == 0:                    
                gray_transparent = Color(0.85,0.85,0.85, alpha=0.5)      
                c.setFillColor(gray_transparent)                 
                c.rect(x+0.5*mm, y - 8*mm - z, 97*mm, 15*mm, fill=1, stroke=1)
                c.setFillColor(black) 
            else:              
                c.rect(x+0.5*mm, y - 8*mm - z,97*mm,15*mm,fill=0,stroke=1)
            num += 1
            z += 16*mm
    if not task_list and not repair:
        c.drawString(x, y - 42 * mm, 'Quarantine: UNVERIFIED PART')
        c.rect(x-0.5*mm,y-61*mm,97*mm,24*mm,fill=0,stroke=1)
    error = c.save() 
    return error 
    
def create_barcodes_bai(barcode_value,ctrl_id,ctrl_number,part_no,desc,serial,qty,wo_number,task_list,repair):
    from reportlab.graphics.barcode import code39
    from reportlab.lib.pagesizes import A6,landscape
    from reportlab.lib.units import mm,inch
    from reportlab.pdfgen import canvas  
    from reportlab.lib.colors import Color,black    
    filepath = "/home/ubuntu/mo_template/uploads/barcode.pdf"   
    c = canvas.Canvas(filepath, pagesize=A6)
    c.setPageSize(landscape((105*mm,153*mm)))
    c.rotate(90)
    barcode39Std = code39.Standard39(barcode_value, humanReadable=True, barHeight=75, barWidth=1.20, stop=1, checksum=0)  
    code = barcode39Std 
    y = -32 * mm
    x = 5 * mm
    code.drawOn(c, -2*mm, y)
    #image = '/home/ubuntu/mo_template/static/bischoff_logo.jpg'
    #c.drawImage(image, 43*mm, y, width=162,height=70,mask='auto')
    y = y - 6 * mm    
    descrip = desc[:22]
    # move the origin up and to the left
    #c.translate(inch,inch)
    # define a large font
    c.setFont("Courier", 14)
    c.drawString(x + 28*mm, y, barcode_value)   
    c.drawString(x, y - 6 * mm, 'CTRL#: ' + str(ctrl_number))
    c.drawString(42 * mm, y - 6 * mm,'CTRL ID: ' + str(ctrl_id))      
    c.drawString(x, y - 12 * mm, 'Part: ' + str(part_no))  
    c.drawString(x, y -18 * mm, 'Desc: ' + descrip)
    #if desc[22:]:   
        #c.drawString(x, y - 24 * mm, desc[22:])     
    c.drawString(x, y - 24 * mm, 'Serial#: ' + str(serial))
    c.drawString(x, y - 30 * mm, 'Qty: ' + str(qty)) 
    c.drawString(28 * mm, y - 30 * mm,' WO#: ' + str(wo_number))
    if repair:
        c.drawString(x, y - 42 * mm, 'Send for Outside Service') 
        c.rect(x-0.5*mm,y-61*mm,97*mm,24*mm,fill=0,stroke=1)   
    else:
        z = 42 * mm
        num = 1
        for task in task_list:
            att_val = str(task.att_value)
            att_val = att_val[:19] + (att_val[19:] and '..')
            c.drawString(x + 1*mm, y - z +2*mm, str(task.att_seq) + '.') 
            c.drawString(x + 5*mm, y - z +2*mm, str(task.att_name))
            c.drawString(x + 6*mm, y - z -4*mm, att_val)             
            c.rect(x + 86*mm, y - z -4* mm, 6*mm, 6*mm, fill=0, stroke=1)  
            if (num % 2) == 0:                    
                gray_transparent = Color(0.85,0.85,0.85, alpha=0.5)      
                c.setFillColor(gray_transparent)                 
                c.rect(x+0.5*mm, y - 8*mm - z, 97*mm, 15*mm, fill=1, stroke=1)
                c.setFillColor(black) 
            else:              
                c.rect(x+0.5*mm, y - 8*mm - z,97*mm,15*mm,fill=0,stroke=1)
            num += 1
            z += 16*mm
    if not task_list and not repair:
        c.drawString(x, y - 42 * mm, 'Quarantine: UNVERIFIED PART')
        c.rect(x-0.5*mm,y-61*mm,97*mm,24*mm,fill=0,stroke=1)
    error = c.save() 
    return error   
  
def create_traveller(barcode_value,ctrl_id,ctrl_number,part_no,desc,serial,qty,wo_number,task_list,repair):
    from reportlab.graphics.barcode import code39
    from reportlab.lib.pagesizes import A6,landscape
    from reportlab.lib.units import mm,inch
    from reportlab.pdfgen import canvas  
    from reportlab.lib.colors import Color,black    
    filepath = "/home/ubuntu/mo_template/uploads/barcode.pdf"   
                                                          
    c = canvas.Canvas(filepath, pagesize=A6)
    c.setPageSize(landscape((105*mm,153*mm)))
    c.rotate(90)
    barcode39Std = code39.Standard39(barcode_value, barHeight=75, barWidth=1.20, stop=1)  
    code = barcode39Std 
    y = -32 * mm
    x = 5 * mm
    code.drawOn(c, -2*mm, y)
    #image = '/home/ubuntu/mo_template/static/bischoff_logo.jpg'
    #c.drawImage(image, 43*mm, y, width=162,height=70,mask='auto')
    y = y - 6 * mm    
    descrip = desc[:22]
    # move the origin up and to the left
    #c.translate(inch,inch)
    # define a large font
    c.setFont("Courier", 14)
    c.drawString(x + 28*mm, y, barcode_value)   
    c.drawString(x, y - 6 * mm, 'CTRL#: ' + str(ctrl_number))
    c.drawString(42 * mm, y - 6 * mm,'CTRL ID: ' + str(ctrl_id))      
    c.drawString(x, y - 12 * mm, 'Part: ' + str(part_no))  
    c.drawString(x, y -18 * mm, 'Desc: ' + descrip)
    if desc[22:]:   
        c.drawString(x, y - 24 * mm, desc[22:])     
    c.drawString(x, y - 24 * mm, 'Serial#: ' + str(serial))
    c.drawString(x, y - 30 * mm, 'Qty: ' + str(qty)) 
    c.drawString(28 * mm, y - 30 * mm,' WO#: ' + str(wo_number))
    if repair:
        c.drawString(x, y - 42 * mm, 'Send for Outside Service') 
        c.rect(x-0.5*mm,y-61*mm,97*mm,24*mm,fill=0,stroke=1)   
    else:
        z = 42 * mm
        num = 1
        for task in task_list:
            att_val = str(task.att_value)
            att_val = att_val[:19] + (att_val[19:] and '..')
            c.drawString(x + 1*mm, y - z +2*mm, str(task.att_seq) + '.') 
            c.drawString(x + 5*mm, y - z +2*mm, str(task.att_name))
            c.drawString(x + 6*mm, y - z -4*mm, att_val)             
            c.rect(x + 86*mm, y - z -4* mm, 6*mm, 6*mm, fill=0, stroke=1)  
            if (num % 2) == 0:                    
                gray_transparent = Color(0.85,0.85,0.85, alpha=0.5)      
                c.setFillColor(gray_transparent)                 
                c.rect(x+0.5*mm, y - 8*mm - z, 97*mm, 15*mm, fill=1, stroke=1)
                c.setFillColor(black) 
            else:              
                c.rect(x+0.5*mm, y - 8*mm - z,97*mm,15*mm,fill=0,stroke=1)
            num += 1
            z += 16*mm
    if not task_list and not repair:
        c.drawString(x, y - 42 * mm, 'Quarantine: UNVERIFIED PART')
        c.rect(x-0.5*mm,y-61*mm,97*mm,24*mm,fill=0,stroke=1)
    error = c.save() 
    return error   

def check_license_expiry(user):
    error = ''
    #check group and get quantum_cmp_key
    user_groups = user.groups.values_list('id',flat=True)
    group = user_groups and user_groups[0] or None
    groups = UserGroupProfile.objects.filter(quantum_cmp_key__isnull=False)
    lic_group = groups and groups[0] or None
    if not lic_group:
        error = 'User must belong to a company group.'
        return error	    
    quantum_cmp_key = lic_group.quantum_cmp_key
    conn_string = lic_group.conn_string
    if quantum_cmp_key and conn_string:
        from portal.tasks import check_exp_date
        res = check_exp_date.delay(conn_string,quantum_cmp_key)
        error = res.get()
    else:
        error = 'Must have active license. Contact MRO Live Account Manager.'
    return error
    
def init_session(session_key):
    """
    Initialize same session as done for ``SessionMiddleware``.
    """
    from django.conf import settings
    import importlib
    engine = importlib.import_module(settings.SESSION_ENGINE)
    return engine.SessionStore(session_key)
    
def check_sessions(request,user_id,logoff='0'):
    error = ''
    session_count = 0 
    from datetime import timezone
    from django.contrib.sessions.models import Session
    from django.contrib.auth import logout
    right_now = datetime.now(timezone.utc)
    all_sessions = Session.objects.filter(expire_date__gte=right_now)    
    for sesh in all_sessions:
        if logoff == '1':
            request.session = init_session(sesh.session_key)
            logout(request)
        if logoff != '1' and sesh.get_decoded().get('_auth_user_id') == str(user_id):
            session_count += 1
            if session_count > 1:
                error = 'You are already logged in on another session...'
    return error
    
def create_lot_teardown(quapi_id,sysur_auto_key,user_id,row):
    error = ''
    req_data = []
    prod_data = {}
    from portal.tasks import lot_teardown
    res = lot_teardown.delay(quapi_id,sysur_auto_key,user_id,row)
    error,msg,show_msg = res.get()     
    return error,msg,show_msg
 
def create_teardown_rows(quapi_id,sysur_auto_key,user_id,row,new_wo,\
    is_default_repair,stm_serials=[],serial_notes=[]):
    error = ''
    req_data = []
    prod_data = {}
    from portal.tasks import teardown_rows
    #for row in recs:
        #1. call procedure for creating teardown rows (WOB)
           #a. first check if serialized and if so, throw an exception
    res = teardown_rows.delay(quapi_id,sysur_auto_key,user_id,row,\
        new_wo,is_default_repair,stm_serials=stm_serials,serial_notes=[])
    error,msg,show_msg,new_woo_key,default_repair,pn_info = res.get()     
    return error,msg,show_msg,new_woo_key,default_repair,pn_info
    
    
def printnode_pdf(printset,auth_key):            
    error = ''
    import requests
    if printset and auth_key and 0:
        printer_name = printset.printer_name or None
        #printer_name = 'Shop Label Printer'
        printers = requests.get('https://api.printnode.com/printers', auth=(auth_key,''))
        printers = printers and printers.json()
        printers = printers and [pri for pri in printers if pri['name'] == printer_name] or []
        printer_id = printers and printers[0]['id'] or '71870030'
        #for printer in printers:
            #if printer[1] == printer_name:
                #printer_id = printer[0]              
        #this is good code.  We just need to create a pdf with the barcode and then run this with the pdf file name and print it.
        if printer_id:    
            #options = '"options": {"copies":1,"pages":"1","duplex":"long-edge","paper":"A4","bin":"Tray 1"}',
            headers = {'Content-Type': 'application/json',}
            data = '{ "printerId":"' + str(printer_id) + '", "title": "MRO Live Teardown barcode label", "contentType": "pdf_uri", "content": "http://18.188.56.36:8029/uploads/barcode.pdf", "source": "MRO Live Portal - Teardown App" }'
            error = requests.post('https://api.printnode.com/printjobs', headers=headers, data=data, auth=(auth_key, ''))
            error = error.json()    
            error = error and type(error) is dict and 'message' in error and error['message'] or ''
    return error
   
def teardown(request,quapi_id='',si_number='',part_number='',condition_code='',notes=''):
    new_status,location,filter_status,show_msg,task_list = '','','','',[]
    user_rec,fail_ms,new_woo_key = None,'',''
    val_dict,form = {},{}  
    total_rows = 0    
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request and request.user or None   
    
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')     
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict)
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='teardown')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    is_default_repair = app_allow and app_allow[0] and app_allow[0].default_repair
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    val_dict['user_apps'] = user_apps
    val_dict['quapi_id'] = quapi_id  
    val_dict['user_name'] = user_name 
    session_id = 'Au23*72334agAHAUas&394'
    from portal.tasks import get_conditions
    res = get_conditions.delay(
        quapi_id,session_id,
    )
    error = res.get()            
    from polls.models import PartConditions
    conditions = PartConditions.objects.filter(
        session_id = session_id
    )
    
    val_dict.update({
        'conditions':conditions,
        'condition_code': 'AR',
    })
    
    if request.method == 'GET':
        form = WODashboardForm()
        import random
        import string
        
        val_dict.update({
            'wo_task': si_number,
            'part_number': part_number,
            'condition_code': condition_code or 'AR',
            'notes': notes,
            'form': form,
        })
        return render(request, 'mrolive/teardown.html', val_dict)
        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)   
        val_dict['form'] = form
        new_wo = req_post.get('new_wo','') or req_post.get('mod_new_wo','')
        val_dict['new_wo'] = new_wo and 'checked' or ''
        val_dict['mod_new_wo'] = new_wo and 'checked' or ''
        wo_task = req_post.get('wo_task','')
        quantity = req_post.get('quantity','')        
        serial_number = req_post.get('serial_number','')
        part_number = req_post.get('part_number','')
        part_number = part_number and part_number.lstrip() or part_number
        part_number = part_number and part_number.rstrip() or part_number
        format_type = req_post.get('type','')
        auto_click = req_post.get('auto_click','')
        notes = req_post.get('notes','')
        session_id = req_post.get('session_id','')
        show_msg = req_post.get('show_modal','')
        modal_pn = req_post.get('modal_pn','')
        modal_pn = modal_pn and modal_pn.lstrip() or modal_pn
        modal_pn = modal_pn and modal_pn.rstrip() or modal_pn
        modal_desc = req_post.get('modal_desc','')
        modal_serialized = req_post.get('modal_serialized','')
        modal_cancel = req_post.get('mod_cancel','')
        serial_cancel = req_post.get('serial_cancel','')
        error_ok = req_post.get('error_ok','')
        app_code = req_post.get('app_code','')
        opm_list = req_post.getlist('opm_list[]')
        pnm_auto_key = req_post.get('pn_sel','')
        condition_code = req_post.get('cond_selector','')
        val_dict.update({
            'condition_code': condition_code,
            'teardown': 'T',
                                   
            })
            
        pn_info = []

        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','')                
        
        if modal_pn and not modal_serialized:
            serial_number = ''
           
        """if modal_cancel or serial_cancel or error_ok:
            val_dict.update({
                'show_modal': '',
                'session_id': session_id,
                'part_number': part_number,
                'quantity': quantity,
                'wo_task': wo_task,
                'serial_number': serial_number,
                'notes': notes,
                'user_name': user_name,  
                'auto_click': 'done', 
                })  
            return render(request, 'mrolive/teardown.html', val_dict)"""
            
        row = []
        row.append(wo_task)
        row.append(quantity)
        row.append(part_number)
        row.append(serial_number)
        row.append(session_id)
        row.append(notes)
        row.append(modal_desc)
        row.append(pnm_auto_key)
        row.append(condition_code)        
        stm_serials = req_post.getlist('modal_serials[]') 
        serial_notes = req_post.getlist('serial_notes[]') 
       
        if wo_task and stm_serials:
            error,msg,show_msg,new_woo_key,default_repair,pn_info=create_teardown_rows(\
                quapi_id,sysur_auto_key,user_name,row,new_wo,is_default_repair,\
                stm_serials=stm_serials,serial_notes = serial_notes)
            show_msg = 'done'  
            val_dict.update({
                'error': error,
                'msg': msg,
                'show_modal': 'done',
                'session_id': session_id,
                'part_number': part_number,
                'quantity': '',
                'condition_code': condition_code,
                'wo_task': wo_task,
                'serial_number': '',
                'notes': notes,
                'user_name': user_name,  
                'auto_click': 'done', 
                })  
        
        elif wo_task and show_msg in ['','done']:   
       
            error,msg,show_msg,new_woo_key,default_repair,pn_info = create_teardown_rows(\
            quapi_id,sysur_auto_key,user_name,row,new_wo,is_default_repair)
            if show_msg == 'show_serials':
                #launch the prompt for serial numbers
                stms = WOStatus.objects.filter(session_id=session_id)
                val_dict.update({
                    'serial_stms':stms,
                    'show_modal': 'show_serials',
                    'session_id': session_id,
                    'part_number': part_number,
                    'quantity': quantity,
                    'wo_task': wo_task,
                    'serial_number': serial_number,
                    'notes': notes,
                    'user_name': user_name,  
                    'auto_click': 'done', 
                    })  
                return render(request, 'mrolive/teardown.html', val_dict)                    
                
            elif show_msg not in ['show_modal','show_pns']:
                show_msg = 'done' 
                auto_click = ''
                
        elif wo_task and show_msg == 'got_data':
            #create the new PN with this data using a new task (method) called create_pn
            from portal.tasks import create_pn          
            res = create_pn.delay(quapi_id,session_id,modal_pn,modal_desc,modal_serialized,sysur_auto_key)
            error,msg = res.get()
            row[2] = modal_pn
            val_dict['error'] = error
            
            if not error:
                if show_msg != 'show_serials':
                    error,msg,show_msg,new_woo_key,default_repair,pn_info = create_teardown_rows(\
                    quapi_id,sysur_auto_key,user_name,row,new_wo,is_default_repair)
                    
                    if show_msg != 'show_serials':
                        show_msg = 'done'
                        auto_click = ''                    
                    
                    else:
                        stms = WOStatus.objects.filter(session_id=session_id)
                        val_dict.update({
                            'serial_stms':stms,
                            'show_modal': 'show_serials',
                            'session_id': session_id,
                            'part_number': part_number,
                            'quantity': quantity,
                            'wo_task': wo_task,
                            'serial_number': serial_number,
                            'notes': notes,
                            'user_name': user_name,  
                            'auto_click': 'done', 
                            })
                        return render(request, 'mrolive/teardown.html', val_dict) 
                
        elif not wo_task:
            error = "WO# is required."
            val_dict.update({
                'error': error,           
                'session_id': session_id,
                'part_number': part_number,
                'quantity': quantity,
                'wo_task': wo_task,
                'serial_number': serial_number,
                'notes': notes,
                'user_name': user_name,  
                'auto_click': 'done',                               
            }) 
            return render(request, 'mrolive/teardown.html', val_dict)   
            
        records = WOStatus.objects.filter(session_id=session_id)
        record = records and records[0] or None
        
        if record and show_msg != 'show_serials':
            description = record.description   
            ctrl_id = record.ctrl_id or 0
            ctrl_number = record.ctrl_number or 0
            stm = record.stm_auto_key or ''
            if not stm or 1:
                element = '0' + str(ctrl_number) 
                element += '00000' + str(ctrl_id)
            else:
                element = 'C' + str(stm)
            pn = record.part_number or ''
            pnm_auto_key = record.pnm_auto_key or ''
            activity = record.activity or ''
            desc_last = ''
            if len(description) > 10:
                if len(description) > 20:
                    desc_last = description[10:20]
                else:
                    desc_last = description[10:]
            description = description[:10]
            serial_no = record.serial_number or serial_number
            qty = record.quantity or 0
            wo_number = record.wo_number or ''
            stock_line = record.stock_line or ''
            exp_date = record.exp_date or ''
            #condition_code = record.condition_code or ''
            consignment_code = record.consignment_code or ''
            loc_code = record.location_code or ''
            mfg_lot_num = record.spn_code or ''
            eng_model = record.int_rank or ''
            esn = record.slug
            repair = activity and activity == 'Repair' or ''
            notes = record.notes or '' 
            customer = record.customer or ''
            ata_code = record.cart or ''
            work_request = record.wh_code or ''            
            
            #get the attributes list to display on the Teardown label
            if pnm_auto_key and not new_wo:
                from portal.tasks import get_part_attributes
                res = get_part_attributes.delay(quapi_id,sysur_auto_key,session_id,pnm_auto_key,group='IN HOUSE',create_anew=False)  
                error,att_msg = res.get()
                msg += att_msg
            if not error:
                from polls.models import UserDefAtts as uda
                task_list = uda.objects.filter(session_id=session_id).order_by('att_seq') 
            printed_date = datetime.now()
            printed_date = printed_date.strftime('%m/%d/%Y %H:%M:%S')            
            val_dict.update({
                'element': element,
                'records': records,
                'stock_recs': records,
                'ctrl_id': ctrl_id,
                'ctrl_number': ctrl_number,
                'stm_auto_key': stm,
                'wo_task': wo_task,
                'pn': pn,
                'part_number': pn,
                'description': description,
                'desc_last': desc_last,
                'serial_no': serial_no,
                'serial_number': serial_no,
                'qty': qty,
                'wo_number': wo_number,  
                'si_number': wo_number,
                'type': format_type, 
                'total_rows': '0',  
                'auto_click': auto_click, 
                'task_list': task_list, 
                'repair': repair,  
                'new_wo': new_wo, 
                'stock_line': stock_line,
                'exp_date': exp_date, 
                'condition_code': condition_code or '    ',
                'consignment_code': consignment_code,
                'loc_code': loc_code, 
                'mfg_lot_num': mfg_lot_num,
                'customer': customer,
                'quantity': quantity,
                'notes': notes or '   ', 
                'eng_model': eng_model,
                'ata_code': ata_code,
                'work_request': work_request,
                'printed_by': user_name,
                'printed_date': printed_date,
                'esn': esn,
                'error': error,
                })
        
        if 'opm_list[]' in req_post and new_woo_key and not is_default_repair:
            #call the task method and pass the opm selected by the user
            if opm_list:
                opm_auto_key = opm_list[0]
                from portal.tasks import create_opm_tasks
                res = create_opm_tasks.delay(session_id,sysur_auto_key,quapi_id,opm_auto_key,new_woo_key) 
                error,opm_msg = res.get() 
                msg += opm_msg 
            else:
                val_dict['error'] = 'No operation selected.'
                return render(request, 'mrolive/teardown.html', val_dict)
             
        if records and show_msg != 'show_serials':     
            msg += ' Successful teardown'                        
            printset = app_allow and app_allow[0] and app_allow[0].printset_id
            auth_key = printset and printset.printnode_auth_key
                                                       
            if not error:                                                                        
                val_dict['msg'] = 'Successful teardown. Barcode label printed.' 
            else:
                val_dict['error'] = error               
            val_dict['app_code'] = 'teardown'
            if new_wo and new_woo_key:
                new_subwoo = WOStatus.objects.filter(session_id=session_id,woo_auto_key=new_woo_key)
                record = new_subwoo and new_subwoo[0] or None
                description = record.description and len(record.description) > 19 and record.description[:19] or record.description
                val_dict.update({
                    'ctrl_id': record.ctrl_id or 0,
                    'ctrl_number': record.ctrl_number or 0,
                    'part_number': record.part_number or '',
                    'si_number': record.wo_number or '',
                    'description': description or '',
                    'serial_no': record.serial_number or '',
                    'qty': record.quantity or 0,
                    'ref': record.task_ref,
                    'stm_auto_key': record.stm_auto_key,
                    'model': record.pnm_modify,
                    'esn': record.slug,
                    #'customer': record.customer,
                    'printed_by': user.first_name + ' ' + user.last_name,
                    'new_wo': new_wo,
                })                                         
                from polls.models import WOTask as wot
                task_list = wot.objects.filter(session_id=session_id).order_by('wot_sequence')
                val_dict['task_list'] = task_list
                val_dict['user'] = user
                right_now = datetime.now()
                val_dict['timestamp'] = right_now.strftime('%m-%d-%Y %H:%M:%S')
                if record and printset and auth_key and default_repair:

                    ctrl_id = record.ctrl_id or 0
                    ctrl_number = record.ctrl_number or 0
                    pn = record.part_number or ''
                    description = record.description or ''
                    serial_no = record.serial_number or ''
                    qty = record.quantity or 0
                    wo_number = record.wo_number or ''
                    
                    error = create_traveller(element,ctrl_id,\
                        ctrl_number,pn,description,serial_no,\
                        qty,wo_number,task_list,repair)

                    if error:
                        val_dict['error'] = error
                        
                    else:
                        #error = printnode_pdf(printset,auth_key)
                        val_dict['msg'] += ' Created new WO. Traveller printed.'
                        
            if 0 and new_wo and not is_default_repair and not opm_list:            
                updated_opms = Operation.objects.filter(session_id=session_id)            
                val_dict.update({        
                    'total_rows': str(len(updated_opms))
                })  
                
            elif printset and auth_key:            
                #return render(request, 'mrolive/teardown.html', val_dict)
                return render(request, 'mrolive/teardown.html', val_dict)
                
            else:
                # and default_repair
                if new_wo and new_woo_key and not error:                 
                    return render(request, 'mrolive/teardown_traveller_mro_header.html',val_dict)
                    
                elif not error:                                 
                    return render(request, 'mrolive/plain_barcode_mro.html',val_dict)
                    
                else:                   
                    val_dict['new_wo'] = 'checked'
                    return render(request, 'mrolive/teardown.html',val_dict)
                    
        if show_msg in ['show_modal']:
            modal_pn = part_number
            
        elif show_msg in ['show_pns'] and pn_info:
            #lookup pn desc and 
            from polls.models import PartNumbers
            part_data,error = [],''
            #[pnm_auto_key,part_num,part_desc,mfg_code]
            for part in pn_info:    
                part_data.append(PartNumbers(
                pnm_auto_key = part[0],
                part_number = part[1],
                description = part[2],
                mfg_code = part[3],
                session_id = session_id,      
                ))
            if part_data:
                try:
                    delete = PartNumbers.objects.filter(session_id=session_id).delete()
                    PartNumbers.objects.bulk_create(part_data) or []    
                except Exception as exc:
                    error += "Error, %s, creating parts."%(exc)
            pnms = PartNumbers.objects.filter(session_id=session_id)
            val_dict['pn_info'] = pnms
        
        if show_msg == 'show_pns':
            if not pn_info:
                show_msg = ''
                
        val_dict.update({
            'teardown': teardown,        
            'session_id': session_id, 
            'quantity': quantity,
            'part_number': part_number or '    ',
            'active_part': part_number or '    ',
            'wo_task': wo_task or '    ',
            'serial_number': serial_number,
            'notes': notes or '    ',
            'condition_code': condition_code or '    ',            
            'show_modal': show_msg,
            'modal_pn': modal_pn,
            'user_name': user_name,  
        })            
        
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/teardown.html', val_dict)
    
def create_barcodes(vals,inspection=False):
    from reportlab.graphics.barcode import code39,code93
    from reportlab.lib.pagesizes import A4,letter,landscape
    from reportlab.lib.units import mm,inch
    from reportlab.pdfgen import canvas                              
    #set the filepath and logo_path for the pdf label file.
    file_path = "/home/ubuntu/mo_template/static/teardown-inspection-report.pdf" 
    logo_path = "/home/ubuntu/mo_template/static/logo.jpg"
    stock_recs = vals['stock_recs']    
    pagesize = (3*inch, 3*len(stock_recs)*inch)
    c = canvas.Canvas(file_path, pagesize=pagesize)
    x = 1*mm                         
    y = 87*len(stock_recs)*mm
    c.setPageSize((105*mm,105*len(stock_recs)*mm))    
    
    for rec in stock_recs: 
        element = '0' + str(rec.ctrl_number)
        element += '00000' + str(rec.ctrl_id)    
        barcode39Std = code39.Extended39(element,\
            barHeight=45, barWidth=1.31, stop=1, checksum=0) 
        code = barcode39Std
        code.drawOn(c, -4*mm, y)                                  
        c.setFont("Helvetica",6)
        c.drawString(10 * mm, y - 3 * mm, 'Ctrl#: ' + str(rec.ctrl_number))
        c.drawString(34 * mm, y - 3 * mm, 'CtrlID: ' + str(rec.ctrl_id))
        c.setFont("Helvetica",18)
        c.drawString(x, y - 9 * mm, 'WO#: ' + str(rec.wo_number))
        c.drawString(x, y - 15 * mm,'COND: ' + str(rec.condition_code))
        c.setFont("Helvetica",10)
        c.drawString(x, y - 21 * mm, 'PN: ' + rec.part_number)    
        c.drawString(x, y - 26 * mm, 'SN: ' + rec.serial_number)   
        c.drawString(x, y - 31 * mm, 'DESC: ' + rec.description[:47]) 
        c.drawString(x, y - 36 * mm, 'NOTES:')
        c.drawString(0.85 * mm, y - 45 * mm, rec.notes[:24])
        if len(rec.notes) > 24:
            c.drawString(0.85 * mm, y - 50 * mm, rec.notes[25:50]) 
        if len(rec.notes) > 50:
            c.drawString(0.85 * mm, y - 55 * mm, rec.notes[51:66])  
        if len(rec.notes) > 66:
            c.drawString(0.85 * mm, y - 60 * mm, rec.notes[67:92])              
        c.rect(0.5 * mm, y - 61* mm, 44.5*mm, 21*mm, fill=0, stroke=1) 
        c.drawImage(logo_path,63 * mm,y-18*mm,width=117,height=49,mask='auto')        
        c.drawString(49 * mm, y - 26 * mm, 'ESN: ' + rec.slug) 
        c.drawString(49 * mm, y - 36 * mm, 'QTY: ' + str(rec.quantity or rec.qty_oh))
        c.drawString(49 * mm, y - 41 * mm, 'SL:' + rec.stock_line) 
        c.drawString(49 * mm, y - 46 * mm, 'PRINTED DATE: ')
        c.drawString(49 * mm, y - 51 * mm, vals['printed_date'])
        print_by = 'PRINTED BY:'    
        if inspection:
            print_by = 'INSPECTED BY:'
        c.drawString(49 * mm, y - 56 * mm, print_by + vals['printed_by'])         
        if inspection:
            c.drawString(49 * mm, y - 61 * mm, 'STAMP:')
            c.rect(49 * mm, y - 78 * mm, 44.5 * mm, 15*mm, fill=0, stroke=1) 
        y -= 81 * mm
        
    error = c.save() 
    return error   
    
def gen_fake_qty():
    # generate random integer values
    from random import seed
    from random import randint
    # generate some integers
    fake_qty = randint(0, 1099)
    return str(fake_qty)

def requests_view(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec,fail_ms = 'user not set',None,''
    val_dict,form = {},{}  
    total_rows = 0    
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='parts-request')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    val_dict['user_apps'] = user_apps
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':
        form = WODashboardForm() 
        from portal.tasks import get_statuses_nsync_beta, get_users_nsync_beta, get_loc_whs_cart_nsync_beta 
        res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=0)
        user_error = res.get()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post, request.FILES)   
        val_dict['form'] = form        
        active_user = 'active_user' in req_post and req_post['active_user'] or ''  
        filter_user = 'filter_user' in req_post and req_post['filter_user'] or ''        
        user_id = 'user_code' in req_post and req_post['user_code'] or ''
        user_name = 'user_name' in req_post and req_post['user_name'] or ''
        user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id) or None
        user_rec = user_rec and user_rec[0] or None
        sysur_auto_key = user_rec and user_rec.user_auto_key or None
        user_name = user_rec and user_rec.user_name or user_name or ''
        if user_id and user_rec and not user_name:
            val_dict['error'] = 'Invalid user id.'
            return render(request, 'mrolive/requests-view.html', val_dict) 
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        wo_number = req_post.get('wo_number','')
        wo_task = req_post.get('wo_task','')
        if 'wo_number' in req_post and not wo_number:
            error = 'Must enter WO#.'
            val_dict.update({
                'error': error,
                'user_id': user_id,
                'user_name': user_name,            
                'session_id': session_id, 
                'wo_task': wo_task,
                'wo_number': wo_number,               
            }) 
            return render(request, 'mrolive/requests-view.html', val_dict)  
        elif 'wo_task' in req_post and wo_number:
            #do the lookup and create grid results records in the wostatus table
            from portal.tasks import get_requests_view
            res = get_requests_view.delay(quapi_id,session_id,user_id,wo_task,wo_number)   
            error,msg = res.get() 
            requests = WOStatus.objects.filter(session_id=session_id)
            total_rows = len(requests)            
        val_dict.update({
            'user_id': user_id,
            'user_name': user_name,            
            'session_id': session_id, 
            'label': '',
            'wo_task': '',  
            'total_rows': total_rows,            
        })           
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/requests-view.html', val_dict) 
    
def parts_request(request,quapi_id=None):
    new_status,location,filter_status,show_conf = '','','','F'
    user_id,user_rec,fail_ms,session_id = 'user not set',None,'',''
    val_dict,form = {},{}   
    error,msg,loc_msg,stat_msg,user_msg = '','','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='parts-request')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    modes = app_id and get_modes(app_id) or []
    if not modes:
        return redirect('/login/') 
    val_dict['modes'] = modes 
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    val_dict['user_apps'] = user_apps
    val_dict['quapi_id'] = quapi_id                                
    if request.method == 'GET':
        form = WODashboardForm() 
        val_dict['user_name'] = user_name
        val_dict['active_mode'] = '1'      
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post, request.FILES)   
        val_dict['form'] = form  
        active_mode = req_post.get('mode_selector','')
        if not active_mode:
            active_mode = req_post.get('active_mode','')
        sel_mode = 'sel_mode' in req_post and req_post['sel_mode'] or ''
        cur_mode = 'cur_mode' in req_post and req_post['cur_mode'] or ''        
        if not (sel_mode or active_mode or cur_mode):
            val_dict['error'] = 'Must select a mode.'
            render(request, 'mrolive/parts-request.html', val_dict)                
        active_user = 'active_user' in req_post and req_post['active_user'] or ''  
        filter_user = 'filter_user' in req_post and req_post['filter_user'] or ''        
        #user_id = 'user_id' in req_post and req_post['user_id'] or ''
        user_rec = QuantumUser.objects.filter(user_auto_key=sysur_auto_key) or None
        user_rec = user_rec and user_rec[0] or None
        #sysur_auto_key = user_rec and user_rec.user_auto_key or None
        user_name = user_rec and user_rec.user_name or user_name or ''
        user_id = user_rec and user_rec.user_id or ''
        if user_id and not user_name:
            val_dict['error'] = 'Invalid user id.'
            return render(request, 'mrolive/parts-request.html', val_dict) 
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id or session_id == '1234567':
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''       
        part_number = req_post.get('part_number','')
        part_number = part_number and part_number.replace(" ", "") or part_number
        quantity = req_post.get('quantity','')
        notes = req_post.get('notes','')
        activity = req_post.get('activity','')
        wo_task = req_post.get('wo_task','')
        wo_number = req_post.get('wo_number','')
        yes_request = req_post.get('yes_request','')
        no_request = req_post.get('no_request','')
        bom_status = req_post.get('bom_status','')
        wo_task_sel = req_post.get('wo_task_sel','')
        modal_pn = req_post.get('modal_pn','')
        modal_pn = modal_pn and modal_pn.replace(" ", "") or modal_pn
        modal_desc = req_post.get('modal_desc','')
        modal_serialized = req_post.get('modal_serialized','')
        pnm_auto_key = req_post.get('pn_sel','')
        pn_info = []

        if active_mode == '1' or sel_mode=='1' or cur_mode =='1':
            from portal.tasks import get_bom_statuses,get_activities_conditions_tasks
            res = get_bom_statuses.delay(quapi_id,session_id)
            res.get()                              
            bom_statuses = StatusSelection.objects.filter(session_id = session_id).order_by('name')
            res = get_activities_conditions_tasks.delay(quapi_id,session_id,si_number = wo_task)
            error,msg = res.get()              
            wo_tasks = WOTask.objects.filter(\
                session_id = session_id,\
                ).order_by('wot_sequence')
                
            val_dict.update({
                'bom_statuses': bom_statuses,
                'wo_tasks': wo_tasks,
            })           
            
            if not part_number and quantity and wo_task:
                error = 'Need Job, Part and Quantity to proceed.'
                val_dict.update({
                    'error': error,
                    'user_id': user_id,
                    'user_name': user_name, 
                    'active_mode': active_mode or sel_mode or cur_mode,
                    'sel_mode': sel_mode or active_mode or cur_mode, 
                    'cur_mode': cur_mode or sel_mode or active_mode,
                    'session_id': session_id, 
                    'part_number': part_number,
                    'wo_task': wo_task,
                    'quantity': quantity,
                    'notes': notes, 
                    'is_submit': '0',                    
                }) 
                return render(request, 'mrolive/parts-request.html', val_dict)
                
            elif modal_pn or ('part_number' in req_post and part_number and quantity and wo_task): 
                
                if modal_pn:
                    from portal.tasks import create_pn
                    res = create_pn.delay(quapi_id,session_id,\
                        modal_pn,modal_desc,modal_serialized,\
                        sysur_auto_key)
                    error,msg = res.get() 
                    
                from portal.tasks import parts_request   
                res = parts_request.delay(quapi_id,user_name,session_id,\
                    sysur_auto_key,part_number,wo_task,quantity,\
                    notes,bom_status,wo_task_upd=wo_task_sel,\
                    yes_request=yes_request,pnm_auto_key=pnm_auto_key) 
                error,user_msg,show_conf,pn_info = res.get()

                if show_conf in ['show_pns'] and pn_info:
                    #lookup pn desc and 
                    from polls.models import PartNumbers
                    part_data,error = [],''
                    #[pnm_auto_key,part_num,part_desc,mfg_code]
                    for part in pn_info:    
                        part_data.append(PartNumbers(
                        pnm_auto_key = part[0],
                        part_number = part[1],
                        description = part[2],
                        mfg_code = part[3],
                        session_id = session_id,      
                        ))
                    if part_data:
                        try:
                            delete = PartNumbers.objects.filter(session_id=session_id).delete()
                            PartNumbers.objects.bulk_create(part_data) or []    
                        except Exception as exc:
                            error += "Error, %s, creating parts."%(exc)
                    pnms = PartNumbers.objects.filter(session_id=session_id)
                    val_dict['pn_info'] = pnms
                    
                val_dict.update({
                    'user_id': user_id,
                    'sysur_auto_key': sysur_auto_key,
                    'user_name': user_name,            
                    'session_id': session_id, 
                    'part_number': part_number,
                    'active_mode': active_mode or sel_mode or cur_mode,
                    'wo_task': wo_task,
                    'wo_number': wo_number,
                    'quantity': quantity,
                    'notes': notes,
                    'is_submit': '1',                        
                    'show_conf': show_conf,                   
                    'msg': user_msg + ' ' + msg,
                    'error': error,                   
                })                
                return render(request, 'mrolive/parts-request.html', val_dict)

            else:                        
                val_dict.update({
                    'user_id': user_id,
                    'user_name': user_name,            
                    'session_id': session_id,
                    'sysur_auto_key': sysur_auto_key,                 
                    'part_number': part_number,
                    'active_mode': active_mode or sel_mode or cur_mode,
                    'sel_mode': sel_mode or active_mode or cur_mode, 
                    'cur_mode': cur_mode or sel_mode or active_mode,
                    'wo_task': wo_task,
                    'is_submit': '0',
                    'quantity': quantity,
                    'notes': notes, 
                    'show_conf': show_conf, 
                    'msg': msg,                    
                })                
                return render(request, 'mrolive/parts-request.html', val_dict)
                           
        elif active_mode == '2' or sel_mode=='2' or cur_mode =='2':
            if wo_task or wo_number or part_number:
                #do the lookup and create grid results records in the wostatus table
                from portal.tasks import get_requests_view               
                res = get_requests_view.delay(quapi_id,session_id,user_id,wo_task,wo_number,part_number)   
                error,msg = res.get()
                requests = WOStatus.objects.filter(session_id=session_id)
                total_rows = len(requests)
                val_dict['total_rows'] = total_rows
                
    val_dict['msg'] = msg + user_msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/parts-request.html', val_dict)

def gen_fake_woo():    
    # generate random integer values
    from random import seed
    from random import randint
    # generate some integers
    fake_woo = randint(100000, 199999)
    return str(fake_woo)
    
def gen_fake_pn():    
    # generate random integer values
    from random import seed
    from random import randint
    # generate some integers
    fake_pn = str(randint(100, 999))
    fake_pn += '-' + str(randint(10000, 42299))
    fake_pn += '-' + str(randint(10, 99))
    return fake_pn 

def tools(request,quapi_id=None):
    new_status,location,filter_status,user_auto_key = '','','',''
    user_id,user_rec,fail_ms = 'user not set',None,''
    val_dict,form = {},{}   
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    
    val_dict['kiosk'] = 'F'
    if user_profile:
        #check apps that are kiosk user apps and return the flag in the view to prompt for user_id
        for app in user_profile.kiosk_apps.all():
            if app.code == 'tools':  
                val_dict['kiosk'] = 'T'
                break                
                
    #user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='tools')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    if not (reg_user_id and dj_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    val_dict['user_apps'] = user_apps
    alloc_app = MLApps.objects.filter(name="Tools")
    alloc_app = alloc_app and alloc_app[0] or None
    modes = alloc_app and get_modes(alloc_app) or []
    if not modes:
        return redirect('/login/') 
    val_dict['modes'] = modes 
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':
        form = WODashboardForm() 
        val_dict['form'] = form
        from portal.tasks import get_users_nsync_beta
        res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=0)
        user_error = res.get()
        val_dict['active_mode'] = '1'     
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)           
        active_user = 'active_user' in req_post and req_post['active_user'] or ''  
        filter_user = 'filter_user' in req_post and req_post['filter_user'] or ''        
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        user_name = 'user_name' in req_post and req_post['user_name'] or ''
        user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id) or None
        user_rec = user_rec and user_rec[0] or None
        active_mode = 'mode_selector' in req_post and req_post['mode_selector'] or ''
        sel_mode = 'sel_mode' in req_post and req_post['sel_mode'] or '' 
        if not (sel_mode or active_mode):
            val_dict['error'] = 'Must select a mode.'
            render(request, 'mrolive/tools.html', val_dict) 
        if active_mode == '1' and user_id and not user_rec:
            val_dict['error'] = 'Invalid user id.'
            return render(request, 'mrolive/tools.html', val_dict) 
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        val_dict['active_mode'] = active_mode or sel_mode
        val_dict['sel_mode'] = sel_mode or active_mode
        wo_task = req_post.get('wo_task','')
        label = req_post.get('label','')
        
        #if 'kiosk' in val_dict and val_dict['kiosk'] == 'T':
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        user_name = 'user_name' in req_post and req_post['user_name'] or ''
        if user_name:
            user_rec = user_name and QuantumUser.objects.filter(quapi_id=quapi_id,user_name__iexact=user_name) or None
            if not user_rec:
                user_rec = user_name and QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_name) or None 
            if not user_rec:
                user_rec = user_name and QuantumUser.objects.filter(quapi_id=quapi_id,employee_code__iexact=user_name) or None        
            user_rec = user_rec and user_rec[0] or None
            user_name = user_rec and user_rec.user_name or user_name or ''
            user_auto_key = user_rec and user_rec.user_auto_key or None
            
            if not user_auto_key:
                error = 'User not found.'
                val_dict.update({
                    'error': error,
                    'user_id': user_id,
                    'user_name': user_name,            
                    'session_id': session_id,
                    'label': '',
                    'wo_task': '',
                    'form': form,                    
                })
                return render(request, 'mrolive/tools.html', val_dict) 
        elif label:
            error = 'Must enter user.'
            val_dict.update({
                'error': error,
                'user_id': user_id,
                'user_name': user_name,            
                'session_id': session_id,
                'label': '',
                'wo_task': '',
                'form': form,                    
            })
            return render(request, 'mrolive/tools.html', val_dict)                    
        if active_mode == '2':
            if 'label' in req_post and not label:
                error = 'Enter the \'Job\'.'
                val_dict.update({
                    'error': error,
                    'user_id': user_id,
                    'user_name': user_name,            
                    'session_id': session_id,
                    'label': '',
                    'wo_task': '',
                    'form': form,                    
                })
                return render(request, 'mrolive/tools.html', val_dict) 
            elif 'label' in req_post and label:  
                ctrl_number,ctrl_id='',''           
                if label and len(label) > 6:
                    ctrl_number = label[:6]               
                    ctrl_id = label[7:] 
                else:
                    val_dict['label'] = ''
                    val_dict['error'] = 'Invalid tool.'
                    val_dict['form'] = form
                    return render(request, 'mrolive/tools.html', val_dict)                     
                from portal.tasks import tools_checkin
                res = tools_checkin.delay(quapi_id,session_id,sysur_auto_key,ctrl_number,ctrl_id,user_auto_key)
                error,msg = res.get()
        if active_mode == '1':
            if not label or not wo_task:
                error = 'Enter tool and job.'
                val_dict.update({
                    'error': error,
                    'user_id': user_id,
                    'user_name': user_name,            
                    'session_id': session_id, 
                    #'label': label,
                    'wo_task': wo_task,   
                    'form': form,                    
                }) 
                return render(request, 'mrolive/tools.html', val_dict) 
            else:  
                if user_rec and not user_name:
                    val_dict['error'] = 'Invalid user id.'
                    #return render(request, 'mrolive/tools.html', val_dict) 
                ctrl_number,ctrl_id='',''           
                if label and len(label) > 6:
                    ctrl_number = label[:6]               
                    ctrl_id = label[7:] 
                else:                                    
                    val_dict['error'] = 'Invalid Tool.',                        
                    
                from portal.tasks import tools_checkout
                res = tools_checkout.delay(quapi_id,session_id,sysur_auto_key,ctrl_number,ctrl_id,wo_task,user_auto_key)
                error,msg = res.get()
        val_dict.update({
            'user_id': user_id,
            'user_name': user_name,            
            'session_id': session_id, 
            'label': '',
            'wo_task': wo_task,
            'form': form,                    
        })  
        
    val_dict['msg'] = msg
    val_dict['error'] = error                
    val_dict['form'] = form  
    return render(request, 'mrolive/tools.html', val_dict)
       
from django.views.decorators.cache import cache_control

@cache_control(max_age=0, no_cache=True, no_store=True, must_revalidate=True)    
def wo_template_import(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec,fail_ms = 'user not set',None,''
    val_dict,form = {},{}   
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':
        form = WODashboardForm() 
        from portal.tasks import get_statuses_nsync_beta, get_users_nsync_beta, get_loc_whs_cart_nsync_beta 
        res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=0)
        user_error = res.get()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post, request.FILES)    
        active_user = 'active_user' in req_post and req_post['active_user'] or ''  
        filter_user = 'filter_user' in req_post and req_post['filter_user'] or ''        
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        user_name = 'user_name' in req_post and req_post['user_name'] or ''
        user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id) or None
        user_rec = user_rec and user_rec[0] or None
        sysur_auto_key = user_rec and user_rec.user_auto_key or None
        user_name = user_rec and user_rec.user_name or user_name or ''
        if user_id and user_rec and not user_name:
            val_dict['error'] = 'Invalid user id.'
            return render(request, 'mrolive/wo_template_import.html', val_dict) 
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        options_col,page_size = get_options(req_post,session_id)            
        val_dict.update({
            'msg': msg,
            'active_user': active_user or user_id,
            'filter_user': filter_user or active_user or user_id,
            'user_id': user_id or filter_user or active_user,
            'user_name': user_name,            
            'options_col': options_col,
            'page_size': page_size, 
            'session_id': session_id,                    
        }) 
        if request.FILES:           
            error,msg,fail_msg = prep_wo_temp_import(quapi_id,sysur_auto_key,request.FILES,session_id)
            val_dict['fail_msg'] = fail_msg
            bad_rows = WOStatus.objects.filter(session_id=session_id) 
            total_rows = len(bad_rows)
            val_dict['total_rows'] = total_rows
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/wo_template_import.html', val_dict) 

@cache_control(max_age=0, no_cache=True, no_store=True, must_revalidate=True)    
def wo_template_measures(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec,fail_ms = 'user not set',None,''
    val_dict,form = {},{}   
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':
        form = WODashboardForm() 
        #from portal.tasks import get_statuses_nsync_beta, get_users_nsync_beta, get_loc_whs_cart_nsync_beta 
        #res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=0)
        #user_error = res.get()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post, request.FILES)    
        active_user = 'active_user' in req_post and req_post['active_user'] or ''  
        filter_user = 'filter_user' in req_post and req_post['filter_user'] or ''        
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        """user_name = 'user_name' in req_post and req_post['user_name'] or ''
        user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id) or None
        user_rec = user_rec and user_rec[0] or None
        sysur_auto_key = user_rec and user_rec.user_auto_key or None
        user_name = user_rec and user_rec.user_name or user_name or ''"""
        if user_id and user_rec and not user_name:
            val_dict['error'] = 'Invalid user id.'
            return render(request, 'mrolive/wo_template_import.html', val_dict) 
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        options_col,page_size = get_options(req_post,session_id)            
        val_dict.update({
            'msg': msg,
            'active_user': active_user or user_id,
            'filter_user': filter_user or active_user or user_id,
            'user_id': user_id or filter_user or active_user,
            'user_name': user_name,            
            'options_col': options_col,
            'page_size': page_size, 
            'session_id': session_id,                    
        })
  
        if request.FILES:           
            error,msg,fail_msg = prep_wo_temp_meas_import(quapi_id,sysur_auto_key,request.FILES,session_id)
            val_dict['fail_msg'] = fail_msg
            bad_rows = WOStatus.objects.filter(session_id=session_id) 
            total_rows = len(bad_rows)
            val_dict['total_rows'] = total_rows
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/wo_template_import.html', val_dict)
    
def prep_wo_temp_meas_import(quapi_id,sysur_auto_key,req_files,session_id):
    error,msg,fail_msg = '','','' 
    up_file = 'loc_whs_file' in req_files or None
    from portal.tasks import wo_template_measures
    if up_file:
        up_file = req_files['loc_whs_file']
    csv,error = None,''
    if up_file:
        csv = Document(session_id=session_id,docfile=up_file)
        csv.save()
    if csv:
        #full_path = os.path.join(FILE_PATH + up_file.name)
        res = wo_template_measures.delay(quapi_id,sysur_auto_key,session_id)
        error,msg,fail_msg = res.get()
        #now, if there is a csv that was saved, delete it!
        #os.remove(os.path.join(FILE_PATH + up_file.name))
    return error,msg,fail_msg  
    
def prep_wo_temp_import(quapi_id,sysur_auto_key,req_files,session_id):
    error,msg,fail_msg = '','','' 
    up_file = 'loc_whs_file' in req_files or None
    from portal.tasks import wo_template_create
    if up_file:
        up_file = req_files['loc_whs_file']
    csv,error = None,''
    if up_file:
        csv = Document(session_id=session_id,docfile=up_file)
        csv.save()
    if csv:
        #full_path = os.path.join(FILE_PATH + up_file.name)
        res = wo_template_create.delay(quapi_id,sysur_auto_key,session_id)
        error,msg,fail_msg = res.get()
        #now, if there is a csv that was saved, delete it!
        #os.remove(os.path.join(FILE_PATH + up_file.name))
    return error,msg,fail_msg    
     
def loc_whs_create(quapi_id,sysur_auto_key,req_files,session_id):
    error,msg,fail_msg = '','','' 
    up_file = 'loc_whs_file' in req_files or None
    if up_file:
        up_file = req_files['loc_whs_file']
    csv,error = None,''
    if up_file:
        csv = Document(docfile=up_file)
    if csv:
        csv.save()
        full_path = "/home/ubuntu/mo_template/media/" + up_file.name
        error,msg,fail_msg = import_loc_whs(quapi_id,session_id,sysur_auto_key,csv,full_path)
        #now, if there is a csv that was saved, delete it!
        os.remove(full_path)
    return error,msg,fail_msg
    
def import_loc_whs(quapi_id,session_id,sysur_auto_key,csv_import,file_path,chunk_size=2500):
    error,msg,fail_msg,dupe,num_failures = '','',False,'',0
    row_list,locations,good_rows,warehouses,loc_data = [],[],[],[],[]
    with open(file_path, mode='r') as csv_file:  
        row_list = [{k: v for k, v in row.items()} for row in csv.DictReader(csv_file)]
        line_count = 0
        bad_rows = 0
        for row in row_list:
            error = ''
            location_code = row.get('location_code','')           
            whs_code = row.get('warehouse_code','')
            location_name = row.get('location_name','')           
            whs_name = row.get('warehouse_name','')
            if error:
                loc_data.append(WarehouseLocation(
                    location_code = location_code,
                    location_name = location_name,
                    whs_code = whs_code,
                    whs_name = whs_name,
                    bulk_imp_error = error,
                    session_id = session_id,
                ))
            locations.append(location_code)
            if not error:
                good_rows.append(row)
            good_rows_cond = good_rows and (((len(row_list) > 1 and line_count) and line_count+1 % int(chunk_size)==0) or line_count==len(row_list)-1) or None
            if good_rows_cond:
                from portal.tasks import loc_whs_bulk
                res = loc_whs_bulk.delay(quapi_id,session_id,sysur_auto_key,good_rows)
                error,msg,task_bad_rows = res.get()
                good_rows = []
            line_count += 1            
        if loc_data:            
            #try:
            WarehouseLocation.objects.filter(session_id=session_id).delete()
            WarehouseLocation.objects.bulk_create(loc_data) or []    
            #except Exception as exc:
            #    error += "Error with creating warehouse/locations locally for display in grid: %s"%(exc)
            fail_msg = ' ' + str(bad_rows + task_bad_rows) + ' rows could not be imported as shown below in the grid. '
    return error,msg,fail_msg

def loc_whs_import(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec,fail_ms = 'user not set',None,''
    val_dict,form = {},{} 
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':
        form = WODashboardForm() 
        from portal.tasks import get_statuses_nsync_beta, get_users_nsync_beta, get_loc_whs_cart_nsync_beta 
        res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=0,app='loc-whs-import')
        user_error,app = res.get()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post, request.FILES)    
        active_user = 'active_user' in req_post and req_post['active_user'] or ''  
        filter_user = 'filter_user' in req_post and req_post['filter_user'] or ''        
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        """user_name = 'user_name' in req_post and req_post['user_name'] or ''
        user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id) or None
        user_rec = user_rec and user_rec[0] or None
        sysur_auto_key = user_rec and user_rec.user_auto_key or None
        user_name = user_rec and user_rec.user_name or user_name or ''
        if user_id and user_rec and not user_name:
            val_dict['error'] = 'Invalid user id.'
            return render(request, 'mrolive/wo_template_import.html', val_dict)"""
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0 
        options_col,page_size = get_options(req_post,session_id)            
        val_dict.update({
            'msg': msg,
            'active_user': active_user or user_id,
            'filter_user': filter_user or active_user or user_id,
            'user_id': user_id or filter_user or active_user,
            'user_name': user_name,            
            'sel_rows': sel_rows,
            'total_rows': total_rows,
            'options_col': options_col,
            'page_size': page_size, 
            'session_id': session_id,                    
        })
        woos_list = req_post.getlist('woos_list[]')
        if woos_list:
            #must send those ids over to the barcode label printing method to print them
            val_dict['woos_list'] = woos_list
            loc_list = WarehouseLocation.objects.filter(id__in=woos_list)
            loc_list = loc_list and loc_list.values_list('location_code',flat=True) or []
            loc_list = list(loc_list)
            #now return barcode labels with template under the guise of never leaving the page.
            return barcode_labels(request,quapi_id,loc_list,filter_user,user_name)
            #path = '/barcode-labels/' + quapi_id + '/' + loc_list + '/'
            #return redirect(path)
        elif request.FILES:         
            error,msg,fail_msg = loc_whs_create(quapi_id,sysur_auto_key,request.FILES,session_id)
            val_dict['fail_msg'] =fail_msg            
            all_lwhs = WarehouseLocation.objects.filter(session_id=session_id) 
            val_dict['total_rows'] = len(all_lwhs)
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/loc_whs_import.html', val_dict)  
                                                                                                                                                                                                                                                    
def jc_import(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec = 'user not set',None
    val_dict,form = {},{}
    all_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id='') 
    updated_woos = all_woos    
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id        
    if request.method == 'GET':
        form = WODashboardForm() 
        from portal.tasks import get_statuses_nsync_beta, get_users_nsync_beta, get_loc_whs_cart_nsync_beta 
        res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=0)
        user_error = res.get()         
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post, request.FILES)
        exact_match = 'exact_match' in req_post and req_post['exact_match'] or ''
        if exact_match:
            exact_match = 'checked'          
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        user_name = 'user_name' in req_post and req_post['user_name'] or ''
        user_rec = user_id and QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id) or None
        user_rec = user_rec and user_rec[0] or None
        sysur_auto_key = user_rec and user_rec.user_auto_key or None
        user_name = user_rec and user_rec.user_name or user_name or ''
        if user_id and user_rec and not user_name:
            val_dict['error'] = 'Invalid user id.'
            return render(request, 'mrolive/jcimport.html', val_dict) 
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0 
        options_col,page_size = get_options(req_post,session_id)            
        val_dict.update({
            'all_woos': updated_woos, 
            'msg': msg,
            'user_id': user_id or '',
            'user_name': user_name or '',            
            'wo_number': wo_number,
            'session_id': session_id,
            'sel_rows': sel_rows,
            'total_rows': total_rows,
            'options_col': options_col,
            'page_size': page_size, 
            'session_id': session_id,
                            
        })
        if wo_number and request.FILES:         
            error,msg,fail_msg = mbr_bulk_create(quapi_id,sysur_auto_key,request.FILES,session_id,wo_number) 
            val_dict['fail_msg'] =fail_msg            
            all_woos = WOStatus.objects.filter(session_id=session_id)
            val_dict['all_woos'] = all_woos   
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['total_rows'] = str(len(updated_woos) or len(all_woos))
    val_dict['form'] = form
    return render(request, 'mrolive/jcimport.html', val_dict)     
     
def mbr_bulk_create(quapi_id,sysur_auto_key,req_files,session_id,wo_number): 
    error,msg = '','' 
    up_file = 'file' in req_files or None
    if up_file:
        up_file = req_files['file']
    csv,error = None,''
    if up_file:
        csv = Document(docfile=up_file)
    if csv:
        csv.save()
        error,msg,fail_msg = convert_csv_dict(quapi_id,sysur_auto_key,os.path.join(FILE_PATH + up_file.name),req_files,session_id,wo_number)
        #now, if there is a csv that was saved, delete it!
        os.remove(os.path.join(FILE_PATH + up_file.name))
    return error,msg,fail_msg
    
def convert_csv_dict(quapi_id,sysur_auto_key,file_path,req_files,session_id,wo_number):
    error,msg,fail_msg,dupe,num_failures = '','',False,'',0
    row_list,titles,good_rows = [],[],[]
    from portal.tasks import task_insertion
    with open(file_path, mode='r') as csv_file:
        #row_list = csv.DictReader(csv_file)
        row_list = [{k: v for k, v in row.items()} for row in csv.DictReader(csv_file)]
        line_count = 0
        mbr_data = []
        for row in row_list:
            error = ''
            start_date = row.get('start_date',None)
            from portal.tasks import format_start_date
            input_format = '%m/%d/%Y' 
            new_format = '%Y-%m-%d'
            start_date = start_date and format_start_date(start_date,new_format) or None            
            wot_sequence = row.get('wot_sequence',0) 
            if not wot_sequence: 
                error += 'Must have a wot_sequence.' 
            else:                
                if not wot_sequence.isdigit():
                    wot_sequence = 0
                    error += 'Sequence must be numeric.'
                    dupe=True
            task_ref = row.get('task_ref',None)
            task_desc = row.get('task_desc',None)
            task_position = row.get('task_position',None)
            task_title = row.get('task_title',None)
            if not task_title: 
                error += 'Must have a task_title.' 
                dupe=True
            task_close_reqs = row.get('task_close_reqs',None)
            skill = row.get('skill','')
            est_hours = row.get('skills_est_hours',0)
            if est_hours and not str(est_hours).replace('.','',1).isdigit():
                est_hours = 0
                error += 'Hours must be numeric.'
                dupe=True
            if skill and not est_hours:
                error += 'Skill was entered but no value found for estimated hours. Must have both or none.'
            if est_hours:
                if not skill:
                    error += 'Estimated hours entered but no value found for skill. Must have both or none.' 
                    dupe = True                    
                else:
                    if est_hours.replace('.','',1).isdigit():
                        est_hours = float(est_hours)
                    else:
                        est_hours = 0
                        error += 'Estimated hours must be numeric.'
                    dupe=True
            if not task_desc:
                error += "Description required."
                dupe = True  
            if line_count==0 and not (start_date or wot_sequence or task_ref or task_title or task_close_reqs):
                error = 'Incorrect column headings.'
                break
            else:                  
                if task_title in titles:
                    error += 'Another row was found with that task_title.'
                    dupe = True                                    
                if dupe and error:  
                    titles.append(task_title)                
                    mbr_data.append(WOStatus(
                        wo_number = wo_number,
                        start_date = start_date,
                        wot_sequence = wot_sequence,
                        task_master_desc = task_desc,
                        task_ref = task_ref,
                        task_position = task_position,
                        task_title = task_title,
                        task_close_reqs = task_close_reqs,
                        skill_desc = skill,
                        wot_est_hours = est_hours,
                        bulk_imp_error = error,
                        session_id = session_id,
                    ))
                else:
                    good_rows.append(row)
            line_count += 1
            dupe = False
        try:
            res = task_insertion.delay(quapi_id,session_id,sysur_auto_key,good_rows,wo_number)
            error,msg,fail_msg,num_failures,wot_auto_key = res.get()
            if error == 'WO does not exist.':
                return error,msg,fail_msg    
        except Exception as exc:
            error += "Error with creating new tasks in Quantum from bulk import: %s"%exc 
        if mbr_data:         
            try:
                rec = WOStatus.objects.bulk_create(mbr_data) or []    
            except Exception as exc:
                error += "%s | Error with creating unimportable tasks for WO# from bulk: %s"%(exc,wo_number)
            #fail_msg += str(len(row_list) - len(good_rows))+ ' rows could not be imported as shown below in the grid.'
            num_failures += len(mbr_data)
            fail_msg += ' ' + str(num_failures) + ' rows could not be imported as shown below in the grid. '
            
    return error,msg,fail_msg
                                         
def wo_labor(request,quapi_id=None):
    new_status,location,filter_status,session_id = '','','',''
    user_id,user_rec,wtms = 'user not set',None,None
    batch_num,batch_no = '',''
    val_dict,form = {},{}
    wot_id_list = []
    error,msg,loc_msg,stat_msg = '','','',''
    recall_mode,create_mode = False,False
    from polls.models import LaborBatch as batches,QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    val_dict['active_mode'] = 'tracking' 
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    val_dict['kiosk'] = 'F'
    
    if user_profile:
        #check apps that are kiosk user apps and return the flag in the view to prompt for user_id
        for app in user_profile.kiosk_apps.all():
            if app.code == 'labor-management':  
                val_dict['kiosk'] = 'T'
                break

    if val_dict['kiosk'] != 'T':                
        from portal.tasks import get_batches
        import random
        import string
        session_id = ''.join(random.choices(string.ascii_uppercase + string.digits, k=9))
        res = get_batches.apply_async(
            queue='labor', 
            priority=1, 
            kwargs={
                'quapi_id':quapi_id,
                'sysur_auto_key':sysur_auto_key,
                'session_id':session_id,
                'per_user': True,
            })
        #res = get_batches.delay(quapi_id,sysur_auto_key,session_id)
        error,msg = res.get()
        all_batches = batches.objects.all().order_by('-batch_id')        
        from portal.tasks import get_wtls
        res = get_wtls.apply_async(
            queue='labor', 
            priority=1, 
            kwargs={
                'quapi_id':quapi_id,
                'sysur_auto_key':sysur_auto_key
            })
        #res = get_wtls.delay(quapi_id,sysur_auto_key)
        error,msg = res.get()
        open_wtls = TaskLabor.objects.filter(sysur_auto_key=sysur_auto_key).order_by('-id') 
        val_dict['button_txt'] = 'Start Task'
        open_batches = all_batches.filter(session_id=session_id,sysur_auto_key=sysur_auto_key,stop_time__isnull=True,start_time__isnull=False)
        open_batch = open_batches and open_batches[0] or None 
        val_dict['batches'] = all_batches        
        if open_batch:
            val_dict['active_mode'] = 'batch-recall'
            val_dict['sel_mode'] = 'batch-recall'
            val_dict['conf_sel_mode'] = 'batch-recall'
            val_dict['button_txt'] = 'Stop Batch'  
            val_dict['batch_no'] = open_batch.batch_id
            start_time = open_batch.start_time 
            start_time = start_time and start_time - timedelta(hours=4)
            start_time = start_time.strftime('%m-%d-%Y %H:%M:%S')              
            val_dict['msg'] = ' started ' + start_time + '.'
            val_dict['description'] = open_batch.description
            val_dict['start_batch'] = 'T'           
            val_dict['batch_recall'] = open_batch.batch_id
            val_dict['total_labor_rows'] = len(open_wtls)
        elif open_wtls:
            val_dict['msg'] = 'WO #: ' + open_wtls[0].wo_number
            val_dict['msg'] += ' | Seq: ' + str(open_wtls[0].sequence)
            val_dict['msg'] += ' | ' + open_wtls[0].task_desc
            start_time = open_wtls[0].start_time 
            start_time = start_time and start_time - timedelta(hours=4)
            start_time = start_time.strftime('%m-%d-%Y %H:%M:%S')              
            val_dict['msg'] += ' started ' + start_time + '.'
            val_dict['button_txt'] = 'Stop Task'
            val_dict['wot_key'] = str(open_wtls[0].wot_auto_key) + 'c'     
                    
    user_name = user and val_dict['kiosk'] == 'F' and user.username or ''
    val_dict['user_name'] = user_name
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id  
    alloc_app = MLApps.objects.filter(name="Labor Tracking").order_by('-id')
    alloc_app = alloc_app and alloc_app[0] or None
    modes = AppModes.objects.filter(ml_apps_id=alloc_app).order_by('-code')
    val_dict['modes'] = modes   
    if request.method == 'GET':
        form = WODashboardForm()
        if val_dict['kiosk'] == 'T':
            val_dict['button_txt'] = 'OK'
            
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        val_dict['form'] = form
        remove_wots = req_post.get('remove_wots','0')
        add_task = req_post.get('add_task','0')
        add_wots = req_post.get('add_wots','0')
        task_submit = req_post.get('task_submit','0') 
        search_user = req_post.get('search_user','F')  
        show_conf = req_post.get('show_conf','F')                
        labor_submit = 'labor_submit' in req_post and req_post['labor_submit'] or ''
        clear_form = 'clear_form' in req_post and req_post['clear_form'] or ''
        start_batch = 'start_batch' in req_post and req_post['start_batch'] or ''
        stop_batch = 'stop_batch' in req_post and req_post['stop_batch'] or ''
        batch_button = 'batch_button' in req_post and req_post['batch_button'] or ''
        button_txt = 'button_txt' in req_post and req_post['button_txt'] or ''
        val_dict['button_txt'] = button_txt
        c_button_txt = 'c_button_txt' in req_post and req_post['c_button_txt'] or ''
        conf_batch_button = 'conf_batch_button' in req_post and req_post['conf_batch_button'] or ''
        conf_batch_no = 'conf_batch_no' in req_post and req_post['conf_batch_no'] or ''
        conf_msg = 'conf_msg' in req_post and req_post['conf_msg'] or ''
        batch_no = 'batch_no' in req_post and req_post['batch_no'] or ''
        description = req_post.get('description','') or req_post.get('conf_description','')                                                                                    
        yes_complete = req_post.get('yes_complete','F')  
        not_complete = req_post.get('not_complete','F')         
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        wot_auto_key = 'wot_auto_key' in req_post and req_post['wot_auto_key'] or ''
        wot_key = 'wot_key' in req_post and req_post['wot_key'] or ''
        single_wot = 'single_wot' in req_post and req_post['single_wot'] or ''
        single_wot_key = 'single_wot_key' in req_post and req_post['single_wot_key'] or ''
        active_mode = 'mode_selector' in req_post and req_post['mode_selector'] or ''
        sel_mode = 'sel_mode' in req_post and req_post['sel_mode'] or ''
        conf_sel_mode = 'conf_sel_mode' in req_post and req_post['conf_sel_mode'] or '' 
        recall_mode = conf_sel_mode == 'batch-recall' or active_mode == 'batch-recall' or sel_mode == 'batch-recall' or ''
        create_mode = conf_sel_mode == 'create-batch' or active_mode == 'create-batch' or sel_mode == 'create-batch' or ''
        user_logged = 'user_logged' in req_post and req_post['user_logged'] or ''
        active_user = 'active_user' in req_post and req_post['active_user'] or ''
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        if not user_name:
            user_name = 'user_name' in req_post and req_post['user_name'] or ''
        if not user_name:
            user_name = user_logged

        if not user_name:
            user_name = active_user

        if button_txt == 'Start Batch' and not labor_submit: 
            
            if batch_no and add_task == '0' and wo_number:
                
                if wo_number[-1] in ['s','S']:
                    add_task = '1'
                    
                else:
                    #lookup the WO's tasks and get the count for total_rows
                    from portal.tasks import get_signoff_tasks
                    res = get_signoff_tasks.apply_async(
                    queue='labor',
                    priority=1, 
                    kwargs={
                        'quapi_id':quapi_id,
                        'session_id':session_id,
                        'wo_number':wo_number,
                    })
                    #res = get_signoff_tasks.delay(quapi_id,session_id,wo_number=wot_auto_key)
                    error,msg = res.get()          
                    wtms = WOTask.objects.filter(session_id=session_id).exclude(status_type='Closed') or []
                    total_rows = len(wtms)
                    val_dict.update({
                        'total_rows': total_rows,
                    })
                    
        if 'kiosk' in val_dict and val_dict['kiosk'] == 'T' and (user_id or user_name or user_logged or active_user):
            user_rec = None
            if user_id:        
                user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id) or None
            if not user_rec and (user_id or user_name or user_logged):
                user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_name__iexact=user_id or user_name or user_logged) or None
            if not user_rec and (user_id or user_name or user_logged):
                user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,employee_code__iexact=user_id or user_name or user_logged) or None
            user_rec = user_rec and user_rec[0] or None
            if not user_rec:
                val_dict['error'] = 'User not found.'
                val_dict['active_mode'] = active_mode or sel_mode or conf_sel_mode
                val_dict['form'] = form
                return render(request, 'mrolive/wolabor.html', val_dict) 
            user_name = user_rec and user_rec.user_name or user_name or ''
            if not user_name:
                val_dict['error'] = 'User name cannot be blank.'
                val_dict['active_mode'] = active_mode or sel_mode or conf_sel_mode
                val_dict['form'] = form
                return render(request, 'mrolive/wolabor.html', val_dict) 
            val_dict['user_name'] = user_name
            val_dict['active_user'] = user_name
            sysur_auto_key = user_rec and user_rec.user_auto_key or None
            from portal.tasks import get_batches
            res = get_batches.apply_async(
                queue='labor', 
                priority=1, 
                kwargs={
                    'quapi_id':quapi_id,
                    'sysur_auto_key':sysur_auto_key,
                    'session_id':session_id,
                })
            #res = get_batches.delay(quapi_id,sysur_auto_key,session_id)
            error,msg = res.get()
            all_batches = batches.objects.filter(session_id=session_id).order_by('-batch_id')
            from portal.tasks import get_wtls
            res = get_wtls.apply_async(
                queue='labor', 
                priority=1, 
                kwargs={
                    'quapi_id':quapi_id,
                    'sysur_auto_key':sysur_auto_key,
                })
            #res = get_wtls.delay(quapi_id,sysur_auto_key)
            error,msg = res.get()
            open_wtls = TaskLabor.objects.filter(sysur_auto_key=sysur_auto_key).order_by('-id') 
            if recall_mode:
                val_dict['button_txt'] = 'Recall Batch'
            elif create_mode:
                val_dict['button_txt'] = 'Create'
            else:
                val_dict['button_txt'] = 'Start'
            open_batch = None
                                                 
            if sysur_auto_key:
                open_batches = all_batches.filter(sysur_auto_key=sysur_auto_key,stop_time__isnull=True,start_time__isnull=False)
                open_batch = open_batches and open_batches[0] or None
            val_dict['batches'] = all_batches            
            if open_batch:
                val_dict['active_mode'] = 'batch-recall'
                val_dict['sel_mode'] = 'batch-recall'
                val_dict['conf_sel_mode'] = 'batch-recall'
                val_dict['button_txt'] = 'Stop Batch'               
                val_dict['batch_no'] = open_batch.batch_id
                start_time = open_batch.start_time 
                start_time = start_time and start_time - timedelta(hours=4)
                start_time = start_time.strftime('%m-%d-%Y %H:%M:%S')              
                val_dict['description'] += open_batch.description + ' started ' + start_time + '.'
                val_dict['start_batch'] = 'T'           
                val_dict['batch_recall'] = open_batch.batch_id
            elif open_wtls:
                val_dict['msg'] = 'WO #: ' + open_wtls[0].wo_number
                val_dict['msg'] += ' | Seq: ' + str(open_wtls[0].sequence)
                val_dict['msg'] += ' | ' + open_wtls[0].task_desc
                start_time = open_wtls[0].start_time 
                start_time = start_time and start_time - timedelta(hours=4)
                start_time = start_time.strftime('%m-%d-%Y %H:%M:%S')              
                val_dict['msg'] += ' started ' + start_time + '.'
                val_dict['button_txt'] = 'Stop Task'
                val_dict['wot_key'] = str(open_wtls[0].wot_auto_key) + 'c'
                val_dict['wot_auto_key'] = str(open_wtls[0].wot_auto_key) + 'c'
            if search_user == 'T' and (open_wtls or open_batch):
                val_dict['form'] = form
                return render(request, 'mrolive/wolabor.html', val_dict)             
        #if not user_id and not user_auto_key and recall_mode:
        user_auto_key = sysur_auto_key
        #elif user_id and recall_mode and not user_auto_key:
        #    val_dict['error'] = 'User not found.'
        #    val_dict['active_mode'] = active_mode or sel_mode or conf_sel_mode
        #    val_dict['sel_mode'] = sel_mode or active_mode or conf_sel_mode  
        #    val_dict['conf_sel_mode'] = conf_sel_mode or active_mode or sel_mode  
        #    render(request, 'mrolive/wolabor.html', val_dict) 
        active_batch = 'batch_recall' in req_post and req_post['batch_recall'] or ''
        """if 'batch_recall' in req_post and recall_mode and not active_batch:
            val_dict['error'] = 'Must select batch for recall.'
            val_dict['active_mode'] = 'batch-recall'
            val_dict['sel_mode'] = 'batch-recall'
            val_dict['conf_sel_mode'] = 'batch-recall' 
 
            render(request, 'mrolive/wolabor.html', val_dict) """           
        if user_auto_key and recall_mode:
            from portal.tasks import get_batches
            res = get_batches.apply_async(
                queue='labor', 
                priority=1, 
                kwargs={
                    'quapi_id':quapi_id,
                    'sysur_auto_key':sysur_auto_key,
                    'session_id':session_id,
                    'per_user': True,
                })
            #res = get_batches.delay(quapi_id,sysur_auto_key,session_id)
            error,msg = res.get()
            all_batches = batches.objects.all().order_by('-batch_id')
            val_dict['batches'] = all_batches
        
        if not wo_number:
            wo_number = wot_auto_key
        if not wot_auto_key:
            wot_auto_key = wo_number
        
        #total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        options_col,page_size = get_options(req_post,session_id)   
        #if user submitted clear list form by pressing button

        if clear_form and req_post['clear_form']=='1':
            #user_logged = 'user_logged' in req_post and req_post['user_logged'] or ''
            val_dict['user_id'] = user_name or user_logged or active_user or user_id             
            val_dict['user_name'] = user_name or user_logged or active_user or user_id       
            val_dict['user_logged'] = user_name or user_logged or active_user or user_id      
            val_dict['active_user'] = user_name or user_logged or active_user or user_id      
            val_dict['user_id'] = user_id            
            val_dict['all_woos'] = []            
            val_dict['msg'] = ''
            val_dict['active_mode'] = sel_mode
            val_dict['show_all'] = 1  
            
            if recall_mode:
                
                val_dict['button_txt'] = 'Recall Batch'        
                all_batches = batches.objects.all().order_by('-batch_id')
                val_dict['batches'] = all_batches

            elif create_mode:
                val_dict['button_txt'] = 'Create'
                
            else:
                val_dict['button_txt'] = 'Start Task'
                
            if open_batch:
                val_dict['button_txt'] = 'Stop Batch'
                val_dict['active_mode'] = 'batch-recall'
            
            elif open_wtls:
                val_dict['button_txt'] = 'Stop Task'
                val_dict['active_mode'] = 'tracking'
                
            form = WODashboardForm(val_dict)
            val_dict['form'] = form          
            return render(request, 'mrolive/wolabor.html', val_dict)             
        
        val_dict.update({
            'all_woos': '',          
            #'button_txt': button_txt or 'Start',
            'user_id': user_name or user_logged or active_user,
            'user_name': user_name or user_logged or active_user, 
            'user_logged': user_logged or user_name or active_user,              
            'wo_number': '',
            'single_wot': '',
            'wot_key': wot_key or wot_auto_key,
            'wot_auto_key': wot_auto_key or wot_key,
            'session_id': session_id,
            'options_col': options_col,
            'page_size': page_size, 
            'show_conf': show_conf,
            'sysur_auto_key': sysur_auto_key,  
            'start_batch': start_batch,
            'stop_batch': stop_batch,
            'batch_button': batch_button or 'Start Batch',
            'main_button': 'Update',
            'batch_no': conf_batch_no or batch_no or description,
            'description': description or batch_no,
            'active_mode': active_mode or sel_mode or conf_sel_mode,
            'sel_mode': sel_mode or active_mode or conf_sel_mode,  
            'conf_sel_mode': conf_sel_mode or active_mode or sel_mode, 
            'conf_msg': conf_msg or msg,
            'msg': msg or conf_msg,            
        })
        
        #import pdb;pdb.set_trace()
        #if the user recalls a batch, we need to send the request to batch_labor
        #and create a new lbh_auto_key with new ldb_auto_key(s)
        wot_id_list = req_post.getlist('woos_list[]')
        
        if add_wots == 'T':
            
            if wot_id_list:              
                from portal.tasks import batch_task_add
                
                res = batch_task_add.apply_async(
                    queue='labor', 
                    priority=1,
                    kwargs={
                        'quapi_id':quapi_id,
                        'session_id':session_id or update_session,
                        'sysur_auto_key':sysur_auto_key,
                        'batch_id':batch_no or conf_batch_no or description,
                        'wot_list': wot_id_list,
                        'user_name': user_name,
                    })
                error,msg = res.get()
                    
                val_dict['error'] = error
                val_dict['button_txt'] = 'Start Batch'
                val_dict['session_id'] = session_id or update_session
                labor_rows = TaskLabor.objects.filter(session_id=session_id)
                val_dict['total_labor_rows'] = len(labor_rows)
                val_dict['total_rows'] = ''
                val_dict['batch_no'] = batch_no or active_batch or conf_batch_no
                
                if not error:
                    val_dict['msg'] = msg
                    
            else:
                val_dict['error'] = 'Must select at least one task.'
                
            val_dict['batch_no'] = batch_no or active_batch or conf_batch_no
            return render(request, 'mrolive/wolabor.html', val_dict)
                
        if remove_wots == 'T':
            
            batch_stopped = False
            if wot_id_list:
               
                no_wtls = False
                if button_txt == 'Start Batch':
                    no_wtls = True

                labor_rows = TaskLabor.objects.filter(session_id=session_id)
                total_rows = len(labor_rows)
                
                stop_batch = False
                
                if total_rows == len(wot_id_list):
                    stop_batch = True
                    
                from portal.tasks import batch_task_remove
                #for wot in wot_id_list:
                res = batch_task_remove.apply_async(
                    queue='labor', 
                    priority=1, 
                    kwargs={
                        'quapi_id':quapi_id,
                        'session_id':session_id or update_session,
                        'sysur_auto_key':sysur_auto_key,
                        'batch_id':batch_no or conf_batch_no or description,
                        'wot_auto_key':'',
                        'wot_list': wot_id_list,
                        'no_wtls':no_wtls,
                        'stop_batch':stop_batch,
                        'user_name': user_name,
                    })
                error,msg,batch_stopped,new_batch_no = res.get()
                    
                val_dict['error'] = error
                val_dict['button_txt'] = 'Start Batch'
                val_dict['session_id'] = session_id or update_session
                labor_rows = TaskLabor.objects.filter(session_id=session_id)
                val_dict['total_labor_rows'] = len(labor_rows)
                
                if batch_stopped:
                    val_dict['button_txt'] = 'Recall Batch'
                    
                else:
                    val_dict['batch_no'] = new_batch_no or batch_no or active_batch or conf_batch_no
                
                if not error:
                    val_dict['msg'] = 'Successful removal of labor from batch. ' + msg
                    
            else:
                val_dict['error'] = 'Must select at least one task.'
                
            return render(request, 'mrolive/wolabor.html', val_dict) 
            
        #if we aren't logging in as a user at Kiosk...
        if open_wtls:
            wot_auto_key = str(open_wtls[0].wot_auto_key)+'c'
            wot_key = wot_auto_key
            single_wot = wot_auto_key
                      
        if search_user != 'T':

            if (wot_auto_key and wot_auto_key[-1] in ['c','C'] or button_txt in ['Stop Batch','Stop Task']\
                or labor_submit in ['Stop Batch','Stop Task']) and show_conf in ['','F','got_data']: 
                    
                if show_conf != 'got_data':
                    val_dict['show_conf']='T'
                    
                if not open_wtls or open_wtls and len(open_wtls) == 1:
                    val_dict['button_txt'] = 'Start Task'
                 
                if open_batch:
                    labor_rows = TaskLabor.objects.filter(session_id=session_id)
                    val_dict['total_labor_rows'] = len(labor_rows)
                
                show_conf = 'got_data'
                val_dict['single_wot_key'] = single_wot
                val_dict['user_id'] = user_name or user_logged\
                    or active_user or user_id             
                val_dict['user_name'] = user_name or user_logged\
                    or active_user or user_id 
                if recall_mode:
                    val_dict['button_txt'] = 'Stop Batch'
                    val_dict['batch_no'] = batch_no or description
                    val_dict['description'] = description or batch_no
            
            #if recall_mode and show_conf in ['','F','got_data']\
            if yes_complete != 'T' and not_complete != 'T'\
            and recall_mode and show_conf in ['','F','got_data']\
            and (task_submit == 'Add Task' or\
            labor_submit not in ['Recall Batch','Stop Batch']):   
                
                if conf_sel_mode:
                    val_dict.update({
                        'button_txt': 'Recall Batch',
                        'error': error,
                        'msg': msg,
                        'form': form
                    })
                    return render(request, 'mrolive/wolabor.html', val_dict)   
                    
                start_batch = False
                start_tasks = False
                from portal.tasks import batch_labor 
                val_dict['button_txt'] = 'Start Batch'
               
                if labor_submit == 'Start Batch':
                    start_tasks = True
                    start_batch = True
                    val_dict['button_txt'] = 'Stop Batch'
                    
                elif button_txt != 'Stop Batch':
                    if button_txt == 'Recall Batch':
                        start_batch = True
                                           
                        wot_auto_key = ''
                        batch_no = active_batch
                    start_tasks = False
                   
                total_rows = 0
                wot_id_list = req_post.getlist('woos_list[]')

                if wot_id_list:

                    res = batch_labor.apply_async(
                        queue='labor', 
                        priority=1, 
                        kwargs={
                            'quapi_id':quapi_id,
                            'session_id':session_id,
                            'sysur_auto_key':sysur_auto_key,
                            'wot_auto_key':'',
                            'add_tasks_only': True,
                            'batch_no':batch_no,
                            'active_batch':batch_no,
                            'start_batch':False,
                            'create_batch':False,
                            'start_tasks':False,
                            'wot_ids': wot_id_list,
                            'user_name':user_name,
                        })                        
                        
                    error,batch_num,description,msg = res.get()
                    labor_rows = TaskLabor.objects.filter(session_id=session_id)
                    labor_rows = len(labor_rows)
                    
                    val_dict.update({                    
                        'total_labor_rows': labor_rows,                
                        'batch_no': batch_no or batch_num,
                        'description': batch_num,
                        'wo_number': '',
                        'wot_auto_key': '',
                        'msg': msg,
                        'active_mode': 'batch-recall',
                        'sel_mode': 'batch-recall',
                        'conf_sel_mode': 'batch-recall',
                        'error': error, 
                        'start_batch': 'T',            
                        'button_txt': 'Start Batch',
                    })
                    
                elif wot_auto_key and not wot_auto_key[-1] in ['s','S'] and not wot_id_list and add_task=='1':
                    error,msg,total_rows=get_tasks(quapi_id,session_id,wo_number=wot_auto_key)
                    wtms = WOTask.objects.filter(session_id=session_id)

                    val_dict.update({                    
                        'batch_no': batch_no or batch_num,
                        'description': batch_num,
                        'wo_number': '',
                        'total_rows': total_rows,
                        'wot_auto_key': '',
                        'msg': msg,
                        'active_mode': 'batch-recall',
                        'sel_mode': 'batch-recall',
                        'conf_sel_mode': 'batch-recall',
                        'error': error, 
                        'start_batch': 'T',            
                        'button_txt': 'Start Batch',
                    })
                
                elif start_batch or add_task == '1': 
                    if add_task != '1':
                        wot_auto_key = ''
                    
                                        
                    res = batch_labor.apply_async(
                        queue='labor', 
                        priority=1, 
                        kwargs={
                            'quapi_id':quapi_id,
                            'session_id':session_id,
                            'sysur_auto_key':sysur_auto_key,
                            'wot_auto_key':wot_auto_key,
                            'batch_no':batch_no,
                            'active_batch':active_batch,
                            'start_batch':start_batch,
                                                
                            'start_tasks':start_tasks,
                                                       
                                                   
                            'user_name':user_name,
                        })
                        
                    #res = batch_labor.delay(quapi_id,session_id,sysur_auto_key,wot_auto_key,batch_no=batch_no,active_batch=active_batch,start_batch=start_batch,start_tasks=start_tasks,user_name=user_name)
                    error,batch_no,description,msg = res.get() 
                    if not error and add_task == '1':
                        val_dict['msg'] = 'Task %s added to batch.'%wot_auto_key[:-1]
                    val_dict['error'] = error                
                    val_dict['batch_no'] = batch_no
                    val_dict['description'] = description or batch_no
                    val_dict['wo_number'] = ''
                    val_dict['start_batch'] = 'T'
                    val_dict['button_text'] = 'Start Batch'
                    val_dict['msg'] = msg  
                    labor_rows = TaskLabor.objects.filter(session_id=session_id)
                    val_dict['total_labor_rows'] = len(labor_rows)
                    
            #elif (create_mode or recall_mode)\                    
            elif ((recall_mode or create_mode) and (yes_complete == 'T' or not_complete == 'T'))\
                or (yes_complete != 'F' and not_complete != 'F' and (create_mode or recall_mode)\
                and (conf_batch_button == 'Stop Batch' or button_txt == 'Stop Batch')\
                and show_conf == 'got_data' and (conf_batch_no or batch_no or description)):
                    
                from portal.tasks import stop_labor
                res = stop_labor.apply_async(
                    queue='labor', 
                    priority=1,
                    kwargs={
                        'quapi_id':quapi_id,
                        'session_id':session_id,
                        'sysur_auto_key':sysur_auto_key,
                        'batch_no': conf_batch_no or batch_no or description,
                        'wot_auto_key': wot_auto_key,
                        'yes_complete': yes_complete,
                        'user_name': user_name,
                        }
                    )                         
                #res = stop_labor.delay(quapi_id,session_id,sysur_auto_key,conf_batch_no,wot_auto_key,yes_complete,user_name=user_name)
                error,msg = res.get() 
                val_dict['error'] = error          
                val_dict['start_batch'] = 'F'
                val_dict['description'] = conf_batch_no or batch_no
                labors = TaskLabor.objects.filter(session_id = session_id)
                total_labor_rows = len(labors)
                val_dict['total_labor_rows'] = total_labor_rows 
                val_dict['stop_batch'] = 'T'
                val_dict['wo_number'] = ''
                val_dict['batch_no'] = ''
                val_dict['button_txt'] = (recall_mode and 'Recall Batch') or (create_mode and 'Create')              
                #'Batch ' + batch_no + ' stopped.'                  
                val_dict['msg'] = msg + conf_msg
                            
#           elif (not create_mode and not recall_mode or (recall_mode and not wot_auto_key))\
#                and (batch_no or active_batch or wot_auto_key or wot_key or show_conf == 'got_data'):
            elif not (create_mode or recall_mode) and (wot_auto_key or wot_key or show_conf == 'got_data'):
                
                if (single_wot_key and single_wot_key[-1] in ['s','S','c','C'])\
                or (wot_auto_key and wot_auto_key[-1] in ['s','S','c','C']):             
                    if not open_wtls and (button_txt == 'Stop Task' or show_conf == 'got_data'):
                        val_dict['button_txt'] = 'Start Task'
                        val_dict['show_task'] = 'T'
                        wot_auto_key = wot_auto_key[:-1] + 'c'
                        if single_wot_key:
                            wot_auto_key = ''
                            wot_key = ''
                            single_wot_key = single_wot_key[:-1] + 'c'
                            
                    if not (yes_complete=='F' and not_complete=='F') or\
                    (wot_auto_key and wot_auto_key[-1] in ['s','S','c','C']):                            
                        
                        from portal.tasks import add_wo_labor
                        res = add_wo_labor.apply_async(
                            queue='labor',
                            priority=1,
                            kwargs={
                                'quapi_id':quapi_id,
                                'session_id':session_id,
                                'sysur_auto_key':sysur_auto_key,
                                'wot_auto_key':wot_auto_key or wot_key or single_wot_key,
                                'yes_complete':yes_complete,
                                'user_name':user_name,
                                }
                            )
                                                       
                        #res = add_wo_labor.delay(quapi_id,session_id,sysur_auto_key,wot_auto_key=wot_auto_key or wot_key or single_wot_key,yes_complete=yes_complete,user_name=user_name)
                        error,msg = res.get()
                        
                    val_dict.update({
                        'error': error,
                        'msg': msg,
                        'form': form,
                        'show_task': 'T',
                    })
                    
                    if error:
                        return render(request, 'mrolive/wolabor.html', val_dict)
                        
                    if not error and button_txt in ('Start Task','Start'):
                        val_dict['button_txt'] = 'Stop Task'                                                                                                                                                                                                                      
                                    
                else:
                    start_tasks = False                    
                    #if recall_mode:
                    #    wot_auto_key = batch_no or active_batch or ''
                        
                    if 'woos_list[]' in req_post:
                        wot_id_list = req_post.getlist('woos_list[]')                    
                    if button_txt != 'OK' and not (single_wot_key or wot_id_list) and show_conf != 'got_data':
                        error = 'Must select one or more task.'
                        val_dict['button_txt'] = 'Start Task'
                        val_dict['show_task'] = 'T'
                        
                    elif (not wot_id_list or len(wot_id_list) == 1) and single_wot:
                                                               
                        if wot_id_list and len(wot_id_list) == 1:
                            single_wot = wot_id_list[0] + 's'
                        from portal.tasks import add_wo_labor
                        res = add_wo_labor.apply_async(
                            queue='labor', 
                            priority=1, 
                            kwargs={
                                'quapi_id':quapi_id,
                                'session_id':session_id,
                                'sysur_auto_key':sysur_auto_key,
                                'wot_auto_key':single_wot,
                                'user_name':user_name,
                                }
                            )                      
                        #res = add_wo_labor.delay(quapi_id,session_id,sysur_auto_key,wot_auto_key=single_wot,user_name=user_name)
                        error,msg = res.get()
                                 
                                                   
                        val_dict['button_txt'] = 'Start Task'
                        #val_dict['total_labor_rows'] = 10                     
                        if not error and button_txt in ['Start','Start Task']:
                            val_dict['button_txt'] = 'Stop Task'
                        val_dict['single_wot'] = single_wot
                        
                    if wot_id_list and not len(wot_id_list) == 1 and button_txt == 'Start':
                        create_batch = True
                        description = wot_key
                        from portal.tasks import batch_labor 
                        res = batch_labor.apply_async(
                            queue='labor', 
                            priority=1, 
                            kwargs={
                                'quapi_id':quapi_id,
                                'session_id':session_id,
                                'sysur_auto_key':sysur_auto_key,
                                'wot_auto_key':'',
                                'batch_no':batch_no,
                                'active_batch':'',
                                'start_batch':False,
                                'create_batch':True,
                                'start_tasks':False,
                                'description': description,
                                'wot_ids': wot_id_list,
                                'user_name':user_name,
                            })                        
                        #res = batch_labor.delay(quapi_id,session_id,sysur_auto_key,\
                        #     '',batch_no=batch_no,\
                        #     active_batch='',\
                        #     start_batch=True,create_batch=True,\
                        #     start_tasks=True,description=description,wot_ids=wot_id_list,
                        #     user_name=user_name)                
                        error,batch_num,description,msg = res.get()
                        val_dict.update({                    
                            'batch_no': batch_no or batch_num,
                            'description': batch_num,
                            'wo_number': '',
                            'wot_auto_key': '',
                            'msg': msg,
                            'active_mode': 'batch-recall',
                            'sel_mode': 'batch-recall',
                            'conf_sel_mode': 'batch-recall',
                            'error': error, 
                            'start_batch': 'T',            
                            'button_txt': 'Start Batch',
                        })
                        
                    #Must be a WO_NUMBER so we will get all tasks for this WO_NUMBER and display them
                    elif labor_submit != 'OK' and val_dict['button_txt'] != 'Stop Task' and not recall_mode:
                        val_dict['button_txt'] = 'Start'
                        
                    if not wot_id_list and wot_auto_key and not recall_mode:
                        from portal.tasks import get_signoff_tasks
                        res = get_signoff_tasks.apply_async(
                        queue='labor', 
                        priority=1, 
                        kwargs={
                            'quapi_id':quapi_id,
                            'session_id':session_id,
                            'wo_number':wot_auto_key,
                        })
                        #res = get_signoff_tasks.delay(quapi_id,session_id,wo_number=wot_auto_key)
                        error,msg = res.get()          
                        wtms = WOTask.objects.filter(session_id=session_id).order_by('wot_sequence','wot_auto_key').exclude(status_type='Closed') or []
                        #wtms = WOTask.objects.filter(session_id=session_id)
                        total_rows = len(wtms)
                        
                        if total_rows == 0 and button_txt != 'Stop Batch':
                            val_dict['error'] = 'No tasks found.'
                            val_dict['button_txt'] = 'Start Task'
                            val_dict['form'] = form
                            return render(request, 'mrolive/wolabor.html', val_dict)  
                        
                        if total_rows == 1:
                            wot_auto_key = wtms and wtms[0] and wtms[0].wot_auto_key or 0
                            if val_dict['button_txt'] != 'Stop Task' and 'c_button_txt' not in req_post:
                                val_dict['button_txt'] = 'Start Task'
                                val_dict['show_task'] = 'F'
                            elif labor_submit == 'Stop Task' or ('c_button_txt' in req_post and req_post['c_button_txt'] == 'Stop Task'):
                                val_dict['show_task'] = 'T'
                                val_dict['button_txt'] = 'Start Task'   
                        
                        val_dict.update({
                            'wtms': wtms,
                            'single_wot': str(wot_auto_key) + 's',
                            'total_rows': total_rows,
                            'sel_rows': 0,
                        })
                        
                res = get_wtls.apply_async(
                queue='labor', 
                priority=1, 
                kwargs={
                    'quapi_id':quapi_id,
                    'sysur_auto_key':sysur_auto_key
                })
                
                error,msg = res.get()
                open_wtls = TaskLabor.objects.filter(sysur_auto_key=sysur_auto_key).order_by('-id') 
                if not open_wtls:                
                    val_dict['msg'] = msg
                else:
                    val_dict['msg'] = 'WO #: ' + open_wtls[0].wo_number
                    val_dict['msg'] += ' | Seq: ' + str(open_wtls[0].sequence)
                    val_dict['msg'] += ' | ' + open_wtls[0].task_desc
                    #start_time = datetime.strptime(open_wtls[0].start_time,'%m-%d-%Y %H:%M:%S')
                    start_time = open_wtls[0].start_time 
                    start_time = start_time and start_time - timedelta(hours=4)
                    start_time = start_time.strftime('%m-%d-%Y %H:%M:%S')
                    val_dict['msg'] += ' started ' + start_time + '.'
                    val_dict['button_txt'] = 'Stop Task'
                    val_dict['wot_key'] = str(open_wtls[0].wot_auto_key) + 'c'
                    val_dict['wot_auto_key'] = str(open_wtls[0].wot_auto_key) + 'c'   
                val_dict['wo_number'] = ''
                
            elif active_batch and labor_submit == 'Recall Batch':
                                                                           
                from portal.tasks import batch_labor
                if task_submit == 'Add Task' and not wot_auto_key.isnumeric():
                    val_dict['error'] = "Task must be an integer followed by \'s\'"
                    return render(request, 'mrolive/wolabor.html', val_dict)  
                                                                                                                     
                if task_submit != 'Add Task' or add_task != '1':
                    wot_auto_key  = ''
                res = batch_labor.apply_async(
                    queue='labor', 
                    priority=1, 
                    kwargs={
                        'quapi_id':quapi_id,
                        'session_id':session_id,
                        'sysur_auto_key':sysur_auto_key,
                        'wot_auto_key':wot_auto_key,
                        'active_batch':active_batch,
                        'create_batch': True,
                        'start_batch':False,
                        'user_name':user_name,
                    })

                #res = batch_labor.delay(quapi_id,session_id,sysur_auto_key,wot_auto_key,active_batch=active_batch,start_batch=False,user_name=user_name)
                error,batch_no,description,msg = res.get()              
                val_dict['button_txt'] = 'Start Batch'             
                        
                if error[:22] == 'No open tasks in batch':
                    val_dict['button_txt'] = 'Recall Batch'
                    msg = 'no tasks found in the batch to recall.'
                    
                val_dict['msg'] = msg
                val_dict['error'] = error                 
                val_dict['batch_no'] = batch_no
                val_dict['description'] = description or batch_no
                labor_rows = TaskLabor.objects.filter(session_id=session_id)
                val_dict['total_labor_rows'] = len(labor_rows)
                val_dict['wo_number'] = ''
                val_dict['wot_auto_key'] = ''            
                val_dict['start_batch'] = 'T' 
                
            elif create_mode:
                
                if wot_auto_key and not wot_auto_key[-1] in ['s','S']:
                    
                    if not wot_id_list and labor_submit != 'Start Batch' and task_submit != 'Add Task':
                        error,msg,total_rows=get_tasks(quapi_id,session_id,wo_number=wot_auto_key)
                        wtms = WOTask.objects.filter(session_id=session_id)

                        val_dict.update({                    
                            'batch_no': batch_no and batch_no.upper() or batch_num and batch_num.upper(),
                            'description': batch_num and batch_num.upper(),
                            'wo_number': '',
                            'total_rows': total_rows,
                            'wot_auto_key': '',
                            'msg': msg,
                            'active_mode': 'batch-recall',
                            'sel_mode': 'batch-recall',
                            'conf_sel_mode': 'batch-recall',
                            'error': error, 
                            'start_batch': 'T',            
                            'button_txt': 'Start Batch',
                        })
                        return render(request, 'mrolive/wolabor.html', val_dict)
                
                start_batch = False
                start_tasks = False           
                create_batch = False             
                             
                #if task_submit == 'Add Task' and not wot_auto_key[:-1].isnumeric():
                    #val_dict['error'] = "Task must be an integer followed by \'s\'"
                    #return render(request, 'mrolive/wolabor.html', val_dict) 
                    
                if labor_submit == 'Start Batch' and task_submit != 'Add Task':
                    start_batch = True
                    start_tasks = True
                    button_txt = 'Stop Batch'  
                    
                if button_txt == 'Create':
                    button_txt = 'Start Batch'
                    create_batch = True
                    val_dict['wo_number'] = ''
                    val_dict['msg'] = msg
                    val_dict['error'] = error 
                    val_dict['start_batch'] = 'T'            
                    val_dict['button_txt'] = button_txt
                    val_dict['wot_auto_key'] = ''
                    
                    if len(active_batch) > 5:
                        val_dict['button_txt'] = 'Create'
                        val_dict['error'] = 'Batch name must be no more than 5 characters.'
                        return render(request, 'mrolive/wolabor.html', val_dict)
                        
                    if '_' in active_batch:
                        val_dict['button_txt'] = 'Create'
                        val_dict['error'] = 'Batch name cannot contain _.'
                        return render(request, 'mrolive/wolabor.html', val_dict)
 
                from portal.tasks import batch_labor
                res = batch_labor.apply_async(
                    queue='labor',
                    priority=1, 
                    kwargs={
                        'quapi_id':quapi_id,
                        'session_id':session_id,
                        'sysur_auto_key':sysur_auto_key,
                        'wot_auto_key':wot_auto_key,
                        'batch_no':batch_no or conf_batch_no,
                        'active_batch':active_batch,
                        'start_batch':start_batch,
                        'create_batch': create_batch,
                        'start_tasks':start_tasks,
                        'description': description,
                        'user_name':user_name,
                    })                
                #res = batch_labor.delay(quapi_id,session_id,sysur_auto_key,\
                #     wot_auto_key,batch_no=batch_no or conf_batch_no,\
                #     active_batch=active_batch or batch_no,\
                #     start_batch=start_batch,create_batch=create_batch,\
                #     start_tasks=start_tasks,description=description,
                #     user_name=user_name)                
                error,batch_num,description,msg = res.get() 
                labor_rows = TaskLabor.objects.filter(session_id=session_id)                
                val_dict['batch_no'] = batch_num or batch_no
                val_dict['description'] = description or batch_num or batch_no
                val_dict['wo_number'] = ''
                val_dict['msg'] = msg
                val_dict['error'] = error 
                val_dict['start_batch'] = 'T'                               
                val_dict['button_txt'] = button_txt
                val_dict['total_labor_rows'] = len(labor_rows)
                val_dict['wot_auto_key'] = ''
                

            if labor_submit == 'Add' and single_wot and wot_id_list or labor_submit == 'Start Batch' and (task_submit not in ['Add Task'] or wot_id_list) or val_dict['single_wot'] == '' and val_dict['show_conf'] != 'T' and val_dict['button_txt'] in ['Stop Task','Start Task'] or conf_batch_button in ['Stop Batch']:
                rows = TaskLabor.objects.filter(session_id=session_id)              
                val_dict['total_labor_rows'] = len(rows)  
                val_dict['total_rows'] = 0                
                if search_user in ['','F'] and 'kiosk' in val_dict and val_dict['kiosk'] == 'T':
                    val_dict['button_txt'] = 'OK'
                    val_dict['user_name'] = ''
                    val_dict['batch_no'] = ''
                    val_dict['single_wot_key'] = ''
                    val_dict['wot_auto_key'] = '' 
    if error:                
        val_dict['error'] = error

    if 'button_txt' in val_dict\
        and not val_dict['button_txt']\
        and not (recall_mode or create_mode):
        val_dict['button_txt'] = 'Start Task'

    val_dict['form'] = form
    val_dict['wo_number'] = ''
    return render(request, 'mrolive/wolabor.html', val_dict)   
  
def get_tasks(quapi_id,session_id,wot_auto_key='',wo_number='',woo_auto_key='',batch_id=''):  
    from portal.tasks import get_signoff_tasks
    res = get_signoff_tasks.apply_async(
    queue='labor', 
    priority=1, 
    kwargs={
        'quapi_id':quapi_id,
        'session_id':session_id,
        'wot_auto_key':wot_auto_key,
        'wo_number': wo_number,
        'woo_auto_key': woo_auto_key,
        'batch_id': batch_id,
    })
    error,msg = res.get()          
    wtms = WOTask.objects.filter(session_id=session_id).order_by('wot_sequence','wot_auto_key').exclude(status_type='Closed') or []
    total_rows = len(wtms)  
    return error,msg,total_rows
    
def full_clean(obj_to_clean):
    error = ''
    try:
        obj_to_clean.full_clean()
    except ValidationError as e:
        # Do something based on the errors contained in e.message_dict.
        # Display them to a user, or handle them programmatically.
        pass
        error = e
    return error

def logout_view(request):
    logout(request)
    # Redirect to a success page.
    val_dict={}
    return redirect('https://portal.mrolive.com/')
    
def account_route(request, logoff='0'):
    #get current user's SQLite db id
    app_sel,quapi_sel,app_selected,quapi_selected,red,val_dict='','','','','',{}
    req_post = None
    quapis = QueryApi.objects.all()
    apps = MLApps.objects.all()
    from datetime import timezone
    from django.contrib.sessions.models import Session
    from django.contrib.auth import logout
    right_now = datetime.now(timezone.utc)
    #all_sessions = Session.objects.filter(expire_date__gte=right_now)
    user = request.user
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    if not user_apps:
        val_dict['error'] = 'No apps assigned. Use user management to assign.'
        return render(request, 'registration/home.html', val_dict)         
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    username = user and user.username or 'No Username'
    user_id = user and user.is_authenticated and user.id or None
    if not user.id or not user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    #check group and get quantum_cmp_key
    #group = user_groups and user_groups[0] or None
    #using the user profile, query Quantum to confirm the sysur_auto_key is correct
    #we could either use the get_users() task or run a query just for our user 
    #look up latest users and then compare the username with the user_id field.    
    from portal.tasks import get_users_nsync_beta
    res = get_users_nsync_beta.delay(1,user_id,is_dashboard=0,app='account-route')
    user_error,app = res.get()
    # | Q(employee_code=username)
    q_user = QuantumUser.objects.filter(user_id=username)
    if not q_user:
        q_user = QuantumUser.objects.filter(employee_code=username)
    if not q_user:
        q_user = QuantumUser.objects.filter(user_name=username)
    q_user = q_user and q_user[0] or None    
    if q_user:
        #check that the sysur_auto_key is correct
        #if not, then we have to correct it on the user profile record
        user_profile = UserProfile.objects.filter(user=user)
        user_profile = user_profile and user_profile[0] or None 
        sysur_auto_key = user_profile.sysur_auto_key
        if sysur_auto_key != q_user.user_auto_key:
            #change the sysur_auto_key
            user_profile.sysur_auto_key = q_user.user_auto_key
    else:
        #prompt the user to synch up the username with that of a user_id from a Quantum user.
        val_dict['error'] = 'User name does not match any USER IDs in Quantum.'
    group = None
    user_groups = user.groups.all()
    for user_group in user_groups:
        group = UserGroupProfile.objects.filter(group=user_group.id)
        group = group and group[0] or None
        if group:
            break
    if not group:
        val_dict['error'] = 'User must belong to the company group.'
        return render(request, 'registration/home.html', val_dict)	    
    quantum_cmp_key = group.quantum_cmp_key
    conn_string = group.conn_string
    if quantum_cmp_key and conn_string:
        from portal.tasks import check_exp_date
        res = check_exp_date.delay(conn_string,quantum_cmp_key)
        error = res.get()
        if error:
            val_dict['error'] = error
            return render(request, 'registration/home.html', val_dict)
    else:
        val_dict['error'] = 'Must have company ID on group profile. Contact MRO Live Account Manager.'
        return render(request, 'registration/home.html', val_dict) 
        
    if request.method == 'POST':
        #based on the user's selection of quapi and app, we will route to the appropriate place
        req_post = request.POST
        app_sel = None
        app_selected = 'app_selector' in req_post and req_post['app_selector'] or None
        """mgmt_app_selected = 'mgmt_app_selector' in req_post and req_post['mgmt_app_selector'] or None
        setup_app_selected = 'setup_app_selector' in req_post and req_post['setup_app_selector'] or None
        dash_app_selected = 'dash_app_selector' in req_post and req_post['dash_app_selector'] or None
        if op_app_selected:
            app_sel = op_app_selected and apps.filter(id=op_app_selected) or '' 
            app_sel = app_sel and app_sel[0] or None
        elif mgmt_app_selected:
            app_sel = mgmt_app_selected and apps.filter(id=mgmt_app_selected) or '' 
            app_sel = app_sel and app_sel[0] or None
        elif dash_app_selected:
            app_sel = dash_app_selected and apps.filter(id=dash_app_selected) or '' 
            app_sel = app_sel and app_sel[0] or None"""
        if app_selected:
            app_sel = app_selected and apps.filter(id=app_selected) or '' 
            app_sel = app_sel and app_sel[0] or None
        quapi_selected = 'quapi_selector' in req_post and req_post['quapi_selector'] or 1
        #quapi_sel = quapi_selected and quapis.filter(id=quapi_selected) or None   
        #quapi_sel = quapi_sel and quapi_sel[0] or None
    quapi_set = user_id and UserQuapiRel.objects.filter(user=user) or []
    if not quapi_set or not user:
        render(request, 'registration/home.html', val_dict)   
    #check user_id in the rel table that associates the Django app user
    #with the quantum user
    session_id = request.session and request.session.session_key or None 
    #quapi_id = user_id and quapis and quapis.filter(id = quapi_id) or None
    #quapi_id = quapi_id and quapi_id[0] or None
    val_dict.update({
        'app_set': user_apps,
        'app_sel': app_sel and app_sel.id or None,
        'user_id': user_id,
        'quapi_set': quapi_set,
        'quapi_sel': quapi_selected and int(quapi_selected) or 0,        
        'session_id': session_id,
        })
    if app_sel and user_id and req_post:
        app_view = app_sel and '/portal/' + str(app_sel.code) + '/' + str(quapi_selected) or None
        #url = app_view and quapi_id and request.build_absolute_uri(reverse(app_view, args=(quapi_id, ))) or None
        #app_view = app_sel and '/portal/' + str(app_sel.code) or None
        if app_view:
            return redirect(app_view)
    return render(request, 'registration/home.html', val_dict) 

def app_mgmt(request):
    #get current user's SQLite db id
    app_sel,app_selected,user_sel,user_selected,error,res,dj_user_name='','','','','','',''
    req_post,dupe_app,dupe_quapi = None,False,False
    apps = MLApps.objects.all()
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    username = user and user.username or 'No Username'
    user = user and user.is_authenticated and user.id or None
    from django.contrib.auth.models import User
    users = User.objects.all()
    quapis = QueryApi.objects.all()
    if request.method == 'POST':
        #based on the user's selection of quapi and app, we will route to the appropriate place
        req_post = request.POST
        app_selected = 'app_selected' in req_post and req_post['app_selected'] or ''
        app_sel = app_selected and apps.filter(id=app_selected) or '' 
        app_sel = app_sel and app_sel[0] or None
        user_selected = 'user_selected' in req_post and req_post['user_selected'] or ''
        user_sel = user_selected and users.filter(id=user_selected) or None   
        user_sel = user_sel and user_sel[0] or 0        
        dj_user_name = user_sel.username
        quapi_selected = 'quapi_selector' in req_post and req_post['quapi_selector'] or ''
        quapi_sel = quapi_selected and quapis.filter(id=quapi_selected) or None   
        quapi_sel = quapi_sel and quapi_sel[0] or None         
    #check user_id in the rel table that associates the Django app user
    #with the quantum user
    session_id = request.session and request.session.session_key or None
    user_perms = user_selected and UserAppPerms.objects.filter(user=user) or []
    val_dict = {
        'app_set': apps,
        'app_sel': app_selected and int(app_selected) or None,
        'user_id': user,
        'user_sel': user_selected and int(user_selected) or None, 
        'quapi_sel': quapi_selected and int(quapi_selected) or None,        
        'user_perms': user_perms,  
        'users': users,        
        'session_id': session_id,
        'error': error,
        'quapi_set': quapis,
        }
    if user_selected and app_selected and app_sel:
        search_dupes = UserAppPerms.objects.filter(user=user,ml_apps_id=app_sel)
        if search_dupes:
            val_dict['error'] = 'This user already has global access to that app'
            dupe_app = True
            #return render(request, 'registration/app_management.html', val_dict)            
        if not dupe_app:
            try:
                res = UserAppPerms(dj_username=dj_user_name,user=user,ml_apps_id=app_sel)   
                res.save()   
            except Exception as exc:
                logger.exception('Creating permissions caused this exception: %r', exc) 
    user_perms = user_selected and UserAppPerms.objects.filter(user=user) or []
    if user_selected and quapi_selected and quapi_sel:
        search_dupes = QueryApi.objects.filter(dj_user_id=user_selected,id=quapi_selected)
        if search_dupes:
            val_dict['error'] = 'This user is already associated with that schema.'
            dupe_quapi = True
            #return render(request, 'registration/app_management.html', val_dict) 
        if not dupe_quapi:            
            try:
                res = UserQuapiRel(dj_username=dj_user_name,dj_user_id=user_selected,quapi_id=quapi_sel)   
                res.save()   
            except Exception as exc:
                logger.exception('Creating API access for the user caused this exception: %r', exc) 
    user_schema = user_selected and UserQuapiRel.objects.filter(dj_user_id=user_selected) or []
    val_dict['user_perms'] = user_perms
    val_dict['user_schema'] = user_schema
    return render(request, 'registration/app_management.html', val_dict)    
#**************************************************End User Accounts and Sessions************************************
   
#****************************************************Begin Dashboard Apps*********************************************                    
def so_management(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec = 'user not set',None
    wo_number = ''
    val_dict,form = {},{}
    all_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id='') 
    updated_sales = all_woos 
    woos_updated = False    
    error,msg,loc_msg,stat_msg = '','','',''
    woos_updated = False
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    user_id = user and user.is_authenticated and user.id or None
    if not user_id:
        return redirect('/login/')   
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps                                           
    val_dict['quapi_id'] = quapi_id        
    from portal.tasks import get_statuses_nsync_beta
    if request.method == 'GET':
        res = get_statuses_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=1,app='wo-management')
        stat_error,app = res.get() 
        form = WODashboardForm()        
    val_dict['status_vals'] = dj_user_id and stat_sel.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id,is_dashboard=1).distinct() or [] 
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        exact_match = 'exact_match' in req_post and req_post['exact_match'] or ''
        if exact_match:
            exact_match = 'checked'
        customer = 'customer' in req_post and req_post['customer'] or ''
        so_number = 'so_number' in req_post and req_post['so_number'] or ''
        new_status = 'new_status' in req_post and req_post['new_status'] or ''
        due_date = 'due_date' in req_post and req_post['due_date'] or ''
        new_due_date = 'get_due_date' in req_post and req_post['get_due_date'] or ''      
        salesperson = 'spn_code' in req_post and req_post['spn_code'] or ''                                                                                    
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0 
        options_col,page_size = get_options(req_post,session_id)            
        val_dict.update({
            'all_woos': updated_sales, 
            'msg': msg,
            'new_status': new_status and int(new_status), 
            'customer': customer,
            'due_date': due_date,
            'new_due_date': new_due_date,
            'user_id': user_id or '',
            'salesperson': salesperson,
            #'new_salesperson': new_salesperson,             
            'so_number': so_number,
            'session_id': session_id,
            'sel_rows': sel_rows,
            'total_rows': total_rows,
            #'warehouse': warehouse,
            #'location': location,
            'exact_match': exact_match,
            'options_col': options_col,
            'page_size': page_size,               
        })
        if not (so_number or salesperson or due_date or customer or new_status):
            val_dict['error'] = 'Must make an entry in one of the search fields to filter.' 
            return render(request, 'mrolive/so_management.html', val_dict)
        from portal.tasks import so_dashboard
        res = so_dashboard.delay(quapi_id,session_id,so_number,salesperson,due_date,exact_match,customer=customer,status=new_status) 
        error,msg = res.get()  
        if error == 'no errors':
            error = ''        
        updated_sales = Sale.objects.filter(session_id=session_id)
        val_dict['all_woos'] = updated_sales
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['total_rows'] = str(len(updated_sales))
    val_dict['form'] = form
    user_profile = UserProfile.objects.filter(user=user)
    user_profile = user_profile and user_profile[0] or None
    val_dict['num_records'] = user_profile.num_records or 10
    return render(request, 'mrolive/so_management.html', val_dict) 
    
def so_dashboard(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec = 'user not set',None
    wo_number = ''
    val_dict,form = {},{}
    all_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id='') 
    updated_sales = all_woos 
    woos_updated = False    
    error,msg,loc_msg,stat_msg = '','','',''
    woos_updated = False
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    user_id = user and user.is_authenticated and user.id or None
    if not user_id:
        return redirect('/login/')   
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps                               
    val_dict['quapi_id'] = quapi_id        
    from portal.tasks import get_statuses_nsync_beta
    if request.method == 'GET':                                  
        form = WODashboardForm()        
    val_dict['status_vals'] = dj_user_id and stat_sel.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id,is_dashboard=1).distinct() or [] 
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        val_dict['form'] = form
        exact_match = 'exact_match' in req_post and req_post['exact_match'] or ''
        if exact_match:
            exact_match = 'checked'
        customer = req_post.get('customer','')
        customer = req_post.get('customer','')
        so_number = 'so_number' in req_post and req_post['so_number'] or ''
        new_status = 'new_status' in req_post and req_post['new_status'] or ''
        due_date = 'due_date' in req_post and req_post['due_date'] or ''
        location = 'location' in req_post and req_post['location'] or ''      
        salesperson = 'spn_code' in req_post and req_post['spn_code'] or ''                                                                                        
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0 
        options_col,page_size = get_options(req_post,session_id)            
        val_dict.update({
            'all_woos': updated_sales, 
            'msg': msg,
            'new_status': new_status and int(new_status), 
            'customer': customer,
            'due_date': due_date,
                       
            'user_id': user_id or '',
            'salesperson': salesperson,             
                                 
            'so_number': so_number,
            'session_id': session_id,
            'sel_rows': sel_rows,
            'total_rows': total_rows,
                                   
            'location': location,
            'exact_match': exact_match,
            'options_col': options_col,
            'page_size': page_size,               
        })
        #if not (so_number or salesperson or due_date or location):
        #    val_dict['error'] = 'Must make an entry in one of the search fields to filter.' 
        #    return render(request, 'mrolive/so_dashboard.html', val_dict) 
        from portal.tasks import so_dashboard
        res = so_dashboard.delay(quapi_id,session_id,\
        so_number,salesperson,due_date,location,exact_match=False) 
        error,msg = res.get()  
        if error == 'no errors':
            error = ''        
        updated_sales = Sale.objects.filter(session_id=session_id)
        val_dict['all_woos'] = updated_sales
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['total_rows'] = str(len(updated_sales))
    val_dict['form'] = form
    user_profile = UserProfile.objects.filter(user=user)
    user_profile = user_profile and user_profile[0] or None
    val_dict['num_records'] = user_profile.num_records or 10
    return render(request, 'mrolive/so_dashboard.html', val_dict) 
    
def stock_picking(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec = 'user not set',None
    wo_number = ''
    val_dict,form = {},{}
    all_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id='') 
    updated_woos = all_woos 
    woos_updated = False    
    error,msg,loc_msg,stat_msg = '','','',''
    woos_updated = False
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    user_id = user and user.is_authenticated and user.id or None
    if not user_id:
        return redirect('/login/')   
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps                               
    val_dict['quapi_id'] = quapi_id        
    from portal.tasks import get_statuses_nsync_beta
    if request.method == 'GET':                                  
        form = WODashboardForm()        
    val_dict['status_vals'] = dj_user_id and stat_sel.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id,is_dashboard=1).distinct() or [] 
    if request.method == 'POST':
        req_post = request.POST
        
        form = WODashboardForm(req_post)
        exact_match = 'exact_match' in req_post and req_post['exact_match'] or ''
        if exact_match:
            exact_match = 'checked'
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''                                                                                     
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0 
        options_col,page_size = get_options(req_post,session_id) 
        wo_reserve = 'wo_reserve' in req_post and req_post['wo_reserve'] or ''        
        val_dict.update({
            'all_woos': updated_woos, 
            'msg': msg,
            'user_id': user_id or '',           
            'wo_number': '',
            'session_id': session_id,
            'sel_rows': sel_rows,
            'total_rows': total_rows,
            'exact_match': exact_match,
            'options_col': options_col,
            'page_size': page_size,               
        })
        wob_id_list = []
        from portal.tasks import stock_picking
        if 'wobs_list[]' in req_post:
            wob_id_list = req_post.getlist('wobs_list[]')
            #print report with all of the stock moves and reserve them all to the bom for the part
            res = stock_picking.delay(quapi_id,sysur_auto_key,session_id,wo_number,exact_match=True,wob_id_list=wob_id_list) 
            error,msg = res.get() 
        else:
            if not wo_number:
                val_dict['error'] = 'Enter WO#.' 
                return render(request, 'mrolive/stock_picking.html', val_dict) 
            res = stock_picking.delay(quapi_id,sysur_auto_key,session_id,wo_number,exact_match=True) 
            error,msg = res.get()               
        if error == 'no errors':
            error = ''        
        updated_woos = WOStatus.objects.filter(session_id=session_id,is_detail=False)
        val_dict['all_woos'] = updated_woos
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['total_rows'] = str(len(updated_woos))
    val_dict['form'] = form
    user_profile = UserProfile.objects.filter(user=user)
    user_profile = user_profile and user_profile[0] or None
    val_dict['num_records'] = user_profile.num_records or 10
    return render(request, 'mrolive/stock_picking.html', val_dict) 
    
def wo_task_detail(request,quapi_id=None,parent_auto_key=None,woo_auto_key=None,ult_parent_woo=None,session_id=None,wo_number=None):
    #this method takes a GET request along with a wo_number to generate all children woos and their tasks under the woo passed in.
    val_dict={}
    total_rows = 0
    if not (parent_auto_key or woo_auto_key):
        val_dict['error'] = 'There is a problem with your workorder. Try again.'
    else:
        new_status,location,filter_status = '','',''
        user_id,user_rec = 'user not set',None
        val_dict,form = {},{}   
        error,msg,loc_msg,stat_msg = '','','',''
        quser = QuantumUser
        wos_obj = WOStatus
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    from portal.tasks import create_wo_tasks
    res = create_wo_tasks.delay(quapi_id,parent_auto_key,session_id,woo_id=woo_auto_key)
    error = res.get()
    detail_rows = not error and WOTask.objects.filter(ult_parent_woo=parent_auto_key,woo_auto_key=woo_auto_key,session_id=session_id) or []
    total_rows = len(detail_rows)        
    if not error and total_rows > 0:
        val_dict['total_rows'] = total_rows
        val_dict['all_woos'] = True
    elif total_rows <= 0:
        error += 'There are no tasks or sub-tasks for this WO'    
    val_dict['error'] = error
    val_dict['quapi_id'] = quapi_id
    val_dict['session_id'] = session_id
    val_dict['parent_auto_key'] = int(parent_auto_key)
    val_dict['woo_auto_key'] = int(woo_auto_key)
    val_dict['wo_number'] = wo_number
    user_profile = UserProfile.objects.filter(user=user)
    user_profile = user_profile and user_profile[0] or None
    val_dict['num_records'] = user_profile.num_records or 10
    return render(request, 'mrolive/wotaskdetail.html', val_dict)

def toll_gate(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec = 'user not set',None
    wo_number = ''
    val_dict,form = {},{}
    all_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id='') 
    updated_woos = all_woos 
    woos_updated = False    
    error,msg,loc_msg,stat_msg = '','','',''
    woos_updated = False
    from polls.models import QuantumUser as quser
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')   
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id  
   
    if request.method == 'GET': 
        form = WODashboardForm()        
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post) 
        exact_match = 'exact_match' in req_post and req_post['exact_match'] and 'checked' or ''
        customer = 'customer' in req_post and req_post['customer'] or ''
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        due_date = 'get_due_date' in req_post and req_post['get_due_date'] or ''
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0  
        options_col,page_size = get_options(req_post,session_id)            
        val_dict.update({
            'all_woos': updated_woos, 
            'msg': msg,
            'customer': customer,
            'get_due_date': due_date,
            'user_id': user_id or '',
            'wo_number': wo_number,
            'session_id': session_id,
            'total_rows': total_rows,
            'exact_match': exact_match,
            'options_col': options_col,
            'page_size': page_size,  
        })
        #function call to add the records         
        from portal.tasks import add_wo_record
        res = add_wo_record.delay(
            is_toll=True,
            is_dashboard=1,
            quapi_id=quapi_id,
            customer=customer,
            due_date=due_date,
            wo_number=wo_number,
            session_id=session_id,
            exact_match=exact_match
        ) 
        error,msg = res.get()  
        if error == 'no errors':
            error = ''               
        updated_woos = WOStatus.objects.filter(is_detail=False,active=1,is_dashboard=1,session_id=session_id)
        val_dict['all_woos'] = updated_woos              
    val_dict['error'] = error
    val_dict['total_rows'] = str(len(updated_woos))
    val_dict['form'] = form
    val_dict['quapi'] = quapi_id and str(quapi_id) or 1
    user_profile = UserProfile.objects.filter(user=user)
    user_profile = user_profile and user_profile[0] or None
    val_dict['num_records'] = user_profile.num_records or 10
    return render(request, 'mrolive/tollgateanalysis.html', val_dict)  

def toll_gate_detail(request,quapi_id=None,parent_auto_key=None,toll_gate=None,session_id=None,wo_number=None):
    #this method takes a GET request along with a wo_number to generate all children woos under the woo passed in.
    val_dict={}
    total_rows = 0
    if not parent_auto_key:
        val_dict['error'] = 'There is a problem with your workorder. Try again.'
    else:
        new_status,location,filter_status = '','',''
        user_id,user_rec = 'user not set',None
        val_dict,form = {},{}   
        error,msg,loc_msg,stat_msg = '','','',''
        from polls.models import QuantumUser as quser,WOStatus as wos_obj
        user = request.user
        if not user.id:
            val_dict['error'] = 'Access denied.'
            return redirect('/login/')  
        reg_user_id = user and user.is_authenticated and user.id or None
        dj_user_id = user and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
        if not reg_user_id or not dj_user_id:
            val_dict['error'] = 'Access denied.'
            return redirect('/login/')  
        user_apps = user and UserAppPerms.objects.filter(user=user) or None
        op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
        val_dict['op_apps'] = op_apps
        mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
        val_dict['mgmt_apps'] = mgmt_apps
        dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
        val_dict['dash_apps'] = dash_apps
        setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
        val_dict['setup_apps'] = setup_apps
        if toll_gate and toll_gate != '0':
            detail_rows = WOStatus.objects.filter(sub_wo_gate=toll_gate,parent_auto_key=parent_auto_key,is_detail=True,active=1,is_dashboard=1,session_id=session_id)
        else:
            detail_rows = WOStatus.objects.filter(parent_auto_key=parent_auto_key,is_detail=True,active=1,is_dashboard=1,session_id=session_id)
        total_rows = len(detail_rows)        
        if not error and total_rows > 0:
            val_dict['total_rows'] = total_rows
            val_dict['all_woos'] = True
        else:
            error += 'There are no sub-WOs for this WO'               
        val_dict['error'] = error
        val_dict['quapi_id'] = quapi_id
        val_dict['session_id'] = session_id
        val_dict['wo_number'] = wo_number
        val_dict['parent_auto_key'] = parent_auto_key
        val_dict['sub_wo_gate'] = toll_gate
        user_profile = UserProfile.objects.filter(user=user)
        user_profile = user_profile and user_profile[0] or None
        val_dict['num_records'] = user_profile.num_records or 10
    return render(request, 'mrolive/tollgatewodetail.html', val_dict)  

def open_labor(request,quapi_id):
    new_status,filter_status = '',''
    user_id,user_rec = 'user not set',None
    wo_number = ''
    val_dict,form,all_wotls = {},{},[]
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':          
        session_id = request.session.session_key          
        val_dict.update({
            'msg': msg,
            'session_id': session_id,      
        })
        existing_wotls = TaskLabor.objects.filter(session_id=session_id)
        if existing_wotls:
            existing_wotls.delete()
        from portal.tasks import open_labor
        res = open_labor.apply_async(
            queue='refresh', 
            priority=1, 
            kwargs={
                'quapi_id':quapi_id,
                'session_id':session_id,
            })
        error,msg = res.get()         
        if error == 'no errors':
            error = '' 
        all_wotls = TaskLabor.objects.filter(session_id=session_id)            
        val_dict['all_wotls'] = all_wotls
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['total_rows'] = str(len(all_wotls))
    val_dict['form'] = form
    return render(request, 'mrolive/open_labor.html', val_dict)
    
def dash_refresh_display(request,quapi_id,location):
    new_status,filter_status = '',''
    user_id,user_rec = 'user not set',None
    wo_number = ''
    val_dict,form,all_woos = {},{},[]
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id     
    if request.method == 'GET':
        #req_post = request.POST
        #form = WODashboardForm(req_post)
        if not location:
            val_dict['error'] = 'Must have a location in the query string URI.' 
            return render(request, 'mrolive/wo_dashboard_refresh.html', val_dict)             
        session_id = request.session.session_key  
        #options_col,page_size = get_options(req_post,session_id)          
        val_dict.update({
            'msg': msg,
            'session_id': session_id,
            'location': location,  
            #'options_col': options_col,
            #'page_size': page_size,            
        })
        existing_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id=session_id)
        if existing_woos:
            existing_woos.delete()
        from portal.tasks import add_wo_record
        #res = add_wo_record.delay(quapi_id=quapi_id,location=location,session_id=session_id,exact_match=False,is_dash=True) 
        res = add_wo_record.apply_async(
            queue='refresh', 
            priority=1, 
            kwargs={
                'quapi_id':quapi_id,
                'location':location,
                'session_id':session_id,
                'exact_match':False,
                'is_dash':True,
            })
        error,msg = res.get()         
        if error == 'no errors':
            error = '' 
        all_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id=session_id)            
        val_dict['all_woos'] = WOStatus.objects.filter(active=1, is_dashboard=1, session_id=session_id)
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['total_rows'] = str(len(all_woos))
    val_dict['form'] = form
    return render(request, 'mrolive/wo_dashboard_refresh_display.html', val_dict)
    
def closed_rec_dash(request,quapi_id,date_from='',date_to=''):
    template = 'mrolive/closed_dock_refresh.html'
    new_status,filter_status = '',''
    user_id,user_rec = 'user not set',None
    val_dict,form,all_woos = {},{},[]
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id
    """
        I need a receiving view like wo dash refresh. See mock up here. Let me know what questions you have
        "Order #" should be the order found in the tracking number lookup (SO/PO/RO)
        "# of Items" should return a count of RCD in the related RCH
        Other than Receiving Label printout that's all we need
        this dashboard should refresh and only filter based on RC_HEADER['OPEN_FLAG'] = 'T'
    """

    if request.method == 'POST': 
        req_post = request.POST
        date_from = req_post.get('date_from','') 
        date_to = req_post.get('date_to','') 
        date_from = date_from.replace("/", "-") 
        date_to = date_to.replace("/", "-") 
        session_id = req_post.get('date_session','')                       
        if date_from or date_to:
            path = '/portal/closed-receivers/%s/%s/%s'%(quapi_id,date_from,date_to)
            return redirect(path)
        else:
            error = 'Must enter from date.'
        
    if request.method == 'GET':            
        session_id = request.session.session_key 
                
        date_from = date_from.replace("/", "-") 
        date_to = date_to.replace("/", "-")          
        val_dict.update({
            'msg': msg,
            'session_id': session_id,
            'show_modal': not date_from and 'T' or 'F', 
            'date_from': date_from,
            'date_to': date_to,            
        })
        if date_from:
            val_dict = rec_dash_lines(quapi_id,session_id,date_from,date_to,val_dict)
            all_woos = WOStatus.objects.filter(session_id=session_id).order_by('priority')           
            val_dict['all_woos'] = all_woos        
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['total_rows'] = str(len(all_woos))
    val_dict['form'] = form
    return render(request, template, val_dict)

def rec_dash_lines(quapi_id,session_id,date_from,date_to,val_dict):
        
    existing_woos = WOStatus.objects.filter(session_id=session_id)
    if existing_woos:
        existing_woos.delete()
    from portal.tasks import get_receivers  
    res = get_receivers.apply_async(
        queue='refresh', 
        priority=1, 
        kwargs={
            'quapi_id':quapi_id,
            'session_id':session_id,
            'closed': True,
            'from_date': date_from,
            'to_date': date_to,
        })
    error,msg = res.get()         
    if error == 'no errors':
        error = ''  
    val_dict['error'] = error    
    val_dict['all_woos'] = WOStatus.objects.filter(session_id=session_id)
    val_dict['total_rows'] = str(len(val_dict['all_woos']))
    return val_dict
        
  
def rec_dash(request,quapi_id,closed='F'):
    template = 'mrolive/dock_refresh.html'
    if closed != 'F':
        template = 'mrolive/closed_dock_refresh.html'
    new_status,filter_status = '',''
    user_id,user_rec = 'user not set',None
    val_dict,form,all_woos = {},{},[]
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id
    """
        I need a receiving view like wo dash refresh. See mock up here. Let me know what questions you have
        "Order #" should be the order found in the tracking number lookup (SO/PO/RO)
        "# of Items" should return a count of RCD in the related RCH
        Other than Receiving Label printout that's all we need
        this dashboard should refresh and only filter based on RC_HEADER['OPEN_FLAG'] = 'T'
    """    
    if request.method == 'GET':            
        session_id = request.session.session_key                
        val_dict.update({
            'msg': msg,
            'session_id': session_id,          
        })
        existing_woos = WOStatus.objects.filter(session_id=session_id)
        if existing_woos:
            existing_woos.delete()
        closed = closed == 't' and True
        from portal.tasks import get_receivers
        res = get_receivers.apply_async(
            queue='refresh', 
            priority=1, 
            kwargs={
                'quapi_id':quapi_id,
                'session_id':session_id,
                'closed': closed,
            })
        error,msg = res.get()         
        if error == 'no errors':
            error = '' 
        all_woos = WOStatus.objects.filter(session_id=session_id).order_by('priority')           
        val_dict['all_woos'] = all_woos
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['total_rows'] = str(len(all_woos))
    val_dict['form'] = form
    return render(request, template, val_dict)     

def refresh_lines(quapi_id,dj_user_id,session_id,location,warehouse,val_dict):
        
    val_dict.update({
        'session_id': session_id,
        'location': location,  
        'warehouse': warehouse,         
    })
    existing_woos = WOStatus.objects.filter(session_id=session_id)
    if existing_woos:
        existing_woos.delete()
    from portal.tasks import add_wo_record  
    res = add_wo_record.apply_async(
        queue='refresh', 
        priority=1, 
        kwargs={
            'quapi_id':quapi_id,
            'location':location,
            'warehouse':warehouse,
            'session_id':session_id,
            'exact_match':False,
            'is_dash':True,
        })
    error,msg = res.get()         
    if error == 'no errors':
        error = ''  
    val_dict['error'] = error    
    val_dict['all_woos'] = WOStatus.objects.filter(session_id=session_id)
    val_dict['total_rows'] = str(len(val_dict['all_woos']))
    return val_dict
    
def dash_refresh(request,quapi_id,warehouse='',location=''):
    new_status,filter_status = '',''
    user_id,user_rec = 'user not set',None
    wo_number = ''
    val_dict,form,all_woos = {},{},[]
    error,msg,loc_msg,stat_msg = '','','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id
    session_id = request.session.session_key
    if not location:
        location = warehouse
    if request.method == 'POST': 
        req_post = request.POST
        location = req_post.get('loc_input','') 
        warehouse = req_post.get('whs_input','')
        session_id = req_post.get('loc_session','')                
        val_dict = refresh_lines(quapi_id,dj_user_id,session_id,location,warehouse,val_dict)        
        if not warehouse:
            path = '/portal/wo-dash-refresh/%s/%s'%(quapi_id,location)
        else:
            path = '/portal/wo-dash-refresh/%s/%s/%s'%(quapi_id,warehouse,location)
        return redirect(path)        
    if request.method == 'GET':           
        if not location:
            #val_dict['error'] = 'Must have a location in the query string URI.'
            from portal.tasks import get_loc_whs_cart_nsync_beta
            res = get_loc_whs_cart_nsync_beta.delay(quapi_id,dj_user_id,loc_ok=True,whs_ok=True,app='dash-refresh') 
            loc_error,app = res.get() 
            if not loc_error:
                loc_vals = Location.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id)          
                val_dict['loc_vals'] = loc_vals
                whs_vals = Warehouse.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id).values('warehouse_code','whs_auto_key').distinct()          
                val_dict['whs_vals'] = whs_vals
                val_dict['session_id'] = session_id
                val_dict['loc_session'] = session_id
                return render(request,'mrolive/wo_dashboard_refresh.html', val_dict) 
            else:
                val_dict['error'] = loc_error
        #elif warehouse:
            #path = '/portal/wo-dash-refresh/%s/%s/%s'%(quapi_id,warehouse,location)
            #return redirect(path) 
        val_dict = refresh_lines(quapi_id,dj_user_id,session_id,location,warehouse,val_dict) 
        if 'error' in val_dict and val_dict['error'] == 'No records found.':
            loc_vals = Location.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id)          
            val_dict['loc_vals'] = loc_vals
            whs_vals = Warehouse.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id).values('warehouse_code','whs_auto_key').distinct()          
            val_dict['whs_vals'] = whs_vals
            val_dict['session_id'] = session_id
            val_dict['loc_session'] = session_id
            return render(request,'mrolive/wo_dashboard_refresh.html', val_dict)         
        #else:
            #path = '/portal/wo-dash-refresh/%s/%s'%(quapi_id,location)                    
    return render(request, 'mrolive/wo_dashboard_refresh.html', val_dict)     
          
def dashboard(request,quapi_id=None):
    new_status,location,filter_status = '','',''
    user_id,user_rec = 'user not set',None
    wo_number = ''
    val_dict,form = {},{}
    all_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id='') 
    updated_woos = all_woos 
    woos_updated = False    
    error,msg,loc_msg,stat_msg = '','','',''
    woos_updated = False
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id        
    from portal.tasks import get_statuses_nsync_beta
    if request.method == 'GET':
        res = get_statuses_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=1)
        stat_error = res.get()  
        form = WODashboardForm()        
    val_dict['status_vals'] = dj_user_id and stat_sel.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id,is_dashboard=1).distinct() or [] 
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        exact_match = 'exact_match' in req_post and req_post['exact_match'] or ''
        if exact_match:
            exact_match = 'checked'
        customer = 'customer' in req_post and req_post['customer'] or ''
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        new_status = 'new_status' in req_post and req_post['new_status'] or ''
        due_date = 'get_due_date' in req_post and req_post['get_due_date'] or ''
        new_due_date = 'due_date' in req_post and req_post['due_date'] or '' 
        search_mgr = 'get_manager' in req_post and req_post['get_manager'] or ''
        manager = 'manager' in req_post and req_post['manager'] or ''
        warehouse = 'warehouse' in req_post and req_post['warehouse'] or ''
        location = 'location' in req_post and req_post['location'] or ''
        if not (wo_number or customer or new_status or due_date or search_mgr or manager or warehouse or location):
            val_dict['error'] = 'Must make an entry in one of the search fields to filter.' 
            return render(request, 'mrolive/wodashboards.html', val_dict)             
        user_id = 'user_id' in req_post and req_post['user_id'] or ''
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0 
        options_col,page_size = get_options(req_post,session_id)            
        val_dict.update({
            'all_woos': updated_woos, 
            'msg': msg,
            'new_status': new_status and int(new_status), 
            'customer': customer,
            'get_due_date': due_date,
            'new_due_date': new_due_date,
            'user_id': user_id or '',
            'manager': manager,
            'get_manager': search_mgr,             
            'wo_number': wo_number,
            'session_id': session_id,
            'sel_rows': sel_rows,
            'total_rows': total_rows,
            'warehouse': warehouse,
            'location': location,
            'exact_match': exact_match,
            'options_col': options_col,
            'page_size': page_size,               
        })
        from portal.tasks import add_wo_record
        res = add_wo_record.delay(quapi_id=quapi_id,user_id='',manager=manager,due_date=due_date,warehouse=warehouse,location=location,wo_number=wo_number,session_id=session_id,exact_match=exact_match,is_dash=True) 
        error,msg = res.get()  
        if error == 'no errors':
            error = ''        
        updated_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id=session_id)
        val_dict['all_woos'] = updated_woos
    val_dict['msg'] = msg
    val_dict['error'] = error
    val_dict['total_rows'] = str(len(updated_woos))
    val_dict['form'] = form
    user_profile = UserProfile.objects.filter(user=user)
    user_profile = user_profile and user_profile[0] or None
    val_dict['num_records'] = user_profile.num_records or 10
    return render(request, 'mrolive/wodashboards.html', val_dict) 

def wo_mgmt_detail(request,quapi_id=None,woo_auto_key=None,session_id=None,wo_number=None):
    #this method takes a GET request along with a wo_number to generate all children woos under the woo passed in.
    val_dict={}
    total_rows = 0
    if not woo_auto_key:
        val_dict['error'] = 'No workorder. Try again.'
    else:
        new_status,location,filter_status = '','',''
        user_id,user_rec = 'user not set',None
        val_dict,form = {},{}   
        error,msg,loc_msg,stat_msg = '','','',''
        from polls.models import QuantumUser as quser,WOStatus as wos_obj
        user = request.user
        if not user.id:
            val_dict['error'] = 'Access denied.'
            return redirect('/login/')  
        reg_user_id = user and user.is_authenticated and user.id or None
        dj_user_id = user and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
        if not reg_user_id or not dj_user_id:
            val_dict['error'] = 'Access denied.'
            return redirect('/login/')  
        user_apps = user and UserAppPerms.objects.filter(user=user) or None
        op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
        val_dict['op_apps'] = op_apps
        mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
        val_dict['mgmt_apps'] = mgmt_apps
        dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
        val_dict['dash_apps'] = dash_apps
        setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
        val_dict['setup_apps'] = setup_apps
        from portal.tasks import create_wo_tasks
        res = create_wo_tasks.delay(quapi_id,1,session_id,woo_id=woo_auto_key)
        error = res.get()
        detail_rows = WOStatus.objects.filter(woo_auto_key=woo_auto_key,session_id=session_id)
        total_rows = len(detail_rows)        
        if not error and total_rows > 0:
            val_dict['total_rows'] = total_rows
            val_dict['all_woos'] = True
        else:
            error += 'There are no tasks.'
        val_dict['quapi_id'] = quapi_id
        val_dict['session_id'] = session_id
        val_dict['woo_auto_key'] = woo_auto_key
        val_dict['wo_number'] = wo_number
        user_profile = UserProfile.objects.filter(user=user)
        user_profile = user_profile and user_profile[0] or None
        val_dict['num_records'] = user_profile.num_records or 50
    return render(request, 'mrolive/wo_mgmt_detail.html', val_dict)  
        
def management(request,quapi_id=None):
    new_status,location,filter_status,error,total_rows = '','','','','0'
    user_id,user_rec,keep_recs = 'user not set',None,False
    wo_number = ''
    val_dict,form = {},{}
    all_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id='') 
    updated_woos = all_woos 
    woos_updated = False    
    msg,loc_msg,stat_msg = '','',''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_id = user.username or 'none'
    user_profile = UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None  
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')   
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id
    from portal.tasks import get_statuses_nsync_beta,get_users_nsync_beta
    if request.method == 'GET':
        res = get_statuses_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=1,app='wo-management')
        stat_error,app = res.get() 
        res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=1,app='wo-management')
        user_error,app = res.get()   
        val_dict['sel_rows'] = 0
        form = WODashboardForm()        
    val_dict['emp_vals'] = dj_user_id and quser.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id).distinct() or [] 
    val_dict['status_vals'] = dj_user_id and stat_sel.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id,is_dashboard=1).distinct()
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        exact_match = 'exact_match' in req_post and req_post['exact_match'] or ''
        if exact_match == '1':
            exact_match = 'checked'
        is_bom_sched = 'bom_sched_val' in req_post and req_post['bom_sched_val'] or '' 
        if not is_bom_sched and is_bom_sched != '0':
            is_bom_sched = 'is_bom_sched' in req_post and req_post['is_bom_sched'] or ''            
        active_user = 'active_user' in req_post and req_post['active_user'] or ''
        part_number = req_post.get('part_number','')
        customer = 'customer' in req_post and req_post['customer'] or ''
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        new_status = 'new_status' in req_post and req_post['new_status'] or ''
        upd_status = 'upd_status' in req_post and req_post['upd_status'] or ''
        active_due_date = 'active_due_date' in req_post and req_post['active_due_date'] or ''
        get_due_date = 'get_due_date' in req_post and req_post['get_due_date'] or ''
        due_date = 'due_date' in req_post and req_post['due_date'] or '' 
        search_mgr = 'get_manager' in req_post and req_post['get_manager'] or ''
        manager = 'manager' in req_post and req_post['manager'] or ''
        warehouse = 'warehouse' in req_post and req_post['warehouse'] or ''
        location = 'location' in req_post and req_post['location'] or ''
        rank = 'rank' in req_post and req_post['rank'] or ''
        update_user = 'update_user' in req_post and req_post['update_user'] or ''
        manual_ecd = req_post.get('arrival_date',None)
        user_name = 'user_name' in req_post and req_post['user_name'] or ''
        #lookup user_id in the database to make sure we can authenticate    
        show_all = 'show_all' in req_post and req_post['show_all'] or (user_name and 1) or None
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0
        #update hidden fields
        filter_status = 'filter_status' in req_post and req_post['filter_status'] or ''
        filter_due_date = 'filter_due_date' in req_post and req_post['filter_due_date'] or ''
        filter_customer = 'filter_customer' in req_post and req_post['filter_customer'] or ''
        filter_number = 'filter_number' in req_post and req_post['filter_number'] or ''
        filter_manager = 'filter_manager' in req_post and req_post['filter_manager'] or ''
        filter_location = 'filter_location' in req_post and req_post['filter_location'] or ''
        filter_warehouse = 'filter_warehouse' in req_post and req_post['filter_warehouse'] or ''
        filter_session = 'filter_session' in req_post and req_post['filter_session'] or ''
        update_session = 'update_session' in req_post and req_post['update_session'] or ''
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id and not (filter_session or update_session):
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
            filter_session = session_id
        if not session_id and (filter_session or update_session):
            session_id = filter_session or update_session                       
        dash_update = 'dash_update' in req_post and req_post['dash_update'] or None     
        options_col,page_size = get_options(req_post,session_id)     
        val_dict.update({
            'all_woos': updated_woos, 
            'msg': msg,
            'new_status':new_status and int(new_status) or filter_status and int(filter_status), 
            'customer': customer or filter_customer,
            'get_due_date': get_due_date or filter_due_date,
            'due_date': due_date or active_due_date,
            'user_id': user_id or active_user or update_user,                  
            'manager': manager,
            'location': location,
            'warehouse': warehouse,
            'part_number': part_number,
            'get_manager': search_mgr or filter_manager,
            'rank': rank,
            'manual_ecd': manual_ecd,
            'upd_status': upd_status,
            'wo_number': wo_number or filter_number,
            'filter_status': filter_status and int(filter_status) or new_status and int(new_status),
            'filter_customer': filter_customer or customer,
            'filter_number': filter_number or wo_number,
            'filter_due_date': filter_due_date or get_due_date,
            'filter_manager': filter_manager or search_mgr,
            'filter_location': filter_location or location,
            'filter_warehouse': filter_warehouse or warehouse,
            'filter_session': filter_session or session_id,#even when the update form is submitted, filter session always takes
            'update_session': update_session or session_id,
            'update_user': update_user or user_id or active_user,
            'session_id': session_id,
            'sel_rows': sel_rows,
            'total_rows': total_rows,
            'user_name': user_name,
            'active_user': active_user or user_id,
            'show_all': show_all or len(user_name) or active_user,
            'is_bom_sched': is_bom_sched, 
            'exact_match': exact_match,
            'options_col': options_col,
            'page_size': page_size,  
        })
        form = WODashboardForm(val_dict)
        val_dict['form'] = form
        """if 'user_id' in req_post and not user_id:
            msg += 'You must enter your Employee ID before updating any WO\'s.'
            val_dict['error'] = msg
            return render(request, 'mrolive/womgmt.html', val_dict)
        if user_id and not user_rec:
            msg += 'Invalid employee number.  Please enter a valid one.'
            val_dict['error'] = msg
            return render(request, 'mrolive/womgmt.html', val_dict)"""
        woo_id_list = []
        if 'woos_list[]' in req_post:
            woo_id_list = req_post.getlist('woos_list[]')                    
        if is_bom_sched and is_bom_sched != '0':
            from portal.tasks import bom_schedule
            res = bom_schedule.delay(woo_id_list,quapi_id,user_id,session_id,1) 
            error,msg = res.get()
            if not error:
                msg = 'Succesful BOM scheduling.'            
            val_dict['error'] = error
            val_dict['msg'] = msg            
        elif manual_ecd or rank or due_date or manager or upd_status:
            #now, set the user on the active woos for the dashboard
            if not woo_id_list:
                error = 'You must select at least one row in the grid to proceed with the update.'
                val_dict['error'] = error
            else:
                #Why do we need to update the woos now?  Just an expensive operation that gives us nothing
                #update_the_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id = (filter_session or update_session)).update(user_id=user_id,reg_user_id = user_rec)                                 
                try:                
                    from portal.tasks import make_updates
                    res = make_updates.delay(
                        sysur_auto_key=sysur_auto_key,                              
                        user_id=user_id, 
                        rank=rank,
                        manager=manager,
                        due_date=get_due_date,
                        new_due_date=due_date or active_due_date,
                        manual_ecd=manual_ecd,
                        customer=filter_customer,
                        status=upd_status,
                        search_mgr=filter_manager or search_mgr,
                        wo_number=filter_number,
                        session_id=session_id,
                        woo_ids=woo_id_list,
                        quapi_id=quapi_id,
                        )
                    error,msg = res.get()                     
                    #updated_woos = WOStatus.objects.filter(active=1, is_dashboard=1, session_id=(filter_session or update_session)) 
                    #val_dict['all_woos'] = updated_woos
                    val_dict['sel_rows'] = 0 
                    val_dict['error'] = error
                except Exception as exc:
                    logger.exception('Sending make_updates task raised this exception: %r', exc)               
        elif customer or part_number or search_mgr or location or warehouse or get_due_date or wo_number or new_status:
            if dash_update:
                keep_recs = True
            from portal.tasks import add_wo_record
                                                   
                                    
                                        
            res = add_wo_record.delay(is_mgmt=True,keep_recs=keep_recs,part_number=part_number,customer=customer,status=new_status,manager=search_mgr,location=location,warehouse=warehouse,due_date=get_due_date,wo_number=wo_number,session_id=session_id,quapi_id=quapi_id,exact_match=exact_match) 
            error,msg = res.get()             
            val_dict['sel_rows'] = 0
        else:
            error = 'Must have an entry in at least one filter.'
        if error and not (rank or due_date or manager or upd_status):
            updated_woos = []
            total_rows = '0'
        else:
            updated_woos = WOStatus.objects.filter(session_id=session_id or filter_session)
            if not updated_woos:
                updated_woos = WOStatus.objects.filter(session_id=filter_session) 
            total_rows = str(len(updated_woos))                
        val_dict['all_woos'] = updated_woos             
    val_dict['msg'] = msg
    val_dict['total_rows'] = total_rows
    val_dict['error'] = error
    val_dict['form'] = form
    return render(request, 'mrolive/womgmt.html', val_dict)  
   
#****************************************************begin wostatus****************************************************************   
def get_control(barcode,delim_str):
    ctrl_number = barcode.partition(delim_str)
    ctrl_id = ctrl_number and ctrl_number[2] or None
    ctrl_number = ctrl_number and ctrl_number[0] or None 
    return ctrl_number,ctrl_id 
    
def get_modes(app_id=None):  
    """a.	Reserve
        b.	Un-Reserve
        c.	Issue
        d.	Un-Issue"""

    modes = app_id and AppModes.objects.filter(ml_apps_id=app_id).order_by('sequence')\
        or AppModes.objects.filter(ml_apps_id=None) or None
    return modes

    
def barcarting_beta(request,quapi_id=None): 
    location,user_id,user_logged,update,warehouse,rack,new_rack,rerack,rack_user = '','','','','','','','',''
    wo_number,user_error,stat_error,loc_error = '','','',''
    val_dict,form,updated_woos,all_woos,woo_num_list,woo_key_list = {},{},[],[],[],[]           
    msg,loc_msg,stat_msg,error,lookup_recs,clear_cart,iq_enable = '','','','',False,False,False
    loc_key,whs_key,cart_key,new_status_name,is_rch,syscm_auto_key,dpt_auto_key=None,None,None,'',False,None,None
    ml_apps_id = MLApps.objects.filter(code='barcoding')
    ml_apps_id = ml_apps_id and ml_apps_id[0] or None
    right_now = datetime.now()
    timestamp = right_now.strftime("%Y-%m-%dT%H:%M")
    modes = []
    if ml_apps_id:
        modes = get_modes(app_id=ml_apps_id)
        val_dict['modes'] = modes
    val_dict['quapi_id'] = quapi_id 
    user = request.user
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    auth_user_name = user.username
    user_profile = UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None  
    reg_user_id = user and user.is_authenticated and user.id or None 
    #check to see if the user has access by searching for userquapirel record
    dj_user_id = user and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')         
    if request.method == 'GET':
        val_dict['total_rows'] = None
        val_dict['active_mode'] = '1'
        if not reg_user_id or not dj_user_id:
            val_dict['error'] = 'Access denied.'
            return redirect('/login/') 
        #from portal.tasks import get_loc_whs_cart_nsync_beta
        form = WODashboardForm()
        val_dict['form'] = form           
        #res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=0,app='barcoding')
        #user_error,app = res.get() 
        #res = get_loc_whs_cart_nsync_beta.delay(quapi_id,dj_user_id,loc_ok=True,whs_ok=True,cart_ok=True,app='barcoding') 
        #loc_error,app = res.get()                       
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    #val_dict['departments'] = Departments.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id)   
    #val_dict['acc_cos'] = Companies.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id,is_acc_co=True)          
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        val_dict['form'] = form
        ctrl_id,ctrl_number = '',''
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        num_rows = 'num_rows' in req_post and req_post['num_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        if not dj_user_id:
            dj_user_id = 'dj_user_id' in req_post and req_post['dj_user_id'] or ''#dj admin user id
        #lookup user_id in the database to make sure we can authenticate
        #user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id)
        #user_rec = user_rec and user_rec[0] or None
        clear_form = 'clear_form' in req_post and req_post['clear_form'] or False       
        lookup_recs = 'lookup_recs' in req_post and req_post['lookup_recs'] or False           
        location = 'location' in req_post and req_post['location'] or '' 
        rack = 'rack' in req_post and req_post['rack'] or '' 
        new_rack = 'new_rack' in req_post and req_post['new_rack'] or '' 
        warehouse = 'warehouse' in req_post and req_post['warehouse'] or ''
        active_mode = 'mode_selector' in req_post and req_post['mode_selector'] or ''
        val_dict['active_mode'] = active_mode
        sel_mode = 'sel_mode' in req_post and req_post['sel_mode'] or '' 
        if not (sel_mode or active_mode):
            val_dict['error'] = 'Must select a mode.'
            render(request, 'mrolive/barcoding.html', val_dict)
        loc_cart = req_post.get('loc_cart','')            
        cart_code = 'cart_code' in req_post and req_post['cart_code'] or ''         
        new_status = 'new_status' in req_post and req_post['new_status'] or ''     
        show_status = 'show_status' in req_post and req_post['show_status'] or ''
        show_user = 'show_user' in req_post and req_post['show_user'] or ''
        show_all = 'show_all' in req_post and req_post['show_all'] or ''
        clear_cart = 'ccart_form' in req_post and True or False
        stock_label = 'stock_label' in req_post and req_post['stock_label'] or ''
        do_status = sel_mode and (sel_mode == '2' or sel_mode == '1') or False
        do_user = sel_mode or False       
        #do_all = user_id or user_logged or rack_user or False 
        quantity='quantity' in req_post and req_post['quantity'] or ''      
        department='dept_input' in req_post and req_post['dept_input'] or ''
        condition='condition_code' in req_post and req_post['condition_code'] or ''      
        account_company = 'acco_input' in req_post and req_post['acco_input'] or ''
        consignment='consignment' in req_post and req_post['consignment'] or ''                                                                       
        clear_rack = 'clear_rack' in req_post and req_post['clear_rack'] or ''
        yes_clear = 'yes_clear' in req_post and req_post['yes_clear'] or False
        no_clear = 'no_clear' in req_post and req_post['no_clear'] or False
        cur_mode = 'cur_mode' in req_post and req_post['cur_mode'] or False
        val_transfer = req_post.get('val_transfer','')
        print_manifest = req_post.get('print_manifest','')
        confirm_transfer = req_post.get('confirm_transfer','')
        cart_transfer = req_post.get('cart_transfer','')
        launch_transfer = req_post.get('is_transfer','')
        do_transfer = req_post.get('do_transfer','')
        val_transfer = req_post.get('val_transfer','')     
        trans_loc = req_post.get('trans_loc','')
        trans_cart = req_post.get('trans_cart','')
        stm_keys = req_post.get('stm_keys','')

                                                                                                                   
                                                                     
        if no_clear:       
            msg = 'Clear cart cancelled.'
            val_dict['active_mode'] = cur_mode
        if active_mode == '1' and not wo_number:
            val_dict['error'] = 'Enter a valid record.'
            return render(request, 'mrolive/barcoding.html', val_dict)              
        if account_company:
            acc_co = Companies.objects.filter(name=account_company,dj_user_id=dj_user_id,quapi_id=quapi_id)
            syscm_auto_key = acc_co and acc_co[0] and acc_co[0].cmp_auto_key or None
            if not syscm_auto_key:
                error = 'Account company not found.'
        if department:
            dept = Departments.objects.filter(name=department,dj_user_id=dj_user_id,quapi_id=quapi_id)
            dpt_auto_key = dept and dept[0] and dept[0].dpt_auto_key or None            
            if not dpt_auto_key:
                error = 'Department not found.' 
                
        val_dict.update({
            'wo_number': '',
            'all_woos': updated_woos,
            'loc_cart': loc_cart,
            'msg': msg,
            
            'warehouse': warehouse,
            'location': location,
            'dj_user_id': dj_user_id,
            #'user_id': user_id or user_logged or rack_user,
            #'user_name': auth_user_name,
            'rack': rack or cart_code or clear_rack,
            'clear_rack': clear_rack or rack or cart_code,
            #'user_logged': user_logged or user_id,
            #'rack_user': user_id,
            'new_rack': rack,
            'modes': modes,
            'active_mode': active_mode or sel_mode or cur_mode or '',
            'sel_mode': sel_mode or active_mode or cur_mode or '',
            'cur_mode': cur_mode or active_mode or sel_mode or '',
            'cart_code': cart_code or rack or '',
            'stock_label': stock_label or wo_number or '',
            'new_status': new_status and int(new_status) or None,
            'lookup_recs': lookup_recs,
            'show_status': show_status or do_status,
            'do_status': do_status or show_status,
            'show_user': show_user or do_user,                                 
            'do_user': do_status or show_status,
            'show_all': show_all and show_all!='0',
            'session_id': session_id,
            'sel_rows': sel_rows,
            'total_rows': total_rows or num_rows,
            'num_rows': num_rows or total_rows,
            'is_rch': is_rch,
            'form': form,                                                                
            })
            
        if new_status:
            new_status_name = val_dict['status_vals'].filter(wos_auto_key = new_status)
            new_status_name = new_status_name and new_status_name[0] and new_status_name[0].name or '' 
        #yes_clear = 'yes_clear' in req_post and req_post['yes_clear'] or False
        #no_clear = 'no_clear' in req_post and req_post['no_clear'] or False 
        
        if clear_cart or yes_clear:     
            rack = clear_rack or cart_code or rack or '' 
            val_dict['total_rows'] = 0        
            clear_cart = yes_clear and True or False
            if not (yes_clear or no_clear):
                val_dict['confirm_clear'] = 'True' 
                val_dict['cur_mode'] = sel_mode
                return render(request, 'mrolive/barcoding.html', val_dict)
                
        #if user submitted clear list form by pressing button
        if clear_form and req_post['clear_form']=='1':         
            WOStatus.objects.filter(user_id=auth_user_name,is_dashboard=0,active=1,is_racking=1).delete()            
            val_dict['msg'] = '' 
            val_dict['active_mode'] = sel_mode 
            val_dict['user_id'] = auth_user_name
            val_dict['show_all'] = 0 
            form = WODashboardForm(val_dict)
            val_dict['form'] = form             
            return render(request, 'mrolive/barcoding.html', val_dict)
            
        wo_number = wo_number or stock_label or '' 
        ctrl_number,ctrl_id,stm_auto_key='','',''       
        if wo_number:
            if wo_number[0] in ['c','C']:
                stm_auto_key = wo_number[1:]               
            elif len(wo_number) < 7:
                error = "Your barcode is not long enough."        
            else:
                ctrl_number = wo_number[:-6]               
                ctrl_id = wo_number[-6:]
                
                if ctrl_number and ctrl_id:
                    from portal.tasks import check_stock
                    res = check_stock.delay(quapi_id,ctrl_id,ctrl_number)
                    error,msg = res.get()
                    
                    if error:
                        val_dict['error'] = error
                        return render(request, 'mrolive/barcoding.html', val_dict)  
                        
        if rack or trans_cart or loc_cart:
            if trans_cart:
                rack = trans_cart
            elif loc_cart:
                rack = loc_cart
            from portal.tasks import lookup_stock_cart
            res = lookup_stock_cart.delay(quapi_id,session_id,rack)
            error,msg = res.get()
            cart = StockCart.objects.filter(session_id=session_id,
                udl_code__iexact=rack
                )  
            cart = cart and cart[0] or None
            if cart:
                cart_code = cart.udl_code
                cart_key = cart.udl_auto_key
            elif not loc_cart:
                val_dict['error'] = 'Cart not found: %s'%rack
                return render(request,'mrolive/barcoding.html', val_dict)                
        

        if location or trans_loc or loc_cart:
          
            if trans_loc:
                location = trans_loc
            elif loc_cart:
                location = loc_cart            
            from portal.tasks import lookup_location
            res = lookup_location.delay(quapi_id,session_id,location)
            error,msg = res.get()
            
            if error and not loc_cart:
                val_dict['error'] = error
                return render(request,'mrolive/barcoding.html', val_dict)
            
            loc_key = Location.objects.filter(session_id=session_id,
                location_code__iexact=location
                )            
            loc_key = loc_key and loc_key[0] or None           
            
            if loc_key:
                location = loc_key.location_code
                iq_enable = loc_key.iq_enable
                loc_key = loc_key.loc_auto_key  

                if loc_cart and cart_code and location and cart_code == location:
                                                                                                                  
                               
                    loc_key = None
                    
            elif not loc_cart:
                val_dict['error'] = 'Location not found: %s'%location
                return render(request, 'mrolive/barcoding.html', val_dict)
                                 
        if warehouse:
            from portal.tasks import lookup_warehouse
            res = lookup_warehouse.delay(quapi_id,session_id,warehouse)
            error,msg = res.get()
            
            if error:
                val_dict['error'] = error
                return render(request,'mrolive/barcoding.html', val_dict)
                
            whs = Warehouse.objects.filter(session_id=session_id,
                warehouse_code__iexact=location
                )  
            whs = whs and whs[0] or None
            if whs:
                warehouse = whs.warehouse_code
                whs_key = whs.whs_auto_key
            else:
                val_dict['error'] = 'Warehouse not found: %s'%warehouse
                return render(request, 'mrolive/barcoding.html', val_dict)  
                
        elif launch_transfer == '1':
            stm_keys = WOStatus.objects.filter(session_id=session_id).values_list('stm_auto_key',flat=True)
            stm_keys = list(stm_keys)
            
            val_dict.update({
                'msg': '', 
                'active_mode': active_mode, 
                'user_id': auth_user_name,
                'launch_transfer': '1',
                'stm_keys': stm_keys,
                'location': location,
                'rack': cart_code,               
            })
            
        elif do_transfer or val_transfer:
            if not (trans_loc or trans_cart):
                error = 'Enter location or cart for transfer.'
                
            val_dict.update({
                'error': error,
                'msg': '', 
                'active_mode': cur_mode, 
                'user_id': auth_user_name,
            })    

            if not error:
                parameters=[loc_key,iq_enable,cart_key,trans_cart]
                #from polls.models import UserInput
                
                #set up the data and create the record
                if not loc_key:
                    location = ''
                    loc_key = ''
                
                #rec = UserInput(
                #    user_name = auth_user_name,
                #    sysur_auto_key = sysur_auto_key,
                #    timestamp = timestamp,
                #    user_inputs = 'Loc/cart: ' + str(location) + '('\
                #        + str(loc_key) + ')' + ' | Cart: ' + str(cart_code)\
                #        + '(' + str(cart_key) + ')' + ' | Label: '\
                #        + str(wo_number or stock_label),
                #    app_mode = active_mode=='1' and 'Update'\
                #        or active_mode=='2' and 'Transfer'\
                #        or active_mode=='3' and 'Validate',
                #    ml_apps_id = ml_apps_id,
                #    app_name = ml_apps_id and ml_apps_id.name or '',
                #)
                #rec.save()
                
                from portal.tasks import transfer_val_stock         
                res = transfer_val_stock.apply_async(
                        args=[quapi_id,auth_user_name,sysur_auto_key,session_id,parameters,stm_keys],
                        queue='bulk',
                        priority=0, 
                        )                            
                error,msg = res.get()
                val_dict.update({
                    'error': error,
                    'msg': msg,
                })
        
        elif confirm_transfer == '1':
        
            if not cart_key:
                error = 'Enter cart to transfer.'
                
            stm_keys = WOStatus.objects.filter(session_id=session_id)
            total_stms = len(stm_keys)
            
            val_dict.update({
                'msg': '',
                'active_mode': active_mode, 
                'user_id': auth_user_name,
                'confirm_transfer': '1',
                'total_stms': total_stms,
                'session_id': session_id,
                'rack': rack,
                'location': location,
            })
            
        elif active_mode in ['3'] and not (stock_label or wo_number):

            if (cart_key or loc_key):
                from portal.tasks import search_val_stock
                parameters=[loc_key,cart_key,cart_code]
                res = search_val_stock.delay(quapi_id,auth_user_name,\
                sysur_auto_key,session_id,parameters    
                )
                error,msg = res.get()
            stm_recs = WOStatus.objects.filter(session_id=session_id)            
            val_dict['total_rows'] = len(stm_recs)
            val_dict['error'] = error
            
            if print_manifest == '1':
            
                print_dict = ({
                    'stm_recs': stm_recs,
                    'quapi_id': quapi_id,
                })
                
                return render(request, 'mrolive/validate_manifest.html', print_dict)                         
            
        elif launch_transfer != '1' and confirm_transfer != '1'\
            and (condition or quantity or\
            consignment or syscm_auto_key or\
            dpt_auto_key or clear_cart or cart_key or loc_key or\
            whs_key or wo_number):
 
            if active_mode == '2' and cart_key and loc_key and not cart_transfer:
                loc_key = ''
                
            user_groups = request.user.groups.values_list('id',flat=True)
            group = user_groups and user_groups[0] or None
            group = UserGroupProfile.objects.filter(group=group)
            group_queue = group and group[0] and group[0].priority or None
            res = None
            from portal.tasks import run_racking_beta       
            #logger.error('User: ' + auth_user_name + ' | ') 
            #logger.error('ctrl#/id: ' + ctrl_number + ' / ' + ctrl_id + ' | ')
            #logger.error('label: ' + wo_number + ' | cart: ' + cart_code + ' | ')
            #logger.error('whs: ' + warehouse + ' | location: ' +  location  + ' | ')
            if group_queue and group_queue == 'highest_priority':
                res = run_racking_beta.apply_async(
                    args=[session_id],
                    queue=group_queue, 
                    priority=0, 
                    kwargs={
                        'rack_auto_key': cart_key,
                        'loc_auto_key': loc_key,
                        'wo_number': wo_number or stock_label,
                        'user_id': auth_user_name,
                        'whs_auto_key': whs_key,
                        'new_status': new_status,
                        'new_status_name': new_status_name,
                        'sysur_auto_key': sysur_auto_key,
                        'woo_key_list': woo_key_list,
                        'ctrl_id': ctrl_id,
                        'ctrl_number': ctrl_number,
                        'stock_label': wo_number or stock_label,
                        'mode': active_mode or sel_mode or cur_mode,
                        'lookup_recs': lookup_recs,
                        'quapi_id': quapi_id,
                        'clear_cart': clear_cart,
                        'dj_user_id': dj_user_id,
                        'rack': cart_code,
                        'location':location,
                        'warehouse':warehouse,
                        'cond_code':condition,
                        'syscm_auto_key':syscm_auto_key,
                        'consignment':consignment,
                        'quantity':quantity,
                        'dpt_auto_key':dpt_auto_key,
                        'iq_enable':iq_enabledef
                    })
                 
            else:
                from polls.models import UserInput
                
                #set up the data and create the record
                if not loc_key:
                    location = ''
                    loc_key = ''
                
                rec = UserInput(
                    user_name = auth_user_name,
                    sysur_auto_key = sysur_auto_key,
                    timestamp = timestamp,
                    user_inputs = 'Loc/cart: ' + str(location) + '('\
                        + str(loc_key) + ')' + ' | Cart: ' + str(cart_code)\
                        + '(' + str(cart_key) + ')' + ' | Label: '\
                        + str(wo_number or stock_label),
                    app_mode = active_mode=='1' and 'Update'\
                        or active_mode=='2' and 'Transfer'\
                        or active_mode=='3' and 'Validate',
                    ml_apps_id = ml_apps_id,
                    app_name = ml_apps_id and ml_apps_id.name or 'None',
                )
                rec.save()
                                             
                res = run_racking_beta.delay(
                    session_id,
                    rack_auto_key = cart_key,
                    loc_auto_key=loc_key,
                    wo_number=wo_number or stock_label,
                    user_id=auth_user_name,
                    whs_auto_key=whs_key,
                    new_status=new_status,
                    new_status_name=new_status_name,
                    sysur_auto_key=sysur_auto_key,
                    woo_key_list=woo_key_list,
                    ctrl_id = ctrl_id,
                    ctrl_number = ctrl_number,
                    stock_label = wo_number or stock_label,
                    mode = active_mode or sel_mode or cur_mode,
                    lookup_recs = lookup_recs,
                    quapi_id = quapi_id,
                    clear_cart = clear_cart,
                    dj_user_id = dj_user_id,
                    rack = cart_code,
                    location=location,
                    warehouse=warehouse,
                    cond_code=condition,
                    syscm_auto_key=syscm_auto_key,
                    consignment=consignment,
                    quantity=quantity,
                    dpt_auto_key=dpt_auto_key,
                    iq_enable = iq_enable,
                )
                
            error,msg,is_rch = res.get()
            val_dict['is_rch'] = is_rch and '1' or '0'
            recs = WOStatus.objects.filter(session_id=session_id)
            val_dict['total_rows'] = len(recs)
            
        #if (active_mode == '3' or cur_mode == '3') and not clear_cart:            
            #updated_woos = not is_rch and WOStatus.objects.filter(session_id=session_id) or []
            #updated_rchs = is_rch and StockReceiver.objects.filter(session_id=session_id) or []            
            #val_dict['total_rows'] = str(len(updated_woos or updated_rchs))
                             
        val_dict['msg'] = msg
        val_dict['error'] = error + user_error + stat_error + loc_error       
        if not wo_number and lookup_recs not in [1,'1']:
            val_dict['lookup_recs'] = 1
        elif wo_number and lookup_recs not in [0,'0']:
            val_dict['lookup_recs'] = 0    
    val_dict['user_name'] = auth_user_name            
    form = WODashboardForm(val_dict)
    val_dict['form'] = form    
    return render(request, 'mrolive/barcoding.html', val_dict)

def prepend(list, str):      
    # Using format() 
    str += '{0}'
    list = [str.format(i) for i in list] 
    return(list) 
    
def gen_range(base,from_num,to_num):
    #we add 1 to the upper bound because the built-in range function doesn't include the to_num in it's output
    list_range = list(range(int(from_num),int(to_num)+1))
    final_list = []
    elements = []
   
    for elem in list_range:
        elem = str(elem)
        #add the number of zeroes 
        #necessary to get to the total 
        #length of from_num
        if len(elem) < len(from_num):
            #leading_zeroes
            diff_length = len(from_num) - len(elem)
            count = 0
            leading_zeroes = ''
            while count < diff_length:
                leading_zeroes += '0'
                count += 1
            elem = leading_zeroes + elem
        elements.append(elem.split(' ', 1)[0])
        final_list.append(elem)     
    final_list = prepend(final_list,base)
    return final_list,elements

def barcode_labels(request,quapi_id=None,loc_list=[],user_id='',user_name=''):
    val_dict,form,cart_labels,elements = {},None,[],''
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    val_dict['cart_labels'] = []   
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None 
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id    
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
    if request.method == 'POST':
        req_post = request.POST
        if not user_id:
            user_id = 'user_id' in req_post and req_post['user_id'] or ''#sysur_auto_key
        if not user_name:
            user_name = 'user_name' in req_post and req_post['user_name'] or ''
        #lookup user_id in the database to make sure we can authenticate
        user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id)
        user_rec = user_rec and user_rec[0] or None 
        code_type = 'type' in req_post and req_post['type'] or ''
        label = 'label' in req_post and req_post['label'] or ''
        from_num = 'from_num' in req_post and req_post['from_num'] or None
        to_num = 'to_num' in req_post and req_post['to_num'] or None       
        if from_num and to_num:
            cart_labels,elements = gen_range(label,from_num,to_num)           
        if user_rec:
            user_name = user_rec and user_rec.user_name or user_name
        if cart_labels:
            from portal.tasks import create_carts  
            res = create_carts.delay(quapi_id,sysur_auto_key,label,cart_labels)    
            error,msg,udls = res.get()             
            val_dict.update({
                'type': 'Code128',
                'label': label,
                'user_name': user_name,
                'user_id': user_id,
                'from_num': from_num,
                'to_num': to_num,
                'elements': elements,
                'cart_labels': cart_labels or [label] or [],
                'user_apps': user_apps,
                'quapi_id': quapi_id,
                'error': error,
                'msg': msg,
            })
            return render(request, 'mrolive/plain_barcode_label.html', val_dict)
        if loc_list:
            val_dict.update({
                'type': 'Code128',
                'user_name': 'barcode_gen',
                'user_id': 'barcode_gen',
                'cart_labels': loc_list,
            })
        form = WODashboardForm(val_dict)             
    if not form:
        form = WODashboardForm() 
    val_dict['form'] = form
    val_dict['user'] = user
    return render(request, 'mrolive/barcode_labels.html', val_dict)
   
def location_labels(request,quapi_id=None,loc_list=[],user_id='',user_name=''):
    val_dict,form,cart_labels,elements = {},None,[],''
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    val_dict['cart_labels'] = []   
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None 
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id    
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')
    #if request.method == 'GET': 
    session_id = 'aufOISs892gas843ag43'
    from portal.tasks import get_warehouses
    res = get_warehouses.delay(quapi_id,session_id)
    error,msg = res.get()
    warehouses = Warehouse.objects.filter(session_id = session_id)
    val_dict.update({
        'error': error,
        'msg': msg,
        'default_whs': 'WH-GLENDALE',
        'warehouse_codes': warehouses,
        'session_id': session_id,
    })             
    if request.method == 'POST':
        req_post = request.POST
        
        if not user_id:
            user_id = 'user_id' in req_post and req_post['user_id'] or ''#sysur_auto_key
            
        if not user_name:
            user_name = 'user_name' in req_post and req_post['user_name'] or ''

        session_id = req_post.get('session_id','')
        user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id)
        user_rec = user_rec and user_rec[0] or None 
        code_type = 'type' in req_post and req_post['type'] or ''
        label = 'label' in req_post and req_post['label'] or ''
        warehouse = 'warehouse' in req_post and req_post['warehouse'] or ''
        from_num = 'from_num' in req_post and req_post['from_num'] or None
        to_num = 'to_num' in req_post and req_post['to_num'] or None 
        
        if from_num and to_num:
            loc_labels,elements = gen_range(label,from_num,to_num)           
        if user_rec:
            user_name = user_rec and user_rec.user_name or user_name
            
        if loc_labels:
            whs_auto_key = ''
            from portal.tasks import create_locations,lookup_warehouse
            if warehouse:
                res = lookup_warehouse.delay(quapi_id,session_id,warehouse) 
                error,msg,whs_auto_key = res.get()
                
                if error:
                    val_dict['error'] = error
                    form = WODashboardForm(val_dict)
                    val_dict['form'] = form
                    return render(request, 'mrolive/location_labels.html', val_dict)
         
            res = create_locations.delay(quapi_id,sysur_auto_key,label,warehouse,whs_auto_key,loc_labels)    
            error,msg,loc_labels = res.get()  
            
            if not loc_labels:
                
                val_dict.update({
                    'error': 'There is a problem. No labels created: ' + error,
                    })
                form = WODashboardForm(val_dict)
                val_dict['form'] = form                
                return render(request, 'mrolive/location_labels.html', val_dict)
                
            loc_labels = loc_labels and [loc[0] for sublist in loc_labels for loc in sublist] or []
                
            val_dict.update({
                'type': 'Code128',
                'label': label,
                'user_name': user_name,
                'user_id': user_id,
                'from_num': from_num,
                'to_num': to_num,
                'elements': elements,
                'loc_labels': loc_labels or [],
                'user_apps': user_apps,
                'quapi_id': quapi_id,
                'error': error,
                'msg': msg,
            })
            return render(request, 'mrolive/plain_location_label.html', val_dict)
        else:
            val_dict['error'] = 'No labels generated.'
            
        form = WODashboardForm(val_dict)             
    if not form:
        form = WODashboardForm() 
    val_dict['form'] = form
    val_dict['user'] = user
    return render(request, 'mrolive/location_labels.html', val_dict)
  
#++++++++++++===================Audit Trail Code============================++++++++++++++++++
def audit_trail(request,quapi_id=None):
    from polls.models import AuditTrail as Adt
    filter_user_id,active_app,msg,error,user_name,show_grid = '','','','','',False
    val_dict,results,req_post,form,date_from,date_to,options_col,session_id = {},[],{},{},'','',[],''
    sel_status,aud_status,page_size = False,False,25
    right_now = datetime.now()
    date_to_input = right_now + timedelta(days=1)
    date_to_input = date_to_input.strftime('%m/%d/%Y')
    date_from_input = right_now - timedelta(days=45)
    date_from_input = date_from_input.strftime('%m/%d/%Y')
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user) or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    audit_apps = user_apps.filter(audit_ok = True)
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        format = '%Y-%m-%d %H:%M:%S'
        new_format = '%m/%d/%Y'
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        filter_user_id = 'user_id' in req_post and req_post['user_id'] or None
        date_from_input = 'date_from' in req_post and req_post['date_from'] or None
        date_from = date_from_input and datetime.strptime(date_from_input, new_format) or None
        date_from = date_from and datetime.strftime(date_from, format) or None
        date_to_input = 'date_to' in req_post and req_post['date_to'] or None
        date_to = date_to_input and datetime.strptime(date_to_input, new_format) or None
        date_to = date_to and datetime.strftime(date_to, format) or None
        active_app = 'app_selector' in req_post and req_post['app_selector'] or None
        user_id = filter_user_id and QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=filter_user_id) or None
        user_id = user_id and user_id[0] or None
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        user_name = user_id and user_id.user_name or ''
        show_grid = True
        #aud_status = 'status' in req_post and req_post['status'] or []
        #options_col,page_size = get_options(req_post,session_id)            
        #if not user_id:
        #    error = 'Invalid user.'
        if not active_app:
            error = 'Please enter an application.' 
        if not (date_from and date_to):
            error = 'Please enter a date from and date to.'        
    else:
        form = WODashboardForm(initial={'date_from': date_from_input, 'date_to': date_to_input})
    app_id = active_app and UserAppPerms.objects.filter(id=active_app) or None
    app_id = app_id and app_id[0] and app_id[0].ml_apps_id

    if app_id and app_id.id == 403:
        app_id = MLApps.objects.filter(code__icontains='Labor')
        app_id = [app.id for app in app_id]
        
    if req_post and app_id:     
        if filter_user_id:
            if date_from and date_to:
                results = Adt.objects.filter(quapi_id=quapi_id,ml_apps_id__in=app_id,user_id__iexact=filter_user_id,create_date__gte=date_from,create_date__lte=date_to) or []  
            elif date_from and not date_to:
                results = Adt.objects.filter(quapi_id=quapi_id,ml_apps_id__in=app_id,user_id__iexact=filter_user_id,create_date__gte=date_from) or []           
            elif not date_from and date_to: 
                results = Adt.objects.filter(quapi_id=quapi_id,ml_apps_id__in=app_id,user_id__iexact=filter_user_id,create_date__lte=date_to) or [] 
            else: 
                results = Adt.objects.filter(quapi_id=quapi_id,ml_apps_id__in=app_id,user_id__iexact=filter_user_id) or []             
        elif date_from and not date_to:
            results = Adt.objects.filter(quapi_id=quapi_id,ml_apps_id__in=app_id,create_date__gte=date_from) or []           
        elif not date_from and date_to: 
            results = Adt.objects.filter(quapi_id=quapi_id,ml_apps_id__in=app_id,create_date__lte=date_to) or [] 
        elif date_from and date_to:
            results = Adt.objects.filter(quapi_id=quapi_id,ml_apps_id__in=app_id,create_date__gte=date_from,create_date__lte=date_to) or [] 
        else: 
            results = Adt.objects.filter(quapi_id=quapi_id,ml_apps_id__in=app_id) or []               
    if not results and req_post and app_id:
        error = 'No audit trail records match your search.'
    val_dict.update({
        'app_set': user_apps,
        'audit_apps': audit_apps,
        'user_name': user_name,
        'user_id': filter_user_id,
        'user': user,
        'msg': msg,
        'error': error,
        'date_from': date_from_input,
        'date_to': date_to_input,
        'active_app': active_app and int(active_app) or None,
        'msg': msg,
        'quapi_id': quapi_id,
        'show_grid': show_grid,
        'total_rows': len(results),
        'form': form,
        #'aud_status': aud_status,
        'user': user,
        'options_col': options_col,
        'page_size': page_size,
        'session_id': session_id,
        })
    return render(request, 'mrolive/audit_trail.html', val_dict) 
    
#**************************************KENDO UI GRID CONTROLLERS********************************************************
from django.http import HttpResponse
import json
def results_grid_pop(request):
    results = []
    for record in WOStatus.objects.filter(is_dashboard=1,active=1):
        results.append({'wo_number':record.wo_number,'status':record.status,
        'time_status':record.time_status,'due_date_var':record.due_date_var,'rank':record.rank,
        'part_number':record.part_number,'description':record.description,'serial_number':record.serial_number,
        'manager':record.manager,'wo_type':record.wo_type,'location_code':record.location_code,
        'time_loc':record.time_loc,'cust_ref_number':record.cust_ref_number})
    data = json.dumps(results)
    return HttpResponse(data)
    #was a parameter passed into HttpResponse:, mimetype='application/javascript'
#from django.shortcuts import render, redirect
from django.views.generic import View
from django.contrib.auth.models import User
from django.views import generic
from django.contrib.auth.models import Group
from django.contrib.auth.forms import UserChangeForm
from django.http import HttpResponse, HttpResponseRedirect
from rest_framework import serializers, generics
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView


class SalePageNumberPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'pageSize'
    
    
class PIPageNumberPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'pageSize'

class OPMSerializer(serializers.ModelSerializer):
    class Meta:
        model = Operation
        fields = '__all__'
        
# Create your views here.
class OPMListView(generic.ListView):
    model = Operation
    def get_context_data(self, **kwargs):       
        context = super(OPMListView, self).get_context_data(**kwargs)
        return context

class OPMJsonView(generics.ListAPIView):
    serializer_class = OPMSerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):
        session_id = 'session_id' in self.request.GET and self.request.GET['session_id']        
        records = Operation.objects.filter(session_id=session_id)     
        field,direction = '',''
        
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records

    
class SrecSerializer(serializers.ModelSerializer):
    class Meta:
        model = StockReceiver
        fields = '__all__'
        
    
# Create your views here.
class SrecListView(generic.ListView):
    model = StockReceiver
    def get_context_data(self, **kwargs):       
        context = super(SrecListView, self).get_context_data(**kwargs)
        return context

class SrecJsonView(generics.ListAPIView):
    serializer_class = SrecSerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):
        session_id = 'session_id' in self.request.GET and self.request.GET['session_id']        
        records = StockReceiver.objects.filter(session_id=session_id)                
        field,direction = '',''
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records
 
class UserSerializer(serializers.ModelSerializer):
    class Meta:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        model = User
        fields = '__all__'

class UserProfileSerializer(serializers.ModelSerializer):
    class Meta:
        from polls.models import UserProfile
        model = UserProfile
        fields = '__all__'
        
# Create your views here.
class UserListView(generic.ListView):
    from django.contrib.auth import get_user_model
    User = get_user_model()
    model = User
    model = User
    def get_context_data(self, **kwargs):       
        context = super(UserListView, self).get_context_data(**kwargs)
        return context

class UserJsonView(generics.ListAPIView):
    serializer_class = UserSerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):
        session_id = 'session_id' in self.request.GET and self.request.GET['session_id']       
        from django.contrib.auth import get_user_model
        User = get_user_model()
        records = User.objects.all().order_by('last_name')
        field,direction = '',''
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records
        

class UproJsonView(generics.ListAPIView):
    serializer_class = UserProfileSerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):      
        from polls.models import UserProfile as upro
        records = upro.objects.all()
        field,direction = '',''
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records
  
class LaborSerializer(serializers.ModelSerializer):
    class Meta:
        model = TaskLabor
        fields = '__all__'

# Create your views here.
class LaborListView(generic.ListView):
    model = TaskLabor
    def get_context_data(self, **kwargs):      
        context = super(LaborListView, self).get_context_data(**kwargs)
        return context

class LaborJsonView(generics.ListAPIView):
    serializer_class = LaborSerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):
        records = []
        session_id = self.request.GET.get('session_id','') 
        batch_no = self.request.GET.get('batch_no','')         
        is_mgmt = self.request.GET.get('is_mgmt','')
        is_detail = self.request.GET.get('is_detail','')  
        user_name = self.request.GET.get('user_name','')
        wo_number = self.request.GET.get('wo_number','')
        wtl_sels = self.request.GET.get('wtl_sels','')
        mod_type = self.request.GET.get('mod_type','')
        num_rows = self.request.GET.get('num_rows',10)
        
        if wtl_sels:
            import ast
            wtl_sels = ast.literal_eval(wtl_sels)
        if batch_no:
            records = TaskLabor.objects.filter(batch_id=batch_no).order_by('-id')
        if mod_type == 'ADD':
            records = TaskLabor.objects.filter(session_id='ay8nNoi80920KHOI:jgals82').orderl_by('id')
        #elif not session_id and num_rows:
        #    records = TaskLabor.objects.all().order_by('-id')[:num_rows]
        elif is_mgmt == '1':
            records = TaskLabor.objects.filter(session_id=session_id,skill_desc='1',batch_id='1',pn='1').order_by('user_name')
        elif is_detail == '1' and not wtl_sels and user_name and not wo_number:
            records = TaskLabor.objects.filter(session_id=session_id,user_name=user_name).order_by('-wtl_auto_key')
            #.exclude(batch_id='1',pn='1',skill_desc='1')
        elif session_id and wtl_sels:
            records = TaskLabor.objects.filter(session_id=session_id,wtl_auto_key__in=wtl_sels)
        elif session_id:
            records = TaskLabor.objects.filter(session_id=session_id)
        else:
            records = TaskLabor.objects.filter(session_id=session_id).order_by('id')
            
        field,direction = '',''
        
        if 'sort[0][dir]' in self.request.GET:
            
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        
        return records
        
   
class MailSerializer(serializers.ModelSerializer):
    class Meta:
        model = MailGroup
        fields = '__all__'

# Create your views here.
class MailListView(generic.ListView):
    model = MailGroup
    def get_context_data(self, **kwargs):       
        context = super(MailListView, self).get_context_data(**kwargs)
        return context


class MailJsonView(generics.ListAPIView):
    serializer_class = MailSerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):
        #records = MailGroup.objects.filter(active=True)
        records = MailGroup.objects.all()
        field,direction = '',''
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records
        

class UinpSerializer(serializers.ModelSerializer):
    class Meta:
        from polls.models import UserInput
        model = UserInput
        fields = '__all__'

# Create your views here.
class UinpListView(generic.ListView):
    from polls.models import UserInput
    model = UserInput
    def get_context_data(self, **kwargs):       
        context = super(EventListView, self).get_context_data(**kwargs)
        return context

class UinpJsonView(generics.ListAPIView):
    serializer_class = UinpSerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):
        format = '%Y-%m-%d %H:%M:%S'
        new_format = '%m/%d/%Y'
        date_from_input = self.request.GET['date_from']
        date_to_input = self.request.GET['date_to']  
        date_from = date_from_input and datetime.strptime(date_from_input, new_format) or None
        date_from = date_from and datetime.strftime(date_from, format) or None
        date_to = date_to_input and datetime.strptime(date_to_input, new_format) or None
        date_to = date_to and datetime.strftime(date_to, format) or None        
        app_id = self.request.GET['app_id']
        app_id = app_id and UserAppPerms.objects.filter(id=app_id) or None
        ml_apps_id = app_id and app_id[0] and app_id[0].ml_apps_id
        ml_apps_id = ml_apps_id and ml_apps_id.id or None
        if ml_apps_id and ml_apps_id == 403:
            ml_apps_id = MLApps.objects.filter(code__icontains='Labor')
        quapi_id = self.request.GET['quapi_id']
        quapi = quapi_id and QueryApi.objects.filter(id=quapi_id) or None
        quapi = quapi and quapi[0] or None
        req_get = self.request.GET
        user_name = req_get and req_get.get('user_name','')
        from polls.models import UserInput as Uinp
        
        if user_name:         
            
            if date_from and date_to:
                records = Uinp.objects.filter(ml_apps_id=ml_apps_id,user_name__iexact=user_name,timestamp__gte=date_from,timestamp__lte=date_to) or []  
            elif date_from and not date_to:
                records = Uinp.objects.filter(ml_apps_id=ml_apps_id,user_name__iexact=user_name,timestamp__gte=date_from) or []           
            elif not date_from and date_to: 
                records = Uinp.objects.filter(ml_apps_id=ml_apps_id,user_name__iexact=user_name,timestamp__lte=date_to) or [] 
            else: 
                records = Uinp.objects.filter(ml_apps_id=ml_apps_id,user_name__iexact=user_name) or []             
        elif date_from and not date_to:
            records = Uinp.objects.filter(ml_apps_id=ml_apps_id,timestamp__gte=date_from) or []           
        elif not date_from and date_to: 
            records = Uinp.objects.filter(ml_apps_id=ml_apps_id,timestamp__lte=date_to) or [] 
        elif date_from and date_to:
            records = Uinp.objects.filter(ml_apps_id=ml_apps_id,timestamp__gte=date_from,timestamp__lte=date_to) or [] 
        else: 
            records = Uinp.objects.filter(quapi_id=quapi_id,ml_apps_id__in=app_id) or []   
            
        field,direction = '',''       
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records
     


class EventSerializer(serializers.ModelSerializer):
    class Meta:
        model = EventNotification
        fields = '__all__'

# Create your views here.
class EventListView(generic.ListView):
    model = EventNotification
    def get_context_data(self, **kwargs):       
        context = super(EventListView, self).get_context_data(**kwargs)
        return context

class EventJsonView(generics.ListAPIView):
    serializer_class = EventSerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):

        req_get = self.request.GET
        
        # Here you list all your filter names
        filter_names = (
            'wo_number',
            'pn','task',
            'cond_code',
            'user_id',
            'mail_group_id',
        )
        
        filter_dict={}
        records = EventNotification.objects.all()

        for param in filter_names:
            
            if param in req_get and req_get[param]:
                filter_dict[param] = req_get[param]  
                   
        records = records.filter(**filter_dict)
            
        field,direction = '',''
        if 'sort[0][dir]' in req_get:
            try:
                direction = req_get['sort[0][dir]']
                field = req_get['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records       
    
class TaskSerializer(serializers.ModelSerializer):
    class Meta:
        model = WOTask
        fields = '__all__'

# Create your views here.
class TaskListView(generic.ListView):
    model = WOTask
    def get_context_data(self, **kwargs):       
        context = super(TaskListView, self).get_context_data(**kwargs)
        return context

class TaskJsonView(generics.ListAPIView):
    serializer_class = TaskSerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):
        session_id = 'session_id' in self.request.GET and self.request.GET['session_id'] 
        exclude_closed = self.request.GET.get('exclude_closed',False) 
        parent_auto_key = 'parent_auto_key' in self.request.GET and self.request.GET['parent_auto_key'] or ''    
        woo_auto_key = 'woo_auto_key' in self.request.GET and self.request.GET['woo_auto_key'] or ''
      
        if woo_auto_key and parent_auto_key:        
            records = WOTask.objects.filter(ult_parent_woo=parent_auto_key,woo_auto_key=woo_auto_key,session_id=session_id).order_by('wot_sequence','wot_auto_key')               
        elif exclude_closed:  
            records = WOTask.objects.filter(session_id=session_id).order_by('wot_sequence','wot_auto_key').exclude(status_type='Closed')
        else:  
            records = WOTask.objects.filter(session_id=session_id).order_by('wot_sequence','wot_auto_key')
        field,direction = '',''
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records
    
class ADTSerializer(serializers.ModelSerializer):
    class Meta:
        model = AuditTrail
        fields = '__all__'

# Create your views here.
class ADTListView(generic.ListView):
    model = AuditTrail
    def get_context_data(self, **kwargs):       
        context = super(ADTListView, self).get_context_data(**kwargs)
        return context

class ADTJsonView(generics.ListAPIView):
    serializer_class = ADTSerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):
        format = '%Y-%m-%d %H:%M:%S'
        new_format = '%m/%d/%Y'
        date_from_input = self.request.GET['date_from']
        date_to_input = self.request.GET['date_to']  
        date_from = date_from_input and datetime.strptime(date_from_input, new_format) or None
        date_from = date_from and datetime.strftime(date_from, format) or None
        date_to = date_to_input and datetime.strptime(date_to_input, new_format) or None
        date_to = date_to and datetime.strftime(date_to, format) or None        
        app_id = self.request.GET['app_id']
        app_id = app_id and UserAppPerms.objects.filter(id=app_id) or None
        ml_apps_id = app_id and app_id[0] and app_id[0].ml_apps_id
        ml_apps_id = ml_apps_id and ml_apps_id.id or None
        if ml_apps_id and ml_apps_id == 403:
            ml_apps_id = MLApps.objects.filter(code__icontains='Labor')
        quapi_id = self.request.GET['quapi_id']
        quapi = quapi_id and QueryApi.objects.filter(id=quapi_id) or None
        quapi = quapi and quapi[0] or None
        req_get = self.request.GET
        user_id = req_get and req_get.get('user_id',None)
        if user_id and user_id != 'None':         
            if date_from and date_to:
                records = AuditTrail.objects.filter(ml_apps_id__in=ml_apps_id,user_id__iexact=user_id,quapi_id=quapi,create_date__gte=date_from,create_date__lte=date_to).order_by('-id')  
            elif date_from and not date_to:
                records = AuditTrail.objects.filter(ml_apps_id__in=ml_apps_id,user_id__iexact=user_id,quapi_id=quapi,create_date__gte=date_from).order_by('-id')        
            elif not date_from and date_to: 
                records = AuditTrail.objects.filter(quapi_id=quapi,ml_apps_id__in=ml_apps_id,user_id__iexact=user_id,create_date__lte=date_to).order_by('-id')  
            else: 
                records = AuditTrail.objects.filter(quapi_id=quapi,ml_apps_id__in=ml_apps_id,user_id__iexact=user_id).order_by('-id')               
        elif date_from and not date_to:
            records = AuditTrail.objects.filter(quapi_id=quapi,ml_apps_id__in=ml_apps_id,create_date__gte=date_from).order_by('-id')          
        elif not date_from and date_to: 
            records = AuditTrail.objects.filter(quapi_id=quapi,ml_apps_id__in=ml_apps_id,create_date__lte=date_to).order_by('-id')
        elif date_from and date_to:
            records = AuditTrail.objects.filter(quapi_id=quapi,ml_apps_id__in=ml_apps_id,create_date__lte=date_to,create_date__gte=date_from).order_by('-id') 
        else: 
            records = AuditTrail.objects.filter(quapi_id=quapi,ml_apps_id__in=ml_apps_id).order_by('-id')      
        field,direction = '',''       
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records
        
class SaleSerializer(serializers.ModelSerializer):
    class Meta:
        model = Sale
        fields = '__all__'

# Create your views here.
class SaleListView(generic.ListView):
    model = Sale
    def get_context_data(self, **kwargs):       
        context = super(SaleListView, self).get_context_data(**kwargs)
        return context
        
class SaleJsonView(generics.ListAPIView):
    serializer_class = SaleSerializer
    pagination_class = SalePageNumberPagination
    
    def get_queryset(self, *args, **kwargs):
        session_id = self.request.GET['session_id']
        records = Sale.objects.filter(session_id=session_id)                  
        field,direction = '',''
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records
               
class PISerializer(serializers.ModelSerializer):
    class Meta:
        model = PILogs
        fields = '__all__'

# Create your views here.
class PIListView(generic.ListView):
    model = PILogs
    def get_context_data(self, **kwargs):       
        context = super(PIListView, self).get_context_data(**kwargs)
        return context

class PIJsonView(generics.ListAPIView):
    serializer_class = PISerializer
    pagination_class = PIPageNumberPagination
    
    def get_queryset(self, *args, **kwargs):
        session_id = self.request.GET['session_id'] 
        records = PILogs.objects.filter(session_id=session_id,active=1).order_by('-id')                   
        field,direction = '',''
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records
#==============================================================================================================

class StockPageNumberPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'pageSize'

class StockSerializer(serializers.ModelSerializer):
    class Meta:
        model = WOStatus
        fields = '__all__'

# Create your views here.
class StockListView(generic.ListView):
    model = WOStatus
    def get_context_data(self, **kwargs):       
        context = super(StockListView, self).get_context_data(**kwargs)
        return context

class StockJsonView(generics.ListAPIView):
    serializer_class = StockSerializer
    pagination_class = StockPageNumberPagination

    def get_queryset(self, *args, **kwargs):
        rec_set = [] 
        keepers = self.request.GET.get('keepers','F')
        session_id = self.request.GET.get('session_id','')         
        field,direction = '',''
        if keepers == 'T':
            wo_type = 'KEEPER'       
            records = session_id and WOStatus.objects.filter(session_id=session_id,wo_type = wo_type).order_by('-id')
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass 
            if isinstance(records,list):
                records = rec_set           
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records

#==============================================================================================================            
class RecordPageNumberPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'pageSize'

class RecordSerializer(serializers.ModelSerializer):
    class Meta:
        model = WOStatus
        fields = '__all__'

# Create your views here.
class RecordListView(generic.ListView):
    model = WOStatus
    def get_context_data(self, **kwargs):       
        context = super(RecordListView, self).get_context_data(**kwargs)
        return context

class RecordJsonView(generics.ListAPIView):
    serializer_class = RecordSerializer
    pagination_class = RecordPageNumberPagination

    def get_queryset(self, *args, **kwargs):
        rec_set = [] 
        shippers = self.request.GET.get('shippers','F')
        session_id = 'session_id' in self.request.GET and self.request.GET['session_id'] 
        user_id = 'user_id' in self.request.GET and self.request.GET['user_id']
        is_wos = 'is_wos' in self.request.GET and self.request.GET['is_wos'] or False
        is_dock = 'is_dock' in self.request.GET and self.request.GET['is_dock'] or False          
        active_mode = 'active_mode' in self.request.GET and self.request.GET['active_mode'] or False          
        is_rack = 'is_rack' in self.request.GET and self.request.GET['is_rack'] or False 
        is_shop = self.request.GET.get('is_shop','')
        sub_wo_gate = 'sub_wo_gate' in self.request.GET and self.request.GET['sub_wo_gate'] or ''
        is_toll_analysis = 'is_toll_analysis' in self.request.GET and self.request.GET['is_toll_analysis'] or ''
        is_toll_detail = 'is_toll_detail' in self.request.GET and self.request.GET['is_toll_detail'] or ''
        parent_auto_key = 'parent_auto_key' in self.request.GET and self.request.GET['parent_auto_key'] or ''
        is_loc_whs = 'is_loc_whs' in self.request.GET and self.request.GET['is_loc_whs'] or '' 
        is_parts_req = 'is_parts_req' in self.request.GET and self.request.GET['is_parts_req'] or ''  
        is_picking = 'is_picking' in self.request.GET and self.request.GET['is_picking'] or '' 
        sub_grid = 'sub_grid' in self.request.GET and self.request.GET['sub_grid'] or ''
        filter_val = 'filter_val' in self.request.GET and self.request.GET['filter_val'] or ''
        
        if shippers == 'T':
            wo_type = 'SHIPPER'       
            records = session_id and WOStatus.objects.filter(session_id=session_id,wo_type = wo_type).order_by('-id')
        elif is_shop == 'nada':          
            records = WOStatus.objects.filter(session_id=session_id).order_by('rank').order_by('due_date')       
        elif is_toll_detail == '1':
            if sub_wo_gate and sub_wo_gate != '0':
                records = is_wos != '0' and session_id and WOStatus.objects.filter(parent_auto_key=parent_auto_key,is_detail=True,sub_wo_gate=sub_wo_gate,session_id=session_id,active=1).order_by('-id') 
            else:
                records = is_wos != '0' and session_id and WOStatus.objects.filter(parent_auto_key=parent_auto_key,is_detail=True,session_id=session_id,active=1).order_by('-id')              
        elif is_toll_analysis == '1':  
            records = is_wos != '0' and session_id and WOStatus.objects.filter(is_detail=False,session_id=session_id,active=1).order_by('-id')
        elif session_id and (is_wos == '1' or is_shop == '1'):
           
            from django.db.models import Q
            rec_set1 = WOStatus.objects.filter(int_rank__lt=11,int_rank__gt=0,session_id=session_id).exclude(int_rank__isnull=True).exclude(int_rank=0).order_by(F('int_rank').asc(nulls_last=True),F('due_date').asc(nulls_last=True)) 
            rec_set2 = WOStatus.objects.filter(Q(int_rank__gte=11) | Q(int_rank=0),session_id=session_id).order_by(F('due_date').asc(nulls_last=True),F('int_rank').asc(nulls_last=True))     
            #rec_set3 = WOStatus.objects.filter(int_rank=0,session_id=session_id).order_by(F('due_date').asc(nulls_last=True))
            #rec_set = rec_set1 | rec_set2 | rec_set3
            rec_set = rec_set1 | rec_set2
            #records = list(rec_set1) + list(rec_set2) + list(rec_set3)
            records = list(rec_set1) + list(rec_set2)

        elif is_dock:           
            records = session_id and WOStatus.objects.filter(session_id=session_id).order_by(F('due_date').asc(nulls_last=True))
        elif is_rack == '1':      
            records = session_id and WOStatus.objects.filter(session_id=session_id).order_by('-id')                     
        elif is_loc_whs == '1':            
            records = session_id and WarehouseLocation.objects.filter(session_id=session_id).order_by('id')  
        elif sub_grid and filter_val:      
            records = session_id and WOStatus.objects.filter(si_number = filter_val, is_detail=True,session_id=session_id).order_by('ctrl_number')
        elif is_picking:    
            records = session_id and WOStatus.objects.filter(is_detail=False,session_id=session_id).order_by('wo_number','part_number')   
        else:      
            records = session_id and WOStatus.objects.filter(session_id=session_id).order_by('id')            
        field,direction = '',''
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass 
            if isinstance(records,list):
                records = rec_set           
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records

class WLSerializer(serializers.ModelSerializer):
    class Meta:
        model = WarehouseLocation
        fields = '__all__'

# Create your views here.
class WLListView(generic.ListView):
    model = WarehouseLocation
    def get_context_data(self, **kwargs):       
        context = super(WLListView, self).get_context_data(**kwargs)
        return context

class WLJsonView(generics.ListAPIView):
    serializer_class = WLSerializer
    pagination_class = RecordPageNumberPagination

    def get_queryset(self, *args, **kwargs):
        session_id = 'session_id' in self.request.GET and self.request.GET['session_id'] or ''
        is_loc_whs = 'is_loc_whs' in self.request.GET and self.request.GET['is_loc_whs'] or ''           
        if is_loc_whs == '1':            
            records = session_id and WarehouseLocation.objects.filter(session_id=session_id).order_by('id')             
        field,direction = '',''
        if 'sort[0][dir]' in self.request.GET:
            try:
                direction = self.request.GET['sort[0][dir]']
                field = self.request.GET['sort[0][field]']
            except:
                pass    
            if direction == 'asc':
                records = records.order_by(field)
            elif direction == 'desc':
                records = records.order_by('-'+field)
            else:
                return records
        return records

"""
Methods for PI_UPDATE:
"""
def contains_zero(num):
    for n in num:
        if n == '0':
            return True 
            break            
    return False
    
def get_control(barcode,delim_str):
    ctrl_number = 0
    ctrl_id = 0
    if barcode:
        ctrl_number = barcode.partition(delim_str)
        ctrl_id = ctrl_number and ctrl_number[2] or None
        ctrl_number = ctrl_number and ctrl_number[0] or None       
    return ctrl_number,ctrl_id
    
def get_control_pi(barcode,delim_str):
    ctrl_id = 0
    if barcode:
        partition_bc = barcode.partition(delim_str)
        ctrl_id = partition_bc and partition_bc[2] or None      
    return ctrl_id
    
def pi_update(request,quapi_id=None):
    val_dict,form = {},{}
    user_id,location_sel,msg,show_modal,error,session_id,user_error,req_post = '','','',False,'','','',None
    locations,user_rec = [],[]
    loc_error,loc_key = '',''
    batch_no,control_no,control_id,quantity,reent = None,None,None,None,False
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')   
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    val_dict['quapi_id'] = quapi_id
        
    #from portal.tasks import get_uoms
    #from polls.models import UomCodes
    #res = get_uoms.delay(quapi_id,session_id)
    #error,msg = res.get()
    #val_dict['uoms'] = UomCodes.objects.filter(session_id=session_id)
    
    if request.method == 'GET':
        #res = get_loc_whs_cart_nsync_beta.delay(quapi_id,dj_user_id,loc_ok=True,whs_ok=False,cart_ok=False,app='physical-inventory') 
        #loc_error,app = res.get() 
        form = PIUpdateForm()          
    if request.method == 'POST':
        req_post = request.POST
        form = PIUpdateForm(req_post)  #if not valid shows error with previous post values in corresponding field
        val_dict['form'] = form
        scan = 'stock_label' in req_post and req_post['stock_label'] or 'wo_number' in req_post and req_post['wo_number']
        if not scan:
            scan = 'scan' in req_post and req_post['scan'] or ''       
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0    
        batch = 'batch_no' in req_post and req_post['batch_no'] or ''
        if not batch:
            batch = 'batch' in req_post and req_post['batch'] or ''
        new_qty = 'quantity' in req_post and req_post['quantity'] or '' 
        if not new_qty:
            new_qty = 'new_qty' in req_post and req_post['new_qty'] or '' 
        #get the user_auto_key from this text
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        loc_session = 'loc_session' in req_post and req_post['loc_session'] or ''
        if not session_id and not loc_session:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
            loc_session = session_id
        elif not session_id and loc_session:
            session_id = loc_session
        elif session_id and not loc_session:
            loc_session = session_id
       
        show_modal = 'show_modal' in req_post and req_post['show_modal'] or ''
        loc_modal = 'loc_modal' in req_post and req_post['loc_modal'] or ''
        show_all = 'show_all' in req_post and req_post['show_all'] and req_post['show_all'] == '1' and '1' or (user_name and '1') or '0'
        #if user submitted clear list form by pressing button
        clear = 'clear_form' in req_post and req_post['clear_form'] or None
        ctrl_number,ctrl_id,stm_auto_key='','',''                                                                                       
        control_number = 'control_number' in req_post and req_post['control_number'] or 0
        control_id = 'control_id' in req_post and req_post['control_id'] or 0
        #if user submitted clear list form by pressing button
        clear = 'clear_formage' in req_post and req_post['clear_formage'] or None          
        loc_input = 'location' in req_post and req_post['location'] or ''
                                                                               
        val_dict.update({
            'error': error,
            'location': loc_input,
            'reent': '1',
            'batch_no': batch,#PI_NUMBER from PI_HEADER
            'quantity': new_qty,
            'loc_input': loc_input,
            'loc_vals': locations,
            'msg': msg,
            'show_modal': show_modal,
            'loc_modal': loc_modal or show_modal,
            'batch': batch,
            'scan': scan,
            'new_qty': new_qty,
            'control_id': control_id,
            'control_number': control_number,
            'ctrl_id': ctrl_id or control_id,
            'ctrl_number': ctrl_number or control_number, 
            'show_all': show_all or (user_name and '1') or '0',
            'user_name': user_name,
            'session_id': session_id, 
            'loc_session': loc_session, 
            'form': form,                     
            'sysur_auto_key': sysur_auto_key,            
        })

        if clear:
            woos_to_remove = req_post.getlist('woos_to_clear[]')                        
            if isinstance(woos_to_remove, list):               
                PILogs.objects.filter(pk__in=woos_to_remove).delete()
            return render(request, 'mrolive/pi_results.html', val_dict)


        if loc_input:
            
            from portal.tasks import lookup_location
            res = lookup_location.delay(quapi_id,session_id,loc_input)
            error,msg = res.get()
            
            loc_key = Location.objects.filter(session_id=session_id,
                location_code__iexact=loc_input
                )             
            loc_key = loc_key and loc_key[0] or None 
            
            if loc_key:
                loc_input = loc_key.location_code
                loc_key = loc_key.loc_auto_key
                
            else:
                val_dict['error'] = 'Location not found: %s'%loc_input
                return render(request, 'mrolive/pi_results.html', val_dict)
                                         
        
        if scan:
            if scan[0] in ['c','C']:
                stm_auto_key = scan[1:] 
                scan = scan[1:]                
            elif len(scan) < 7:
                error = "Your barcode is not long enough."
                      
        if not stm_auto_key:
            
            if user_name and scan:
                ctrl_number = scan[0:6]               
                ctrl_id = scan[7:len(scan)].partition('0')
                ctrl_id = ctrl_id and ctrl_id[2]
                
                if not ctrl_number and ctrl_id:
                    error = "Something is wrong with the control#/control id combination."
                   
                else:
                    from portal.tasks import check_stock
                    res = check_stock.delay(quapi_id,ctrl_id,ctrl_number)
                    error,msg = res.get()
               
                    if error:
                        val_dict.update({
                            'error': error,
                            'ctrl_id': ctrl_id or control_id,
                            'ctrl_number': ctrl_number or control_number,
                        })
                        
                        return render(request, 'mrolive/pi_results.html', val_dict)                      
        if error:
            val_dict.update({'error': error})
            return render(request, 'mrolive/pi_results.html', val_dict)

        if 'woos_to_clear[]' in req_post and clear:
            woos_to_remove = req_post.getlist('woos_to_clear[]')                        
            if isinstance(woos_to_remove, list):               
                PILogs.objects.filter(pk__in=woos_to_remove).delete()
            return render(request, 'mrolive/pi_results.html', val_dict)
                                 
        from portal.tasks import make_pi_updates            
        res = make_pi_updates.delay(\
            loc_session or session_id,\
            batch,ctrl_id or control_id,\
            ctrl_number or control_number,\
            new_qty,scan,user_name,\
            sysur_auto_key,quapi_id=quapi_id,\
            loc_input=loc_input,location_key=loc_key,stm=stm_auto_key)                
        error,msg = res.get()
        val_dict.update({
            'quantity': '',
            'location': loc_input, 
            'wo_number':scan,
            'stock_label':'',
            'error': error,
            'msg': msg,
            'show_modal':'False'
            })                                
    all_woos = PILogs.objects.filter(active=1,session_id=session_id) or []
    val_dict['total_rows'] = len(all_woos) 
    val_dict['all_woos'] = all_woos  
    val_dict['error'] = (error != 'show_modal' and error + str(user_error) + str(loc_error)) or ''
    val_dict['msg'] = msg
    val_dict['user_name'] = user_name
    form = PIUpdateForm(val_dict)
    val_dict['form'] = form
    val_dict['user'] = user
    return render(request, 'mrolive/pi_results.html', val_dict)

def is_integer(string):
    res = string
    try:
        res = string and int(string) or None
    except Exception as exc:
        logger.exception('Not authenticated due to invalid character in load view query string - Not an integer: %r', exc)
    return res 
    
def stock_reserve(request,quapi_id=None):     
    location,user_id,user_logged,update,warehouse,rack,new_rack,rerack,rack_user = '','','','','','','','',''
    wo_number,user_error,stat_error,loc_error = '','','',''
    val_dict,form,updated_woos,all_woos,woo_num_list,woo_key_list = {},{},[],[],[],[]           
    msg,loc_msg,stat_msg,error,lookup_recs,clear_cart = '','','','',False,False
    loc_key,whs_key,cart_key,new_status_name=None,None,None,''
    from polls.models import StatusSelection as stat_sel,QuantumUser as quser
    val_dict['quapi_id'] = quapi_id 
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    val_dict['user_name'] = user_name
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')        
    val_dict['emp_vals'] = quapi_id and QuantumUser.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id).distinct() or ''
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    alloc_app = MLApps.objects.filter(code="stock-issue")
    alloc_app = alloc_app and alloc_app[0] or None
    modes = alloc_app and get_modes(alloc_app) or []
    if not modes:
        return redirect('/login/') 
    val_dict['modes'] = modes    
    if request.method == 'GET':
        if not reg_user_id or not dj_user_id:
            val_dict['error'] = 'Access denied.'
            return redirect('/login/')  
        from portal.tasks import get_users_nsync_beta
        form = WODashboardForm()   
        res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=0,app='stock-allocation')
        user_error,app = res.get() 
        val_dict['active_mode'] = '3'
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        quantity = 'quantity' in req_post and req_post['quantity'] or ''
        total_rows = 'total_rows' in req_post and req_post['total_rows'] or 0
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        if not session_id:
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
        if not dj_user_id:
            dj_user_id = 'dj_user_id' in req_post and req_post['dj_user_id'] or ''#dj admin user id
        user_id = 'user_id' in req_post and req_post['user_id'] or ''#sysur_auto_key
        user_logged = 'user_logged' in req_post and req_post['user_logged'] or ''
        rack_user = 'rack_user' in req_post and req_post['rack_user'] or '' 
        user_id = user_id or user_logged or rack_user or ''
        #lookup user_id in the database to make sure we can authenticate
        user_rec = QuantumUser.objects.filter(quapi_id=quapi_id,user_id__iexact=user_id)
        user_rec = user_rec and user_rec[0] or None
        clear_form = 'clear_form' in req_post and req_post['clear_form'] or False       
        lookup_recs = 'lookup_recs' in req_post and req_post['lookup_recs'] or False           
        location = 'location' in req_post and req_post['location'] or '' 
        rack = 'rack' in req_post and req_post['rack'] or '' 
        new_rack = 'new_rack' in req_post and req_post['new_rack'] or '' 
        warehouse = 'warehouse' in req_post and req_post['warehouse'] or ''
        active_mode = 'mode_selector' in req_post and req_post['mode_selector'] or ''
        sel_mode = 'sel_mode' in req_post and req_post['sel_mode'] or ''  
        if not (sel_mode or active_mode):
            val_dict['error'] = 'Must select a mode.'
            render(request, 'mrolive/stock_reserve.html', val_dict) 
        cart_code = 'cart_code' in req_post and req_post['cart_code'] or ''         
        new_status = 'new_status' in req_post and req_post['new_status'] or ''     
        show_status = 'show_status' in req_post and req_post['show_status'] or ''
        show_user = 'show_user' in req_post and req_post['show_user'] or ''
        show_all = 'show_all' in req_post and req_post['show_all'] or ''
        clear_cart = 'ccart_form' in req_post and True or False
        label = 'label' in req_post and req_post['label'] or ''
        wo_task = 'wo_task' in req_post and req_post['wo_task'] or ''
        do_status = sel_mode and (sel_mode == '2' or sel_mode == '1') or False
        do_user = sel_mode or False       
        do_all = user_id or user_logged or rack_user or False  
        options_col,page_size = get_options(req_post,session_id)            
        val_dict.update({
            'wo_number': '',
            'all_woos': updated_woos,
            'msg': msg,
            'warehouse': warehouse,
            'location': location,
            'dj_user_id': dj_user_id,
            'user_id': user_id or (user_rec and user_rec.user_id) or user_logged or rack_user,
            'user_name': user_name,
            'rack': rack or new_rack,
            'user_logged': user_logged or user_id,
            'rack_user': user_id,
            'new_rack': rack,
            'modes': modes,
            'active_mode': active_mode or sel_mode or '',
            'sel_mode': sel_mode or active_mode or '',
            'cart_code': cart_code or rack or '',
            'label': '',
            'new_status': new_status and int(new_status) or None,
            'lookup_recs': lookup_recs,
            'show_status': show_status or do_status,
            'do_status': do_status or show_status,
            'show_user': show_user or do_user,
            'do_user': do_status or show_status,
            'show_all': show_all and show_all!='0' or do_all,
            'do_all': do_all or show_all,
            'session_id': session_id,
            'sel_rows': sel_rows,
            'total_rows': total_rows,
            'quantity': '',
            'form': form,  
            'options_col': options_col,
            'page_size': page_size,             
            })          
        #if user submitted clear list form by pressing button
        if clear_form and req_post['clear_form']=='1':       
            WOStatus.objects.filter(user_id=user_logged,is_dashboard=0,active=1,is_racking=1).delete()
            val_dict['all_woos'] = []            
            val_dict['msg'] = '' 
            val_dict['active_mode'] = sel_mode
            val_dict['show_all'] = 1  
            form = WODashboardForm(val_dict)
            val_dict['form'] = form                         
            return render(request, 'mrolive/stock_reserve.html', val_dict)              
        wo_number = wo_number or label or ''
        ctrl_number,ctrl_id = '',''        
        if wo_number and len(wo_number) > 6:
            ctrl_number = wo_number[:6]               
            ctrl_id = wo_number[7:]
        #create the demo wostatus objec
        #TODO:
        #   1. Add a new field for task + s
        #   2. Lookup task
        #   3. Find the ctrl#/id and then add a new reservation to the task for the stock move?
        #   4. Find the lowest wob_auto_key bom from the pnm (from the stm entered by user) and then 
        """
            
        """
        app_mode = sel_mode or active_mode or ''
        from portal.tasks import stock_reserve,stock_issue,stock_unissue
        if app_mode == '1':
            res = stock_reserve.delay(quapi_id,session_id,sysur_auto_key,user_id,wo_task,quantity,active_mode,ctrl_number,ctrl_id)
            error,msg,qty_res = res.get()
        elif app_mode == '3':
            qty_res = 'quantity' in req_post and req_post['quantity']
            must_reserve = not qty_res and 'T' or 'F'
            must_reserve = 'must_reserve' in req_post and req_post['must_reserve']
            res = stock_issue.delay(quapi_id,session_id,sysur_auto_key,user_name,ctrl_number,ctrl_id,wo_task,quantity,app_mode,must_reserve)
            error,msg,qty_res,must_reserve = res.get()  
            #val_dict['must_reserve'] = qty_res or 'T'
            #if qty_res:
            val_dict['quantity'] = ''
            val_dict['label'] = ''
            val_dict['wo_task'] = wo_task
            val_dict['must_reserve'] = must_reserve
        elif app_mode == '4':
            res = stock_unissue.delay(quapi_id,session_id,sysur_auto_key,user_name,quantity,wo_task,ctrl_number,ctrl_id)
            error,msg,quantity = res.get()
            val_dict['quantity'] = '' 
            val_dict['label'] = ''
            val_dict['wo_task'] = wo_task            
        updated_woos = WOStatus.objects.filter(session_id=session_id)
        val_dict['all_woos'] = updated_woos
        val_dict['total_rows'] = str(len(updated_woos))
        val_dict['msg'] = msg   
        val_dict['error'] = error + user_error + stat_error + loc_error       
        if not wo_number and lookup_recs not in [1,'1']:
            val_dict['lookup_recs'] = 1
        elif wo_number and lookup_recs not in [0,'0']:
            val_dict['lookup_recs'] = 0            
    form = WODashboardForm(val_dict)
    val_dict['form'] = form  
    val_dict['user'] = user    
    return render(request, 'mrolive/stock_reserve.html', val_dict) 
    
def construct_akl(woo_ids):
    woo_id_list = woo_ids and '(' + str(woo_ids[0]) or ''
    woo_lists = []
    if len(woo_ids) == 1:
        woo_id_list += ')'
        woo_lists = [woo_id_list]
        return woo_lists
    count = 1
    if woo_ids and len(woo_ids) > 1:
        for wak in woo_ids[1:]:
            woo_id_list += ',' + str(wak)
            #if we get to the 496th WOO, then we close out the string and will begin with another element to start the next list of 496.
            if (count+1)%495 == 0 and woo_ids[count+1]: 
                woo_id_list += ')'           
                woo_lists.append(woo_id_list)
                woo_id_list = '(' + str(woo_ids[count+1])
            count += 1 
        woo_id_list += ')'
        woo_lists.append(woo_id_list) 
    return woo_lists
    
def save_grid_options(request):
    
                              
    req_post = request.POST
    if req_post:
        col_options = req_post.get('col_options',[])
        pagesize = req_post.get('pagesize',10000)
        session_id = req_post.get('session_id','')
        quapi_id = req_post.get('quapi_id','')
        user_id = req_post.get('user_id','')
        ml_app_id = req_post.get('ml_app_id','')
        options_col = []
    
                   
        
            
                                                    
        if col_options and user_id:
                                                                
                                                                     
         
                                    
                                                         
            
            try:
                #insert the code you wrote today on Dev 
                #that inserts the col options records and saves them.
                options_col,pagesize = store_grid_options(req_post,\
                    session_id,'','',user_id=user_id,ml_app_id=ml_app_id\
            )
            except Exception:           
                return JsonResponse('error retrieving data.')
                
        result = {'message': 'Grid settings saved.','options_col': options_col}       
        return JsonResponse(result, safe = False)

    return render(request, 'mrolive/repair_order_mgmt.html')
 
def get_popup_stock(stm_key_list,session_id): 
    stms = WOStatus.objects.filter(session_id = session_id,stm_auto_key__in=stm_key_list)
    prod_pn,prod_desc,prod_cons,sl_list = '','','',''
    count = 0
    
    for stm in stms:
        if count == 0:
            prod_pn = stm.part_number
            prod_desc = stm.description
            prod_cons = stm.consignment_code
            
        else:                    
            if prod_pn != stm.part_number:
                prod_pn = 'Multiple'
            
            if prod_desc != stm.description:
                prod_desc = 'Multiple'
                
            if prod_cons != stm.consignment_code:
                prod_cons = 'Multiple'
        
        sl_list += stm.stock_line + ' | '
        count += 1
    return [prod_pn,prod_desc,prod_cons,sl_list]
    
def repair_order_mgmt(request,quapi_id=None):
    
    location,customer,user_id = '','',''
    wo_number,user_error,stat_error,loc_error = '','','',''
    val_dict,form,updated_woos,all_woos,woo_num_list,woo_key_list = {},{},[],[],[],[]           
    msg,loc_msg,stat_msg,error,lookup_recs,clear_cart = '','','','',False,False
    loc_key,whs_key,cart_key,new_status_name=None,None,None,''
    quser=QuantumUser
    val_dict['quapi_id'] = quapi_id 
    user = request.user
    
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
        
    user_id = user.username
    user_profile = UserProfile.objects.filter(user=user)
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile.sysur_auto_key
    reg_user_id = user and user.is_authenticated and user.id or None 
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None
    
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
        
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    alloc_app = MLApps.objects.filter(name="RO Management")
    alloc_app = alloc_app and alloc_app[0] or None
    modes = alloc_app and get_modes(alloc_app) or []
    val_dict['sel_rows'] = 0
    
    if request.method == 'GET':
        form = WODashboardForm() 
        sel_rows = 0
        
    from portal.tasks import get_stock_status,get_uda_status,get_consignment_codes
    session_id = 'annW8NK23R2ifOSALKO8234LAKASGLKL3OI;R'
    res = get_uda_status.delay(quapi_id,session_id,app='ro-management')
    stat_error,app = res.get()
    statuses = StatusSelection.objects.filter(session_id=session_id)
    val_dict['status_vals'] = statuses
    session_id = 'auaawe()agnsdlkghw^234Neir908OI23'
    res = get_stock_status.delay(quapi_id,session_id)
    error = res.get()        
    val_dict['stock_statuses'] = StatusSelection.objects.filter(session_id=session_id)   
    res = get_consignment_codes.delay(quapi_id,session_id)        
    error = res.get()
    val_dict['consignments'] = Consignments.objects.filter(session_id=session_id) 
    
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        part_number = 'part_number' in req_post and req_post['part_number'] or ''
        label = 'label' in req_post and req_post['label'] or ''
        condition_code = 'condition_code' in req_post and req_post['condition_code'] or ''
        location = 'location' in req_post and req_post['location'] or '' 
        customer = 'customer' in req_post and req_post['customer'] or '' 
        if not dj_user_id:
            dj_user_id = 'dj_user_id' in req_post and req_post['dj_user_id'] or ''#dj admin user id
        ctrl_number,ctrl_id = get_control(label,'000000')
        if not ctrl_id:
            ctrl_number,ctrl_id = get_control(label,'00000')
        #update hidden fields
        stock_status = 'stock_status' in req_post and req_post['stock_status'] or ''
        new_status = 'new_status' in req_post and req_post['new_status'] or ''
        filter_status = 'filter_status' in req_post and req_post['filter_status'] or ''
        filter_stock_status = 'filter_stock_status' in req_post and req_post['filter_stock_status'] or ''
        show_modal = 'show_modal' in req_post and req_post['show_modal'] or None
        filter_label = 'filter_label' in req_post and req_post['filter_label'] or ''
        filter_customer = 'filter_customer' in req_post and req_post['filter_customer'] or ''
        filter_wo_number = 'filter_wo_number' in req_post and req_post['filter_wo_number'] or ''
        filter_part_number = 'filter_part_number' in req_post and req_post['filter_part_number'] or ''
        filter_location = 'filter_location' in req_post and req_post['filter_location'] or ''
        filter_condition_code = 'filter_condition_code' in req_post and req_post['filter_condition_code'] or ''
        filter_session = 'filter_session' in req_post and req_post['filter_session'] or ''
        update_session = 'update_session' in req_post and req_post['update_session'] or ''
        user_session = 'user_session' in req_post and req_post['user_session'] or ''
        active_user = 'active_user' in req_post and req_post['active_user'] or ''
        update_user = 'update_user' in req_post and req_post['update_user'] or ''
        #user_id = 'user_id' in req_post and req_post['user_id'] or None       
        user_in = 'user_in' in req_post and req_post['user_in'] or ''
        #user_name = 'user_name' in req_post and req_post['user_name'] or '' 
        #fields from pop-up to carry back to main form.
        wo_stat = 'wo_stat' in req_post and req_post['wo_stat'] or ''
        wo_stock_stat = 'wo_stock_stat' in req_post and req_post['wo_stock_stat'] or ''
        cond_code = 'cond_code' in req_post and req_post['cond_code'] or ''  
        location_code = 'location_code' in req_post and req_post['location_code'] or ''
        wo_num = 'wo_num' in req_post and req_post['wo_num'] or ''
        part_num = 'part_num' in req_post and req_post['part_num'] or ''
        stock_lab = 'stock_lab' in req_post and req_post['stock_lab'] or '' 
        mode_code = 'mode_code' in req_post and req_post['mode_code'] or ''        
        clear_form = 'clear_form' in req_post and req_post['clear_form'] or False 
        vendor_input = 'vend_input' in req_post and req_post['vend_input'] or '' 
        stm_keys = 'stm_keys' in req_post and req_post['stm_keys'] or []
        active_mode = 'mode_selector' in req_post and req_post['mode_selector'] or ''
        ro_number = 'ro_number' in req_post and req_post['ro_number'] or ''
        sel_mode = 'sel_mode' in req_post and req_post['sel_mode'] or '' 
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or ''
        socond_code = req_post.get('socond_code','')
        filter_socond_code = req_post.get('filter_socond_code','')
        socondition_code = req_post.get('socondition_code','')
        active_mode = active_mode or sel_mode 
        stock_line = 'label' in req_post and req_post['label'] or ''                                                                 
        wo_update = 'wo_update' in req_post and req_post['wo_update'] or False
        show_all = 1
        is_search = req_post.get('is_search','')
        search_stock = 'search_stock' in req_post and req_post['search_stock'] or False   
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        upd_status = req_post.get('upd_status','')
        stock_status = req_post.get('stock_status','') 
        serial_number = req_post.get('serial_number','')
        instr = req_post.get('instr','')
        is_search = req_post.get('is_search','0')  
        launch_update = req_post.get('launch_update','0') 
        certifs = request.POST.getlist('cert',[])
        certifs = ', '.join(certifs)    
        upd_notes = "Please %s- provide %s.  %s"%(instr,certifs,req_post.get('upd_notes',''))
        upd_status = req_post.get('upd_status','')
        repair_type = req_post.get('repair_type','')
        log_choice = req_post.get('sysnl','')
        next_num = req_post.get('next_num','')
        cons_code = req_post.get('cons_selector','')
        separate_ros = req_post.get('separate_ros','') 
        yes_proceed = req_post.get('yes_proceed','')
        not_proceed = req_post.get('not_proceed','')  
        ro_warning = req_post.get('ro_warning_form','')         

        if not session_id and not (filter_session or user_session):
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
            filter_session = session_id
            user_session = session_id
            
        elif not session_id and filter_session:
            session_id = filter_session
            user_session = filter_session
        
        elif not session_id and not filter_session:
            session_id = user_session
            filter_session = user_session   
        
        val_dict.update({
            'wo_number': wo_number,
            'all_woos': updated_woos,
            'log_choice': log_choice,
            'msg': msg,
            'customer': customer,
            'condition_code': condition_code or cond_code,
            'socondition_code': socondition_code or filter_socond_code or socond_code,
            'socond_code': socond_code or filter_socond_code or socondition_code,
            'cond_code': cond_code or filter_condition_code or condition_code,
            'location': location or location_code,
            'label': label or stock_lab,
            'ctrl_id': ctrl_id,
            'ctrl_number': ctrl_number,
            'dj_user_id': dj_user_id,
            'user_id': user_id or active_user or update_user or user_in,
            'user_in': user_in or user_id or active_user or update_user,
            'active_user': active_user or user_id,
            'show_all': show_all,
            'show_modal': show_modal,
            'session_id': session_id or user_session or filter_session or update_session,
            'sel_rows': sel_rows or 0,
            'vendor_input': vendor_input,
            'form': form, 
            'quapi_id': quapi_id, 
            'active_mode': active_mode or sel_mode or mode_code or '',
            'mode_code': mode_code or active_mode or sel_mode or '',
            'sel_mode': sel_mode or active_mode or mode_code or '',
            'new_status':new_status or filter_status or wo_stat or '', 
            'filter_status': filter_status or new_status or wo_stat,  
            'wo_stat': wo_stat or new_status or filter_status, 
            'stock_status': stock_status or filter_stock_status or wo_stock_stat, 
            'filter_stock_status': filter_stock_status or stock_status or wo_stock_stat,  
            'wo_stock_stat': wo_stock_stat or stock_status or filter_stock_status,             
            'filter_wo_number': filter_wo_number or wo_number or wo_num,
            'wo_num': wo_num or filter_wo_number or wo_number,
            'part_num': part_num or filter_part_number or part_number,
            'part_number': part_number or filter_part_number or part_num,
            'filter_part_number': filter_part_number or part_num or part_number,
            'filter_label': filter_label or label or stock_lab,
            'stock_lab': stock_lab or filter_label or label,
            'filter_location': filter_location or location or location_code,
            'filter_condition_code': filter_condition_code or condition_code or cond_code, 
            'filter_socond_code': filter_socond_code or socondition_code or socond_code,
            'filter_session': filter_session or session_id or user_session or update_session,
            'update_session': update_session or session_id or user_session or filter_session, 
            'user_session': user_session or update_session or session_id or filter_session,
            'ro_number': ro_number,   
            'next_num': next_num,  
            'separate_ros': separate_ros,    
            'cons_code': cons_code,            
            })

        #if user submitted clear list form by pressing button
        if clear_form and req_post['clear_form']=='1':       
            WOStatus.objects.filter(user_id=user_logged,is_dashboard=0,active=1,is_racking=1).delete()
            val_dict['all_woos'] = []            
            val_dict['msg'] = '' 
            val_dict['show_all'] = 1  
            form = WODashboardForm(val_dict)
            val_dict['form'] = form                         
            return render(request, 'mrolive/repair_order_mgmt.html', val_dict)
            
        stm_key_list = []
        woo_keys = []
        selection = None
        
        if 'stm_sels[]' in req_post:
            stm_list = req_post.getlist('stm_sels[]') 
            stm_key_list = req_post.getlist('stm_sels[]') 
            selection = WOStatus.objects.filter(\
               stm_auto_key__in = stm_key_list,session_id=session_id)  
            woo_keys = [woo.woo_auto_key for woo in selection]
            
            #update with the data from the pop up
            from portal.tasks import update_lots
            parameters = [upd_status,upd_notes,'','','']              
            res = update_lots.delay(quapi_id,user_id,\
                session_id,sysur_auto_key,\
                stm_key_list,parameters,[],is_mgmt=True)
            error,msg = res.get()

        if 'stm_key_list[]' in req_post:
        
            stm_key_list = req_post.getlist('stm_key_list[]') 
            val_dict['stm_keys'] = stm_key_list
            stm_key_tuple = tuple([int(stm) for stm in stm_key_list])
            selection = WOStatus.objects.filter(stm_auto_key__in = stm_key_tuple,session_id=session_id)  
            woo_keys = [woo.woo_auto_key for woo in selection]

            if not_proceed:
                #return the grid and the rest of the data
                grid_rows = WOStatus.objects.filter(session_id=session_id)
                
                val_dict.update({
                    'msg': 'you chose not to proceed.',
                    'error': '',
                    'total_rows': len(grid_rows),
                })
                
                return render(request, 'mrolive/repair_order_mgmt.html', val_dict) 
                
            if launch_update == '1':
            
                stms = WOStatus.objects.filter(session_id = session_id,stm_auto_key__in=stm_key_list)
                prod_pn,prod_desc,prod_cons,sl_list = '','','',''
                count = 0
                for stm in stms:
                    if count == 0:
                        prod_pn = stm.part_number
                        prod_desc = stm.description
                        prod_cons = stm.consignment_code
                        
                    else:                    
                        if prod_pn != stm.part_number:
                            prod_pn = 'Multiple'
                        
                        if prod_desc != stm.description:
                            prod_desc = 'Multiple'
                            
                        if prod_cons != stm.consignment_code:
                            prod_cons = 'Multiple'
                    
                    sl_list += stm.stock_line + ' | '
                    count += 1
                
                all_woos = WOStatus.objects.filter(session_id=session_id) 
                total_rows = len(all_woos)
                
                val_dict.update({
                    'show_modal': 'update_stock',
                    'prod_pn': prod_pn,
                    'prod_desc': prod_desc,
                    'prod_cons': prod_cons,
                    'stm_list': stm_key_list,
                    'sl_list': sl_list,
                    'all_woos': '1',
                    'modes': modes,
                    'total_rows': total_rows,
                    'all_woos': True,
                    
                })
                
                return render(request, 'mrolive/repair_order_mgmt.html', val_dict)
        
        if not (wo_update or ro_number) and ((ctrl_number and ctrl_id)\
            or stock_status or socondition_code or wo_number or part_number\
            or condition_code or location or customer or new_status\
            or stock_line or cons_code):
            
            from portal.tasks import run_ro_mgmt
            res = run_ro_mgmt.delay(
                session_id,
                wo_number=wo_number,
                part_number=part_number,
                customer=customer,
                location=location,
                condition_code=condition_code,
                socond_code=socondition_code,
                user_id=user_id,
                sysur_auto_key=sysur_auto_key,
                stock_label = label,
                ctrl_id = ctrl_id,
                ctrl_number = ctrl_number,
                quapi_id = quapi_id,
                clear_cart = clear_cart,
                dj_user_id = dj_user_id,
                wos_auto_key = new_status,
                stock_status = stock_status,
                stock_line = stock_line,
                cons_code = cons_code,
            )
            error,msg = res.get()
            val_dict['sel_rows'] = 0
                            
        elif vendor_input and repair_type and not not_proceed:
            
            if 'stm_key_list[]' in req_post:
                stm_key_list = req_post.getlist('stm_key_list[]') 
                
            from portal.tasks import add_new_ro
             #get vendors now that we know we're going to need them
            #!!!!fix this!!!!
            from portal.tasks import get_companies_n_sync
            res = get_companies_n_sync.delay(quapi_id,dj_user_id,is_vendor=True)
            error = res.get()
            vendor_list = Companies.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id,is_vendor=True)           
            vendor = vendor_list.filter(name=vendor_input)            
            vendor = vendor and vendor[0]
            vendor_id = vendor and vendor.cmp_auto_key or 0

            all_woos = WOStatus.objects.filter(session_id=session_id) 
            total_rows = len(all_woos)           
            
            if not (not_proceed or yes_proceed) and log_choice:
                
                if vendor and vendor.allow_ro == 'Warn':
                    #launch the popup to display the warning
                                   
                    val_dict.update({
                        'warn_user':'1',
                        'ro_warn_msg': vendor.ro_warning,
                        'stm_keys': stm_key_list,
                        'msg': '',
                        'all_woos': all_woos,
                        'repair_type': repair_type,
                        'vend_input': vendor_input,
                    })
                    
                    return render(request, 'mrolive/repair_order_mgmt.html', val_dict) 
                    
                elif vendor and vendor.allow_ro == 'Never':
                    #launch an error pop-up
                    val_dict.update({
                        'error': 'Vendor not allowed for RO: ' + vendor.ro_warning,
                        'stm_keys': stm_key_list,
                        'msg': '',
                        'all_woos': all_woos,
                        'repair_type': repair_type,
                        'vend_input': vendor_input,
                    })
                    
                    return render(request, 'mrolive/repair_order_mgmt.html', val_dict) 
                    
            val_dict['vendor_vals'] = vendor_list
            val_dict['separate_ros'] = separate_ros
            val_dict['active_mode'] = '2'           
            #check the vendor selection to see if they are eligible for repairs
            #vendor_id = vendors and vendors[0] or None
            #vendor_id = vendor_id and vendor_id.cmp_auto_key or None                       
            val_dict['session_id'] = user_session
            session_id = user_session 
            selection = stm_key_list and WOStatus.objects.filter(session_id=session_id,stm_auto_key__in=stm_key_list) or [] 
            woo_keys = [woo.woo_auto_key for woo in selection]
            
            if vendor_id:

                if log_choice and log_choice.isnumeric() and not next_num:
                    from polls.models import NumberLog
                    log_chosen = NumberLog.objects.filter(sysnl_auto_key=log_choice)
                    next_num = log_chosen and log_chosen[0] and log_chosen[0].next_number
                 
                res = add_new_ro.delay(quapi_id,session_id,\
                sysur_auto_key,stm_keys=stm_key_list,\
                vendor=vendor_id,last_vendor=False,\
                woo_keys=woo_keys,separate_ros=separate_ros,\
                repair_type=repair_type,next_num=next_num,\
                sysnl_auto_key=log_choice,stock_status=upd_status,\
                notes=upd_notes,yes_proceed=yes_proceed)
                error,msg,next_nl = res.get()                 

                if upd_status or upd_notes and msg not in ['show_sysnl','warn_user']:
                    from portal.tasks import update_lots
                    parameters = [upd_status,upd_notes,'','','']              
                    res = update_lots.delay(quapi_id,user_id,\
                        session_id,sysur_auto_key,\
                        stm_key_list,parameters,[],is_mgmt=True)
                    upderror,updmsg = res.get() 
 
                if msg == 'show_sysnl' and not log_choice:
                    from polls.models import NumberLog
                    sysnl = NumberLog.objects.filter(session_id=session_id).order_by('sequence')                              
                    #update the values to show the sequences
                    
                    val_dict.update({
                        'show_sysnl':'1',
                        'stm_keys': stm_key_list,
                        'msg': '',
                        'sysnl': sysnl,
                        'repair_type': repair_type,
                        'vend_input': vendor_input,
                    })
                    
                if msg[:9] == 'warn_user' and not (not_proceed or yes_proceed):                    
                    val_dict.update({
                        'warn_user':'1',
                        'ro_warn_msg': msg[9:],
                        'stm_keys': stm_key_list,
                        'msg': '',
                        'repair_type': repair_type,
                        'vend_input': vendor_input,
                    })
                    
                    return render(request, 'mrolive/repair_order_mgmt.html', val_dict)                   

                if not error and log_choice: 
                    from portal.tasks import update_sysnl
                    res = update_sysnl.delay(quapi_id,session_id,log_choice,next_num)
                    upderror,sysnl_msg = res.get()   
                    
                val_dict['msg'] = msg
                val_dict['error'] = error                 
                val_dict['show_modal'] = '0'
                
            else:
                val_dict['error'] = 'Vendor not found. Must select from the list' 
           
             
        elif active_mode in ['1','2'] or ro_number:
               
            if stm_key_list:            
                from portal.tasks import add_new_ro

                if ro_number:
                    
                    if repair_type:
                            
                        res = add_new_ro.delay(quapi_id,session_id,sysur_auto_key,\
                        ro_number=ro_number,stm_keys=stm_key_list,last_vendor=False,\
                        woo_keys=woo_keys,separate_ros=False,repair_type=repair_type,
                        next_num=0)
                        error,msg,next_nl = res.get()
                        
                        if msg[:9] == 'warn_user':                   
                            val_dict.update({
                                'warn_user':'1',
                                'ro_warn_msg': msg[9:],
                                'stm_list': stm_key_list,
                                'msg': '',
                                'repair_type': repair_type,
                                'vend_input': vendor_input,
                            }) 
                        
                        if upd_status or upd_notes:
                            from portal.tasks import update_lots
                            parameters = [upd_status,upd_notes,'','','']              
                            res = update_lots.delay(quapi_id,user_id,\
                                session_id,sysur_auto_key,\
                                stm_key_list,parameters,[],is_mgmt=True)
                            upderror,updmsg = res.get()
                        
                    else:
                        prod_pn,prod_desc,prod_cons,sl_list = get_popup_stock(stm_key_list,session_id)
                        val_dict.update({
                            'ro_number': ro_number,
                            'separate_ros': separate_ros,
                            'show_reps':'1',
                            'stm_list': stm_key_list,
                            'prod_pn': prod_pn,
                            'prod_desc': prod_desc,
                            'prod_cons': prod_cons,
                            'stm_list': stm_key_list,
                            'sl_list': sl_list,
                        })
                        
                elif active_mode == '2' and is_search != '1' and not repair_type:
                    
                    #need to raise the pop-up to prompt for repair type.
                    prod_pn,prod_desc,prod_cons,sl_list = get_popup_stock(stm_key_list,session_id)
                    val_dict.update({
                        'user_session': session_id,
                        'show_reps':'1',
                        'stm_keys': stm_key_list,
                        'prod_pn': prod_pn,
                        'prod_desc': prod_desc,
                        'prod_cons': prod_cons,
                        'stm_list': stm_key_list,
                        'sl_list': sl_list,
                        'separate_ros': separate_ros,
                        'active_mode': '2',
                    })                                                      
                    
                elif active_mode == '2' and is_search != '1' and repair_type:
                    
                    #need to raise the pop-up to prompt for vendor_input.
                    val_dict['show_modal'] = '1'    
                    val_dict['repair_type'] = repair_type                   
                    val_dict['user_session'] = session_id                 
                    val_dict['stm_keys'] = stm_key_list
                    #get vendors now that we know we're going to need them
                    from portal.tasks import get_companies_n_sync
                    res = get_companies_n_sync.delay(quapi_id,dj_user_id,is_vendor=True)
                    error = res.get()
                    vendor_vals = Companies.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id,is_vendor=True)
                    val_dict['vendor_vals'] = vendor_vals
                    val_dict['separate_ros'] = separate_ros
                    val_dict['active_mode'] = '2'
                    
                elif active_mode == '1':

                    if repair_type:
                        
                        if log_choice and log_choice.isnumeric():
                            from polls.models import NumberLog
                            log_chosen = NumberLog.objects.filter(sysnl_auto_key=log_choice)
                            next_num = log_chosen and log_chosen[0] and log_chosen[0].next_number
                            
                        #get the last vendor via sql query and create a new RO with that vendor
                        res = add_new_ro.delay(quapi_id,session_id,sysur_auto_key,\
                            stm_keys=stm_key_list,last_vendor=True,woo_keys=woo_keys,\
                            separate_ros=separate_ros,repair_type=repair_type,next_num=next_num,\
                            sysnl_auto_key=log_choice,yes_proceed=yes_proceed)                      
                        error,msg,next_nl = res.get() 
                        
                        if msg[:9] == 'warn_user':
                            if not yes_proceed:                    
                                val_dict.update({
                                    'warn_user':'1',
                                    'ro_warn_msg': msg[9:],
                                    'stm_keys': stm_key_list,
                                    'msg': '',
                                    'repair_type': repair_type,
                                    'vend_input': vendor_input,
                                    'next_num': next_num,
                                })
                            
                            return render(request, 'mrolive/repair_order_mgmt.html', val_dict)   
                    
                        if upd_status or upd_notes:
                            from portal.tasks import update_lots
                            parameters = [upd_status,upd_notes,'','','']              
                            res = update_lots.delay(quapi_id,user_id,\
                                session_id,sysur_auto_key,\
                                stm_key_list,parameters,[],is_mgmt=True)
                            upderror,updmsg = res.get()
                        
                        if not error and log_choice and log_choice.isnumeric(): 
                            from portal.tasks import update_sysnl
                            res = update_sysnl.delay(quapi_id,session_id,log_choice,next_num)
                            sysnl_error,sysnl_msg = res.get()
                            
                        if msg == 'show_sysnl':
                            from polls.models import NumberLog
                            sysnl = NumberLog.objects.filter(session_id=session_id).order_by('sequence')                              
                            #update the values to show the sequences
                            
                            val_dict.update({
                                'show_sysnl':'1',
                                'stm_list': stm_key_list,
                                'msg': msg,
                                'sysnl': sysnl,
                                'repair_type': repair_type,
                                'error': error,
                            })
                    else:
                        prod_pn,prod_desc,prod_cons,sl_list = get_popup_stock(stm_key_list,session_id)
                        val_dict.update({
                            'separate_ros': separate_ros,
                            'show_reps':'1',
                            'stm_list': stm_key_list,
                            'prod_pn': prod_pn,
                            'prod_desc': prod_desc,
                            'prod_cons': prod_cons,
                            'stm_list': stm_key_list,
                            'sl_list': sl_list,
                        })
            elif not not_proceed:
                error = 'Must select row(s).'
     
        #if user has already entered a vendor:
        #the user clicked the "New RO" button so we add a new one with the stock lines as RO's
        stm_key_list =[int(stm) for stm in stm_key_list]
        updated_woos = WOStatus.objects.filter(session_id=session_id or user_session or filter_session or update_session)
        if not error and 'stm_sels[]' in req_post:
            updated_woos = updated_woos.exclude(stm_auto_key__in=stm_key_list)
        val_dict['all_woos'] = updated_woos
        val_dict['total_rows'] = str(len(updated_woos))
        val_dict['msg'] = msg
        val_dict['error'] = error

        options_col,pagesize = store_grid_options(req_post,session_id,user_profile,alloc_app)
        if options_col:
            val_dict.update({
                'options_col': options_col,
                'pagesize': pagesize,
            })
            
    form = WODashboardForm(val_dict)
    val_dict['form'] = form
                                                            
    val_dict['modes'] = modes
    val_dict['error'] = error
    
    return render(request, 'mrolive/repair_order_mgmt.html', val_dict)
    
def store_grid_options(req_post,\
    session_id,user_profile,ml_app,\
    user_id='',ml_app_id=''):        

    """
        store the local db for stored grid settings.
        return the grid id and col thru the get_options()        

    """
    options_col = {}
    page_size = 10000

    from polls.models import GridOptions
    if not user_profile:
                                                      
        user_profile = user_id and UserProfile.objects.filter(user_id = user_id)
        
    if user_profile:   
        if ml_app_id:
            ml_app = MLApps.objects.filter(code = ml_app_id)

            grid = GridOptions.objects.create(
                user_profile_id=user_profile,
                app_id = ml_app,
                session_id=session_id
            )
            
        else:

            grid = GridOptions.objects.create(
                user_profile_id=user_profile,
                session_id=session_id
            )
            
        grid = grid.save()

    if req_post:
        options_col,page_size = get_options(
            req_post,session_id,
            user_profile_id = user_profile,
            app_id = ml_app,
        )
        
    return options_col,page_size
  
def repair_order_edit(request,quapi_id=None): 
    location,customer,user_id = '','',''
    wo_number,user_error,stat_error,loc_error = '','','',''
    val_dict,form,updated_woos,all_woos,woo_num_list,woo_key_list = {},{},[],[],[],[]           
    msg,loc_msg,stat_msg,error,lookup_recs,clear_cart = '','','','',False,False
    loc_key,whs_key,cart_key,new_status_name=None,None,None,''
    quser=QuantumUser
    val_dict['quapi_id'] = quapi_id 
    user = request.user
    if not user.id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_id = user.username
    user_profile = UserProfile.objects.filter(user=user)
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile.sysur_auto_key
    reg_user_id = user and user.is_authenticated and user.id or None 
    dj_user_id = user and quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None 
    dj_user_id = dj_user_id and user.id or None
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    val_dict['emp_vals'] = quapi_id and QuantumUser.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id).distinct() or ''
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    alloc_app = MLApps.objects.filter(name="RO Management")
    alloc_app = alloc_app and alloc_app[0] or None
    modes = alloc_app and get_modes(alloc_app) or []
    val_dict['sel_rows'] = 0 
    from portal.tasks import get_users_nsync_beta,get_statuses_nsync_beta,get_uda_status,get_consignment_codes
    if request.method == 'GET':
        if not reg_user_id or not dj_user_id:
            val_dict['error'] = 'Access denied.'
            return redirect('/login/')  
        #form = WODashboardForm()   
        #res = get_users_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=0,app='ro-edit')
        #user_error,app = res.get()
                    
        res = get_statuses_nsync_beta.delay(quapi_id,dj_user_id,is_dashboard=1,app='ro-edit',object_type='RO')
        stat_error,app = res.get()
        form = WODashboardForm()        
    session_id = 'annW8NK23R2ifOSALKO8234LAKASGLKL3OI;R'    
    val_dict['status_vals'] = StatusSelection.objects.filter(quapi_id=quapi_id,dj_user_id=dj_user_id,is_dashboard=1).distinct() or [] 
    res = get_consignment_codes.delay(quapi_id,session_id)        
    error = res.get()
    val_dict['consignments'] = Consignments.objects.filter(session_id=session_id) 
    
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        split_user_id = 'split_user_id' in req_post and req_post['split_user_id'] or ''
        split_ro_submit = 'split_ro_submit' in req_post and req_post['split_ro_submit'] or ''
        split_ro_number = 'split_ro_number' in req_post and req_post['split_ro_number'] or ''
        split_wo_number = 'split_wo_number' in req_post and req_post['split_wo_number'] or ''
        split_vendor = 'split_vendor' in req_post and req_post['split_vendor'] or ''
        split_new_status = 'split_new_status' in req_post and req_post['split_new_status'] or ''
        split_upd_status = 'split_upd_status' in req_post and req_post['split_upd_status'] or ''
        split_entry_date= 'split_entry_date' in req_post and req_post['split_entry_date'] or ''
        split_pn = 'split_pn' in req_post and req_post['split_pn'] or ''
        split_session = 'split_session' in req_post and req_post['split_session'] or ''
        split_ro = 'split_ro' in req_post and req_post['split_ro'] == 'Split' or False
        split_misc_cost = 'split_misc_cost' in req_post and req_post['split_misc_cost'] or ''
        split_parts_cost = 'split_parts_cost' in req_post and req_post['split_parts_cost'] or ''
        split_labor_cost = 'split_labor_cost' in req_post and req_post['split_labor_cost'] or ''
        split_quoted = 'split_quoted' in req_post and req_post['split_quoted'] or ''
        split_approved = 'split_approved' in req_post and req_post['split_approved'] or ''
        split_next_dlv_date = 'split_next_dlv_date' in req_post and req_post['split_next_dlv_date'] or ''
        split_cond_code = 'split_quoted' in req_post and req_post['split_quoted'] or ''
        split_pnm_modify = 'split_pnm_modify' in req_post and req_post['split_pnm_modify'] or ''
        split_airway_bill = 'split_airway_bill' in req_post and req_post['split_airway_bill'] or ''
        split_cond_code = 'split_cond_code' in req_post and req_post['split_cond_code'] or ''
        split_notes = 'split_notes' in req_post and req_post['split_notes'] or ''
        split_new_status = req_post.get('split_new_status','')
        quantity = 'quantity' in req_post and req_post['quantity'] or 0
        is_update = req_post.get('is_update','')           
        ro_number = 'ro_number' in req_post and req_post['ro_number'] or ''
        wo_number = 'wo_number' in req_post and req_post['wo_number'] or ''
        new_status = req_post.get('new_status','')
        uda_status = req_post.get('uda_status','')
        repair_type = req_post.get('repair_type','')
        
        if not dj_user_id:
            dj_user_id = 'dj_user_id' in req_post and req_post['dj_user_id'] or ''#dj admin user id
            
        #user_id = 'user_id' in req_post and req_post['user_id'] or None
        user_name = 'user_name' in req_post and req_post['user_name'] or ''
        clear_form = 'clear_form' in req_post and req_post['clear_form'] or False
        upd_status = req_post.get('upd_status','0')
        header_notes = req_post.get('header_notes','')
        upd_header_notes = req_post.get('upd_header_notes','')
        split_header_notes = req_post.get('split_header_notes','')
        receiver_instr = req_post.get('receiver_instr','')
        upd_receiver_instr = req_post.get('upd_receiver_instr','')
        split_receiver_instr = req_post.get('split_receiver_instr','')
        vendor = 'vendor' in req_post and req_post['vendor'] or '' 
        notes = 'notes' in req_post and req_post['notes'] or '' 
        part_number = 'part_number' in req_post and req_post['part_number'] or '' 
        airway_bill = 'airway_bill' in req_post and req_post['airway_bill'] or '' 
        cond_code = 'condition_code' in req_post and req_post['condition_code'] or '' 
        pnm_modify = 'pnm_modify' in req_post and req_post['pnm_modify'] or ''
        entry_date = 'entry_date' in req_post and req_post['entry_date'] or ''
        ship_date = 'ship_date' in req_post and req_post['ship_date'] or ''
        ro_categ = req_post.get('ro_categ','')
        notes = 'notes' in req_post and req_post['notes'] or ''
        sel_rows = 'sel_rows' in req_post and req_post['sel_rows'] or 0
        upd_ro_submit = req_post.get('upd_ro_submit','')
        approved = 'approved' in req_post and req_post['approved'] or ''
        quoted = 'quoted' in req_post and req_post['quoted'] or ''
        next_dlv_date = 'next_dlv_date' in req_post and req_post['next_dlv_date'] or ''
        misc_cost = 'misc_cost' in req_post and req_post['misc_cost'] or ''
        parts_cost = 'parts_cost' in req_post and req_post['parts_cost'] or ''
        labor_cost = 'labor_cost' in req_post and req_post['labor_cost'] or ''
        upd_approved = 'upd_approved' in req_post and req_post['upd_approved'] or ''
        upd_quoted = 'upd_quoted' in req_post and req_post['upd_quoted'] or ''
        upd_ship_date = 'upd_ship_date' in req_post and req_post['upd_ship_date'] or ''
        upd_ndlv_date = 'upd_ndlv_date' in req_post and req_post['upd_ndlv_date'] or ''
        upd_misc_cost = 'upd_misc_cost' in req_post and req_post['upd_misc_cost'] or ''
        upd_parts_cost = 'upd_parts_cost' in req_post and req_post['upd_parts_cost'] or ''
        upd_labor_cost = 'upd_labor_cost' in req_post and req_post['upd_labor_cost'] or ''
        upd_pnm_modify = 'upd_pnm_modify' in req_post and req_post['upd_pnm_modify'] or ''
        upd_airway_bill = 'upd_airway_bill' in req_post and req_post['upd_airway_bill'] or ''
        upd_cond_code = 'upd_cond_code' in req_post and req_post['upd_cond_code'] or ''
        upd_notes = 'upd_notes' in req_post and req_post['upd_notes'] or ''
        show_all = 1
        search_stock = 'search_stock' in req_post and req_post['search_stock'] or False  
        filter_vendor = 'filter_vendor' in req_post and req_post['filter_vendor'] or ''
        filter_uda_status = 'filter_uda_status' in req_post and req_post['filter_uda_status'] or ''
        filter_upd_status = 'filter_upd_status' in req_post and req_post['filter_upd_status'] or ''
        filter_new_status = 'filter_new_status' in req_post and req_post['filter_new_status'] or ''
        filter_entry_date = 'filter_entry_date' in req_post and req_post['filter_entry_date'] or ''
        filter_ro_number = 'filter_ro_number' in req_post and req_post['filter_ro_number'] or ''
        filter_wo_number = 'filter_wo_number' in req_post and req_post['filter_wo_number'] or ''
        filter_part_number = 'filter_part_number' in req_post and req_post['filter_part_number'] or ''
        filter_session = 'filter_session' in req_post and req_post['filter_session'] or ''
        update_session = 'update_session' in req_post and req_post['update_session'] or ''
        update_user = 'update_user' in req_post and req_post['update_user'] or ''
        active_user = 'active_user' in req_post and req_post['active_user'] or ''       
        session_id = 'session_id' in req_post and req_post['session_id'] or ''
        cons_code = req_post.get('cons_selector','')
        
        if not session_id and not (filter_session or split_session):
            session_id = 'csrfmiddlewaretoken' in req_post and req_post['csrfmiddlewaretoken'] or ''
            filter_session = session_id
            split_session = session_id
            
        if not session_id and filter_session:
            session_id = filter_session
            
        elif not session_id and split_session:
            session_id = split_session
           
        val_dict.update({
            'cons_code': cons_code,
            'header_notes': '',
            'receiver_instr': '',
            'approved': '',
            'quoted': '',
            'next_dlv_date': '',
            'misc_cost': '',
            'parts_cost': '',
            'labor_cost': '',
            'ro_number': ro_number or filter_ro_number or split_ro_number,
            'wo_number': wo_number or filter_wo_number or split_wo_number,
            'new_status': new_status or filter_new_status or split_new_status,
            'all_woos': updated_woos,
            'msg': msg,           
            'entry_date': entry_date or filter_entry_date or split_entry_date,
            'ship_date': ship_date,
            'part_number': part_number or filter_part_number or split_pn,
            'dj_user_id': dj_user_id,
            'user_id': user_id or active_user or update_user or split_user_id,
            'active_user': active_user or user_id or split_user_id,
            'update_user': update_user or user_name or split_user_id,
            'session_id': session_id,
            'sel_rows': sel_rows or 0,
            'vendor': vendor or filter_vendor,
            'split_user_id': split_user_id or active_user or user_id or update_user or '',
            'pnm_modify': '',
            'airway_bill': '',
            'condition_code': '',
            'notes': '',              
            'upd_cond_code': upd_cond_code or cond_code or split_cond_code,
            'upd_airway_bill': upd_airway_bill or airway_bill or split_airway_bill,
            'upd_approved': upd_approved or approved or split_approved,
            'upd_labor_cost': upd_labor_cost or labor_cost or split_labor_cost,
            'upd_misc_cost': upd_misc_cost or misc_cost or split_misc_cost,
            'upd_ndlv_date': upd_ndlv_date or next_dlv_date or split_next_dlv_date,
            'upd_ship_date': upd_ship_date,
            'upd_status': upd_status or filter_upd_status or split_upd_status,
            'upd_notes': upd_notes or notes or split_notes,
            'upd_header_notes': upd_header_notes or header_notes or split_header_notes,
            'upd_receiver_instr': upd_receiver_instr or receiver_instr or split_receiver_instr,
            'form': form,
            'filter_new_status': filter_new_status or new_status or split_new_status,            
            'filter_ro_number': filter_ro_number or ro_number or split_ro_number,
            'filter_wo_number': filter_wo_number or wo_number or split_wo_number,
            'filter_entry_date': filter_entry_date or entry_date or split_entry_date,
            'filter_part_number': filter_part_number or part_number or split_pn,
            'filter_vendor': filter_vendor or vendor or split_vendor, 
            'split_ro_number': filter_ro_number or ro_number or split_ro_number,
            'split_wo_number': filter_wo_number or wo_number or split_wo_number,
            'split_entry_date': filter_entry_date or entry_date or split_entry_date,
            'split_part_number': filter_part_number or part_number or split_pn,
            'split_vendor': filter_vendor or vendor or vendor or split_vendor, 
            'split_new_status': filter_new_status or new_status or split_new_status,
            'filter_session': filter_session or session_id,
            'split_session': session_id or filter_session,                      
            })  
            
        show_modal = '0'
        ro_id_list = []               
        selection = None
        
        if is_update:
            if 'woos_list[]' in req_post:
            
                from portal.tasks import get_conditions,get_categories
                res = get_conditions.delay(quapi_id,session_id)
                error = res.get()                
                res = get_categories.delay(quapi_id,session_id)
                error = res.get()
                ro_id_list = req_post.getlist('woos_list[]')
                from polls.models import PartConditions as pcc, Categories as rct
                ro_categs = rct.objects.filter(session_id=session_id)
                conditions = pcc.objects.filter(session_id=session_id).order_by('condition_code')
                val_dict['conditions'] = conditions
                val_dict['ro_categs'] = ro_categs
                val_dict['rod_keys'] = ro_id_list
                val_dict['update_modal'] = '1'
                val_dict['form'] = form           
                return render(request, 'mrolive/repair_order_edit.html', val_dict)
                
            else:  
                error += 'No records selected.'
                val_dict['error'] = error
                
        if split_ro:
            stm_sels = req_post.getlist('stm_sels[]')
            if not stm_sels:
                error = 'No records selected.'
            else:                
                val_dict.update({
                    'show_modal':'1',
                    'rod_keys': stm_sels,
                })                    
            
        if (is_update or split_ro) and not user_id and clear_form != '1':
            error += 'Invalid employee number.'
            val_dict['error'] = error
            return render(request, 'mrolive/repair_order_edit.html', val_dict)   
            
        #if user submitted clear list form by pressing button
        if clear_form and req_post['clear_form']=='1':       
            #WOStatus.objects.filter(session_id=session_id,is_dashboard=0,active=1,is_racking=1).delete()
            val_dict['all_woos'] = []            
            val_dict['msg'] = '' 
            val_dict['show_all'] = 1  
            form = WODashboardForm(val_dict)
            val_dict['form'] = form                         
            return render(request, 'mrolive/repair_order_edit.html', val_dict)                         

        fields_update = misc_cost or parts_cost or labor_cost\
            or airway_bill or pnm_modify or cond_code\
            or approved or quoted or next_dlv_date\
            or notes or upd_status or header_notes or receiver_instr\
            or ro_categ or ship_date or repair_type
        
        if not upd_ro_submit:
            from portal.tasks import run_ro_edit
            res = run_ro_edit.delay(
                session_id,
                quapi_id,
                ro_number = ro_number,
                wo_number = wo_number,
                vendor = vendor,
                part_number = part_number,
                entry_date = entry_date,
                rst_auto_key = new_status,
                uda_status = uda_status,
                cons_code = cons_code,
            )
            error,msg = res.get()
            
        else:
            stm_sels = req_post.getlist('stm_sels[]')
            if not stm_sels:
                error = 'No records selected.' 
            else:
                from portal.tasks import ro_update_costs_dates
                res = ro_update_costs_dates.delay(
                    session_id,
                    quapi_id,
                    misc_cost=misc_cost,
                    parts_cost=parts_cost,
                    labor_cost=labor_cost,
                    quoted_date=quoted,
                    approved_date=approved,
                    next_dlv_date=next_dlv_date,
                    ro_id_list=stm_sels,
                    airway_bill=airway_bill,
                    pnm_modify=pnm_modify,
                    cond_code=cond_code,
                    notes=notes,
                    uda_status=upd_status,
                    header_notes=header_notes,
                    sysur_auto_key=sysur_auto_key,
                    new_status=new_status,
                    ro_number=ro_number,
                    vendor=vendor,
                    part_number=part_number,
                    entry_date=entry_date,
                    wo_number=wo_number,
                    receiver_instr=receiver_instr,
                    ship_date=ship_date,
                    ro_categ=ro_categ,
                    repair_type=repair_type,
                )
                error,msg = res.get()
                from portal.tasks import run_ro_edit
                res = run_ro_edit.delay(
                    filter_session,
                    quapi_id,
                    ro_number = filter_ro_number,
                    wo_number = filter_wo_number,
                    vendor = filter_vendor,
                    part_number = filter_part_number,
                    entry_date = filter_entry_date,
                    rst_auto_key = filter_new_status,
                    uda_status = filter_uda_status,
                )
                error,search_msg = res.get()
                
        if quantity and split_ro_submit == '1':
            rod_keys_sel = req_post.getlist('rod_keys_sel[]')
            if len(rod_keys_sel) > 1:
                error = 'Must select only 1 RO to split.'
            elif len(rod_keys_sel) == 1:
                rod_auto_key = rod_keys_sel and rod_keys_sel[0] or []
                rod = WOStatus.objects.filter(rod_auto_key=rod_auto_key,session_id=session_id)
                rod = rod and rod[0] or None
                stm_auto_key = rod and rod.stm_auto_key or None
                qty_reserved = rod and rod.qty_reserved or 0
                #user just submitted for an RO with multiple stock reservations/stm_auto_keys.
                #now we submit everything and the message we will get backs is 'show_grid_modal'               
                from portal.tasks import split_ro
                res = split_ro.delay(
                    session_id,
                    quapi_id,
                    quantity,
                    qty_reserved,
                    misc_cost=split_misc_cost,
                    parts_cost=split_parts_cost,
                    labor_cost=split_labor_cost,
                    approved_date=split_approved,
                    quoted_date=split_quoted,
                    next_dlv_date=split_next_dlv_date,
                    notes = split_notes,
                    ro_id_list=rod_keys_sel,
                    stm_auto_key=stm_auto_key,
                    sysur_auto_key=sysur_auto_key,
                    ro_number = split_ro_number,
                    vendor = split_vendor,
                    part_number = split_pn,
                    entry_date = split_entry_date)
                error,msg = res.get()
                if msg == 'show_grid_modal':
                    #count total row for grid to display on modal pop-up
                    updated_woos = WOStatus.objects.filter(is_modal_grid=True,session_id=session_id)
                    val_dict['all_woos'] = updated_woos
                    val_dict['total_rows'] = str(len(updated_woos))
            else:
                error = 'Must select an RO to split.'
        #else:
            #error = 'Enter a value in the filters to search.'
        updated_woos = WOStatus.objects.filter(session_id=session_id)
        val_dict['all_woos'] = updated_woos
        val_dict['total_rows'] = str(len(updated_woos))
        val_dict['msg'] = msg
        val_dict['error'] = error      
    form = WODashboardForm(val_dict)
    val_dict['form'] = form
    user_profile = UserProfile.objects.filter(user=user)
    user_profile = user_profile and user_profile[0] or None
    val_dict['num_records'] = user_profile.num_records or 10
    val_dict['modes'] = modes
    return render(request, 'mrolive/repair_order_edit.html', val_dict)
    
def get_options(req_post,session_id,user_profile_id=None,app_id=None):
    
    page_size = 'options_pagesize' in req_post and req_post['options_pagesize'] or 25
    options_col = 'options_col' in req_post and req_post['options_col'] or '' 
    
    if not options_col:
         options_col = 'update_col' in req_post and req_post['update_col'] or ''
         page_size = 'update_pagesize' in req_post and req_post['update_pagesize'] or 25
         
    if not options_col:
         options_col = 'vendor_col' in req_post and req_post['vendor_col'] or ''
         page_size = 'vendor_pagesize' in req_post and req_post['vendor_pagesize'] or 25
         
    if not options_col:
         options_col = 'split_ro_col' in req_post and req_post['split_ro_col'] or ''
         page_size = 'split_ro_pagesize' in req_post and req_post['split_ro_pagesize'] or 25
         
    if not options_col:
         options_col = 'loc_col' in req_post and req_post['loc_col'] or ''
         page_size = 'loc_pagesize' in req_post and req_post['loc_pagesize'] or 25
         
    if options_col:
        convert_to_dict = options_col.replace('true','True')
        convert_to_dict = convert_to_dict.replace('false','False') or None
        options_col = convert_to_dict and eval(convert_to_dict) or []
        num = 0
        col_recs = []
        
        cols_to_del = ColumnSettings.objects.filter(session_id=session_id)

        if app_id and user_profile_id:
            from polls.models import GridOptions
            grid = GridOptions.objects.filter(
                app_id=app_id,
                user_profile_id = user_profile_id
            )
            
            grid = grid and grid[0]
           
            if grid:
                cols_to_del = ColumnSettings.objects.filter(groptions_id=grid)
            
        cols_to_del.delete()

        for col in options_col:
            #store each column with its name and widths
            col_recs.append(ColumnSettings(
                    name = 'title' in col and col['title'] or '',
                    field = 'field' in col and col['field'] or '',
                    width = 'width' in col and col['width'] and float(col['width']) or 80,
                    tmpl_text = 'template' in col and col['template'] or '',
                    obj_type = 'grid',
                    session_id = session_id,
                    seq_num = num,
                )
            )  
            num += 1

        try:
            ColumnSettings.objects.bulk_create(col_recs)
        except Exception as exc:
            error = "\r\Problem with creating column settings locally: %s"%exc 
        options_col = ColumnSettings.objects.filter(session_id=session_id)
        
    return options_col,page_size
            
def dock_receiving(request,quapi_id=None):
    val_dict,msg,message = {},'',''
    sysnl = []
    next_num = 0
    val_dict['quapi_id'] = quapi_id 
    user = request and request.user or None
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')  
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    user_name = user and user.username or 'No Username'
    reg_user_id = user and user.is_authenticated and user.id or None
    dj_user_id = quapi_id and UserQuapiRel.objects.filter(user=user,quapi_id=quapi_id) or None
    dj_user_id = dj_user_id and user.id or None 
    if not reg_user_id or not dj_user_id:
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')      
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    show_modal,error = '0',''
    if request.method == 'GET': 
        from portal.tasks import get_users_nsync_beta
        form = WODashboardForm()   
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)
        print_label = req_post.get('print_label','')
        label_num = req_post.get('label_num','')
        session_id = req_post.get('session_id','')
        user_session = req_post.get('user_session','')
        if not session_id and not user_session:
            session_id = req_post.get('csrfmiddlewaretoken','')
        elif not session_id and user_session:
            session_id = user_session
        elif not user_session and session_id:
            user_session = session_id           
        if not dj_user_id:
            dj_user_id = req_post.get('dj_user_id','')
        airway_bill = req_post.get('airway_bill','')
        arrival_date = req_post.get('arrival_date','')
        location = req_post.get('location','')                                                       
        search = req_post.get('search',None)   
        loc_wo_type = req_post.get('loc_wo_type','')
        loc_wo_number = req_post.get('loc_wo_number','')
        loc_priority = req_post.get('loc_priority','')
        loc_rc_number = req_post.get('loc_rc_number','')
        loc_company = req_post.get('loc_company','')          
        wo_type = req_post.get('wo_type','')
        wo_number = req_post.get('number','')
        priority = req_post.get('priority','')
        rc_number = req_post.get('rc_number','')  
        company = req_post.get('company','')
        print_label = req_post.get('print_label',False)
        label = req_post.get('label','') or airway_bill
        message = req_post.get('msg','')
        log_choice = req_post.get('sysnl','')
                   
        val_dict.update({
            'msg': message,
            'airway_bill': airway_bill,
            'arrival_date': arrival_date,
            'location': location,
            'msg': msg,
            'dj_user_id': dj_user_id,
            'session_id': session_id or user_session,
            'user_session': user_session or session_id,                                             
            'form': form,
            'label': label or label_num, 
            'label_num': label_num or label,  
            'show_data': 0, 
            'wo_type': wo_type or loc_wo_type,
            'wo_number': wo_number or loc_wo_number,
            'priority': priority or loc_priority,
            'rc_number': rc_number or loc_rc_number,
            'company': company or loc_company,           
            }) 
            
        if label:
            
            #first, need search mechanism for SO/RO/PO (priority)
            #====================================================

            if log_choice:
                
                from portal.tasks import get_soropos,update_sysnl
                from polls.models import NumberLog
                log_chosen = NumberLog.objects.filter(sysnl_auto_key=log_choice)
                next_num = log_chosen and log_chosen[0] and log_chosen[0].next_number
                res = get_soropos.delay(quapi_id,label,session_id,next_num=next_num)
                error,msg = res.get()       

            else:
                from portal.tasks import get_soropos
                res = get_soropos.delay(quapi_id,label,session_id)
                error,msg = res.get()  
                #check to see if multiple sequences
                
                from polls.models import NumberLog
                sysnl = NumberLog.objects.filter(session_id=session_id).order_by('sequence')  
                
            if not error:          
                soh_auto_key,roh_auto_key,poh_auto_key = '','',''
                order = WOStatus.objects.filter(session_id=session_id)
                order = order and order[0] or None
                order_type = order.wo_type
                if order_type == 'SO':
                    soh_auto_key = order.parent_auto_key
                elif order_type == 'RO':
                    roh_auto_key = order.parent_auto_key
                elif order_type == 'PO':
                    poh_auto_key = order.parent_auto_key
                syscm_auto_key = order.syscm_auto_key
                cmp_auto_key = order.cmp_auto_key
                dpt_auto_key = order.dpt_auto_key
                next_number = order.next_num
                next_num = not next_num and next_number.replace(' ','') or next_num.replace(' ','')
                order_num = order.wo_number               
                priority = order.priority
                airway_bill = order.airway_bill or airway_bill
                cust_ref_number = order.cust_ref_number
                pnm_auto_key = order.pnm_modify
                arrival_date = order.arrival_date
                sowo_number = order.si_number
                due_date = order.due_date
                vendor = order.vendor
                tracking = order.spn_code
                open_flag = order.is_toll
                lot_number = order.stock_owner
                
                #vendor and date were submitted so we can create the RC now
                if open_flag and not sysnl:
                    from portal.tasks import create_rc

                    res = create_rc.delay(quapi_id,session_id,airway_bill,order_num,order_type,arrival_date,location,syscm_auto_key,cmp_auto_key,dpt_auto_key,next_num,cust_ref_number,pnm_auto_key,priority,soh_auto_key,roh_auto_key,poh_auto_key,sysur_auto_key,tracking,log_choice)
                    error,msg,location_code,existing_rc = res.get()

                    val_dict.update({
                        'element': next_num,
                        'arrival_date': arrival_date,
                        'location': location_code,
                        'wo_type': order_type,
                        'wo_number': order_num,
                        'sowo_number': sowo_number,
                        'vendor': vendor,
                        'due_date': due_date,
                        'priority': priority,
                        'rc_number': existing_rc,
                        'company': order.customer or order.vendor,
                        'order_type':order_type,
                        'show_data': 1,
                        'lot_number': lot_number,
                        'session_id': session_id,
                        'sysnl': sysnl,
                        'show_sysnl': sysnl and 'show' or 'nope',
                                                                 
                        })
                        
                    if not error and log_choice: 
                        
                        res = update_sysnl.delay(quapi_id,session_id,log_choice,next_num)
                        error,msg = res.get()
                        
                                                               
                if not error and not sysnl:                    
                    return render(request, 'mrolive/plain_barcode_dock.html', val_dict)
    
    show_sysnl = not error and sysnl and 'show' or 'nope'                                                 
    val_dict.update({
        'msg': msg or message,
        'error': error,
        'sysnl': not error and sysnl or '',
        'show_sysnl': show_sysnl,
    })
    form = WODashboardForm(val_dict)
    val_dict['form'] = form
    return render(request, 'mrolive/dock_receiving.html', val_dict)
                                                              
def timeclock(request,quapi_id):
    error,msg,warning_msg = '','',''
    user = request and request.user or None
    user_profile = user and UserProfile.objects.filter(user=user) or None   
    user_profile = user_profile and user_profile[0] or None   
    sysur_auto_key = user_profile and user_profile.sysur_auto_key or None
    val_dict = {}
    if not (user and user.id):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/')   
    lic_error = check_license_expiry(user)
    if lic_error:
        val_dict['lic_error'] = lic_error
        return render(request, 'registration/home.html', val_dict)
    form = WODashboardForm()  
    reg_user_id = user and user.is_authenticated and user.id or None
    user_apps = user and UserAppPerms.objects.filter(user=user) or None
    op_apps = user_apps and user_apps.filter(ml_apps_id__app_type='operations').order_by('ml_apps_id__menu_seq') or None
    val_dict['op_apps'] = op_apps
    mgmt_apps = user_apps and user_apps.filter(ml_apps_id__app_type='management').order_by('ml_apps_id__menu_seq') or None
    val_dict['mgmt_apps'] = mgmt_apps
    dash_apps = user_apps and user_apps.filter(ml_apps_id__app_type='dashboards').order_by('ml_apps_id__menu_seq') or None
    val_dict['dash_apps'] = dash_apps
    setup_apps = user_apps and user_apps.filter(ml_apps_id__app_type='setup').order_by('ml_apps_id__menu_seq') or None
    val_dict['setup_apps'] = setup_apps
    app_id = MLApps.objects.filter(code='timeclock')[0]
    app_allow = user_apps and user_apps.filter(ml_apps_id=app_id)
    val_dict['quapi_id'] = quapi_id
    
    if not (reg_user_id and app_allow):
        val_dict['error'] = 'Access denied.'
        return redirect('/login/') 
        
    if request.method == 'GET':
        form = WODashboardForm()
 
    if request.method == 'POST':
        req_post = request.POST
        form = WODashboardForm(req_post)      
        user_id = req_post.get('user_id','')
        val_dict['user_id'] = user_id
        
        if not user_id:
            val_dict['form'] = form
            val_dict['error'] = 'Enter user ID.'
            return render(request, 'mrolive/timeclock.html', val_dict)
            
        user_name = req_post.get('user_name','')
        quser = QuantumUser.objects.filter(user_name=user_id)
        
        if not quser:
            quser = QuantumUser.objects.filter(user_id=user_id)
            if not quser:
                quser = QuantumUser.objects.filter(employee_code=user_id)
                
        user_name = quser and quser[0] and quser[0].user_name or ''
        val_dict['user_name'] = user_name
        
        if not user_name:
            val_dict['form'] = form
            val_dict['error'] = 'No user found.'
            return render(request, 'mrolive/timeclock.html', val_dict) 
            
        mode = req_post.get('mode_sel','') 
        clock_in = mode == 'clock-in' and 'clocked in.'
        clock_out = mode == 'clock-out' and 'clocked out.'
        lunch_in = mode == 'lunch-in' and 'clocked back in from lunch.'
        lunch_out = mode == 'lunch-out' and 'clocked out for lunch.'
                
        session_id = req_post.get('session_id','')

        if not mode:
            modes = app_id and get_modes(app_id) or []  
            
            if not modes:
                val_dict['form'] = form
                val_dict['error'] = 'Contact administrator to define timeclock modes'
                return render(request, 'mrolive/timeclock.html', val_dict)
                
            val_dict['modes'] = modes
            val_dict['show_conf'] = 'T'
            
        else:
            mode_msg = clock_in or clock_out or lunch_in or lunch_out
            msg = 'Successfully ' + mode_msg
            
        if not session_id:
            session_id = req_post.get('csrfmiddlewaretoken','')
            
        val_dict.update({
            'session_id': session_id,   
            'msg': msg,  
            'sel_mode': mode,            
        })    
        
    form = WODashboardForm(val_dict)
    val_dict['form'] = form
    return render(request, 'mrolive/timeclock.html', val_dict)